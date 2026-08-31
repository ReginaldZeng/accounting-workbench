# -*- coding: utf-8 -*-
# [Change Log]
# Date: 2026-08-07 | Author: Claude / c | Version: V2.241
# Description: 报表导出·工作簿组装。一个主体一个文件，五个表页：
#              资产负债表 / 利润表 / 现金流量表（来自 KDS_Report）+ 科目余额 / 序时账簿（来自 GL_BALANCE / GL_VOUCHER）。
#              版式对齐金蝶原生导出，但**修掉它的列宽毛病**（业务方明确要求"注意列宽"）：
#                · 序时簿原生 14 列全是 30.6 宽（金蝶把 defaultColWidth 设成了 30.625）——
#                  摘要实测需 217 字符、核算维度需 109，反而不够；期间只要 4、凭证字只要 6，白占一屏。
#                · 科目余额原生默认 9.0 宽——核算维度名称需 75、8 个金额列各需 16，全被截成 ####。
#                · 三大报表原生把"会企01表/单位：元"那列留 8.6（需 15，被切掉），还各拖着 49~56 个幻影空列。
#              另给两张长表加冻结表头 + 自动筛选（序时簿两万多行，没这个没法看）。
#
# ⚠ 只写**值**、不写公式。金蝶原生导出里三大报表有 56 个公式格（合计行），
#   而 openpyxl 写公式时**只留公式串、丢掉缓存值**——Excel 打开会自己重算看不出来，
#   但 pandas / Power BI / 任何读缓存值的程序读到的就是空白（V2.241 开发中实测：
#   利润表 B23 净利润 -4,002,793.58 变空）。合计值本来就由金蝶算好随 <Data> 一起给了，
#   直接落值最稳，也不需要在服务器上装 LibreOffice 去重算。
import os
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter as _gl

import kingdee_client as kc

# 列宽：按参考文件**实测内容宽度**定，不是拍脑袋。
# 短列给内容宽+2；超长文本列（摘要/核算维度）不追内容，给个舒服的阅读宽度，要看全的人自己拉；
# 金额列按 #,##0.0000 带千分位的最长实测值 16 位 +1。
WIDTHS = {
    "资产负债表": [38, 16, 16, 40, 16, 16],
    "利润表": [45, 16, 18],
    "现金流量表": [60, 16, 18],
    #  科目编码 科目名称 维度编码 维度名称 + 期初/本期/本年/期末 各借贷两列
    "科目余额": [16, 22, 34, 34, 17, 17, 17, 17, 17, 17, 17, 17],
    #  账簿 日期 期间 凭证字 凭证号 摘要 科目编码 科目名称 核算维度 借方 贷方 制单 审核 来源系统
    "序时账簿": [26, 11, 6, 7, 8, 45, 13, 20, 42, 16, 16, 11, 11, 12],
}
# 三大报表的合并区（会企01/02/03 表版式固定，照参考文件原样复刻）
MERGES = {"资产负债表": ["A1:F1", "A3:D3"], "利润表": ["A1:C1", "A2:B2", "A3:B3"],
          "现金流量表": ["A1:C1", "A3:B3"]}
# 字体：业务方定用微软雅黑（金蝶原生导出是宋体，屏幕上发虚、数字辨识度差）。
# 字号沿用金蝶：三大报表 9、两张账表 11——雅黑同字号比宋体显大，再调大就撑破列宽了。
FACE = "微软雅黑"
F_RPT = Font(name=FACE, size=9)
F_RPT_B = Font(name=FACE, size=9, bold=True)
F_LGR = Font(name=FACE, size=11)
F_LGR_B = Font(name=FACE, size=11, bold=True)
CENTER = Alignment(horizontal="center", vertical="center")
# 边框只画三大报表（会企01/02/03 表是**正式报表**，本来就带表格线；金蝶原件数据区 100% 有 thin 框）。
# 科目余额与序时账簿**刻意不画**——金蝶原件那两张也是一格框都没有，两万多行画满格线只会更难读，
# 可读性靠冻结表头＋筛选＋列宽解决。
THIN = Side(style="thin")
BOX = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
FMT2, FMT4 = "#,##0.00", "#,##0.0000"
SUBJECT_HDR2 = ["科目编码", "科目名称", "核算维度编码", "核算维度名称",
                "期初余额", None, "本期发生", None, "本年累计", None, "期末余额", None]


def _put(ws, r, c, v, font, fmt=None, center=False):
    cell = ws.cell(r, c, v)
    cell.font = font
    if fmt:
        cell.number_format = fmt
    if center:
        cell.alignment = CENTER
    return cell


def _widths(ws, name):
    for i, w in enumerate(WIDTHS[name], start=1):
        ws.column_dimensions[_gl(i)].width = w


def _write_fin_sheet(wb, name, grid):
    """三大报表：Spread 网格 {(行,列): 值} → 表页。行列 0 基转 1 基。"""
    ws = wb.create_sheet(name)
    ncol = len(WIDTHS[name])
    maxr = 0
    for (r, c), v in grid.items():
        if c >= ncol:                                    # 幻影列：金蝶一路拖到 BJ，不搬过来
            continue
        isnum = isinstance(v, (int, float))
        # 表头四行（标题/期间/编制单位/列名）加粗居中，与参考文件一致
        hdr = r <= 3 and not isnum
        _put(ws, r + 1, c + 1, v, F_RPT_B if hdr else F_RPT,
             FMT2 if isnum else None, center=hdr)
        maxr = max(maxr, r + 1)
    # 表格线：从列名行（第 4 行）画到最后一行，**整片都画**——包括值为空的格。
    # 金蝶原件就是这样（数据区 100% 覆盖）；只给有值的格画框会画出一张缺牙的表。
    # 同时给空格子也设字体：openpyxl 新建的格默认 Calibri，不设的话往里一打字就冒出个西文字体。
    for r in range(4, maxr + 1):
        for c in range(1, ncol + 1):
            cell = ws.cell(r, c)
            cell.border = BOX
            if cell.value is None:
                cell.font = F_RPT
    for m in MERGES.get(name, []):
        ws.merge_cells(m)
    # 关掉网格线（业务方定，只关三大报表）：这三张自己带表格线，再叠一层灰网格，
    # 表外那片空白会被切成豆腐块，正式报表的样子就没了。
    # 科目余额/序时账簿**不关**——它们刻意不画边框，全靠网格线分行分列，关了就成一片糊字。
    ws.sheet_view.showGridLines = False
    _widths(ws, name)
    return ws


def _write_subject(wb, rows, book_name, year, period, cur):
    """科目余额：三行表头（币别/账簿/期间 → 分组名 → 借贷）+ 明细 + 合计行。版式照参考文件。"""
    ws = wb.create_sheet("科目余额")
    _put(ws, 1, 1, "币别:%s" % cur, F_LGR)
    _put(ws, 1, 2, "账簿:%s" % book_name, F_LGR)
    _put(ws, 1, 3, "期间:%d.%d -- %d.%d" % (year, period, year, period), F_LGR)
    for i, t in enumerate(SUBJECT_HDR2, start=1):
        if t:
            _put(ws, 2, i, t, F_LGR_B, center=True)
    for i in range(5, 13):                                # 8 个金额列的第二行表头：借方/贷方
        _put(ws, 3, i, "借方" if i % 2 else "贷方", F_LGR_B, center=True)
    for i in range(1, 5):
        ws.merge_cells(start_row=2, start_column=i, end_row=3, end_column=i)   # 前四列纵向合并
    for i in range(5, 13, 2):
        ws.merge_cells(start_row=2, start_column=i, end_row=2, end_column=i + 1)
    for ri, row in enumerate(rows, start=4):
        for ci, v in enumerate(row, start=1):
            if v is None:
                continue
            _put(ws, ri, ci, v, F_LGR, FMT4 if ci >= 5 else None)
    ws.freeze_panes = "A4"
    _widths(ws, "科目余额")
    return ws


def _write_journal(wb, rows):
    """序时账簿：一行表头 + 明细。两万多行——冻结表头并挂自动筛选，否则根本没法用。"""
    ws = wb.create_sheet("序时账簿")
    for i, t in enumerate(kc.JOURNAL_COLS, start=1):
        _put(ws, 1, i, t, F_LGR_B, center=True)
    for ri, row in enumerate(rows, start=2):
        for ci, v in enumerate(row, start=1):
            if v is None or v == "":
                continue
            fmt = FMT4 if ci in (10, 11) else ("#,##0" if ci == 3 else None)
            _put(ws, ri, ci, v, F_LGR, fmt)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:%s%d" % (_gl(len(kc.JOURNAL_COLS)), max(ws.max_row, 1))
    _widths(ws, "序时账簿")
    return ws


_BAD_FN = re.compile(r'[\\/:*?"<>|]')


def month_dir(year, period):
    """按月分子目录：2026年07月（业务方定）。
    一个月 8 个文件、一年 96 个堆在一个目录里没法看；分了月，找哪个月一目了然，
    也方便整月归档／整月删除。"""
    return "%d年%02d月" % (int(year), int(period))


def file_name(year, period, org, org_name):
    """2026年06期_101_深圳市星期零食品科技有限公司_财务报表.xlsx
    （比金蝶原生少了尾部那串时间戳；主体名里的非法文件名字符替换成 _）
    文件名里**保留年期**——文件被单独拷走、脱离了月份目录，还能自证是哪个期间的。"""
    return "%d年%02d期_%s_%s_财务报表.xlsx" % (year, int(period), org, _BAD_FN.sub("_", org_name or ""))


def build_workbook(year, period, rpt, sheets, subject_rows, journal_rows, cur):
    """组装一个主体的工作簿。表页顺序＝三大报表在前、两张账表在后（业务方口径：
    "三大报表，和科目余额表，序时账簿"），**不用金蝶那个 资产负债表→序时簿→科目余额→利润表→现金流量表 的怪序。**"""
    wb = Workbook()
    wb.remove(wb.active)
    for name in kc.FIN_RPT_SHEETS:                        # 资产负债表 / 利润表 / 现金流量表
        if name in sheets:
            _write_fin_sheet(wb, name, sheets[name])
    _write_subject(wb, subject_rows, rpt["org_name"], year, period, cur)
    _write_journal(wb, journal_rows)
    return wb


def export_one(year, period, rpt, out_dir, s=None, conf=None):
    """取一个主体的五张表 → 落一个 xlsx。返回 (文件全路径, 各表行数)。"""
    sheets = kc.fetch_fin_report_sheets(rpt["rid"], s, conf)
    cur = rpt.get("cur") or "人民币"
    subject = kc.fetch_subject_balance_full(year, period, rpt["org"], cur=cur, s=s, conf=conf)
    journal = kc.fetch_journal_full(year, period, rpt["org"], s=s, conf=conf)
    wb = build_workbook(year, period, rpt, sheets, subject, journal, cur)
    mdir = os.path.join(out_dir, month_dir(year, period))
    os.makedirs(mdir, exist_ok=True)
    path = os.path.join(mdir, file_name(year, period, rpt["org"], rpt["org_name"]))
    # 先写临时文件、再原子改名，**不要直接 wb.save(目标路径)**：
    # 取件机每分钟来扫一次，openpyxl 往目标路径写的那几秒里文件是残缺的——
    # 正好被扫到就会取走半个 Excel。改名是原子操作，外界要么看到旧的、要么看到完整的新的，
    # 没有中间态。同名覆盖＝重导即更新，不留新旧两份让人猜哪个准。
    tmp = path + ".part"
    wb.save(tmp)
    os.replace(tmp, path)
    return path, {"三大报表": len(sheets), "科目余额": len(subject), "序时账簿": len(journal)}

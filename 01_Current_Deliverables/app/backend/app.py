# -*- coding: utf-8 -*-
# [Change Log]
# Date: 2026-07-03 | Author: Claude / c | Version: V1.1 (was V1.10)
# Description: 财务核算工作台 本地应用 后端(FastAPI)。复用内核；数据源可切换 样例/金蝶。
#              金蝶模式经 kingdee_client 真实拉数(GL_BALANCE/CN_BANKACNT/GL_VOUCHER)→同一批内核。
#              /api/config 切换源与期间，/api/kingdee/test 测连接。只读，不写金蝶。
#              V1.1: 逐笔稽核接引擎 v2(七态,接匹配桥); 资金看板/稽核加缓存(GET 读缓存 / 仅 sync 重取, 进入秒开)。
import os
import json
import shutil
import datetime
import time
import threading
import calendar
import hashlib
import re
import urllib.parse
from io import BytesIO
from fastapi import FastAPI, Request
from fastapi.responses import Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import JSONResponse

from kernels import balance_dashboard as bd
from kernels import reconcile as rc
from kernels import account_ledger as al
from kernels import bank_import as bimp
from kernels import subject_balance as sb
from kernels import wealth_statement as wstmt
from kernels import wealth_recon as wr
from kernels import logistics_accrual as la
from kernels import logistics_recon as lrc
from kernels import cost_ledger as clg
from kernels import fx_rate as fx
import sample_data as S
import kingdee_client as kc
import db
import mailer
import notifier

# ── 共享内核（V2.172 从本文件拆出；见 core.py 头部说明）──
from core import (  # noqa: F401  部分名供 routers/ 与本文件共用
    AUTH_LEDGER_PATH, BASE, CFG, DIST, KdNotFetched, LEDGER_PATH, SAMPLE_YM, _BADJ_CACHE,
    _CH_CACHE, _DS_CACHE, _FUND_CACHE, _OPEN_API, _RECON_CACHE, _SBAL_CACHE, _SYNC_AT,
    _WR_CACHE, _cache_clear, _cache_get, _cache_key, _closed_block, _closed_info, _current_user,
    _is_closed, _kd_fetch_store, _kd_get, _kd_sync_info, _kd_synced_at, _now, _period_bank,
    _period_data_status, _period_str, _require_perm, _user_public, save_cfg, sid_name, VERSION_INFO,
    _PULL_PATHS, pull_token_ok, can_enter_dev, dev_users_info,
)


_CAT2SUBJ = {"银行账户": "银行存款", "理财产品": "交易性金融资产",
             "电商渠道": "其它货币资金", "现金": "库存现金"}
# 认领操作人名单（认领=账务处理，只核算组；BP组曾禹锡/吴卓文不涉及账务、不进认领）。
# 真登录/账号系统为后续一期，届时再纳全部门与权限。
ROSTER = [
    {"组": "核算组", "成员": ["小朋友A", "小朋友B", "李志鹏", "陈梓华", "黄春艳", "冀欣欣", "冯辉"]},
]


def _load_claims() -> dict:
    return db.load_claims()             # 多人共享：认领状态存 DB


def _assign_keys(results):
    """给每条逐笔结果配稳定 key（内容哈希+同键序号），供认领状态挂靠；确定性引擎→跨次稳定。"""
    seen = {}
    for r in results:
        raw = "|".join(str(r.get(k)) for k in
                       ("账号", "日期", "方向", "借方金额", "贷方金额", "摘要", "金蝶凭证", "status"))
        k = hashlib.md5(raw.encode("utf-8")).hexdigest()[:12]
        seen[k] = seen.get(k, 0) + 1
        r["key"] = k if seen[k] == 1 else f"{k}{seen[k]}"


def _overlay_claims(results):
    """把 claims.json 的认领状态叠加到结果上（每次 GET 现读，认领即时生效，不吃缓存）。"""
    claims = _load_claims()
    for r in results:
        c = claims.get(r.get("key"))
        r["认领状态"] = (c.get("状态") if c else "待认领")
        r["认领人"] = (c.get("操作人", "") if c else "")
        r["认领时间"] = (c.get("时间", "") if c else "")
        r["认领备注"] = (c.get("备注", "") if c else "")


def _load_overrides() -> dict:
    return db.load_overrides()          # 多人共享：存 DB（SQLite本地 / MySQL服务器）


def _save_overrides(ov: dict):
    db.save_overrides(ov)


def _audit_scheme(cat: str) -> str:
    """稽核方案：银行账户 → 逐笔明细稽核；其余(理财/电商/现金) → 余额稽核。"""
    return "明细" if cat == "银行账户" else "余额"


def _auth_ledger_records(period_str):
    """金蝶出纳银行账号(CN_BANKACNT)建的权威台账 -> 前端账户台账记录。"""
    if not os.path.exists(AUTH_LEDGER_PATH):
        return None
    try:
        rows = json.load(open(AUTH_LEDGER_PATH, encoding="utf-8"))
    except Exception:
        return None
    overrides = _load_overrides()
    out = []
    for r in rows:
        cat = r.get("类别", "")
        acct = r.get("账号", "")
        ov_acct = overrides.get(str(acct), {}) if acct else {}
        kd_active = bool(r.get("_active", r.get("状态") == "生效"))
        manual_off = bool(ov_acct.get("失效"))
        scheme = ov_acct.get("稽核方案") if ov_acct.get("稽核方案") in ("明细", "余额") else _audit_scheme(cat)
        if not kd_active:
            status = "已销户"        # 金蝶侧已销户
        elif manual_off:
            status = "失效"          # 手工标失效（金蝶不维护此步）
        else:
            status = "生效"
        out.append({
            "账号": acct,
            "账户全名": r.get("账户全名", ""),
            "开户行": r.get("开户行", ""),
            "主体": r.get("主体", ""),
            "类别": cat,
            "科目大类": _CAT2SUBJ.get(cat, r.get("科目大类", "银行存款")),
            "币种": r.get("币种", "CNY"),
            "稽核方案": scheme,
            "稽核方案_手工": bool(ov_acct.get("稽核方案")),
            "状态": status,
            "手工失效": manual_off,
            "_active": kd_active and not manual_off,   # 生效 = 金蝶生效 且 未手工失效
            "本月新增": False,
            "首次出现期间": r.get("第一笔动账日期", "") or "",
            "最近同步期间": period_str,
            "来源": "金蝶出纳·银行账号",
        })
    return out

app = FastAPI(title="财务核算工作台", version="1.1")


@app.exception_handler(Exception)
def _clg_config_missing_handler(request: Request, exc: Exception):
    """成本台账配置缺失 → 说人话，别甩 500（V2.136）。
    只认领 ClgConfigMissing，其余异常照常按 500 抛出（不吞别人的错）。"""
    if type(exc).__name__ != "ClgConfigMissing":
        raise exc
    return JSONResponse({"ok": False, "msg":
        "服务器缺少成本台账配置文件 sample_data/cost_ledger_config.json（存货类别↔总账科目对照）——"
        "它是代码配置、不是数据，请把部署包里的这个文件放到 backend/sample_data/ 下再重启。"},
        status_code=500)
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])


@app.middleware("http")
async def _api_no_store(request, call_next):
    """所有 /api/ 响应禁缓存：后端更新字段后，浏览器普通刷新即拿最新，不会拿旧接口数据(字段对不上)。"""
    resp = await call_next(request)
    if request.url.path.startswith("/api/"):
        resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate"
        resp.headers["Pragma"] = "no-cache"
    return resp


# ---------------- 登录 / 账号（阶段2） ----------------


@app.middleware("http")
async def _auth_gate(request, call_next):
    """登录门：除白名单外的 /api/* 都要登录；未登录返回 401（前端据此弹登录页）。
    例外（V2.241）：报表取件通道——内网取件机没有会话、只揣令牌，令牌对得上才放行；
    令牌没配或对不上，照旧走登录门（见 core.pull_token_ok）。"""
    p = request.url.path
    if p.startswith("/api/") and p not in _OPEN_API:
        u = _current_user(request)
        # 例外（V2.443）：BP 后端同机调 BOM 消费口/导出——内部令牌对得上且来源回环才放过登录门，
        # 之后由路由自己再验一遍（导出只准已审核版）。此前漏了这条，令牌请求在这里就被 401「未登录」。
        bp_internal = p.startswith(bom_quote.INTERNAL_PATH_PREFIXES) and bom_quote.internal_token_ok(request)
        if not u and not (p in _PULL_PATHS and pull_token_ok(request)) and not bp_internal:
            return JSONResponse({"ok": False, "msg": "未登录"}, status_code=401)
        # 初始密码闸（V2.330）：账号被新建/重置密码后 must_change_pwd=1——改密之前除 /api/change-pwd
        # 外一律 403（含 /api/bp-authz，BP 也进不去）。前端据 code 弹强制改密页；服务端拦，直连 API 也绕不过。
        if u and u.get("must_change_pwd") and p != "/api/change-pwd":
            return JSONResponse({"ok": False, "code": "must_change_pwd",
                                 "msg": "首次登录（或密码被重置后）需先设置新密码"}, status_code=403)
    return await call_next(request)


@app.post("/api/login")
def api_login(body: dict, request: Request, response: Response):
    name = str(body.get("name", "") or "").strip()
    pwd = str(body.get("password", "") or "")
    u = db.verify_login(name, pwd)
    if not u:
        return JSONResponse({"ok": False, "msg": "姓名或密码错误，或账号已被禁用"}, status_code=401)
    tok = db.create_session(name)
    # cookie 名按端口隔离（V2.196，见 core.sid_name）——本机多实例并行时各端口互不顶号
    response.set_cookie(sid_name(request), tok, httponly=True, max_age=7 * 24 * 3600, samesite="lax")
    db.audit(name, "登录")
    return {"ok": True, "user": _user_public(u)}


@app.post("/api/logout")
def api_logout(request: Request, response: Response):
    for nm in {sid_name(request), "sid"}:
        tok = request.cookies.get(nm)
        if tok:
            db.delete_session(tok)
        response.delete_cookie(nm)
    return {"ok": True}


@app.get("/api/me")
def api_me(request: Request):
    u = _current_user(request)
    if not u:
        return JSONResponse({"ok": False}, status_code=401)
    return {"ok": True, "user": _user_public(u)}


@app.post("/api/change-pwd")
def api_change_pwd(body: dict, request: Request):
    """本人修改密码（V2.330）。新建/重置后 must_change_pwd=1，_auth_gate 会把其余接口全拦到这来。
    纪律：密码值不进任何日志/审计——audit 只记「谁改了密码」。"""
    u = _current_user(request)
    if not u:
        return JSONResponse({"ok": False, "msg": "未登录"}, status_code=401)
    ok, msg = db.change_own_pwd(u["name"], str(body.get("oldPwd", "") or ""), str(body.get("newPwd", "") or ""))
    if not ok:
        return {"ok": False, "msg": msg}
    db.audit(u["name"], "修改密码")
    return {"ok": True}


# ---------------- 账号管理（主管理员 / 工作台子管理员分级） ----------------
def _admin(request):
    """主管理员（role=admin，全权）。删除账号、任命子管理员等只给主管理员。"""
    u = _current_user(request)
    return u if db.is_super(u) else None


def _acct_admin(request):
    """能进账号管理的人：主管理员 或 工作台子管理员。"""
    u = _current_user(request)
    return u if db.can_admin_accounts(u) else None


@app.get("/api/users")
def api_users(request: Request):
    adm = _acct_admin(request)
    if not adm:
        return JSONResponse({"ok": False, "msg": "无账号管理权限"}, status_code=403)
    users = db.list_users()
    if not db.is_super(adm):                       # 子管理员只看自己管辖工作台内的普通账号
        users = [u for u in users if db.can_manage_user(adm, u)]
    return {"ok": True, "users": users, "scope": {
        "is_super": db.is_super(adm),
        "managed_ws": db.managed_workspaces(adm),
        "assignable": sorted(db.assignable_caps(adm)),                        # 可授予的权限点
        "manageable_grps": (None if db.is_super(adm) else sorted(db.manageable_grps(adm) or [])),
    }}


@app.post("/api/users/create")
def api_user_create(body: dict, request: Request):
    adm = _acct_admin(request)
    if not adm:
        return JSONResponse({"ok": False, "msg": "无账号管理权限"}, status_code=403)
    name = str(body.get("name", "") or "").strip()
    pwd = str(body.get("password", "") or "")
    grp = str(body.get("grp", "核算组") or "核算组")
    post = _norm_post(body.get("post"))     # 岗位：识别标签 + 一键套用模板的依据，本身不判权（D4）
    role = "admin" if body.get("role") == "admin" else "normal"
    perms = body.get("perms") if isinstance(body.get("perms"), dict) else None
    if not name or not pwd:
        return {"ok": False, "msg": "姓名和密码都要填"}
    if db.get_user(name):
        return {"ok": False, "msg": "该姓名已存在"}
    if not db.is_super(adm):
        # 子管理员：只能建普通用户、只能建到自己管辖的组、权限只能给自己可授范围
        role = "normal"
        grps = db.manageable_grps(adm) or set()
        if grp not in grps:
            return {"ok": False, "msg": f"你只能建 {('/'.join(sorted(grps)) or '（无）')} 的账号"}
        allowed = db.assignable_caps(adm)
        perms = {k: bool(v) for k, v in (perms or {}).items() if k in allowed}
    db.create_user(name, pwd, grp, role, perms=perms, post=post)
    db.audit(adm["name"], "建账号", name, f"{grp}/{post}/{role}")
    return {"ok": True}


@app.get("/api/bp-perm-drift")
def api_bp_perm_drift(request: Request):
    """BP 权限码表对账（V2.106）：拉 BP 的 /api/perms/registry 与本地 CAP_META 比对。
    给账号管理页提示「BP 新增了码但这里没登记」——漂移不会报错，只会静默锁人，故要显性提示。
    仅账号管理员可见；BP 不可达 → available=False（静默降级）。"""
    if not _acct_admin(request):
        return JSONResponse({"ok": False, "msg": "无权限"}, status_code=403)
    d = db.bp_registry_drift()
    if d is None:
        return {"ok": True, "available": False}
    return {"ok": True, "available": True, **d}


@app.get("/api/bp-authz")
def api_bp_authz(request: Request):
    """BP 反代鉴权探针（Nginx auth_request 调用）：校验 BP 准入并把该用户的 BP 权限码回吐到响应头，
    Nginx 用 auth_request_set 取 X-BP-User/X-BP-Perms 再注入到 /bp/ 上游请求。
    未登录→401(Nginx 弹登录)；登录但无 enter_bp→403(拒绝，堵掉"登录即进整个BP"窟窿)；通过→200+头。"""
    u = _current_user(request)
    if not u:
        return JSONResponse({"ok": False, "msg": "未登录"}, status_code=401)
    if not (db.is_super(u) or db.user_can(u, "enter_bp")):
        return JSONResponse({"ok": False, "msg": "无BP工作台准入"}, status_code=403)
    codes = db.bp_perm_codes(u)
    resp = JSONResponse({"ok": True})
    # HTTP 头须 latin-1：中文用户名 percent-encode（BP 侧仅作展示/审计，可解码）
    resp.headers["X-BP-User"] = urllib.parse.quote(str(u["name"]), safe="")
    resp.headers["X-BP-Perms"] = ",".join(codes)
    # V2.324（BP 侧 Owner 要求：工作台左下角显示"姓名 + 岗位"）：岗位只存在核算这边（users.post），
    #   BP 自身无登录、拿不到，故随准入探针一并透传。发**中文名**而非 key（BP 不该反查我们的岗位字典），
    #   同样 percent-encode（HTTP 头只认 latin-1，中文岗位不编码会 500）。
    #   ⚠ Nginx 需同步加一行 auth_request_set/proxy_set_header 转发 X-BP-Post，否则 BP 收不到。
    try:
        resp.headers["X-BP-Post"] = urllib.parse.quote(_post_label(str(u.get("post") or "")), safe="")
    except Exception:
        pass                      # 岗位只是展示信息，取不到不能影响 BP 准入
    return resp


@app.get("/api/perms/caps")
def api_perms_caps(request: Request):
    """权限能力清单（键+中文名+分组），给账号管理主从页按组渲染。附当前管理员的可授范围。"""
    u = _current_user(request)
    return {"ok": True, "caps": _caps_for_ui(),
            "scope": {"is_super": db.is_super(u),
                      "assignable": sorted(db.assignable_caps(u)),
                      "escalated": sorted(db.escalated_caps()),             # 主管理员加严的码（前端叠加高亮）
                      "manageable_grps": (None if db.is_super(u) else sorted(db.manageable_grps(u) or []))}}


@app.post("/api/perms/sensitivity/toggle")
def api_sensitivity_toggle(body: dict, request: Request):
    """主管理员把某常规权限单向升为敏感/降回常规（棘轮：代码定死项改不了）。"""
    adm = _admin(request)
    if not adm:
        return JSONResponse({"ok": False, "msg": "仅主管理员可调整敏感级别"}, status_code=403)
    cap = str(body.get("cap", "") or "")
    on = bool(body.get("on"))
    if not db.set_cap_escalated(cap, on, adm["name"]):
        return {"ok": False, "msg": "该权限不可调整（不存在/已是系统固定敏感/属准入或任命）"}
    db.audit(adm["name"], "敏感级别-" + ("升为敏感" if on else "降回常规"), cap)
    return {"ok": True, "escalated": sorted(db.escalated_caps())}


# ---------------- 门户管理（工具卡片 CMS） ----------------
# V2.331 状态自动联动：卡片挂了 mods（核算导航模块 key 列表）→ 状态由「系统设置 › 导航模块上线管理」
# 的六档进度推导，不再看手工档。六档 → 门户四档：
#   敬请期待→soon；开发中/测试验证→beta(开发中)；人工并行/待验收→par(人工并行)；引擎正常→ok(已上线闪绿灯)。
# 多模块合成＝**木桶取最低**（有一个模块还没到位，卡就不能亮更高档——不虚标）。
# 「隐藏」模块与树上已不存在的 key 不参与推导（后者以 autoMissing 露出，不静默）；
# 全部不可推导 → 回落手工档。BP 泳道无模块状态体系 → mods 空 = 维持手工（BP 联动记 backlog）。
_PORTAL_TIER = {"敬请期待": "soon", "开发中": "beta", "测试验证": "beta",
                "人工并行": "par", "待验收": "par", "引擎正常": "ok"}
_TIER_RANK = {"soon": 0, "beta": 1, "par": 2, "ok": 3}


def _portal_autolink(tools):
    state = _nav_state(False)
    labels = {m["key"]: m["label"] for m in _all_modules()}
    for t in tools:
        mods = t.get("mods") or []
        if not mods:
            t["statusSrc"] = "manual"
            continue
        tiers, detail, missing = [], [], []
        for k in mods:
            if k not in labels:
                missing.append(k)
                continue
            nav_st = (state.get(k) or {}).get("status")
            tier = _PORTAL_TIER.get(nav_st)
            if tier is None:                    # 「隐藏」等：不参与推导
                continue
            tiers.append(tier)
            detail.append({"key": k, "label": labels[k], "navStatus": nav_st, "tier": tier})
        t["autoDetail"] = detail
        if missing:
            t["autoMissing"] = missing
        if tiers:
            t["statusManual"] = t["status"]
            t["status"] = min(tiers, key=lambda x: _TIER_RANK[x])
            t["statusSrc"] = "auto"
        else:
            t["statusSrc"] = "manual"
    return tools


@app.get("/api/portal/tools")
def api_portal_tools(request: Request):
    """门户工具卡片列表（已登录即可读，供门户渲染 + 管理页）。V2.331 起状态经自动联动推导后下发。"""
    return {"ok": True, "tools": _portal_autolink(db.list_portal_tools())}


@app.post("/api/portal/tools/save")
def api_portal_tool_save(body: dict, request: Request):
    adm = _admin(request)
    if not adm:
        return JSONResponse({"ok": False, "msg": "需管理员"}, status_code=403)
    if not str(body.get("name", "") or "").strip():
        return {"ok": False, "msg": "工具名称不能为空"}
    tid = db.save_portal_tool(body)
    db.audit(adm["name"], "门户工具-保存", str(body.get("name", "")), str(body.get("lane", "")))
    return {"ok": True, "id": tid}


@app.post("/api/portal/tools/delete")
def api_portal_tool_delete(body: dict, request: Request):
    adm = _admin(request)
    if not adm:
        return JSONResponse({"ok": False, "msg": "需管理员"}, status_code=403)
    db.delete_portal_tool(body.get("id"))
    db.audit(adm["name"], "门户工具-删除", str(body.get("id", "")))
    return {"ok": True}


@app.post("/api/portal/tools/reset-defaults")
def api_portal_tools_reset(request: Request):
    """同步出厂工具集（V2.328）：同名覆盖、缺失补插、手工卡保留。人点按钮才跑，不在启动时静默改数据。"""
    adm = _admin(request)
    if not adm:
        return JSONResponse({"ok": False, "msg": "需管理员"}, status_code=403)
    r = db.reset_portal_tools_defaults()
    db.audit(adm["name"], "门户工具-同步出厂集",
             f"新增{len(r['added'])} 覆盖{len(r['updated'])} 保留{len(r['kept'])}")
    return {"ok": True, **r}


def _scoped_target(request, name):
    """取 (adm, target)；校验当前管理员能否管理 target。返回 (adm, target, err_resp)。"""
    adm = _acct_admin(request)
    if not adm:
        return None, None, JSONResponse({"ok": False, "msg": "无账号管理权限"}, status_code=403)
    u = db.get_user(name)
    if not u:
        return adm, None, JSONResponse({"ok": False, "msg": "账号不存在"}, status_code=404)
    if not db.can_manage_user(adm, u):
        return adm, u, JSONResponse({"ok": False, "msg": "你没有管理该账号的权限"}, status_code=403)
    return adm, u, None


@app.post("/api/users/perms")
def api_user_perms(body: dict, request: Request):
    name = str(body.get("name", "") or "").strip()
    adm, u, err = _scoped_target(request, name)
    if err:
        return err
    if u.get("role") == "admin":
        return {"ok": False, "msg": "管理员默认拥有全部权限，无需设置"}
    req = body.get("perms") if isinstance(body.get("perms"), dict) else {}
    if db.is_super(adm):
        perms = req
    else:
        # 子管理员：只允许改自己可授的权限点，其余(别工作台权限、manage_* 任命权)保留目标原值不动
        allowed = db.assignable_caps(adm)
        cur = db.parse_perms(u.get("perms"))
        perms = dict(cur)
        for k, v in req.items():
            if k in allowed:
                perms[k] = bool(v)
    # 进不去的菜单，底下的动作一律收回（业务方定）。前端已经这么显示了，但**真相以后端为准**——
    # 直接调接口也绕不过去，否则库里会攒下一堆屏幕上看不见的幽灵权限。
    before = dict(perms)
    perms = _cascade_revoke(perms)
    dropped = [k for k in before if before.get(k) and not perms.get(k)]
    db.set_user_perms(name, perms)
    db.audit(adm["name"], "改权限", name, ",".join(k for k, v in perms.items() if v))
    if dropped:
        db.audit(adm["name"], "级联收回动作权限", name, "（因无对应菜单准入）" + "、".join(dropped)[:260])
    return {"ok": True, "cascadeRevoked": dropped}


@app.post("/api/users/post")
def api_user_post(body: dict, request: Request):
    """设账号岗位；apply=True 时顺带按岗位模板一键套用权限（存量迁移的主要工具，确认书第六节）。

    套用 = **只勾上、不取消**。理由：模板是「批量勾选」的便利，不是权限本身（D4）——
    静默收回管理员手工开过的点，比少给更危险（人在月结中途突然没了权限，且没人知道为什么）。
    要收回请到权限清单里手动取消。
    敏感点永远不进模板（_template_caps 已滤），故套用不可能提权。"""
    name = str(body.get("name", "") or "").strip()
    adm, u, err = _scoped_target(request, name)
    if err:
        return err
    post = _norm_post(body.get("post"))
    if post and post not in _post_keys():
        return {"ok": False, "msg": "没有「%s」这个岗位，先去系统设置的岗位名单里加" % post}
    db.set_user_post(name, post)
    db.audit(adm["name"], "改岗位", name, _post_label(post) if post else "（清空）")

    granted = []
    if body.get("apply") and post:
        if u.get("role") == "admin":
            return {"ok": True, "post": post, "granted": [],
                    "msg": "主管理员本来就有全部权限，不用套用模板"}
        want = _template_caps(post)
        if not db.is_super(adm):
            want &= db.assignable_caps(adm)     # 子管理员只能套自己可授范围内的点
        cur = db.parse_perms(u.get("perms"))
        granted = sorted(c for c in want if not cur.get(c))
        if granted:
            db.set_user_perms(name, {**cur, **{c: True for c in granted}})
            db.audit(adm["name"], "套用岗位模板", name,
                     "%s：新增 %s" % (_post_label(post), "、".join(granted))[:300])
    return {"ok": True, "post": post, "granted": granted}


@app.post("/api/users/active")
def api_user_active(body: dict, request: Request):
    name = str(body.get("name", "") or "").strip()
    adm, u, err = _scoped_target(request, name)
    if err:
        return err
    active = bool(body.get("active"))
    db.set_user_active(name, active)
    if not active:
        db.delete_user_sessions(name)      # 禁用即踢下线
    db.audit(adm["name"], "启用账号" if active else "禁用账号", name)
    return {"ok": True}


@app.post("/api/users/reset-pwd")
def api_user_reset(body: dict, request: Request):
    name = str(body.get("name", "") or "").strip()
    adm, u, err = _scoped_target(request, name)
    if err:
        return err
    pwd = str(body.get("password", "") or "")
    if not pwd:
        return {"ok": False, "msg": "新密码不能为空"}
    db.reset_pwd(name, pwd)
    db.audit(adm["name"], "重置密码", name)
    return {"ok": True}


@app.post("/api/users/delete")
def api_user_delete(body: dict, request: Request):
    adm = _admin(request)                  # 删除账号（不可逆）仅【主管理员】
    if not adm:
        return JSONResponse({"ok": False, "msg": "删除账号仅主管理员可操作（子管理员请用禁用）"}, status_code=403)
    name = str(body.get("name", "") or "").strip()
    if name == adm["name"]:
        return {"ok": False, "msg": "不能删除自己"}
    db.delete_user(name)
    db.delete_user_sessions(name)
    db.audit(adm["name"], "删除账号", name)
    return {"ok": True}


# ---------------- 数据源：样例 / 金蝶（金蝶走"取数一次定格"缓存，不再每次直连）----------------
def _balance_rows():
    if CFG["source"] == "kingdee":
        return _kd_get("gl_balance")     # 读本期已存；没取过抛 KdNotFetched（页面转友好提示）
    return S.sample_balance_rows()


def _bank_accounts():
    if CFG["source"] == "kingdee":
        return kc.fetch_bank_accounts()
    return S.sample_kd_accounts_jun()


# ---------------- 通用 ----------------
@app.get("/api/health")
def health():
    return {"ok": True, "period": _period_str(), "source": CFG["source"],
            "conf": kc.conf_path() or "(未找到 conf.ini)", "db": db.backend_info()}


@app.get("/api/config")
def get_config():
    # 带上封存态 + 本期数据状态：前端各页据此显示徽标/胶囊，不必额外请求
    # 版本：并行开发后一台机器同时跑多条线，界面左下角自报「版本·分支·提交」（V2.176）
    return {**CFG, "period_str": _period_str(), "conf": kc.conf_path(),
            "封存": _closed_info(), "数据状态": _period_data_status(), "版本": VERSION_INFO}


@app.get("/api/period-statuses")
def period_statuses(year: int = 0, source: str = ""):
    """某年度 12 个月各自的数据状态（期间下拉逐月标注用）：已封存/数据已上传/数据未上传/样例数据。

    source 缺省＝全局账本（银行对账那一套，历史行为不变）；
    传 "cl:<账簿代码>" ＝成本台账某主体的账本——它按主体各记各的期间，
    不带这个参数就会把银行对账的月份状态显示到成本台账页上（看着还挺权威，实则张冠李戴）。"""
    y = int(year or CFG["year"])
    src = (source or "").strip()
    if src.startswith("cl:"):
        org = src[3:]
        # ⚠**函数内导入**：`_cl_period_data_status` 住在 routers/cost_ledger.py，
        #   拆 routers 那次这里的引用断了却没人发现——本函数一被调用就 NameError → 500，
        #   而前端 PeriodPicker 的 `.catch(() => setSt({}))` 把它吞成"没有状态"，
        #   于是成本台账右上角**十二个月一个胶囊都不显示**，看着像"这功能没做"（V2.299 才查出来）。
        #   放函数内而非文件顶部：那个模块 import 时会读配置/建连接，顶部导入会拖慢启动并引入耦合。
        from routers.cost_ledger import _cl_period_data_status
        return {"year": y, "source": src,
                "statuses": {str(m): _cl_period_data_status(y, m, org) for m in range(1, 13)}}
    if src == "logi":
        # 物流计提月份状态（V2.219，V2.221 加「已上传」态）：封存 > 已计提(录入台账有凭证) > 已上传(有账单批次) > 未计提
        posted = db.logistics_posted_periods(y)
        uploaded = db.bill_upload_periods(y)
        def _logi_st(m):
            if _is_closed(y, m):
                return "已封存"
            if posted.get(m):
                return "已计提"
            return "已上传" if uploaded.get(m) else "未计提"
        return {"year": y, "source": src,
                "statuses": {str(m): _logi_st(m) for m in range(1, 13)},
                "counts": {str(m): posted.get(m, 0) for m in range(1, 13)}}
    return {"year": y, "source": CFG["source"],
            "statuses": {str(m): _period_data_status(y, m) for m in range(1, 13)}}


@app.post("/api/config")
def set_config(body: dict, request: Request):
    # 数据源/会计期间是全局共享状态，改动影响所有人 → 限核算工作台准入者（挡住非核算账号）
    if not _require_perm(request, "enter_accounting"):
        return JSONResponse({"ok": False, "msg": "无权限修改数据源/会计期间（需核算工作台准入）"}, status_code=403)
    # 先全部校验、后一次性写入：非法值直接 400，绝不把坏值部分写进全局 CFG（否则 _period_str 之后全站 500）
    updates = {}
    if body.get("year") is not None:
        try:
            y = int(body["year"])
        except (TypeError, ValueError):
            return JSONResponse({"ok": False, "msg": "会计年度必须是数字"}, status_code=400)
        if not (2000 <= y <= 2100):
            return JSONResponse({"ok": False, "msg": "会计年度超出合理范围（2000–2100）"}, status_code=400)
        updates["year"] = y
    if body.get("period") is not None:
        try:
            pnum = int(body["period"])
        except (TypeError, ValueError):
            return JSONResponse({"ok": False, "msg": "会计期间必须是数字"}, status_code=400)
        if not (1 <= pnum <= 12):
            return JSONResponse({"ok": False, "msg": "会计期间必须在 1–12 之间"}, status_code=400)
        updates["period"] = pnum
    if body.get("source") is not None:
        if body["source"] not in ("sample", "kingdee"):
            return JSONResponse({"ok": False, "msg": "数据源只能是 sample 或 kingdee"}, status_code=400)
        updates["source"] = body["source"]
    if body.get("bank_import_dir") is not None:
        updates["bank_import_dir"] = body["bank_import_dir"]
    CFG.update(updates)     # 校验全过后一次性更新，避免部分写入污染全局态
    # 样例=演示数据，固定 2026 年 6 月：切到样例源即把期间锁到 6 月，杜绝"头上写5月、底下是6月假数据"的错位
    if CFG.get("source") == "sample":
        CFG["year"], CFG["period"] = SAMPLE_YM
    save_cfg(CFG)
    _cache_clear()          # 源/期间变了，缓存作废
    return {**CFG, "period_str": _period_str(), "封存": _closed_info(), "数据状态": _period_data_status(),
            "sample_locked": CFG["source"] == "sample"}


# ---------------- 导航模块上线管理（全员生效，存 app_settings，限主管理员）----------------
# 每个模块一个【状态】（V2.173 按需求方口径重排）：敬请期待 → 开发中 → 测试验证 → 人工并行 → 待验收 → 引擎正常。
# 「测试验证 / 人工并行 / 待验收 / 引擎正常」能进入（V2.174 需求方定：测试验证期也要进去看；
# 并行期天天在用、待验收得让人进去才验得了）；其余侧栏灰显、不可点。
# 「引擎正常」= 正式运行（侧栏显呼吸绿灯）；可挂【岗位】标签说明它服务谁——标签不是权限门，挡人仍用账号管理权限点。
# 旧状态平滑迁移：已上线→引擎正常、等待部署→测试验证（存量 DB 值读/存时自动映射，无需重设）。
#
# V2.52 导航与权限架构重构（确认书 20260717，D1–D12）：
#   · 一级升格为**实体**（NAV_SECTIONS，可增删改名排序），不再从模块的 grp 字符串反推
#   · 二级/三级挂 sec(一级key) + parent(父模块key) + order(排序)，均可在设置页改
#   · 内置模块的**位置可覆盖**（代码给默认值、DB 存覆盖值），否则这次搬完家以后照样搬不动
#   · **准入点跟菜单树自动生成**（kind="nav"，见 _nav_cap_meta），动作点仍由 db.CAP_META_STATIC 持有
#   · 岗位**key 化**：{key,label}，label 随便改、绑定不丢（D11）
NAV_STATUSES = ["敬请期待", "开发中", "测试验证", "人工并行", "待验收", "引擎正常", "隐藏"]
_NAV_LEGACY = {"已上线": "引擎正常", "等待部署": "测试验证"}   # 旧值（DB存量/代码default/旧前端缓存）→ 新值
# 「隐藏」＝从左侧导航整个移除（不进 itemsOf/childrenOf），只在系统设置本页可见可改回；不在 NAV_OPEN，故不可进入。
NAV_OPEN = ("测试验证", "人工并行", "待验收", "引擎正常")
# 「开发中」＝**只有指定的人进得去，其他人连菜单都点不动**（V2.242 业务方定）。
# 为什么要单独开这一档：在建的模块得有人能进去看效果，但半成品不该让同事撞见——
# 撞见了就会问"这个怎么用/怎么报错了"，白白产生一轮解释。
#
# 谁能进由 core.can_enter_dev() 定：conf.ini `[nav] dev_users` 有名单就只认名单，
# 留空则回落"所有主管理员"。名单放 conf.ini 不放 DB，是因为它要把范围收得**比主管理员还窄**——
# 能在页面上改的话，任何主管理员都能顺手把自己加回去，这个限制就等于没有。
#
# ⚠ 这仍是**上线状态、不是权限**：它只管左侧导航点不点得动，**不挡接口**。
#   真要挡人访问接口，仍然靠「账号管理」里的权限点（同 NAV_OPEN 那几档的既有约定）。
NAV_OPEN_ADMIN = ("开发中",)


def _nav_norm(st):
    return _NAV_LEGACY.get(st, st)

# ── 一级板块（实体）──
# bottom=True：钉在侧栏底部滚动区之外（基础数据/系统设置），不参与六大板块
NAV_SECTIONS = [
    {"key": "report", "label": "报表板块", "order": 10},
    {"key": "gl", "label": "总账板块", "order": 20},
    {"key": "ap", "label": "应付板块", "order": 30},
    {"key": "cost", "label": "成本模块", "order": 40},
    {"key": "ar", "label": "应收模块", "order": 50},
    {"key": "misc", "label": "其它模块", "order": 60},
    {"key": "common", "label": "通用", "order": 999, "bottom": True},
]
_SEC_KEYS = {s["key"] for s in NAV_SECTIONS}

# ── 二级 / 三级 ──
# group_only=True → 纯分组父项（点了展开，不直接进入，不设准入点）
# 有 parent 且父项非 group_only → 父项是「可进入且有子项」的第三种节点（如成本台账）
# cap=xxx → 复用一个**已存在的**动作点当准入点（不自动生成 enter:<key>）
# default = 未设过时的初始状态，如实反映当前进度
NAV_MODULES = [
    # ── 报表板块（V2.240 按业务方定的新结构重排）──
    # 两个二级都是**纯分组**：财务报表、源单列表本身不进页面，内容全在三级里（业务方定）。
    # 撤下的旧三级：科目余额(sbal)、序时账簿(journal)——业务方定「直接删掉」，两个 key 就此退出菜单树，
    #   其自动生成的 enter:sbal / enter:journal 准入点随之消失（存量账号上那两个键成为死键，不再渲染）。
    #   ⚠ 科目余额**不是空占位**：SubjectBalance.jsx 与 /api/subject-balance* 是在跑的真页面，
    #     本次只把它从菜单摘下（页面代码与接口原样留着、不可达），未删文件、未删 API。
    #     要不要连页面带接口一起删，属治理红线「删文件/删既有 API」，须先出影响分析再动。
    {"key": "fiacc", "label": "财务报表", "sec": "report", "order": 10, "default": "已上线", "group_only": True},
    {"key": "rptdash", "label": "报表仪表盘", "sec": "report", "order": 11, "parent": "fiacc", "default": "待验收"},
    # 报表导出/源单导出页内各两个页签：①一键导出 ②通知设置（业务方定放页面内，不做四级菜单）
    {"key": "rptexport", "label": "报表导出", "sec": "report", "order": 12, "parent": "fiacc", "default": "待验收"},
    # srcbill 由「二级直接进入」改为纯分组（业务方定）：它的 enter:srcbill 准入点因此消失，
    # 准入闸落到三级 enter:srcexport 上——与 fiacc 一族的处理一致。
    {"key": "srcbill", "label": "源单列表", "sec": "report", "order": 20, "default": "已上线", "group_only": True},
    {"key": "srcexport", "label": "源单导出", "sec": "report", "order": 21, "parent": "srcbill", "default": "敬请期待"},
    # ── 总账板块 ──
    {"key": "bankrecon", "label": "银行对账", "sec": "gl", "order": 10, "default": "已上线", "group_only": True},
    {"key": "reconcile", "label": "对账程序", "sec": "gl", "order": 11, "parent": "bankrecon", "default": "已上线"},
    {"key": "fxrate", "label": "汇率录入", "sec": "gl", "order": 12, "parent": "bankrecon", "default": "待验收"},
    {"key": "wealth", "label": "理财对账", "sec": "gl", "order": 13, "parent": "bankrecon", "default": "已上线"},
    {"key": "fundboard", "label": "资金看板", "sec": "gl", "order": 14, "parent": "bankrecon", "default": "已上线"},
    {"key": "ledger", "label": "账户台账", "sec": "gl", "order": 15, "parent": "bankrecon", "default": "已上线"},
    # 月结看板【暂留】（确认书 D10）：它是封存/解封的唯一入口，而封存拦截遍布全台（_closed_block）。
    # 现在摘掉 = 某期一旦封存就永远解不开。等「封存拆到各二级」验收后再下线，中间不留空窗。
    {"key": "periodclose", "label": "月结看板", "sec": "gl", "order": 20, "default": "待验收"},
    # ── 应付板块 ──
    {"key": "logisticsrecon", "label": "物流对账", "sec": "ap", "order": 10, "default": "已上线", "group_only": True},
    # V2.198 业务方定的二级顺序：基础数据 → 物流计提 → 账单核对 → 单据运费。
    # 基础数据=物流线共用维表（供应商列表/费用归属映射/业务线/标注翻译/税率），从物流计提步②搬出独立成页。
    {"key": "logibase", "label": "基础数据", "sec": "ap", "order": 10, "parent": "logisticsrecon", "default": "已上线"},
    # V2.226 业务方定：账单上传=物流部专属工作台（传账单/长表→质检→补维度→提交给核算组），
    # 与核算组的「物流计提」（复核/去向费率/录金蝶）按角色隔离——物流部账号只开这个菜单即可干完他们的活。
    {"key": "logiupload", "label": "账单上传", "sec": "ap", "order": 11, "parent": "logisticsrecon", "default": "已上线"},
    {"key": "logistics", "label": "物流计提", "sec": "ap", "order": 12, "parent": "logisticsrecon", "default": "已上线"},
    # 下面两个只改 label（付款对账→账单核对、单据物流成本→单据运费），**key 不动**：
    # 数据/权限/审计历史全绑在 key 上，改 key 等于搬家，改 label 零迁移（确认书 D6/D7）
    {"key": "logisticspay", "label": "账单核对", "sec": "ap", "order": 13, "parent": "logisticsrecon", "default": "待验收"},
    {"key": "logisticscost", "label": "单据运费", "sec": "ap", "order": 14, "parent": "logisticsrecon", "default": "待验收"},
    # ── 成本模块 ──
    # V2.254（业务方定）：成本台账 → **存货台账**，并由「第三种节点」改回**纯分组 + 三个三级**：
    #   台账导出 clexport（原来挂在二级上的八步工作流页，一行代码没动）／存货看板 cldash（新）／
    #   基础资料 clwh（原「仓库类型」，改名后要装的不止仓库类型，还有类别↔科目对照）。
    #   **这条调整 D8**：D8 当年把 clrecon 那层取消、内容上提，理由是"只有一个子项，多一层没意义"；
    #   现在有三个并列的三级，分组这层重新成立。不是把 D8 改回去，是前提变了。
    #   ⚠ 二级 key `costledger` **不动**（D6/D7：数据/权限/审计历史全绑在 key 上，改 label 零迁移）。
    #   ⚠ 纯分组按原逻辑不设闸、子项各长各的闸 → 会把 V2.142「一组一个闸」推翻。
    #      故这里给分组**显式声明共用闸** `cap="enter:costledger"`（配合 _enter_cap V2.254 的改动），
    #      三个三级全部继承它；进组后各页见不见，仍由各自的 act_cap 决定。
    #
    # ⚠ 这两个**不复用** cost_ledger / cost_ledger_wh 当准入点（确认书第四节原本这么定，实测是错的）：
    #   那俩是【动作点】，非敏感 → default_perms 会把它们自动发给每个核算组账号（db.py:default_perms）。
    #   复用＝这两个菜单人人自动可见，业务方定的「存量全部不给」(D5) 在此当场失效——端到端实测抓到。
    #   改用自动生成的 enter:costledger（kind="nav"，默认不给）；
    #   cost_ledger / cost_ledger_wh 退回本职，继续当页面里的动作点用。
    #   enter_settings 仍可复用——它是**敏感点**，本来就默认不给，不存在这个问题。
    {"key": "costledger", "label": "存货台账", "sec": "cost", "order": 10, "default": "待验收",
     "group_only": True, "cap": "enter:costledger"},
    {"key": "clexport", "label": "台账导出", "sec": "cost", "order": 11, "parent": "costledger",
     "act_cap": "cost_ledger", "default": "待验收"},
    {"key": "cldash", "label": "存货看板", "sec": "cost", "order": 12, "parent": "costledger",
     "act_cap": "cost_ledger", "default": "开发中"},
    # V2.142：不再有自己的准入点（enter:clwh 取消）——「存货台账这一组共用一个闸」（业务方定）：
    # 准入闸只有 enter:costledger 一个；进组后基础资料见不见，看 act_cap=cost_ledger_wh。
    {"key": "clwh", "label": "基础资料", "sec": "cost", "order": 13, "parent": "costledger",
     "act_cap": "cost_ledger_wh", "default": "待验收"},
    # BOM报价审核（确认书 v1.0，2026-09-03）：二级分组 + 两个三级。分组不设共用闸 → 两子各自生成准入点
    #   （可见性口径不同：待办与复核=未审核只权限人看、标准成本台账=已审核公开）。
    #   待办与复核=未审核工作台（钉钉抓取/入账/复核/定稿/价格校验，enter:bomdraft 敏感、只给成本会计/Owner）；
    #   标准成本台账=已审核成品库（只放已定稿、公开可查、供 BP 消费，enter:bomstd 广授）。
    {"key": "bomprice", "label": "BOM报价审核", "sec": "cost", "order": 20, "default": "待验收", "group_only": True},
    {"key": "bomdraft", "label": "待办与复核", "sec": "cost", "order": 21, "parent": "bomprice", "default": "待验收"},
    {"key": "bomstd", "label": "标准成本台账", "sec": "cost", "order": 22, "parent": "bomprice", "default": "待验收"},
    {"key": "bomconfig", "label": "基础设置", "sec": "cost", "order": 23, "parent": "bomprice", "default": "待验收"},
    {"key": "prodbrief", "label": "生产简报复核", "sec": "cost", "order": 30, "default": "敬请期待"},
    # V2.318：临时工考勤升为纯分组父项（group_only＝本身没页面、没准入点），下挂两个三级：
    #   复核工具   tempattrev   —— 上报工时 vs 打卡，逐日重算与四档判定
    #   临时工看板 tempattboard —— 全年用工结构（车间/月度/夜班占比/派遣方）
    # 拆 key 而不是复用 tempatt 当其中一页：key 是权限与审计的锚，一个 key 只能对一个页面。
    # 此前 tempatt 一直是「敬请期待」占位、无任何存量权限与数据绑定，故此时改造零迁移成本。
    {"key": "tempatt", "label": "临时工考勤", "sec": "cost", "order": 40, "default": "待验收", "group_only": True},
    {"key": "tempattrev", "label": "复核工具", "sec": "cost", "order": 41, "parent": "tempatt", "default": "待验收"},
    {"key": "tempattboard", "label": "临时工看板", "sec": "cost", "order": 42, "parent": "tempatt", "default": "待验收"},
    # ── 应收模块 ──
    {"key": "revledger", "label": "收入台账", "sec": "ar", "order": 10, "default": "敬请期待"},
    {"key": "custrecon", "label": "客户对账", "sec": "ar", "order": 20, "default": "敬请期待"},
    {"key": "ecompromo", "label": "电商推广", "sec": "ar", "order": 30, "default": "敬请期待"},
    # 电商对账（V2.250 条目⑤一期）：ecom 升级为分组父项，**key 不动**（同 D6/D7 教训：数据/权限/
    # 审计全绑 key，改 label/结构零迁移）。三级=收款核销+基础资料；发货核对（条目①）后续第三个三级。
    {"key": "ecom", "label": "电商对账", "sec": "ar", "order": 40, "default": "待验收", "group_only": True},
    {"key": "ecomsettle", "label": "收款核销", "sec": "ar", "order": 41, "parent": "ecom", "default": "待验收"},
    {"key": "ecombase", "label": "基础资料", "sec": "ar", "order": 42, "parent": "ecom", "default": "待验收"},
    # ── 其它模块 ──
    {"key": "archive", "label": "凭证归档", "sec": "misc", "order": 10, "default": "待验收"},
    # ── 通用（钉底部）──
    # 基础数据/基础设置=平台基础设施（配主体档案、数据源、金蝶连接），恒常可用，不参与上线开关（always）
    {"key": "basicdata", "label": "基础数据", "sec": "common", "order": 10, "default": "已上线", "always": True},
    {"key": "settings", "label": "系统设置", "sec": "common", "order": 20, "default": "已上线", "always": True, "cap": "enter_settings"},
]
_NAV_KEY = "nav_modules"                  # {key: {status, posts}}
_SEC_OVR_KEY = "nav_section_overrides"    # {sec_key: {label?, order?}}
_USEC_KEY = "nav_user_sections"           # [{key,label,order}] 自建一级
_MOD_OVR_KEY = "nav_module_overrides"     # {key: {sec?, parent?, order?, label?}} 内置模块位置覆盖

# ── 岗位（key 化，D11）──
# 拿名字当主键的老做法：改名＝删旧+建新，挂在模块上的标签会被读取过滤静默丢弃（旧 app.py:817）。
# 岗位升级成模板、账号又要绑岗位之后，改一次名会把模板绑定和账号绑定一起弄丢。故 key 稳定、label 随便改。
NAV_POSTS_DEFAULT = [
    {"key": "fin_mgr", "label": "财务经理"},
    {"key": "gl_acc", "label": "总账会计"},
    {"key": "cost_acc", "label": "成本会计"},
    {"key": "ap_acc", "label": "应付会计"},
    {"key": "ar_acc", "label": "应收会计"},
    {"key": "tax_acc", "label": "税务会计"},
    {"key": "fund_spec", "label": "资金专员"},
    {"key": "intern", "label": "实习生"},
]
_POSTS_KEY = "nav_posts"
NAV_POSTS_MAX = 20
# 老岗位名单（字符串）→ 新 key。用于把存量模块上挂的岗位标签平移过来，不丢。
_POST_ALIAS = {"成本岗": "cost_acc", "应收岗": "ar_acc", "总账岗": "gl_acc",
               "资金岗": "fund_spec", "财务经理": "fin_mgr",
               # 账号 post 字段是手打的自由文本，常见写法一并认（确认书 Q4）
               "总账会计": "gl_acc", "总账": "gl_acc", "总帐": "gl_acc", "总帐会计": "gl_acc",
               "成本会计": "cost_acc", "成本": "cost_acc",
               "应付会计": "ap_acc", "应付": "ap_acc",
               "应收会计": "ar_acc", "应收": "ar_acc",
               "税务会计": "tax_acc", "税务": "tax_acc",
               "资金专员": "fund_spec", "资金": "fund_spec", "出纳": "fund_spec",
               "实习生": "intern", "实习": "intern"}

# ── 岗位模板（D4：岗位不判权，只是「一键套用」的模板）──
# 模板 = 准入点 + 非敏感动作点。**敏感点一律不进模板**，永远手工显式给，否则地板被绕过。
# 按【板块/模块】写而不是按权限码写：菜单树一变，模板自动跟上，不用回来改代码。
# secs=整个板块全给；mods=额外单点给；acts=非敏感动作点（显式列，看得见给了什么）。
# ⚠ 内容是确认书第五节的**草案**（Q1 税务会计无专属工具 / Q2 资金专员vs总账会计分工 / Q3 实习生只读范围
#   均待业务方确认）。主管理员可在设置页调，调完存 DB，代码这份只当种子。
_TPL_KEY = "nav_post_templates"
NAV_POST_TEMPLATES_DEFAULT = {
    # secs 带上 common＝把「基础数据」给他；「系统设置」同在 common，但它的 cap 是敏感点 enter_settings，
    # 会被 _template_caps 挡掉——模板永远给不出敏感点，地板不破（确认书第五节）。
    "fin_mgr": {"secs": ["report", "gl", "ap", "cost", "ar", "misc", "common"],
                "acts": ["bank_upload", "kingdee_refresh", "claim", "ledger_override", "subject_upload",
                         "logistics_upload", "cost_ledger", "cost_ledger_wh", "archive_edit"]},
    # 总账会计 / 资金专员在【银行对账】内部的分工，照业务方**已有的岗位标注**来（V2.51 前就挂在模块上）：
    #   对账程序＋理财对账 = 总账岗    资金看板＋账户台账 = 资金岗
    # 汇率录入/月结看板是新增的、没标注过，暂归总账会计（确认书 Q2 待业务方确认）。
    "gl_acc": {"secs": ["report", "misc"], "mods": ["reconcile", "wealth", "fxrate", "periodclose"],
               "acts": ["bank_upload", "kingdee_refresh", "claim", "subject_upload", "archive_edit"]},
    "fund_spec": {"mods": ["fundboard", "ledger"], "acts": ["ledger_override"]},
    # V2.240 报表板块改名后按语义平移：科目余额+序时账簿 → 报表仪表盘；源单列表(现纯分组) → 源单导出。
    # 「报表导出」暂不进这三个模板（一键导出是新动作，给谁待业务方定）；gl_acc/tax_acc 走 secs=["report"] 整给，自动含它。
    "cost_acc": {"secs": ["cost"], "mods": ["rptdash"], "acts": ["cost_ledger", "cost_ledger_wh"]},
    "ap_acc": {"secs": ["ap"], "mods": ["rptdash", "srcexport"], "acts": ["logistics_upload"]},
    "ar_acc": {"secs": ["ar"], "mods": ["rptdash", "srcexport"], "acts": []},
    # 税务会计**不是没有工具**：DB 里早有个自建占位 tax「进项税核对」（原挂"其它小工具"）。
    # 确认书 Q1 据此修正——问题不是"缺工具"，而是"这个工具该归哪个板块"，待业务方定。
    "tax_acc": {"secs": ["report"], "mods": ["tax"], "acts": []},
    "intern": {"mods": [], "acts": []},                                 # Q3 待确认：给哪些菜单的只读
}


def _nav_can_edit(u):
    return bool(u and db.is_super(u))          # 只有主管理员能改


_KEY_RE = re.compile(r"^[a-z][a-z0-9_]{1,23}$")   # key 要当路由/状态键用，限小写字母数字下划线
_UMOD_KEY = "nav_user_modules"     # 自建模块（主管理员在设置页加的「规划中」占位，不绑代码）
_BUILTIN_KEYS = {m["key"] for m in NAV_MODULES}


# ================= 进程内缓存 =================
# caps() 会被 parse_perms 高频调用（list_users 里每个用户一次），而菜单树在 DB 里——
# 裸调 get_setting 等于每次判权都打库。故缓存 + 变更时显式失效。
# 与本模块既有的 _SYNC_AT/_*_CACHE 一样假设**单进程**（uvicorn 单 worker）。
_NAV_CACHE = {}


def _nav_cache_clear():
    _NAV_CACHE.clear()


def _nav_cached(k, fn):
    if k not in _NAV_CACHE:
        _NAV_CACHE[k] = fn()
    return _NAV_CACHE[k]


# ================= 岗位 =================
def _clean_posts(raw):
    """岗位名单 [{key,label}] 去重去空、限长限量；保持录入顺序。
    兼容老格式（纯字符串名单）：按 _POST_ALIAS 平移成 key，认不出的用 p<n> 兜底、label 保留原文。"""
    out, seen = [], set()
    for i, p in enumerate(raw or []):
        if isinstance(p, str):                       # 老格式：字符串岗位名
            label = p.strip()[:10]
            if not label:
                continue
            key = _POST_ALIAS.get(label) or ("p%d" % (i + 1))
        elif isinstance(p, dict):
            key = str(p.get("key") or "").strip()[:24]
            label = str(p.get("label") or "").strip()[:10]
            if not key or not label:
                continue
        else:
            continue
        if key in seen:
            continue
        seen.add(key)
        out.append({"key": key, "label": label})
    return out[:NAV_POSTS_MAX]


def _nav_posts():
    """岗位名单（主管理员可增删改名）。没设过就用种子。"""
    return _nav_cached("posts", lambda: (_clean_posts(db.get_setting(_POSTS_KEY, None))
                                         or [dict(p) for p in NAV_POSTS_DEFAULT]))


def _post_keys():
    return {p["key"] for p in _nav_posts()}


def _post_label(key):
    return next((p["label"] for p in _nav_posts() if p["key"] == key), key)


def _norm_post(raw):
    """账号 post 字段 → 岗位 key。存量是手打的自由文本，按别名表尽力认；认不出原样留着，
    前端显示成「待认领」由管理员在下拉里选（确认书 Q4）——不静默清空，不丢信息。"""
    s = str(raw or "").strip()
    if not s:
        return ""
    if s in _post_keys():
        return s
    return _POST_ALIAS.get(s, s)


# ================= 一级板块 =================
def _user_sections():
    raw = db.get_setting(_USEC_KEY, None) or []
    out, seen = [], set(_SEC_KEYS)
    for s in raw:
        if not isinstance(s, dict):
            continue
        k = str(s.get("key") or "").strip()
        if not _KEY_RE.match(k) or k in seen:
            continue
        seen.add(k)
        out.append({"key": k, "label": str(s.get("label") or k).strip()[:20],
                    "order": int(s.get("order") or 500), "builtin": False})
    return out


def _all_sections():
    """内置 + 自建，套上覆盖值，按 order 排。bottom 的永远最后。"""
    def build():
        ovr = db.get_setting(_SEC_OVR_KEY, {}) or {}
        out = []
        for s in NAV_SECTIONS:
            o = ovr.get(s["key"]) if isinstance(ovr.get(s["key"]), dict) else {}
            out.append({**s, "builtin": True,
                        "label": str(o.get("label") or s["label"])[:20],
                        "order": int(o.get("order", s["order"]))})
        out += _user_sections()
        return sorted(out, key=lambda s: (bool(s.get("bottom")), s["order"], s["label"]))
    return _nav_cached("sections", build)


# ================= 模块 =================
def _user_modules():
    """[{key,label,sec,parent?,order}]。过滤掉与内置撞 key 的、非法的、指向不存在一级的。"""
    raw = db.get_setting(_UMOD_KEY, None) or []
    sec_keys = {s["key"] for s in _all_sections()}
    out, seen = [], set(_BUILTIN_KEYS)
    for m in raw:
        if not isinstance(m, dict):
            continue
        k = str(m.get("key") or "").strip()
        if not _KEY_RE.match(k) or k in seen:
            continue
        seen.add(k)
        # V2.52 迁移：老自建模块存的是 grp(一级中文名)，新结构用 sec(一级 key)。
        # 老的四个一级（月结与结账/报表与台账/常规对账/其它小工具）已撤销，认不出的一律归「其它模块」，
        # 免得模块指着一个不存在的一级、在侧栏里凭空消失。
        sec = str(m.get("sec") or "").strip()
        if sec not in sec_keys:
            sec = "misc"
        out.append({"key": k, "label": str(m.get("label") or k).strip()[:20], "sec": sec,
                    "parent": str(m.get("parent") or "").strip() or None,
                    "order": int(m.get("order") or 500)})
    return out


def _all_modules():
    """内置 + 用户自建，套上位置覆盖，按 (sec, order) 排。builtin=True 的由代码持有、不可删。"""
    def build():
        ovr = db.get_setting(_MOD_OVR_KEY, {}) or {}
        sec_keys = {s["key"] for s in _all_sections()}
        mods = []
        for m in NAV_MODULES:
            o = ovr.get(m["key"]) if isinstance(ovr.get(m["key"]), dict) else {}
            sec = str(o.get("sec") or m["sec"])
            mods.append({**m, "builtin": True,
                         "label": str(o.get("label") or m["label"])[:20],
                         "sec": sec if sec in sec_keys else m["sec"],
                         "parent": o.get("parent", m.get("parent")) or None,
                         "order": int(o.get("order", m.get("order", 500)))})
        for m in _user_modules():
            mods.append({**m, "default": "开发中", "builtin": False})
        # 父项被挪走/删了 → 子项别成孤儿（会整棵消失）：认不出的 parent 一律置空，降级成二级
        keys = {m["key"] for m in mods}
        by_key = {m["key"]: m for m in mods}
        for m in mods:
            if m.get("parent") and m["parent"] not in keys:
                m["parent"] = None
            m["cap"] = _enter_cap(m, by_key)   # 准入点算好一起发：前端别自己拼 "enter:"+key，拼错＝静默放行
            # act ＝ 组内第二道门（V2.142）：第三种节点的子项，进组后见不见还要看这个动作点。
            # 前端 canEnter = 授了 cap AND（无 act 或授了 act）。
            m["act"] = m.get("act_cap") or None
        return sorted(mods, key=lambda m: (m["sec"], m["order"], m["label"]))
    return _nav_cached("modules", build)


def _mod_by_key(k):
    return next((m for m in _all_modules() if m["key"] == k), None)


# ================= 准入点自动生成（kind="nav"）=================
def _enter_cap(m, by_key=None):
    """该菜单的准入点码。纯分组父项不设（它没页面，可见性由子项决定）；
    声明了 cap 的复用现成点（系统设置=enter_settings，敏感、默认不给，可复用）。

    V2.142（业务方定）：**准入闸设在二级、不下沉到三级**——「第三种节点」（可进入且有子项，
    如成本台账）的子项【不生成自己的 enter:】，准入跟父项那一个闸走；子项在组内是否可见，
    由它声明的 act_cap（动作点）决定（见 _all_modules 的 act 字段）。
    「成本台账+仓库类型是一组」：勾一个「进入 成本台账」即可，仓库类型见不见看「维护仓库类型」动作权限。
    ⚠这不与确认书第四节的教训冲突：当年的坑是把非敏感动作点**直接**当准入闸（自动发放→人人可见）；
    现在动作点只做**组闸之内**的第二道门——没有组闸的人整组不可见，D5「存量全部不给」仍然成立。
    group_only 组（银行对账等）的子项维持各自准入不变——那些是并列工具页，本就要分岗授权。

    V2.254：**纯分组也可以声明一个共用闸**（`group_only=True` + `cap="enter:<自己key>"`）。
    起因是存货台账由「可进入且有子项」改成纯分组（三级：台账导出/存货看板/基础资料）——
    照原逻辑纯分组不设闸、子项各长各的闸，等于把 V2.142「一组一个闸」当场推翻。
    改法是把 `cap` 的判断提到 `group_only` 之前，并让子项**按父项算出来的闸**继承（父项没闸才自立门户）。
    对现有 group_only 组零影响：银行对账等都没声明 cap → 仍返回 None → 子项照旧各设准入。"""
    if m.get("cap"):
        return m["cap"]
    if m.get("group_only"):
        return None
    p = by_key.get(m.get("parent")) if (by_key and m.get("parent")) else None
    if p:
        pc = _enter_cap(p, by_key)           # 父项有闸就跟父闸（第三种节点／声明了共用闸的分组）
        if pc:
            return pc
    return "enter:" + m["key"]


def _nav_cap_meta():
    """菜单树派生的准入点，注册给 db.caps()。**只产自动生成的那些**——
    声明了 cap 的复用已在 CAP_META_STATIC 里的码，再造一条就重了。"""
    def build():
        sec_label = {s["key"]: s["label"] for s in _all_sections()}
        out = []
        for m in _all_modules():
            cap = m.get("cap")
            if not cap or not cap.startswith("enter:"):
                continue
            # V2.142：只登记【自己的】闸。子项继承父闸（cap != enter:<自己key>）时跳过——
            # 否则同一个 enter:costledger 会被注册两次、第二条还顶着「进入 仓库类型」的错标签。
            if cap != "enter:" + m["key"]:
                continue
            out.append({"key": cap, "label": "进入 " + m["label"], "ws": "accounting",
                        "group": sec_label.get(m["sec"], "其它模块"), "kind": "nav", "tier": "nav",
                        "mod": m["key"]})
        return out
    return _nav_cached("caps", build)


def _caps_for_ui():
    """账号页三栏用的权限点清单：在 cap_meta() 基础上补两样前端算不出来的东西——
      gate    ＝ 这个动作归哪个菜单的准入点管（第③栏「只显示进得去的」靠它过滤、后端级联收回也靠它）
      modLabel／secLabel ＝ 该动作所属菜单/板块的中文名（第③栏按菜单归堆、并标出处）
    **别让前端自己拼 "enter:"+mod**：拼错＝该挡的没挡住，而且是静默的。"""
    mods = {m["key"]: m for m in _all_modules()}
    secs = _all_sections()
    sec_label = {s["key"]: s["label"] for s in secs}
    sec_ix = {s["key"]: i for i, s in enumerate(secs)}          # 板块的**展示顺序**（与侧栏一致）
    all_caps = {c["key"]: c for c in db.cap_meta()}
    out = []
    for c in db.cap_meta():
        c = dict(c)
        m = mods.get(c.get("mod")) if c.get("mod") else None
        if c.get("tier") == "act" and m:
            c["gate"] = _enter_cap(m)                  # 纯分组父项没准入点 → None → 该动作不受菜单门控
            c["modLabel"] = m["label"]
            c["secLabel"] = sec_label.get(m["sec"], "")
        elif c.get("tier") == "act" and c.get("mod") in all_caps:
            # BP 的菜单不在核算菜单树里，它的「板块码」本身就是准入点 → mod 即 gate
            g = all_caps[c["mod"]]
            c["gate"] = g["key"]
            c["modLabel"] = g["label"]
            c["secLabel"] = g.get("group", "")
        if c.get("tier") == "nav" and m:
            c["group"] = sec_label.get(m["sec"], c.get("group") or "其它模块")
            c["_ix"] = (sec_ix.get(m["sec"], 99), m.get("order", 500), m["label"])
        out.append(c)
    # 第②栏按【板块的展示顺序】出，别让「系统设置」这类静态码因为在注册表里排得早就窜到最前面
    nav = sorted([c for c in out if c.get("tier") == "nav"], key=lambda c: c.get("_ix", (99, 999, "")))
    for c in nav:
        c.pop("_ix", None)
    return ([c for c in out if c.get("kind") in ("enter", "manage")] + nav
            + [c for c in out if c.get("tier") == "act"])


def _cascade_revoke(perms):
    """准入点没给 → 它底下的动作点一律置 False（业务方定：跟着一起收回）。
    为什么必须后端做：屏幕上看到的就得是库里存的。留着「看不见却仍然授予」的幽灵权限，
    管理员既审计不到、也想不起何时给过；等哪天重新给了准入，旧动作权限会静悄悄自己回来。
    V2.323 第二级：动作点也可挂 parent（如 发送周报←周报查看、导出Excel←项目视图）——
    父动作没给，子动作同收（先跑菜单级、再跑 parent 级，父项被菜单级收掉时子项同轮跟上）。"""
    ui = _caps_for_ui()
    out = dict(perms)
    for c in ui:
        g = c.get("gate")
        if c.get("tier") == "act" and g and not out.get(g):
            out[c["key"]] = False
    for c in ui:
        p = c.get("parent")
        if c.get("tier") == "act" and p and not out.get(p):
            out[c["key"]] = False
    return out


db.set_nav_caps_provider(_nav_cap_meta)


def _migrate_nav_posts_once():
    """V2.52 一次性迁移：岗位名单 老格式(字符串) → key 化 [{key,label}]（确认书 D11）。

    为什么非迁不可：老名单存的是「成本岗/应收岗/总账岗/资金岗/财务经理」这套旧命名。因为 _nav_posts()
    是「DB 设过就不用种子」，不迁的话业务方新定的 8 个岗位**永远不会生效**——平移只救得回 5 个，
    应付会计/税务会计/实习生 三个新岗位根本冒不出来。
    key 按别名表平移（模块上挂的标签因此不丢）、label 用种子里业务方给的新名字（成本岗→成本会计）。
    认不出的自定义岗位原样保留。幂等：已是新格式就不动。"""
    raw = db.get_setting(_POSTS_KEY, None)
    if not isinstance(raw, list) or not raw or not any(isinstance(p, str) for p in raw):
        return
    old = _clean_posts(raw)                       # 字符串 → key（别名表）
    seed = {p["key"]: p["label"] for p in NAV_POSTS_DEFAULT}
    have = {p["key"] for p in old}
    merged = [dict(p) for p in NAV_POSTS_DEFAULT]                       # 种子 8 个，按业务方给的顺序
    merged += [p for p in old if p["key"] not in seed]                  # 自定义岗位接在后面，不丢
    merged = _clean_posts(merged)
    db.set_setting(_POSTS_KEY, merged, "系统迁移")
    _nav_cache_clear()
    db.audit("系统迁移", "岗位名单key化", "核算工作台",
             "平移 %d 个、新增 %d 个 → %s"
             % (len(have & set(seed)), len([p for p in merged if p["key"] not in have]),
                "、".join(p["label"] for p in merged))[:300])


_migrate_nav_posts_once()


def _nav_one(m, saved, post_keys, dev_ok=False):
    s = saved.get(m["key"])
    if isinstance(s, bool):                    # V2.63 老格式(True/False 开关) → 平滑迁移到状态
        s = {"status": "引擎正常" if s else "敬请期待"}
    if not isinstance(s, dict):
        s = {}
    raw = _nav_norm(s.get("status"))           # V2.173：旧值（已上线/等待部署）读时平移
    st = raw if raw in NAV_STATUSES else _nav_norm(m.get("default", "引擎正常"))
    if m.get("always"):
        st = "引擎正常"
    # 只有「引擎正常」能挂岗位；岗位被删掉后，挂在模块上的那条自动消失（读时按名单过滤）。
    # V2.52 起挂的是岗位 key，老数据里的中文名按别名表平移（改名不再丢标签，见 D11）。
    posts = []
    if st == "引擎正常":
        for p in (s.get("posts") or []):
            k = _POST_ALIAS.get(p, p) if isinstance(p, str) else None
            if k in post_keys and k not in posts:
                posts.append(k)
    # 「开发中」只对名单里的人放行；别人拿到的是可进入=false，菜单点不动
    ok = st in NAV_OPEN or (dev_ok and st in NAV_OPEN_ADMIN)
    return {"status": st, "posts": posts, "可进入": ok, "仅开发者": st in NAV_OPEN_ADMIN}


def _nav_state(dev_ok=False):
    saved = db.get_setting(_NAV_KEY, {}) or {}
    pk = _post_keys()
    return {m["key"]: _nav_one(m, saved, pk, dev_ok) for m in _all_modules()}


# ================= 岗位模板 =================
def _post_templates():
    def build():
        t = db.get_setting(_TPL_KEY, None)
        return t if isinstance(t, dict) and t else {k: dict(v) for k, v in NAV_POST_TEMPLATES_DEFAULT.items()}
    return _nav_cached("tpl", build)


def _template_caps(post_key):
    """某岗位模板 → 应勾上的权限码集合。
    模板 = 准入点 + 显式列的非敏感动作点。**敏感点一律不进**（确认书第五节：地板不能被模板绕过）——
    即便有人往模板里塞了敏感码，这里也会滤掉。"""
    t = _post_templates().get(post_key) or {}
    secs = set(t.get("secs") or [])
    mods = set(t.get("mods") or [])
    out = set()
    for m in _all_modules():
        # 不按 always 跳过：always 管的是【上线状态】(恒常可用)，不是权限。
        # 基础数据 always=True 但准入点非敏感，该给就给；系统设置 always=True 但 cap=enter_settings 是
        # 敏感点，被下面 is_sensitive 挡掉。用 always 当权限闸门会让「基础数据」谁都进不去（改造前它人人可见）。
        if m["sec"] in secs or m["key"] in mods:
            cap = _enter_cap(m)
            if cap and not db.is_sensitive(cap):
                out.add(cap)
    static = {c["key"] for c in db.CAP_META_STATIC}
    for a in (t.get("acts") or []):
        if a in static and not db.is_sensitive(a):
            out.add(a)
    return out


@app.get("/api/nav-modules")
def nav_modules(request: Request):
    # 侧栏读的就是这个接口。「开发中」是否可进入**因人而异**，所以状态必须按当前登录人算，
    # 不能是一份全局静态表。
    u = _current_user(request)
    dev = can_enter_dev(u)
    # 名单 + **服务器到底读到了什么**一起下发：这个闸读不到配置时是放宽的（回落全体主管理员），
    # 不把原因摆出来的话，现场表现就是"我明明配了却不生效"而页面上毫无线索。
    # 名单本身不是机密（就是几个账号名），机密的是 conf.ini 里的口令/密钥，那些一律不下发。
    info = dev_users_info()
    return {"modules": _all_modules(), "sections": _all_sections(), "statuses": NAV_STATUSES,
            "posts": _nav_posts(), "posts_max": NAV_POSTS_MAX, "templates": _post_templates(),
            "可进入状态": list(NAV_OPEN) + ([*NAV_OPEN_ADMIN] if dev else []),
            "state": _nav_state(dev), "我能进开发中": dev,
            "开发者名单": info["names"], "开发者名单说明": info["note"],
            "我是谁": (u or {}).get("name", ""),
            "可编辑": _nav_can_edit(u)}


@app.post("/api/nav-modules/save")
def nav_modules_save(body: dict, request: Request):
    """一次保存：岗位名单(posts) + 各模块状态与所挂岗位(state) + 岗位模板(templates)。"""
    u = _current_user(request)
    if not _nav_can_edit(u):
        return JSONResponse({"ok": False, "msg": "只有主管理员能改导航模块的状态"}, status_code=403)

    roster = _nav_posts()
    if isinstance(body.get("posts"), list):     # 岗位名单有变更
        new_roster = _clean_posts(body["posts"])
        if not new_roster:
            return {"ok": False, "msg": "岗位名单至少保留 1 个"}
        old = {p["key"]: p["label"] for p in roster}
        new_keys = {p["key"] for p in new_roster}
        removed = [lb for k, lb in old.items() if k not in new_keys]
        renamed = ["%s→%s" % (old[p["key"]], p["label"]) for p in new_roster
                   if p["key"] in old and old[p["key"]] != p["label"]]
        db.set_setting(_POSTS_KEY, new_roster, u["name"])
        # V2.146：删除岗位【后端连带】从各模块的挂载里摘掉——原先这步靠核算系统设置页的前端
        # 顺手做（delPost 同时改 st 一起提交）；岗位维护搬进门户权限中枢后是只发 {posts} 的调用，
        # 不下沉到后端就会在 state 里残留死岗位 key（侧栏岗位标签显示成裸 key）。谁调都安全。
        removed_keys = {k for k in old if k not in new_keys}
        if removed_keys:
            st_cur = db.get_setting(_NAV_KEY, {}) or {}
            changed = False
            for mk, mv in st_cur.items():
                if isinstance(mv, dict) and mv.get("posts"):
                    kept = [p for p in mv["posts"] if p not in removed_keys]
                    if len(kept) != len(mv["posts"]):
                        mv["posts"] = kept
                        changed = True
            if changed:
                db.set_setting(_NAV_KEY, st_cur, u["name"])
        # 同理：模板里被删岗位的条目一并清掉（不清也只是死数据，但留着会在导出/审计里冒出来）
        if removed_keys:
            tpl_cur = db.get_setting(_TPL_KEY, None)
            if isinstance(tpl_cur, dict) and any(k in tpl_cur for k in removed_keys):
                db.set_setting(_TPL_KEY, {k: v for k, v in tpl_cur.items() if k not in removed_keys}, u["name"])
        _nav_cache_clear()
        roster = _nav_posts()
        if removed:
            db.audit(u["name"], "岗位名单", "核算工作台", "删除：" + "、".join(removed))
        if renamed:                             # 改名不丢绑定（key 没动），但要留痕
            db.audit(u["name"], "岗位改名", "核算工作台", "、".join(renamed))

    if isinstance(body.get("templates"), dict):
        pk = _post_keys()
        tpl = {k: {"secs": [str(x) for x in (v.get("secs") or [])],
                   "mods": [str(x) for x in (v.get("mods") or [])],
                   "acts": [str(x) for x in (v.get("acts") or [])]}
               for k, v in body["templates"].items()
               if k in pk and isinstance(v, dict)}
        db.set_setting(_TPL_KEY, tpl, u["name"])
        db.audit(u["name"], "岗位模板", "核算工作台", "已更新 %d 个岗位" % len(tpl))

    # V2.144：state 只在【显式传了】时才写——原先 `body.get("state") or {}` 无条件重写，
    # 谁只想存模板（不带 state）就会把全部模块状态静默重置回 default。权限中枢的
    # 「岗位模板设置」只发 {templates}，踩上这雷＝改个模板顺手把上线开关全冲掉。
    if isinstance(body.get("state"), dict):
        st = body["state"]
        pk = _post_keys()
        all_mods = _all_modules()
        new = {}
        for m in all_mods:
            if m.get("always"):
                continue                            # 恒常模块(基础数据/设置)不存、不可改
            cur = st.get(m["key"]) or {}
            raw = _nav_norm(cur.get("status"))          # V2.173：旧前端缓存可能还发旧值，存时也平移
            status = raw if raw in NAV_STATUSES else _nav_norm(m.get("default", "引擎正常"))
            posts = [p for p in (cur.get("posts") or []) if p in pk] if status == "引擎正常" else []
            new[m["key"]] = {"status": status, "posts": posts}
        db.set_setting(_NAV_KEY, new, u["name"])
        _nav_cache_clear()
        brief = "；".join("%s=%s" % (m["label"], new[m["key"]]["status"])
                          for m in all_mods if not m.get("always"))
        db.audit(u["name"], "导航模块状态", "核算工作台", brief[:300])
    _nav_cache_clear()
    return {"ok": True, "state": _nav_state(can_enter_dev(u)), "posts": _nav_posts(), "templates": _post_templates()}


@app.post("/api/nav-modules/move")
def nav_module_move(body: dict, request: Request):
    """改一个模块的位置：挂到哪个一级(sec)、挂在谁下面(parent)、排第几(order)、叫什么(label)。
    内置模块也能改——存的是**覆盖值**，代码里那份只当默认。不然这次搬完家，以后照样搬不动。"""
    u = _current_user(request)
    if not _nav_can_edit(u):
        return JSONResponse({"ok": False, "msg": "只有主管理员能调整菜单"}, status_code=403)
    key = str(body.get("key") or "").strip()
    m = _mod_by_key(key)
    if not m:
        return {"ok": False, "msg": "没有这个模块：%s" % key}

    patch = {}
    if "sec" in body:
        sec = str(body.get("sec") or "").strip()
        if sec not in {s["key"] for s in _all_sections()}:
            return {"ok": False, "msg": "没有这个一级板块：%s" % sec}
        patch["sec"] = sec
    if "parent" in body:
        p = str(body.get("parent") or "").strip() or None
        if p:
            pm = _mod_by_key(p)
            if not pm:
                return {"ok": False, "msg": "没有这个父模块：%s" % p}
            if p == key:
                return {"ok": False, "msg": "不能挂在自己下面"}
            if pm.get("parent"):
                return {"ok": False, "msg": "「%s」已经是三级了，菜单只到三级" % pm["label"]}
            # 自己底下还有子项 → 再挂到别人下面就成四级了
            if any(x.get("parent") == key for x in _all_modules()):
                return {"ok": False, "msg": "「%s」底下还挂着子项，不能再挂到别人下面（菜单只到三级）" % m["label"]}
            patch["sec"] = pm["sec"]            # 子项必须跟父项同一个一级，否则会在别的板块里凭空冒出来
        patch["parent"] = p
    if "order" in body:
        try:
            patch["order"] = max(0, min(9999, int(body.get("order"))))
        except Exception:
            return {"ok": False, "msg": "排序要填数字"}
    if "label" in body:
        lb = str(body.get("label") or "").strip()[:20]
        if not lb:
            return {"ok": False, "msg": "名称不能为空"}
        patch["label"] = lb
    if not patch:
        return {"ok": False, "msg": "没有要改的内容"}

    if m.get("builtin"):
        ovr = db.get_setting(_MOD_OVR_KEY, {}) or {}
        cur = ovr.get(key) if isinstance(ovr.get(key), dict) else {}
        ovr[key] = {**cur, **patch}
        db.set_setting(_MOD_OVR_KEY, ovr, u["name"])
    else:
        umods = _user_modules()
        for x in umods:
            if x["key"] == key:
                x.update(patch)
        db.set_setting(_UMOD_KEY, umods, u["name"])
    _nav_cache_clear()
    db.audit(u["name"], "菜单位置调整", key,
             "、".join("%s=%s" % (k, v) for k, v in patch.items())[:200])
    return {"ok": True, "modules": _all_modules(), "sections": _all_sections(), "state": _nav_state(can_enter_dev(u))}


@app.post("/api/nav-sections/save")
def nav_sections_save(body: dict, request: Request):
    """一级板块：改名/排序（内置存覆盖值）、新增/删除（仅自建）。"""
    u = _current_user(request)
    if not _nav_can_edit(u):
        return JSONResponse({"ok": False, "msg": "只有主管理员能改一级板块"}, status_code=403)
    secs = body.get("sections")
    if not isinstance(secs, list):
        return {"ok": False, "msg": "参数不对"}

    ovr, usecs, seen = {}, [], set()
    for s in secs:
        if not isinstance(s, dict):
            continue
        k = str(s.get("key") or "").strip()
        lb = str(s.get("label") or "").strip()[:20]
        if not k or not lb or k in seen:
            continue
        seen.add(k)
        try:
            order = max(0, min(9999, int(s.get("order") or 500)))
        except Exception:
            order = 500
        if k in _SEC_KEYS:
            ovr[k] = {"label": lb, "order": order}
        else:
            if not _KEY_RE.match(k):
                return {"ok": False, "msg": "标识「%s」要 2–24 位、小写字母开头、只含小写字母/数字/下划线" % k}
            usecs.append({"key": k, "label": lb, "order": order})

    # 内置一级不可删（模块还挂在上面，删了整片菜单会消失）；自建的删掉前得先腾空
    gone = {s["key"] for s in _all_sections() if not s.get("builtin")} - {s["key"] for s in usecs}
    for g in gone:
        用着的 = [m["label"] for m in _all_modules() if m["sec"] == g]
        if 用着的:
            return {"ok": False, "msg": "「%s」下面还有 %s，先把它们挪走再删" % (g, "、".join(用着的[:3]))}
    missing = _SEC_KEYS - seen
    if missing:
        return {"ok": False, "msg": "内置板块删不得：%s（要隐藏就把它下面的模块设成「敬请期待」）"
                                    % "、".join(sorted(missing))}

    db.set_setting(_SEC_OVR_KEY, ovr, u["name"])
    db.set_setting(_USEC_KEY, usecs, u["name"])
    _nav_cache_clear()
    db.audit(u["name"], "一级板块调整", "核算工作台",
             "；".join("%s=%s(%d)" % (s["key"], s["label"], s["order"]) for s in _all_sections())[:300])
    return {"ok": True, "sections": _all_sections(), "modules": _all_modules()}


@app.post("/api/nav-modules/add-module")
def nav_module_add(body: dict, request: Request):
    """主管理员自建一个「规划中」模块（不绑代码）。key 唯一、限小写字母数字下划线。
    没接代码前它在侧栏灰显、点不进去（默认「开发中」）；开发接上同名 key 即自动可用。"""
    u = _current_user(request)
    if not _nav_can_edit(u):
        return JSONResponse({"ok": False, "msg": "只有主管理员能加模块"}, status_code=403)
    key = str(body.get("key") or "").strip().lower()
    label = str(body.get("label") or "").strip()
    sec = str(body.get("sec") or "").strip()
    parent = str(body.get("parent") or "").strip() or None
    status = _nav_norm(body.get("status")) if _nav_norm(body.get("status")) in NAV_STATUSES else "开发中"
    if not label:
        return {"ok": False, "msg": "模块名称不能为空"}
    if not _KEY_RE.match(key):
        return {"ok": False, "msg": "标识（key）要 2–24 位、小写字母开头、只含小写字母/数字/下划线（它当路由用）"}
    if key in _BUILTIN_KEYS:
        return {"ok": False, "msg": "标识「%s」是内置工具已占用，换一个" % key}
    umods = _user_modules()
    if any(m["key"] == key for m in umods):
        return {"ok": False, "msg": "标识「%s」已存在" % key}
    # parent＝建三级（D2：二级也可以新增三级）。子项跟父项同一个一级，否则会在别的板块里凭空冒出来。
    if parent:
        pm = _mod_by_key(parent)
        if not pm:
            return {"ok": False, "msg": "没有这个父模块：%s" % parent}
        if pm.get("parent"):
            return {"ok": False, "msg": "「%s」已经是三级了，菜单只到三级" % pm["label"]}
        sec = pm["sec"]
    if sec not in {s["key"] for s in _all_sections()}:
        return {"ok": False, "msg": "要先选一个一级板块"}
    try:
        order = max(0, min(9999, int(body.get("order") or 500)))
    except Exception:
        order = 500
    umods.append({"key": key, "label": label[:20], "sec": sec, "parent": parent, "order": order})
    db.set_setting(_UMOD_KEY, umods, u["name"])
    # 初始状态写进状态表（非「设置」类，可存）
    saved = db.get_setting(_NAV_KEY, {}) or {}
    saved[key] = {"status": status, "posts": []}
    db.set_setting(_NAV_KEY, saved, u["name"])
    _nav_cache_clear()      # 准入点跟菜单树走：新菜单的 enter:<key> 要立刻能在账号管理页勾到
    sl = next((s["label"] for s in _all_sections() if s["key"] == sec), sec)
    db.audit(u["name"], "导航模块新增", key,
             "%s @ %s%s · %s" % (label, sl, ("／" + _mod_by_key(parent)["label"]) if parent else "", status))
    return {"ok": True, "modules": _all_modules(), "sections": _all_sections(), "state": _nav_state(can_enter_dev(u))}


@app.post("/api/nav-modules/del-module")
def nav_module_del(body: dict, request: Request):
    """删除一个自建模块。内置（代码持有）不可删——要隐藏内置工具请把状态设「敬请期待」。"""
    u = _current_user(request)
    if not _nav_can_edit(u):
        return JSONResponse({"ok": False, "msg": "只有主管理员能删模块"}, status_code=403)
    key = str(body.get("key") or "").strip().lower()
    if key in _BUILTIN_KEYS:
        return {"ok": False, "msg": "「%s」是内置工具，删不得（要隐藏就把状态设「敬请期待」）" % key}
    kids = [m["label"] for m in _all_modules() if m.get("parent") == key]
    if kids:
        return {"ok": False, "msg": "「%s」下面还挂着 %s，先把它们挪走或删掉" % (key, "、".join(kids[:3]))}
    umods = [m for m in _user_modules() if m["key"] != key]
    db.set_setting(_UMOD_KEY, umods, u["name"])
    saved = db.get_setting(_NAV_KEY, {}) or {}
    if key in saved:
        del saved[key]
        db.set_setting(_NAV_KEY, saved, u["name"])
    _nav_cache_clear()      # 菜单没了，它的准入点也跟着退休
    db.audit(u["name"], "导航模块删除", key)
    return {"ok": True, "modules": _all_modules(), "sections": _all_sections(), "state": _nav_state(can_enter_dev(u))}


@app.get("/api/kingdee/test")
def kingdee_test():
    ok, msg = kc.test_connection()
    return {"ok": ok, "msg": msg, "conf": kc.conf_path()}


# ---------------- 资金看板 ----------------
# 账户性质：比"科目大类"更贴业务的分类（库存现金/基本户/一般户/通知存款/理财/第三方支付…），
# 取自出纳主数据的 类别+账户类型；主数据没有则按 科目大类+名称 兜底。理财类归并（结构性存款/1101）。
def _acct_nature(cat, category, atype, name):
    cat, category, atype, name = (cat or ""), (category or ""), (atype or ""), str(name or "")
    if cat == "库存现金":
        return "库存现金"
    if category == "电商渠道" or any(k in name for k in ("支付宝", "微信", "抖音", "淘宝", "天猫", "京东", "有赞", "小红书", "网商", "@")):
        return "第三方支付"
    if cat == "交易性金融资产" or atype == "结构性存款" or any(k in name for k in ("理财", "结构性", "基金", "债", "增利", "现金添利", "天天")):
        return "理财"
    if atype:                    # 银行账户：直接用出纳的账户类型（基本户/一般户/通知存款/资本金户…）
        return atype
    if "基本户" in name:
        return "基本户"
    if "一般户" in name:
        return "一般户"
    if cat == "其它货币资金":
        return "其它货币"
    return cat or "其他"


def _fund():
    try:
        rows = _balance_rows()
    except KdNotFetched:
        return {"未取数": True, "source": CFG["source"], "period": _period_str(),
                "note": "本期未取数：请到「数据接入」点【刷新金蝶数据】",
                "集团合计": 0, "科目大类": [], "主体": [], "accounts": [], "guardrail": {}}
    except kc.KingdeeError as e:
        return {"error": str(e), "source": CFG["source"], "period": _period_str(),
                "集团合计": 0, "科目大类": [], "主体": [], "accounts": [], "guardrail": {}}
    bank_map = {al.norm_acct(r.get("核算维度.银行账号.名称") or ""):
                al.parse_bank(r.get("核算维度.银行账号.名称") or "") for r in rows}
    # 未过账期间 GL_BALANCE 期末停在期初，须用序时账还原期末（期初+本期序时账净），与科目余额表同口径。
    # 样例模式无序时账、样例余额自带真实期末 → 传 None 走期末字段。
    vou = None
    if CFG["source"] == "kingdee":
        try:
            vou = _kd_get("gl_subjects")
        except KdNotFetched:
            vou = None
    bal = bd.load_balance(rows, vou)
    ledger = al.to_ledger_map(al.load_ledger(LEDGER_PATH))
    kd_last = S.sample_kd_last() if CFG["source"] == "sample" else {}
    flow_last = S.sample_flow_last() if CFG["source"] == "sample" else {}
    d = bd.build_dashboard(bal, kd_last=kd_last, flow_last=flow_last, ledger=ledger or None)
    # 开户行 + 账户性质 取自【出纳管理·银行账号】主数据（刷新时定格），台账/名称解析兜底
    brec = db.get_period_input(CFG["source"], CFG["year"], CFG["period"], "kd:bank_master")
    bmaster = (brec["payload"].get("map") if brec else {}) or {}
    for a in d["accounts"]:
        ac = al.norm_acct(a.get("账号") or "")
        info = bmaster.get(ac) or {}
        if isinstance(info, str):
            info = {"开户行": info}          # 兼容旧格式（曾只存开户行字符串）
        a["开户行"] = info.get("开户行") or (ledger.get(a["账号"]) or {}).get("开户行") or bank_map.get(a["账号"], "")
        a["账户性质"] = _acct_nature(a.get("科目大类"), info.get("类别"), info.get("账户类型"), a.get("账号"))
    # 当月新增账户：本期账号在【上一期金蝶科目余额】里不存在＝本月首次出现(*New)。
    # 靠上期已取数的持久化余额判断；上期没取过数则无法判断（不标、给提示）。
    prior_known, prior_set = False, set()
    if CFG["source"] == "kingdee":
        py, pp = (CFG["year"] - 1, 12) if int(CFG["period"]) == 1 else (CFG["year"], int(CFG["period"]) - 1)
        rec = db.get_period_input("kingdee", py, pp, "kd:gl_balance")
        if rec:
            prior_known = True
            for r in rec["payload"].get("rows", []):
                a0 = al.norm_acct(r.get("核算维度.银行账号.编码") or r.get("核算维度.银行账号.名称") or "")
                if a0:
                    prior_set.add(a0)
    for a in d["accounts"]:
        acct = al.norm_acct(a.get("账号") or "")
        a["新增"] = bool(prior_known and acct and acct not in prior_set)
    # 样例演示：无上期数据，指定一户标新增，便于演示 *New 功能
    if CFG["source"] == "sample":
        prior_known = True
        for a in d["accounts"]:
            a["新增"] = "755953100010001" in str(a.get("账号") or "")
    d["新增可判断"] = prior_known
    d["金蝶取数"] = _kd_sync_info()          # V2.176：资金看板也明示"谁于何时刷新"
    d["period"] = _period_str(); d["source"] = CFG["source"]
    d["synced_at"] = "样例数据" if CFG["source"] == "sample" else "金蝶"
    return d


@app.get("/api/fund-dashboard")
def fund_dashboard():
    return _cache_get(_FUND_CACHE, _fund)                 # 普通进入：读缓存秒开


@app.post("/api/fund-dashboard/sync")
def fund_sync():
    return _closed_block() or _cache_get(_FUND_CACHE, _fund, force=True)   # 一键接入金蝶：强制重取


@app.get("/api/debug/balance-raw")
def debug_balance_raw(prefix: str = "1101", acct: str = ""):
    """排查用：看本期某科目前缀的【原始 GL_BALANCE 行】——核算维度到底落在哪个字段、值是什么。
    例：/api/debug/balance-raw?prefix=1101 → 确认交易性金融资产的"63672948信益嘉321号6单元"
    有没有取到（若「核算维度.银行账号.名称/编码」有值＝FF100002 命中；若为空＝该科目用了别的核算维度）。"""
    try:
        rows = _balance_rows()
    except KdNotFetched:
        return {"note": "本期未取数：请先到「数据接入」刷新金蝶", "source": CFG["source"]}
    except kc.KingdeeError as e:
        return {"error": str(e), "source": CFG["source"]}
    hit = [r for r in rows if str(r.get("科目编码") or r.get("FAccountID.FNumber") or "").startswith(prefix)]
    if acct:
        hit = [r for r in hit if acct in str(r.get("核算维度.银行账号.编码") or r.get("核算维度.银行账号.名称") or "")]
    show = ["账簿", "科目编码", "科目名称", "核算维度.银行账号.编码", "币别",
            "期初原币", "本期借方原币", "本期贷方原币", "期末原币", "期末本位币"]
    return {"source": CFG["source"], "period": _period_str(), "prefix": prefix, "count": len(hit),
            "命中核算维度名的行数": sum(1 for r in hit if str(r.get("核算维度.银行账号.名称") or r.get("核算维度.银行账号.编码") or "").strip()),
            "rows": [{k: r.get(k) for k in show} for r in hit[:60]]}


# ---------------- 账户台账 ----------------
@app.get("/api/account-ledger")
def account_ledger():
    auth = _auth_ledger_records(_period_str())
    if auth is not None:
        return {"period": _period_str(), "source": "金蝶权威台账(出纳·银行账号)", "records": auth}
    recs = al.load_ledger(LEDGER_PATH)
    for r in recs:
        r["本月新增"] = bool(r.get("首次出现期间") == _period_str() and r.get("_active"))
    return {"period": _period_str(), "source": CFG["source"], "records": recs}


@app.post("/api/account-ledger/sync")
def account_ledger_sync():
    blocked = _closed_block()
    if blocked:
        return blocked
    try:
        kd = al.kd_accounts_from_cn(_bank_accounts())
    except kc.KingdeeError as e:
        return {"error": str(e), "source": CFG["source"], "period": _period_str(),
                "records": al.load_ledger(LEDGER_PATH), "report": {}}
    prev = al.load_ledger(LEDGER_PATH)
    recs, report = al.sync_ledger(prev, kd, _period_str())
    al.save_ledger(LEDGER_PATH, recs)
    return {"period": _period_str(), "source": CFG["source"], "records": recs,
            "report": {k: v for k, v in report.items() if k != "changes"}}


@app.post("/api/account-ledger/override")
def account_ledger_override(body: dict, request: Request):
    """手工覆盖账户：失效/恢复、稽核方案(明细/余额)。金蝶不维护此步，本地持久化，不写金蝶。"""
    if not _require_perm(request, "ledger_override"):
        return JSONResponse({"ok": False, "msg": "无「账户台账改动」权限，请联系管理员"}, status_code=403)
    acct = str(body.get("账号", "") or "").strip()
    if not acct:
        return {"ok": False, "msg": "缺账号"}
    ov = _load_overrides()
    cur = dict(ov.get(acct, {}))
    if "失效" in body:
        if body.get("失效"):
            cur["失效"] = True
        else:
            cur.pop("失效", None)
    if body.get("稽核方案") in ("明细", "余额"):
        cur["稽核方案"] = body["稽核方案"]
    if cur:
        ov[acct] = cur
    else:
        ov.pop(acct, None)
    _save_overrides(ov)
    db.audit(str(body.get("operator", "") or ""), "账户覆盖", acct, json.dumps(cur, ensure_ascii=False))
    return {"ok": True, "账号": acct, "override": cur}


# ---------------- 逐笔稽核 v2（引擎七态 + 匹配桥；银行流水待上传归一化明细，暂用样例）----------------
def _auth_ledger_rows():
    """加载权威台账原始行（供匹配桥 ledger_index 用）。"""
    if os.path.exists(AUTH_LEDGER_PATH):
        try:
            return json.load(open(AUTH_LEDGER_PATH, encoding="utf-8"))
        except Exception:
            return []
    return []


def _scheme_map():
    """{账号数字键: 稽核方案(明细/余额)} —— 取自权威台账 + 手工覆盖。用于稽核方案联动。"""
    recs = _auth_ledger_records(_period_str()) or []
    m = {}
    for r in recs:
        d = al.norm_acct(r.get("账号", ""))
        if d:
            m[d] = r.get("稽核方案", "明细")
    return m


def _real_bank_kd():
    """金蝶模式真数据：本期已定格的 金蝶1002序时账 + 已上传解析好的银行流水。
    返回 (bank_rows, kd_rows, manifest, bank_src)。金蝶未取数→KdNotFetched；银行未上传→空+提示。"""
    kd_rows = _kd_get("gl_voucher:1002")          # 本期已存的1002；没取过抛 KdNotFetched
    bank_rows, manifest, bmeta = _period_bank()   # 本期已上传的流水（严格按期间，不串月、不用样例）
    if bank_rows is None:
        bank_rows, manifest = [], []
        bank_src = "本期未上传银行流水（到「数据接入」上传本月流水包）"
    else:
        bank_src = "本期银行流水（%s 上传于 %s，并入逐笔 %d 笔）" % (
            bmeta.get("updated_by", "?"), bmeta.get("updated_at", "?"), len(bank_rows))
    return bank_rows, kd_rows, manifest, bank_src


def _reconcile():
    idx = rc.ledger_index(_auth_ledger_rows())
    manifest, balance_accts = [], []
    if CFG["source"] == "kingdee":
        try:
            bank_rows, kd_rows, manifest, bank_src = _real_bank_kd()
        except KdNotFetched:
            return {"period": _period_str(), "source": CFG["source"], "未取数": True,
                    "bank_source": "", "results": [], "summary": {}, "guardrail": {}, "manifest": [],
                    "余额稽核账户": [], "note": "本期未取数：请到「数据接入」点【刷新金蝶数据】"}
        except kc.KingdeeError as e:
            return {"period": _period_str(), "source": CFG["source"], "error": str(e),
                    "bank_source": "", "results": [], "summary": {}, "guardrail": {},
                    "manifest": [], "余额稽核账户": []}
        # ④ 稽核方案联动：只对"明细"方案账户逐笔；"余额"账户(台账可点调)排除，走余额调节。
        scheme = _scheme_map()
        def _is_detail(acct_raw):
            d = al.norm_acct(acct_raw or "")
            return scheme.get(d, "明细") == "明细"          # 台账未收录默认明细，照常逐笔
        def _kd_acct(r):
            return r.get("核算维度.银行账号.编码") or r.get("FDetailID.FF100002.FNumber") or ""
        excluded = {al.norm_acct(r.get("账号", "")) for r in bank_rows if not _is_detail(r.get("账号", ""))}
        excluded |= {al.norm_acct(_kd_acct(r)) for r in kd_rows if not _is_detail(_kd_acct(r))}
        bank_rows = [r for r in bank_rows if _is_detail(r.get("账号", ""))]
        kd_rows = [r for r in kd_rows if _is_detail(_kd_acct(r))]
        balance_accts = sorted(a for a in excluded if a)
    else:
        kd_rows = S.sample_kd_rows()
        bank_rows = S.sample_bank_rows()
        bank_src = "sample"
    bank = rc.bank_to_recs(bank_rows, idx or None)
    kd = rc.kd_to_recs(kd_rows, idx or None)
    # V2.171：集团主体名单（主体档案 全称/简称/别名 + 流水本方户名）→「内部往来」侦测的对方户名闸——
    # 对方是外部名称（个人/税务局/供应商）时不猜内部往来，金额巧合不再误标（7月实测曹水英/税款两例）
    gn = {str(r.get("户名") or "").strip() for r in bank_rows}
    try:
        for o in db.list_orgs():
            gn.add(str(o.get("full_name") or "").strip())
            gn.add(str(o.get("short_name") or "").strip())
            gn.update(str(a).strip() for a in (o.get("aliases") or []))
    except Exception:
        pass
    gn.discard("")
    results, summary = rc.reconcile(bank, kd, group_names=gn or None)
    _assign_keys(results)
    return {"period": _period_str(), "source": CFG["source"],
            "bank_source": bank_src,
            "results": results,
            "summary": {k: v for k, v in summary.items() if k != "guardrail"},
            "guardrail": summary["guardrail"],
            "manifest": manifest,
            "金蝶取数": _kd_sync_info(),      # V2.176：页面明示"谁于何时刷新"，点了刷新有据可查
            "余额稽核账户": balance_accts}


def _data_sources():
    """数据接入页：银行导入 manifest + 每家银行 金蝶↔银行 覆盖对照（不跑逐笔）。"""
    base = {"source": CFG["source"], "period": _period_str(),
            "bank_import_dir": CFG.get("bank_import_dir", ""), "updated_at": _now()}
    if CFG["source"] != "kingdee":
        return {**base, "bank_source": "样例数据", "manifest": [], "coverage": [],
                "kd_count": 0, "balance_count": 0, "balance_by_subject": []}
    try:
        bank_rows, kd_rows, manifest, bank_src = _real_bank_kd()
    except KdNotFetched:
        # 金蝶未取数：银行流水可能已上传，仍显示它的状态，只是金蝶侧空着
        bank_rows, manifest, bmeta = _period_bank()
        return {**base, "未取数": True, "bank_source":
                ("本期银行流水（%s 上传于 %s）" % (bmeta.get("updated_by", "?"), bmeta.get("updated_at", "?"))
                 if bank_rows is not None else "本期未上传银行流水"),
                "bank_meta": bmeta, "kd_synced_at": "", "manifest": manifest or [], "coverage": [],
                "kd_count": 0, "balance_count": None, "balance_by_subject": [],
                "note": "本期未取数：请点【刷新金蝶数据】"}
    except kc.KingdeeError as e:
        return {**base, "error": str(e), "bank_source": "", "manifest": [], "coverage": [], "kd_count": 0}
    _bmeta = _period_bank()[2]
    from collections import Counter
    bank_by, kd_by = Counter(), Counter()
    # V2.168：归行按【账号→官方开户行】（出纳管理主数据>账户台账），查无才按行内"银行"标签认字——
    # 财资平台导出无开户行列，行标签是"宁波/招商"混合值，纯按标签认会把招商的笔数全记到宁波头上（招商恒"待补"）。
    brec = db.get_period_input(CFG["source"], CFG["year"], CFG["period"], "kd:bank_master")
    bmaster = (brec["payload"].get("map") if brec else {}) or {}
    lmap = al.to_ledger_map(al.load_ledger(LEDGER_PATH))
    for r in bank_rows:
        bank_by[al.bank_of_row(r.get("账号", ""), r.get("银行", ""), bmaster, lmap) or "其他"] += 1
    for r in kd_rows:
        nm = r.get("核算维度.银行账号.编码") or r.get("FDetailID.FF100002.FNumber") or ""
        kd_by[al.bank_of(nm) or "其他"] += 1
    coverage = [{"银行": b, "金蝶笔数": kd_by.get(b, 0), "银行笔数": bank_by.get(b, 0),
                 "状态": "已覆盖" if bank_by.get(b, 0) > 0 else "待补"}
                for b in sorted(set(bank_by) | set(kd_by),
                                key=lambda x: -(kd_by.get(x, 0) + bank_by.get(x, 0)))]
    bal_count, balance_by_subject = None, []
    try:
        bal_rows = _kd_get("gl_balance")   # 四类资金科目 1001/1002/1012/1101（本期已存的）
        bal_count = len(bal_rows)
        sc = Counter()
        for r in bal_rows:
            cat = al.cat_from_code(str(r.get("科目编码") or ""))
            if cat:
                sc[cat] += 1
        balance_by_subject = [{"科目": s, "行数": sc[s]} for s in
                              ("库存现金", "银行存款", "其它货币资金", "交易性金融资产") if sc.get(s)]
    except (KdNotFetched, kc.KingdeeError):
        pass
    _si = _kd_sync_info()
    return {**base, "bank_source": bank_src, "manifest": manifest, "bank_meta": _bmeta,
            "kd_synced_at": _si["at"], "kd_synced_by": _si["by"],
            "coverage": coverage, "kd_count": len(kd_rows),
            "balance_count": bal_count, "balance_by_subject": balance_by_subject}


def _acct_info(a):
    """账号 → (主体, 开户行, 账户名称, 币别)，取自权威台账。"""
    for r in _auth_ledger_rows():
        la = al.norm_acct(r.get("账号", ""))
        if la and (la == a or la.endswith(a) or a.endswith(la)):
            cur = str(r.get("币别") or r.get("币种") or "").strip()
            return (r.get("主体", ""), r.get("开户行", ""), r.get("账户名称", "") or r.get("户名", ""), cur)
    return ("", "", "", "")


def _balance_adjust():
    """银行存款余额调节表：每户 银行对账单余额 / 金蝶账面余额 / 未达账项(来自逐笔稽核) → 调节后对平。
    未达口径：金蝶单边=企业已记银行未记(调银行侧)；疑似漏账+内部往来未做账=银行已记企业未记(调账面侧)。"""
    base = {"source": CFG["source"], "period": _period_str(), "updated_at": _now()}
    if CFG["source"] != "kingdee":
        return {**base, "accounts": [], "对平户数": 0, "不平户数": 0, "note": "样例数据"}
    try:
        bank_rows, kd_rows, _mf, _src = _real_bank_kd()
        bal_rows = _kd_get("gl_balance")
    except KdNotFetched:
        return {**base, "未取数": True, "accounts": [], "对平户数": 0, "不平户数": 0,
                "note": "本期未取数：请到「数据接入」点【刷新金蝶数据】"}
    except kc.KingdeeError as e:
        return {**base, "error": str(e), "accounts": [], "对平户数": 0, "不平户数": 0}
    # 金蝶账面余额 = 期初(dedup) + 本期序时账净发生。
    # 不用 GL_BALANCE 期末：凭证若未过账，期末=期初、不含本期变动，会与逐笔稽核的未达口径打架。
    # GL_BALANCE 会返回重复行(同科目+同维度)，须按(科目,维度)去重，否则相加→翻倍。
    # open_bal=原币账面(逐户对平用)；open_bal_base=本位币账面(本位币列用，人民币户两者相等)。
    # ⚠外币户 GL_BALANCE 有两行：币别=空 的本位币合计行(原币栏其实是本位币) + 币别=具体 的原币明细行。
    # 原币账面必须取「具体币别」行(否则外币户原币会错取成本位币，数飘)；本位币两行一致。故按(科目,维度)选具体币别行。
    open_rows = {}
    for r in bal_rows:
        code = str(r.get("科目编码") or "")
        if not code.startswith("1002"):
            continue
        dim = str(r.get("核算维度.银行账号.编码") or r.get("核算维度.银行账号.名称") or "")
        if not dim:
            continue
        key = (code, dim)
        cur = str(r.get("币别") or "").strip()
        prev = open_rows.get(key)
        if prev is None or (cur and not prev[2]):   # 首次；或本行有具体币别而已存行没有 → 用本行(原币才准)
            open_rows[key] = (rc.to_float(r.get("期初原币") or 0), rc.to_float(r.get("期初本位币") or 0), cur)
    open_bal, open_bal_base = {}, {}
    for (code, dim), (yo, ben, _cur) in open_rows.items():
        a = al.norm_acct(dim)
        if a:
            open_bal[a] = open_bal.get(a, 0.0) + yo
            open_bal_base[a] = open_bal_base.get(a, 0.0) + ben
    kd_move, kd_move_base, acct_kd_cur = {}, {}, {}
    for r in kd_rows:
        a = al.norm_acct(r.get("核算维度.银行账号.编码") or r.get("FDetailID.FF100002.FNumber") or "")
        if not a:
            continue
        # 原币口径(V2.32)：期初取的是「期初原币」，本期发生也必须原币——外币户 FDEBIT/FCREDIT 是
        # 账簿本位币(境外簿=美元)，港币户按本位币加会得出「港币期初+美元发生」的错账面。
        kd_move[a] = kd_move.get(a, 0.0) + rc.kd_delta_for(r)
        # 本位币净发生(本位币列用)：FDEBIT/FCREDIT 本就是账簿本位币，直接借-贷
        kd_move_base[a] = kd_move_base.get(a, 0.0) + rc.to_float(r.get("FDEBIT") or 0) - rc.to_float(r.get("FCREDIT") or 0)
        c = str(r.get("FCURRENCYID.FName") or r.get("币别") or "").strip()   # 币别兜底：台账没有时用金蝶序时账
        if c and a not in acct_kd_cur:
            acct_kd_cur[a] = c
    book = {a: round(open_bal.get(a, 0.0) + kd_move.get(a, 0.0), 2) for a in set(open_bal) | set(kd_move)}
    # 金蝶账面·本位币：期初本位币+本期序时账本位币净，还原实时本位币账面。口径=账簿本位币
    # （境内簿=人民币，境外簿 Sinkio/Starfield=美元；与逐笔稽核 V2.32「本位币金额」列一致）。人民币户 == book(原币)。
    book_base = {a: round(open_bal_base.get(a, 0.0) + kd_move_base.get(a, 0.0), 2) for a in set(open_bal_base) | set(kd_move_base)}
    # 银行对账单期末余额：每户取交易日期最晚一笔的余额
    bank_last = {}
    for r in bank_rows:
        a = al.norm_acct(r.get("账号") or "")
        if not a or r.get("余额") is None:
            continue
        d = r.get("交易日期") or ""
        if a not in bank_last or d >= bank_last[a][0]:
            bank_last[a] = (d, r.get("余额"))
    # 未达账项：来自逐笔稽核
    from collections import defaultdict
    adj = defaultdict(lambda: {"单边收": 0.0, "单边支": 0.0, "未记收": 0.0, "未记支": 0.0})
    recon = _cache_get(_RECON_CACHE, _reconcile)
    for r in recon.get("results", []):
        a = al.norm_acct(r.get("账号") or "")
        if not a:
            continue
        v = r["借方金额"] if r["借方金额"] is not None else (r["贷方金额"] or 0)
        st = r["status"]
        if st in ("kd_only", "kd_xfer"):   # 内部划转·对应他账户：本户金蝶已记、本户银行未见 → 口径同金蝶单边，仍计银行侧未达
            adj[a]["单边收" if r["方向"] == "收" else "单边支"] += (v or 0)
        elif st in ("bank_leak", "xfer_unbooked"):
            adj[a]["未记收" if r["方向"] == "收" else "未记支"] += (v or 0)
    notes = db.list_balance_notes(CFG["year"], CFG["period"])   # 会计填的未达原因，按账号
    accounts, ping = [], 0
    # 以银行为锚点：银行流水是真实的，金蝶账面要向银行看齐。（V2.36 口径改）
    _SEV = {"不明差异": 0, "待金蝶更正": 1, "缺账面": 2, "账实相符": 3}
    for a in sorted(bank_last):
        bb = bank_last[a][1]              # 银行对账单余额（锚点·真实）
        kb = book.get(a)                 # 金蝶账面余额（可能虚高/虚低）
        kb_base = book_base.get(a)
        j = adj[a]
        # 金蝶待更正(疑似做错)＝金蝶单边净：金蝶已记、银行没有 → 很大概率金蝶做错(重复/错账户/错记)，金蝶应冲/改。
        # 金蝶应补记(漏账/内部往来未做账)＝银行已记、金蝶未记 → 金蝶应补做账。二者都是「金蝶向银行看齐」。
        kd_fix = round(j["单边收"] - j["单边支"], 2)
        kd_add = round(j["未记收"] - j["未记支"], 2)
        book_fixed = round((kb or 0.0) - kd_fix + kd_add, 2) if kb is not None else None   # 金蝶冲错+补漏后
        diff = round(book_fixed - bb, 2) if book_fixed is not None else None                # 对银行差额，应=0
        if kb is None:
            status = "缺账面"
        elif abs(kd_fix) > 0.01:
            status = "待金蝶更正"        # 有金蝶单边=疑似做错，倒逼金蝶冲/改 → 不算对平
        elif diff is not None and abs(diff) > 0.01:
            status = "不明差异"          # 冲错补漏后仍≠银行，真·不明差异待查
        else:
            status = "账实相符"          # 金蝶(经漏账时间差补记后)=银行，且无金蝶单边错误
        ok = (status == "账实相符")
        ping += 1 if ok else 0
        sub, bank, acct_name, cur = _acct_info(a)
        cur = cur or acct_kd_cur.get(a, "") or "人民币"        # 台账币别 → 金蝶兜底 → 默认人民币
        nt = notes.get(a, {})
        is_rmb_acct = cur in ("人民币", "CNY", "RMB", "")
        kb_base_out = (round(kb_base, 2) if (kb_base is not None and not is_rmb_acct) else None)
        accounts.append({
            "账号": a, "主体": sub, "开户行": bank, "账户名称": acct_name, "币别": cur,
            "银行对账单余额": round(bb, 2), "金蝶账面余额": (round(kb, 2) if kb is not None else None),
            "金蝶账面本位币": kb_base_out,
            "金蝶待更正": kd_fix, "金蝶应补记": kd_add,
            "更正后账面": book_fixed, "对银行差额": diff,
            "状态": status, "对平": ok, "有未达": bool(kd_fix or kd_add),
            # 兼容旧字段（导出/明细展开用）
            "企业已记银行未记_收": round(j["单边收"], 2), "企业已记银行未记_支": round(j["单边支"], 2),
            "银行已记企业未记_收": round(j["未记收"], 2), "银行已记企业未记_支": round(j["未记支"], 2),
            "差额": diff,
            "未达原因": nt.get("note", ""), "原因填写人": nt.get("operator", ""), "原因时间": nt.get("ts", ""),
        })
    accounts.sort(key=lambda x: (_SEV.get(x["状态"], 9), -abs(x.get("对银行差额") or 0)))   # 待处理的排前

    # ---------- 钩稽关系（做法2：把核对链摊开，每环带勾稽结果）----------
    g = recon.get("guardrail", {}) or {}
    covered = set(bank_last.keys())
    # 科目余额表 1002 口径：book=期初+本期序时账净(与「科目余额表」视图同算法)，有余额的户
    kd_accts = {a for a in book if abs(book.get(a, 0.0)) > 0.01}
    # 未纳入余额调节的户：金蝶 1002 有余额、但没导银行流水(未做银行余额调节)——审计缺口
    not_covered = []
    for a in sorted(kd_accts - covered):
        sub, bank, acct_name, cur = _acct_info(a)
        not_covered.append({"账号": a, "主体": sub, "开户行": bank,
                            "币别": cur or acct_kd_cur.get(a, "") or "人民币",
                            "金蝶账面": round(book.get(a, 0.0), 2)})
    from collections import Counter as _Counter
    scnt = _Counter(x["状态"] for x in accounts)
    tie = {
        # 环1 笔数勾稽（账证）：银行/金蝶每一笔都归了一类
        "笔数": {"银行笔数": g.get("银行笔数"), "银行已归类": g.get("银行已归类"),
                 "金蝶笔数": g.get("金蝶笔数"), "金蝶已归类": g.get("金蝶已归类"),
                 "银行对平": bool(g.get("银行笔数核对一致")), "金蝶对平": bool(g.get("金蝶笔数核对一致"))},
        # 环2 余额调节勾稽（账实·以银行为锚点）：金蝶更正后应=银行余额
        "余额调节": {"覆盖户数": len(accounts), "账实相符": scnt.get("账实相符", 0),
                     "待金蝶更正": scnt.get("待金蝶更正", 0), "不明差异": scnt.get("不明差异", 0),
                     "缺账面": scnt.get("缺账面", 0),
                     "对平户数": ping, "不平户数": len(accounts) - ping},
        # 环3 科目余额表勾稽（账账）：1002 有余额户 = 已调节户 + 未纳入户
        "科目余额表": {"有余额户数": len(kd_accts), "已纳入调节": len(kd_accts & covered),
                       "未纳入户数": len(not_covered), "未纳入户": not_covered,
                       "口径": "期初+本期序时账净（与「科目余额表」视图一致）"},
    }
    return {**base, "accounts": accounts, "对平户数": ping, "不平户数": len(accounts) - ping, "钩稽": tie}


def _channel_adjust():
    """第三方渠道(支付宝等,1012)余额勾稽：渠道对账单期末余额+本期收支 vs 金蝶1012该维度账面(期初+序时账)。
    逐笔对不了(海量微交易 vs 汇总)，故核对总额：本期净是否一致 + 期末余额差(=期初跨期差)。"""
    base = {"source": CFG["source"], "period": _period_str(), "updated_at": _now()}
    if CFG["source"] != "kingdee":
        return {**base, "channels": [], "note": "样例数据"}
    rec = db.get_period_input(CFG["source"], CFG["year"], CFG["period"], "bank")
    channels = (rec["payload"].get("channels") if rec else None) or []   # 上传时已解析并定格
    try:
        vou = _kd_get("gl_voucher:1012")
        bal = _kd_get("gl_balance")
    except KdNotFetched:
        return {**base, "未取数": True, "channels": [], "note": "本期未取数：请点【刷新金蝶数据】"}
    except kc.KingdeeError as e:
        return {**base, "error": str(e), "channels": []}
    # 金蝶 1012 各维度：期初(去重) + 本期序时账净
    kd_open, _seen = {}, set()
    for r in bal:
        code = str(r.get("科目编码") or "")
        if not code.startswith("1012"):
            continue
        dim = str(r.get("核算维度.银行账号.编码") or "")
        key = (code, dim)
        if key in _seen or not dim:
            continue
        _seen.add(key)
        kd_open[dim] = kd_open.get(dim, 0.0) + rc.to_float(r.get("期初原币") or 0)
    kd_net = {}
    for r in vou:
        dim = str(r.get("核算维度.银行账号.编码") or r.get("FDetailID.FF100002.FNumber") or "")
        if not dim:
            continue
        kd_net[dim] = kd_net.get(dim, 0.0) + rc.kd_delta_for(r)   # 原币口径（1012 全人民币，两口径同值）
    kd = {d: {"账面": round(kd_open.get(d, 0.0) + kd_net.get(d, 0.0), 2), "净": round(kd_net.get(d, 0.0), 2)}
          for d in set(kd_open) | set(kd_net)}
    out = []
    for ch in channels:
        email = ch.get("email") or ""
        net = round(ch["收"] - ch["支"], 2)
        cands = [d for d in kd if email and email in d]
        match = next((d for d in cands if abs(kd[d]["净"] - net) < 0.01), (cands[0] if cands else None))
        ki = kd.get(match) if match else None
        out.append({
            "渠道": ch.get("渠道", "支付宝"), "支付宝账户": email or ch.get("acct", ""), "笔数": ch["笔数"],
            "渠道期末余额": ch["期末余额"], "本期收": ch["收"], "本期支": ch["支"], "本期净": net,
            "金蝶维度": match or "(未匹配金蝶维度)", "金蝶账面": (ki["账面"] if ki else None), "金蝶本期净": (ki["净"] if ki else None),
            "净一致": bool(ki and abs(ki["净"] - net) < 0.01),
            "余额差": (round(ch["期末余额"] - ki["账面"], 2) if (ch["期末余额"] is not None and ki) else None),
        })
    out.sort(key=lambda x: (x["净一致"], -x["笔数"]))
    return {**base, "channels": out,
            "净一致户数": sum(1 for r in out if r["净一致"]), "总户数": len(out)}


@app.get("/api/channel-adjust")
def channel_adjust():
    return _cache_get(_CH_CACHE, _channel_adjust)


@app.post("/api/channel-adjust/sync")
def channel_adjust_sync():
    return _closed_block() or _cache_get(_CH_CACHE, _channel_adjust, force=True)


def _recon_data():
    """逐笔稽核结果。已封存 → 读快照（数字定格，不再碰金蝶）；进行中 → 实时算 + 叠加认领态。"""
    if _is_closed():
        d = db.load_snapshot(CFG["source"], CFG["year"], CFG["period"], "reconcile")
        if d:
            return {**d, "封存": True, "封存信息": _closed_info(), "cached": True}
    d = _cache_get(_RECON_CACHE, _reconcile)             # 读缓存秒开
    _overlay_claims(d.get("results", []))                # 认领态现读，即时生效
    return {**d, "封存": False}


def _badj_data():
    """余额调节表（含钩稽）。已封存 → 读快照。"""
    if _is_closed():
        d = db.load_snapshot(CFG["source"], CFG["year"], CFG["period"], "balance_adjust")
        if d:
            return {**d, "封存": True, "封存信息": _closed_info(), "cached": True}
    return {**_cache_get(_BADJ_CACHE, _balance_adjust), "封存": False}


@app.get("/api/reconcile")
def reconcile():
    return _recon_data()


@app.post("/api/reconcile/sync")
def reconcile_sync():
    blocked = _closed_block()
    if blocked:
        return blocked
    d = _cache_get(_RECON_CACHE, _reconcile, force=True)
    _overlay_claims(d.get("results", []))
    return {**d, "封存": False}


@app.get("/api/operators")
def operators():
    return {"roster": ROSTER}


@app.post("/api/reconcile/claim")
def reconcile_claim(body: dict, request: Request):
    """认领工作流：key + action(认领/已调整/识别有误/撤销)。操作人=登录用户(防冒名)。"""
    u = _current_user(request)
    if not db.user_can(u, "claim"):
        return JSONResponse({"ok": False, "msg": "无「认领/处理差异」权限，请联系管理员"}, status_code=403)
    blocked = _closed_block()          # 已封存的期间：认领态定格，不许再动
    if blocked:
        return blocked
    op = u["name"] if u else ""
    key = str(body.get("key", "") or "").strip()
    action = body.get("action")
    if not key or action not in ("认领", "已调整", "识别有误", "撤销"):
        return {"ok": False, "msg": "参数不全"}
    note = str(body.get("备注", "") or "")
    if action == "撤销":
        db.del_claim(key)
        db.audit(op, "撤销认领", key)
        return {"ok": True, "key": key, "claim": None}
    st = {"认领": "已认领", "已调整": "已调整", "识别有误": "识别有误"}[action]
    ts = _now()
    db.set_claim(key, st, op, ts, note)
    db.audit(op, action, key, note)      # 审计留痕（财务必备）
    return {"ok": True, "key": key, "claim": {"状态": st, "操作人": op, "时间": ts, "备注": note}}


def _build_report_xlsx(recon, badj):
    """对账底稿 xlsx（4 表：对账汇总 / 差异清单 / 余额调节表 / 全部逐笔明细）→ bytes。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="4B5563")
    title_font = Font(bold=True, size=13)

    def style_header(ws, ncol):
        for c in range(1, ncol + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"

    def autowidth(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    wb = Workbook()
    # ① 对账汇总
    ws0 = wb.active; ws0.title = "对账汇总"
    ws0.append(["银行—金蝶对账底稿"]); ws0["A1"].font = title_font
    ws0.append(["期间", _period_str(), "生成时间", _now(), "数据源", CFG["source"]])
    if recon.get("封存"):
        fi = recon.get("封存信息") or {}
        ws0.append(["本期状态", "已封存（数据定格，下方为封存当时的结果）",
                    "封存人", fi.get("封存人", ""), "封存时间", fi.get("封存时间", "")])
        ws0.append(["金蝶取数时点", fi.get("金蝶取数时点", "")])
    else:
        ws0.append(["本期状态", "进行中（数据随金蝶变动，未封存）"])
    ws0.append([])
    ws0.append(["逐笔稽核 · 各状态笔数"]); ws0.cell(row=ws0.max_row, column=1).font = Font(bold=True)
    for k, v in (recon.get("summary") or {}).items():
        ws0.append([k, v])
    ws0.append([])
    g = recon.get("guardrail") or {}
    ws0.append(["护栏 · 银行笔数", g.get("银行笔数"), "金蝶笔数", g.get("金蝶笔数"),
                "两侧各归一类", "是" if (g.get("银行笔数核对一致") and g.get("金蝶笔数核对一致")) else "否"])
    ws0.append(["余额调节 · 对平户数", badj.get("对平户数", 0), "不平户数", badj.get("不平户数", 0)])
    autowidth(ws0, [22, 16, 14, 16, 12, 10])

    # ② 差异清单（非"已匹配"的逐笔，即需人工处理的）
    # 序号=在稽核结果里的全局位置(与逐笔稽核页「序号」一致，办公室对号沟通用)
    cols = ["序号", "状态", "日期", "金蝶日期", "账号", "开户行", "主体", "方向", "借方金额", "贷方金额",
            "对方户名", "摘要", "金蝶凭证", "制单人", "差额", "内部往来对应", "组合候选说明"]
    dcols = cols[1:]
    results = recon.get("results") or []
    ws1 = wb.create_sheet("差异清单")
    ws1.append(cols); style_header(ws1, len(cols))
    for i, r in enumerate(results, 1):
        if r.get("status") != "matched":
            ws1.append([i] + [r.get(c) for c in dcols])
    autowidth(ws1, [7, 14, 11, 11, 20, 12, 18, 6, 14, 14, 18, 30, 12, 10, 12, 22, 20])

    # ③ 余额调节表
    # 以银行为锚点：银行对账单余额=真实，金蝶更正后应=银行。
    bcols = ["主体", "账户名称", "开户行", "账号", "币别", "银行对账单余额", "金蝶账面余额", "金蝶账面本位币",
             "金蝶待更正", "金蝶应补记", "更正后账面", "对银行差额", "状态", "未达原因", "原因填写人"]
    ws2 = wb.create_sheet("余额调节表")
    ws2.append(bcols); style_header(ws2, len(bcols))
    for a in (badj.get("accounts") or []):
        ws2.append([a.get(c) for c in bcols])
    autowidth(ws2, [18, 22, 10, 20, 8, 16, 16, 16, 16, 16, 16, 14, 12, 40, 12])

    # ④ 全部逐笔明细
    ws3 = wb.create_sheet("全部逐笔明细")
    ws3.append(cols); style_header(ws3, len(cols))
    for i, r in enumerate(results, 1):
        ws3.append([i] + [r.get(c) for c in dcols])
    autowidth(ws3, [7, 14, 11, 11, 20, 12, 18, 6, 14, 14, 18, 30, 12, 12, 22, 20])

    bio = BytesIO(); wb.save(bio); return bio.getvalue()


@app.get("/api/export/report")
def export_report(request: Request):
    """结果出具：导出对账底稿 xlsx（浏览器直接下载）。已封存期间导出的是封存那一刻的快照。"""
    recon = _recon_data()
    badj = _badj_data()
    data = _build_report_xlsx(recon, badj)
    u = _current_user(request)
    if u:
        db.audit(u["name"], "导出对账底稿", _period_str())     # 月结看板据此判断「底稿已导出」
    fname = f"对账底稿_{_period_str()}.xlsx"
    disp = "attachment; filename=report.xlsx; filename*=UTF-8''" + urllib.parse.quote(fname)
    return Response(content=data,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": disp})


@app.get("/api/balance-adjust")
def balance_adjust():
    return _badj_data()


@app.post("/api/balance-adjust/sync")
def balance_adjust_sync():
    blocked = _closed_block()
    if blocked:
        return blocked
    return {**_cache_get(_BADJ_CACHE, _balance_adjust, force=True), "封存": False}


@app.post("/api/balance-adjust/note")
def balance_adjust_note(body: dict, request: Request):
    """会计填/改某账户的未达原因（供领导核查）。用「认领/处理差异」权限；填写人服务端认登录用户。"""
    u = _require_perm(request, "claim")
    if not u:
        return JSONResponse({"ok": False, "msg": "无「认领/处理差异」权限，不能填未达原因"}, status_code=403)
    blocked = _closed_block()
    if blocked:
        return blocked
    acct = str(body.get("acct", "") or "").strip()
    if not acct:
        return {"ok": False, "msg": "缺账号"}
    note = str(body.get("note", "") or "")
    db.set_balance_note(CFG["year"], CFG["period"], acct, note, u["name"])
    db.audit(u["name"], "余额调节-未达原因", acct, note[:80])
    _BADJ_CACHE.clear()          # 缓存里带的是旧原因，清掉让下次读到最新
    return {"ok": True, "operator": u["name"], "ts": db._now()}


# ==================== 月结看板 / 期间封存 ====================
# 月结是总纲，银行对账、物流计提都是它下面的一环；看板=本月各环做到哪一步，封存=全做完、结果存死、本期只读。
# 成本台账不进清单：它属于结账【后】的报表系列，不是封账前的必做项（业务方 2026-07-10 定）。
# 这几项不过不让封存（其余项只提示）：金蝶数据 / 银行流水 / 护栏 / 差异处理 / 余额对平 / 无待更正
_CL_BLOCK = ("kd", "bank", "guard", "claim", "badj", "fix")
_CL_GROUPS = ("数据准备", "环① 银行对账", "环② 物流计提")     # 看板分组顺序


def _checklist():
    """月结清单：ok=已完成 / warn=提示(不挡封存) / fail=未完成(挡封存)。"""
    recon, badj = _recon_data(), _badj_data()
    g = recon.get("guardrail") or {}
    tie = badj.get("钩稽") or {}
    t_bal, t_sub = (tie.get("余额调节") or {}), (tie.get("科目余额表") or {})
    results = recon.get("results") or []
    manifest = recon.get("manifest") or []
    y, p = int(CFG["year"]), int(CFG["period"])

    diffs = [r for r in results if r.get("status") != "matched"]
    undone = [r for r in diffs if r.get("认领状态") not in ("已调整", "识别有误")]
    unping = int(badj.get("不平户数") or 0)
    fixcnt = int(t_bal.get("待金蝶更正") or 0)
    nocov = int(t_sub.get("未纳入户数") or 0)
    posts = len(db.list_logistics_posts(y, p) or [])
    exported = db.audit_exists("导出对账底稿", _period_str())
    sync_at = _closed_info().get("金蝶取数时点") if _is_closed() else _kd_synced_at()
    is_kd = (CFG["source"] == "kingdee")
    guard_ok = bool(g.get("银行笔数核对一致") and g.get("金蝶笔数核对一致"))

    def it(grp, key, title, ok, txt, warn=False):
        return {"环": grp, "key": key, "标题": title, "说明": txt,
                "状态": "ok" if ok else ("warn" if warn else "fail"),
                "阻断封存": (key in _CL_BLOCK) and not ok}

    G0, G1, G2 = _CL_GROUPS
    items = [
        it(G0, "kd", "金蝶数据已刷新", is_kd and bool(sync_at),
           ("数据取自 " + sync_at) if (is_kd and sync_at) else
           ("当前用的是样例数据，不能封存" if not is_kd else "本期还没从金蝶取过数")),
        it(G0, "bank", "银行流水已导入", bool(manifest),
           ("已并入 %d 个流水文件" % len(manifest)) if manifest else "还没上传银行流水包"),

        it(G1, "guard", "逐笔稽核护栏通过", guard_ok,
           ("银行 %s 笔 / 金蝶 %s 笔，每一笔都归了类，不重不漏" % (g.get("银行笔数"), g.get("金蝶笔数")))
           if guard_ok else "有笔数没归类，稽核结果不可信"),
        it(G1, "claim", "差异已全部处理", not undone,
           ("%d 笔差异全部处理完" % len(diffs)) if not undone else
           ("还有 %d 笔没处理（本期共 %d 笔差异）" % (len(undone), len(diffs)))),
        it(G1, "badj", "余额调节已对平", unping == 0,
           ("%d 个账户全部账实相符" % (t_bal.get("覆盖户数") or 0)) if unping == 0 else
           ("还有 %d 个账户对不上银行余额" % unping)),
        it(G1, "fix", "无「金蝶待更正」账户", fixcnt == 0,
           "没有疑似做错的金蝶单边" if fixcnt == 0 else
           ("有 %d 个账户等金蝶更正（疑似做错账户/重复记账），改完再封" % fixcnt)),
        it(G1, "cover", "科目余额表无缺口", nocov == 0,
           "1002 有余额的账户都纳入了余额调节" if nocov == 0 else
           ("%d 个账户金蝶有余额、但没导银行流水，未做调节" % nocov), warn=True),
        it(G1, "export", "对账底稿已导出", exported,
           "已导出过对账底稿" if exported else "还没导出底稿（封存后仍可从快照导出）", warn=True),

        it(G2, "accrual", "计提凭证已录入", posts > 0,
           ("本期已录 %d 张计提凭证" % posts) if posts else "本期没有经工具录入的计提凭证（本月无需计提可忽略）",
           warn=True),
    ]
    blockers = [i for i in items if i["阻断封存"]]
    return {"期间": _period_str(), "数据源": CFG["source"], "items": items,
            "分组": list(_CL_GROUPS),
            "可封存": not blockers, "未完成项": [i["标题"] for i in blockers],
            "金蝶取数时点": sync_at or ""}


@app.get("/api/period")
def period_state():
    """月结看板：本期状态 + 清单 + 历史封存记录。
    已封存 → 直接读封存那一刻的清单快照（定格；也不必重算，服务器重启后照样准）。"""
    cl = db.load_snapshot(CFG["source"], CFG["year"], CFG["period"], "checklist") if _is_closed() else None
    return {"year": CFG["year"], "period": CFG["period"], "period_str": _period_str(),
            "source": CFG["source"], "状态": _closed_info(), "清单": cl or _checklist(),
            "历史": db.list_closed_periods(CFG["source"])}


@app.post("/api/period/close")
def period_close(body: dict, request: Request):
    """封存本期：清单全绿 → 把结果拍照落库 → 本期转只读。
    清单没全绿时，主管理员/子管理员可填理由强制封存（force=true）——真实月结总有"就是有个说不清的差异"。"""
    u = _require_perm(request, "close_period")
    if not u:
        return JSONResponse({"ok": False, "msg": "无「月结看板·封存本期」权限，请联系管理员"}, status_code=403)
    if _is_closed():
        return {"ok": False, "msg": "本期已封存，无需重复操作"}
    if CFG["source"] != "kingdee":
        return {"ok": False, "msg": "样例数据不能封存，请先在「设置」把数据源切到金蝶"}
    note = str(body.get("note", "") or "").strip()
    cl = _checklist()
    if not cl["可封存"]:
        if not body.get("force"):
            return {"ok": False, "msg": "月结清单还有未完成项：" + "、".join(cl["未完成项"]), "清单": cl}
        if not db.can_admin_accounts(u):
            return JSONResponse({"ok": False, "msg": "清单未全绿，只有主管理员/核算工作台子管理员能强制封存"},
                                status_code=403)
        if len(note) < 5:
            return {"ok": False, "msg": "强制封存必须填理由（至少5个字），供领导与审计核查"}
        note = "【强制封存】未完成：%s。理由：%s" % ("、".join(cl["未完成项"]), note)

    # 先存快照、再落封存标记：万一中途挂了，本期仍是进行中，不会出现"已封存却读到空快照"
    recon, badj = _recon_data(), _badj_data()
    sync_at = _kd_synced_at()
    n = db.save_snapshot(CFG["source"], CFG["year"], CFG["period"], "reconcile", recon)
    n += db.save_snapshot(CFG["source"], CFG["year"], CFG["period"], "balance_adjust", badj)
    db.save_snapshot(CFG["source"], CFG["year"], CFG["period"], "checklist", cl)
    info = db.close_period(CFG["source"], CFG["year"], CFG["period"], u["name"], note, sync_at)
    db.audit(u["name"], "月结封存", _period_str(), note or "清单全绿")
    _cache_clear()
    return {"ok": True, "msg": "%s 已封存，本期转为只读。" % _period_str(), "状态": info, "快照字节": n}


@app.post("/api/period/reopen")
def period_reopen(body: dict, request: Request):
    """解封：高危操作，限主管理员/核算工作台子管理员，必须填理由，全程留痕。"""
    u = _current_user(request)
    if not u or not (db.is_super(u) or db.user_can(u, "manage_accounting")):
        return JSONResponse({"ok": False, "msg": "解封是高危操作，只有主管理员或核算工作台子管理员可以执行"},
                            status_code=403)
    if not _is_closed():
        return {"ok": False, "msg": "本期是进行中，无需解封"}
    reason = str(body.get("reason", "") or "").strip()
    if len(reason) < 5:
        return {"ok": False, "msg": "解封必须填写理由（至少5个字），供领导与审计核查"}
    info = db.reopen_period(CFG["source"], CFG["year"], CFG["period"], u["name"], reason)
    db.audit(u["name"], "月结解封", _period_str(), reason)
    _cache_clear()      # 解封后回到实时算：重新从金蝶取数
    return {"ok": True, "msg": "%s 已解封，恢复可编辑。改完记得重新封存。" % _period_str(), "状态": info}


@app.get("/api/data-sources")
def data_sources():
    return _cache_get(_DS_CACHE, _data_sources)


@app.post("/api/data-sources/sync")
def data_sources_sync():
    return _closed_block() or _cache_get(_DS_CACHE, _data_sources, force=True)


# ---------------- 科目余额表视图（账面核对：API 取数 + 手工上传金蝶报表逐项核对）----------------
def _subject_balance():
    """科目余额表（四类资金科目）。金蝶模式：期初(GL_BALANCE去重) + 本期序时账借贷 = 实时期末
    （与金蝶科目余额表报表同公式；接口"期末"字段凭证未过账时停在期初，不能直接用，见 V1.7）。"""
    base = {"source": CFG["source"], "period": _period_str(), "updated_at": _now()}
    if CFG["source"] != "kingdee":
        return {**base, "rows": sb.build_rows_sample(S.sample_balance_rows()), "note": "样例数据"}
    try:
        bal = _kd_get("gl_balance")
        vou = _kd_get("gl_subjects")
    except KdNotFetched:
        return {**base, "未取数": True, "rows": [], "note": "本期未取数：请到「数据接入」点【刷新金蝶数据】"}
    except kc.KingdeeError as e:
        return {**base, "error": str(e), "rows": []}
    return {**base, "rows": sb.build_rows_kingdee(bal, vou)}


@app.get("/api/subject-balance")
def subject_balance():
    return _cache_get(_SBAL_CACHE, _subject_balance)


@app.post("/api/subject-balance/sync")
def subject_balance_sync():
    return _closed_block() or _cache_get(_SBAL_CACHE, _subject_balance, force=True)


def _sbal_upload_path():
    return os.path.join(UPLOAD_DIR, "科目余额表_核对上传.xlsx")


def _sbal_check():
    """人眼核对：解析已上传的金蝶科目余额表 → 与工具数逐科目对照。没上传过返回 compare=None。"""
    p = _sbal_upload_path()
    if not os.path.exists(p):
        return {"ok": True, "compare": None}
    uploaded, err = sb.parse_report_xlsx(p)
    if err:
        return {"ok": False, "msg": err}
    tool = _cache_get(_SBAL_CACHE, _subject_balance)
    cmp_ = sb.compare(tool.get("rows", []), uploaded)
    return {"ok": True, "compare": cmp_, "period": _period_str(),
            "uploaded_at": datetime.datetime.fromtimestamp(os.path.getmtime(p)).strftime("%Y-%m-%d %H:%M")}


@app.get("/api/subject-balance/check")
def subject_balance_check():
    return _sbal_check()


@app.post("/api/subject-balance/upload")
async def subject_balance_upload(request: Request):
    """上传金蝶导出的《科目余额表》Excel（原始字节走请求体），存档后立即返回逐科目核对结果。"""
    blocked = _closed_block()
    if blocked:
        return blocked
    if not _require_perm(request, "subject_upload"):
        return JSONResponse({"ok": False, "msg": "无「上传科目余额表」权限，请联系管理员"}, status_code=403)
    data = await request.body()
    if not data:
        return {"ok": False, "msg": "空文件（请选择金蝶导出的科目余额表 Excel）"}
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    with open(_sbal_upload_path(), "wb") as f:
        f.write(data)
    return _sbal_check()


# ---------------- 理财产品对账（1101/1012理财腿 + 6xxx收益，产品维度聚合；含 PDF OCR）----------------
_WSTMT_PARSE_CACHE: dict = {}   # (路径, mtime) -> 解析记录：文件没变不重复 OCR


def _parse_wealth_pdf(path):
    key = (path, os.path.getmtime(path))
    if key in _WSTMT_PARSE_CACHE:
        return _WSTMT_PARSE_CACHE[key]
    rec = wstmt.parse(pdf_path=path)
    rec["_文件"] = os.path.basename(path)
    _WSTMT_PARSE_CACHE[key] = rec
    return rec


def _wealth_recon():
    """理财对账：流水目录里的理财对账单 PDF(OCR解析) × 金蝶 1101/1012理财腿 + 6xxx收益 → 产品维度勾稽。"""
    base = {"source": CFG["source"], "period": _period_str(), "updated_at": _now()}
    if CFG["source"] != "kingdee":
        return {**base, "rows": [], "产品数": 0, "已勾稽": 0, "有差异": 0, "note": "理财对账仅金蝶模式可用"}
    bdir = str(CFG.get("bank_import_dir", "") or "")
    pdfs = wstmt.find_wealth_pdfs(bdir) if os.path.isdir(bdir) else []
    stmts, parse_errs = [], []
    for p in pdfs:
        try:
            stmts.append(_parse_wealth_pdf(p))
        except wstmt.OCRUnavailable as e:
            return {**base, "error": f"OCR 组件缺失：{e}", "rows": [], "产品数": 0}
        except Exception as e:
            parse_errs.append({"文件": os.path.basename(p), "错误": str(e)[:120]})
    try:
        vou = kc.fetch_gl_voucher_subjects(CFG["year"], CFG["period"])
        inc = kc.fetch_gl_voucher_income(CFG["year"], CFG["period"])
    except kc.KingdeeError as e:
        return {**base, "error": str(e), "rows": [], "产品数": 0}
    kd = wr.kd_wealth_by_product(vou, inc)
    res = wr.reconcile_wealth(stmts, kd)
    return {**base, **res, "对账单文件": [os.path.basename(p) for p in pdfs],
            "对账单笔数": sum(len(s.get("交易明细", [])) for s in stmts), "parse_errors": parse_errs}


@app.get("/api/wealth-recon")
def wealth_recon_get():
    """读缓存秒回；不触发 OCR。首次/未对账返回占位，点『开始理财对账』(POST /sync) 才跑 OCR。"""
    key = _cache_key()
    if key in _WR_CACHE:
        d = dict(_WR_CACHE[key]); d["cached"] = True
        return d
    return {"source": CFG["source"], "period": _period_str(), "rows": [], "产品数": 0,
            "未对账": True, "note": "理财对账含理财对账单 PDF 的 OCR 识别（较慢），点『开始理财对账』运行。"}


@app.post("/api/wealth-recon/sync")
def wealth_recon_sync():
    """跑理财对账（OCR 解析理财 PDF + 金蝶取数 + 产品维度勾稽）。"""
    blocked = _closed_block()
    if blocked:
        return blocked
    d = _wealth_recon()
    _WR_CACHE[_cache_key()] = d
    out = dict(d); out["cached"] = False
    return out


UPLOAD_DIR = os.path.join(BASE, "bank_uploads")


@app.post("/api/bank-import/upload")
async def bank_import_upload(request: Request):
    """出纳流水包(zip) 上传 → 解压到本期专属目录 → 解析并【按期间定格入库】→ 返回解析清单。
    上传一次就定格：之后各页直接读库里这份，不再每次进来重新解析。原始 zip 字节走请求体(免 multipart 依赖)。"""
    u = _require_perm(request, "bank_upload")
    if not u:
        return JSONResponse({"ok": False, "msg": "无「上传资金流水」权限，请联系管理员"}, status_code=403)
    blocked = _closed_block()
    if blocked:
        return blocked
    data = await request.body()
    if not data:
        return {"ok": False, "msg": "空文件（请选择出纳导出的流水压缩包）"}
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    tag = "%s_%d_%02d" % (CFG["source"], CFG["year"], CFG["period"])   # 每期一个目录，不互相覆盖
    zip_path = os.path.join(UPLOAD_DIR, tag + ".zip")
    with open(zip_path, "wb") as f:
        f.write(data)
    extract_dir = os.path.join(UPLOAD_DIR, tag, "extracted")
    if os.path.isdir(extract_dir):
        shutil.rmtree(extract_dir, ignore_errors=True)
    # 密码走请求头(不进 URL/日志)；前端 encodeURIComponent 过，这里还原
    from urllib.parse import unquote
    pwd = unquote(request.headers.get("x-zip-password", "")) or None
    try:
        bimp.extract_archive(zip_path, extract_dir, password=pwd)   # 自动识别 zip / rar
    except Exception as e:
        s = str(e).lower()
        if "no_rar_tool" in s:
            return {"ok": False, "msg": "这是 RAR 包，服务器未装解压工具。请安装 7-Zip（Windows）/ p7zip（Linux）后重试；或把流水包改存为 ZIP。"}
        if pwd and ("wrong password" in s or "password" in s or "bad" in s or "mac" in s):
            return {"ok": False, "msg": "解压失败：压缩包密码错误（RAR/ZIP 均已尝试）"}
        if (not pwd) and ("password" in s or "encrypt" in s):
            return {"ok": False, "msg": "该压缩包已加密，请在下方「压缩包密码」填入密码后再上传"}
        return {"ok": False, "msg": f"解压失败：{e}"}
    # 解析并按期间定格入库：逐笔 rows + manifest + 第三方渠道 channels，一次算好存死
    rows, manifest = bimp.load_bank_dir(extract_dir)
    try:
        channels = bimp.parse_channels(extract_dir)
    except Exception:
        channels = []
    # 侦测流水实际月份（交易日期占比最高的 YYYY-MM），跟所选期间对不上就警告——防"选4月却传了6月流水"
    from collections import Counter as _C
    ymc = _C()
    for r in rows:
        ym = str(r.get("交易日期") or "")[:7]
        if len(ym) == 7 and ym[4] == "-":
            ymc[ym] += 1
    bank_ym = ymc.most_common(1)[0][0] if ymc else ""
    sel_ym = _period_str()
    mismatch = bool(bank_ym and bank_ym != sel_ym)
    # 财资多份导出出现让位/重复/疑漏 → 打"重复待确认"标记：查重只是系统初核，须人工弹窗确认留痕
    dup_pending = bimp.needs_dup_confirm(manifest)
    meta = {"笔数": len(rows), "文件数": len(manifest), "流水月份": bank_ym}
    if dup_pending:
        meta["重复待确认"] = True
    db.set_period_input(CFG["source"], CFG["year"], CFG["period"], "bank",
                        {"rows": rows, "manifest": manifest, "channels": channels, "dir": extract_dir},
                        meta, u["name"])
    db.audit(u["name"], "上传银行流水", _period_str(),
             "并入 %d 笔 / %d 个文件%s%s" % (len(rows), len(manifest),
                                            ("（⚠流水月份=%s）" % bank_ym) if mismatch else "",
                                            "（财资重复判定待人工确认）" if dup_pending else ""))
    CFG["bank_import_dir"] = extract_dir     # 兼容旧字段（渠道/理财等仍读它当本期目录）
    save_cfg(CFG)
    _cache_clear()
    return {"ok": True, "并入笔数": len(rows), "manifest": manifest, "dir": extract_dir,
            "period": sel_ym, "updated_by": u["name"], "updated_at": _now(),
            "bank_ym": bank_ym, "sel_ym": sel_ym, "period_mismatch": mismatch,
            "need_dup_confirm": dup_pending}


@app.post("/api/bank-import/confirm-dup")
async def bank_import_confirm_dup(request: Request):
    """人工确认财资归并的重复/让位判定（需求方 2026-09-01 定：系统查重只是初核，不独自背锅——
    必须有人点确认，确认人/时间写进本期 meta 并入审计；只改 meta，不动数据与上传留痕）。"""
    u = _require_perm(request, "bank_upload")
    if not u:
        return JSONResponse({"ok": False, "msg": "无「上传资金流水」权限，请联系管理员"}, status_code=403)
    rec = db.get_period_input(CFG["source"], CFG["year"], CFG["period"], "bank")
    if rec is None:
        return {"ok": False, "msg": "本期还没有已上传的银行流水"}
    if not (rec.get("meta") or {}).get("重复待确认"):
        return {"ok": True, "msg": "本期没有待确认的重复判定"}
    now = _now()
    db.update_period_input_meta(CFG["source"], CFG["year"], CFG["period"], "bank",
                                {"重复待确认": None, "重复确认人": u["name"], "重复确认时间": now})
    db.audit(u["name"], "确认财资重复", _period_str(), "人工确认多份财资导出的重复/让位判定（弹窗逐条核对后确认）")
    _cache_clear()
    return {"ok": True, "确认人": u["name"], "确认时间": now}


@app.post("/api/kingdee/refresh")
def kingdee_refresh(request: Request):
    """从金蝶总闸刷新：本期真取一次金蝶数据并【定格入库】——科目余额 / 1002序时账 / 1012序时账 / 四类科目序时账。
    这是【唯一】真去金蝶取数的地方；之后各页直接读库里这份，进页面不再打金蝶（取数一次就定格）。"""
    u = _require_perm(request, "kingdee_refresh")
    if not u:
        return JSONResponse({"ok": False, "msg": "无「从金蝶更新」权限，请联系管理员"}, status_code=403)
    if CFG["source"] != "kingdee":
        return {"ok": False, "msg": "当前是样例数据源，无需从金蝶取数（到「设置」切换到金蝶真数据）"}
    blocked = _closed_block()          # 已封存 = 数据定格，不再从金蝶取新数
    if blocked:
        return blocked
    y, p = CFG["year"], CFG["period"]
    jobs = [
        ("gl_balance", "科目余额", lambda: kc.fetch_gl_balance(y, p)),
        ("gl_voucher:1002", "1002序时账", lambda: kc.fetch_gl_voucher(y, p, "1002")),
        ("gl_voucher:1012", "1012序时账", lambda: kc.fetch_gl_voucher(y, p, "1012")),
        ("gl_subjects", "四类科目序时账", lambda: kc.fetch_gl_voucher_subjects(y, p)),
    ]
    out = {"updated": True, "period": _period_str(), "updated_at": _now(), "取回": {}, "errors": {}}
    for call_key, label, fn in jobs:
        try:
            out["取回"][label] = _kd_fetch_store(call_key, fn, u["name"])
        except kc.KingdeeError as e:
            out["errors"][label] = str(e)
    # 出纳管理·银行账号主数据：每个账号的 开户行 + 账户类型(基本户/一般户/通知存款/结构性存款…) + 类别(银行账户/电商渠道)，定格入库供资金看板读
    try:
        bm = al.kd_accounts_from_cn(kc.fetch_bank_accounts())
        bmap = {}
        for r in bm:
            ac = al.norm_acct(r.get("账号") or "")
            if ac:
                bmap[ac] = {"开户行": r.get("开户行") or "", "账户类型": r.get("账户类型") or "", "类别": r.get("类别") or ""}
        db.set_period_input(CFG["source"], y, p, "kd:bank_master", {"map": bmap}, {"账户数": len(bmap)}, u["name"])
        out["取回"]["出纳银行账号"] = len(bmap)
    except kc.KingdeeError as e:
        out["errors"]["出纳银行账号"] = str(e)
    _cache_clear()                     # 内存派生缓存作废，下次读到新定格的库数据
    db.audit(u["name"], "刷新金蝶数据", _period_str(),
             "；".join("%s=%s" % (k, v) for k, v in out["取回"].items()))
    out["ok"] = not out["errors"]
    return out


# ---------------- 基础数据 › 主体档案（平台级；编辑限主管理员） ----------------
# 这张表是「册号首段」与「金蝶账簿代码」的共同源头，改错一行的破坏面横跨凭证归档与物流计提，
# 故只有主管理员能写；其余角色只读（各工具的主体下拉框都从这里取）。
@app.get("/api/orgs")
def orgs_list(request: Request):
    """主体档案列表（登录即可看）。locked=该主体已有凭证册在册、简码不可再改。
    palette=标签纸色板（前端不另抄一份，色值口径以后端为准）。"""
    if not _current_user(request):
        return JSONResponse({"ok": False}, status_code=401)
    rows = db.list_orgs()
    for r in rows:
        r["locked"] = db.org_code_locked(r["short_name"])
    return {"ok": True, "orgs": rows,
            "palette": [{"name": n, "hex": h} for n, h in db.LABEL_PALETTE]}


@app.post("/api/orgs/save")
def orgs_save(body: dict, request: Request):
    u = _admin(request)
    if not u:
        return JSONResponse({"ok": False, "msg": "主体档案只有主管理员能改（它同时决定册号前缀与金蝶账簿）"},
                            status_code=403)
    try:
        oid = db.save_org(body, u["name"])
    except ValueError as e:
        return {"ok": False, "msg": str(e)}
    db.audit(u["name"], "主体档案-保存", str(body.get("short_name", "")),
             "简码=%s 账簿=%s" % (body.get("code", ""), body.get("book_code", "")))
    return {"ok": True, "id": oid}


@app.post("/api/orgs/delete")
def orgs_delete(body: dict, request: Request):
    u = _admin(request)
    if not u:
        return JSONResponse({"ok": False, "msg": "主体档案只有主管理员能改"}, status_code=403)
    if not body.get("id"):
        return {"ok": False, "msg": "缺 id"}
    try:
        db.delete_org(body["id"])
    except ValueError as e:
        return {"ok": False, "msg": str(e)}
    db.audit(u["name"], "主体档案-删除", str(body["id"]))
    return {"ok": True}


# ── 各工具线路由（V2.172 拆出）──────────────────────────────
# 一条工具线一个模块，改某条线只动 routers/<线>.py，app.py 不再是并行开发的冲突源。
# 必须在下面的 SPA 兜底路由之前注册：兜底吃掉所有非 /api 路径，注册晚了会被它抢走。
from routers import (logistics_accrual, archive, fxrate, logistics_recon, cost_ledger,
                     rptexport, report_dashboard, ec, llm_hub, temp_attendance, bom_quote)

app.include_router(logistics_accrual.router)
app.include_router(archive.router)
app.include_router(fxrate.router)
app.include_router(logistics_recon.router)
app.include_router(cost_ledger.router)
app.include_router(rptexport.router)
app.include_router(report_dashboard.router)
app.include_router(ec.router)
app.include_router(llm_hub.router)   # V2.301 门户模型配置 P0.5 聚合看板
app.include_router(temp_attendance.router)
app.include_router(bom_quote.router)   # V-draft BOM报价审核


# 托管 React 构建产物 (SPA: /api/* 优先; 真实静态文件直接给; 其余非API路径回退 index.html,
# 使前端子路径/刷新不再 404 —— {"detail":"Not Found"} 即此前缺兜底所致)
from fastapi.responses import FileResponse
INDEX = os.path.join(DIST, "index.html")
NOCACHE = {"Cache-Control": "no-cache, no-store, must-revalidate"}   # index.html 不缓存：前端更新即生效，避免浏览器旧缓存
if os.path.isdir(DIST) and os.path.exists(INDEX):
    @app.get("/")
    def _index():
        return FileResponse(INDEX, headers=NOCACHE)

    @app.get("/{full_path:path}")
    def _spa(full_path: str):
        # /api/* 已在上面定义、优先匹配; 这里只兜底其它 GET 路径
        cand = os.path.normpath(os.path.join(DIST, full_path))
        if full_path and cand.startswith(DIST) and os.path.isfile(cand):
            return FileResponse(cand)          # 真实静态文件(assets/js/css, 带 hash 可长缓存)
        return FileResponse(INDEX, headers=NOCACHE)   # 前端路由 -> 回 index.html（不缓存）
else:
    @app.get("/")
    def _root():
        return JSONResponse({"msg": "后端已启动。前端构建产物未就绪(static/)。API 见 /api/*"})

# -*- coding: utf-8 -*-
# [Change Log]
# Date: 2026-07-09 | Author: Claude / c | Version: V2.57
# Description: 成本台账内核 5 月真底稿 DoD 跑分器。读《成本台账-星期九5月.xlsx》的
#   库存（5月）=跨维度、成本（5月）按时间=按日期，跑三道勾稽/透视/异常/损益，
#   断言与底稿一致：三组账实勾稽差0、类别透视逐类一致、5负结存+镜像尾差捕获、损益 -4874.95/8912.37。
#   用法：python tools/run_cost_ledger_5月.py <底稿.xlsx路径>
import io
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
import openpyxl
from kernels import cost_ledger as cl

DEFAULT_SRC = r"C:\Users\94899\AppData\Local\Temp\claude\D--0-Claude----03--------\e025ffc8-b8cc-4e6b-9ca8-e34e391a26ce\scratchpad\成本台账-星期九5月.xlsx"
CFG = cl.load_config(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                                  "sample_data", "cost_ledger_config.json"))

# 5 月底稿科目余额（金蝶科目余额表，按科目去重后；见 V2.53 实测）
GL_5 = {"库存商品": 4153701.90, "原材料": 6364996.13, "周转材料": 439353.22,
        "在途物资": 4465.13, "委托加工物资": 31519.64}
# 底稿类别透视结存（DoD 对照基准）
EXPECT_CAT_END = {"产成品": 3944172.41, "自制半成品": 209529.49, "委外半成品": 0.0,
                  "原材料": 4695704.81, "包材": 1669291.33, "低值易耗品": 435941.50,
                  "广宣品": 3411.72}


def rows_of(ws):
    return list(ws.iter_rows(values_only=True))


def main():
    src = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_SRC
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    cross = cl.parse_cross_report(rows_of(wb["库存（5月）"]))
    bydate = cl.parse_bydate_report(rows_of(wb["成本（5月）按时间"]))
    # 货损/盘盈亏（损益归集）
    ws_loss = wb["货损明细-管理费 （5） "]
    loss = []
    for r in ws_loss.iter_rows(min_row=2, values_only=True):
        if r[1]:
            loss.append({"cat": r[1], "amount": r[9]})
    ws_disp = wb["盘盈亏 （5)  "]
    disp = []
    for r in ws_disp.iter_rows(min_row=2, values_only=True):
        if r[3] and r[9]:
            disp.append({"amount": r[9]})
    wb.close()

    res = cl.build_cost_ledger(cross, CFG, gl_balance=GL_5, bydate=bydate,
                               loss_rows=loss, disposal_rows=disp)

    fails = []

    def check(name, cond, detail=""):
        print(f"  [{'PASS' if cond else 'FAIL'}] {name}  {detail}")
        if not cond:
            fails.append(name)

    print(f"明细行数：跨维度 {len(cross)} / 按日期 {len(bydate)}")

    print("\n勾稽① 两表互勾：")
    t1 = res["ties"]["two_reports"]
    for k in ("期初金额", "收入金额", "发出金额", "结存金额"):
        check(k, t1[k]["pass"], f"跨维度 {t1[k]['cross']:,.2f} vs 按日期 {t1[k]['bydate']:,.2f} 差 {t1[k]['diff']}")

    print("\n勾稽② 收发存自平（总额）：")
    t2 = res["ties"]["self_balance"]
    check("自平总额", t2["合计"]["pass"], f"期初+收入-发出-结存 差 {t2['合计']['diff']}")

    print("\n勾稽③ 账实勾稽：")
    t3 = res["ties"]["book_vs_actual"]
    for subj in ("库存商品", "原材料", "周转材料"):
        s = t3["subjects"][subj]
        check(subj, s["pass"], f"账面 {s['book']:,.2f} vs 收发存 {s['actual']:,.2f} 差 {s['diff']}")
    check("总账存货合计", abs(t3["book_total"] - 10994036.02) < 0.05, f"{t3['book_total']:,.2f}")
    check("无对照缺失", len(t3["unmapped"]) == 0, str(t3["unmapped"]))

    print("\n类别透视结存 vs 底稿：")
    pc = res["pivot_category"]
    for cat, exp in EXPECT_CAT_END.items():
        got = pc.get(cat, {}).get("ea", -1)
        check(cat, abs(got - exp) < 0.05, f"{got:,.2f}（底稿 {exp:,.2f}）")

    print("\n异常扫描：")
    an = res["anomalies"]
    neg = [i for i in an["items"] if i["status"] == cl.ST_NEG]
    tail = [i for i in an["items"] if i["status"] == cl.ST_TAILDIFF]
    check("负结存=5", len(neg) == 5, f"{len(neg)} 个：" + "、".join(i["name"][:10] for i in neg))
    check("挂账尾差捕获", len(tail) >= 2, f"{len(tail)} 个")
    check("异常计数不重不漏", sum(an["counts"].values()) == an["total_rows"],
          f"Σ{sum(an['counts'].values())}={an['total_rows']}")

    print("\n损益归集：")
    pnl = res["pnl"]
    check("货损→管理费用 -4874.95", abs(pnl["loss"]["total"] - (-4874.95)) < 0.1, f"{pnl['loss']['total']}")
    check("处置→营业外 8912.37", abs(pnl["disposal"]["total"] - 8912.37) < 0.1, f"{pnl['disposal']['total']}")

    print("\n可信度结论：", "可信（三道全过）" if res["credible"] else "待复核")

    print("\n" + ("=" * 48))
    if fails:
        print(f"DoD 未达标，失败 {len(fails)} 项：{fails}")
        sys.exit(1)
    print("DoD 全部达标 ✓")


if __name__ == "__main__":
    main()

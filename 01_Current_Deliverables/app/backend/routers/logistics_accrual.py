# -*- coding: utf-8 -*-
# [Change Log]
# Date: 2026-08-05 | Author: Claude / c | Version: V2.172
# Description: 【物流计提】路由（V2.172 从 app.py 拆出）。
#              本文件是「物流计提」这条工具线在后端的唯一落点：改这条线的接口只动本文件，
#              不再碰 app.py —— 这样多条需求并行开发时互不冲突。
#              共享的配置/期间/权限判定见 core.py；算法在 kernels/logistics_accrual.py。
#              app.py 只负责 include_router(router)，不感知本文件内部。

import os

from fastapi import APIRouter
from fastapi import Request
from fastapi import File, UploadFile
from fastapi.responses import FileResponse
from io import BytesIO
from typing import List
from kernels import logistics_accrual as la
from kernels import logistics_bills as lb
import kingdee_client as kc

from core import (
    JSONResponse, _KD_STATUS_CN, _closed_block, _now, _require_perm, db,
)

router = APIRouter()


# ==================== 物流计提（月结与结账 · 通用技能）V2.20→V2.29 ====================
def _la_sheet(sheet, month):
    if sheet:
        return sheet
    return f"{int(month)}月" if month else 0


@router.post("/api/logistics-accrual/parse")
async def logistics_accrual_parse(request: Request, sheet: str = "", month: int = 0):
    """上传物流计提表(xlsx 字节) → 解析 → 生成计提凭证 + 自校验。不写金蝶。税率优先取维表。"""
    if not _require_perm(request, "logistics_upload"):
        return JSONResponse({"ok": False, "msg": "无「上传物流计提表」权限，请联系管理员"}, status_code=403)
    data = await request.body()
    if not data:
        return {"ok": False, "msg": "空文件（请选择物流计提表 .xlsx）"}
    try:
        res = la.process_workbook(data, _la_sheet(sheet, month), int(month or 0), rates=db.tax_rate_lookup())
    except Exception as e:
        return {"ok": False, "msg": f"解析失败：{e}"}
    return {"ok": True, "sheet": _la_sheet(sheet, month), **res}


@router.post("/api/logistics-accrual/suppliers-check")
async def logistics_accrual_suppliers(request: Request, sheet: str = "", month: int = 0):
    """核对计提表(月结)供应商是否都已在金蝶建档。列出缺的（需人工先去金蝶建）。"""
    if not _require_perm(request, "logistics_upload"):
        return JSONResponse({"ok": False, "msg": "无「上传物流计提表」权限，请联系管理员"}, status_code=403)
    data = await request.body()
    if not data:
        return {"ok": False, "msg": "空文件"}
    try:
        import pandas as pd
        df = pd.read_excel(BytesIO(data), sheet_name=_la_sheet(sheet, month), header=None)
        tbl = la.list_suppliers_in_table(la.parse_accrual_df(df))
    except Exception as e:
        return {"ok": False, "msg": f"解析失败：{e}"}
    try:
        sup = kc.fetch_suppliers()
    except kc.KingdeeError as e:
        return {"ok": False, "msg": f"金蝶取供应商失败：{e}"}
    names = set(s["供应商名称"] for s in sup)
    missing = [t for t in tbl if t not in names]
    return {"ok": True, "金蝶供应商数": len(sup), "计提表供应商数": len(tbl),
            "已建档": len(tbl) - len(missing), "缺建档": missing}


# ---------------- 税率维表（供应商×费用类型，不含税口径） ----------------
@router.get("/api/logistics-accrual/tax-rates")
def logistics_tax_rates(request: Request):
    """税率维表列表（登录即可看）。"""
    return {"ok": True, "rates": db.list_tax_rates()}


@router.post("/api/logistics-accrual/tax-rates/save")
def logistics_tax_rate_save(body: dict, request: Request):
    u = _require_perm(request, "logistics_upload")
    if not u:
        return JSONResponse({"ok": False, "msg": "无「上传物流计提表·维护税率」权限，请联系管理员"}, status_code=403)
    supplier = str(body.get("supplier", "") or "").strip()
    fee_type = str(body.get("fee_type", "") or "").strip()
    if not supplier:
        return {"ok": False, "msg": "供应商全名不能为空"}
    try:
        rate = float(body.get("rate"))
    except (TypeError, ValueError):
        return {"ok": False, "msg": "税率要填数字（如 9 代表 9%）"}
    if rate > 1:                      # 前端按百分数填，后端统一存小数
        rate = rate / 100.0
    if not (0 <= rate <= 0.5):
        return {"ok": False, "msg": "税率超出常理范围（0%~50%），请检查"}
    rid = db.save_tax_rate(supplier, fee_type, rate, u["name"])
    db.audit(u["name"], "物流计提-税率维护", supplier, f"{fee_type or '(默认)'}={rate}")
    return {"ok": True, "id": rid}


@router.post("/api/logistics-accrual/tax-rates/delete")
def logistics_tax_rate_delete(body: dict, request: Request):
    u = _require_perm(request, "logistics_upload")
    if not u:
        return JSONResponse({"ok": False, "msg": "无「上传物流计提表·维护税率」权限，请联系管理员"}, status_code=403)
    rid = body.get("id")
    if not rid:
        return {"ok": False, "msg": "缺 id"}
    db.delete_tax_rate(rid)
    db.audit(u["name"], "物流计提-税率删除", str(rid))
    return {"ok": True}

# ==================== 账单直采（V2.195 重构方案 v2.1 §9：核对后账单 → 计提明细活表） ====================
@router.post("/api/logistics-accrual/bills-parse")
async def logistics_bills_parse(request: Request, month: int = 0, year: int = 2026,
                                files: List[UploadFile] = File(...)):
    """上传核对后账单包（多文件）→ 按商解析 → 标注翻译 → 聚合成计提明细活表行（维度已预填，前端可改）。
    解析成功即落上传批次留痕（V2.221）——谁/何时/哪几份+活表行全量，切回该月可恢复现场不必重传。"""
    u = _require_perm(request, "logistics_upload")
    if not u:
        return JSONResponse({"ok": False, "msg": "无「上传物流计提表」权限，请联系管理员"}, status_code=403)
    if not (1 <= int(month or 0) <= 12):
        return {"ok": False, "msg": "请先选择计提月份(1-12)"}
    pairs = []
    for f in files:
        data = await f.read()
        if data:
            pairs.append((f.filename or "", data))
    if not pairs:
        return {"ok": False, "msg": "没有收到账单文件"}
    try:
        res = lb.parse_bill_files(pairs, int(month), rates=db.tax_rate_lookup(),
                                  fee_lk=db.fee_map_lookup(), biz_lk=db.bizline_lookup(),
                                  tmap=db.type_map_lookup(), suppliers=db.list_logi_suppliers())
    except Exception as e:
        return {"ok": False, "msg": f"账单解析失败：{e}"}
    # V2.229 按供应商合并进本月工作批次（到几家传几家，晚到补传不顶别家），并自动落"初始账单"档案
    uid, merged = db.merge_bill_upload(int(year), int(month), u["name"],
                                       res.get("rows"), res.get("per_file"), res.get("tickets"))
    _upsert_docs_from_rows(int(year), int(month), res.get("rows") or [], res.get("per_file") or [], u["name"])
    db.audit(u["name"], "物流计提-上传账单包", f"{year}年{month}月",
             f"{res['stats']['文件数']}个文件 {res['stats']['票数']}票 {res['stats']['含税合计']}")
    # 归集完成判定（V2.225 核算组定稿）：无认不出的账单 + 无待人工行 + 无缺税率 → 归集完成（按合并后全量判）
    res.update(rows=merged["rows"], per_file=merged["per_file"], stats=merged["stats"])
    rows = res.get("rows") or []
    no_rate = sum(1 for v in rows if v.get("税率来源") == "缺税率")
    issues = []
    if res.get("unknown_files"):
        issues.append(f"有 {len(res['unknown_files'])} 个账单没认出物流商（可能是新供应商）")
    if res["stats"].get("待人工行"):
        issues.append(f"{res['stats']['待人工行']} 行待人工补维度（{res['stats']['待人工金额']:,.2f} 元）")
    if no_rate:
        issues.append(f"{no_rate} 行缺税率")
    complete = not issues
    # 认不出的账单文件（=疑似新供应商）→ 自动 邮件+钉钉 通知核算组建档（失败只回执不拦）
    notify_res = None
    if res.get("unknown_files"):
        import notifier
        fl = "\n".join(f"  · {x}" for x in res["unknown_files"])
        subject = f"【物流计提】{year}年{month}月账单包发现 {len(res['unknown_files'])} 个新供应商账单，请建档"
        text = (f"{u['name']} 于 {_now()} 上传 {year}年{month}月账单包（{res['stats']['文件数']} 个文件），"
                f"以下账单没认出对应物流商：\n{fl}\n"
                f"请到 财务核算工作台 → 物流对账 → 基础数据·供应商列表 建档（简称要能在文件名里认出）并维护税率；"
                f"全新格式的账单需开发解析器再上传。")
        try:
            notify_res = notifier.notify(subject, text, scene="新供应商建档")
        except Exception as e:
            notify_res = {"error": str(e)}
    return {"ok": True, "month": int(month), "upload_id": uid,
            "complete": complete, "issues": issues, "notify": notify_res, **res}


@router.post("/api/logistics-accrual/bill-uploads/submit")
def logistics_bill_upload_submit(body: dict, request: Request):
    """物流部「提交给核算组」（V2.225 定稿流程收口）：批次标已提交 + 邮件/钉钉通知核算组检查录入。"""
    u = _require_perm(request, "logistics_upload")
    if not u:
        return JSONResponse({"ok": False, "msg": "无权限"}, status_code=403)
    info = db.submit_bill_upload(body.get("id") or 0, u["name"])
    if not info:
        return {"ok": False, "msg": "找不到该上传批次"}
    st = info["stats"]
    files = "、".join(p["文件"] for p in info["per_file"][:8]) + ("…" if len(info["per_file"]) > 8 else "")
    subject = f"【物流计提】{info['year']}年{info['month']}月计提长表已归集完成，请检查并录入金蝶"
    text = (f"{u['name']} 于 {_now()} 提交 {info['year']}年{info['month']}月物流计提：\n"
            f"  {st.get('文件数')} 个账单 · {st.get('票数')} 票 → {st.get('明细行数')} 行计提明细，"
            f"含税合计 {st.get('含税合计'):,.2f} 元；待人工 {st.get('待人工行', 0)} 行。\n"
            f"  文件：{files}\n"
            f"请到 财务核算工作台 → 物流计提 → 选 {info['month']} 月「载入这批」→ 活表复核 → 做账去向与费率 → 勾选录入金蝶。")
    import notifier
    try:
        notify_res = notifier.notify(subject, text, scene="提交核算组")
    except Exception as e:
        notify_res = {"error": str(e)}
    db.audit(u["name"], "物流计提-提交给核算组", f"{info['year']}年{info['month']}月",
             f"批次{info['id']}：{st.get('明细行数')}行 {st.get('含税合计')}")
    return {"ok": True, "notify": notify_res}


_NOTIFY_SCENES = [   # V2.230 分场景通知收件人（账单上传④通知设置页维护；空=回落 conf.ini 公共名单）
    ("新供应商建档", "账单/长表里认出新物流商时 → 提醒核算组建档并维护税率"),
    ("提交核算组", "物流部提交计提表后 → 提醒核算组检查并录入金蝶"),
    ("付款提醒", "汇总表勾选发起付款时 → 提醒安排付款（出纳/主管）"),
]


def _notify_passcode():
    """改收件人的口令——conf.ini [notify] passcode（机密、不下发前端；与汇率线通知设置同一把口令）。未配置=空串。"""
    import configparser
    try:
        c = configparser.ConfigParser()
        c.read(kc.conf_path(), encoding="utf-8")
        return (c.get("notify", "passcode", fallback="") or "").strip()
    except Exception:
        return ""


@router.get("/api/logistics-accrual/notify-recipients")
def logistics_notify_recipients(request: Request):
    """三个场景的收件人配置 + conf.ini 公共名单（留空场景实际发给谁，页面要能看见）。"""
    u = _require_perm(request, "logistics_upload")
    if not u:
        return JSONResponse({"ok": False, "msg": "无权限"}, status_code=403)
    import notifier
    import mailer
    saved = {r["scene"]: r for r in db.list_notify_recipients()}
    scenes = [{"scene": s, "desc": d,
               "mobiles": (saved.get(s) or {}).get("mobiles") or "",
               "emails": (saved.get(s) or {}).get("emails") or "",
               "updated_by": (saved.get(s) or {}).get("updated_by") or "",
               "updated_at": (saved.get(s) or {}).get("updated_at") or ""} for s, d in _NOTIFY_SCENES]
    dt = notifier.load_dingtalk_conf()
    sm = mailer.load_smtp_conf()
    fallback = {"mobiles": (dt or {}).get("mobiles") or [], "userids": (dt or {}).get("userids") or [],
                "emails": (sm or {}).get("to") or [],
                "dingtalk_ready": bool(dt), "email_ready": bool(sm)}
    return {"ok": True, "scenes": scenes, "fallback": fallback,
            "passcode_set": bool(_notify_passcode())}   # V2.231：改收件人须口令；口令没配→页面锁改


@router.post("/api/logistics-accrual/notify-recipients")
def logistics_notify_recipients_save(body: dict, request: Request):
    u = _require_perm(request, "logistics_upload")
    if not u:
        return JSONResponse({"ok": False, "msg": "无权限"}, status_code=403)
    scene = str(body.get("scene") or "").strip()
    if scene not in {s for s, _d in _NOTIFY_SCENES}:
        return {"ok": False, "msg": f"不认识的场景：{scene}"}
    # V2.231 改收件人须口令（conf.ini [notify] passcode，与汇率线同一把）——防止随手改动把通知发丢
    pc = _notify_passcode()
    if not pc:
        return {"ok": False, "msg": "后端未设置通知口令（conf.ini [notify] passcode 为空），暂不能从页面改收件人，请联系管理员配置。"}
    if str(body.get("passcode") or "").strip() != pc:
        db.audit(u["name"], "物流计提-通知收件人", scene, "口令错误，未保存")
        return JSONResponse({"ok": False, "msg": "口令错误，未保存"}, status_code=403)
    db.save_notify_recipients(scene, body.get("mobiles") or "", body.get("emails") or "", u["name"])
    db.audit(u["name"], "物流计提-通知收件人", scene,
             f"钉钉[{body.get('mobiles') or '（空→公共名单）'}] 邮件[{body.get('emails') or '（空→公共名单）'}]")
    return {"ok": True}


@router.post("/api/logistics-accrual/notify-test")
def logistics_notify_test(body: dict, request: Request):
    """按当前配置真发一条测试消息（收到即配置正确）——本机通知通道禁用时只回执不发。"""
    u = _require_perm(request, "logistics_upload")
    if not u:
        return JSONResponse({"ok": False, "msg": "无权限"}, status_code=403)
    scene = str(body.get("scene") or "").strip()
    import notifier
    try:
        res = notifier.notify(f"【测试】财务核算工作台·{scene}通知连通测试",
                              f"{u['name']} 于 {_now()} 在「账单上传 › 通知设置」发送测试。收到本条即该场景收件人配置正确。",
                              scene=scene)
    except Exception as e:
        res = {"error": str(e)}
    db.audit(u["name"], "物流计提-通知测试", scene, str(res)[:200])
    return {"ok": True, "notify": res}


def _upsert_docs_from_rows(year, month, rows, per_file, operator):
    """解析成功的行按 商×主体 聚合 → 自动落"初始账单"档（V2.232 分主体）。
    只处理本次解析到的商（以 per_file 已解析为准）；同商的账单文件挂到它全部主体行。"""
    files_by_short = {}
    for p in per_file or []:
        s = p.get("物流商") or ""
        if s and not s.startswith("（") and p.get("状态") == "已解析":
            files_by_short.setdefault(s, []).append(p.get("文件") or "")
    grp = {}
    for r in rows or []:
        s = r.get("物流商") or ""
        if s not in files_by_short:
            continue
        sub = r.get("主体") or "待补"
        g = grp.setdefault(s, {}).setdefault(sub, {"票数": 0, "金额": 0.0})
        g["票数"] += int(r.get("票数") or 0)
        g["金额"] += float(r.get("含税") or 0)
    for s, groups in grp.items():
        db.upsert_supplier_bills(year, month, s, groups, "；".join(files_by_short[s]), operator)


@router.get("/api/logistics-accrual/supplier-matrix")
def logistics_supplier_matrix(request: Request, year: int = 2026, month: int = 0):
    """供应商月度生命周期矩阵（V2.229）：一行一商，格子=初始账单→计提→核对定稿→发票→付款。
    行来源=本月档案 ∪ 本月批次行 ∪ 上月有费用的商（应到未到→催单）。"""
    u = _require_perm(request, "logistics_upload")
    if not u:
        return JSONResponse({"ok": False, "msg": "无权限"}, status_code=403)
    year, month = int(year), int(month or 0)
    if not (1 <= month <= 12):
        return {"ok": False, "msg": "缺月份"}
    docs = db.get_supplier_docs(year, month)                 # {商: {主体: 档案}}
    inv_files = db.list_invoice_files(year, month)           # {(商, 主体): [附件]}
    base, base_status = db._latest_bill_upload_with_rows(year, month)
    rows = (base or {}).get("rows") or []
    per_file = (base or {}).get("per_file") or []
    py, pm = (year, month - 1) if month > 1 else (year - 1, 12)
    prev, _st = db._latest_bill_upload_with_rows(py, pm)
    prev_combos = {(r.get("物流商"), r.get("主体") or "待补")
                   for r in ((prev or {}).get("rows") or []) if r.get("物流商")}
    sups = {s.get("short"): s for s in db.list_logi_suppliers()}
    files_by_short = {}
    for p in per_file:
        if p.get("物流商") and p.get("状态") == "已解析":
            files_by_short.setdefault(p["物流商"], []).append(p.get("文件") or "")
    combos = {(r.get("物流商"), r.get("主体") or "待补") for r in rows if r.get("物流商")}
    combos |= {(s, sub) for s, m in docs.items() for sub in m}
    combos |= prev_combos
    out = []
    for s, sub in combos:
        d = (docs.get(s) or {}).get(sub) or {}
        srows = [r for r in rows if r.get("物流商") == s and (r.get("主体") or "待补") == sub]
        pend = [r for r in srows if r.get("可录入") is False]
        if d.get("bill_file"):
            bill = {"已到": True, "文件": d["bill_file"], "时间": d.get("bill_ts") or "",
                    "票数": d.get("bill_tickets") or 0, "金额": d.get("bill_amount") or 0}
        elif srows and s in files_by_short:   # 分主体前传的存量批次：按该主体的行兜底展示
            bill = {"已到": True, "文件": "；".join(files_by_short[s]),
                    "时间": (base or {}).get("ts") or "",
                    "票数": sum(int(r.get("票数") or 0) for r in srows),
                    "金额": round(sum(float(r.get("含税") or 0) for r in srows), 2)}
        else:
            bill = {"已到": False}
        sup = sups.get(s) or {}
        atts = inv_files.get((s, sub)) or []
        accr = round(sum(float(r.get("含税") or 0) for r in srows), 2)
        out.append({
            "主体": sub, "简称": s, "全名": sup.get("full") or s, "可解析": s in lb.ADAPTERS,
            "应到": (s, sub) in prev_combos, "初始账单": bill,
            "计提": {"行数": len(srows), "待补": len(pend), "含税": accr,
                     "已提交": bool(srows) and base_status == "已提交"},
            "定稿": {"已确认": bool(d.get("verified")), "金额": d.get("verified_amount"),
                     "时间": d.get("verified_ts") or "", "人": d.get("verified_by") or "",
                     "说明": d.get("verify_note") or "",
                     "附件": [f for f in atts if f.get("kind") == "结算账单"],
                     "与计提差": (round(float(d.get("verified_amount")) - accr, 2)
                                  if d.get("verified") and srows and d.get("verified_amount") is not None else None)},
            "发票": {"票号": d.get("invoice_no") or "", "金额": d.get("invoice_amount"),
                     "日期": d.get("invoice_date") or "",
                     "附件": [f for f in atts if (f.get("kind") or "发票") == "发票"],
                     "差异": (round(float(d.get("invoice_amount")) - float(d.get("verified_amount")), 2)
                              if d.get("invoice_no") and d.get("invoice_amount") is not None
                              and d.get("verified_amount") is not None else None)},
            "付款": {"已提醒": bool(d.get("pay_requested")), "时间": d.get("pay_ts") or "",
                     "可申请": bool(d.get("verified")) and bool(d.get("invoice_no")) and not pend and bool(srows)},
        })
    out.sort(key=lambda x: (not x["初始账单"]["已到"], x["简称"], -x["计提"]["含税"]))
    return {"ok": True, "suppliers": out, "submitted": base_status == "已提交",
            "upload_id": (base or {}).get("id")}


@router.post("/api/logistics-accrual/supplier-doc")
def logistics_supplier_doc(body: dict, request: Request):
    """档案状态推进（V2.232 分主体）：verify定稿 / unverify撤定稿 / invoice登记发票 / clear_invoice / unpay撤付款标记。"""
    u = _require_perm(request, "logistics_upload")
    if not u:
        return JSONResponse({"ok": False, "msg": "无权限"}, status_code=403)
    year, month = int(body.get("year") or 0), int(body.get("month") or 0)
    short, action = str(body.get("short") or ""), str(body.get("action") or "")
    subject = str(body.get("subject") or "待补")
    if not (short and 1 <= month <= 12):
        return {"ok": False, "msg": "缺供应商或月份"}
    if action == "verify":
        # 分主体前传的存量批次没有"初始账单"档——定稿时按该商的行回填缺的主体行（只补缺，不动已有）
        have = db.get_supplier_docs(year, month).get(short) or {}
        if not (have.get(subject) or {}).get("bill_file"):
            base, _st = db._latest_bill_upload_with_rows(year, month)
            need = {}
            for r in ((base or {}).get("rows") or []):
                if r.get("物流商") != short:
                    continue
                sub = r.get("主体") or "待补"
                if (have.get(sub) or {}).get("bill_file"):
                    continue
                g = need.setdefault(sub, {"票数": 0, "金额": 0.0})
                g["票数"] += int(r.get("票数") or 0)
                g["金额"] += float(r.get("含税") or 0)
            files = [p.get("文件") or "" for p in ((base or {}).get("per_file") or [])
                     if p.get("物流商") == short and p.get("状态") == "已解析"]
            if need and files:
                db.upsert_supplier_bills(year, month, short, need, "；".join(files), u["name"], cleanup=False)
    ok, msg = db.set_supplier_doc(year, month, short, subject, action, u["name"],
                                  invoice_no=body.get("invoice_no") or "",
                                  invoice_amount=body.get("invoice_amount"),
                                  invoice_date=body.get("invoice_date") or "",
                                  final_amount=body.get("final_amount"),
                                  verify_note=body.get("verify_note") or "")
    if ok:
        adj = action == "verify" and body.get("final_amount") not in (None, "")
        label = {"verify": "按异议定稿(手登金额)" if adj else "确认无误(定稿)", "unverify": "撤销定稿",
                 "invoice": "登记发票", "clear_invoice": "清除发票登记", "unpay": "撤销付款提醒标记"}.get(action, action)
        detail = f"{subject}×{short}：{label}"
        if action == "invoice":
            detail += f" {body.get('invoice_no')}"
        if adj:
            detail += f" 定稿{body.get('final_amount')}（{body.get('verify_note') or '未填原因'}）"
        db.audit(u["name"], "物流计提-供应商档案", f"{year}年{month}月", detail)
    return {"ok": ok, "msg": msg}


@router.post("/api/logistics-accrual/invoice-file")
async def logistics_invoice_file_upload(request: Request, year: int = 2026, month: int = 0,
                                        short: str = "", subject: str = "", kind: str = "发票",
                                        file: UploadFile = File(...)):
    """结算凭证附件上传（V2.233 发票；V2.235 扩 kind=结算账单——正确版/盖章版账单，只存档不解析，计提行不动）。"""
    u = _require_perm(request, "logistics_upload")
    if not u:
        return JSONResponse({"ok": False, "msg": "无权限"}, status_code=403)
    if not (short and 1 <= int(month or 0) <= 12):
        return {"ok": False, "msg": "缺供应商或月份"}
    kind = "结算账单" if kind == "结算账单" else "发票"
    fn = file.filename or kind
    ext = os.path.splitext(fn)[1].lower()
    allowed = (".pdf", ".jpg", ".jpeg", ".png", ".ofd") + ((".xlsx", ".xls") if kind == "结算账单" else ())
    if ext not in allowed:
        hint = "PDF/JPG/PNG/OFD/Excel" if kind == "结算账单" else "PDF/JPG/PNG/OFD"
        return {"ok": False, "msg": f"{kind}只收 {hint}，收到 {ext or '未知格式'}"}
    data = await file.read()
    if not data:
        return {"ok": False, "msg": "空文件"}
    if len(data) > 15 * 1024 * 1024:
        return {"ok": False, "msg": "文件超过 15MB——请压缩后再传"}
    fid = db.save_invoice_file(int(year), int(month), short, subject, fn, data, u["name"], kind=kind)
    db.audit(u["name"], f"物流计提-{kind}附件", f"{year}年{month}月", f"{subject}×{short}：{fn}（{len(data)//1024}KB）")
    return {"ok": True, "id": fid, "filename": fn, "kind": kind}


@router.get("/api/logistics-accrual/invoice-file")
def logistics_invoice_file_download(request: Request, fid: int = 0):
    u = _require_perm(request, "logistics_upload")
    if not u:
        return JSONResponse({"ok": False, "msg": "无权限"}, status_code=403)
    f = db.get_invoice_file(fid)
    if not f:
        return JSONResponse({"ok": False, "msg": "附件不存在（可能已删除）"}, status_code=404)
    from urllib.parse import quote
    from fastapi.responses import Response
    mt = {".pdf": "application/pdf", ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
          ".png": "image/png"}.get(os.path.splitext(f["filename"])[1].lower(), "application/octet-stream")
    return Response(content=f["content"], media_type=mt,
                    headers={"Content-Disposition": f"attachment; filename*=utf-8''{quote(f['filename'])}"})


@router.post("/api/logistics-accrual/invoice-file-delete")
def logistics_invoice_file_delete(body: dict, request: Request):
    u = _require_perm(request, "logistics_upload")
    if not u:
        return JSONResponse({"ok": False, "msg": "无权限"}, status_code=403)
    f = db.get_invoice_file(body.get("id") or 0)
    if not f:
        return {"ok": False, "msg": "附件不存在"}
    db.delete_invoice_file(f["id"])
    db.audit(u["name"], "物流计提-发票附件删除", f"{f['year']}年{f['month']}月", f"{f['subject']}×{f['short']}：{f['filename']}")
    return {"ok": True}


@router.post("/api/logistics-accrual/diff-parse")
async def logistics_diff_parse(request: Request, year: int = 2026, month: int = 0,
                               short: str = "", subject: str = "", file: UploadFile = File(...)):
    """修正版账单解析比对（V2.236 业务方定）：对账吵出差异后，传修正版 Excel →
    同一套解析引擎沙箱跑一遍（**只比对，不进批次、计提数据零改动**）→ 逐类列出 计提 vs 修正版 差在哪；
    修正版文件自动存档为「结算账单」附件。"""
    u = _require_perm(request, "logistics_upload")
    if not u:
        return JSONResponse({"ok": False, "msg": "无权限"}, status_code=403)
    if not (short and 1 <= int(month or 0) <= 12):
        return {"ok": False, "msg": "缺供应商或月份"}
    data = await file.read()
    if not data:
        return {"ok": False, "msg": "空文件"}
    fn = file.filename or "修正版账单.xlsx"
    try:
        res = lb.parse_bill_files([(fn, data)], int(month), rates=db.tax_rate_lookup(),
                                  fee_lk=db.fee_map_lookup(), biz_lk=db.bizline_lookup(),
                                  tmap=db.type_map_lookup(), suppliers=db.list_logi_suppliers())
    except Exception as e:
        return {"ok": False, "msg": f"修正版解析失败：{e}"}
    if res.get("unknown_files"):
        return {"ok": False, "msg": "认不出这份账单属于哪家物流商（文件名要含供应商简称）。也可以直接手登定稿金额，文件用「传正确版/盖章版账单」存档。"}
    pf = (res.get("per_file") or [{}])[0]
    if pf.get("物流商") != short:
        return {"ok": False, "msg": f"这份账单认出的是「{pf.get('物流商')}」，不是当前行的「{short}」——请核对文件"}
    if pf.get("状态") != "已解析":
        return {"ok": False, "msg": f"解析失败：{pf.get('状态')}。可先手登定稿金额，文件用「传正确版/盖章版账单」存档。"}
    new_rows = [r for r in (res.get("rows") or []) if r.get("物流商") == short]
    base, _st = db._latest_bill_upload_with_rows(int(year), int(month))
    old_rows = [r for r in ((base or {}).get("rows") or []) if r.get("物流商") == short]

    def _k(r):
        return (r.get("主体") or "待补", r.get("费用归属") or "（待人工）",
                r.get("业务线") or "—", r.get("业务描述") or "")
    agg = {}
    for r in old_rows:
        agg.setdefault(_k(r), [0.0, 0.0])[0] += float(r.get("含税") or 0)
    for r in new_rows:
        agg.setdefault(_k(r), [0.0, 0.0])[1] += float(r.get("含税") or 0)
    diff = [{"主体": k[0], "费用归属": k[1], "业务线": k[2], "描述": k[3],
             "计提": round(a, 2), "修正": round(b, 2), "差额": round(b - a, 2)}
            for k, (a, b) in sorted(agg.items(), key=lambda kv: -abs(kv[1][1] - kv[1][0]))]
    st = {}
    for d in diff:
        t = st.setdefault(d["主体"], [0.0, 0.0])
        t[0] += d["计提"]
        t[1] += d["修正"]
    subject_totals = {k: {"计提": round(a, 2), "修正": round(b, 2), "差额": round(b - a, 2)}
                      for k, (a, b) in st.items()}
    # 修正版自动存档（结算账单）——解析主流程一步没走，批次/档案/计提行零改动
    fid = db.save_invoice_file(int(year), int(month), short, subject, fn, data, u["name"], kind="结算账单")
    tot = round(sum(d["差额"] for d in diff), 2)
    db.audit(u["name"], "物流计提-修正版比对", f"{year}年{month}月", f"{subject}×{short}：{fn} 总差额{tot}")
    return {"ok": True, "file_id": fid, "diff": diff, "subject_totals": subject_totals, "总差额": tot}


@router.get("/api/logistics-accrual/annotation-spec")
def logistics_annotation_spec(request: Request):
    """账单标注规范下载（V2.237 业务方定）：物流部填单号 VLOOKUP 出单据类型后，照本表选标注写进账单"类型"列。
    实时从「基础数据·标注翻译表」生成——核算组在维表里增改，物流部下载的规范永远是最新版。"""
    u = _require_perm(request, "logistics_upload")
    if not u:
        return JSONResponse({"ok": False, "msg": "无权限"}, status_code=403)
    from io import BytesIO as _B
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    rows = db.list_type_map()
    fee_order = {f: i for i, f in enumerate([
        "销售出库费用", "成品入库费用", "原料入库费用", "成品仓储费用", "原料仓储费用",
        "成品调拨费用", "原料调拨费用", "出库装卸费用", "成品入库装卸费用", "原料入库装卸费用",
        "研发设备采购", "设备调拨费用", "其它"])}
    rows.sort(key=lambda r: (fee_order.get(r.get("fee") or "", 99), r.get("bizline") or "", r.get("pattern") or ""))
    wb = Workbook()
    ws = wb.active
    ws.title = "标注规范"
    bold = Font(name="微软雅黑", bold=True)
    ws["A1"] = "物流账单标注规范（写在账单「类型」列——系统按此自动生成计提维度）"
    ws["A1"].font = Font(name="微软雅黑", bold=True, size=13)
    notes = [
        "① 填写单号 → VLOOKUP 出单据类型/销售组织 → 从「标注清单」页选一个标准标注，原样写进账单「类型」列；",
        "② 命名法＝费用类型-业务线(-产品线)，如：销售出库单-零售-山姆；",
        "③ 仅两类允许自由后缀：设备调拨-需求部门（如 设备调拨-永续研发中心，科目部门由核算组逐笔定）、研发费用-项目名(-TOB)；",
        "④ 到付运费单独标「到付」；没有合适标注的先空着（系统进待人工），并联系核算组在「基础数据·标注翻译表」加一行；",
        "⑤ 本表由系统实时生成——每月下载最新版，别用旧文件。",
    ]
    for i, t in enumerate(notes):
        ws.cell(row=2 + i, column=1).value = t
    hr = 8
    for ci, h in enumerate(["标准标注（照抄这列）", "→ 费用归属", "→ 业务线", "→ 摘要用语"], start=1):
        cell = ws.cell(row=hr, column=ci)
        cell.value = h
        cell.font = bold
        cell.fill = PatternFill("solid", fgColor="DDE6F5")
        cell.alignment = Alignment(horizontal="center")
    for ri, r in enumerate(rows, start=hr + 1):
        ws.cell(row=ri, column=1).value = r.get("pattern")
        ws.cell(row=ri, column=2).value = r.get("fee")
        ws.cell(row=ri, column=3).value = r.get("bizline") or "—"
        ws.cell(row=ri, column=4).value = r.get("descr") or ""
    for col, w in (("A", 30), ("B", 18), ("C", 12), ("D", 18)):
        ws.column_dimensions[col].width = w
    ws2 = wb.create_sheet("标注清单")   # 纯清单页——给物流部做下拉/VLOOKUP 源
    ws2["A1"] = "标准标注"
    ws2["A1"].font = bold
    for ri, r in enumerate(rows, start=2):
        ws2.cell(row=ri, column=1).value = r.get("pattern")
    ws2.column_dimensions["A"].width = 30
    bio = _B()
    wb.save(bio)
    from urllib.parse import quote
    from fastapi.responses import Response
    fn = "物流账单标注规范.xlsx"
    return Response(content=bio.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename*=utf-8''{quote(fn)}"})


@router.get("/api/logistics-accrual/export-long-form")
def logistics_export_long_form(request: Request, year: int = 2026, month: int = 0):
    """导出当前批次为 24 列计提长表（V2.234 兜底闭环）：解析不了的新商（开发没空做解析器时），
    物流部导出本表 → Excel 里手工补行 → 从通道二重新上传——导出格式与模板完全一致，可直接回环解析。"""
    u = _require_perm(request, "logistics_upload")
    if not u:
        return JSONResponse({"ok": False, "msg": "无权限"}, status_code=403)
    if not (1 <= int(month or 0) <= 12):
        return {"ok": False, "msg": "缺月份"}
    base, _st = db._latest_bill_upload_with_rows(int(year), int(month))
    if not base:
        return JSONResponse({"ok": False, "msg": "本月还没有解析数据，先上传账单"}, status_code=404)
    rows = base["rows"]
    tpl = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                       "templates", "物流费用计提表模板.xlsx")
    from io import BytesIO as _B
    from openpyxl import load_workbook
    wb = load_workbook(tpl)
    ws = None
    for nm in wb.sheetnames:
        if "计提明细" in nm:
            ws = wb[nm]
            break
    if ws is None:
        ws = wb[wb.sheetnames[0]]
    hdr = {str(ws.cell(row=3, column=c).value or "").strip(): c for c in range(1, 25)}
    for r in range(4, ws.max_row + 1):        # 清掉模板示例行/合计行，再灌真数据
        for c in range(1, 25):
            ws.cell(row=r, column=c).value = None
    r = 4
    for v in rows:
        ws.cell(row=r, column=1).value = f"{int(year)}-{int(month):02d}"
        ws.cell(row=r, column=2).value = v.get("主体") or ""
        ws.cell(row=r, column=3).value = v.get("物流商") or ""
        ws.cell(row=r, column=4).value = v.get("费用归属") or ""
        ws.cell(row=r, column=5).value = v.get("业务线") or "—"
        ws.cell(row=r, column=6).value = v.get("业务描述") or ""
        ws.cell(row=r, column=7).value = v.get("结算类型") or "月结"
        ws.cell(row=r, column=9).value = v.get("含税") or 0
        for name, key in (("税率", "税率"), ("备注", "备注"), ("票数", "票数")):
            if name in hdr and v.get(key) is not None:
                ws.cell(row=r, column=hdr[name]).value = v.get(key)
        # T-X 做账维度（列位与解析器约定一致：20-24）——人工/映射已定的值全带出去，回环解析时人工优先
        for k, c in (("科目", 20), ("部门", 21), ("费用项目", 22), ("产品分类编码", 23), ("产品项目", 24)):
            ws.cell(row=r, column=c).value = v.get(k) or ""
        r += 1
    ws.cell(row=r, column=1).value = "合计"    # 解析器以"合计"为数据区终点
    ws.cell(row=r, column=9).value = round(sum(float(v.get("含税") or 0) for v in rows), 2)
    bio = _B()
    wb.save(bio)
    db.audit(u["name"], "物流计提-导出长表", f"{year}年{month}月", f"{len(rows)}行")
    from urllib.parse import quote
    from fastapi.responses import Response
    fn = f"物流费用计提表_{int(year)}-{int(month):02d}_导出.xlsx"
    return Response(content=bio.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename*=utf-8''{quote(fn)}"})


@router.get("/api/logistics-accrual/template")
def logistics_template_download(request: Request):
    """计提长表模板下载（V2.227 账单上传页第一页）。物流部按模板整理「计提账单」再上传。"""
    u = _require_perm(request, "logistics_upload")
    if not u:
        return JSONResponse({"ok": False, "msg": "无权限"}, status_code=403)
    p = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                     "templates", "物流费用计提表模板.xlsx")
    if not os.path.isfile(p):
        return JSONResponse({"ok": False, "msg": "服务器缺模板文件 templates/物流费用计提表模板.xlsx——请管理员重新部署"}, status_code=404)
    return FileResponse(p, filename="物流费用计提表模板.xlsx",
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@router.get("/api/logistics-accrual/parse-support")
def logistics_parse_support(request: Request):
    """解析能力清单（V2.227 账单上传页第一页）：维表里每家物流商 × 有没有自动解析器。
    可解析=账单直接传「复核无误的账单」通道；不可解析=用模板整理成「计提账单」上传。"""
    u = _require_perm(request, "logistics_upload")
    if not u:
        return JSONResponse({"ok": False, "msg": "无权限"}, status_code=403)
    out = [{"简称": s.get("short") or "", "全名": s.get("full") or "", "渠道": s.get("channel") or "线下",
            "可解析": (s.get("short") or "") in lb.ADAPTERS}
           for s in db.list_logi_suppliers() if s.get("short")]
    return {"ok": True, "suppliers": out}


@router.post("/api/logistics-accrual/pay-remind")
def logistics_pay_remind(body: dict, request: Request):
    """汇总表勾选 → 发起钉钉付款提醒（V2.227）。当前=通知核算组/出纳按清单安排付款；
    对接钉钉付款审批流（自动起单）留到对账线三期，与账单核对打通后一起做。"""
    u = _require_perm(request, "logistics_upload")
    if not u:
        return JSONResponse({"ok": False, "msg": "无权限"}, status_code=403)
    year, month = int(body.get("year") or 0), int(body.get("month") or 0)
    items = body.get("items") or []
    if not items:
        return {"ok": False, "msg": "请先在汇总表勾选要发起付款的物流商"}
    # 付款闸门（V2.229，V2.232 起分主体）：该主体×商 核对定稿 + 发票已登记 才放行——服务端也拦
    docs = db.get_supplier_docs(year, month)
    blocked = []
    for x in items:
        d = (docs.get(str(x.get("物流商") or "")) or {}).get(str(x.get("主体") or "待补")) or {}
        tag = f"{x.get('主体')}×{x.get('物流商')}"
        if not d.get("verified"):
            blocked.append(f"{tag}（还没确认无误定稿）")
        elif not d.get("invoice_no"):
            blocked.append(f"{tag}（还没登记发票）")
    if blocked:
        return {"ok": False, "msg": "以下还没走完流程，不能发起付款：" + "；".join(blocked)}
    total = sum(float(x.get("含税") or 0) for x in items)
    lines = "\n".join(f"  · {x.get('主体', '')} 付 {x.get('全名') or x.get('物流商', '')}："
                      f"{float(x.get('含税') or 0):,.2f} 元（{x.get('行数', '?')} 行明细）" for x in items)
    subject = f"【物流付款】{year}年{month}月物流费付款提醒（{len(items)} 笔 · 合计 {total:,.2f} 元）"
    text = (f"{u['name']} 于 {_now()} 在「账单上传 › 汇总表」发起付款提醒：\n{lines}\n"
            f"  合计：{total:,.2f} 元\n"
            f"请按计提汇总核对后安排付款。（钉钉付款审批自动起单待对账线三期接入，当前为提醒）")
    import notifier
    try:
        notify_res = notifier.notify(subject, text, scene="付款提醒")
    except Exception as e:
        notify_res = {"error": str(e)}
    for x in items:   # 档案点亮"已发起付款提醒"（矩阵付款格，分主体）
        db.set_supplier_doc(year, month, str(x.get("物流商") or ""), str(x.get("主体") or "待补"), "pay", u["name"])
    db.audit(u["name"], "物流计提-付款提醒", f"{year}年{month}月", f"{len(items)}笔 合计{round(total, 2)}")
    return {"ok": True, "notify": notify_res}


@router.post("/api/logistics-accrual/long-form-parse")
async def logistics_long_form_parse(request: Request, month: int = 0, year: int = 2026,
                                    file: UploadFile = File(...)):
    """物流部上传 24 列长表（V2.224 核算组定稿流程）→ 解析 → 质检（新增供应商 + 干净度）→ 活表行。
    发现新增供应商：自动 邮件+钉钉机器人 通知核算组（去基础数据建档+维护税率）；通知失败不拦上传、回执随响应。"""
    u = _require_perm(request, "logistics_upload")
    if not u:
        return JSONResponse({"ok": False, "msg": "无「上传物流计提表」权限，请联系管理员"}, status_code=403)
    if not (1 <= int(month or 0) <= 12):
        return {"ok": False, "msg": "请先选择计提月份(1-12)"}
    data = await file.read()
    if not data:
        return {"ok": False, "msg": "空文件"}
    try:
        res = lb.parse_long_form(data, int(month), rates=db.tax_rate_lookup(),
                                 fee_lk=db.fee_map_lookup(), biz_lk=db.bizline_lookup(),
                                 suppliers=db.list_logi_suppliers())
    except ValueError as e:
        return {"ok": False, "msg": str(e)}
    except Exception as e:
        return {"ok": False, "msg": f"长表解析失败：{e}"}
    fn = file.filename or "长表"
    # V2.229 长表 per_file 按供应商拆行（原来一条"(物流部长表)"总行）——合并/生命周期档案都以商为单位
    grp = {}
    for r in res["rows"]:
        s = r.get("物流商") or "（未填物流商）"
        g = grp.setdefault(s, {"票数": 0, "金额": 0.0})
        g["票数"] += 1
        g["金额"] += float(r.get("含税") or 0)
    per_file = [{"文件": fn, "物流商": s, "状态": "已解析", "票数": g["票数"], "金额": round(g["金额"], 2)}
                for s, g in grp.items()]
    uid, merged = db.merge_bill_upload(int(year), int(month), u["name"], res["rows"], per_file, [])
    _upsert_docs_from_rows(int(year), int(month), res["rows"], per_file, u["name"])
    db.audit(u["name"], "物流计提-上传长表", f"{year}年{month}月",
             f"{fn}：{res['stats']['票数']}行 {res['stats']['含税合计']}；新供应商{res['stats']['新供应商']} 不干净{res['stats']['不干净行']}")
    # 新增供应商 → 邮件 + 钉钉机器人 通知核算组（notifier 未配置/失败只回执不拦）
    notify_res = None
    if res["new_suppliers"]:
        import notifier
        lines = "\n".join(f"  · {x['简称']}：{x['行数']} 行 · {x['金额']:,.2f} 元" for x in res["new_suppliers"])
        rows_html = "".join(f"<tr><td style='padding:4px 12px;border:1px solid #ddd'>{x['简称']}</td>"
                            f"<td style='padding:4px 12px;border:1px solid #ddd;text-align:right'>{x['行数']}</td>"
                            f"<td style='padding:4px 12px;border:1px solid #ddd;text-align:right'>{x['金额']:,.2f}</td></tr>"
                            for x in res["new_suppliers"])
        subject = f"【物流计提】{year}年{month}月长表发现 {len(res['new_suppliers'])} 家新供应商，请建档并维护税率"
        text = (f"{u['name']} 于 {_now()} 上传《{fn}》（{year}年{month}月，{res['stats']['票数']} 行，"
                f"合计 {res['stats']['含税合计']:,.2f}）。\n以下供应商不在「物流基础数据·供应商列表」：\n{lines}\n"
                f"请到 财务核算工作台 → 物流对账 → 基础数据：①供应商列表建档（简称/全名/渠道）②维护税率；"
                f"金蝶未建档的先在金蝶建供应商档案。建完物流计提页「载入这批」即可继续。")
        html = (f"<p><b>{u['name']}</b> 于 {_now()} 上传《{fn}》（{year}年{month}月，{res['stats']['票数']} 行，"
                f"合计 <b>{res['stats']['含税合计']:,.2f}</b>）。</p><p>以下供应商不在「物流基础数据·供应商列表」：</p>"
                f"<table style='border-collapse:collapse'><tr><th style='padding:4px 12px;border:1px solid #ddd'>简称</th>"
                f"<th style='padding:4px 12px;border:1px solid #ddd'>行数</th><th style='padding:4px 12px;border:1px solid #ddd'>金额</th></tr>{rows_html}</table>"
                f"<p>请到 财务核算工作台 → 物流对账 → <b>基础数据</b>：①供应商列表建档（简称/全名/渠道）②维护税率；"
                f"金蝶未建档的先在金蝶建供应商档案。建完在物流计提页「载入这批」即可继续。</p>")
        try:
            notify_res = notifier.notify(subject, text, html=html, scene="新供应商建档")
        except Exception as e:
            notify_res = {"error": str(e)}
    # 通知/audit 用完本文件口径后，再把"按商合并后的全量现场"回给前端（含税合计=全月，不只这份长表）
    res.update(rows=merged["rows"], stats={**res["stats"], **merged["stats"]})
    return {"ok": True, "month": int(month), "upload_id": uid, "per_file": merged["per_file"],
            "notify": notify_res, **res}


@router.get("/api/logistics-accrual/bill-uploads")
def logistics_bill_uploads(request: Request, year: int = 2026, month: int = 0):
    """某月账单上传批次列表（谁/何时/哪几份，最新在前）——月份胶囊配套（V2.221）。"""
    if not _require_perm(request, "logistics_upload"):
        return JSONResponse({"ok": False, "msg": "无权限"}, status_code=403)
    if not (1 <= int(month or 0) <= 12):
        return {"ok": False, "msg": "缺月份"}
    return {"ok": True, "uploads": db.list_bill_uploads(int(year), int(month))}


@router.post("/api/logistics-accrual/bill-uploads/save-rows")
def logistics_bill_upload_save_rows(body: dict, request: Request):
    """活表改动自动保存回上传批次（V2.223）——前端每次改维度成功即调，静默高频，不逐次 audit。"""
    if not _require_perm(request, "logistics_upload"):
        return JSONResponse({"ok": False, "msg": "无权限"}, status_code=403)
    uid = body.get("id") or 0
    ok = db.update_bill_upload_rows(uid, body.get("rows") or [])
    return {"ok": bool(ok), "msg": "" if ok else "批次不存在"}


@router.post("/api/logistics-accrual/bill-uploads/load")
def logistics_bill_upload_load(body: dict, request: Request):
    """载入某上传批次（活表行全量）——恢复现场，不用重传文件。"""
    if not _require_perm(request, "logistics_upload"):
        return JSONResponse({"ok": False, "msg": "无权限"}, status_code=403)
    d = db.get_bill_upload(body.get("id") or 0)
    if not d:
        return {"ok": False, "msg": "找不到该上传批次（可能已被清理）"}
    return {"ok": True, **d}


@router.post("/api/logistics-accrual/row-refresh")
def logistics_row_refresh(body: dict, request: Request):
    """活表改维度后重算一行（税率/未税/税额/摘要/分录/可录入）。服务端是唯一事实源，防前端各拼各的。
    body: {month, row, requery: bool}——requery=真 时按 费用归属×主体×业务线 重查映射默认(改了归属/业务线后用)。"""
    if not _require_perm(request, "logistics_upload"):
        return JSONResponse({"ok": False, "msg": "无权限"}, status_code=403)
    row = body.get("row") or {}
    month = int(body.get("month") or 0)
    if not row or not (1 <= month <= 12):
        return {"ok": False, "msg": "缺 row/month"}
    if body.get("requery"):
        # 待人工占位描述(无标注/到付/翻译失败)在补齐维度后清掉,别拼进摘要
        d = str(row.get("业务描述") or "")
        if d in ("账单无类型标注", "到付·是否计提待定") or d.startswith("标注无法翻译"):
            row["业务描述"] = ""
        lk = db.fee_map_lookup()
        m, tier = db.resolve_fee_map(lk, row.get("费用归属") or "", row.get("主体") or "", row.get("业务线") or "")
        if m:
            row.update(科目=m["account"] or "", 部门=m["dept"] or "", 费用项目=m["item"] or "",
                       摘要用语=m["sword"] or "", manual=bool(m["manual"]), 映射层级=tier)
        biz_lk = db.bizline_lookup()
        acc4 = (row.get("科目") or "")[:4]
        cpfl, cpxm = biz_lk.get(row.get("业务线") or "", ("", ""))
        row["产品分类编码"], row["产品项目"] = (cpfl, cpxm) if acc4 in ("6601", "6401") else ("", "")
    try:
        row = lb.finalize_row(row, month, rates=db.tax_rate_lookup())
    except Exception as e:
        return {"ok": False, "msg": f"重算失败：{e}"}
    return {"ok": True, "row": row}


@router.post("/api/logistics-accrual/adopt")
def logistics_adopt(body: dict, request: Request):
    """「采纳入维表」：把活表里人工改过的维度存为 fee_map 例外行，下月自动预填（复用税率表一键采纳模式）。
    body: {fee, subject, bizline, account, dept, item, sword}——subject/bizline 传空=改默认行。"""
    u = _require_perm(request, "logistics_upload")
    if not u:
        return JSONResponse({"ok": False, "msg": "无权限"}, status_code=403)
    try:
        rid = db.save_fee_map(body, u["name"])
    except ValueError as e:
        return {"ok": False, "msg": str(e)}
    db.audit(u["name"], "物流计提-采纳入维表", str(body.get("fee")),
             f'{body.get("subject") or "(默认)"}×{body.get("bizline") or "(默认)"}→{body.get("account")}/{body.get("dept")}')
    return {"ok": True, "id": rid}


def _map_crud(list_fn, save_fn, del_fn, label):
    """三张映射维表的同构 CRUD 出口。"""
    def _list(request: Request):
        return {"ok": True, "rows": list_fn()}

    def _save(body: dict, request: Request):
        u = _require_perm(request, "logistics_upload")
        if not u:
            return JSONResponse({"ok": False, "msg": "无权限"}, status_code=403)
        try:
            rid = save_fn(body, u["name"])
        except ValueError as e:
            return {"ok": False, "msg": str(e)}
        db.audit(u["name"], f"物流计提-{label}维护", str(body)[:120])
        return {"ok": True, "id": rid}

    def _delete(body: dict, request: Request):
        u = _require_perm(request, "logistics_upload")
        if not u:
            return JSONResponse({"ok": False, "msg": "无权限"}, status_code=403)
        if not body.get("id"):
            return {"ok": False, "msg": "缺 id"}
        del_fn(body["id"])
        db.audit(u["name"], f"物流计提-{label}删除", str(body.get("id")))
        return {"ok": True}
    return _list, _save, _delete


for _path, _fns, _label in (
    ("fee-map", (db.list_fee_map, db.save_fee_map, db.delete_fee_map), "费用归属映射"),
    ("bizlines", (db.list_bizlines, db.save_bizline, db.delete_bizline), "业务线维表"),
    ("type-map", (db.list_type_map, db.save_type_map, db.delete_type_map), "标注翻译表"),
    ("suppliers", (db.list_logi_suppliers, db.save_logi_supplier, db.delete_logi_supplier), "供应商列表"),
):
    _l, _sv, _d = _map_crud(*_fns, _label)
    router.get(f"/api/logistics-accrual/{_path}")(_l)
    router.post(f"/api/logistics-accrual/{_path}/save")(_sv)
    router.post(f"/api/logistics-accrual/{_path}/delete")(_d)


# ---------------- B 期·费用率（V2.197 第一刀）：分母=BP 各业务线不含税收入 ----------------
@router.post("/api/logistics-accrual/expense-ratio")
def logistics_expense_ratio(body: dict, request: Request):
    """费率看板：分子=本月活表行未税(前端传 rows)，分母=BP 工作台不含税收入；
    环比基数=金蝶序时账上月已录计提(可选,慢)。BP 不可达 → available:false 降级提示,不报错。
    body: {year, month, rows, with_prev: bool}"""
    if not _require_perm(request, "logistics_upload"):
        return JSONResponse({"ok": False, "msg": "无权限"}, status_code=403)
    year = int(body.get("year") or 2026)
    month = int(body.get("month") or 0)
    rows = body.get("rows") or []
    if not (1 <= month <= 12):
        return {"ok": False, "msg": "缺计提月份"}
    bp_rev = lb.fetch_bp_revenue(year)
    if bp_rev is None:
        return {"ok": True, "available": False,
                "msg": "BP 工作台不可达（收入数取不到）。请先启动 BP 后端(8010)或检查 BP_API_BASE 配置。"}
    prev_exp = None
    if body.get("with_prev"):
        pm_y, pm_m = (year - 1, 12) if month == 1 else (year, month - 1)
        prev_exp = lb.fetch_ledger_expense(pm_y, pm_m)   # 金蝶不可达→None,环比列留空
    res = lb.compute_expense_ratio(rows, year, month, bp_rev=bp_rev, prev_exp=prev_exp)
    cur = bp_rev.get(res["period"], {}) or {}
    res.update(ok=True, available=True, prev_loaded=prev_exp is not None,
               bp_unmapped=cur.get("unmapped", 0),
               bp_meta={"fetched_at": cur.get("fetched_at", ""), "batch": cur.get("batch"),
                        "pending": cur.get("pending", 0), "pending_rows": cur.get("pending_rows", 0),
                        "unassigned": cur.get("unassigned", 0)})
    return res


# ---------------- 一键录入金蝶（存草稿 + 回传凭证号；提交/审核始终人在金蝶做） ----------------
@router.post("/api/logistics-accrual/post")
def logistics_accrual_post(body: dict, request: Request):
    """把勾选的计提凭证录入金蝶（草稿态）。服务端重新校验每张凭证，摘要重复的自动跳过防重录。
    body: {year, month(计提月), period(落账期间1-12), vouchers:[解析返回的凭证对象]}"""
    u = _require_perm(request, "logistics_post")
    if not u:
        return JSONResponse({"ok": False, "msg": "无「一键录入金蝶」权限（写正式账套的敏感操作，请联系管理员开通）"}, status_code=403)
    vouchers = body.get("vouchers") or []
    if not vouchers:
        return {"ok": False, "msg": "没有勾选任何凭证"}
    if len(vouchers) > 300:
        return {"ok": False, "msg": "一次最多录 300 张，请分批"}
    try:
        year = int(body.get("year") or 2026)
        period = int(body.get("period") or 0)
    except (TypeError, ValueError):
        return {"ok": False, "msg": "落账年份/月份格式不对"}
    if not (1 <= period <= 12):
        return {"ok": False, "msg": "请先选择凭证落账月份（1-12）"}
    blocked = _closed_block(year, period)    # 落账月份已封存 → 不许再往那个月录凭证
    if blocked:
        return blocked
    # 账单直采/活表行（带「摘要用语」字段）：服务端按行维度重拼摘要——摘要是防重幂等键，唯一事实源在服务端
    accrual_month = int(body.get("month") or period)
    for v in vouchers:
        if v.get("摘要用语") is not None:
            v["摘要"] = lb.make_summary(v.get("公司全名", ""), accrual_month, v.get("渠道"),
                                        v.get("业务线"), v.get("业务描述"), v.get("摘要用语"))

    # 连金蝶：供应商编码 + 防重录双闸——①金蝶已提交/已审核的同摘要凭证（草稿在列表查询里不可见）
    # ②工具自己的录入台账（草稿被人在金蝶删了会经 View 自愈放行）
    try:
        s, conf = kc.login()
        sup_codes = kc.supplier_code_map(kc.fetch_suppliers(s, conf))
        books = sorted({la.BOOK_CODE[v["主体"]] for v in vouchers if v.get("主体") in la.BOOK_CODE})
        existing = set()
        if books:
            ors = " or ".join(f"FACCOUNTBOOKID.FNumber='{b}'" for b in books)
            rows = kc._query(s, conf, "GL_VOUCHER", [("FEXPLANATION", "摘要")],
                             f"FYear={year} and FPeriod={period} and FEXPLANATION like '计提%' and ({ors})", "")
            existing = {r["摘要"] for r in rows if r.get("摘要")}
    except kc.KingdeeError as e:
        return {"ok": False, "msg": f"连接金蝶失败，本次没有录入任何凭证：{e}"}
    posted = db.logistics_posted(year, period)

    results = []
    for v in vouchers:
        zy = v.get("摘要", "")
        item = {"摘要": zy, "主体": v.get("主体", ""), "公司全名": v.get("公司全名", ""),
                "费用归属": v.get("费用归属", ""), "含税": v.get("含税")}
        problems = la.check_voucher_for_post(v, sup_codes)
        dup_log = posted.get(zy)
        if dup_log:                                   # 台账命中 → View 看草稿是否还在金蝶
            chk = kc.view_voucher(dup_log["kd_id"], s, conf)
            if not chk["exists"]:
                db.delete_logistics_post_log(dup_log["id"])   # 草稿已被删，放行重录
                dup_log = None
        if problems:
            item.update(status="blocked", msg="；".join(problems))
        elif zy in existing:
            item.update(status="skipped", msg="金蝶该月已有同摘要凭证（已提交/已审核），自动跳过防录重")
        elif dup_log:
            old_no = ("记-" + dup_log["vno"]) if dup_log["vno"] else ("单据号" + dup_log["billno"])
            item.update(status="skipped", 凭证号=old_no,
                        msg=f"本工具已录过（{dup_log['ts']} {dup_log['operator']} 录，凭证号 {old_no}，草稿在金蝶待提交），自动跳过防录重")
        else:
            try:
                model = la.build_kd_model(v, year, period, sup_codes)
                r = kc.save_voucher(model, s, conf)
                info = kc.view_voucher(r["id"], s, conf)      # 保存即分配记-字号，View 回查
                vno = info.get("vno") or ""
                item.update(status="saved", 单据编号=r["billno"], 内码=r["id"],
                            凭证号=("记-" + vno) if vno else r["billno"])
                db.log_logistics_post(year, period, zy, r["billno"], r["id"], vno, u["name"])
                posted[zy] = {"id": -1, "kd_id": r["id"], "billno": r["billno"], "vno": vno,
                              "operator": u["name"], "ts": _now()}   # 同批内重复摘要也拦
            except kc.KingdeeError as e:
                item.update(status="failed", msg=f"金蝶保存失败：{e}")
            except Exception as e:
                item.update(status="failed", msg=f"生成凭证报文失败：{e}")
        results.append(item)

    n_saved = sum(1 for r in results if r["status"] == "saved")
    n_skip = sum(1 for r in results if r["status"] == "skipped")
    n_fail = len(results) - n_saved - n_skip
    db.audit(u["name"], "物流计提-一键录入", f"{year}年{period}期",
             f"勾选{len(vouchers)}张：成功{n_saved} 跳过{n_skip} 失败{n_fail}")
    return {"ok": True, "year": year, "period": period,
            "成功": n_saved, "跳过": n_skip, "失败或拦下": n_fail, "results": results}


@router.get("/api/logistics-accrual/posted")
async def logistics_accrual_posted(request: Request, year: int = 2026, period: int = 0):
    """列出本工具在某落账期间录入金蝶的凭证（台账 + 金蝶实时状态），供撤销/删除草稿。"""
    if not _require_perm(request, "logistics_post"):
        return JSONResponse({"ok": False, "msg": "无「物流计提·一键录入金蝶」权限"}, status_code=403)
    if not (1 <= int(period or 0) <= 12):
        return {"ok": False, "msg": "请指定落账月份(1-12)"}
    logs = db.list_logistics_posts(int(year), int(period))
    if not logs:
        return {"ok": True, "year": year, "period": period, "items": []}
    try:
        s, conf = kc.login()
    except kc.KingdeeError as e:
        return {"ok": False, "msg": f"连接金蝶失败：{e}"}
    items = []
    for lg in logs:
        chk = kc.view_voucher(lg["kd_id"], s, conf)
        st = chk.get("status") or ""
        exists = chk.get("exists")
        items.append({
            "id": lg["id"], "摘要": lg["zhaiyao"], "单据编号": lg["billno"],
            "凭证号": ("记-" + lg["vno"]) if lg["vno"] else lg["billno"],
            "录入人": lg["operator"], "录入时间": lg["ts"],
            "金蝶状态": ("已删除" if not exists else _KD_STATUS_CN.get(st, st or "草稿")),
            # 可撤销 = 金蝶里已不存在(清台账即可) 或 仍是草稿态(暂存/创建/重新审核)，已提交/已审核不给删
            "可撤销": (not exists) or st in ("Z", "A", "D", ""),
        })
    return {"ok": True, "year": year, "period": period, "items": items}


@router.post("/api/logistics-accrual/unpost")
def logistics_accrual_unpost(body: dict, request: Request):
    """撤销录入：删除本工具录入金蝶的【草稿】凭证（并清台账）。已提交/已审核的不删，提示去金蝶反审核。
    body: {ids:[台账id...]}"""
    u = _require_perm(request, "logistics_post")
    if not u:
        return JSONResponse({"ok": False, "msg": "无「物流计提·一键录入金蝶」权限"}, status_code=403)
    ids = body.get("ids") or []
    if not ids:
        return {"ok": False, "msg": "没有勾选要撤销的凭证"}
    # 撤销的是"某张凭证落在哪个月"，按台账里记的落账期间判断——那个月封存了就不许再删（全有或全无，先查后删）
    for lid in ids:
        lg = db.get_logistics_post(lid)
        if lg:
            blocked = _closed_block(lg["year"], lg["period"])
            if blocked:
                return blocked
    try:
        s, conf = kc.login()
    except kc.KingdeeError as e:
        return {"ok": False, "msg": f"连接金蝶失败，未删除任何凭证：{e}"}
    results = []
    for lid in ids:
        lg = db.get_logistics_post(lid)
        if not lg:
            results.append({"id": lid, "status": "skipped", "msg": "台账无此记录（可能已撤销）"})
            continue
        item = {"id": lid, "摘要": lg["zhaiyao"], "凭证号": ("记-" + lg["vno"]) if lg["vno"] else lg["billno"]}
        chk = kc.view_voucher(lg["kd_id"], s, conf)
        if not chk.get("exists"):
            db.delete_logistics_post_log(lid)          # 金蝶里已没了 → 只清台账
            item.update(status="cleared", msg="金蝶里已不存在，已清理台账记录")
        elif (chk.get("status") or "") in ("B", "C"):  # 已提交/已审核 → 不删
            item.update(status="blocked", msg=f"金蝶状态『{_KD_STATUS_CN.get(chk.get('status'), chk.get('status'))}』，不能从这里删，请先去金蝶反审核/撤销提交")
        else:
            try:
                kc.delete_vouchers(lg["kd_id"], s, conf)
                db.delete_logistics_post_log(lid)
                item.update(status="deleted", msg="草稿已删除")
            except kc.KingdeeError as e:
                item.update(status="failed", msg=f"金蝶删除失败：{e}")
        results.append(item)
    n_del = sum(1 for r in results if r["status"] in ("deleted", "cleared"))
    n_block = sum(1 for r in results if r["status"] == "blocked")
    db.audit(u["name"], "物流计提-撤销录入", f"{len(ids)}张",
             f"删除{n_del} 拦下{n_block}")
    return {"ok": True, "删除": n_del, "拦下": n_block, "失败": len(results) - n_del - n_block, "results": results}

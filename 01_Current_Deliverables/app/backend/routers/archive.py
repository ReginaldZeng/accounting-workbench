# -*- coding: utf-8 -*-
# [Change Log]
# Date: 2026-08-05 | Author: Claude / c | Version: V2.172
# Description: 【凭证归档】路由（V2.172 从 app.py 拆出）。
#              本文件是「凭证归档」这条工具线在后端的唯一落点：改这条线的接口只动本文件，
#              不再碰 app.py —— 这样多条需求并行开发时互不冲突。
#              共享的配置/期间/权限判定见 core.py；算法在 kernels/archive.py。
#              app.py 只负责 include_router(router)，不感知本文件内部。

from fastapi import APIRouter
from fastapi import Request
from fastapi.responses import Response
from io import BytesIO

from core import (
    JSONResponse, _current_user, _require_perm, db,
)

router = APIRouter()


# ---------------- 凭证归档（其它小工具；登记/转移/借出限总账会计，销毁执行限财务经理） ----------------
@router.get("/api/archive/orgs")
def archive_orgs(request: Request):
    """登记新册用的主体下拉：只列启用且有简码的（没简码不能生成册号）。"""
    if not _current_user(request):
        return JSONResponse({"ok": False}, status_code=401)
    return {"ok": True, "orgs": [{"short_name": o["short_name"], "code": o["code"], "color": o["color"]}
                                 for o in db.list_orgs() if o["active"] and o["code"]]}


@router.get("/api/archive/find")
def archive_find(request: Request, org: str = "", year: int = 0, month: int = 0, no: int = 0):
    """按主体+年月+凭证号定位册子。三者缺一不可（否则会查出好几本）。"""
    if not _current_user(request):
        return JSONResponse({"ok": False}, status_code=401)
    if not (org and year and month and no):
        return {"ok": False, "msg": "主体、年月、凭证号三项缺一不可"}
    hits = db.find_volume(org, year, month, no)
    for h in hits:
        h["trail"] = db.volume_trail(h["vol_no"])
    return {"ok": True, "hits": hits}


@router.get("/api/archive/volumes")
def archive_volumes_list(request: Request, org: str = "", year: int = 0, status: str = ""):
    if not _current_user(request):
        return JSONResponse({"ok": False}, status_code=401)
    return {"ok": True, "volumes": db.list_volumes(org or None, year or None, status or None)}


@router.get("/api/archive/period-info")
def archive_period_info(request: Request, org: str = "", year: int = 0, month: int = 0):
    """登记新册的自动带出：该主体该期间已登记册子 → 下一册序、起号自动接、号段占用。"""
    if not _current_user(request):
        return JSONResponse({"ok": False}, status_code=401)
    if not (org and year and month):
        return {"ok": False, "msg": "缺参数"}
    ranges, seqs = db._period_ranges(org, year, month)
    top = max([b for _, b in ranges], default=0)
    from kernels import archive as arch
    return {"ok": True, "next_seq": arch.next_seq(seqs), "count": len(seqs),
            "suggest_from": top + 1, "used": sorted(ranges)}


@router.post("/api/archive/register")
def archive_register(body: dict, request: Request):
    u = _require_perm(request, "archive_edit")
    if not u:
        return JSONResponse({"ok": False, "msg": "无「凭证归档·登记」权限（限总账会计），请联系管理员"}, status_code=403)
    try:
        vol_no = db.register_volume(body, u["name"])
    except ValueError as e:
        return {"ok": False, "msg": str(e)}
    db.audit(u["name"], "凭证归档-登记新册", vol_no)
    return {"ok": True, "vol_no": vol_no}


@router.get("/api/archive/locations")
def archive_locations(request: Request):
    if not _current_user(request):
        return JSONResponse({"ok": False}, status_code=401)
    return {"ok": True, "locations": db.list_locations()}


@router.post("/api/archive/locations/save")
def archive_location_save(body: dict, request: Request):
    u = _require_perm(request, "archive_edit")
    if not u:
        return JSONResponse({"ok": False, "msg": "无「凭证归档·维护位置」权限"}, status_code=403)
    try:
        lid = db.save_location(body, u["name"])
    except ValueError as e:
        return {"ok": False, "msg": str(e)}
    db.audit(u["name"], "凭证归档-位置维护", str(body.get("name", "")))
    return {"ok": True, "id": lid}


@router.post("/api/archive/transfer")
def archive_transfer(body: dict, request: Request):
    u = _require_perm(request, "archive_edit")
    if not u:
        return JSONResponse({"ok": False, "msg": "无「凭证归档·转移」权限"}, status_code=403)
    vol_nos = body.get("vol_nos") or []
    expected = {k: (v or None) for k, v in (body.get("expected") or {}).items()}
    to_id = body.get("to_id")
    reason = str(body.get("reason") or "转移")
    if not vol_nos:
        return {"ok": False, "msg": "没选册子"}
    try:
        ok, conflicts, tno = db.transfer_volumes(vol_nos, expected, to_id, reason, u["name"])
    except ValueError as e:
        return {"ok": False, "msg": str(e)}
    if not ok:
        return {"ok": False, "conflict": True, "conflicts": conflicts,
                "msg": "其中 %d 本在你勾选之后被别人挪走了，请刷新后重试" % len(conflicts)}
    db.audit(u["name"], "凭证归档-批量转移", tno, "%d 本 → %s" % (len(vol_nos), reason))
    return {"ok": True, "transfer_no": tno}


@router.post("/api/archive/borrow")
def archive_borrow(body: dict, request: Request):
    u = _require_perm(request, "archive_edit")
    if not u:
        return JSONResponse({"ok": False, "msg": "无「凭证归档·借出」权限"}, status_code=403)
    try:
        tno = db.borrow_volumes(body.get("vol_nos") or [], body.get("borrower"), body.get("due_date"), u["name"])
    except ValueError as e:
        return {"ok": False, "msg": str(e)}
    db.audit(u["name"], "凭证归档-借出", str(body.get("borrower", "")), tno)
    return {"ok": True, "transfer_no": tno}


@router.post("/api/archive/return")
def archive_return(body: dict, request: Request):
    u = _require_perm(request, "archive_edit")
    if not u:
        return JSONResponse({"ok": False, "msg": "无「凭证归档·归还」权限"}, status_code=403)
    try:
        tno = db.return_volumes(body.get("vol_nos") or [], u["name"])
    except ValueError as e:
        return {"ok": False, "msg": str(e)}
    db.audit(u["name"], "凭证归档-归还", "", tno)
    return {"ok": True, "transfer_no": tno}


@router.post("/api/archive/destroy/apply")
def archive_destroy_apply(body: dict, request: Request):
    u = _require_perm(request, "archive_edit")
    if not u:
        return JSONResponse({"ok": False, "msg": "无「凭证归档·销毁申请」权限（限总账会计）"}, status_code=403)
    try:
        tno = db.destroy_apply(body.get("vol_nos") or [], body.get("approve_no"), body.get("batch_name"), u["name"])
    except ValueError as e:
        return {"ok": False, "msg": str(e)}
    db.audit(u["name"], "凭证归档-销毁申请", str(body.get("approve_no", "")), tno)
    return {"ok": True, "transfer_no": tno}


@router.post("/api/archive/destroy/cancel")
def archive_destroy_cancel(body: dict, request: Request):
    u = _require_perm(request, "archive_edit")
    if not u:
        return JSONResponse({"ok": False, "msg": "无权限"}, status_code=403)
    try:
        tno = db.destroy_cancel(body.get("vol_nos") or [], u["name"])
    except ValueError as e:
        return {"ok": False, "msg": str(e)}
    db.audit(u["name"], "凭证归档-撤回销毁", "", tno)
    return {"ok": True, "transfer_no": tno}


@router.post("/api/archive/destroy/execute")
def archive_destroy_execute(body: dict, request: Request):
    u = _require_perm(request, "archive_destroy")
    if not u:
        return JSONResponse({"ok": False, "msg": "销毁执行限【财务经理】（archive_destroy 权限），总账会计只能申请"}, status_code=403)
    try:
        tno = db.destroy_execute(body.get("vol_nos") or [], u["name"])
    except ValueError as e:
        return {"ok": False, "msg": str(e)}
    db.audit(u["name"], "凭证归档-销毁执行", "", tno)
    return {"ok": True, "transfer_no": tno}


# ---------------- 凭证归档：Excel 批量导入（期初存量）+ 号段体检 ----------------
_ARCH_IMPORT_COLS = ["主体", "年", "月", "凭证号起", "凭证号止", "存放位置", "凭证类型", "备注"]


@router.get("/api/archive/import-template")
def archive_import_template(request: Request):
    """下载批量导入模板 xlsx。列：主体/年/月/凭证号起/凭证号止/存放位置/凭证类型/备注。"""
    if not _current_user(request):
        return JSONResponse({"ok": False}, status_code=401)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook(); ws = wb.active; ws.title = "凭证册导入"
    ws.append(_ARCH_IMPORT_COLS)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="4B53C4")
    ws.append(["深圳星期零", 2026, 1, 1, 50, "档案室B › 3号柜 › 第1层", "记账凭证", "示例行，可删"])
    ws.append(["深圳星期零", 2026, 1, 51, 96, "档案室B › 3号柜 › 第1层", "记账凭证", ""])
    widths = [14, 6, 6, 9, 9, 30, 12, 20]
    for i, wd in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = wd
    bio = BytesIO(); wb.save(bio)
    return Response(content=bio.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=archive_import_template.xlsx"})


@router.post("/api/archive/import")
async def archive_import(request: Request):
    """上传填好的模板，逐行校验+登记，返回成功/失败清单。写操作限 archive_edit。"""
    u = _require_perm(request, "archive_edit")
    if not u:
        return JSONResponse({"ok": False, "msg": "无「凭证归档·登记」权限，请联系管理员"}, status_code=403)
    data = await request.body()
    import openpyxl
    try:
        wb = openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=True)
    except Exception:
        return {"ok": False, "msg": "文件打不开，请确认是 .xlsx 格式（用下载的模板填）"}
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    header = next(it, None)
    if not header:
        return {"ok": False, "msg": "表是空的"}
    hmap = {str(h).strip(): i for i, h in enumerate(header) if h is not None}
    need = ["主体", "年", "月", "凭证号起", "凭证号止"]
    miss = [n for n in need if n not in hmap]
    if miss:
        return {"ok": False, "msg": "缺少必填列：%s（请用下载的模板）" % "、".join(miss)}
    rows = []
    for r in it:
        if r is None or all(v is None or str(v).strip() == "" for v in r):
            continue
        rows.append({col: (r[hmap[col]] if col in hmap and hmap[col] < len(r) else None) for col in _ARCH_IMPORT_COLS})
    if not rows:
        return {"ok": False, "msg": "没有数据行（只有表头）"}
    res = db.import_volumes(rows, u["name"])
    db.audit(u["name"], "凭证归档-批量导入", "", "成功 %d / 失败 %d" % (len(res["成功"]), len(res["失败"])))
    return res


@router.get("/api/archive/checkup")
def archive_checkup(request: Request, org: str = "", year: int = 0, month: int = 0, total: int = 0):
    """号段体检：给定该主体该期间的凭证总张数 total（手工填 或 从金蝶取），列出没有任何册子覆盖的缺口。
    total 来源双模式：①手工填（会计自己看金蝶或凭证封面数）②从金蝶取（/api/archive/kingdee-count）。"""
    if not _current_user(request):
        return JSONResponse({"ok": False}, status_code=401)
    if not (org and year and month and total):
        return {"ok": False, "msg": "主体、年月、凭证总张数都要有"}
    gaps = db.range_checkup(org, year, month, total)
    return {"ok": True, "gaps": gaps, "total": total,
            "covered_ok": len(gaps) == 0}


@router.get("/api/archive/kingdee-count")
def archive_kingdee_count(request: Request, year: int = 0, month: int = 0, book_code: str = ""):
    """从金蝶取该账簿该期间的凭证张数（记字最大号）。只读、best-effort；连不上则报错让人改用手工填。"""
    if not _current_user(request):
        return JSONResponse({"ok": False}, status_code=401)
    if not (year and month and book_code):
        return {"ok": False, "msg": "缺年月或账簿代码"}
    try:
        import kingdee_client as kd
        cnt = kd.fetch_voucher_count(year, month, book_code)
        return {"ok": True, "count": cnt}
    except Exception as e:
        return {"ok": False, "msg": "金蝶取数失败（%s）——可改用手工填张数" % (str(e)[:80])}

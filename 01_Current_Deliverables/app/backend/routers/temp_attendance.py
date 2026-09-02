# -*- coding: utf-8 -*-
# [Change Log]
# Date: 2026-08-18 | Author: Claude / c | Version: V2.318
# Description: 【临时工考勤】路由（样机）。本文件是这条工具线在后端的唯一落点：
#              改这条线的接口只动本文件，不碰 app.py（app.py 只 include_router）。
#              算法在 kernels/temp_attendance.py，本层只做「收文件 → 调内核 → 出 JSON/Excel」。
#              对金蝶只读、不写、不产生凭证。
#
#              V2.320 起按期留档（2026-08-19 定案，改掉原先「算完即走什么都不留」）：
#                · 复核结论 → 长期落库，选月份即可回看，不必重新上传
#                · 原始两张表 → 只按期落盘保留最近 KEEP_MONTHS 期，到期自动清
#              分开处理的理由：打卡表含全员姓名与每天的上下班时刻，是人事敏感数据，
#              没必要长期堆在服务器上；而「哪一期核过、结论是什么、谁核的」必须长期可查
#              ——一个复核工具不留复核记录，本身就说不过去。
#              代价说清楚：过了留存窗口就只能看结论，不能改参数重跑（原表没了）。

import calendar
import gzip
import json
import os
import re
import shutil
import threading
import time
from urllib.parse import quote
from base64 import b64decode, b64encode
from datetime import datetime, timedelta
from io import BytesIO

from fastapi import APIRouter, Request
from fastapi.responses import Response

from kernels import temp_attendance as ta

try:
    from kernels import dingtalk_attendance as dda
except Exception:                       # 缺 requests / 没配钉钉都不该拖垮整条线
    dda = None

from core import JSONResponse, _require_perm, db

router = APIRouter()

CAP = "tempatt_review"
CAP_RATES = "tempatt_rates"
CAP_BOARD = "tempatt_board"
_RESULT_KEY = "tempatt_result_"        # + YYYY-MM，压缩后的结论快照（长期）
_META_KEY = "tempatt_meta_"            # + YYYY-MM，小摘要，给期次列表用（长期）
_KEEP_KEY = "tempatt_keep_months"      # 原表保留期数，可在库里改
_BOARD_KEY = "tempatt_board_cache"     # 看板汇总缓存，按期次签名自动失效
KEEP_MONTHS_DEFAULT = 6

_BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
UPLOAD_DIR = os.path.join(_BASE, "tempatt_uploads")     # tempatt_uploads/{YYYY-MM}/
_MONTH_RE = re.compile(r"^20\d{2}-\d{2}$")


# ==================== 合同价：按行各自设生效期 ====================
# 成本会计维护的合同价，**每一行（派遣方 × 岗位）各自带生效日与失效日**（2026-08-22 定案）。
#
# 为什么不是「整表一个版本」：各家派遣方的续签时间本来就不同。锦绣 8 月调价，
# 不该逼着把华顺、恒祺的价也跟着重述一遍——那样既是无谓的重复录入，
# 又会让人误以为所有家都在 8 月动过价。
#
# 失效日可以留空：留空＝**同一行（同派遣方同岗位）下一条的生效日前一天**；再没有下一条就是「至今」。
# 但允许显式填——合同到期不再续、这家停止合作时，失效日是真实存在的，推不出来。
# ⚠ 显式填了就可能出现**重叠**（两条区间盖在一起）或**空档**（中间没有任何一条覆盖），
#   这两种都会让某期取不到价或取错价，所以逐条检出并在页面上标红，不静默放过。
_CONTRACT_KEY = "tempatt_contract"      # {"行": [...]}
_DATE_RE = re.compile(r"^20\d{2}-\d{2}-\d{2}$")


def _band(v):
    """[员工工资, 管理费] → 规范成 (w, m)；空/无 → None。"""
    if not v or not isinstance(v, (list, tuple)) or len(v) < 2:
        return None
    try:
        return [float(v[0]), float(v[1])]
    except Exception:
        return None


def _row_id(r):
    return f"{r.get('派遣方','')}|{r.get('岗位','')}|{r.get('生效日','')}"


def _contract_rows():
    """规范化 + 算出实际失效日 + 检出重叠/空档。返回按(派遣方,岗位,生效日)排好的行。"""
    raw = (db.get_setting(_CONTRACT_KEY) or {}).get("行")
    if not isinstance(raw, list):
        return []
    rows = []
    for x in raw:
        if not isinstance(x, dict):
            continue
        if not _DATE_RE.match(str(x.get("生效日") or "")):
            continue
        if not str(x.get("派遣方") or "").strip():
            continue
        eff_end = str(x.get("失效日") or "").strip()
        rows.append({
            "派遣方": str(x["派遣方"]).strip(),
            "岗位": str(x.get("岗位") or ta.POST_DEFAULT).strip() or ta.POST_DEFAULT,
            "生效日": x["生效日"],
            "失效日": eff_end if _DATE_RE.match(eff_end) else None,
            "day": _band(x.get("day")), "night": _band(x.get("night")),
            "备注": str(x.get("备注") or ""),
            # 谁、什么时间录的，逐行记（需求方 2026-08-22 追加）。旧字段名一并认，免得历史数据丢人名
            "录入人": x.get("录入人") or x.get("维护人") or "",
            "录入时间": x.get("录入时间") or x.get("维护时间") or "",
        })
    rows.sort(key=lambda r: (r["派遣方"], r["岗位"], r["生效日"]))
    by = {}
    for r in rows:
        by.setdefault((r["派遣方"], r["岗位"]), []).append(r)
    for key, g in by.items():
        for i, r in enumerate(g):
            nxt = g[i + 1] if i + 1 < len(g) else None
            if r["失效日"]:
                r["实际失效日"], r["失效日来源"] = r["失效日"], "手填"
            elif nxt:
                r["实际失效日"] = (datetime.strptime(nxt["生效日"], "%Y-%m-%d")
                                   - timedelta(days=1)).strftime("%Y-%m-%d")
                r["失效日来源"] = "自动（下一条生效日前一天）"
            else:
                r["实际失效日"], r["失效日来源"] = None, "至今"
            # 当前状态＝相对**今天**（跟「本期生效」不是一回事，那个相对所选期次）。
            # 已经失效的行留着是有用的——历史期次要按当时的价核，所以不能删；
            # 但摆在一起容易被当成还在用的价，所以标出来、导出里还打灰底。
            today = datetime.now().strftime("%Y-%m-%d")
            if r["生效日"] > today:
                r["当前状态"] = "未生效"
            elif r["实际失效日"] and r["实际失效日"] < today:
                r["当前状态"] = "已失效"
            else:
                r["当前状态"] = "生效中"
            r["问题"] = []
            if r["失效日"] and r["失效日"] < r["生效日"]:
                r["问题"].append("失效日早于生效日")
            if nxt and r["实际失效日"]:
                if r["实际失效日"] >= nxt["生效日"]:
                    r["问题"].append(f"与下一条（{nxt['生效日']} 起）区间重叠")
                else:
                    gap0 = (datetime.strptime(r["实际失效日"], "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")
                    if gap0 < nxt["生效日"]:
                        r["问题"].append(f"到下一条之间有空档：{gap0} ~ "
                                         + (datetime.strptime(nxt["生效日"], "%Y-%m-%d")
                                            - timedelta(days=1)).strftime("%Y-%m-%d"))
            r["id"] = _row_id(r)
    return rows


def _month_span(month):
    d0 = datetime.strptime(month + "-01", "%Y-%m-%d")
    last = calendar.monthrange(d0.year, d0.month)[1]
    return month + "-01", f"{month}-{last:02d}"


def _contract_for(month):
    """某期适用的合同价表 + 逐格说明。返回 (table or None, info)。"""
    rows = _contract_rows()
    if not rows or not _MONTH_RE.match(month or ""):
        return None, {"有行": bool(rows)}
    d0, d1 = _month_span(month)
    table, mid, applied = {}, [], []
    by = {}
    for r in rows:
        by.setdefault((r["派遣方"], r["岗位"]), []).append(r)
    for (agency, post), g in by.items():
        hit = [r for r in g
               if r["生效日"] <= d1 and (r["实际失效日"] is None or r["实际失效日"] >= d0)]
        if not hit:
            continue
        if len(hit) > 1:
            # 期内换过价：本工具按整月一个价核对（人力结算表本身一人一价按月结），取覆盖期末那条
            mid.append({"派遣方": agency, "岗位": post,
                        "生效日": [r["生效日"] for r in hit]})
        cur = sorted(hit, key=lambda r: r["生效日"])[-1]
        table.setdefault(agency, {})[post] = {"day": cur["day"], "night": cur["night"]}
        applied.append({"派遣方": agency, "岗位": post, "生效日": cur["生效日"],
                        "失效日": cur["实际失效日"], "备注": cur["备注"],
                        "录入人": cur["录入人"], "录入时间": cur["录入时间"]})
    return (table or None), {"有行": True, "适用": applied, "期中调价": mid,
                             "问题行": [{"id": r["id"], "派遣方": r["派遣方"], "岗位": r["岗位"],
                                        "生效日": r["生效日"], "问题": r["问题"]}
                                       for r in rows if r["问题"]]}


def _contract_with_info(month):
    """该期适用的合同价表 + 挑行时的说明（期中调价、问题行）。
    登记表一行都没覆盖 → (None, info) → 内核按「全部缺档／待核」处理，**不拿任何别的表顶上**
    （V2.344 前「按月保存」的旧表回读已删：它没有生效期，回读了还得把来源标成"登记表"，张冠李戴）。"""
    if not month:
        return None, {}
    try:
        return _contract_for(month)
    except Exception:
        return None, {}


# ==================== 按期留档 ====================
def _keep_months():
    try:
        return max(1, int(db.get_setting(_KEEP_KEY) or KEEP_MONTHS_DEFAULT))
    except Exception:
        return KEEP_MONTHS_DEFAULT


def _period_dir(month):
    return os.path.join(UPLOAD_DIR, month) if _MONTH_RE.match(month or "") else None


def _pack(obj):
    """结论快照压缩后再进 KV。一期 300 人日的完整结果 JSON 是 MB 级，
    原文塞 app_settings 会把这张表撑得没法看；gzip 后通常只剩十分之一。"""
    raw = json.dumps(obj, ensure_ascii=False).encode("utf-8")
    return {"z": b64encode(gzip.compress(raw, 6)).decode("ascii"), "n": len(raw)}


def _unpack(v):
    if not isinstance(v, dict) or "z" not in v:
        return v if isinstance(v, dict) else None
    try:
        return json.loads(gzip.decompress(b64decode(v["z"])).decode("utf-8"))
    except Exception:
        return None


def _purge_old_uploads(keep=None):
    """只清原始表，不动结论。清掉的期次在 meta 里标成「原表已过留存期」，
    页面据此把「改参数重跑」按钮置灰——而不是让人点了才发现跑不了。"""
    keep = keep or _keep_months()
    if not os.path.isdir(UPLOAD_DIR):
        return []
    months = sorted((d for d in os.listdir(UPLOAD_DIR) if _MONTH_RE.match(d)), reverse=True)
    dropped = []
    for m in months[keep:]:
        try:
            shutil.rmtree(os.path.join(UPLOAD_DIR, m))
            dropped.append(m)
            meta = db.get_setting(_META_KEY + m) or {}
            if meta:
                meta["原表在库"] = False
                db.set_setting(_META_KEY + m, meta, "系统·到期清理")
        except Exception:
            pass
    return dropped


def _save_period(month, summary_bytes, punch_bytes, names, res, user):
    """存一期：原表落盘（留窗口）＋ 结论进库（长期）。存不下也不能让核对本身失败。"""
    if not _MONTH_RE.match(month or ""):
        return {"已留档": False, "原因": "月份不是 YYYY-MM，未留档"}
    try:
        d = _period_dir(month)
        os.makedirs(d, exist_ok=True)
        for fn, data in (("summary.xlsx", summary_bytes), ("punch.xlsx", punch_bytes)):
            with open(os.path.join(d, fn), "wb") as f:
                f.write(data)
        st = res["stats"]
        tot = (res.get("settle") or {}).get("合计") or {}
        meta = {
            "月份": month,
            "跑批时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "跑批人": user.get("name", ""),
            "人数": st.get("人数"), "比对人日": st.get("比对人日"),
            "上报总工时": st.get("上报总工时"), "重算总工时": st.get("重算总工时"),
            "应付合计": tot.get("表上合计") if tot.get("表上合计") is not None else tot.get("应付合计"),
            "结论": tot.get("结论"), "缺合同价人数": st.get("缺合同价人数"),
            "异常多记日次": st.get("异常多记日次"),
            "金额核对条数": st.get("金额核对条数"),
            "合同外调整异常": (st.get("合同外调整合计") or {}).get("异常"),
            "同名跨派遣方数": st.get("同名跨派遣方数"),
            "归属与打卡不符": len(st.get("归属与打卡不符") or []),
            "原表在库": True,
            "原表文件名": {"汇总表": names.get("summary", ""), "打卡表": names.get("punch", "")},
            "跑批次数": int((db.get_setting(_META_KEY + month) or {}).get("跑批次数") or 0) + 1,
        }
        db.set_setting(_RESULT_KEY + month, _pack(res), user.get("name", ""))
        db.set_setting(_META_KEY + month, meta, user.get("name", ""))
        dropped = _purge_old_uploads()
        return {"已留档": True, "到期清理": dropped, "跑批次数": meta["跑批次数"]}
    except Exception as e:
        return {"已留档": False, "原因": f"留档失败（不影响本次核对结果）：{e}"}


def _load_period(month):
    """读某期已留档的结论。返回 (结果, meta)；没留过返回 (None, None)。"""
    if not _MONTH_RE.match(month or ""):
        return None, None
    res = _unpack(db.get_setting(_RESULT_KEY + month))
    meta = db.get_setting(_META_KEY + month) or None
    return res, meta


def _period_files(month):
    """取该期落盘的原表字节；过了留存期返回 (None, None)。"""
    d = _period_dir(month)
    if not d or not os.path.isdir(d):
        return None, None
    try:
        with open(os.path.join(d, "summary.xlsx"), "rb") as f1, \
                open(os.path.join(d, "punch.xlsx"), "rb") as f2:
            return f1.read(), f2.read()
    except Exception:
        return None, None


# ==================== 认定：确认某条可疑项无误 ====================
# ⚠ 是「认定」，不是「删除」。认定后那条**仍在**，只是移进「已认定」区并标明谁、什么时候、因为什么认的。
#    一个能把发现无痕抹掉的审计工具，本身就没有价值——所以这里不提供删除，只提供表态与撤销。
#
# 范围两档：
#   · 本期    —— 只对这一期生效（逐日类只能选这档：日期是期内的，跨期认定没有意义）
#   · 长期    —— 以后每期同一条都不再报。给「浮动组」这种长期安排用；
#                但长期认定会持续压住提示，所以页面上永远显示「另有 N 项已认定」并可展开，不让它彻底消失。
_ACK_KEY = "tempatt_ack_"              # + YYYY-MM，本期认定
_ACK_STANDING = "tempatt_ack_standing"  # 长期认定
# 能长期认定的只有「身份类」发现；逐日类带日期，长期无意义
_ACK_LONG_OK = ("归属不符", "同名", "打卡重名")
_ACK_TYPES = _ACK_LONG_OK + ("多记", "待查", "金额核对", "奖罚")


def _acks(month):
    """该期生效的认定 = 本期认定 ∪ 长期认定。返回 {(类型, 键): 认定记录}。"""
    out = {}
    for src, scope in ((db.get_setting(_ACK_STANDING) or {}, "长期"),
                       (db.get_setting(_ACK_KEY + str(month or "")) or {}, "本期")):
        if not isinstance(src, dict):
            continue
        for k, v in src.items():
            if isinstance(v, dict) and "|" in k:
                t, _, key = k.partition("|")
                out[(t, key)] = dict(v, 范围=scope)
    return out


def _mark(item, t, key, acks):
    """给一条发现盖上认定信息（盖上就是了，不删）。返回是否已认定。"""
    a = acks.get((t, key))
    if a:
        item["已认定"] = a
    return bool(a)


def _mark_legacy(res):
    """V2.346 之前的留档快照里没有「缺合同价」「单价核对」——那是按旧规则（默认表兜底）算出来的，
    结论列满屏「正常」是假的。读出来/导出时打上标记，结论一律改「待核」并说明要重跑。"""
    st = res.get("stats") or {}
    if "缺合同价人数" in st:
        # 这一期做过合同价核对（V2.346+），但**逐人的**「单价不符」是 V2.349 才挂到人身上的。
        # 缺这个键时导出③那一列不能写「✓ 一致」——那等于给一份没核过的名单盖清白章（V2.349 审出）。
        ppl = res.get("people") or []
        if ppl and not any("单价不符" in x for x in ppl):
            res["缺逐人单价核对"] = True
        # V2.352 之前跑的留档没有「上报口径」这个键，它的四档是按 punch 口径算的。
        # 标出来，让页面按 punch 渲染并提示重跑——否则「少记 1410」会被印成「✓ 撑得住 1410」。
        if "上报口径" not in st:
            res["旧口径留档"] = True
            st["上报口径"] = "punch"
        return res
    res["旧版留档"] = True
    res["缺逐人单价核对"] = True
    se = res.get("settle") or {}
    for r in (se.get("明细") or []) + (se.get("派遣方小计") or []) + (se.get("业务线小计") or []) + \
             ([se["合计"]] if se.get("合计") else []):
        r["结论"] = "待核"
        r["异常原因"] = ["旧版留档：按旧规则算的，没有合同价核对，请点「按当前参数重跑」"]
    return res


def _apply_acks(res, month):
    """把认定信息盖到结果上，并统计未认定/已认定条数。
    **不从列表里剔除任何一条**——页面自己按「已认定」分区展示。"""
    st = res["stats"]
    # 「打卡表重名」内核出的是纯姓名列表，这里统一成对象——**不管有没有认定都转**，
    # 否则前端要同时应付两种形态（早先只在有认定时转，页面渲染必然出岔子）
    dups = st.get("打卡表重名") or []
    if dups and isinstance(dups[0], str):
        st["打卡表重名"] = [{"姓名": d} for d in dups]
    acks = _acks(month)
    if not acks:
        st["认定"] = {"已认定": 0, "明细": []}
        st["异常多记未认定"] = st.get("异常多记日次") or 0
        st["归属不符未认定"] = len(st.get("归属与打卡不符") or [])
        ta.settle_verdict(res)
        return res
    n = 0
    for x in st.get("归属与打卡不符") or []:
        n += _mark(x, "归属不符", x.get("姓名", ""), acks)
    for x in st.get("同名多行") or []:
        n += _mark(x, "同名", x.get("归一姓名", ""), acks)
    for x in st.get("打卡表重名") or []:
        if isinstance(x, dict):
            n += _mark(x, "打卡重名", x.get("姓名", ""), acks)
    for x in st.get("金额核对") or []:
        n += _mark(x, "金额核对", f"{x.get('姓名','')}|{x.get('项目','')}", acks)
    for x in st.get("合同外调整") or []:
        n += _mark(x, "奖罚", f"{x.get('姓名','')}|{x.get('项目','')}", acks)
    for r in res.get("rows") or []:
        if r.get("档") == "over_out":
            # 键含归属：同名不同派遣方的两人同一天不再撞键、互相误认（复查揪出）。
            # 老格式 `姓名|日` 的历史认定用 or 兜底读回，不至于因换键丢掉已认定。
            n += (_mark(r, "多记", f"{r.get('姓名','')}|{r.get('归属','')}|{r.get('日','')}", acks)
                  or _mark(r, "多记", f"{r.get('姓名','')}|{r.get('日','')}", acks))
        elif r.get("档") in ("hard", "thin"):
            n += (_mark(r, "待查", f"{r.get('姓名','')}|{r.get('归属','')}|{r.get('日','')}", acks)
                  or _mark(r, "待查", f"{r.get('姓名','')}|{r.get('日','')}", acks))
    st["认定"] = {
        "已认定": n,
        "明细": [{"类型": t, "键": k, **v} for (t, k), v in sorted(acks.items())],
    }
    ta.settle_verdict(res)          # 必须在盖完认定之后：认定过的不再算异常
    # 四档统计里把已认定的多记单列出来，页面顶部的状态条才不会一边说「0 异常」一边亮红
    st["异常多记未认定"] = sum(1 for r in (res.get("rows") or [])
                              if r.get("档") == "over_out" and not r.get("已认定"))
    st["归属不符未认定"] = sum(1 for x in (st.get("归属与打卡不符") or []) if not x.get("已认定"))
    return res


@router.get("/api/tempatt/contract")
def tempatt_contract(request: Request, month: str = ""):
    """合同价明细行（每行自带生效/失效日）+ 某期适用哪些行。"""
    if not _require_perm(request, CAP):
        return JSONResponse({"ok": False, "msg": "无「临时工考勤」权限，请联系管理员"}, status_code=403)
    rows = _contract_rows()
    out = {"ok": True, "行": rows, "本期": month, "可维护": bool(_require_perm(request, CAP_RATES))}
    if _MONTH_RE.match(month or ""):
        t, info = _contract_for(month)
        out.update({"适用表": t or {}, **info})
        d0, d1 = _month_span(month)
        out["期间"] = {"起": d0, "止": d1}
        for r in rows:                      # 标出哪几行在本期生效，页面据此高亮
            r["本期生效"] = bool(r["生效日"] <= d1 and (r["实际失效日"] is None or r["实际失效日"] >= d0))
    return out


@router.post("/api/tempatt/contract/row/save")
async def tempatt_contract_row_save(request: Request):
    """新增或改一行合同价。键＝派遣方 + 岗位 + 生效日；同键再存＝覆盖那一行。"""
    u = _require_perm(request, CAP_RATES)
    if not u:
        return JSONResponse({"ok": False, "msg": "无「维护合同价登记表」权限，请联系管理员"}, status_code=403)
    b = await request.json()
    agency = str(b.get("派遣方") or "").strip()
    post = str(b.get("岗位") or "").strip() or ta.POST_DEFAULT
    eff = str(b.get("生效日") or "").strip()
    end = str(b.get("失效日") or "").strip()
    if not agency:
        return {"ok": False, "msg": "请填派遣方"}
    if not _DATE_RE.match(eff):
        return {"ok": False, "msg": "生效日请填 YYYY-MM-DD"}
    if end and not _DATE_RE.match(end):
        return {"ok": False, "msg": "失效日请填 YYYY-MM-DD，或留空＝到下一条生效日前一天"}
    if end and end < eff:
        return {"ok": False, "msg": "失效日不能早于生效日"}
    day, night = _band(b.get("day")), _band(b.get("night"))
    if not day and not night:
        return {"ok": False, "msg": "白班、夜班至少要填一档单价"}
    cur = db.get_setting(_CONTRACT_KEY) or {}
    raw = [x for x in (cur.get("行") or []) if isinstance(x, dict)]
    key = (agency, post, eff)
    raw = [x for x in raw if (str(x.get("派遣方") or "").strip(),
                              str(x.get("岗位") or ta.POST_DEFAULT).strip(),
                              str(x.get("生效日") or "")) != key]
    raw.append({"派遣方": agency, "岗位": post, "生效日": eff, "失效日": end or None,
                "day": day, "night": night, "备注": str(b.get("备注") or "").strip(),
                "录入人": u["name"], "录入时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S")})
    db.set_setting(_CONTRACT_KEY, {"行": raw}, u["name"])
    try:
        db.audit(u["name"], "临时工考勤-维护合同价行", f"{agency}·{post} {eff} 起",
                 f"白{day}／夜{night}｜失效 {end or '自动'}｜{b.get('备注') or ''}")
    except Exception:
        pass
    rows = _contract_rows()
    bad = [r for r in rows if r["问题"] and (r["派遣方"], r["岗位"]) == (agency, post)]
    return {"ok": True, "行": rows,
            "提醒": [f"{r['生效日']} 起那一行：{'；'.join(r['问题'])}" for r in bad] or None}


# ── 合同价登记表：导出 / 导入 ────────────────────────────────────────
# 为什么导入要分两步（先预览、再确认写）：这张表是所有钱的唯一基准，
# 一次盲写可能把整月的应付算错，而且错了不容易看出来——页面上仍然一片绿，
# 只是绿得不对。所以先把「新增哪几行、覆盖哪几行、旧值→新值是什么」摆出来。
#
# 合并语义：键＝派遣方＋岗位＋生效日，同键覆盖、新键追加。
# **导入不删行**——文件里没有、表里已有的行原样保留。要删就在页面上删，
# 免得有人拿一份旧文件一导，把别人后来加的价悄悄清掉。
# 合计两列是**公式**（员工工资＋管理费＝结算表上那个含管理费单价），改了分项自动跟着变；
# 导入时忽略它们，免得有人只改合计不改分项、两边对不上还不知道以哪个为准。
# 列名与列序按使用者 2026-08-29 改过的那张来：**指标在上、班次在下**，备注挪到最右。
# 导入不认死这套写法（_hdr_key 按词判），换个写法照样读得出。
_C_COLS = ("派遣方", "岗位",
           "员工工资\n白班", "管理费\n白班", "合计\n白班",
           "员工工资\n夜班", "管理费\n夜班", "合计\n夜班",
           "生效日", "失效日")
# 「当前状态」是算出来的，不是填的——导入时忽略这一列。放在失效日后面，跟它推出来的东西挨着。
_C_INFO = ("当前状态", "实际失效日", "录入人", "录入时间", "备注")
_C_STAT = {"生效中": "✔ 生效中", "已失效": "✖ 已失效", "未生效": "○ 未生效"}
_FILL_GREY = "D9D9D9"       # 已失效那几行打灰底


def _c_date(v):
    """单元格 → YYYY-MM-DD。Excel 里可能是日期型也可能是文本，两种都认。"""
    if v in (None, ""):
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    t = str(v).strip().replace("/", "-").replace(".", "-")
    t = re.sub(r"^(\d{4})-(\d{1,2})-(\d{1,2}).*$",
               lambda m: f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}", t)
    return t if _DATE_RE.match(t) else str(v).strip()


def _c_num(v):
    if v in (None, ""):
        return None
    try:
        return round(float(str(v).strip()), 4)
    except Exception:
        return "?"                                   # 填了但不是数字，要报错，不能当空


@router.get("/api/tempatt/contract/export")
def tempatt_contract_export(request: Request):
    """导出成 Excel。导出的这张**原样改完就能导回来**，列名和顺序都是导入认的。"""
    if not _require_perm(request, CAP_RATES):
        return JSONResponse({"ok": False, "msg": "无「维护合同价登记表」权限，请联系管理员"}, status_code=403)
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    rows = _contract_rows()
    wb = Workbook()
    ws = wb.active
    ws.title = "合同价登记表"
    ws.cell(1, 1, "临时工合同价登记表　·　成本会计维护　·　导出时间 "
                  + datetime.now().strftime("%Y-%m-%d %H:%M"))
    ws.cell(1, 1).font = Font(bold=True, size=12)
    head = list(_C_COLS) + list(_C_INFO)
    for c, v in enumerate(head, start=1):
        cell = ws.cell(2, c, v)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=_FILL_HEAD)
        # 表头两行字（指标在上、班次在下），不开自动换行就只显示一行
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 32
    grey = PatternFill("solid", fgColor=_FILL_GREY)
    for i, r in enumerate(rows):
        rr = 3 + i
        d, ni = r.get("day") or [None, None], r.get("night") or [None, None]
        st = r.get("当前状态") or ""
        for c, v in enumerate((r["派遣方"], r["岗位"], d[0], d[1],
                               f'=IF(COUNT(C{rr}:D{rr})=0,"",SUM(C{rr}:D{rr}))',
                               ni[0], ni[1],
                               f'=IF(COUNT(F{rr}:G{rr})=0,"",SUM(F{rr}:G{rr}))',
                               r["生效日"], r["失效日"] or "",
                               _C_STAT.get(st, st), r.get("实际失效日") or "至今",
                               r.get("录入人") or "", r.get("录入时间") or "",
                               r["备注"]), start=1):
            cell = ws.cell(rr, c, v)
            if st == "已失效":
                cell.fill = grey                        # 整行打灰，一眼看出这价已经不用了
            elif st == "未生效":
                cell.font = Font(italic=True, color="7F7F7F")
    ws.freeze_panes = "A3"
    for col, w in zip("ABCDEFGHIJKLMNO",
                      (14, 10, 11, 10, 10, 11, 10, 10, 13, 13, 12, 13, 12, 20, 26)):
        ws.column_dimensions[col].width = w

    w2 = wb.create_sheet("怎么改怎么导回来")
    for i, line in enumerate((
        "改完这张表，回页面「合同价（成本会计维护）」点「导入」就能传回去。",
        "",
        "【键】派遣方 ＋ 岗位 ＋ 生效日。同键＝覆盖那一行，新键＝新增一行。",
        "【不删行】文件里没有、系统里已有的行会原样保留。要删请在页面上删——"
        "否则拿一份旧文件一导，会把别人后来加的价悄悄清掉。",
        "【先预览】导入会先列出「新增哪几行、覆盖哪几行、旧值→新值」，确认了才真正写进去。",
        "",
        "【单价】白班、夜班至少填一档。某一班不做就整档留空（两个格都空）。",
        "        员工工资 ＋ 管理费 ＝ 结算表上的含管理费单价，「合计」两列是公式，改分项会自动跟着变。",
        "        **只改合计不改分项没用**——导入只看分项，合计那两列会被忽略。",
        "【生效日】必填，YYYY-MM-DD。",
        "【失效日】可以留空＝到同一行（同派遣方同岗位）下一条的生效日前一天；再没有下一条就是「至今」。",
        "        合同到期不再续、这家停止合作，才显式填——那种失效日是推不出来的。",
        "        显式填了可能出现区间重叠或空档，导入预览里会逐条标出来。",
        "",
        "【当前状态 / 实际失效日 / 录入人 / 录入时间】这四列是**算出来的或系统记的**，",
        "        导入时一律忽略，填了也不算数。",
        "        当前状态＝拿今天跟生效日、实际失效日比出来的：生效中 / 已失效（整行灰底）/ 未生效。",
        "        已失效的行**不要删**——历史期次要按当时的价核，删了那几个月就没价了。",
        "",
        "⚠ 这张表是复核工具里所有「钱对不对」的唯一基准。没登记的格子＝缺档，",
        "  工具会把那一档判成「待核」，不会拿别的价兜底。",
    )):
        w2.cell(i + 1, 1, line)
    w2.column_dimensions["A"].width = 100

    buf = BytesIO()
    wb.save(buf)
    fn = f"临时工合同价登记表_{datetime.now():%Y%m%d}.xlsx"
    return Response(buf.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename*=UTF-8\'\'{quote(fn)}"})


_HDR_JUNK = re.compile(r"[\s·・／/（）()\[\]【】、,，:：\-—_]+")


def _hdr_key(v):
    """表头单元格 → 认得出的列名。**按词判，不按整串比**。

    使用者会照自己习惯改标题：「白班·员工工资」改成「员工工资（换行）白班」、
    「归属」代替「派遣方」……认死名字的话，单价列一个都找不到，
    整表还会被报成「白班、夜班一档单价都没填」——比直接读不出文件更误导人。
    """
    t = _HDR_JUNK.sub("", str(v or "")).strip()
    if not t:
        return None
    if "实际失效" in t or "当前状态" in t or "录入" in t or "合计" in t:
        return None                                   # 这几列是算出来的，导入一律不看
    for shift in ("白班", "夜班"):
        if shift in t:
            if "管理" in t:
                return f"{shift}·管理费"
            if "员工" in t or "工资" in t:
                return f"{shift}·员工工资"
    if "派遣方" in t or "归属" in t or t == "供应商":
        return "派遣方"
    if "岗位" in t or "工种" in t:
        return "岗位"
    if "生效" in t:
        return "生效日"
    if "失效" in t or "到期" in t or "终止" in t:
        return "失效日"
    if "备注" in t or "说明" in t:
        return "备注"
    return None


def _parse_contract_book(data):
    """Excel → [行]，逐行带「问题」。列名认导出的那套，顺序随便，多几列少几列也不管。"""
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(data), data_only=True)
    ws = hdr = None
    for x in wb.worksheets:                          # 哪一页有「派遣方」和「生效日」就用哪一页
        for r in range(1, 9):
            keys = {_hdr_key(x.cell(r, c).value) for c in range(1, 26)}
            if "派遣方" in keys and "生效日" in keys:
                ws, hdr = x, r
                break
        if ws:
            break
    if ws is None:
        raise ValueError("这个文件里找不到「派遣方」和「生效日」两列表头，确认传的是合同价登记表？"
                         "（列名可以改，但得还看得出是这两样）")
    col = {}
    for c in range(1, 26):
        k = _hdr_key(ws.cell(hdr, c).value)
        if k:
            col.setdefault(k, c)
    g = lambda r, name: ws.cell(r, col[name]).value if name in col else None

    out = []
    for r in range(hdr + 1, ws.max_row + 1):
        agency = str(g(r, "派遣方") or "").strip()
        if not agency:
            continue
        post = str(g(r, "岗位") or "").strip() or ta.POST_DEFAULT
        eff, end = _c_date(g(r, "生效日")), _c_date(g(r, "失效日"))
        dw, dm = _c_num(g(r, "白班·员工工资")), _c_num(g(r, "白班·管理费"))
        nw, nm = _c_num(g(r, "夜班·员工工资")), _c_num(g(r, "夜班·管理费"))
        bad = []
        if not _DATE_RE.match(eff):
            bad.append(f"生效日「{eff or '空'}」不是 YYYY-MM-DD")
        if end and not _DATE_RE.match(end):
            bad.append(f"失效日「{end}」不是 YYYY-MM-DD")
        if end and _DATE_RE.match(end) and _DATE_RE.match(eff) and end < eff:
            bad.append("失效日早于生效日")
        if "?" in (dw, dm, nw, nm):
            bad.append("单价格子里有不是数字的内容")
        day = [dw or 0.0, dm or 0.0] if (dw is not None or dm is not None) else None
        night = [nw or 0.0, nm or 0.0] if (nw is not None or nm is not None) else None
        if day and "?" in day:
            day = None
        if night and "?" in night:
            night = None
        if not day and not night and not bad:
            bad.append("白班、夜班一档单价都没填")
        out.append({"行号": r, "派遣方": agency, "岗位": post, "生效日": eff,
                    "失效日": end or None, "day": day, "night": night,
                    "备注": str(g(r, "备注") or "").strip(), "问题": bad})
    if not out:
        raise ValueError("表里一行数据都没有")
    return out


def _contract_diff(items):
    """把导入的行和库里现有的比，算出新增/覆盖/不变/有问题。只算不写。"""
    cur = {}
    for x in (db.get_setting(_CONTRACT_KEY) or {}).get("行") or []:
        if isinstance(x, dict):
            cur[(str(x.get("派遣方") or "").strip(),
                 str(x.get("岗位") or ta.POST_DEFAULT).strip(),
                 str(x.get("生效日") or ""))] = x
    money = lambda b: "—" if not b else f"{b[0]:g}+{b[1]:g}={b[0] + b[1]:g}"
    seen, out = set(), []
    for it in items:
        key = (it["派遣方"], it["岗位"], it["生效日"])
        rec = {**it, "键": "|".join(key)}
        if it["问题"]:
            rec["动作"] = "✗ 有问题"
        elif key in seen:
            rec["动作"], rec["问题"] = "✗ 有问题", ["同一个键在文件里出现了两次"]
        else:
            old = cur.get(key)
            seen.add(key)
            if not old:
                rec["动作"] = "新增"
            else:
                o = (money(_band(old.get("day"))), money(_band(old.get("night"))),
                     str(old.get("失效日") or ""), str(old.get("备注") or ""))
                w = (money(it["day"]), money(it["night"]),
                     it["失效日"] or "", it["备注"])
                if o == w:
                    rec["动作"] = "不变"
                else:
                    rec["动作"] = "覆盖"
                    rec["旧"] = {"白班": o[0], "夜班": o[1], "失效日": o[2] or "（自动）", "备注": o[3]}
        rec["新"] = {"白班": money(it["day"]), "夜班": money(it["night"]),
                     "失效日": it["失效日"] or "（自动）", "备注": it["备注"]}
        out.append(rec)
    keep = [k for k in cur if k not in {(x["派遣方"], x["岗位"], x["生效日"]) for x in items}]
    return out, len(keep)


@router.post("/api/tempatt/contract/import")
async def tempatt_contract_import(request: Request):
    """只预览，不写。列出新增/覆盖/不变/有问题，以及导入后会不会出现区间重叠或空档。"""
    if not _require_perm(request, CAP_RATES):
        return JSONResponse({"ok": False, "msg": "无「维护合同价登记表」权限，请联系管理员"}, status_code=403)
    form = await request.form()
    uf = form.get("file")
    if uf is None or not hasattr(uf, "read"):
        return {"ok": False, "msg": "请选择要导入的 Excel 文件"}
    try:
        items = _parse_contract_book(await uf.read())
    except Exception as e:
        return {"ok": False, "msg": f"读不了这个文件：{e}"}
    rows, keep = _contract_diff(items)
    cnt = {k: sum(1 for x in rows if x["动作"] == k) for k in ("新增", "覆盖", "不变", "✗ 有问题")}
    return {"ok": True, "行": rows, "计数": cnt, "保留未提及": keep,
            "可写": cnt["新增"] + cnt["覆盖"] > 0 and cnt["✗ 有问题"] == 0}


@router.post("/api/tempatt/contract/import/apply")
async def tempatt_contract_import_apply(request: Request):
    """确认后真正写。有任何一行有问题就整批不写——半批写进去比不写更难收拾。"""
    u = _require_perm(request, CAP_RATES)
    if not u:
        return JSONResponse({"ok": False, "msg": "无「维护合同价登记表」权限，请联系管理员"}, status_code=403)
    form = await request.form()
    uf = form.get("file")
    if uf is None or not hasattr(uf, "read"):
        return {"ok": False, "msg": "请选择要导入的 Excel 文件"}
    try:
        items = _parse_contract_book(await uf.read())
    except Exception as e:
        return {"ok": False, "msg": f"读不了这个文件：{e}"}
    rows, keep = _contract_diff(items)
    bad = [x for x in rows if x["动作"] == "✗ 有问题"]
    if bad:
        return {"ok": False, "msg": f"有 {len(bad)} 行有问题，整批没写。先改好文件再导。",
                "行": rows}
    cur = db.get_setting(_CONTRACT_KEY) or {}
    raw = [x for x in (cur.get("行") or []) if isinstance(x, dict)]
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    wrote = 0
    for x in rows:
        if x["动作"] not in ("新增", "覆盖"):
            continue
        key = (x["派遣方"], x["岗位"], x["生效日"])
        raw = [y for y in raw if (str(y.get("派遣方") or "").strip(),
                                  str(y.get("岗位") or ta.POST_DEFAULT).strip(),
                                  str(y.get("生效日") or "")) != key]
        raw.append({"派遣方": x["派遣方"], "岗位": x["岗位"], "生效日": x["生效日"],
                    "失效日": x["失效日"], "day": x["day"], "night": x["night"],
                    "备注": x["备注"], "录入人": f'{u["name"]}（导入）', "录入时间": now})
        wrote += 1
    db.set_setting(_CONTRACT_KEY, {"行": raw}, u["name"])
    try:
        c = {k: sum(1 for x in rows if x["动作"] == k) for k in ("新增", "覆盖", "不变")}
        db.audit(u["name"], "临时工考勤-导入合同价",
                 f"新增{c['新增']} 覆盖{c['覆盖']} 不变{c['不变']}", f"未提及保留 {keep} 行")
    except Exception:
        pass
    out = _contract_rows()
    return {"ok": True, "写入": wrote, "行": out,
            "提醒": [f"{r['派遣方']}·{r['岗位']} {r['生效日']} 起：{'；'.join(r['问题'])}"
                     for r in out if r["问题"]] or None}


@router.post("/api/tempatt/contract/row/delete")
async def tempatt_contract_row_delete(request: Request):
    """删一行。删完该期可能落到更早的一行、或变成没有合同价——页面已做二次确认。"""
    u = _require_perm(request, CAP_RATES)
    if not u:
        return JSONResponse({"ok": False, "msg": "无「维护合同价登记表」权限，请联系管理员"}, status_code=403)
    b = await request.json()
    rid = str(b.get("id") or "")
    if str(b.get("confirm") or "") != rid:
        return {"ok": False, "msg": "删除需要二次确认"}
    cur = db.get_setting(_CONTRACT_KEY) or {}
    raw = [x for x in (cur.get("行") or []) if isinstance(x, dict)]
    left = [x for x in raw if _row_id({"派遣方": str(x.get("派遣方") or "").strip(),
                                       "岗位": str(x.get("岗位") or ta.POST_DEFAULT).strip(),
                                       "生效日": str(x.get("生效日") or "")}) != rid]
    if len(left) == len(raw):
        return {"ok": False, "msg": "没找到这一行"}
    db.set_setting(_CONTRACT_KEY, {"行": left}, u["name"])
    try:
        db.audit(u["name"], "临时工考勤-删除合同价行", rid, "")
    except Exception:
        pass
    return {"ok": True, "行": _contract_rows()}


@router.get("/api/tempatt/acks")
def tempatt_acks(request: Request):
    """认定清单：跨期回答「谁、什么时候、因为什么、放过了什么」。
    V2.346 起不再单独占一步——页面把「长期认定」折在结算风险卡里展开；本接口仍是它唯一的数据源。

    ⚠ 这一页存在的直接原因是一个设计缺口：**长期认定会每期压住提示，却没有任何地方能列出它们**。
    只有在「某期恰好又冒出同一条发现」时才会露面——万一那个人离职、或那条风险不再出现，
    这条认定就永远躺在库里、一直生效、谁也看不见。一个能永久压住告警又没有清单的开关，迟早出事。

    另：撤销认定＝把那条记录删掉，只有审计日志还留着。所以这里把审计流水一并摆出来，
    否则「谁曾经认定过、后来又撤了」这条线索是断的。"""
    if not _require_perm(request, CAP):
        return JSONResponse({"ok": False, "msg": "无「临时工考勤」权限，请联系管理员"}, status_code=403)
    now = datetime.now()

    def age_days(t):
        try:
            return (now - datetime.strptime(str(t)[:19], "%Y-%m-%d %H:%M:%S")).days
        except Exception:
            return None

    standing = []
    for k, v in (db.get_setting(_ACK_STANDING) or {}).items():
        if not isinstance(v, dict) or "|" not in k:
            continue
        t, _, key = k.partition("|")
        standing.append({"类型": t, "键": key, "范围": "长期", **v, "距今天数": age_days(v.get("时间"))})
    standing.sort(key=lambda x: x.get("时间") or "", reverse=True)

    periods, flow_keys = [], set()
    for mk in db.list_settings(prefix=_META_KEY):
        meta = db.get_setting(mk) or {}
        m = meta.get("月份")
        if not m:
            continue
        res, _ = _load_period(m)
        row = {"月份": m, "跑批时间": meta.get("跑批时间"), "跑批人": meta.get("跑批人"),
               "待认": None, "已认定": 0, "明细": []}
        cur = db.get_setting(_ACK_KEY + m) or {}
        for k, v in cur.items():
            if isinstance(v, dict) and "|" in k:
                t, _, key = k.partition("|")
                row["明细"].append({"类型": t, "键": key, "范围": "本期", **v,
                                    "距今天数": age_days(v.get("时间"))})
        if res:
            _apply_acks(res, m)
            st = res["stats"]
            open_n = 0
            for x in st.get("归属与打卡不符") or []:
                open_n += 0 if x.get("已认定") else 1
            for x in st.get("同名多行") or []:
                open_n += 0 if x.get("已认定") else 1
            for x in st.get("打卡表重名") or []:
                open_n += 0 if (isinstance(x, dict) and x.get("已认定")) else 1
            for x in st.get("金额核对") or []:
                open_n += 0 if x.get("已认定") else 1
            for x in st.get("合同外调整") or []:
                if x.get("级别") == "异常":
                    open_n += 0 if x.get("已认定") else 1
            for r in res.get("rows") or []:
                if r.get("档") in ("over_out", "hard", "thin"):
                    open_n += 0 if r.get("已认定") else 1
            row["待认"] = open_n
            row["已认定"] = (st.get("认定") or {}).get("已认定") or 0
        periods.append(row)
        flow_keys.add(m)
    periods.sort(key=lambda x: x["月份"], reverse=True)

    flow = []
    try:
        for a in db.recent_audit(400):
            act = str(a.get("action") or "")
            if act.startswith("临时工考勤-") and ("认定" in act):
                flow.append({"时间": a.get("ts"), "操作人": a.get("operator"), "动作": act.split("-", 1)[-1],
                             "对象": a.get("target"), "说明": a.get("detail") or ""})
    except Exception:
        pass

    return {"ok": True, "长期": standing, "按期": periods, "流水": flow[:200],
            "复核提示天数": 180,
            "说明": "认定不是删除：条目仍在，只是不再计入异常。长期认定会每期生效，"
                    "情况变了记得回来撤销——本页是唯一能看全它们的地方。"}


@router.post("/api/tempatt/ack")
async def tempatt_ack(request: Request):
    """认定某条可疑项无误。必须填理由——「为什么不是问题」才是这条记录的价值所在。"""
    u = _require_perm(request, CAP)
    if not u:
        return JSONResponse({"ok": False, "msg": "无「临时工考勤」权限，请联系管理员"}, status_code=403)
    b = await request.json()
    month, t = str(b.get("month") or ""), str(b.get("类型") or "")
    key, why = str(b.get("键") or ""), str(b.get("理由") or "").strip()
    scope = str(b.get("范围") or "本期")
    if t not in _ACK_TYPES or not key:
        return {"ok": False, "msg": "认定对象不对"}
    if not why:
        return {"ok": False, "msg": "请填一句理由——将来翻这份底稿的人要知道当时为什么认为它不是问题"}
    if scope == "长期" and t not in _ACK_LONG_OK:
        return {"ok": False, "msg": f"「{t}」带日期，只能按本期认定；长期认定仅适用于{ '、'.join(_ACK_LONG_OK) }"}
    if scope != "长期" and not _MONTH_RE.match(month):
        return {"ok": False, "msg": "本期认定需要月份"}
    store = _ACK_STANDING if scope == "长期" else _ACK_KEY + month
    cur = db.get_setting(store) or {}
    if not isinstance(cur, dict):
        cur = {}
    cur[f"{t}|{key}"] = {"理由": why, "认定人": u["name"],
                         "时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                         "月份": month}
    db.set_setting(store, cur, u["name"])
    try:
        db.audit(u["name"], "临时工考勤-认定无误", f"{month} {t} {key}", f"{scope}｜{why}")
    except Exception:
        pass
    return {"ok": True, "范围": scope}


@router.post("/api/tempatt/ack/batch")
async def tempatt_ack_batch(request: Request):
    """一次认定 / 撤销多条同类型的可疑项。

    为什么要有批量：合同外调整这类东西是**成批出现**的——2026-06 十四笔里十三笔是蒸练补贴，
    同一张审批单管全部。逐条点十四次、逐条抄同一句理由，人只会越抄越敷衍。
    理由照样必填，且**一句理由对应这一批**，写进每一条记录里，事后翻得出是哪次、批了哪些。
    """
    u = _require_perm(request, CAP)
    if not u:
        return JSONResponse({"ok": False, "msg": "无「临时工考勤」权限，请联系管理员"}, status_code=403)
    b = await request.json()
    month, t = str(b.get("month") or ""), str(b.get("类型") or "")
    act = str(b.get("动作") or "认定")
    keys = [str(k) for k in (b.get("键") or []) if str(k).strip()]
    why = str(b.get("理由") or "").strip()
    if t not in _ACK_TYPES:
        return {"ok": False, "msg": "认定类型不对"}
    if not keys:
        return {"ok": False, "msg": "一条都没选"}
    if len(keys) > 500:
        return {"ok": False, "msg": f"一次最多 500 条，这次选了 {len(keys)} 条"}
    if not _MONTH_RE.match(month):
        return {"ok": False, "msg": "本期认定需要月份"}
    store = _ACK_KEY + month
    cur = db.get_setting(store) or {}
    if not isinstance(cur, dict):
        cur = {}
    if act == "撤销":
        # 撤销只动本期的。长期认定是跨期的东西，不该被某一期的批量操作顺手抹掉——
        # 要撤长期认定，去第④步结算风险页那份长期清单里单条撤。
        hit = [k for k in keys if f"{t}|{k}" in cur]
        for k in hit:
            cur.pop(f"{t}|{k}")
        db.set_setting(store, cur, u["name"])
        try:
            db.audit(u["name"], "临时工考勤-批量撤销认定", f"{month} {t} {len(hit)} 条", "；".join(hit[:20]))
        except Exception:
            pass
        return {"ok": True, "条数": len(hit),
                "msg": f"撤销了 {len(hit)} 条" + (f"，另有 {len(keys) - len(hit)} 条本就没有本期认定（可能是长期认定，请到第④步单条撤）"
                                                if len(hit) < len(keys) else "")}
    if not why:
        return {"ok": False, "msg": "请填一句理由——将来翻这份底稿的人要知道当时为什么认为它们不是问题"}
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for k in keys:
        cur[f"{t}|{k}"] = {"理由": why, "认定人": u["name"], "时间": now,
                           "月份": month, "批量": len(keys)}
    db.set_setting(store, cur, u["name"])
    try:
        db.audit(u["name"], "临时工考勤-批量认定无误", f"{month} {t} {len(keys)} 条",
                 f"{why}｜{'；'.join(keys[:20])}")
    except Exception:
        pass
    return {"ok": True, "条数": len(keys)}


@router.post("/api/tempatt/ack/undo")
async def tempatt_ack_undo(request: Request):
    """撤销认定——认错了要能改回来，否则没人敢点第一下。"""
    u = _require_perm(request, CAP)
    if not u:
        return JSONResponse({"ok": False, "msg": "无「临时工考勤」权限，请联系管理员"}, status_code=403)
    b = await request.json()
    month, t, key = str(b.get("month") or ""), str(b.get("类型") or ""), str(b.get("键") or "")
    hit = False
    for store in (_ACK_STANDING, _ACK_KEY + month):
        cur = db.get_setting(store) or {}
        if isinstance(cur, dict) and f"{t}|{key}" in cur:
            cur.pop(f"{t}|{key}")
            db.set_setting(store, cur, u["name"])
            hit = True
    if not hit:
        return {"ok": False, "msg": "没找到这条认定"}
    try:
        db.audit(u["name"], "临时工考勤-撤销认定", f"{month} {t} {key}", "")
    except Exception:
        pass
    return {"ok": True}


# ==================== 复核结论·确认无误 ====================
# 用工成本汇总是要发给经理、往上报的数。第⑨步在确认之前不开——
# 免得复核还没做完，成本表已经被复制走了（使用者 2026-08-29 提的）。
_SIGNOFF = "tempatt_signoff"


@router.get("/api/tempatt/signoff")
def tempatt_signoff_get(request: Request):
    u = _require_perm(request, CAP)
    if not u:
        return JSONResponse({"ok": False, "msg": "无「临时工考勤」权限，请联系管理员"}, status_code=403)
    month = str(request.query_params.get("month") or "")
    cur = db.get_setting(_SIGNOFF) or {}
    rec = cur.get(month) if isinstance(cur, dict) else None
    return {"ok": True, "记录": rec if isinstance(rec, dict) else None}


@router.post("/api/tempatt/signoff")
async def tempatt_signoff_set(request: Request):
    """确认本期复核结论无误。同时把当时那一版的关键数字（人数/请款合计/按合同价应付/结论）
    存成「指纹」——之后若重跑出不一样的数，页面据此提示「结果变了，要重新确认」。
    不存指纹的话，确认就成了一个跟数据脱钩的戳，重跑改了金额它照样显示已确认。"""
    u = _require_perm(request, CAP)
    if not u:
        return JSONResponse({"ok": False, "msg": "无「临时工考勤」权限，请联系管理员"}, status_code=403)
    b = await request.json()
    month = str(b.get("month") or "")
    if not _MONTH_RE.match(month):
        return {"ok": False, "msg": "需要月份"}
    rec = {"确认人": u["name"], "时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
           "指纹": str(b.get("指纹") or ""), "摘要": str(b.get("摘要") or ""),
           "备注": str(b.get("备注") or "").strip()}
    cur = db.get_setting(_SIGNOFF) or {}
    if not isinstance(cur, dict):
        cur = {}
    cur[month] = rec
    db.set_setting(_SIGNOFF, cur, u["name"])
    try:
        db.audit(u["name"], "临时工考勤-确认复核结论", month, rec["摘要"])
    except Exception:
        pass
    return {"ok": True, "记录": rec}


@router.post("/api/tempatt/signoff/undo")
async def tempatt_signoff_undo(request: Request):
    """撤销确认。确认是给人签的字，签错了要能撤，否则没人敢签第一次。"""
    u = _require_perm(request, CAP)
    if not u:
        return JSONResponse({"ok": False, "msg": "无「临时工考勤」权限，请联系管理员"}, status_code=403)
    b = await request.json()
    month = str(b.get("month") or "")
    cur = db.get_setting(_SIGNOFF) or {}
    if not isinstance(cur, dict) or month not in cur:
        return {"ok": False, "msg": "本期还没有确认记录"}
    cur.pop(month)
    db.set_setting(_SIGNOFF, cur, u["name"])
    try:
        db.audit(u["name"], "临时工考勤-撤销确认", month, "")
    except Exception:
        pass
    return {"ok": True}


# ==================== 合同外调整·本期奖惩已核对 ====================
# 跟⑧复核结论的确认是两回事：那个签的是「工时/应付这张结论表」，这个签的是
# 「本期奖/罚/补贴我看过了」——尤其**本期一笔都没有**时，得让人正向签一句「确认无奖惩」，
# 而不是只有工具说「没找到」（使用者 2026-08-29 提的）。逐笔认定管的是「某一笔像有问题但没事」，
# 这个管的是「整期奖惩已核对完毕」，两者不冲突。
_ADJSIGN = "tempatt_adjsign"


@router.get("/api/tempatt/adjsign")
def tempatt_adjsign_get(request: Request):
    u = _require_perm(request, CAP)
    if not u:
        return JSONResponse({"ok": False, "msg": "无「临时工考勤」权限，请联系管理员"}, status_code=403)
    month = str(request.query_params.get("month") or "")
    cur = db.get_setting(_ADJSIGN) or {}
    rec = cur.get(month) if isinstance(cur, dict) else None
    return {"ok": True, "记录": rec if isinstance(rec, dict) else None}


@router.post("/api/tempatt/adjsign")
async def tempatt_adjsign_set(request: Request):
    """确认本期奖惩已核对。存当时那版的指纹（笔数/净额/异常/存疑），奖惩一变这枚确认自动失效。"""
    u = _require_perm(request, CAP)
    if not u:
        return JSONResponse({"ok": False, "msg": "无「临时工考勤」权限，请联系管理员"}, status_code=403)
    b = await request.json()
    month = str(b.get("month") or "")
    if not _MONTH_RE.match(month):
        return {"ok": False, "msg": "需要月份"}
    rec = {"确认人": u["name"], "时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
           "指纹": str(b.get("指纹") or ""), "摘要": str(b.get("摘要") or ""),
           "备注": str(b.get("备注") or "").strip()}
    cur = db.get_setting(_ADJSIGN) or {}
    if not isinstance(cur, dict):
        cur = {}
    cur[month] = rec
    db.set_setting(_ADJSIGN, cur, u["name"])
    try:
        db.audit(u["name"], "临时工考勤-确认奖惩已核对", month, rec["摘要"])
    except Exception:
        pass
    return {"ok": True, "记录": rec}


@router.post("/api/tempatt/adjsign/undo")
async def tempatt_adjsign_undo(request: Request):
    u = _require_perm(request, CAP)
    if not u:
        return JSONResponse({"ok": False, "msg": "无「临时工考勤」权限，请联系管理员"}, status_code=403)
    b = await request.json()
    month = str(b.get("month") or "")
    cur = db.get_setting(_ADJSIGN) or {}
    if not isinstance(cur, dict) or month not in cur:
        return {"ok": False, "msg": "本期还没有奖惩确认记录"}
    cur.pop(month)
    db.set_setting(_ADJSIGN, cur, u["name"])
    try:
        db.audit(u["name"], "临时工考勤-撤销奖惩确认", month, "")
    except Exception:
        pass
    return {"ok": True}


# ==================== 上传与解析 ====================
async def _read_two(request: Request):
    """multipart 取两份表：summary=人力上报汇总表，punch=考勤系统打卡时刻表。"""
    form = await request.form()
    out, names = {}, {}
    for k in ("summary", "punch"):
        uf = form.get(k)
        if uf is not None and hasattr(uf, "read"):
            out[k] = await uf.read()
            names[k] = getattr(uf, "filename", "") or ""
        else:
            out[k] = None
    params, rates = {}, None
    for key in ("params", "rates"):
        raw = form.get(key)
        if not raw:
            continue
        try:
            v = json.loads(raw if isinstance(raw, str) else await raw.read())
        except Exception:
            continue
        if key == "params":
            params = v
        else:
            rates = v
    return (out.get("summary"), out.get("punch"), params, rates,
            str(form.get("month") or ""), names)


_RULE_SRC = ("人力 2026-08-10 答复：下班打卡−上班打卡，半小时向下取整，白班扣1小时/夜班扣0.5小时；"
             "多记每天 0.5 小时以内视为正常波动。")


def _run(summary_bytes, punch_bytes, params, rates=None, month=""):
    """rates 参数只为兼容旧调用保留，**不再生效**：单价唯一来源是合同价登记表（按期挑行）。
    页面早就不发它了；真有人发过来也当没看见，否则又绕开了生效期机制。"""
    sm = ta.parse_summary(summary_bytes)
    pk = ta.parse_punch(punch_bytes)
    # 月份优先用前端选的；没选就用汇总表标题里的，再不行用打卡表统计区间的
    month = month or sm.get("period") or pk.get("period") or ""
    contract, info = _contract_with_info(month)
    res = ta.compute(sm, pk, params, contract=contract)
    res["month"] = month
    # 期中调价（本期内换过价，按覆盖期末那行核）和登记表里的重叠/空档行，页面与报告都要讲清楚
    res["rates"]["期中调价"] = info.get("期中调价") or []
    res["rates"]["问题行"] = info.get("问题行") or []
    res["outsiders"] = ta.outsiders(sm, pk)
    res["source"] = {"汇总表页签": sm["sheet"], "可选页签": sm["sheets"], "计价规则": sm.get("rules", "")}
    return res


# ==================== 从钉钉取打卡 ====================
# 取一个月要一万次调用、约 6 分钟，撑不住一个同步请求，所以走后台任务 + 轮询。
# 产出是一张**与人力导出格式一致的打卡表**，不是直接塞进内核——
# 取数结果必须能人眼核对，这是这条线一开始就定的规矩。
_DING_JOBS = {}                     # {id: {状态, 进度, 说明, xlsx, 结果, 错}}
_DING_LOCK = threading.Lock()


def _ding_gc():
    """只留最近 6 个任务，且完成超过 2 小时的丢掉——打卡表是人事敏感数据，别在内存里堆着。"""
    now = time.time()
    with _DING_LOCK:
        for k in [k for k, v in _DING_JOBS.items()
                  if v.get("完成于") and now - v["完成于"] > 7200]:
            _DING_JOBS.pop(k, None)
        for k in sorted(_DING_JOBS, key=lambda x: _DING_JOBS[x]["起于"])[:-6]:
            _DING_JOBS.pop(k, None)


@router.get("/api/tempatt/ding/status")
def tempatt_ding_status(request: Request, month: str = ""):
    """钉钉这条路通不通、这一期还取不取得到。页面据此决定按钮是亮是灰。"""
    if not _require_perm(request, CAP):
        return JSONResponse({"ok": False, "msg": "无「临时工考勤」权限，请联系管理员"}, status_code=403)
    if dda is None:
        return {"ok": False, "可用": False, "说明": "后端没装钉钉取数模块"}
    try:
        pr = dda.probe()
    except Exception as e:
        return {"ok": False, "可用": False, "说明": f"体检失败：{e}"}
    out = {"ok": True, "可用": bool(pr.get("ok")), "配置": pr.get("配置"),
           "项": pr.get("项") or [], "说明": pr.get("说明") or ""}
    try:
        out["缓存"] = dda.cache_info()      # 离职花名册缓存了多少，页面上讲清楚为什么第一次慢
    except Exception:
        pass
    if month:
        good, why = dda.month_reachable(month)
        out["期次可取"], out["期次说明"] = good, why
        out["可用"] = out["可用"] and good
    return out


def _ding_worker(jid, month, names, worked, force=False, scope="worked"):
    def say(msg, pct=0):
        with _DING_LOCK:
            j = _DING_JOBS.get(jid)
            if j:
                j["说明"], j["进度"] = msg, pct
    try:
        r = dda.pull_month(month, names, progress=say, worked_days=worked,
                           force_roster=force, scope=scope)
        with _DING_LOCK:
            j = _DING_JOBS.get(jid) or {}
            j.update(状态="完成", 进度=100, xlsx=r.pop("xlsx"), 结果=r, 完成于=time.time())
    except Exception as e:
        with _DING_LOCK:
            j = _DING_JOBS.get(jid) or {}
            j.update(状态="失败", 错=str(e), 完成于=time.time())


@router.post("/api/tempatt/ding/pull")
async def tempatt_ding_pull(request: Request):
    """传汇总表 + 期次 → 后台去钉钉取这些人的打卡 → 生成打卡表。返回任务号，前端轮询。"""
    u = _require_perm(request, CAP)
    if not u:
        return JSONResponse({"ok": False, "msg": "无「临时工考勤」权限，请联系管理员"}, status_code=403)
    if dda is None:
        return {"ok": False, "msg": "后端没装钉钉取数模块"}
    form = await request.form()
    uf = form.get("summary")
    if uf is None or not hasattr(uf, "read"):
        return {"ok": False, "msg": "请先上传「人力上报汇总表」——要照着它上面的人去钉钉取"}
    month = str(form.get("month") or "")
    try:
        sm = ta.parse_summary(await uf.read())
    except Exception as e:
        return {"ok": False, "msg": f"汇总表解析失败：{e}"}
    # ⚠ 取数要花五六分钟，出发前必须先拦住「汇总表和期间对不上」这一种——
    #   否则拿 6 月的人名去取 7 月的打卡，跑完得到一份谁也对不上的表，白等一场。
    #   /review 那道拦截是在提交之后才生效，这里得提前一步。
    found = sm.get("period") or ""
    if month and found and month != found:
        return {"ok": False,
                "msg": f"右上角选的是 {month}，但这份汇总表识别为 {found}。"
                       f"取数会照汇总表上的人去拿 {month} 的打卡，两边对不上，跑完也没法用。"
                       f"请把右上角切到 {found}，或换一份 {month} 的汇总表。"}
    month = month or found
    good, why = dda.month_reachable(month)
    if not good:
        return {"ok": False, "msg": why}

    names, worked = set(), {}
    for x in sm["people"]:
        names.add(x["name"])
        worked.setdefault(x["name"], set()).update(
            d for d, v in (x.get("days") or {}).items() if v)
    _ding_gc()
    jid = f"{month}-{int(time.time() * 1000)}"
    with _DING_LOCK:
        _DING_JOBS[jid] = {"状态": "进行中", "进度": 0, "说明": "排队中…",
                           "月份": month, "起于": time.time(), "谁": u["name"]}
    threading.Thread(target=_ding_worker,
                     args=(jid, month, sorted(names), worked,
                           str(form.get("refresh") or "") in ("1", "true"),
                           "full" if str(form.get("scope") or "") == "full" else "worked"),
                     daemon=True).start()
    db.audit(u["name"], "临时工考勤-钉钉取数", f"{month} {len(names)}人", "发起")
    return {"ok": True, "任务": jid, "月份": month, "人数": len(names),
            "取数范围": "full" if str(form.get("scope") or "") == "full" else "worked",
            "月份来源": ("右上角所选期间" if str(form.get("month") or "") else "汇总表标题识别"),
            "预计": f"约 {max(1, round(len(names) * 31 / 1800))} 分钟", "说明": why}


@router.get("/api/tempatt/ding/job")
def tempatt_ding_job(request: Request, id: str = ""):
    """查进度。完成后 xlsx 不走这个口，另开下载口，免得把几百 KB 塞进轮询响应。"""
    if not _require_perm(request, CAP):
        return JSONResponse({"ok": False, "msg": "无「临时工考勤」权限，请联系管理员"}, status_code=403)
    with _DING_LOCK:
        j = _DING_JOBS.get(id)
        if not j:
            return {"ok": False, "msg": "任务不存在或已过期，请重新取数"}
        out = {k: v for k, v in j.items() if k not in ("xlsx",)}
    out["ok"] = True
    return out


@router.get("/api/tempatt/ding/file")
def tempatt_ding_file(request: Request, id: str = ""):
    """下载钉钉取数生成的打卡表——先让人眼核对，再拿去复核。"""
    if not _require_perm(request, CAP):
        return JSONResponse({"ok": False, "msg": "无「临时工考勤」权限，请联系管理员"}, status_code=403)
    with _DING_LOCK:
        j = _DING_JOBS.get(id)
        blob = (j or {}).get("xlsx")
    if not blob:
        return JSONResponse({"ok": False, "msg": "任务不存在、未完成或已过期"}, status_code=404)
    fn = f"{(j.get('月份') or '')}钉钉打卡记录.xlsx".encode("utf-8").decode("latin-1")
    return Response(blob, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(fn)}"})


@router.get("/api/tempatt/params")
def tempatt_params(request: Request):
    """默认口径参数 + 计价规则表。前端据此渲染参数区，改完随请求回传。"""
    if not _require_perm(request, CAP):
        return JSONResponse({"ok": False, "msg": "无「临时工考勤」权限，请联系管理员"}, status_code=403)
    return {"ok": True, "params": ta.DEFAULT_PARAMS,
            "单价说明": "单价按「派遣方 × 岗位 × 白班/夜班」分档，唯一来源是第③步「合同价（成本会计维护）」登记表，"
                        "按行带生效期。岗位列为空＝普工；保洁这类单独定价的岗位挂在各自派遣方名下"
                        "（规则原文就是「锦绣保洁」），不会跨派遣方套用。没登记的档一律「缺档／待核」，不拿任何表顶上。",
            "口径来源": _RULE_SRC}


@router.post("/api/tempatt/review")
async def tempatt_review(request: Request):
    """上传两张表 → 逐日重算比对 → 四档判定 → 按期留档。对金蝶只读、不写、不产生凭证。"""
    u = _require_perm(request, CAP)
    if not u:
        return JSONResponse({"ok": False, "msg": "无「临时工考勤」权限，请联系管理员"}, status_code=403)
    summary, punch, params, rates, month, names = await _read_two(request)
    if not summary or not punch:
        return {"ok": False, "msg": "请同时上传「人力上报汇总表」和「打卡时刻表」两个文件"}
    try:
        res = _run(summary, punch, params, rates, month)
    except Exception as e:
        return {"ok": False, "msg": f"解析失败：{e}"}
    # 页面传来的月份优先于表标题——那就必须和表标题对得上。否则「看完 7 月留档去传 8 月的表」
    # 会按 7 月合同价算 8 月的数，还把结果写进 7 月留档把它覆盖掉（review 实测的坑，V2.347 拦住）
    found = res["stats"].get("汇总表月份") or res["stats"].get("打卡表月份") or ""
    if month and found and month != found:
        return {"ok": False,
                "msg": f"右上角选的是 {month}，但上传的表识别为 {found}。请把期间切到 {found}（或清掉再传），"
                       f"以免按 {month} 的合同价算、又把 {month} 的留档覆盖掉。"}
    res["留档"] = _save_period(res.get("month") or month, summary, punch, names, res, u)
    # 认定信息**不进留档快照**——它独立存、独立撤，重跑之后照样生效
    _apply_acks(res, res.get("month") or month)
    _attach_cost(res, res.get("month") or month)      # 用工成本汇总：页面第⑨步直接复制
    try:
        s = res["stats"]
        db.audit(u["name"], "临时工考勤-复核", f"{res.get('month','')} {s['人数']}人/{s['比对人日']}人日",
                 f"少记{s['少记小时']}h 异常多记{s['异常多记日次']}日次")
    except Exception:
        pass
    return {"ok": True, **res}


# ==================== 历史期次：选月份直接看，不必重传 ====================
@router.get("/api/tempatt/periods")
def tempatt_periods(request: Request):
    """已核过的期次清单（倒序）。「可重跑」＝原表还在留存期内，过期的只能看结论。"""
    if not _require_perm(request, CAP):
        return JSONResponse({"ok": False, "msg": "无「临时工考勤」权限，请联系管理员"}, status_code=403)
    out = []
    try:
        for k in db.list_settings(prefix=_META_KEY):
            m = db.get_setting(k) or {}
            if not m:
                continue
            m["可重跑"] = bool(_period_dir(m.get("月份", "")) and
                             os.path.isfile(os.path.join(_period_dir(m["月份"]), "summary.xlsx")))
            m["原表在库"] = m["可重跑"]
            out.append(m)
    except Exception:
        pass
    out.sort(key=lambda x: x.get("月份", ""), reverse=True)
    return {"ok": True, "periods": out, "保留期数": _keep_months(),
            "可删": bool(_require_perm(request, CAP_RATES)),   # 没这个权限的人不显示删除按钮
            "说明": "结论长期保留；原始两张表只留最近 %d 期，过期后仍可看结论，但改参数重跑需重新上传。"
                    % _keep_months()}


@router.get("/api/tempatt/period")
def tempatt_period(request: Request, month: str = ""):
    """读某期留档的完整结论——页面选月份即可直接展示，不必重新上传。"""
    if not _require_perm(request, CAP):
        return JSONResponse({"ok": False, "msg": "无「临时工考勤」权限，请联系管理员"}, status_code=403)
    res, meta = _load_period(month)
    if not res:
        return {"ok": False, "msg": f"{month} 没有留档记录，请上传两张表跑一次"}
    d = _period_dir(month)
    _apply_acks(res, month)
    _mark_legacy(res)
    _attach_cost(res, month)
    return {"ok": True, "读自留档": True, "留档信息": meta,
            "可重跑": bool(d and os.path.isfile(os.path.join(d, "summary.xlsx"))), **res}


@router.get("/api/tempatt/board")
def tempatt_board(request: Request, fresh: int = 0):
    """临工看板的数据源 —— **历次复核留档，不是另外上传一张结构表**（2026-08-22 定案）。

    缓存：按「期次签名」（各期月份＋跑批时间）存一份算好的结果。签名一变（新核了一期、
    改参数重跑、删了一期）自动重算，所以不需要任何人手动刷新；`fresh=1` 可强制重算。
    不做定时失效——数据只在有人跑复核时才变，按时间过期只会白算。"""
    if not _require_perm(request, CAP_BOARD):
        return JSONResponse({"ok": False, "msg": "无「临时工看板」权限，请联系管理员"}, status_code=403)
    metas = []
    try:
        for k in db.list_settings(prefix=_META_KEY):
            m = db.get_setting(k) or {}
            if m.get("月份"):
                metas.append(m)
    except Exception:
        pass
    metas.sort(key=lambda x: x["月份"])
    sig = "|".join(f"{m['月份']}@{m.get('跑批时间','')}" for m in metas)
    if not metas:
        return {"ok": True, "空": True, "期数": 0,
                "msg": "还没有任何已核期次。到「复核工具」跑一期，看板就会自动出来——本看板取的是历次复核的结果，不需要另外上传结构表。"}
    cached = db.get_setting(_BOARD_KEY) or {}
    if not fresh and cached.get("sig") == sig and cached.get("data"):
        d = _unpack(cached["data"])
        if d:
            return {"ok": True, "命中缓存": True, **d}
    data = []
    for m in metas:
        res, _ = _load_period(m["月份"])
        if res:
            data.append((m["月份"], res))
    if not data:
        return {"ok": True, "空": True, "期数": 0, "msg": "留档的结论读不出来，请重新跑一期"}
    out = ta.board_from_periods(data)
    try:
        db.set_setting(_BOARD_KEY, {"sig": sig, "data": _pack(out)}, "系统·看板缓存")
    except Exception:
        pass
    return {"ok": True, "命中缓存": False, **out}


@router.post("/api/tempatt/period/delete")
async def tempatt_period_delete(request: Request):
    """删掉一期留档（结论 + 原表）。调试期难免留下几期试跑的数，得能清掉。

    ⚠ 不可逆，所以：
      · 走 CAP_RATES（维护合同价登记表那个权限），不是人人能删——与「改共享数据」同一档
        （本该单开一个 tempatt_del 权限点，但 db.py 眼下有平台通用线的未提交改动，
         不去动它；等那条线并回来再拆，见变更记录）
      · 必须带 confirm=期间原文，防止误点：前端已做两步确认，这里再挡一道
      · 留审计：谁、什么时候、删了哪一期、那一期原本是什么结论"""
    u = _require_perm(request, CAP_RATES)
    if not u:
        return JSONResponse({"ok": False, "msg": "无「维护合同价登记表」权限，不能删留档，请联系管理员"},
                            status_code=403)
    body = await request.json()
    month = str(body.get("month") or "")
    if not _MONTH_RE.match(month):
        return {"ok": False, "msg": "月份格式应为 YYYY-MM"}
    if str(body.get("confirm") or "") != month:
        return {"ok": False, "msg": "删除需要二次确认"}
    meta = db.get_setting(_META_KEY + month) or {}
    d = _period_dir(month)
    # ⚠ _period_dir 只要月份格式合法就返回路径，不代表目录存在——这里必须用 isdir 判，
    #   否则「删一个本来就不存在的期次」会回一句「已删」，读起来像真删掉了什么
    if not meta and not (d and os.path.isdir(d)):
        return {"ok": False, "msg": f"{month} 没有留档记录，无需删除"}
    if d and os.path.isdir(d):
        try:
            shutil.rmtree(d)
        except Exception as e:
            return {"ok": False, "msg": f"原表删除失败：{e}"}
    # db 里没有删 setting 的辅助函数（不去动 db.py，见上），置空即可：
    # get_setting 拿到 None，期次列表里 `if not m: continue` 会跳过，等同于删掉。
    try:
        db.set_setting(_RESULT_KEY + month, None, u["name"])
        db.set_setting(_META_KEY + month, None, u["name"])
    except Exception as e:
        return {"ok": False, "msg": f"结论删除失败：{e}"}
    try:
        db.audit(u["name"], "临时工考勤-删除留档", month,
                 f"原{meta.get('人数','?')}人/应付{meta.get('应付合计','?')}/"
                 f"{meta.get('跑批时间','')}由{meta.get('跑批人','')}跑批")
    except Exception:
        pass
    return {"ok": True, "month": month, "已删": True}


@router.post("/api/tempatt/rerun")
async def tempatt_rerun(request: Request):
    """拿留档的原表按新参数重跑。过了留存期就没有原表可跑，明确报出来而不是静默失败。"""
    u = _require_perm(request, CAP)
    if not u:
        return JSONResponse({"ok": False, "msg": "无「临时工考勤」权限，请联系管理员"}, status_code=403)
    body = await request.json()
    month = str(body.get("month") or "")
    summary, punch = _period_files(month)
    if not summary:
        return {"ok": False, "msg": f"{month} 的原始两张表已过留存期（只保留最近 {_keep_months()} 期），"
                                    f"改参数重跑请重新上传"}
    try:
        res = _run(summary, punch, body.get("params") or {}, body.get("rates"), month)
    except Exception as e:
        return {"ok": False, "msg": f"解析失败：{e}"}
    meta = db.get_setting(_META_KEY + month) or {}
    res["留档"] = _save_period(month, summary, punch, meta.get("原表文件名") or {}, res, u)
    _apply_acks(res, month)
    _attach_cost(res, month)
    try:
        db.audit(u["name"], "临时工考勤-改参数重跑", month, "")
    except Exception:
        pass
    return {"ok": True, **res}


# ==================== 导出（带公式、异常行标底色）====================
# 需求方在 RPA 调研表里反复写了两件事：「异常行单元格标黄底颜色」「最终输出的表格需保留单元格公式」。
# 所以导出不是把屏幕上的数贴成死值——跨度/重算/差异/金额/判定全部写成公式，并引用「⑤口径参数」页；
# 改参数或改打卡时刻，整表连同底色一起重算，需求方拿去能继续往下核。
_F_HEAD = {"bold": True, "color": "FFFFFF"}
_FILL_HEAD = "4472C4"
_FILL_RED = "FFC7CE"     # ⚠ 异常（含超弹性多记、硬伤）
_FILL_YEL = "FFFF00"     # △ 少记
_FILL_BLU = "DDEBF7"     # ○ 弹性内多记
_FILL_GRN = "E2EFDA"     # 合计行
_FILL_PAY = "FFF2CC"     # 「结算表应付＝请款金额」那一组——真要付出去的钱，与右侧校验列区分开


def _verdict_fill(ws, r, keys, verdict, dark=False):
    """小计/合计行的「结论」格按三态上色——整行刷成小计底色后，异常/待核的小计看不出来。"""
    from openpyxl.styles import PatternFill, Font
    if "结论" not in keys or not verdict:
        return
    i = keys.index("结论") + 1
    c = ws.cell(r, i)
    c.fill = PatternFill("solid", fgColor=(_FILL_RED if verdict == "异常" else _FILL_YEL if verdict == "待核" else _FILL_GRN))
    if dark:
        c.font = Font(name="Arial", size=10, bold=True)


def _xlsx(res, month_label="", punch_bytes=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import FormulaRule

    p = res["params"]
    st = res["stats"]
    _BASIS = "punch" if str(p.get("report_basis")) == "punch" else "shift"
    thin = Side(style="thin", color="BFBFBF")
    BD = Border(left=thin, right=thin, top=thin, bottom=thin)
    CTR = Alignment(horizontal="center", vertical="center")
    # 表头必须开 wrap_text：不开的话「归属\n（派遣方）」这种带换行的表头会被当成一行，
    # 超出列宽就直接截断（实测②页表头显示成「属（派遣方」「上报白班工」）。
    CTRW = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
    LFTN = Alignment(horizontal="left", vertical="center", wrap_text=False)  # 左对齐但**不换行**
    FB = Font(name="Arial", size=10)
    FBB = Font(name="Arial", size=10, bold=True)
    FH = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    FT = Font(name="Arial", size=14, bold=True)

    _setw = {}                      # {(页签, 列字母): 已设过的宽度}

    def head(ws, row, cols, widths):
        """画一行表头。

        ⚠ 列宽**只加不减**：一页上常有好几张表，各自 head() 一次；
        直接赋值的话后一张表会把前一张的列宽压掉——①核对概览的「含义」列本来给 60，
        被后面的表压成 8，一段话撑成二十几行高（实测）。
        """
        for i, (c, w) in enumerate(zip(cols, widths), 1):
            x = ws.cell(row, i, c)
            x.font, x.fill, x.alignment, x.border = FH, PatternFill("solid", fgColor=_FILL_HEAD), CTRW, BD
            L = get_column_letter(i)
            # ⚠ 不能读 ws.column_dimensions[L].width 来比：openpyxl 对没设过的列返回默认值 13，
            #   于是任何小于 13 的宽度都会被「max」吃掉（②页 C 列该 9、D 列该 7，全被顶成 13）。
            #   自己记一份账，只在**我们设过**的宽度之间取大。
            key = (ws.title, L)
            w = max(_setw.get(key, 0), w)
            _setw[key] = w
            ws.column_dimensions[L].width = w
        lines = max(str(c).count("\n") + 1 for c in cols) if cols else 1
        ws.row_dimensions[row].height = 15 * lines + 8

    def note_h(text, span):
        """合并单元格不会自动调高，得自己估：一列宽约放 0.5 个汉字，最多给四行。
        不封顶的话一段长说明能撑掉大半屏（使用者反馈「好多全挤在一起」）。"""
        per = max(40, int(span * 6))
        return 15 * max(1, min(4, -(-len(str(text)) // per))) + 8

    def title(ws, text, span, sub=""):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
        c = ws.cell(1, 1, text); c.font = FT; c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 26
        if sub:
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
            c = ws.cell(2, 1, sub); c.font = Font(name="Arial", size=9, italic=True, color="595959"); c.alignment = LFT
            ws.row_dimensions[2].height = note_h(sub, span)

    wb = Workbook()

    # ---- ⑤ 口径参数（先建，供公式引用）----
    wsP = wb.active
    wsP.title = "⑤口径参数"
    title(wsP, "⑤ 口径参数 — 蓝色单元格可改，全表随之重算", 3, "工时口径来源：" + _RULE_SRC)
    head(wsP, 4, ["参数", "取值", "说明"], [24, 14, 78])
    PARAMS = [
        ("取整粒度（小时）", p["round_step"], "「半小时取整」的粒度"),
        ("取整方向", "向下取整" if p["round_mode"] == "floor" else "四舍五入",
         "向下取整来自人力答复与成本会计底表公式；改这里需同步改本页公式"),
        ("白班扣减（小时）", p["day_break"], "午饭"),
        ("夜班扣减（小时）", p["night_break"], "夜宵"),
        ("多记弹性（小时/天）", p["tolerance"], "人力上报高于打卡口径时，每天此值以内视为正常波动；超过判异常"),
        ("上报工时的口径", "排班班次时长" if _BASIS != "punch" else "按打卡重算",
         "排班班次时长＝2026-06 全量 448 人实证（跨度 12.5h 与 13.0h 的日子上报同为 11.0h，即标准班扣 1 小时休息）；"
         "此口径下只问「打卡撑不撑得起上报的班次」。按打卡重算＝人力 2026-08-10 口头口径，实证不成立，保留可切回"),
    ]
    for i, (a, b, c_) in enumerate(PARAMS):
        r = 5 + i
        wsP.cell(r, 1, a).font = FBB
        x = wsP.cell(r, 2, b); x.font = Font(name="Arial", size=11, bold=True, color="0000FF")
        x.fill = PatternFill("solid", fgColor=_FILL_YEL); x.alignment = CTR
        wsP.cell(r, 3, c_).font = FB
        for cc in range(1, 4):
            wsP.cell(r, cc).border = BD
            if cc != 2:
                wsP.cell(r, cc).alignment = LFT
    C_STEP, C_DAY, C_NIGHT, C_TOL = "'⑤口径参数'!$B$5", "'⑤口径参数'!$B$7", "'⑤口径参数'!$B$8", "'⑤口径参数'!$B$9"

    # ---- ⑤ 下半页：合同价（派遣方 × 岗位 × 班次）----
    # 单独成表的理由：员工工资、管理费、白班夜班三个维度都按派遣方分档，用一个全局单价必错。
    # ⚠ 这里写的是**合同价登记表**（成本会计登记的行），不是汇总表表头解析出来的规则——
    #    表头是人力自己写的，只能作参考，放在最下面单独一段。
    rr = res.get("rates") or {}
    r = 11
    wsP.cell(r, 1, "合同价（派遣方 × 岗位 × 班次）——应付与结算表自查的唯一基准").font = FT
    r += 1
    wsP.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    c = wsP.cell(r, 1, f"来源：{rr.get('合同来源') or '—'}。没登记的档在②页记为「待核」，不拿任何别的价顶上。")
    c.font = FB; c.alignment = LFT; c.border = BD
    r += 1
    head(wsP, r, ["派遣方", "岗位", "白班（员工＋管理费＝合计）", "夜班（员工＋管理费＝合计）"], [16, 12, 32, 32])
    r += 1

    def _band(v):
        if not v:
            return "—（无此班次）"
        return f"{v[0]} ＋ {v[1]} ＝ {round(v[0] + v[1], 2)}"

    def _table_rows(tbl):
        nonlocal r
        for agency, posts in (tbl or {}).items():
            for post, band in posts.items():
                wsP.cell(r, 1, agency).font = FBB
                wsP.cell(r, 2, post).font = FB
                wsP.cell(r, 3, _band(band.get("day"))).font = FB
                wsP.cell(r, 4, _band(band.get("night"))).font = FB
                for cc in range(1, 5):
                    x = wsP.cell(r, cc); x.border = BD; x.alignment = CTR if cc > 2 else LFT
                r += 1

    if rr.get("合同表"):
        _table_rows(rr["合同表"])
    else:
        wsP.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        c = wsP.cell(r, 1, "⚠ 本期一行合同价都没登记——②页「按合同价应付」全部为空、结论全部「待核」。"
                           "请成本会计到复核工具第③步登记后重跑。")
        c.font = FBB; c.fill = PatternFill("solid", fgColor=_FILL_RED); c.alignment = LFT; c.border = BD
        r += 1

    mid = rr.get("期中调价") or []
    badrows = rr.get("问题行") or []
    for x in mid:
        wsP.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        c = wsP.cell(r, 1, f"⚠ 本期期中调价：{x.get('派遣方')}·{x.get('岗位')}（{' → '.join(x.get('生效日') or [])}）——"
                           f"本工具按整月一个价核，取覆盖期末那一行；调价当月的差异请人工复核")
        c.font = FBB; c.fill = PatternFill("solid", fgColor=_FILL_YEL); c.alignment = LFT; c.border = BD
        r += 1
    for x in badrows:
        wsP.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        c = wsP.cell(r, 1, f"⚠ 登记表里有问题的行：{x.get('派遣方')}·{x.get('岗位')} {x.get('生效日')} 起——"
                           f"{'；'.join(x.get('问题') or [])}")
        c.font = FBB; c.fill = PatternFill("solid", fgColor=_FILL_RED); c.alignment = LFT; c.border = BD
        r += 1
    miss = st.get("缺合同价") or []
    if miss:
        r += 1
        wsP.cell(r, 1, f"合同价缺档（{st.get('缺合同价人数', len(miss))} 人）——这些人的应付算不出来").font = FBB
        r += 1
        head(wsP, r, ["姓名", "归属", "岗位", "原因"], [16, 12, 32, 32]); r += 1
        for x in miss:
            for i, k in enumerate(["姓名", "归属", "岗位", "原因"], 1):
                c = wsP.cell(r, i, x.get(k) or "—"); c.font = FB; c.border = BD
                c.alignment = LFT if i == 4 else CTR
                c.fill = PatternFill("solid", fgColor=_FILL_YEL)
            r += 1

    # 参考段：汇总表表头写的计价规则（人力自己写的）
    r += 1
    wsP.cell(r, 1, "参考：汇总表表头写的计价规则（人力自己写的，只作参考，不是基准）").font = FBB
    r += 1
    hl = rr.get("对外总价") or {}
    hl_txt = (f"白班 {hl['day']} 元/小时·人、夜班 {hl['night']} 元/小时·人" if hl.get("day") else "表头未写")
    wsP.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    c = wsP.cell(r, 1, f"表头写的合计价（员工工资＋管理费）：{hl_txt}。下面几行是表头规则文字解析出来的分档，与上面的合同价不一致时以合同价为准。")
    c.font = Font(name="Arial", size=9, italic=True, color="595959"); c.alignment = LFT; c.border = BD
    r += 1
    _table_rows(rr.get("表头解析"))
    for w in (rr.get("表头未解析行") or []):
        wsP.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        c = wsP.cell(r, 1, "表头有一行没看懂：" + w); c.font = FB; c.alignment = LFT; c.border = BD
        r += 1

    # ---- ④ 逐日核对 ----
    ws3 = wb.create_sheet("④逐日核对")
    title(ws3, "④ 逐日核对 — 按口径重算每一天", 14,
          "「跨度」「重算工时」「差异」「金额影响」「判定」全部是公式，引用⑤页参数；改参数或改打卡时刻，数与底色一起重算。"
          "「单价」＝该人结算表上实际套用的含管理费单价（公司按它付钱）；它与合同价对不对见③页「与合同价」列。"
          "底色：红＝⚠异常（含超弹性多记）、黄＝△少记、浅蓝＝○弹性内多记。")
    C3 = ["姓名", "部门", "归属", "岗位", "班型", "日", "上班打卡", "下班打卡", "打卡\n次数",
          "跨度\n(小时)", "上报\n工时", "重算\n工时", "差异\n(重算-上报)", "判定", "已认定\n（人·时间·理由）", "键",
          "单价\n(含管理费)", "金额影响\n(差异×单价)", "无效卡\n(上下班之外)", "整期净多记键"]
    # ⚠ 新列追加在隐藏键列 P 之后（Q、R、S、T），**不插在中间**——③页的 SUMIFS/COUNTIFS 按列字母引用
    #    K/L/M/N/P，插一列就全错位了。无效卡本该挨着上下班，但受此约束只能摆最后。
    # T＝隐藏辅助列：每天「判过班日」的净多记贡献(上报−重算)，供判定公式按人 SUMIF 算整期、判超弹性。
    # 上/下班打卡列给 13：夜班次日下班显示成「次日 08:33」，窄了会撑成 ########
    head(ws3, 3, C3, [12, 9, 8, 7, 7, 5, 13, 13, 7, 9, 9, 9, 11, 30, 28, 4, 11, 13, 16, 4])
    ws3.column_dimensions["P"].hidden = True      # 键＝姓名|归属|部门，给③页 SUMIFS 用：同名两行不互相串
    ws3.column_dimensions["T"].hidden = True      # 整期净多记辅助列，只给判定公式用，不给人看
    _last3 = 3 + len(res.get("rows") or [])       # 末行号（供判定公式里按人 SUMIF 的绝对区间）
    r = 4
    for x in res["rows"]:
        ws3.cell(r, 1, x["姓名"]); ws3.cell(r, 2, x["部门"]); ws3.cell(r, 3, x["归属"])
        ak = x.get("已认定") or {}
        ws3.cell(r, 15, f"{ak.get('认定人', '')} · {ak.get('时间', '')}　{ak.get('理由', '')}".strip(" ·　") if ak else "")
        ws3.cell(r, 16, f"{x['姓名']}|{x['归属']}|{x['部门']}")
        # 单价写死值、金额留公式：改了打卡时刻，差异重算，钱跟着重算
        cq = ws3.cell(r, 17, x.get("单价") or 0); cq.number_format = "0.0"
        if not x.get("单价"):
            cq.fill = PatternFill("solid", fgColor=_FILL_YEL)      # 结算表无单价、合同也缺档 → 钱算不出来
        ws3.cell(r, 18, f"=M{r}*Q{r}").number_format = "#,##0.00;[Red]-#,##0.00"
        # 无效卡＝上下班之外的打卡（午休/宵夜等），跟网页逐日同一份数据，只摆出来让这行读得通
        ws3.cell(r, 19, "  ".join(x.get("无效卡") or []))
        ws3.cell(r, 4, x["岗位"]); ws3.cell(r, 5, x["班型"]); ws3.cell(r, 6, x["日"])
        # 打卡时刻写成 Excel 时间（跨零点＝时刻＋1天，底层仍是真时间，跨度公式 =(H-G)*24 才能算）；
        # 显示用条件格式：值≥1（跨了天）就带「次日」前缀显示成「次日 08:33」，不再甩个累计的 32:33
        for col, key in ((7, "上班打卡"), (8, "下班打卡")):
            v = x[key]
            if v:
                nxt = v.startswith("次日")
                hh, mm = v.replace("次日", "").split(":")
                cell = ws3.cell(r, col, (int(hh) * 60 + int(mm) + (1440 if nxt else 0)) / 1440)
                cell.number_format = '[>=1]"次日 "hh:mm;hh:mm'
            else:
                ws3.cell(r, col, "—")
        ws3.cell(r, 9, x["打卡次数"])
        ws3.cell(r, 10, f"=IF(OR(G{r}=\"—\",H{r}=\"—\"),0,(H{r}-G{r})*24)").number_format = "0.00"
        ws3.cell(r, 11, x["上报工时"]).number_format = "0.0"
        ws3.cell(r, 12, f'=IF(I{r}<2,0,MAX(FLOOR(J{r},{C_STEP})-IF(E{r}="夜班",{C_NIGHT},{C_DAY}),0))').number_format = "0.0"
        ws3.cell(r, 13, f"=L{r}-K{r}").number_format = "0.0;[Red]-0.0"
        # T（隐藏）：这天若是「判过班日」（上报>0、有打卡、非白夜混合）就记净多记贡献(上报−重算)，否则 0。
        # 判定公式按人 SUMIF 它 → 得整期净多记，据此判超弹性（与内核 compute 的净多记口径同构）。
        ws3.cell(r, 20, f'=IF(AND(K{r}>0,I{r}>0,E{r}<>"白夜混合"),K{r}-L{r},0)').number_format = "0.0"
        # 判定公式：已认定的先盖成「✓ 已认定」（不再标红、③页也不再数它）；白夜混合不判档；
        # 再往下按**上报口径**分两套（与内核 compute 完全同构，改一处必须改两处）
        # shift 口径：异常与否看**这个人整期净多记**（按人 SUMIF T 列）是否超弹性——与内核逐人判定同构。
        # 逐日冒尖(M<0)但整期没超 → 「○…整期在弹性内，已消化」(中性)；整期超 → 「⚠…整期超弹性，异常」。
        _pnet = f'SUMIF($P$4:$P${_last3},P{r},$T$4:$T${_last3})'
        _tail = (f'IF(AND(M{r}<-0.001,{_pnet}>{C_TOL}+0.001),"⚠ 多记 "&TEXT(-M{r},"0.0")&" 小时(整期超弹性,异常)",'
                 f'IF(M{r}<-0.001,"○ 多记 "&TEXT(-M{r},"0.0")&" 小时(整期在弹性内,已消化)",'
                 f'IF(M{r}>0.001,"✓ 打卡撑得住上报（在厂 "&TEXT(L{r},"0.0")&"h ≥ 上报 "&TEXT(K{r},"0.0")&"h）",'
                 f'"✓ 打卡撑得住上报")))'
                 if _BASIS != "punch" else
                 f'IF(M{r}>0.001,"△ 少记 "&TEXT(M{r},"0.0")&" 小时",'
                 f'IF(M{r}<-{C_TOL}-0.001,"⚠ 多记 "&TEXT(-M{r},"0.0")&" 小时(超弹性,异常)",'
                 f'IF(M{r}<-0.001,"○ 多记 "&TEXT(-M{r},"0.0")&" 小时(弹性内)","✓ 与口径一致")))')
        ws3.cell(r, 14, f'=IF(O{r}<>"","✓ 已认定："&O{r},'
                        f'IF(E{r}="白夜混合","◇ 白夜混合，本口径下不逐日判，待人工",'
                        f'IF(AND(I{r}=0,K{r}>0),"⚠ 记了工时但无打卡",'
                        f'IF(AND(K{r}=0,I{r}<=1),"△ 仅1次打卡·未记工时，疑似无效",'
                        f'IF(K{r}=0,"◇ 有打卡·当天未计临时工工时",'
                        # ⚠ 五层外层 IF（O/E/AND-I/AND-K/K）要五个右括号收尾；_tail 自身闭合。
                        #   早先只写了四个 → 公式 13 开 12 闭不平衡，Excel 判整列无效、判定全空（使用者实测）。
                        + _tail + ')))))')
        # ⚠ 白夜混合班在 shift 口径下是内核**按日切白/夜**算的，上面这套简单公式复刻不出来：
        #   重算公式会把白班扣减套到混合日、判定又被「E=白夜混合→◇」那层无条件盖成中性，
        #   于是混合工的「打卡撑不起上报(多付钱)」在网页标红、导出却成无色◇被隐藏（复查揪出）。
        #   对混合行改写**内核算好的静态值**（重算/差异/判定），保证导出与网页一致；混合行本就无法随参数联动重算。
        if x.get("班型") == "白夜混合" and _BASIS != "punch":
            ws3.cell(r, 12, x.get("重算工时") or 0).number_format = "0.0"
            ws3.cell(r, 13, x.get("差异") or 0).number_format = "0.0;[Red]-0.0"
            _ak = x.get("已认定") or {}
            ws3.cell(r, 14, ("✓ 已认定：" + f"{_ak.get('认定人','')} · {_ak.get('时间','')}　{_ak.get('理由','')}".strip(" ·　"))
                     if _ak else (x.get("判定") or ""))
        r += 1
    LAST3 = r - 1
    if LAST3 >= 4:
        ws3.cell(r, 1, "合计").font = FBB
        for col in (11, 12, 13, 18):
            L = get_column_letter(col)
            c = ws3.cell(r, col, f"=SUM({L}4:{L}{LAST3})"); c.font = FBB
            c.fill = PatternFill("solid", fgColor=_FILL_GRN)
            c.number_format = "#,##0.00" if col == 18 else "0.0"
        for rr in range(4, r + 1):
            for cc in range(1, 20):
                cell = ws3.cell(rr, cc); cell.border = BD
                if not cell.font.bold:
                    cell.font = FB
                cell.alignment = LFT if cc in (14, 15, 19) else CTR
        ws3.freeze_panes = "A4"
        ws3.auto_filter.ref = f"A3:S{LAST3}"      # 含 Q单价 / R金额影响 / S无效卡；P 是隐藏键列
        # ◇＝中性档（未计工时 / 白夜混合），**不上色**——它们不是问题，标了色就把真问题淹了
        for pref, fill in (("⚠", _FILL_RED), ("△", _FILL_YEL), ("○", _FILL_BLU)):
            ws3.conditional_formatting.add(
                f"A4:S{LAST3}", FormulaRule(formula=[f'LEFT($N4,1)="{pref}"'],
                                            fill=PatternFill("solid", fgColor=fill), stopIfTrue=True))

    # ---- ③ 逐人核对 ----
    ws2 = wb.create_sheet("③逐人核对")
    title(ws2, "③ 逐人核对 — 少记与多记分开列，多记再按弹性分档；「与合同价」列标出这个人的单价对不对", 17,
          "各列由④页按判定分类汇总（SUMIFS/COUNTIFS）。红底＝该人有超弹性异常多记；黄底＝有少记但无异常多记。"
          "「偏离计价单价」＝该人结算表上实际套用的含管理费单价（没有金额列时用合同价），差额金额＝差异×它。")
    C2 = ["序号", "姓名", "部门", "归属", "岗位", "班型", "上报\n总工时", "重算\n总工时", "差异",
          "偏离计价\n单价", "差额金额", "△少记\n日次", "△少记\n小时", "○弹性内\n多记日次",
          "⚠异常多记\n日次", "⚠异常多记\n小时", "与合同价", "键"]
    head(ws2, 3, C2, [6, 12, 9, 8, 7, 7, 9, 9, 9, 9, 10, 8, 8, 10, 11, 11, 34, 4])
    ws2.column_dimensions["R"].hidden = True
    # 匹配键＝姓名|归属|部门（④页 P 列）——只按姓名 SUMIFS，结算表同名两行会互相重复计入、合计被放大
    A = f"'④逐日核对'!$P$4:$P${max(LAST3, 4)}"
    NN = f"'④逐日核对'!$N$4:$N${max(LAST3, 4)}"
    MM = f"'④逐日核对'!$M$4:$M${max(LAST3, 4)}"
    KK = f"'④逐日核对'!$K$4:$K${max(LAST3, 4)}"
    LL = f"'④逐日核对'!$L$4:$L${max(LAST3, 4)}"
    r = 4
    for x in res["people"]:
        ws2.cell(r, 1, x["序号"]); ws2.cell(r, 2, x["姓名"]); ws2.cell(r, 3, x["部门"])
        ws2.cell(r, 4, x["归属"]); ws2.cell(r, 5, x["岗位"]); ws2.cell(r, 6, x["班型"])
        # 「与合同价」：不符要写清差在哪（这一列是给拿去跟人力对话的人看的，只写"不符"没法谈）。
        # ⚠ 不符与缺档可以同时成立（白班对不上、夜班压根没登记），两句都得写，不能 elif 掉一半。
        gaps = x.get("单价不符") or []
        legacy_p = res.get("缺逐人单价核对") or ("单价不符" not in x)
        parts = []
        if legacy_p:
            parts.append("—（旧版留档未做逐人单价核对，请点「按当前参数重跑」后重新导出）")
        elif gaps:
            amt = sum(g.get("金额") or 0.0 for g in gaps)
            t1 = "⚠ 不符：" + "；".join(
                f"{g['项目']} 表上 {g['表上']}、合同 {g['合同']}（{'多' if g['差'] > 0 else '少'} {abs(g['差'])}）"
                for g in gaps)
            if abs(amt) > 0.005:
                t1 += f"　→ 单价算错{'多付' if amt > 0 else '少付'} {abs(amt):,.2f} 元"
            parts.append(t1)
        if x.get("合同缺档"):
            parts.append("△ 合同缺档：" + str(x["合同缺档"]))
        if not parts:
            parts.append("—（结算表无单价列）" if x.get("表上合计") is None else "✓ 一致")
        cg = ws2.cell(r, 17, "　".join(parts))
        if legacy_p:
            cg.fill = PatternFill("solid", fgColor=_FILL_YEL)
        elif gaps:
            cg.fill = PatternFill("solid", fgColor=_FILL_RED)
        elif x.get("合同缺档"):
            cg.fill = PatternFill("solid", fgColor=_FILL_YEL)
        ws2.cell(r, 18, f"{x['姓名']}|{x['归属']}|{x['部门']}")
        ws2.cell(r, 7, f"=SUMIFS({KK},{A},$R{r})").number_format = "0.0"
        ws2.cell(r, 8, f"=SUMIFS({LL},{A},$R{r})").number_format = "0.0"
        ws2.cell(r, 9, f"=H{r}-G{r}").number_format = "0.0;[Red]-0.0"
        cj = ws2.cell(r, 10, x["含管理费单价"]); cj.number_format = "0.0"
        if not x["含管理费单价"]:
            # 结算表没单价、合同价也缺：差额算不出来。写 0 会被当成"没差额"，标黄提醒
            cj.fill = PatternFill("solid", fgColor=_FILL_YEL)
        ws2.cell(r, 11, f"=I{r}*J{r}").number_format = "#,##0.0;[Red]-#,##0.0"
        ws2.cell(r, 12, f'=COUNTIFS({A},$R{r},{NN},"△ 少记*")')
        ws2.cell(r, 13, f'=SUMIFS({MM},{A},$R{r},{NN},"△ 少记*")').number_format = "0.0"
        ws2.cell(r, 14, f'=COUNTIFS({A},$R{r},{NN},"○ 多记*")')
        ws2.cell(r, 15, f'=COUNTIFS({A},$R{r},{NN},"⚠ 多记*")')
        ws2.cell(r, 16, f'=-SUMIFS({MM},{A},$R{r},{NN},"⚠ 多记*")').number_format = "0.0"
        r += 1
    LAST2 = r - 1
    if LAST2 >= 4:
        ws2.cell(r, 2, "合计").font = FBB
        for col in (7, 8, 9, 11, 12, 13, 14, 15, 16):
            L = get_column_letter(col)
            c = ws2.cell(r, col, f"=SUM({L}4:{L}{LAST2})"); c.font = FBB
            c.fill = PatternFill("solid", fgColor=_FILL_GRN)
            c.number_format = "#,##0.0" if col == 11 else "0.0"
        for rr in range(4, r + 1):
            for cc in range(1, 19):
                cell = ws2.cell(rr, cc); cell.border = BD
                if not cell.font.bold:
                    cell.font = FB
                cell.alignment = LFT if cc == 17 else CTR
        ws2.freeze_panes = "C4"
        ws2.auto_filter.ref = f"A3:Q{LAST2}"      # Q＝与合同价（可见列必须进筛选区；R 是隐藏键列，留在区外）
        # 行底色只讲工时（红＝有异常多记、黄＝有少记）；单价对不对由 Q 列自己的底色讲，不混
        ws2.conditional_formatting.add(f"A4:P{LAST2}", FormulaRule(
            formula=["$O4>0"], fill=PatternFill("solid", fgColor=_FILL_RED), stopIfTrue=True))
        ws2.conditional_formatting.add(f"A4:P{LAST2}", FormulaRule(
            formula=["$L4>0"], fill=PatternFill("solid", fgColor=_FILL_YEL)))

    # ---- ① 概览 ----
    ws1 = wb.create_sheet("①核对概览", 0)
    # 副标题按口径分：默认 shift（上报＝班次时长）下没有「少记/弹性内多记」这些档，
    # 别再写 punch 口径那套「少记全部列出、多记弹性」措辞，否则跟下方四档表自相矛盾（复查揪出）。
    _sub1 = ("上报工时 ＝ 排班班次时长（标准班 11 小时），不是按打卡算的；本表只问「打卡在厂时长撑不撑得起上报」。"
             "「差异」＝在厂重算 − 上报，仅供参考——为正是在厂比上报多（正常），撑不起上报的才要查。"
             if _BASIS != "punch" else
             "差异 ＝ 按口径重算 − 人力上报。正数＝人力少记（工人少拿）；负数＝人力多记（公司多付）。"
             "少记全部列出不设弹性；多记每天弹性以内视为正常波动。")
    title(ws1, f"临时工考勤工时核对报告{('　' + month_label) if month_label else ''}", 5, _sub1)
    r = 4
    ws1.cell(r, 1, "一、四档判定").font = FT; r += 1
    head(ws1, r, ["判定档", "日次", "工时(小时)", "金额(元)", "含义"], [30, 10, 13, 13, 72]); r += 1
    _bands = ([("✓ 打卡撑得住上报", st["一致日次"], "—", "—",
                "在厂时长撑得起上报的班次。上报是排班班次时长（标准班 11 小时），本就小于在厂时长，不是问题；"
                f"其中 {st.get('打卡多于上报日次', 0)} 天在厂比上报多，合计 {st.get('打卡多于上报小时', 0)} 小时，仅供参考"),
               ("⚠ 打卡撑不起上报（异常）", st["异常多记日次"], st["异常多记小时"], st["异常多记金额"],
                "唯一必须查的一档：在厂时长撑不起上报的班次，公司可能多付了钱，须业务逐笔说明（实务上多为漏打卡）")]
              if _BASIS != "punch" else
              [("✓ 与口径一致", st["一致日次"], "—", "—", "重算与上报分毫不差"),
               ("○ 多记（弹性内）", st["弹性内多记日次"], st["弹性内多记小时"], st["弹性内多记金额"], "打卡分钟级抖动，视为正常，不追"),
               ("⚠ 多记（超弹性，异常）", st["异常多记日次"], st["异常多记小时"], st["异常多记金额"],
                "唯一必须查的一档：公司多付了钱，须业务逐笔说明（实务上多为漏打卡）"),
               ("△ 少记（提示，不要求补付）", st["少记日次"], st["少记小时"], st["少记金额"],
                "上报少于打卡口径。财务可接受——上报工时本就不是按打卡算出来的")])
    for a, n, h, m, note in _bands + [
        ("⚠ 待查（报了工时却没打卡）", st["待查日次"], "—", "—", "报了工时、却一次卡都没有——先查打卡是否完整再谈工时"),
        ("◇ 有打卡·未计工时", st.get("未计工时日次", 0), "—", "—",
         f"有打卡、当天没算临时工工时（其中 {st.get('未计工时·仅1次卡', 0)} 天只有一次卡），"
         f"涉及 {st.get('未计工时人数', 0)} 人。打卡表是全厂门禁数据，这些天多半是这人在别的名目下上班——不是漏记"),
        ("◇ 白夜混合（待人工）", st.get("白夜混合日次", 0), "—", "—",
         "同月既有白班又有夜班的人。切班规则已按 2026-06 全量实证定案（首卡落在 16:00 之后按夜班切，下班取次日 11:30 前的末卡），这些日子已逐日判过档，正常情况下这一档是 0；"
         "只有切回「按打卡重算」口径时才会有数——那个口径要精确到小时，混合日撑不起。"),
    ]:
        ws1.cell(r, 1, a).font = FBB
        for i, v in enumerate((n, h, m), start=2):
            c = ws1.cell(r, i, v); c.font = FBB
            if isinstance(v, (int, float)):
                c.number_format = "#,##0.0" if i > 2 else "#,##0"
        ws1.cell(r, 5, note).font = FB
        fill = _FILL_RED if a.startswith("⚠") else (_FILL_YEL if a.startswith("△") else None)   # ◇ 中性档不上色
        for cc in range(1, 6):
            x = ws1.cell(r, cc); x.border = BD; x.alignment = LFT if cc in (1, 5) else CTR
            if fill:
                x.fill = PatternFill("solid", fgColor=fill)
        r += 1
    r += 1
    ws1.cell(r, 1, "二、总量").font = FT; r += 1
    # 说明列合并 C:E —— 与上面「四档判定」的「含义」列对齐到同一段，
    # 两张表就不会抢同一列的宽度（列宽只加不减，抢起来会把「工时」这种数字列撑成 60）
    head(ws1, r, ["项目", "数值", "说明", "", ""], [30, 16, 13, 13, 72]); r += 1
    ws1.merge_cells(start_row=r - 1, start_column=3, end_row=r - 1, end_column=5)
    for a, v, note in [
        ("人数 / 比对人日", f"{st['人数']} / {st['比对人日']}", "结算表名单为准，反查打卡"),
        ("上报总工时（小时）", st["上报总工时"], "人力汇总表原数"),
        ("按口径重算（小时）", st["重算总工时"],
         f"(下班−上班) 按 {p['round_step']} 小时{'向下取整' if p['round_mode'] == 'floor' else '四舍五入'} "
         f"− 白班 {p['day_break']} 小时/夜班 {p['night_break']} 小时"),
        ("其中：异常多记已认定（日次）", (st.get("异常多记日次") or 0) - (st.get("异常多记未认定") if st.get("异常多记未认定") is not None else (st.get("异常多记日次") or 0)),
         "已由人确认无误的超弹性多记，④页判定列显示「✓ 已认定」，不再计入③页异常数"),
        ("单价与合同价不符人数",
         "—（旧版留档未核）" if res.get("缺逐人单价核对") else st.get("单价不符人数", 0),
         "结算表给这些人套用的单价与合同价登记表对不上（③页「与合同价」列标红，写明差在哪）"),
        ("偏离未计价人数", st.get("偏离未计价人数", 0), "结算表没有单价、合同价也没登记，差额金额按 0 计（③页单价格标黄）"),
        ("差异（小时）", st["差异小时"], "正＝少记；负＝多记"),
        ("差额金额（元）", st["差额金额"], "差异 × 该人结算表上实际套用的含管理费单价（没有金额列时用合同价）"),
        ("合同价缺档人数", st.get("缺合同价人数", 0), "所属档位没登记合同价，应付算不出来；②页结论为「待核」"),
        ("白夜混合人数", st["白夜混合人数"], "同月既有白班又有夜班。切班规则已按 2026-06 全量实证定案（首卡落在 16:00 之后按夜班切，下班取次日 11:30 前的末卡），逐日已正常判档"),
        ("未匹配到打卡", st["未匹配人数"], "结算表有此人、打卡表找不到"),
        ("待人工指认", st["待指认人数"], "姓名归一后撞上多个人，工具不猜，需人工指定"),
    ]:
        ws1.cell(r, 1, a).font = FBB
        c = ws1.cell(r, 2, v)
        if isinstance(v, (int, float)):
            c.number_format = "#,##0.00" if "元" in a else "#,##0.0"
        c3 = ws1.cell(r, 3, note); c3.font = FB; c3.alignment = LFT
        ws1.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        for cc in range(1, 4):
            x = ws1.cell(r, cc); x.border = BD; x.alignment = CTR if cc == 2 else LFT
        r += 1

    # ---- ② 复核结论（应付）----
    # 这一页是给需求方直接对着请款单核的：一家派遣方一条业务线，该付多少工资、多少管理费。
    # 基数是**人力上报的工时**，不是打卡重算的——付款本来就按上报的数走（2026-08-18 业务定案）。
    se = res.get("settle") or {}
    det = se.get("明细") or []
    tot = se.get("合计") or {}
    has_tbl = (st.get("有表上金额") or 0) > 0
    wsC = wb.create_sheet("②复核结论")
    span = 15 if has_tbl else 10
    title(wsC, f"② 复核结论 — 每家派遣方 · 每条业务线应付多少{('　' + month_label) if month_label else ''}", span,
          ("⚠⚠ 本期是旧版留档：按旧规则算的，没有合同价核对，下表结论一律「待核」——请在复核工具里点「按当前参数重跑」后重新导出。　"
           if res.get("旧版留档") else "") +
          "⚠ 主列是【结算表应付】，也就是各家请款单上的金额；右侧【按合同价应付】是校验尺，"
          "用来量人力有没有按合同的价算，本身不是要付的钱。"
          "口径：上报白班工时 × 白班合同价 ＋ 上报夜班工时 × 夜班合同价，工资与管理费分开算，"
          "合同价取「派遣方 × 岗位 × 班次」（见⑤页合同价表）；没登记合同价的格子应付为空、结论「待核」。"
          "基数是人力上报的工时，不是按打卡重算的——付款本就按上报的数走，工时偏离另见③④页。"
          "结论三态：正常＝各项都对得上；异常＝有任何一项对不上；待核＝没发现问题但合同价缺档、没法核。"
          "本工具只读不写账、不产生凭证。"
          "　▸ 本表按派遣方分块：每家「派遣方小计」在上、其业务线明细默认折叠，点表格左侧的 + 号展开明细。")
    C = ["归属\n（派遣方）", "部门\n（业务线）", "岗位", "人数", "上报白班\n工时", "上报夜班\n工时", "上报\n总工时"]
    W = [12, 12, 9, 7, 11, 11, 10]
    if has_tbl:
        C += ["结算表\n工资", "结算表\n管理费", "补贴/奖/罚", "结算表应付\n＝请款金额"]
        W += [12, 12, 11, 14]
    C += ["按合同价\n工资", "按合同价\n管理费", "按合同价\n合计"]
    W += [12, 12, 12]
    if has_tbl:
        C += ["结算−合同价"]
        W += [11]
    C += ["结论", "原因"]          # 结论由工具给，不让人看着「偏差 0」自己去推断
    W += [9, 36]
    span += 2
    # 本期复核结论「确认无误」签字留痕（第⑧步）——尤其下表有「异常」时，这行就是「谁看过并接受了这些异常」的账。
    # 不留痕的话，导出的报告里异常摆在那儿却没人签字，等于没人为它负责（使用者 2026-08-30 提）。
    _csign = (db.get_setting(_SIGNOFF) or {}).get(res.get("month") or month_label or "")
    if isinstance(_csign, dict) and _csign.get("确认人"):
        _ct = (f"✓ 本期复核结论已确认无误　——　{_csign.get('确认人','')}　{_csign.get('时间','')}"
               + (f"　（{_csign.get('摘要','')}）" if _csign.get('摘要') else ""))
        _cc = wsC.cell(3, 1, _ct); _cc.font = Font(name="Arial", size=10, bold=True, color="15803D")
    else:
        _cc = wsC.cell(3, 1, "○ 本期复核结论尚未确认（复核工具第⑧步点「确认无误」后此处留痕；下表若有异常，更该有人签字）")
        _cc.font = Font(name="Arial", size=10, italic=True, color="B91C1C")
    wsC.merge_cells(start_row=3, start_column=1, end_row=3, end_column=span)
    head(wsC, 4, C, W)
    r = 5
    KEYS = ["归属", "部门", "岗位", "人数", "上报白班工时", "上报夜班工时", "上报总工时"]
    KEYS += (["表上工资", "表上管理费", "补贴奖罚", "表上合计"] if has_tbl else [])
    KEYS += ["应付工资", "应付管理费", "应付合计"] + (["应付偏差"] if has_tbl else [])
    KEYS += ["结论", "异常原因"]
    # 请款金额那一组给底色、合计列加粗——拿 OA 请款单来对的人得一眼找到该盯哪一列
    PAY_COLS = range(8, 12) if has_tbl else range(0, 0)
    PAY_TOTAL = 11 if has_tbl else 0
    # 总分结构（按派遣方）：一家一块——先它的「派遣方小计」(总)，紧接它的业务线明细(分)，
    # 明细用 Excel 分组，可在左侧 +/− 折叠（使用者 2026-08-30 提：原来「全部明细在上、全部小计在下」，
    # 要对某一家的请款单得翻半天才凑到它的小计）。summaryBelow=False：折叠钮落在上面的小计行，
    # 跟网页「点小计看明细」一致；明细默认收起，只看各家小计，点 + 展开。
    if wsC.sheet_properties.outlinePr is None:
        from openpyxl.worksheet.properties import Outline
        wsC.sheet_properties.outlinePr = Outline()
    wsC.sheet_properties.outlinePr.summaryBelow = False
    _by_ag = {}
    for d in det:
        _by_ag.setdefault(d.get("归属"), []).append(d)

    def _wd(r, d, grouped=True):        # 写一行明细（分行）；grouped=可折叠、默认收起
        for i, k in enumerate(KEYS, 1):
            v = d.get(k)
            if isinstance(v, list):                 # 异常原因是列表
                v = "；".join(str(x) for x in v)
            c = wsC.cell(r, i, "—" if v in (None, "") else v)
            c.font, c.border = FB, BD
            c.alignment = LFT if k == "异常原因" else CTR
            if isinstance(v, (int, float)) and i >= 4:
                c.number_format = "#,##0" if i == 4 else ("#,##0.0" if i <= 7 else "#,##0.00")
            if i in PAY_COLS:
                c.fill = PatternFill("solid", fgColor=_FILL_PAY)
            if i == PAY_TOTAL:
                c.font = FBB
            if k == "结论":
                c.font = FBB
                c.fill = PatternFill("solid", fgColor=(_FILL_RED if v == "异常"
                                                       else _FILL_YEL if v == "待核" else _FILL_GRN))
        if has_tbl and abs(d.get("应付偏差") or 0) > 0.01:
            for cc in range(1, span + 1):
                wsC.cell(r, cc).fill = PatternFill("solid", fgColor=_FILL_RED)
        if grouped:
            wsC.row_dimensions[r].outlineLevel = 1
            wsC.row_dimensions[r].hidden = True

    def _wsub(r, sub):                  # 写一行派遣方小计（总行 / 分组头）
        wsC.cell(r, 1, sub["归属"]); wsC.cell(r, 2, "派遣方小计"); wsC.cell(r, 4, sub["人数"])
        for i, k in enumerate(KEYS[4:], 5):
            v = sub.get(k)
            if isinstance(v, list):
                v = "；".join(str(x) for x in v)
            c = wsC.cell(r, i, v)
            if k not in ("结论", "异常原因"):
                c.number_format = "#,##0.0" if i <= 7 else "#,##0.00"
        for cc in range(1, span + 1):
            x = wsC.cell(r, cc); x.font = FBB; x.border = BD; x.alignment = CTR
            x.fill = PatternFill("solid", fgColor=_FILL_GRN)
        _verdict_fill(wsC, r, KEYS, sub.get("结论"))
        wsC.row_dimensions[r].collapsed = True        # 分组头：其下明细默认折叠

    _seen = set()
    for sub in (se.get("派遣方小计") or []):
        ag = sub["归属"]; _seen.add(ag)
        _wsub(r, sub); r += 1
        for d in _by_ag.get(ag, []):
            _wd(r, d); r += 1
    # 兜底：万一某条明细的归属不在小计名单里（理论上不会，小计是按归属滚的），也原样摆出、不折叠、别丢
    for d in det:
        if d.get("归属") not in _seen:
            _wd(r, d, grouped=False); r += 1
    wsC.cell(r, 1, "全表合计"); wsC.cell(r, 4, tot.get("人数"))
    for i, k in enumerate(KEYS[4:], 5):
        v = tot.get(k)
        if isinstance(v, list):
            v = "；".join(str(x) for x in v)
        c = wsC.cell(r, i, v)
        if k not in ("结论", "异常原因"):
            c.number_format = "#,##0.0" if i <= 7 else "#,##0.00"
    for cc in range(1, span + 1):
        x = wsC.cell(r, cc); x.font = FBB; x.border = BD; x.alignment = CTR
        x.fill = PatternFill("solid", fgColor=_FILL_HEAD); x.font = FH
    _verdict_fill(wsC, r, KEYS, tot.get("结论"), dark=True)
    wsC.freeze_panes = "D5"
    r += 2

    # ---- 合同外调整：奖 / 罚 / 蒸练补贴 ----
    # 单独成页（使用者 2026-08-29 提）：这块有自己的十列表头和一大段说明，
    # 塞在②页里会把②的列宽也带偏，且②本身是给人对请款单用的，混着看容易串。
    # ②这里只留一行指路。
    adj = st.get("合同外调整") or []
    adjT = st.get("合同外调整合计") or {}
    if adj:
        wsC.cell(r, 1, f"合同外调整（奖 / 罚 / 蒸练补贴）：{adjT.get('笔数', 0)} 笔，"
                       f"净额 {adjT.get('净额', 0):,.2f} 元"
                       + (f"　⚠ {adjT.get('异常', 0)} 异常 / {adjT.get('存疑', 0)} 存疑"
                          if (adjT.get("异常") or adjT.get("存疑")) else "　全部正常")
                       + "　——逐笔明细见「⑦合同外调整」页").font = FT
        r += 2

    chk = st.get("金额核对") or []
    if not has_tbl:
        wsC.cell(r, 1, "结算表自查：结算表没有金额列，自查未做").font = FT
    else:
        wsC.cell(r, 1, "结算表自查" + ("：全部通过" if not chk else f"：{len(chk)} 处对不上")).font = FT
    r += 1
    note = ("① 单价：表上单价 vs 合同价登记表（没登记的档跳过）　② 金额：表上金额 vs 表上工时×表上单价　"
            "③ 勾稽：员工工资＝白工资+夜工资+补贴+奖+罚，合计＝员工工资+管理费。"
            + ("本期结算表没有单价/金额列（可能是按派遣方拆出来的简表），这三道都没法做。" if not has_tbl
               else "三项全部相符，结算表内部自洽。" if not chk else "工具只指出对不上，不改数。"))
    wsC.merge_cells(start_row=r, start_column=1, end_row=r, end_column=span)
    c = wsC.cell(r, 1, note); c.font = Font(name="Arial", size=9, italic=True, color="595959"); c.alignment = LFT
    wsC.row_dimensions[r].height = 30
    r += 1
    if chk:
        head(wsC, r, ["姓名", "归属", "岗位", "类型", "项目", "表上", "应为", "差", "说明"],
             [12, 10, 9, 8, 14, 12, 12, 11, 40]); r += 1
        for x in chk[:500]:
            for i, k in enumerate(["姓名", "归属", "岗位", "类型", "项目", "表上", "应为", "差", "说明"], 1):
                c = wsC.cell(r, i, x[k])
                c.font, c.border = FB, BD
                c.alignment = LFT if i == 9 else CTR
                c.fill = PatternFill("solid", fgColor=_FILL_RED)
                if isinstance(x[k], (int, float)):
                    c.number_format = "#,##0.00"
            r += 1

    # ---- ⑥ 表外临时工 ----
    if res.get("outsiders"):
        ws5 = wb.create_sheet("⑥表外临时工")
        title(ws5, "⑥ 打卡表里有、结算名单上没有的临时工", 5,
              "筛选：打卡表部门含「临时普工」、本月有打卡、且不在结算名单内。"
              "打卡表对临时工的部门只写到派遣方，看不出车间，故本页只提示、不下结论。")
        head(ws5, 3, ["姓名", "派遣方", "考勤组", "本月出勤天数", "其中仅1-2次打卡"], [14, 14, 22, 14, 16])
        r = 4
        for o in res["outsiders"]:
            for i, k in enumerate(["姓名", "派遣方", "考勤组", "出勤天数", "仅1-2次打卡天数"], 1):
                x = ws5.cell(r, i, o[k]); x.font = FB; x.border = BD; x.alignment = CTR
            r += 1
        ws5.freeze_panes = "A4"

    # 建表顺序是「先建被引用的页」，与阅读顺序不同，最后统一排成 ①②③④⑤⑥
    # ---- ②页尾：用工成本汇总（派遣方 × 车间）+ 可直接转发的文字 ----
    # 为什么加这一块：每月复核完还要往群里发一条「本月临工总用工成本多少、分车间多少、
    # 每家多少、平均多少元每小时」。原来对着报告手敲，2026-08 那条实测就敲错了两个车间的
    # 平均单价（写 19.47 / 20.46，而用它自己列的各家数字反推是 18.86 / 20.66）——
    # 金额对、单价错，这种错最难看出来。现在直接出表出话，元/h 全做成公式，改了数自己重算。
    if det:
        # ⚠ 起点用 ws.max_row 而不是上面的 r：②页在这之前还有派遣方小计、全表合计、
        #   合同外调整、结算表自查好几块，各块各自维护 r，接着用会压到别人的行上（踩过）。
        _cost_summary(wsC, wsC.max_row + 1, det, month_label, span, FB, FBB, FH, FT, CTR, LFT, BD)

    # ---- ⑦ 合同外调整（奖 / 罚 / 蒸练补贴）----
    if adj:
        wsA = wb.create_sheet("⑦合同外调整")
        title(wsA, f"⑦ 合同外调整 — 奖 / 罚 / 蒸练补贴{('　' + month_label) if month_label else ''}", 12,
              f"{adjT.get('笔数', 0)} 笔，净额 {adjT.get('净额', 0):,.2f} 元。"
              "⚠ 这三项是全表唯一没有对照源的钱：工时有打卡可比、单价有合同可比，"
              "奖罚补贴只有结算表这一处孤证。本工具验不了金额对不对，只做三件事——逐笔列出、"
              "验符号（罚≤0、奖≥0、补贴≥0）、验占比（单笔占当月工资超阈值即提示）。"
              "金额本身请附审批单/处罚通知。**过了检查 ≠ 这笔奖罚是对的**。")
        # 本期奖惩「已核对」确认（第⑦步 period 级签字）——导出也要留痕：谁、什么时候、确认时是几笔多少钱。
        _asign = (db.get_setting(_ADJSIGN) or {}).get(res.get("month") or month_label or "")
        _ar = 3
        if isinstance(_asign, dict) and _asign.get("确认人"):
            _txt = (f"✓ 本期奖惩已核对无误　——　{_asign.get('确认人','')}　{_asign.get('时间','')}"
                    + (f"　（确认时：{_asign.get('摘要','')}）" if _asign.get('摘要') else ""))
            _c = wsA.cell(_ar, 1, _txt)
            _c.font = Font(name="Arial", size=10, bold=True, color="15803D")
        else:
            _c = wsA.cell(_ar, 1, "○ 本期奖惩尚未整体确认（第⑦步「确认本期奖惩已核对」后此处留痕）")
            _c.font = Font(name="Arial", size=10, italic=True, color="92400E")
        wsA.merge_cells(start_row=_ar, start_column=1, end_row=_ar, end_column=12)
        # 加两列：状态（✓已确认/待核）+ 确认（人·时间·理由）——逐笔认定的痕迹跟着报告走
        head(wsA, 4, ["级别", "状态", "姓名", "归属", "部门", "岗位", "项目", "金额", "当月工资", "占工资%", "说明", "确认（人·时间·理由）"],
             [8, 10, 12, 11, 12, 9, 12, 12, 12, 10, 44, 30])
        ra = 5
        for a in adj[:300]:
            _ak = a.get("已认定") or {}
            fill = None if _ak else (_FILL_RED if a["级别"] == "异常" else _FILL_YEL if a["级别"] == "存疑" else None)
            _akinfo = (f"{_ak.get('认定人','')} · {_ak.get('时间','')}　{_ak.get('理由','')}".strip(" ·　")) if _ak else ""
            _rowvals = [a["级别"], ("✓ 已确认" if _ak else "待核"), a["姓名"], a["归属"], a["部门"],
                        a["岗位"], a["项目"], a["金额"], a["当月工资"], a["占工资"], a["说明"], _akinfo]
            for i, v in enumerate(_rowvals, 1):
                c = wsA.cell(ra, i, "—" if v is None else v)
                c.font = FBB if (i == 1 and a["级别"] != "提示" and not _ak) else FB
                c.border = BD
                c.alignment = LFT if i in (11, 12) else CTR
                if isinstance(v, (int, float)):
                    c.number_format = "#,##0.00" if i in (8, 9) else "0.0"
                if i == 2 and _ak:
                    c.font = Font(name="Arial", size=10, color="15803D")   # 已确认标绿
                if i == 12 and _ak:
                    c.font = Font(name="Arial", size=9, color="15803D")
                if fill:
                    c.fill = PatternFill("solid", fgColor=fill)
            ra += 1
        wsA.freeze_panes = "A5"
        if len(adj) > 300:
            wsA.cell(ra + 1, 1, f"（只列前 300 笔，本期共 {len(adj)} 笔）").font = FB

    # ---- ⑧ 打卡原始表（结算名单内）----
    # 报告里其它每一页的工时/判定都从打卡推来，却唯独没有打卡本身。审计或领导要问
    # 「凭什么说某人这天只打了一次卡」，光看报告验不了，得回服务器翻原始文件——
    # 尤其钉钉取数时这张表根本不落到经理手上（现生成、只存服务器）。把它随报告带出，
    # 这份报告才自成闭环、能被人眼复核（而不是只能信工具切班切得对）。
    # 只列结算名单内的人：表外的人不背这份报告里任何要付的数，已在⑥单列；
    # 过滤到名单内也顺带把体量压住（不会是几千行的全厂长表）。
    if punch_bytes:
        try:
            _pk = ta.parse_punch(punch_bytes)
        except Exception:
            _pk = None
        _listed = {ta.norm_name(x.get("姓名") or "") for x in (res.get("people") or [])}
        _recs = [rec for k, rs in ((_pk or {}).get("by_key") or {}).items()
                 if k in _listed for rec in rs]
        if _recs:
            _has_id = any(r.get("标识") for r in _recs)
            _maxday = min(31, max((max(r["days"]) for r in _recs if r.get("days")), default=0))
            _dcols = list(range(1, _maxday + 1))
            # 跨月边界（上月末 / 次月初）：整月取数才有；有就各加一列，摆在日号区两端
            _hp = any((r.get("bnd") or {}).get(ta._BND_PREV) for r in _recs)
            _hn = any((r.get("bnd") or {}).get(ta._BND_NEXT) for r in _recs)
            _recs.sort(key=lambda r: ((r.get("部门") or ""), r.get("raw") or ""))
            # 部门给足宽度：源表若是考勤系统全称（「…公司-生产制造部-临时普工-华顺人力」）
            # 一行摆得下、不转行、不删减（使用者：完整全称 / 不做任何动作 / 不要转行）
            _labels = ["姓名", "部门（全称）", "考勤组", "手机尾号"]
            _widths = [12, 42, 22, 10]
            if _hp:
                _labels.append("上月末"); _widths.append(15)
            _labels += [f"{d}日" for d in _dcols]; _widths += [15] * len(_dcols)
            if _hn:
                _labels.append("次月初"); _widths.append(15)
            ws8 = wb.create_sheet("⑧打卡原始表")
            _src = (_pk or {}).get("title") or ""
            title(ws8, f"⑧ 打卡原始表 — 结算名单内人员 · 逐日打卡时刻{('　' + month_label) if month_label else ''}", max(len(_labels), 6),
                  ("原样带出，未做取整 / 切班 / 判定——它是④逐日、③逐人各项数字的原始依据，可逐人对照。"
                   "只含结算名单内的人；表外（有打卡但不在名单上）见⑥，此处不列。"
                   + ("　含「上月末 / 次月初」两列：夜班跨零点，边界那班的卡落在隔月，一并带出才完整。" if (_hp or _hn) else "")
                   + (f"　来源：{_src}。" if _src else "")
                   + ("　带手机尾号（钉钉取数，同名可辨识）。" if _has_id
                      else "　无手机尾号（人力导出）。")))
            head(ws8, 3, _labels, _widths)
            _r = 4
            for rec in _recs:
                for i, v in enumerate([rec.get("raw") or "", rec.get("部门") or "",
                                       rec.get("组") or "", rec.get("标识") or ""], 1):
                    c = ws8.cell(_r, i, v); c.font = FB; c.border = BD
                    # 部门＝完整全称、**不转行**（使用者 2026-08-29：「完整的全称，不做任何动作，不要转行」）
                    c.alignment = LFTN if i == 2 else CTR
                _j = 5
                _bnd = rec.get("bnd") or {}
                if _hp:
                    ts = _bnd.get(ta._BND_PREV)
                    c = ws8.cell(_r, _j, "  ".join(ta.fmt_hm(t) for t in ts) if ts else "")
                    c.font = FB; c.border = BD; c.alignment = CTR; _j += 1
                for d in _dcols:
                    ts = (rec.get("days") or {}).get(d)
                    c = ws8.cell(_r, _j, "  ".join(ta.fmt_hm(t) for t in ts) if ts else "")
                    c.font = FB; c.border = BD; c.alignment = CTR; _j += 1
                if _hn:
                    ts = _bnd.get(ta._BND_NEXT)
                    c = ws8.cell(_r, _j, "  ".join(ta.fmt_hm(t) for t in ts) if ts else "")
                    c.font = FB; c.border = BD; c.alignment = CTR; _j += 1
                _r += 1
            ws8.freeze_panes = "E4"

    # ---- 结算风险（同名重复计费 / 归属与打卡不符）----
    # 网页④结算风险有、导出一直缺（使用者 2026-08-30 提）。这是「钱被重复付、或付给了不该付的
    # 那一家」的风险，比工时偏离更要紧；连认定痕迹一起带出：谁、什么时候、因为什么认为它没事。
    _mm = st.get("归属与打卡不符") or []
    _dups = st.get("同名多行") or []
    _blind = st.get("归属无法核对名单") or []
    if _mm or _dups or _blind:
        wsR = wb.create_sheet("结算风险")
        title(wsR, f"结算风险 — 同名重复计费 / 归属与打卡不符{('　' + month_label) if month_label else ''}", 8,
              "⚠ 这几项不是工时偏离，是【钱被重复付、或付给了不该付的那一家】，比少记多记要紧。"
              "工具只指认并给判据，是不是同一个人、是不是内部调配由人来认；认了留痕（谁·时间·理由），撤销随时可以。"
              "②复核结论里「归属与打卡不符 N 处」的明细就在这一页。")

        def _rk(rr, vals, ak, badmark=False, lft=()):
            for i, v in enumerate(vals, 1):
                c = wsR.cell(rr, i, "—" if v in (None, "") else v)
                c.font, c.border = FB, BD
                c.alignment = LFT if i in lft else CTR
                if isinstance(v, (int, float)):
                    c.number_format = "#,##0.0"
                if badmark and not ak:
                    c.fill = PatternFill("solid", fgColor=_FILL_RED)
            if ak:                                       # 状态列(倒数第2)标绿
                wsR.cell(rr, len(vals) - 1).font = Font(name="Arial", size=10, color="15803D")

        rr = 3
        wsR.cell(rr, 1, "一、归属与打卡不符（结算算 A 家的钱，考勤却挂在 B）").font = FBB
        rr += 1
        head(wsR, rr, ["姓名（手机尾号）", "结算归属", "打卡部门（派遣方）", "打卡部门原文",
                       "总工时", "状态", "确认（人·时间·理由）"],
             [20, 12, 16, 26, 9, 10, 30])
        rr += 1
        if not _mm:
            wsR.cell(rr, 1, "本期没有归属与打卡不符。").font = FB; rr += 1
        for x in _mm:
            ak = x.get("已认定") or {}
            nm = x.get("姓名", "") + (f"（手机尾号{x.get('手机尾号')}）" if x.get("手机尾号") else "")
            _rk(rr, [nm, x.get("结算归属"), x.get("打卡部门派遣方"), x.get("打卡部门原文"),
                     x.get("总工时"), ("✓ 已确认" if ak else "待认"),
                     (f"{ak.get('认定人','')} · {ak.get('时间','')}　{ak.get('理由','')}".strip(" ·　") if ak else "")],
                ak, lft=(4, 7))
            rr += 1

        rr += 1
        wsR.cell(rr, 1, "二、同名多行（同一姓名多行，可能被重复计费）").font = FBB
        rr += 1
        head(wsR, rr, ["归一姓名", "各写法", "派遣方", "各行工时", "合计工时", "判据 / 风险", "状态", "确认（人·时间·理由）"],
             [12, 18, 14, 14, 9, 30, 10, 30])
        rr += 1
        if not _dups:
            wsR.cell(rr, 1, "本期没有同名多行。").font = FB; rr += 1
        for x in _dups:
            ak = x.get("已认定") or {}
            _rk(rr, [x.get("归一姓名"), " / ".join(x.get("原名") or []), "、".join(x.get("派遣方") or []),
                     " + ".join(str(h) for h in (x.get("各行工时") or [])), x.get("合计工时"),
                     (x.get("风险", "") + (("　判据：" + x["打卡判据"]) if x.get("打卡判据") else "")),
                     ("✓ 已确认" if ak else "待认"),
                     (f"{ak.get('认定人','')} · {ak.get('时间','')}　{ak.get('理由','')}".strip(" ·　") if ak else "")],
                ak, badmark=bool(x.get("高风险")), lft=(2, 3, 4, 6, 8))
            rr += 1

        if _blind:
            rr += 1
            wsR.cell(rr, 1, f"三、归属无法核对（打卡表这几行没写部门，无从跟结算归属比）：{len(_blind)} 人").font = FBB
            rr += 1
            wsR.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=8)
            c = wsR.cell(rr, 1, "、".join(_blind)); c.font, c.alignment = FB, LFT
        wsR.freeze_panes = "A4"

    order = ["①核对概览", "②复核结论", "结算风险", "③逐人核对", "④逐日核对", "⑤口径参数",
             "⑥表外临时工", "⑦合同外调整", "⑧打卡原始表"]
    for i, nm in enumerate([n for n in order if n in wb.sheetnames]):
        wb.move_sheet(nm, offset=i - wb.sheetnames.index(nm))
    bio = BytesIO(); wb.save(bio)
    return bio.getvalue()


def _cost_agg(det):
    """按 派遣方 × 车间 汇总用工成本。**Excel 和网页共用这一份**，别各算各的。

    金额取【结算表应付】——那才是各家请款单上的钱；工时取【上报总工时】，
    与单价口径一致（元/h ＝ 该格金额 ÷ 该格上报工时）。
    ⚠ 不能用「按合同价应付」：合同价缺档的格子那一列是空的，加总会少算。
    """
    amt = "表上合计" if any(x.get("表上合计") is not None for x in det) else "应付合计"
    lines = [x for x in det if x.get("归属") and x.get("部门")]
    shops = sorted({str(x["部门"]) for x in lines})
    cell = {}
    for x in lines:
        v = cell.setdefault((str(x["归属"]), str(x["部门"])), [0.0, 0.0])
        v[0] += float(x.get(amt) or 0)
        v[1] += float(x.get("上报总工时") or 0)
    agencies = sorted({k[0] for k in cell},
                      key=lambda a: -sum(v[0] for k, v in cell.items() if k[0] == a))
    return shops, agencies, cell


def _cost_money(v):
    t = f"{v:,.1f}"
    return t[:-2] if t.endswith(".0") else t


def _cost_text(month_label, shops, agencies, cell):
    """照钉钉通报的写法出一段话，页面和 Excel 都用它。"""
    rate = lambda m, h: f"{(m / h):.2f}" if h else "—"
    tm = sum(v[0] for v in cell.values())
    shop_tot = {sh: (sum(v[0] for k, v in cell.items() if k[1] == sh),
                     sum(v[1] for k, v in cell.items() if k[1] == sh)) for sh in shops}
    out = [f"{month_label or ''}临工总用工成本{_cost_money(tm)}元（"
           + "；".join(f"{sh}车间{_cost_money(shop_tot[sh][0])}元、{rate(*shop_tot[sh])}元/h"
                       for sh in shops) + "）。"]
    for i, a in enumerate(agencies, 1):
        parts = [f"{sh}车间{_cost_money(cell[(a, sh)][0])}元、{rate(*cell[(a, sh)])}元/h"
                 for sh in shops if cell.get((a, sh)) and cell[(a, sh)][0]]
        out.append(f"{i}、{a}：{_cost_money(sum(v[0] for k, v in cell.items() if k[0] == a))}元"
                   f"（{'，'.join(parts)}）；")
    out.append("具体明细详见复核报告「③逐人核对」「④逐日核对」两页。")
    return out


def _attach_cost(res, month=""):
    """把用工成本汇总挂到结果上，页面第⑨步直接渲染、直接复制。

    做进结果而不是让前端自己算：一段要发给领导的通报，两边各算一次迟早会对不上
    （2026-08 那条手敲的通报就是金额对、两个车间的平均单价错，照搬了上个月）。
    """
    try:
        det = (res.get("settle") or {}).get("明细") or []
        if not det:
            return res
        shops, agencies, cell = _cost_agg(det)
        if not shops:
            return res
        rate = lambda m, h: round(m / h, 2) if h else None
        rows = []
        for a in agencies:
            tm = sum(v[0] for k, v in cell.items() if k[0] == a)
            th = sum(v[1] for k, v in cell.items() if k[0] == a)
            rows.append({"派遣方": a, "金额": round(tm, 2), "工时": round(th, 1), "元每小时": rate(tm, th),
                         "车间": {sh: {"金额": round(cell[(a, sh)][0], 2),
                                       "工时": round(cell[(a, sh)][1], 1),
                                       "元每小时": rate(*cell[(a, sh)])}
                                  for sh in shops if cell.get((a, sh)) and cell[(a, sh)][0]}})
        tm = sum(v[0] for v in cell.values())
        th = sum(v[1] for v in cell.values())
        res["用工成本"] = {
            "车间": shops, "行": rows,
            "合计": {"金额": round(tm, 2), "工时": round(th, 1), "元每小时": rate(tm, th),
                     "车间": {sh: {"金额": round(sum(v[0] for k, v in cell.items() if k[1] == sh), 2),
                                   "工时": round(sum(v[1] for k, v in cell.items() if k[1] == sh), 1),
                                   "元每小时": rate(sum(v[0] for k, v in cell.items() if k[1] == sh),
                                                    sum(v[1] for k, v in cell.items() if k[1] == sh))}
                              for sh in shops}},
            "文字": _cost_text(month or res.get("month") or "", shops, agencies, cell)}
    except Exception:
        pass                                   # 汇总算不出来不该让整个复核结果发不出去
    return res


def _cost_summary(ws, r, det, month_label, span, FB, FBB, FH, FT, CTR, LFT, BD):
    """按派遣方 × 车间汇总用工成本，并出一段可直接发群的文字。

    金额取【结算表应付】——那才是各家请款单上的钱；工时取【上报总工时】，
    与单价口径一致（元/h ＝ 该格金额 ÷ 该格上报工时）。
    ⚠ 不能用「按合同价应付」：合同价缺档的格子那一列是空的，加总会少算。
    ⚠ 这里**不调 head()**——那个函数会改列宽，会把上面主表的版式搞乱。
    """
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter as L

    shops, agencies, cell = _cost_agg(det)
    if not shops:
        return
    NC = 4 + len(shops) * 3

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max(span, NC))
    c = ws.cell(r, 1, f"用工成本汇总 —— 按派遣方 × 车间{('　' + month_label) if month_label else ''}")
    c.font, c.alignment = FT, Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 24
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max(span, NC))
    c = ws.cell(r, 1, "金额取【结算表应付】（各家请款单上的钱），工时取【上报总工时】；"
                      "元/h ＝ 该格金额 ÷ 该格上报工时，全部是公式，改了数会自己重算。"
                      "不用「按合同价应付」——合同价缺档的格子是空的，加总会少算。"
                      "最底下那几行文字可直接复制到群里。")
    c.font, c.alignment = Font(name="Arial", size=9, italic=True, color="595959"), LFT
    ws.row_dimensions[r].height = 30
    r += 1

    cols = ["派遣方", "合计\n金额（元）", "合计\n工时", "合计\n元/h"]
    for sh in shops:
        cols += [f"{sh}车间\n金额（元）", f"{sh}车间\n工时", f"{sh}车间\n元/h"]
    # ⚠ 表头必须**开换行**，否则「植物肉车间\n金额（元）」不折两行、在窄列里被截成「肉车间金额…」
    #   （使用者 2026-08-30：注意列宽和转行）。再按最长那段把列宽补够。
    from openpyxl.styles import Alignment as _Al
    from openpyxl.utils import get_column_letter as _gcl
    _CTRW = _Al(horizontal="center", vertical="center", wrap_text=True)
    _cw = [12, 12, 9, 8]                     # 派遣方 / 合计金额 / 合计工时 / 合计元/h
    for _ in shops:
        _cw += [12, 11, 11]                  # 车间金额 / 车间工时 / 车间元/h（「植物肉车间」5 字要 ≥12）
    for i, t in enumerate(cols, 1):
        x = ws.cell(r, i, t)
        x.font, x.fill, x.alignment, x.border = (
            FH, PatternFill("solid", fgColor=_FILL_HEAD), _CTRW, BD)
        # 与②结算表共列，只加不减：head() 已给这些列设过宽度，这里读得到实值
        L2 = _gcl(i)
        _cur = ws.column_dimensions[L2].width
        _need = _cw[i - 1] if i - 1 < len(_cw) else 11
        if _cur is None or _cur < _need:
            ws.column_dimensions[L2].width = _need
    ws.row_dimensions[r].height = 32
    r += 1

    first = r
    for a in agencies:
        ws.cell(r, 1, a)
        for j, sh in enumerate(shops):
            m, h = cell.get((a, sh), [0.0, 0.0])
            c0 = 5 + j * 3
            # ⚠ 必须显式写 0，不能写 None：openpyxl 的 ws.cell(r,c,None) **不赋值**
            #   （源码里 if value is not None 才写），底下若有旧内容会原样露出来——
            #   实测「鑫路达·植物肉」本来没有数据，却显示成 1985.25 元 / 1.5 小时。
            ws.cell(r, c0, round(m, 2))
            ws.cell(r, c0 + 1, round(h, 2))
            ws.cell(r, c0 + 2, f'=IF(N({L(c0+1)}{r})=0,"",{L(c0)}{r}/{L(c0+1)}{r})')
        ws.cell(r, 2, "=" + "+".join(f"{L(5+j*3)}{r}" for j in range(len(shops))))
        ws.cell(r, 3, "=" + "+".join(f"{L(6+j*3)}{r}" for j in range(len(shops))))
        ws.cell(r, 4, f'=IF(N(C{r})=0,"",B{r}/C{r})')
        r += 1
    ws.cell(r, 1, "合计")
    for c2 in [2, 3] + [5 + j * 3 + k for j in range(len(shops)) for k in (0, 1)]:
        ws.cell(r, c2, f"=SUM({L(c2)}{first}:{L(c2)}{r-1})")
    ws.cell(r, 4, f'=IF(N(C{r})=0,"",B{r}/C{r})')
    for j in range(len(shops)):
        c0 = 5 + j * 3
        ws.cell(r, c0 + 2, f'=IF(N({L(c0+1)}{r})=0,"",{L(c0)}{r}/{L(c0+1)}{r})')
    last = r
    for rr in range(first, last + 1):
        for c2 in range(1, NC + 1):
            x = ws.cell(rr, c2)
            x.font, x.border, x.alignment = (FBB if rr == last or c2 == 1 else FB), BD, CTR
            if c2 >= 2:
                # 零显示成「—」：某家没在这个车间用工时，摆一串 0.0 反而像有数据
                x.number_format = ('#,##0.00;;"—"' if (c2 == 4 or (c2 >= 5 and (c2 - 4) % 3 == 0))
                                   else '#,##0.0;;"—"')
            if rr == last:
                x.fill = PatternFill("solid", fgColor=_FILL_PAY)

    # ---- 可直接转发的一段话（与页面第⑨步同一份计算）----
    txt = _cost_text(month_label, shops, agencies, cell)
    r = last + 2
    x = ws.cell(r, 1, "↓ 下面这几行可直接复制到群里")
    x.font = FBB
    r += 1
    for line in txt:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max(span, NC))
        x = ws.cell(r, 1, line)
        x.alignment = Alignment(horizontal="left", vertical="center")
        x.font = FB
        r += 1


@router.post("/api/tempatt/export")
async def tempatt_export(request: Request):
    """导出核对报告 Excel（带公式、异常行标底色）。
    两种来路都认：① 现场上传两张表 ② 只给 month，直接拿该期留档的结论出表——
    历史期次不该为了导一份报告再去翻原始文件。"""
    u = _require_perm(request, CAP)
    if not u:
        return JSONResponse({"ok": False, "msg": "无「临时工考勤」权限，请联系管理员"}, status_code=403)
    summary, punch, params, rates, month, _names = await _read_two(request)
    try:
        if summary and punch:
            res = _run(summary, punch, params, rates, month)
            punch_raw = punch
        else:
            res, _meta = _load_period(month)
            if not res:
                return JSONResponse(
                    {"ok": False, "msg": "请上传两张表，或选一个已留档的月份"}, status_code=400)
            # 留档路：原始打卡表从落盘的那份读回来（过了留存期就没有，⑧页自动省略）
            _, punch_raw = _period_files(res.get("month") or month)
        _apply_acks(res, res.get("month") or month)
        if not (summary and punch):
            _mark_legacy(res)
        data = _xlsx(res, res.get("month") or month, punch_raw)
    except Exception as e:
        return JSONResponse({"ok": False, "msg": f"导出失败：{e}"}, status_code=400)
    try:
        db.audit(u["name"], "临时工考勤-导出", month or "", f"{res['stats']['人数']}人")
    except Exception:
        pass
    return Response(content=data,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=temp_attendance_review.xlsx"})


# ==================== 临工看板：全年工资结构 ====================
@router.post("/api/tempatt/structure")
async def tempatt_structure(request: Request):
    """上传《临工结构》→ 解析「工资结构综合」页 → 出看板要用的几组序列。只读，不落库。"""
    u = _require_perm(request, CAP_BOARD)
    if not u:
        return JSONResponse({"ok": False, "msg": "无「临时工看板」权限，请联系管理员"}, status_code=403)
    form = await request.form()
    uf = form.get("file")
    data = await uf.read() if uf is not None and hasattr(uf, "read") else None
    if not data:
        return {"ok": False, "msg": "请上传《临工结构》Excel（含「工资结构综合」页）"}
    try:
        res = ta.parse_structure(data, str(form.get("sheet") or "") or None)
    except Exception as e:
        return {"ok": False, "msg": f"解析失败：{e}"}
    try:
        db.audit(u["name"], "临时工看板-解析结构表", res.get("sheet", ""),
                 f"{len(res['months'])}个月/{len(res['company'])}家派遣方")
    except Exception:
        pass
    return {"ok": True, **res}

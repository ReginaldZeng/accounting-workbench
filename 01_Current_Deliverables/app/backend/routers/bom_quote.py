# -*- coding: utf-8 -*-
# [Change Log]
# Date: 2026-09-02 | Author: Claude / c | Version: V-draft(BOM报价审核)
# Description: 【BOM报价审核】路由——这条工具线在后端的唯一落点（改这条线只动本文件，不碰 app.py）。
#              链路：钉钉「BOM表报价」审批附件 抓取/手工上传 → 解析预检（多产品勾选、勾稽全平才可入账）
#                    → 入账（版本粒度、排版差异版去重）→ 复核（改费用参数/渠道，逐项留痕）→ 定稿（独立指针）
#                    → BP 消费（GET final 拉定稿版分项）。
#              算法在 kernels/bom_quote.py（解析+口径+导出）与 kernels/dingtalk_bom.py（抓取）。
#              权限：录入两人(dingtalk:fetch)、复核仅会计(bom_review)、定稿仅会计(bom_finalize)。
#              数据源命名空间：跟平台 CFG.source（sample/kingdee）走——样例种子只在 sample 可见。
#              红线：勾稽不平不准静默入账；0 数据上 git（源附件落 bom_uploads/，已 gitignore）。

import hashlib
import json
import os
import shutil
import time
import uuid
from io import BytesIO

from fastapi import APIRouter, Request
from fastapi.responses import Response, HTMLResponse

from kernels import bom_quote as bq

try:
    from kernels import dingtalk_bom as dtb
except Exception:                       # 缺 requests / 没配钉钉都不该拖垮整条线
    dtb = None

from core import CFG, JSONResponse, _current_user, _require_perm, db

router = APIRouter()

CAP_FETCH = "dingtalk:fetch"            # 抓取/上传/入账（Owner + 成本会计）
CAP_AUDIT = "bom:audit"                # 审核＝复核+定稿合一（仅成本会计）
CAP_PRICE = "bom:price_check"          # 查金蝶实采价
CAP_EXPORT = "bom:export"              # 导出
CAP_ATTACH = "bom:attach_bom"          # 补挂BOM清单
CAP_VIEW_SHEET = "bom:view_sheet"      # 核算表查阅（下钻）
CAP_CONFIG = "bom:config"              # 基础设置（公开版脱敏规则等）
CAP_FINAL_REVIEW = "bom:final_review"  # 财务BP终审：①终审通过盖已审核戳（只有终审的才对外开放）②作废批准
ENTER_DRAFT = "enter:bomdraft"         # 进「待办与复核」＝看未审核（敏感）
ENTER_STD = "enter:bomstd"             # 进「标准成本台账」＝查已定稿公开

_BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
UPLOAD_DIR = os.path.join(_BASE, "bom_uploads")          # 源附件永久区 bom_uploads/{source}/{id}__{file}
STAGING_DIR = os.path.join(UPLOAD_DIR, "_staging")       # 预检暂存 bom_uploads/_staging/{sid}/
_STAGING_TTL = 6 * 3600                                  # 暂存 6 小时自动清

CH_LABELS = {"ecom": "电商", "common": "通品", "tob": "TOB", "toc": "TOC"}
# 复核四个页签（业务方定 2026-09-03）：①BOM清单 ②工艺流程 ③用量自洽 ④报价核算。
# ⚠ 但**①②只是看的、不用确认**（业务方定 2026-09-04）——它们是研发给的参考材料，不是会计要签字的判断；
#   要签字确认的只有 ③用量自洽 ④报价核算。两步确认 + 审核定性（物料类别/是否允许报价）＝可定稿（定性即定稿）。
REVIEW_STEPS = ("qty", "price")
STEP_LABELS = {"qty": "用量自洽", "price": "报价核算"}
STEP_NO = {"qty": "③", "price": "④"}


def _src():
    return CFG.get("source", "sample")


# ---------------- 台账口径投影：db entry → 前端记录 ----------------
def _fee_of(e):
    """entry 的台账费用参数；未复核过（为 None）时用源表推的默认值。"""
    src = bq.fee_from_summary(e.get("summary"))
    return {"mfg": e["fee_mfg"] if e.get("fee_mfg") is not None else src["mfg"],
            "load": e["fee_load"] if e.get("fee_load") is not None else src["load"],
            "adm": e["fee_adm"] if e.get("fee_adm") is not None else src["adm"]}


def _rec_from_entry(e):
    """把 db entry 还原成 kernel 认的 camelCase 记录（导出/compose 用）。"""
    return {"srcFile": e.get("src_file"), "sheet": e.get("sheet"), "supplier": e.get("supplier"),
            "productName": e.get("product_name"), "erpCode": e.get("erp_code"), "cpCode": e.get("cp_code"),
            "packSpec": e.get("pack_spec"), "calcDate": e.get("calc_date"), "orderQty": e.get("order_qty"),
            "customer": e.get("customer"), "matSubtotal": e.get("mat_subtotal_excl"),
            "packSubtotal": e.get("pack_subtotal_excl"), "summary": e.get("summary") or {},
            "materials": e.get("materials") or [], "checks": e.get("checks") or []}


def _upstream_status(e, finals=None, others=None):
    """已入账记录的**上游链路**：料行里引用的半成品/复配料，在台账里是什么状态。
    口径 quirk#5：下层「全成本含税」＝上层料行的「含税价」，对不上就是链路串了。

    ⚠ 只认**在台账里真能对上同名产品**的料行（业务方 2026-09-04 纠偏）。
      早先按名字关键词猜（含「半成品/复配料/复合」就当上游）**误报严重**——实证：
      「复合宝A1」是外购原料、「复配料纸箱」是包材，都被误判成「链路不通」。
      名字带这些字的外购件很常见，光靠名字分不出「未入账的半成品」和「外购件」，所以宁可不报。
      同一核算表文件内的上下游由 bq.upstream_refs 按页名精确配（那条路无歧义），入账闸走那条。
    返回 [{matName, priceUsed, found:True, entryId, upFull, priceOk, status, isFinal}]。"""
    src = e.get("source")
    finals = finals if finals is not None else db.bom_finals(src)
    by_name = {}
    for x in (others if others is not None else db.bom_list_entries(src)):
        if x["id"] == e["id"]:
            continue
        pn = (x.get("product_name") or "").strip()
        if pn:
            by_name.setdefault(pn, x)
    out = []
    for m in (e.get("materials") or []):
        nm = (m.get("matName") or "").strip()
        if not nm or m.get("seg") == "包材":       # 包材不可能是半成品
            continue
        up = by_name.get(nm)
        if not up:                                 # 台账里没有同名产品 → 就是外购料，不当上游
            continue
        comp = bq.compose(_rec_from_entry(up), _fee_of(up))
        out.append({"matName": nm, "priceUsed": m.get("priceIncl"), "found": True, "entryId": up["id"],
                    "upFull": comp["full"], "status": up.get("status"),
                    "isFinal": finals.get(up.get("product_key")) == up["id"],
                    "priceOk": (m.get("priceIncl") is not None
                                and abs((comp["full"] or 0) - float(m.get("priceIncl"))) < 0.01)})
    return out


def _upstream_block(ups):
    """定稿硬闸的理由（只在**能确定**时拦）：上游在台账里但未定稿 / 价格与上游全成本对不上。"""
    bad = []
    for u in ups:
        if not u.get("isFinal"):
            bad.append("「%s」尚未定稿（当前 %s）" % (u["matName"], u.get("status") or "未复核"))
        elif not u.get("priceOk"):
            bad.append("「%s」本品用价 %s ≠ 其全成本含税 %s" % (u["matName"], u.get("priceUsed"), round(u.get("upFull") or 0, 4)))
    return bad


def _kind_of(e):
    """记录的最终基础分类：人工定性的物料类别优先，否则编码/名字建议值。"""
    return bq.effective_kind(e.get("cp_code"), e.get("product_name"), e.get("mat_category"))


def _invalidate_review(e, fields):
    """成本被改（复核改费用/采纳商品版）后**失效审核态**：已初审/已终审 → 降回「已复核」、清终审戳(ack)、
    并**按记录id校验后**清定稿指针（若它正是本产品定稿版），强制重新初审→终审（审查 H1/M8）。
    直接往传入的 fields 里写降级字段（随后一并 update），并即时清指针。返回一句话说明（无需失效则 None）。"""
    st = e.get("status")
    if st not in ("初审", "已审核"):
        return None
    fields["status"] = "已复核"
    fields["ack"] = None
    fields["finalized_by"] = ""
    fields["finalized_at"] = ""
    db.bom_clear_final_if(e.get("source"), e.get("product_key"), e["id"])
    db.bom_clear_obsolete_by(e["id"])       # 本版退出审核态 → 它替代过的旧版恢复为当前版（换码承接 V2.440）
    return "%s → 已复核（改了成本）" % st


def _classified(e):
    """已完成审核定性？＝物料类别已指定 且 是否允许报价已决（不建议时须有原因）。定稿前置。"""
    cat = bq.cat_to_kind(e.get("mat_category"))
    q = e.get("quotable")
    if not cat or q is None:
        return False
    return bool(q) or bool((e.get("quote_reason") or "").strip())


def _net_weight(e):
    """单位净重 kg（BP 元/kg→袋/盒换算，对接 2026-09-05 §2）：存值(成本会计确认) 优先，否则按 pack_spec 预填（待确认）。
    → (kg 或 None, 'manual' | 'auto' | '')"""
    v = e.get("net_weight_kg")
    if v is not None:
        try:
            f = float(v)
            if f > 0:
                return f, "manual"
        except (TypeError, ValueError):
            pass
    g = bq.net_weight_from_spec(e.get("pack_spec"))
    return (g, "auto") if g else (None, "")


# ---- 换码承接（业务方定 2026-09-05，V2.440）----
# 规则：① CP 码正常不重复；同 CP 再来一张核算表＝同一产品重核，成本会计初审时**要跳出来问「原来那个是否失效」**。
#      ② 不同 CP 但**同 ERP 物料编码**（研发改配方换 CP、卖的还是同一 SKU）：后审核的替代先审核的；
#         BP 眼里物料编码才是身份 → 凡引用旧 CP 的定价都要重新提示 BP（对外口带 supersedes）。
# 落法：初审（定性即定稿 / finalize）与补物料编码两处检测冲突 → 前端确认 → 旧版 obsolete_by=新版（留痕双向）。
#      对外（/api/bomcost/final）只发当前版：旧版在其替代者**终审通过**那一刻退出——终审前 BP 仍拿旧版，不会出现「没成本」的空档。
def _audit_at(x):
    """一条记录的审核时刻（终审 ack.at 优先，其次初审 finalized_at）——「后审核的替代先审核的」按它比。"""
    return ((x.get("ack") or {}).get("at") or x.get("finalized_at") or "")


def _obs_brief(x, why=""):
    comp = bq.compose(_rec_from_entry(x), _fee_of(x))
    return {"entryId": x["id"], "cpCode": (x.get("cp_code") or "").strip(), "productKey": x.get("product_key"),
            "productName": (x.get("product_name") or "").strip(), "erpCode": (x.get("erp_code") or "").strip(),
            "status": x.get("status") or "", "auditAt": _audit_at(x)[:10], "fullIncl": comp["full"],
            "approvalNo": x.get("approval_no") or "", "why": why}


def _obsolete_candidates(e, others=None):
    """本版定稿会**替代失效**哪些已审核版本：同 CP，或（本版有物料编码时）同物料编码而 CP 不同。
    只认当前有效(active)、已初审/已审核、尚未失效、非本条的记录。→ [brief]（why=同CP / 同物料编码）。"""
    cp = (e.get("cp_code") or "").strip()
    erp = (e.get("erp_code") or "").strip()
    out = []
    for x in (others if others is not None else db.bom_list_entries(e.get("source"))):
        if x["id"] == e["id"] or x.get("obsolete_by") or x.get("status") not in ("初审", "已审核"):
            continue
        if cp and (x.get("cp_code") or "").strip() == cp:
            out.append(_obs_brief(x, "同CP码"))
        elif erp and (x.get("erp_code") or "").strip() == erp:
            out.append(_obs_brief(x, "同物料编码 %s" % erp))
    return out


def _mark_obsolete(e, cands, user):
    """把候选旧版标失效（obsolete_by=本版），两边都留痕（旧版记「被谁替代」、新版记「替代了谁」）。"""
    me = "#%d %s %s" % (e["id"], (e.get("cp_code") or "").strip(), (e.get("product_name") or "").strip())
    for c in cands:
        db.bom_mark_obsolete(c["entryId"], e["id"], "被 %s 替代（%s · %s）" % (me, c["why"], user))
        db.bom_add_audit(c["entryId"], user, "失效·被新版替代", c["status"], "被 %s 替代 · %s" % (me, c["why"]))
        db.bom_add_audit(e["id"], user, "替代旧版", "",
                         "#%d %s（%s · 审核 %s · 全成本 %s）→ 失效" % (c["entryId"], c["cpCode"], c["why"], c["auditAt"] or "—", c["fullIncl"]))


def _live_finals(src):
    """对外版集合：定稿指针 + 状态已审核 + **没被一条当前对外的新版替代**。
    → (served: [entry], entries: {id: entry}(active 全体))。链式（CP3 替 CP2 替 CP1）也正确：CP1 的替代者 CP2 仍是已审核指针 → CP1 退出。"""
    entries = {x["id"]: x for x in db.bom_list_entries(src)}
    live = {}
    for pkey, eid in db.bom_finals(src).items():
        e = entries.get(eid)
        if e and e.get("status") == "已审核":
            live[eid] = e
    served = [e for e in live.values() if not (e.get("obsolete_by") and e.get("obsolete_by") in live)]
    return served, entries


def _goods_view(gv):
    """成本会计商品版留档 → 前端投影（diff/是否已采纳/源文件；不外发全量 materials，明细走 diff 即可）。"""
    if not gv:
        return None
    diff = gv.get("diff") or {}
    return {"srcFile": gv.get("srcFile") or "", "srcLabel": gv.get("srcLabel") or "",
            "hasDiff": bool(diff.get("hasDiff")), "diffCount": diff.get("count", 0),
            "diffRows": diff.get("rows") or [], "applied": bool(gv.get("applied")),
            "matCount": len(gv.get("materials") or [])}


def _entry_view(e, finals):
    fee = _fee_of(e)
    comp = bq.compose(_rec_from_entry(e), fee)
    is_final = finals.get(e["product_key"]) == e["id"]
    others = db.bom_list_entries(e.get("source"))       # 同源 active 全体：上游链路 + 换码承接 共用一次查询
    ups = _upstream_status(e, finals, others)
    bom_mats = e.get("bom_list")
    bom_check = bq.compare_bom(e.get("materials") or [], bom_mats) if bom_mats else None
    nw = _net_weight(e)          # 单位净重(kg)：manual=成本会计确认 / auto=按规格预填(待确认) / ''=未填；空或≤0 不能定稿
    # 换码承接（V2.440）：obsoleteBy=本版被谁替代（live=替代者已终审→本版已退出对外；否则「待替代」仍对外）
    #                   replaces=本版替代了谁；obsoleteCandidates=本版若现在定稿会让谁失效（初审弹窗据此问「原版是否失效」）
    ob = e.get("obsolete_by")
    succ = (next((x for x in others if x["id"] == ob), None) or db.bom_get_entry(ob)) if ob else None
    return {
        "obsoleteBy": ({"entryId": succ["id"], "cpCode": (succ.get("cp_code") or "").strip(),
                        "productName": (succ.get("product_name") or "").strip(), "status": succ.get("status") or "",
                        "live": succ.get("status") == "已审核" and succ.get("active") in (1, None),
                        "at": e.get("obsolete_at") or "", "note": e.get("obsolete_note") or ""} if succ else None),
        "replaces": [_obs_brief(x) for x in others if x.get("obsolete_by") == e["id"]],
        "obsoleteCandidates": ([] if ob else _obsolete_candidates(e, others)),
        "netWeightKg": nw[0], "netWeightSrc": nw[1],
        "bomCheck": bom_check, "hasBomList": bool(bom_mats),
        # ①BOM清单 整表原样（研发出品）：类型/编码/物料/型号/规格/单位/供应商/用量（业务方 2026-09-04 定列序）
        "bomList": [{"seg": b.get("seg"), "matType": b.get("matType"), "matCode": b.get("matCode"),
                     "matName": b.get("matName"), "model": b.get("model"), "spec": b.get("spec"),
                     "unit": b.get("unit"), "brand": b.get("brand"), "qty": b.get("qty"),
                     "codeTBD": bool(bq._real_code(b) is None and (b.get("matCode") or ""))}
                    for b in (bom_mats or [])],
        "craft": e.get("craft") or None,          # BOM 文件里的工艺流程页（复核②）
        "upstream": ups, "upstreamBlock": _upstream_block(ups),   # 上游链路（半成品/复配料）状态与定稿硬闸理由
        "id": e["id"], "productKey": e["product_key"], "cpCode": e["cp_code"], "erpCode": e["erp_code"],
        "productName": (e.get("product_name") or "").strip(), "customer": e.get("customer") or "",
        "packSpec": e.get("pack_spec") or "", "supplier": e.get("supplier") or "",
        "calcDate": e.get("calc_date") or "", "orderQty": e.get("order_qty"),
        "channel": e.get("channel") or "", "channelLabel": CH_LABELS.get(e.get("channel") or "", ""),
        # 复核四步确认状态（业务方 2026-09-03 定）：①BOM清单 ②工艺流程 ③用量自洽 ④报价核算
        "steps": {k: bool((e.get("review_steps") or {}).get(k)) for k in REVIEW_STEPS},
        "stepsInfo": e.get("review_steps") or {},
        "stepsOk": all((e.get("review_steps") or {}).get(k) for k in REVIEW_STEPS),
        # 审核定性（业务方 2026-09-03）：物料类别五选一 + 是否允许对外报价（不建议须写原因）。
        # 编码判定只作建议值；kindDoubt=名字与编码打架待确认；needClassify=定稿前必须先定性。
        "kind": _kind_of(e), "semi": _kind_of(e) != "成品",
        "kindAuto": bq.classify(e.get("cp_code"), e.get("product_name")),
        "matCategory": (e.get("mat_category") or "").strip(),
        "catSuggest": bq.suggest_category(e.get("cp_code"), e.get("product_name"), e.get("supplier")),
        "outsourced": bq.is_outsourced(e.get("mat_category")),
        "quotable": (None if e.get("quotable") is None else bool(e.get("quotable"))),
        "quoteReason": e.get("quote_reason") or "",
        "classifiedBy": e.get("classified_by") or "", "classifiedAt": e.get("classified_at") or "",
        "classified": _classified(e), "needClassify": not _classified(e),
        "kindDoubt": bq.kind_doubt(e.get("cp_code"), e.get("product_name"), e.get("mat_category")),
        "matSubtotal": e.get("mat_subtotal_excl"),
        "packSubtotal": e.get("pack_subtotal_excl"), "summary": e.get("summary") or {},
        "materials": e.get("materials") or [], "checks": e.get("checks") or [],
        "fee": fee, "srcFee": comp["srcFee"], "comp": comp,
        "sourceType": e.get("source_type") or "", "approval": e.get("approval_no") or "",
        "groupId": e.get("group_id") or "", "active": e.get("active") in (1, None),
        "inactiveKind": e.get("inactive_kind") or "", "voidReq": e.get("void_req") or None,
        # 两个戳：初审（成本会计，finalized_by/at）+ 已审核（财务BP终审，ack）
        "ack": e.get("ack") or None,
        "firstPassed": e.get("status") in ("初审", "已审核"),
        "finalPassed": e.get("status") == "已审核",
        "needFinalReview": bool(e.get("status") == "初审"),   # 已初审、待财务BP终审
        "voidPending": (e.get("void_req") or {}).get("state") == "pending",
        "supersededAt": e.get("superseded_at") or "", "supersedeReason": e.get("supersede_reason") or "",
        "origin": e.get("origin") or "", "originLabel": bq.ORIGIN_LABELS.get(e.get("origin") or "", ""),
        "srcLabel": e.get("src_label") or "",
        "hasGoodsVersion": bool(e.get("goods_version")),
        "goodsVersion": _goods_view(e.get("goods_version")),
        "srcFile": e.get("src_file") or "", "sheet": e.get("sheet") or "",
        "status": e.get("status") or "未复核", "isFinal": is_final,
        "staleNote": e.get("stale_note") or "",     # 上游被替换→打回未复核时的提醒
        "createdBy": e.get("created_by") or "", "createdAt": e.get("created_at") or "",
        "reviewedBy": e.get("reviewed_by") or "", "reviewedAt": e.get("reviewed_at") or "",
        "finalizedBy": e.get("finalized_by") or "", "finalizedAt": e.get("finalized_at") or "",
    }


def _inherit_bom(e):
    """本单未附 BOM清单 时，按**产品编码**在台账历史里找带清单的同编码记录沿用。
    业务规则（业务方定）：研发改配方**必换编码**→ 同编码 = 配方未变 = 历史清单可复用。
    编码精确匹配为主、归组键兜底；返回最近一条带清单的记录（bom_list_entries 已按日期↓排），找不到 → None（疑似漏传）。"""
    src, cp, pk = e.get("source"), (e.get("cp_code") or "").strip(), e.get("product_key")
    for o in db.bom_list_entries(src):
        if o["id"] == e["id"] or not o.get("bom_list"):
            continue
        if (cp and (o.get("cp_code") or "").strip() == cp) or (o.get("product_key") == pk):
            return o
    return None


def _default_channel(rec):
    """渠道默认推断（可在复核期改）：产品名/客户含「通品」→通品；半成品→随成品(空)；其余→TOB定制。"""
    if bq.is_semi(rec.get("cpCode")):
        return ""
    blob = (rec.get("productName") or "") + (rec.get("customer") or "")
    if "通品" in blob:
        return "common"
    return "tob"


def _group_id(src, approval, anchor_pk):
    """组锚点：一个核算表文件（成品+半成品+复配料）＝一组。用 hash(源+审批号+成品归组键) 定，
    替换核算表后同一成品落回同组。审批号缺失（手工无单号）时用 anchor_pk 本身兜底。"""
    key = "%s|%s|%s" % (src, approval or "manual", anchor_pk or "")
    return "g" + hashlib.md5(key.encode("utf-8")).hexdigest()[:16]


def _workbook_anchor(records):
    """一个核算表工作簿里的「成品」归组键作组锚——非成品(半成品/复配料)不作锚；全是半成品时用第一条兜底。
    records=同一 stagedFile 的预检项列表（含 rec/productKey/semi）。"""
    fin = next((it for it in records if not it.get("semi")), None)
    return (fin or (records[0] if records else {})).get("productKey")


def _num_fp(rec, comp):
    """数字指纹：quirk#9 只吞「数字完全一致的排版差异版」。归组键+日期+全成本+小计+逐料成本 一致 = 同一版。"""
    mats = sorted(round((m.get("costExcl") or 0), 4) for m in rec.get("materials", []))
    key = "|".join([bq.product_key(rec), rec.get("calcDate") or "",
                    "%.4f" % (comp.get("srcFull") or 0), "%.4f" % (rec.get("matSubtotal") or 0),
                    "%.4f" % (rec.get("packSubtotal") or 0), ",".join("%.4f" % x for x in mats)])
    return hashlib.md5(key.encode("utf-8")).hexdigest()


# ---------------- 暂存区（预检→入账）----------------
def _purge_staging():
    try:
        now = time.time()
        if not os.path.isdir(STAGING_DIR):
            return
        for sid in os.listdir(STAGING_DIR):
            p = os.path.join(STAGING_DIR, sid)
            if os.path.isdir(p) and now - os.path.getmtime(p) > _STAGING_TTL:
                shutil.rmtree(p, ignore_errors=True)
    except Exception:
        pass


def _stage_files(files, approval_no, source_type):
    """files=[(filename, bytes, label)]（label＝钉钉控件标注，手工上传传 ""）。解析→暂存→返回预检结果。
    来源方按标注判定：商务输出/商务复核→采购商务版(procurement，全量·复核底稿)、商品版本→成本会计商品版(costacct)。
    **商品版不单独入账**：它删了型号/规格/供应商三列，只作同产品商务版底稿的「留档+价/税 diff」；
    同产品有商务版又有商品版时，商品版挂到商务版记录上；仅商品版无商务版时降级作底稿并标缺列。"""
    _purge_staging()
    sid = uuid.uuid4().hex[:16]
    sdir = os.path.join(STAGING_DIR, sid)
    os.makedirs(sdir, exist_ok=True)
    parsed, goods, warnings, bom_lists = [], [], [], []
    for entry in files:
        fname, data = entry[0], entry[1]
        label = entry[2] if len(entry) > 2 else ""
        fsrc = entry[3] if len(entry) > 3 else source_type      # 每文件渠道（评论区附件 dingtalk_comment）
        safe = "".join(ch if ch not in '\\/:*?"<>|' else "_" for ch in (fname or "attach.xlsx"))
        with open(os.path.join(sdir, safe), "wb") as fh:
            fh.write(data)
        try:
            recs = bq.parse_workbook(data, fname)
        except Exception as e:
            warnings.append("%s 解析失败：%s" % (fname, e))
            continue
        if not recs:
            # 不是成本核算表 → 试作研发 BOM清单（供核算表 vs BOM清单 自洽校验；来源方=研发BOM）
            try:
                bl = bq.parse_bom_list(data, fname)
            except Exception:
                bl = []
            if bl:
                try:            # BOM 文件里的「工艺流程」页 → 挂到本文件的每条清单上（复核②用）
                    craft = bq.parse_craft(data, fname)
                except Exception:
                    craft = None
                for b in bl:
                    b["craft"] = craft
                    b["srcFile"] = fname
                bom_lists.extend(bl)
            else:
                warnings.append("%s 里既没找到成本核算样表页、也不是 BOM清单。" % fname)
            continue
        origin = bq.origin_from_label(label, fsrc, is_bom_list=False)
        for rec in recs:
            it = {"stagedFile": safe, "rec": rec, "comp": bq.compose(rec),
                  "origin": origin, "srcLabel": label, "productKey": bq.product_key(rec)}
            (goods if origin == "costacct" else parsed).append(it)
    # 商品版配对到同产品商务版底稿；配不上（仅商品版）→ 降级入 parsed 并标缺列
    for g in goods:
        base = next((p for p in parsed if p["productKey"] == g["productKey"] and "goods" not in p), None)
        if base:
            base["goods"] = g
        else:
            g["goodsOnly"] = True
            parsed.append(g)
    # 同一核算表工作簿内的产品互为上下游（复配料→半成品→成品），先按文件聚 recs 供链路检查
    wb_recs = {}
    for it in parsed:
        wb_recs.setdefault(it.get("stagedFile"), []).append(it["rec"])
    records = []
    for it in parsed:
        rec, comp = it["rec"], it["comp"]
        it["idx"] = len(records)
        it["numFp"] = _num_fp(rec, comp)
        it["checksOk"] = bq.all_checks_ok(rec)
        it["semi"] = bq.is_semi(rec.get("cpCode"))
        it["channel"] = _default_channel(rec)
        g = it.get("goods")
        it["goodsDiff"] = bq.diff_goods(rec.get("materials"), g["rec"].get("materials")) if g else None
        # 上游链路：本品用到的半成品/复配料若**自身勾稽不平（或它的更上游不平）**，本品的投入价就来自一张算错的表 → 连带拦截。
        # ⚠ **传导式**（审查 H2）：复配料C不平 → 半成品S(自身平但用了C) → 成品P(自身平但用了S)，三级全拦。
        #   实证 251965：成品自身六项全平却用了不平半成品/复配料的价，成本少算 5.7%、整单少 8882 元。
        file_recs = wb_recs.get(it.get("stagedFile")) or []
        ups = bq.upstream_refs(rec, file_recs)
        it["upstream"] = ups
        it["blockedBy"] = bq.upstream_bad_chain(rec, file_recs)   # 递归收全部不平的上游名（含隔级）
        it["priceMismatch"] = [u for u in ups if u["upChecksOk"] and not u["priceOk"]]
        records.append(it)
    preview = {"sid": sid, "approvalNo": approval_no or "", "sourceType": source_type,
               "records": records, "bomLists": bom_lists, "warnings": warnings, "createdAt": time.time()}
    with open(os.path.join(sdir, "preview.json"), "w", encoding="utf-8") as fh:
        json.dump(preview, fh, ensure_ascii=False, default=str)
    # 前端预检卡片不需要 rec 全文，回精简投影 + 保留 rec 供入账时后端自己读 preview.json
    out = []
    for it in records:
        r = it["rec"]
        dup = db.bom_find_dup(_src(), it["numFp"])
        has_bom = bq.match_bom_list(r, bom_lists) is not None
        gd = it.get("goodsDiff")
        out.append({"idx": it["idx"], "cpCode": r.get("cpCode"), "productName": (r.get("productName") or "").strip(),
                    "erpCode": r.get("erpCode"), "customer": r.get("customer") or "", "calcDate": r.get("calcDate"),
                    "packSpec": r.get("packSpec"), "matCount": len(r.get("materials", [])),
                    "full": it["comp"]["full"], "srcFull": it["comp"]["srcFull"], "diff": it["comp"]["diff"],
                    "checksOk": it["checksOk"], "checks": r.get("checks"),
                    "failedChecks": bq.failed_checks(r) if not it["checksOk"] else [], "semi": it["semi"],
                    "channel": it["channel"], "channelLabel": CH_LABELS.get(it["channel"], ""),
                    "isDup": dup is not None, "dupId": dup["id"] if dup else None, "hasBomList": has_bom,
                    "upstream": it.get("upstream") or [], "blockedBy": it.get("blockedBy") or [],
                    "priceMismatch": it.get("priceMismatch") or [],
                    "origin": it["origin"], "originLabel": bq.ORIGIN_LABELS.get(it["origin"], ""),
                    "srcLabel": it.get("srcLabel") or "", "goodsOnly": bool(it.get("goodsOnly")),
                    "hasGoods": bool(it.get("goods")),
                    "goodsDiffCount": (gd or {}).get("count", 0), "goodsHasDiff": bool((gd or {}).get("hasDiff"))})
    return {"ok": True, "stagingId": sid, "approvalNo": approval_no or "",
            "records": out, "warnings": warnings, "bomListCount": len(bom_lists)}


def _load_staging(sid):
    p = os.path.join(STAGING_DIR, sid, "preview.json")
    if not sid or not os.path.isfile(p):
        return None
    try:
        return json.load(open(p, encoding="utf-8"))
    except Exception:
        return None


# ============ 配置 / 台账列表 / 详情 ============
@router.get("/api/bom/config")
async def bom_config(request: Request):
    u = _current_user(request)
    if not u:
        return JSONResponse({"ok": False, "msg": "未登录"}, status_code=401)
    def can(c):
        return bool(db.user_can(u, c))
    return {"ok": True, "source": _src(),
            "canFetch": can(CAP_FETCH), "canAudit": can(CAP_AUDIT), "canPrice": can(CAP_PRICE),
            "canExport": can(CAP_EXPORT), "canAttach": can(CAP_ATTACH), "canViewSheet": can(CAP_VIEW_SHEET),
            "canDraft": can(ENTER_DRAFT), "canStd": can(ENTER_STD), "canConfig": can(CAP_CONFIG),
            "canFinalReview": can(CAP_FINAL_REVIEW),
            "originLabels": bq.ORIGIN_LABELS,
            "invoiceRules": _invoice_rules(),          # ④发票类型下拉 + 改发票重算成本用
            "invoiceModes": list(bq.INVOICE_MODES),
            "dingtalkConfigured": bool(dtb and dtb.configured())}


# ---------------- 基础设置：公开版脱敏规则（第三页；暂全不遮，日后在此配）----------------
# 公开台账(标准成本台账/给BP)相比复核底稿要遮哪些敏感列。默认口径＝成本会计手工「商品版」的脱敏法：
#   删 型号(model)/规格(spec)/供应商品牌(brand) 三列。现阶段全 False＝不遮（先占位，用户日后开）。
_CFG_DEFAULT = {"hideModel": False, "hideSpec": False, "hideSupplier": False, "hidePriceNote": False}


def _bom_settings():
    s = db.get_setting("bom_config", None) or {}
    return {**_CFG_DEFAULT, **s}


@router.get("/api/bom/settings")
async def bom_get_settings(request: Request):
    u = _current_user(request)
    if not u:
        return JSONResponse({"ok": False, "msg": "未登录"}, status_code=401)
    return {"ok": True, "config": _bom_settings(), "canConfig": bool(db.user_can(u, CAP_CONFIG)),
            "hint": "默认口径＝成本会计商品版脱敏法：遮 型号/规格/供应商。现全不遮（占位），按需开启后公开台账即隐藏对应列。"}


@router.post("/api/bom/settings")
async def bom_set_settings(request: Request):
    u = _require_perm(request, CAP_CONFIG)
    if not u:
        return JSONResponse({"ok": False, "msg": "无「基础设置」权限"}, status_code=403)
    body = await request.json()
    cfg = {k: bool(body.get(k, _CFG_DEFAULT[k])) for k in _CFG_DEFAULT}
    db.set_setting("bom_config", cfg, operator=u["name"])
    db.audit(u["name"], "bom_set_config", target="-", detail=json.dumps(cfg, ensure_ascii=False))
    return {"ok": True, "config": cfg}


# ---------------- 基础设置：发票类型 → 成本不含税 算法（业务方 2026-09-04：维护到基础数据）----------------
def _invoice_rules():
    r = db.get_setting("bom_invoice_rules", None)
    if isinstance(r, list) and r:
        out = []
        for x in r:
            t = str((x or {}).get("type") or "").strip()
            mode = (x or {}).get("mode") if (x or {}).get("mode") in bq.INVOICE_MODES else "价税分离"
            if t:
                out.append({"type": t, "mode": mode, "rate": round(float((x or {}).get("rate") or 0), 4)})
        if out:
            return out
    return [dict(x) for x in bq.INVOICE_RULE_DEFAULTS]


@router.get("/api/bom/invoice-rules")
async def bom_get_invoice_rules(request: Request):
    u = _current_user(request)
    if not u:
        return JSONResponse({"ok": False, "msg": "未登录"}, status_code=401)
    return {"ok": True, "rules": _invoice_rules(), "modes": list(bq.INVOICE_MODES),
            "canConfig": bool(db.user_can(u, CAP_CONFIG)),
            "hint": "对应成本核算表 N 列公式：专票=价/(1+税率)；普票=全额；自产自销农产品=价×(1−扣除率)；农产品专票=有税率则价税分离后计算抵扣。扣除率(农产品)默认9%，可改。"}


@router.post("/api/bom/invoice-rules")
async def bom_set_invoice_rules(request: Request):
    u = _require_perm(request, CAP_CONFIG)
    if not u:
        return JSONResponse({"ok": False, "msg": "无「基础设置」权限"}, status_code=403)
    body = await request.json()
    rows = body.get("rules")
    if not isinstance(rows, list) or not rows:
        return JSONResponse({"ok": False, "msg": "请至少配一条发票规则"}, status_code=400)
    clean, seen = [], set()
    for x in rows:
        t = str((x or {}).get("type") or "").strip()
        mode = (x or {}).get("mode")
        if not t or t in seen:
            continue
        if mode not in bq.INVOICE_MODES:
            return JSONResponse({"ok": False, "msg": "「%s」算法不对，只能是：%s" % (t, "/".join(bq.INVOICE_MODES))}, status_code=400)
        seen.add(t)
        clean.append({"type": t, "mode": mode, "rate": round(max(0.0, float((x or {}).get("rate") or 0)), 4)})
    if not clean:
        return JSONResponse({"ok": False, "msg": "请至少配一条有效发票规则"}, status_code=400)
    db.set_setting("bom_invoice_rules", clean, operator=u["name"])
    db.audit(u["name"], "bom_set_invoice_rules", target="-", detail=json.dumps(clean, ensure_ascii=False))
    return {"ok": True, "rules": clean}


@router.get("/api/bom/ledger")
async def bom_ledger(request: Request):
    """台账两模式（确认书 §2.1/§8）：
    - mode=std 标准成本台账：出**已初审 + 已终审**版（终审通过才对外/供BP，准入 enter:bomstd）；
    - mode=draft 待办与复核：出**未定稿（未审核）**版，只权限人看（准入 enter:bomdraft，敏感）。
    rows=各产品当模式下的展示版；all=该模式可见的版本全集（版本对比用，不越界给草稿）；finals=定稿指针。"""
    u = _current_user(request)
    if not u:
        return JSONResponse({"ok": False, "msg": "未登录"}, status_code=401)
    mode = (request.query_params.get("mode") or "std").strip()
    gate = ENTER_STD if mode == "std" else ENTER_DRAFT
    if not db.user_can(u, gate):
        return JSONResponse({"ok": False, "msg": "无「%s」权限" % ("标准成本台账查阅" if mode == "std" else "待办与复核")}, status_code=403)
    src = _src()
    finals = db.bom_finals(src)
    entries = db.bom_list_entries(src)
    views = [_entry_view(e, finals) for e in entries]
    by_key = {}
    for v in views:
        by_key.setdefault(v["productKey"], []).append(v)
    rows, visible = [], []
    for key, vs in by_key.items():
        vs_sorted = sorted(vs, key=lambda x: (x["calcDate"], x["id"]))
        latest = vs_sorted[-1]
        fin_id = finals.get(key)
        fin = next((x for x in vs if x["id"] == fin_id), None)
        if mode == "std":
            visible += [x for x in vs if x["status"] in ("初审", "已审核")]   # 标准台账见已初审/已审核
            if fin:
                cur = dict(fin); cur["versionCount"] = sum(1 for x in vs if x["status"] in ("初审", "已审核"))
                rows.append(cur)
        else:
            visible += vs                                          # 待办与复核见全部（工作区）
            if latest["status"] not in ("初审", "已审核"):          # 最新版还没初审 = 待办
                cur = dict(latest); cur["versionCount"] = len(vs)
                cur["pending"] = latest["status"]
                rows.append(cur)
    rows.sort(key=lambda x: (x["calcDate"], x["id"]), reverse=True)
    approvals = _approval_summ(views, db.bom_pending_list(src)) if mode != "std" else []
    return {"ok": True, "source": src, "mode": mode, "rows": rows, "all": visible,
            "finals": finals, "approvals": approvals,
            # 标准台账：已初审、待财务BP终审的条数（BP 要拿这数报价，得知道哪些还没终审、还没对外）
            "needAck": sum(1 for r in rows if r.get("needFinalReview")),
            "canFinalReview": bool(db.user_can(u, CAP_FINAL_REVIEW))}


def _approval_summ(views, pendings=None):
    """待办按钉钉单号汇总（列表用）：每单号 组数/产品数/日期/待复核数。无单号归「手工」。倒序。
    ⚠ 还要并入「整组被拦、一条没入账」的待修批次——否则那种单在台账里查无记录，会从待办里整单消失。"""
    by_ap = {}
    for v in views:
        by_ap.setdefault(v.get("approval") or "", []).append(v)
    out = []
    for ap, vs in by_ap.items():
        groups = set(v.get("groupId") or v["id"] for v in vs)
        dates = sorted([v["calcDate"] for v in vs if v.get("calcDate")])
        out.append({"approvalNo": ap, "groupCount": len(groups), "productCount": len(vs),
                    "date": dates[-1] if dates else "",
                    "pending": sum(1 for v in vs if v["status"] not in ("初审", "已审核")),
                    "finalized": sum(1 for v in vs if v["status"] in ("初审", "已审核")), "blocked": 0,
                    "products": [{"cpCode": v["cpCode"], "productName": v["productName"], "kind": v["kind"]} for v in vs]})
    ix = {o["approvalNo"]: o for o in out}
    for p in (pendings or []):
        ap = p.get("approval_no") or ""
        rs = p.get("reasons") or []
        o = ix.get(ap)
        if not o:
            o = {"approvalNo": ap, "groupCount": 0, "productCount": 0, "date": p.get("created_at", "")[:10],
                 "pending": 0, "finalized": 0, "blocked": 0, "products": []}
            out.append(o); ix[ap] = o
        o["groupCount"] += 1
        o["blocked"] += len(rs)
        o["products"] += [{"cpCode": r.get("cpCode"), "productName": r.get("productName"), "kind": "待修"} for r in rs]
    out.sort(key=lambda x: x["date"] or "", reverse=True)
    return out


@router.get("/api/bom/entry/{entry_id}")
async def bom_entry(request: Request, entry_id: int):
    u = _current_user(request)
    if not u:
        return JSONResponse({"ok": False, "msg": "未登录"}, status_code=401)
    e = db.bom_get_entry(entry_id)
    if not e or e.get("source") != _src():
        return JSONResponse({"ok": False, "msg": "记录不存在"}, status_code=404)
    # 可见性（确认书 §8）：已定稿→需核算表查阅或进标准台账；未定稿草稿→需 enter:bomdraft（未审核只权限人看）
    if e.get("status") in ("初审", "已审核", "已定稿"):
        if not (db.user_can(u, CAP_VIEW_SHEET) or db.user_can(u, ENTER_STD)):
            return JSONResponse({"ok": False, "msg": "无「核算表查阅」权限"}, status_code=403)
    elif not db.user_can(u, ENTER_DRAFT):
        return JSONResponse({"ok": False, "msg": "该记录未审核，仅有「待办与复核」权限者可查看"}, status_code=403)
    finals = db.bom_finals(_src())
    v = _entry_view(e, finals)
    v["audits"] = db.bom_audits(entry_id)
    # 本单未附BOM清单 → 按产品编码回查台账历史沿用（改配方必换编码，同编码=配方未变）
    if not v["hasBomList"]:
        src_e = _inherit_bom(e)
        if src_e:
            v["bomCheck"] = bq.compare_bom(e.get("materials") or [], src_e["bom_list"])
            v["bomInherited"] = {"fromEntryId": src_e["id"], "fromCp": src_e.get("cp_code"),
                                 "fromApproval": src_e.get("approval_no"), "fromDate": src_e.get("calc_date"),
                                 "fromProduct": (src_e.get("product_name") or "").strip()}
        else:
            v["bomInherited"] = None      # 编码无历史 → 疑似漏传
    # 同产品全部版本（版本对比/涨跌标）
    sibs = db.bom_list_entries(_src(), e["product_key"])
    v["versions"] = [_entry_view(s, finals) for s in sibs]
    return {"ok": True, "entry": v}


def _src_path(src, entry_id):
    """某记录的源核算表留档路径（跳过商品版留档）。找不到 → None。"""
    pdir = os.path.join(UPLOAD_DIR, src)
    if not os.path.isdir(pdir):
        return None
    for fn in os.listdir(pdir):
        if fn.startswith("%d__" % entry_id) and "__商品版__" not in fn:
            return os.path.join(pdir, fn)
    return None


def _group_file(src, entries, group_id=None, approval_no=None):
    """该组的核算表源文件：先找已入账记录的留档，没有就找「待修」批次的暂存（整组被拦时用）。"""
    for e in entries or []:
        p = _src_path(src, e["id"])
        if p:
            return p
    if group_id:
        for pd in db.bom_pending_list(src, approval_no):
            if pd.get("group_id") == group_id and pd.get("stash_path") and os.path.isfile(pd["stash_path"]):
                return pd["stash_path"]
    return None


def _group_roster(src, entries, booked_views, exclude_pks=None, group_id=None, approval_no=None):
    """组内**全量产品名册**：重解析该组核算表留档，把「勾稽不平·未入账」的产品也列出来。
    否则成本会计只看到入账成功的那一个（如只有成品），看不出复配料/半成品差在哪、没法找上游改（业务方 2026-09-03 提）。
    返回 [已入账的 entry view … , {notBooked:True, 诊断…} …]；文件找不到/解析失败 → 只回已入账的。"""
    path = _group_file(src, entries, group_id, approval_no)
    if not path:
        return booked_views
    try:
        recs = bq.parse_workbook(open(path, "rb").read(), os.path.basename(path))
    except Exception:
        return booked_views
    # 排除本组已入账的；也排除**本审批下别的组**已入账的——同一核算表文件被拆到多组时（如样例种子按产品分组），
    # 否则会把别组入账成功的产品误报成「未入账」。
    booked_pk = {v["productKey"] for v in booked_views} | set(exclude_pks or ())
    extra = []
    for rec in recs:
        pk = bq.product_key(rec)
        if pk in booked_pk:
            continue
        comp = bq.compose(rec)
        ups = bq.upstream_refs(rec, recs)
        blocked = bq.upstream_bad_chain(rec, recs)   # 传导式（审查 H2）
        extra.append({"notBooked": True, "id": None, "productKey": pk,
                      "blockedBy": blocked,   # 自身平但上游不平 → 连带拦下（成本建在错数上）
                      "productName": (rec.get("productName") or "").strip(), "cpCode": rec.get("cpCode"),
                      "sheet": rec.get("sheet"), "calcDate": rec.get("calcDate"),
                      "kind": bq.classify(rec.get("cpCode"), rec.get("productName")),
                      "kindAuto": bq.classify(rec.get("cpCode"), rec.get("productName")),
                      "matCategory": "", "comp": comp, "checksOk": bq.all_checks_ok(rec),
                      "failedChecks": bq.failed_checks(rec), "matCount": len(rec.get("materials") or [])})
    out = list(booked_views) + extra
    # **按依赖深度自下而上排**（业务方定 2026-09-04）：审核顺序＝先复配料、再半成品、最后成品。
    # 深度由「谁引用谁」算：没有上游的是 0 层（最底），引用了 n 层的是 n+1 层。同时带出 uses/usedBy 供画结构。
    by_name = {bq.norm(r.get("productName")): r for r in recs}
    pk_of = {bq.norm(r.get("productName")): bq.product_key(r) for r in recs}
    uses = {}
    for r in recs:
        uses[bq.product_key(r)] = [bq.norm(u["upProductName"]) for u in bq.upstream_refs(r, recs)]

    def depth(pk, seen=()):
        if pk in seen:
            return 0                      # 防环
        ups = [pk_of.get(n) for n in uses.get(pk, []) if pk_of.get(n)]
        return 1 + max([depth(u, seen + (pk,)) for u in ups], default=-1)

    for x in out:
        pk = x.get("productKey")
        x["depth"] = depth(pk) if pk in uses else 0
        x["uses"] = uses.get(pk, [])
        x["usedBy"] = [bq.norm(r.get("productName")) for r in recs
                       if bq.norm(x.get("productName")) in uses.get(bq.product_key(r), [])]
        # **建议分类改用组内 BOM 结构**（业务方 2026-09-04 纠偏，编码把 251965 的成品/半成品判反了）：
        # 只覆盖「建议值」，人工已定性(mat_category)的一律不动。审核顺序链条/组内列/筛选都读 kind，一处改全好。
        if not (x.get("matCategory") or "").strip():
            gk = bq.group_kind(x.get("productName"), x.get("uses"), x.get("usedBy"))
            x["kind"] = gk
            x["kindAuto"] = gk
            x["semi"] = gk != "成品"
            x["catSuggest"] = bq.group_category(x.get("productName"), x.get("uses"),
                                                x.get("usedBy"), x.get("supplier") or "")
            x["kindDoubt"] = False        # 结构定的建议，不再是编码打架的存疑
    out.sort(key=lambda x: (x.get("depth", 0), 1 if x.get("notBooked") else 0, x.get("productName") or ""))
    return out


def _with_bomcheck(e, finals):
    """entry view + 本单未附清单时按编码沿用历史校验（与详情同口径），供处理页组内产品用。"""
    v = _entry_view(e, finals)
    if not v["hasBomList"]:
        se = _inherit_bom(e)
        if se:
            v["bomCheck"] = bq.compare_bom(e.get("materials") or [], se["bom_list"])
            v["bomInherited"] = {"fromEntryId": se["id"], "fromCp": se.get("cp_code"),
                                 "fromApproval": se.get("approval_no"), "fromDate": se.get("calc_date"),
                                 "fromProduct": (se.get("product_name") or "").strip()}
        else:
            v["bomInherited"] = None
    return v


@router.get("/api/bom/approval")
async def bom_approval(request: Request):
    """处理页（业务方定 A 方案）：一个钉钉单号 → 若干「组」（一个核算表文件=一组）。
    每组：当前版产品（成品/半成品/复配料，含勾稽/用量校验/来源方/商品版）+ 被替换旧版留痕（审核历史）。"""
    u = _current_user(request)
    if not u:
        return JSONResponse({"ok": False, "msg": "未登录"}, status_code=401)
    if not db.user_can(u, ENTER_DRAFT):
        return JSONResponse({"ok": False, "msg": "无「待办与复核」权限"}, status_code=403)
    no = (request.query_params.get("no") or "").strip()
    src = _src()
    finals = db.bom_finals(src)
    ents = [e for e in db.bom_list_entries(src, include_superseded=True) if (e.get("approval_no") or "") == no]
    gmap = {}
    for e in ents:
        gmap.setdefault(e.get("group_id") or ("solo%d" % e["id"]), []).append(e)
    # 「待修」组：整组被拦、一条也没入账 → 台账里没有它，但必须在处理页看得见、能替换修复
    pend = {p["group_id"]: p for p in db.bom_pending_list(src, no)}
    for gid in pend:
        gmap.setdefault(gid, [])
    all_booked_pk = {e.get("product_key") for e in ents if e.get("active") in (1, None)}
    groups = []
    for gid, es in gmap.items():
        active = [x for x in es if x.get("active") in (1, None)]
        superseded = [x for x in es if x.get("active") == 0]
        pd = pend.get(gid)
        if not active and not superseded and not pd:
            continue
        anchor = next((x for x in active if not x.get("semi")), (active or es or [{}])[0])
        pv = [_with_bomcheck(x, finals) for x in sorted(active, key=lambda a: (0 if not a.get("semi") else 1, a["id"]))]
        # 组内全量名册：把同一核算表里「不平未入账」的产品也列出来（会计要看到差异在哪、好找上游改）
        pv = _group_roster(src, active + superseded, pv, exclude_pks=all_booked_pk,
                           group_id=gid, approval_no=no)
        hist = [{"id": x["id"], "productName": (x.get("product_name") or "").strip(), "cpCode": x.get("cp_code"),
                 "srcFile": x.get("src_file"), "supersededAt": x.get("superseded_at"),
                 "reason": x.get("supersede_reason"), "createdAt": x.get("created_at"),
                 "createdBy": x.get("created_by")} for x in sorted(superseded, key=lambda a: a["id"])]
        booked = [v for v in pv if not v.get("notBooked")]
        # 整组待修（无任何入账）时，用名册里最深那层（成品）当组名、用待修批次的文件名与日期
        core = next((v for v in reversed(pv) if not v.get("notBooked")), None) or (pv[-1] if pv else {})
        groups.append({"groupId": gid,
                       "coreName": (anchor.get("product_name") or "").strip() or core.get("productName") or "（待修）",
                       "coreCp": anchor.get("cp_code") or core.get("cpCode") or "",
                       "coreFile": anchor.get("src_file") or (pd or {}).get("src_file") or "",
                       "allPending": not active and bool(pd),
                       # 纯历史组：无 active、无待修，只剩被替换/作废的旧记录 → 处理页不当编号「组」，收进「历史留痕」块
                       "historyOnly": (not active) and (not pd),
                       "pendingAt": (pd or {}).get("created_at") or "", "pendingBy": (pd or {}).get("created_by") or "",
                       "date": anchor.get("calc_date") or core.get("calcDate") or "",
                       "products": pv, "superseded": hist,
                       "bookedCount": len(booked), "pendingCount": sum(1 for v in pv if v.get("notBooked")),
                       # 组算齐＝全部产品都入账了（无不平待修）且各自四步已确认/已定稿
                       "allOk": bool(booked) and not any(v.get("notBooked") for v in pv)
                                and all(v["stepsOk"] or v["isFinal"] for v in booked),
                       "anyFinal": any(v["isFinal"] for v in booked),
                       "voidPendingCount": sum(1 for v in booked if v.get("voidPending")),
                       "pendVoid": (pd or {}).get("void_req")})
    groups.sort(key=lambda g: g["date"] or "", reverse=True)
    def can(c):
        return bool(db.user_can(u, c))
    return {"ok": True, "approvalNo": no, "groups": groups, "source": src,
            "canAudit": can(CAP_AUDIT), "canFetch": can(CAP_FETCH), "canAttach": can(CAP_ATTACH),
            "canFinalReview": can(CAP_FINAL_REVIEW),
            "dingtalkConfigured": bool(dtb and dtb.configured())}


@router.get("/api/bom/pending")
async def bom_pending_detail(request: Request):
    """下钻看「勾稽不平·未入账」的产品**到底哪儿不对**（业务方 2026-09-03 提）：
    重解析该组核算表留档，回该产品的逐料明细 + 6 项勾稽 + 不平诊断，
    并标出**没被计进申报小计的那几味料**（missingNames），前端把这些行标红。"""
    u = _current_user(request)
    if not u:
        return JSONResponse({"ok": False, "msg": "未登录"}, status_code=401)
    if not db.user_can(u, ENTER_DRAFT):
        return JSONResponse({"ok": False, "msg": "无「待办与复核」权限"}, status_code=403)
    gid = (request.query_params.get("groupId") or "").strip()
    pkey = (request.query_params.get("productKey") or "").strip()
    appno = (request.query_params.get("no") or "").strip() or None
    src = _src()
    ents = db.bom_group_entries(src, gid)
    path = _group_file(src, ents, gid, appno)     # 已入账留档 → 没有就用待修批次的暂存
    if not path:
        return JSONResponse({"ok": False, "msg": "该组的核算表源文件未留档，无法下钻。"}, status_code=404)
    try:
        recs = bq.parse_workbook(open(path, "rb").read(), os.path.basename(path))
    except Exception as ex:
        return JSONResponse({"ok": False, "msg": "源文件解析失败：%s" % ex}, status_code=400)
    rec = next((r for r in recs if bq.product_key(r) == pkey), None)
    if not rec:
        return JSONResponse({"ok": False, "msg": "源文件里找不到该产品。"}, status_code=404)
    fails = bq.failed_checks(rec)
    miss = []
    for c in fails:
        miss += [m.get("matName") for m in (c.get("missing") or [])]
    comp = bq.compose(rec)
    return {"ok": True, "srcFile": os.path.basename(path).split("__", 1)[-1],
            "product": {"productName": (rec.get("productName") or "").strip(), "cpCode": rec.get("cpCode"),
                        "sheet": rec.get("sheet"), "calcDate": rec.get("calcDate"),
                        "supplier": rec.get("supplier"), "packSpec": rec.get("packSpec"),
                        "matSubtotal": rec.get("matSubtotal"), "packSubtotal": rec.get("packSubtotal"),
                        "materials": rec.get("materials") or [], "checks": rec.get("checks") or [],
                        "summary": rec.get("summary") or {}, "comp": comp,
                        "failedChecks": fails, "missingNames": [m for m in miss if m]}}


# ============ 价格校验（阶段②）：金蝶应付实采 + BOM反查 ============
@router.get("/api/bom/kd-purchase")
async def bom_kd_purchase(request: Request):
    """按物料编码查金蝶近期已审核采购价（应付单为主）。金蝶连不上/未配置 → offline，前端提示需服务器联调。"""
    u = _require_perm(request, CAP_PRICE)
    if not u:
        return JSONResponse({"ok": False, "msg": "无「查金蝶实采价」权限"}, status_code=403)
    code = (request.query_params.get("code") or "").strip()
    months = int(request.query_params.get("months") or 12)
    if not code:
        return JSONResponse({"ok": False, "msg": "缺物料编码"}, status_code=400)
    try:
        import kingdee_client as kc
        rows = kc.fetch_material_prices(code, months=months)
        return {"ok": True, "code": code, "months": months, "rows": rows}
    except Exception as e:
        # worktree 连不上金蝶属正常（凭据/网络在服务器）——回 offline，前端友好提示，不当报错
        return {"ok": True, "offline": True, "code": code, "rows": [],
                "msg": "金蝶未连接或字段待联调：%s" % str(e)[:160]}


# ---- 按 CP 码到金蝶物料档案反查物料编码（业务方 2026-09-05 定：「检测到就提示成本会计确认」，不自动写）----
# 金蝶物料档案自定义字段「研发编码」(kingdee_client.MATERIAL_RD_CODE_FIELD) 登的就是 CP 码；一个 CP 可能挂多个物料编码
# （正式码 + T 开头临时码并存）→ 只出候选让成本会计选。10 分钟进程内缓存，避免每次开页都打金蝶。
_ERP_LOOKUP_CACHE = {}
_ERP_LOOKUP_TTL = 600


def _erp_lookup(cp):
    import kingdee_client as kc
    key = kc.normalize_rd_code(cp)
    hit = _ERP_LOOKUP_CACHE.get(key)
    if hit and time.time() - hit[0] < _ERP_LOOKUP_TTL:
        return hit[1]
    res = kc.fetch_materials_by_rd_code(cp)
    _ERP_LOOKUP_CACHE[key] = (time.time(), res)
    return res


@router.get("/api/bom/erp-lookup")
async def bom_erp_lookup(request: Request):
    """按记录的 CP 码查金蝶物料档案「研发编码」字段，回候选物料编码供成本会计确认（补物料编码 / 初审弹窗 / 右栏都用它）。
    每个候选附 `inLedger`＝台账里已挂该编码的其它记录（选它会触发换码承接的「后审核替代先审核」确认）。
    金蝶连不上 → offline（本机 worktree 常态），前端提示手填。"""
    u = _require_perm(request, CAP_AUDIT)
    if not u:
        return JSONResponse({"ok": False, "msg": "无「审核」权限"}, status_code=403)
    eid = request.query_params.get("entryId")
    cp = (request.query_params.get("cp") or "").strip()
    e = db.bom_get_entry(eid) if eid else None
    if eid and (not e or e.get("source") != _src()):
        return JSONResponse({"ok": False, "msg": "记录不存在"}, status_code=404)
    if e:
        cp = (e.get("cp_code") or "").strip()
    if not cp:
        return {"ok": True, "cp": "", "candidates": [], "msg": "本记录没有 CP 码，无法反查"}
    try:
        cands = _erp_lookup(cp)
    except Exception as ex:
        return {"ok": True, "offline": True, "cp": cp, "candidates": [], "msg": "金蝶未连接或字段待联调：%s" % str(ex)[:160]}
    if cands:
        others = [x for x in db.bom_list_entries(_src()) if not (e and x["id"] == e["id"])]
        for c in cands:
            c["inLedger"] = [{"entryId": x["id"], "cpCode": (x.get("cp_code") or "").strip(), "status": x.get("status") or "",
                              "productName": (x.get("product_name") or "").strip()}
                             for x in others if (x.get("erp_code") or "").strip() == c["erpCode"]]
    return {"ok": True, "cp": cp, "current": (e.get("erp_code") or "").strip() if e else "", "candidates": cands}


@router.get("/api/bom/material-usage")
async def bom_material_usage(request: Request):
    """BOM反查：本数据源下，其他记录用同一物料编码时研发填的含税价（看研发跨产品定价是否一致）。
    ⚠ 会返回**未审核草稿**里的研发采购价/税率——须「看未审核」准入(ENTER_DRAFT)，否则任何登录用户越权可读（审查 H6）。"""
    u = _require_perm(request, ENTER_DRAFT)
    if not u:
        return JSONResponse({"ok": False, "msg": "无「待办与复核」权限（BOM反查会带出未审核价，仅本组可查）"}, status_code=403)
    code = (request.query_params.get("code") or "").strip()
    exclude = request.query_params.get("exclude")
    if not code:
        return JSONResponse({"ok": False, "msg": "缺物料编码"}, status_code=400)
    out = []
    for e in db.bom_list_entries(_src()):
        if exclude and str(e["id"]) == str(exclude):
            continue
        for m in (e.get("materials") or []):
            if str(m.get("matCode") or "").strip() == code:
                out.append({"entryId": e["id"], "productName": (e.get("product_name") or "").strip(),
                            "cpCode": e.get("cp_code"), "calcDate": e.get("calc_date"),
                            "matName": m.get("matName"), "priceIncl": m.get("priceIncl"),
                            "taxRate": m.get("taxRate"), "qtyPerKg": m.get("qtyPerKg"),
                            "brand": m.get("brand"), "seg": m.get("seg"),
                            "priceNote": m.get("priceNote")})
    out.sort(key=lambda r: (r["calcDate"] or "", r["entryId"]), reverse=True)
    return {"ok": True, "code": code, "rows": out}


# 原料内部子类（业务方 2026-09-04）：④报价可把误判的复配料改回原辅料等；**只在原料内部改**（原辅料/复配料/自产半成品），
# 不动原料/包材归属（那按采购核算表为准），故五分项小计不变——只影响是否下钻子核算表/上游链，不重算成本、不失效审核。
MAT_SUBTYPES = ("原辅料", "复配料", "自产半成品")


@router.post("/api/bom/set-mat-type")
async def bom_set_mat_type(request: Request):
    u = _require_perm(request, CAP_AUDIT)
    if not u:
        return JSONResponse({"ok": False, "msg": "无「复核」权限"}, status_code=403)
    body = await request.json()
    e = db.bom_get_entry(body.get("entryId"))
    if not e or e.get("source") != _src():
        return JSONResponse({"ok": False, "msg": "记录不存在"}, status_code=404)
    sub = str(body.get("subType") or "").strip()
    if sub not in MAT_SUBTYPES:
        return JSONResponse({"ok": False, "msg": "物料子类只能是：%s" % "/".join(MAT_SUBTYPES)}, status_code=400)
    code = str(body.get("matCode") or "").strip()
    name = str(body.get("matName") or "").strip()
    seg = str(body.get("seg") or "").strip()
    mats = e.get("materials") or []
    hit = next((m for m in mats if (str(m.get("matCode") or "").strip() == code
                                    and (m.get("matName") or "").strip() == name
                                    and (m.get("seg") or "") == seg)), None)
    if hit is None:
        return JSONResponse({"ok": False, "msg": "未找到该物料行"}, status_code=404)
    if (hit.get("seg") or "") == "包材":
        return JSONResponse({"ok": False, "msg": "只改原料内部子类，包材不在此改"}, status_code=400)
    old = (hit.get("subType") or "").strip() or "（自动判定）"
    hit["subType"] = sub
    db.bom_update_entry(e["id"], {"materials": mats})
    db.bom_add_audit(e["id"], u["name"], "物料子类·" + name, old, sub)
    return {"ok": True, "entry": _entry_view(db.bom_get_entry(e["id"]), db.bom_finals(_src()))}


@router.post("/api/bom/set-erp-code")
async def bom_set_erp_code(request: Request):
    """补/改产品的 ERP 物料编码（钉钉解析常缺此码，成本会计在台账行手工补录）。
    只动标识不动成本；改动逐条留痕（谁、原值→新值），已审核记录也可补——ERP 码是身份不是金额。"""
    u = _require_perm(request, CAP_AUDIT)
    if not u:
        return JSONResponse({"ok": False, "msg": "无「复核」权限"}, status_code=403)
    body = await request.json()
    e = db.bom_get_entry(body.get("entryId"))
    if not e or e.get("source") != _src():
        return JSONResponse({"ok": False, "msg": "记录不存在"}, status_code=404)
    code = str(body.get("erpCode") or "").strip()
    old = (e.get("erp_code") or "").strip()
    if code == old:
        return {"ok": True, "entry": _entry_view(db.bom_get_entry(e["id"]), db.bom_finals(_src()))}
    # 换码承接（V2.440）：给**已审核**的行补物料编码时，若别的 CP 已挂同一编码且也在审核态 →
    # 「后审核的替代先审核的」——先问确认（confirmObsolete），确认后写码并把先审核的标失效。未审核的草稿行留到初审时再判。
    clash = []
    if code and e.get("status") in ("初审", "已审核"):
        clash = [x for x in db.bom_list_entries(_src())
                 if x["id"] != e["id"] and not x.get("obsolete_by") and x.get("status") in ("初审", "已审核")
                 and (x.get("erp_code") or "").strip() == code]
    if clash and not bool(body.get("confirmObsolete")):
        ranked = sorted(clash + [e], key=lambda x: (_audit_at(x), x["id"]))
        newest = ranked[-1]
        losers = [_obs_brief(x, "同物料编码 %s" % code) for x in ranked[:-1]]
        return {"ok": False, "needConfirm": losers, "newest": _obs_brief(newest),
                "msg": "物料编码 %s 已挂在 %s。按规则「后审核的替代先审核的」：%s 为当前版，%s 将失效退出对外台账（BP 引用它的定价会收到成本更新提示）。确认？"
                       % (code, "、".join("%s（审核 %s）" % ((x.get("cp_code") or "").strip(), _audit_at(x)[:10] or "—") for x in clash),
                          (newest.get("cp_code") or "").strip(),
                          "、".join(c["cpCode"] for c in losers))}
    db.bom_update_entry(e["id"], {"erp_code": code})
    db.bom_add_audit(e["id"], u["name"], "补物料编码", old or "（空）", code or "（清空）")
    if clash:
        e2 = db.bom_get_entry(e["id"])
        ranked = sorted(clash + [e2], key=lambda x: (_audit_at(x), x["id"]))
        newest, losers = ranked[-1], ranked[:-1]
        _mark_obsolete(newest, [_obs_brief(x, "同物料编码 %s" % code) for x in losers], u["name"])
    return {"ok": True, "entry": _entry_view(db.bom_get_entry(e["id"]), db.bom_finals(_src()))}


@router.post("/api/bom/set-net-weight")
async def bom_set_net_weight(request: Request):
    """成本会计填/确认**单位净重(kg)**（BP 对接 2026-09-05 §2）：一个销售单位(袋/盒)的净重，BP 元/kg→元/袋 换算用。
    按规格预填的值是「待确认」，确认即存库；空或≤0 不能定稿。改动逐条留痕。"""
    u = _require_perm(request, CAP_AUDIT)
    if not u:
        return JSONResponse({"ok": False, "msg": "无「复核」权限"}, status_code=403)
    body = await request.json()
    e = db.bom_get_entry(body.get("entryId"))
    if not e or e.get("source") != _src():
        return JSONResponse({"ok": False, "msg": "记录不存在"}, status_code=404)
    try:
        kg = float(body.get("netWeightKg"))
    except (TypeError, ValueError):
        return JSONResponse({"ok": False, "msg": "单位净重须为数字（kg）"}, status_code=400)
    if not kg > 0:
        return JSONResponse({"ok": False, "msg": "单位净重须大于 0（kg）"}, status_code=400)
    kg = round(kg, 4)
    old_v, old_src = _net_weight(e)
    db.bom_update_entry(e["id"], {"net_weight_kg": kg})
    db.bom_add_audit(e["id"], u["name"], "单位净重(kg)",
                     ("%s（%s）" % (old_v, "按规格预填" if old_src == "auto" else "已确认")) if old_v else "（未填）", str(kg))
    return {"ok": True, "entry": _entry_view(db.bom_get_entry(e["id"]), db.bom_finals(_src()))}


@router.get("/api/bom/usage-spreads")
async def bom_usage_spreads(request: Request):
    """BOM反查·批量价差（④报价行上红点用）：本记录每个**真实编码**物料，在本数据源全部记录里研发填的含税价
    最高/最低差多少。差异大＝研发跨产品同料定价不一致，值得点开核价看。占位码(XX系列)/空码跳过（会撞不同料）。"""
    u = _require_perm(request, ENTER_DRAFT)
    if not u:
        return JSONResponse({"ok": False, "msg": "无「待办与复核」权限"}, status_code=403)
    e = db.bom_get_entry(request.query_params.get("entryId"))
    if not e or e.get("source") != _src():
        return JSONResponse({"ok": False, "msg": "记录不存在"}, status_code=404)
    my_codes = {c for c in (bq._real_code(m) for m in (e.get("materials") or [])) if c}
    if not my_codes:
        return {"ok": True, "spreads": {}}
    prices = {c: [] for c in my_codes}
    for ent in db.bom_list_entries(_src()):
        for m in (ent.get("materials") or []):
            c = bq._real_code(m)
            if c in prices and m.get("priceIncl"):
                prices[c].append(float(m.get("priceIncl")))
    spreads = {}
    for c, arr in prices.items():
        if len(arr) >= 2:
            lo, hi = min(arr), max(arr)
            spreads[c] = {"min": round(lo, 4), "max": round(hi, 4),
                          "spread": round((hi - lo) / lo, 4) if lo else 0.0, "count": len(arr)}
    return {"ok": True, "spreads": spreads}


# ============ 抓取 / 上传 / 入账 ============
@router.post("/api/bom/fetch-approval")
async def bom_fetch(request: Request):
    u = _require_perm(request, CAP_FETCH)
    if not u:
        return JSONResponse({"ok": False, "msg": "无「抓取/录入」权限"}, status_code=403)
    body = await request.json()
    appno = str(body.get("approvalNo") or "").strip()
    if not appno:
        return JSONResponse({"ok": False, "msg": "请填审批编号"}, status_code=400)
    if not (dtb and dtb.configured()):
        return JSONResponse({"ok": False, "msg": "未配置钉钉应用——请在服务器 conf.ini [dingtalk] 配 appkey/appsecret 后再取数。"}, status_code=400)
    res = dtb.fetch_approval(appno)
    if not res.get("ok"):
        return JSONResponse({"ok": False, "msg": res.get("msg") or "取数失败"}, status_code=400)
    files, skipped, comment_pending = [], [], []
    for a in res.get("attachments", []):
        if a.get("bytes") and str(a.get("fileName") or "").lower().endswith((".xlsx", ".xls")):
            # 带上钉钉控件标注(label)+渠道(source)，供来源方判定：商务输出→采购商务版、商品版本→成本会计商品版
            files.append((a["fileName"], a["bytes"], a.get("label") or "", a.get("source") or "dingtalk_form"))
        elif a.get("source") == "dingtalk_comment":
            # 评论区补传的附件：钉钉把评论区文件存在钉盘/IM 空间，当前应用接口权限够不着（返「无访问权限」）。
            # 不静默丢——显性抛给前端：知道有补传、取不到，请手工下载后用「上传」补入（或后台给应用补授评论区文件下载权限）。
            comment_pending.append({"fileName": a.get("fileName"), "reason": a.get("error") or "评论区文件当前权限取不到"})
        elif a.get("error"):
            skipped.append("%s：%s" % (a.get("fileName"), a["error"]))
    if not files:
        msg = "该审批未取到可解析的 xlsx 表单附件。"
        if comment_pending:
            msg += "另有 %d 个评论区补传附件当前钉钉权限取不到，请手工下载后上传。" % len(comment_pending)
        return JSONResponse({"ok": False, "msg": msg + ("；".join(skipped) if skipped else ""),
                             "commentPending": comment_pending}, status_code=400)
    out = _stage_files(files, appno, "dingtalk_form")
    out["title"] = res.get("title")
    out["skipped"] = skipped
    out["commentPending"] = comment_pending   # 评论区补传但取不到 → 前端提示手工上传
    db.audit(u["name"], "bom_fetch", target=appno, detail="附件 %d 个，评论区待补 %d" % (len(files), len(comment_pending)))
    return out


@router.post("/api/bom/upload")
async def bom_upload(request: Request):
    u = _require_perm(request, CAP_FETCH)
    if not u:
        return JSONResponse({"ok": False, "msg": "无「抓取/录入」权限"}, status_code=403)
    form = await request.form()
    appno = str(form.get("approvalNo") or "").strip()
    files = []
    for key, val in form.multi_items():
        if hasattr(val, "read"):
            data = await val.read()
            if data:
                files.append((getattr(val, "filename", "attach.xlsx"), data))
    if not files:
        return JSONResponse({"ok": False, "msg": "请上传至少一个 xlsx 文件"}, status_code=400)
    out = _stage_files(files, appno, "manual_upload")
    db.audit(u["name"], "bom_upload", target=appno or "-", detail="%d 个文件" % len(files))
    return out


def _book_staged(prev, sid, idxs, u):
    """入账核心（抽出来给「立项」与「入账」共用）。勾稽不平/上游不平的一律拒（红线）；
    排版差异版（数字指纹重复）跳过；整组一条没入 → 记「待修」批次，留在待办里可见可修。"""
    src = _src()
    appno = prev.get("approvalNo") or ""
    # 组锚：同一核算表文件(stagedFile)的产品共用 group_id（成品归组键作锚，替换后落回同组）
    by_wb = {}
    for it in prev["records"]:
        by_wb.setdefault(it.get("stagedFile"), []).append(it)
    gid_of = {wb: _group_id(src, appno, _workbook_anchor(recs)) for wb, recs in by_wb.items()}
    booked, skipped, rejected, booked_gids = [], [], [], set()
    for it in prev["records"]:
        if it["idx"] not in idxs:
            continue
        rec, comp = it["rec"], it["comp"]
        if not it.get("checksOk"):
            rejected.append({"productName": (rec.get("productName") or "").strip(),
                             "cpCode": rec.get("cpCode"), "reason": "勾稽不平，不予入账"})
            continue
        # 上游链路闸（红线延伸）：自身勾稽平，但用到的半成品/复配料自身不平 → 本品成本建在错数上，连带拒收。
        if it.get("blockedBy"):
            rejected.append({"productName": (rec.get("productName") or "").strip(), "cpCode": rec.get("cpCode"),
                             "reason": "上游「%s」勾稽不平，其成本未经确认；本品用了它的价，连带不予入账" % "、".join(it["blockedBy"]),
                             "blockedBy": it["blockedBy"]})
            continue
        dup = db.bom_find_dup(src, it["numFp"])
        if dup:
            # 排版差异版：数字一致。若新版更全（更多物料编码）且旧版**仍是草稿**（未复核/已复核）→ 用新版明细覆盖（价格校验要靠编码），否则跳过。
            # ⚠ 旧版已初审/已终审（对外开放）的**绝不静默覆盖明细**——那会让已审核对外的成本被无痕改写、零留痕（审查 H3）。
            if bq.richness(rec) > bq.richness({"materials": dup.get("materials")}) and dup.get("status") in ("未复核", "已复核", None):
                upd = {"materials": rec.get("materials"), "erp_code": rec.get("erpCode") or dup.get("erp_code")}
                if not dup.get("bom_list"):
                    upd["bom_list"] = bq.match_bom_list(rec, prev.get("bomLists") or [])
                db.bom_update_entry(dup["id"], upd)
                db.bom_add_audit(dup["id"], u["name"], "排版差异版补明细", "旧明细", "更全版覆盖(补齐物料编码)")
                skipped.append({"productName": (rec.get("productName") or "").strip(),
                                "cpCode": rec.get("cpCode"), "reason": "排版差异版·已用更全版更新第 %d 号明细（补齐物料编码）" % dup["id"]})
            else:
                skipped.append({"productName": (rec.get("productName") or "").strip(),
                                "cpCode": rec.get("cpCode"), "reason": "排版差异版（数字与已入账第 %d 号完全一致）" % dup["id"]})
            continue
        fee = comp["srcFee"]
        bom_ent = bq.match_bom_entry(rec, prev.get("bomLists") or [])
        bom_mats = bom_ent["materials"] if bom_ent else None
        craft = (bom_ent or {}).get("craft")      # 工艺流程随 BOM 文件一起挂（复核②）
        # 成本会计商品版（若同产品配到）→ 留档 + 记价/税 diff（供复核时「采纳商品版调整」；无差异静默）
        g = it.get("goods")
        goods_ver = None
        if g:
            gr = g["rec"]
            goods_ver = {"materials": gr.get("materials"), "summary": gr.get("summary"),
                         "srcFile": gr.get("srcFile"), "srcLabel": g.get("srcLabel"),
                         "diff": it.get("goodsDiff") or {"rows": [], "hasDiff": False, "count": 0},
                         "applied": False}
        eid = db.bom_insert_entry({
            "source": src, "product_key": it["productKey"], "cp_code": rec.get("cpCode"),
            "bom_list": bom_mats, "craft": craft,
            "erp_code": rec.get("erpCode"), "product_name": (rec.get("productName") or "").strip(),
            "customer": rec.get("customer") or "", "pack_spec": rec.get("packSpec") or "",
            "supplier": rec.get("supplier") or "", "calc_date": rec.get("calcDate") or "",
            "order_qty": rec.get("orderQty"), "channel": it.get("channel") or "",
            "semi": 1 if it.get("semi") else 0, "mat_subtotal_excl": rec.get("matSubtotal"),
            "pack_subtotal_excl": rec.get("packSubtotal"), "fee_mfg": fee["mfg"], "fee_load": fee["load"],
            "fee_adm": fee["adm"], "full_cost_incl": comp["full"], "src_full": comp["srcFull"],
            "src_fee": comp["srcFee"], "summary": rec.get("summary"), "materials": rec.get("materials"),
            "checks": rec.get("checks"), "num_fp": it["numFp"], "source_type": prev.get("sourceType") or "manual_upload",
            "origin": it.get("origin") or "", "src_label": it.get("srcLabel") or "", "goods_version": goods_ver,
            "group_id": gid_of.get(it.get("stagedFile")), "active": 1,
            "approval_no": appno, "src_file": rec.get("srcFile"), "sheet": rec.get("sheet"),
            "status": "未复核", "created_by": u["name"],
        })
        # 源附件永久留档（供「原版核算表」导出）；商品版另存一份留档
        try:
            pdir = os.path.join(UPLOAD_DIR, src)
            os.makedirs(pdir, exist_ok=True)
            spath = os.path.join(STAGING_DIR, sid, it["stagedFile"])
            if os.path.isfile(spath):
                shutil.copy2(spath, os.path.join(pdir, "%d__%s" % (eid, it["stagedFile"])))
            if g and g.get("stagedFile"):
                gpath = os.path.join(STAGING_DIR, sid, g["stagedFile"])
                if os.path.isfile(gpath):
                    shutil.copy2(gpath, os.path.join(pdir, "%d__商品版__%s" % (eid, g["stagedFile"])))
        except Exception:
            pass
        booked_gids.add(gid_of.get(it.get("stagedFile")))
        booked.append({"id": eid, "productName": (rec.get("productName") or "").strip(), "cpCode": rec.get("cpCode"),
                       "origin": bq.ORIGIN_LABELS.get(it.get("origin") or "", ""),
                       "goodsDiff": (it.get("goodsDiff") or {}).get("count", 0) if g else None})
    # 「待修」登记：整组一条也没入账（全被勾稽/上游闸拦下）时，台账里查无此单 → 待办会整单消失。
    # 把它记进 bom_quote_pending 并留档源文件，让它**留在待办里可见可修**（但绝不进台账/标准库，红线不破）。
    rej_by_wb = {}
    for it in prev["records"]:
        if it["idx"] not in idxs:
            continue
        nm = (it["rec"].get("productName") or "").strip()
        r = next((x for x in rejected if x["productName"] == nm and x["cpCode"] == it["rec"].get("cpCode")), None)
        if r:
            rej_by_wb.setdefault(it.get("stagedFile"), []).append(r)
    for wb, rs in rej_by_wb.items():
        gid = gid_of.get(wb)
        if gid in booked_gids:
            db.bom_pending_clear(src, appno, gid)      # 该组已有产品入账 → 不算待修
            continue
        stash = ""
        try:
            pdir = os.path.join(UPLOAD_DIR, src)
            os.makedirs(pdir, exist_ok=True)
            spath = os.path.join(STAGING_DIR, sid, wb)
            if os.path.isfile(spath):
                stash = os.path.join(pdir, "pending__%s__%s" % (gid, wb))
                shutil.copy2(spath, stash)
        except Exception:
            pass
        db.bom_pending_upsert(src, appno, gid, wb, stash, rs, operator=u["name"])
    db.audit(u["name"], "bom_book", target=prev.get("approvalNo") or "-",
             detail="入账 %d，跳过 %d，拒 %d" % (len(booked), len(skipped), len(rejected)))
    return {"ok": True, "booked": booked, "skipped": skipped, "rejected": rejected}


@router.post("/api/bom/book")
async def bom_book(request: Request):
    """入账所选产品（保留：手工上传后仍可勾选入账）。"""
    u = _require_perm(request, CAP_FETCH)
    if not u:
        return JSONResponse({"ok": False, "msg": "无「抓取/录入」权限"}, status_code=403)
    body = await request.json()
    sid = str(body.get("stagingId") or "")
    idxs = set(int(i) for i in (body.get("indexes") or []))
    prev = _load_staging(sid)
    if not prev:
        return JSONResponse({"ok": False, "msg": "预检数据已过期，请重新取数/上传"}, status_code=400)
    return _book_staged(prev, sid, idxs, u)


@router.post("/api/bom/intake")
async def bom_intake(request: Request):
    """**立项**（业务方定 2026-09-04）：录钉钉单号 → 抓附件 → 解析 → 能入的入账、不能入的记「待修」→ 生成待办。
    这一步只负责把单立起来；哪些能入账、哪里不对、怎么修，统统到「处理页」去看去办，不在录入弹窗里判。"""
    u = _require_perm(request, CAP_FETCH)
    if not u:
        return JSONResponse({"ok": False, "msg": "无「抓取/录入」权限"}, status_code=403)
    body = await request.json()
    appno = str(body.get("approvalNo") or "").strip()
    if not appno:
        return JSONResponse({"ok": False, "msg": "请填钉钉审批编号"}, status_code=400)
    if not (dtb and dtb.configured()):
        return JSONResponse({"ok": False, "msg": "未配置钉钉应用——请在服务器 conf.ini [dingtalk] 配 appkey/appsecret，或改用手工上传。"}, status_code=400)
    res = dtb.fetch_approval(appno)
    if not res.get("ok"):
        return JSONResponse({"ok": False, "msg": res.get("msg") or "取数失败"}, status_code=400)
    files, comment_pending = [], []
    for a in res.get("attachments", []):
        if a.get("bytes") and str(a.get("fileName") or "").lower().endswith((".xlsx", ".xls")):
            files.append((a["fileName"], a["bytes"], a.get("label") or "", a.get("source") or "dingtalk_form"))
        elif a.get("source") == "dingtalk_comment":
            comment_pending.append({"fileName": a.get("fileName"), "reason": a.get("error") or "评论区文件当前权限取不到"})
    if not files:
        msg = "该审批未取到可解析的 xlsx 表单附件。"
        if comment_pending:
            msg += "另有 %d 个评论区补传附件当前钉钉权限取不到，请手工下载后上传。" % len(comment_pending)
        return JSONResponse({"ok": False, "msg": msg, "commentPending": comment_pending}, status_code=400)
    stg = _stage_files(files, appno, "dingtalk_form")
    r = _book_staged(_load_staging(stg["stagingId"]), stg["stagingId"],
                     set(x["idx"] for x in stg["records"]), u)
    db.audit(u["name"], "bom_intake", target=appno,
             detail="立项：附件 %d、入账 %d、待修 %d" % (len(files), len(r["booked"]), len(r["rejected"])))
    return {"ok": True, "approvalNo": appno, "title": res.get("title"),
            "booked": r["booked"], "rejected": r["rejected"], "skipped": r["skipped"],
            "commentPending": comment_pending, "warnings": stg.get("warnings") or []}


def _do_replace_sheet(src, gid, data, fname, label, user, appno, via):
    """用新核算表替换一个组：新文件里勾稽平的产品 → 顶替同组同产品旧版（旧版标 active=0 留痕、退出标准库）。
    仍不平的产品不入、回报（供再修）；新增产品（旧组没有的、如原本不平未入的半成品）直接入组。返回结果字典。"""
    try:
        recs = bq.parse_workbook(data, fname)
    except Exception as e:
        return {"ok": False, "msg": "替换文件解析失败：%s" % e}
    if not recs:
        return {"ok": False, "msg": "这份不是成本核算表（找不到核算样表页）。"}
    old_active = {x.get("product_key"): x for x in db.bom_group_entries(src, gid, include_superseded=False)}
    # 组内的 BOM清单（新文件常只含核算表；用同组既有 bom_list + 本次一并解析的清单兜底）
    bom_pool = []
    for x in db.bom_group_entries(src, gid):
        if x.get("bom_list"):
            bom_pool.append({"productName": x.get("product_name"), "cpCode": x.get("cp_code"),
                             "customer": x.get("customer"), "sheet": x.get("sheet"),
                             "materials": x["bom_list"], "craft": x.get("craft")})
    replaced, added, still_bad = [], [], []
    origin = bq.origin_from_label(label, "dingtalk_form", is_bom_list=False)
    pdir = os.path.join(UPLOAD_DIR, src)
    os.makedirs(pdir, exist_ok=True)
    for rec in recs:
        pk = bq.product_key(rec)
        if not bq.all_checks_ok(rec):
            still_bad.append({"productName": (rec.get("productName") or "").strip(), "cpCode": rec.get("cpCode"),
                              "failedChecks": bq.failed_checks(rec)})
            continue
        # 上游链路闸（同 book）：上游半成品/复配料自身不平 → 本品连带不入，别让替换口绕过红线
        blocked = bq.upstream_bad_chain(rec, recs)   # 传导式（审查 H2）
        if blocked:
            still_bad.append({"productName": (rec.get("productName") or "").strip(), "cpCode": rec.get("cpCode"),
                              "blockedBy": blocked, "failedChecks": []})
            continue
        comp = bq.compose(rec)
        fee = comp["srcFee"]
        old = old_active.get(pk)
        eid = db.bom_insert_entry({
            "source": src, "product_key": pk, "cp_code": rec.get("cpCode"),
            "bom_list": (bq.match_bom_entry(rec, bom_pool) or {}).get("materials"),
            "craft": (bq.match_bom_entry(rec, bom_pool) or {}).get("craft") or (old.get("craft") if old else None),
            "erp_code": rec.get("erpCode"),
            "product_name": (rec.get("productName") or "").strip(), "customer": rec.get("customer") or "",
            "pack_spec": rec.get("packSpec") or "", "supplier": rec.get("supplier") or "",
            "calc_date": rec.get("calcDate") or "", "order_qty": rec.get("orderQty"),
            "channel": (old.get("channel") if old else "") or _default_channel(rec),
            "semi": 1 if bq.is_semi(rec.get("cpCode")) else 0, "mat_subtotal_excl": rec.get("matSubtotal"),
            "pack_subtotal_excl": rec.get("packSubtotal"), "fee_mfg": fee["mfg"], "fee_load": fee["load"],
            "fee_adm": fee["adm"], "full_cost_incl": comp["full"], "src_full": comp["srcFull"],
            "src_fee": comp["srcFee"], "summary": rec.get("summary"), "materials": rec.get("materials"),
            "checks": rec.get("checks"), "num_fp": _num_fp(rec, comp), "source_type": "dingtalk_form",
            "origin": origin, "src_label": label or (old.get("src_label") if old else ""),
            "group_id": gid, "active": 1, "approval_no": appno,
            "src_file": rec.get("srcFile"), "sheet": rec.get("sheet"), "status": "未复核", "created_by": user,
        })
        try:
            safe = "".join(ch if ch not in '\\/:*?"<>|' else "_" for ch in (fname or "replace.xlsx"))
            with open(os.path.join(pdir, "%d__%s" % (eid, safe)), "wb") as fh:
                fh.write(data)
        except Exception:
            pass
        if old:
            db.bom_supersede_entry(old["id"], "被替换（%s）→ 第 %d 号，操作人 %s" % (via, eid, user))
            db.bom_add_audit(eid, user, "替换核算表·" + (rec.get("productName") or ""),
                             "旧 #%d(%s)" % (old["id"], old.get("src_file") or ""), "新 #%d(%s)" % (eid, fname))
            replaced.append({"id": eid, "old": old["id"], "productName": (rec.get("productName") or "").strip()})
        else:
            db.bom_add_audit(eid, user, "组内新增·" + (rec.get("productName") or ""), "（原不平/缺）", "新 #%d" % eid)
            added.append({"id": eid, "productName": (rec.get("productName") or "").strip()})
    # 部分替换后把**依赖被替换产品**的下游打回未复核（业务方 2026-09-04）——防「表面已审、底下换了」的陈旧依赖。
    # 依赖判定**按价配为主**（口径 quirk#5：成品料行「含税价」＝其半成品「全成本含税」，如 17.5577）——名字常对不上
    #   （半成品名「…陈皮辣卤风味…半成品」vs 成品料行「陈皮•梅子豆腐 半成品」），只按名会漏。用**替换前**的组快照配价、口径一致。
    stale = []
    replaced_names = {bq.norm(r["productName"]) for r in replaced}
    if replaced_names:
        pre = list(old_active.values())                       # 替换前的组内 active（含被替换旧版，价配口径一致）
        prods = [(bq.norm(x.get("product_name")), (x.get("summary") or {}).get("全成本含税")) for x in pre]

        def _deps_of(x):
            ups, me = set(), bq.norm(x.get("product_name"))
            for m in (x.get("materials") or []):
                pr, nm = m.get("priceIncl"), bq.norm(m.get("matName"))
                for upn, upfull in prods:
                    if upn == me or not upn:
                        continue
                    if (pr is not None and upfull is not None and abs(float(pr) - float(upfull)) < 0.02) or (nm and nm == upn):
                        ups.add(upn)
            return ups
        uses = {bq.norm(x.get("product_name")): _deps_of(x) for x in pre}

        def _dep_on_replaced(nm, seen=()):
            if nm in seen:
                return False
            for up in uses.get(nm, ()):
                if up in replaced_names or _dep_on_replaced(up, seen + (nm,)):
                    return True
            return False
        for x in db.bom_group_entries(src, gid, include_superseded=False):   # 现在仍 active 的下游
            nm = bq.norm(x.get("product_name"))
            if nm in replaced_names:
                continue                              # 被替换本体已是未复核
            if x.get("status") in ("已复核", "初审", "已审核") and _dep_on_replaced(nm):
                db.bom_update_entry(x["id"], {"status": "未复核", "review_steps": {}, "ack": None,
                                              "finalized_by": "", "finalized_at": "", "stale_note": "上游已更新·请复核"})
                db.bom_clear_final_if(src, x["product_key"], x["id"])       # 若它正是定稿版，撤下指针、退出标准库
                db.bom_add_audit(x["id"], user, "上游已更新·打回未复核", x.get("status") or "", "未复核")
                stale.append({"id": x["id"], "productName": (x.get("product_name") or "").strip()})
    # 待修标记：这次有产品入账 或 组内本就还有 active 记录 → 撤下待修（待修＝整组一条没入账；审查 L11：
    #   否则「新版全不平但旧版仍 active」会写出与台账并存的幽灵待修、下钻还回读旧的平衡文件显无差异）。
    group_has_active = bool(db.bom_group_entries(src, gid, include_superseded=False))
    if replaced or added or group_has_active:
        db.bom_pending_clear(src, appno, gid)
    elif still_bad:
        stash = ""
        try:
            pdir2 = os.path.join(UPLOAD_DIR, src)
            os.makedirs(pdir2, exist_ok=True)
            safe2 = "".join(ch if ch not in '\\/:*?"<>|' else "_" for ch in (fname or "replace.xlsx"))
            stash = os.path.join(pdir2, "pending__%s__%s" % (gid, safe2))
            with open(stash, "wb") as fh:
                fh.write(data)
        except Exception:
            pass
        db.bom_pending_upsert(src, appno, gid, fname, stash,
                              [{"productName": b["productName"], "cpCode": b.get("cpCode"),
                                "reason": ("上游「%s」不平，连带拦下" % "、".join(b["blockedBy"])) if b.get("blockedBy") else "勾稽不平",
                                "blockedBy": b.get("blockedBy") or []} for b in still_bad], operator=user)
    db.audit(user, "bom_replace_sheet", target="%s/%s" % (appno, gid),
             detail="替换 %d、新增 %d、仍不平 %d（%s）" % (len(replaced), len(added), len(still_bad), via))
    return {"ok": True, "replaced": replaced, "added": added, "stillBad": still_bad,
            "staleDownstream": stale, "via": via}


@router.post("/api/bom/replace-sheet")
async def bom_replace_sheet(request: Request):
    """手动上传修正后的核算表替换一个组（组内文件出错时）。旧版留痕、不进标准库。"""
    u = _require_perm(request, CAP_FETCH)
    if not u:
        return JSONResponse({"ok": False, "msg": "无「抓取/录入」权限"}, status_code=403)
    form = await request.form()
    gid = str(form.get("groupId") or "").strip()
    appno = str(form.get("approvalNo") or "").strip()
    if not gid:
        return JSONResponse({"ok": False, "msg": "缺组标识"}, status_code=400)
    uf = next((v for _k, v in form.multi_items() if hasattr(v, "read")), None)
    if uf is None:
        return JSONResponse({"ok": False, "msg": "请上传修正后的核算表 xlsx"}, status_code=400)
    data = await uf.read()
    res = _do_replace_sheet(_src(), gid, data, getattr(uf, "filename", "replace.xlsx"), "", u["name"], appno, "手动上传")
    return JSONResponse(res, status_code=200 if res.get("ok") else 400)


@router.post("/api/bom/refetch-replace")
async def bom_refetch_replace(request: Request):
    """重连钉钉重拉商务版核算表，替换一个组（当研发/工厂在钉钉里改好了重新提交时）。"""
    u = _require_perm(request, CAP_FETCH)
    if not u:
        return JSONResponse({"ok": False, "msg": "无「抓取/录入」权限"}, status_code=403)
    body = await request.json()
    gid = str(body.get("groupId") or "").strip()
    appno = str(body.get("approvalNo") or "").strip()
    if not (gid and appno):
        return JSONResponse({"ok": False, "msg": "缺组标识或审批编号"}, status_code=400)
    if not (dtb and dtb.configured()):
        return JSONResponse({"ok": False, "msg": "未配置钉钉应用，改用手动上传替换。"}, status_code=400)
    res = dtb.fetch_approval(appno)
    if not res.get("ok"):
        return JSONResponse({"ok": False, "msg": res.get("msg") or "取数失败"}, status_code=400)
    biz = next((a for a in res.get("attachments", []) if a.get("bytes") and "商务" in (a.get("label") or "")), None)
    if not biz:
        biz = next((a for a in res.get("attachments", []) if a.get("bytes")
                    and str(a.get("fileName") or "").lower().endswith((".xlsx", ".xls"))
                    and "商品版" not in (a.get("label") or "")), None)
    if not biz:
        return JSONResponse({"ok": False, "msg": "该审批未取到商务版核算表附件。"}, status_code=400)
    out = _do_replace_sheet(_src(), gid, biz["bytes"], biz["fileName"], biz.get("label") or "", u["name"], appno, "重连钉钉")
    return JSONResponse(out, status_code=200 if out.get("ok") else 400)


@router.post("/api/bom/attach-bomlist")
async def bom_attach_bomlist(request: Request):
    """给已入账记录补挂研发 BOM清单（核算表先入账、BOM清单后到时用）。解析后按产品名对齐本记录。"""
    u = _require_perm(request, CAP_ATTACH)
    if not u:
        return JSONResponse({"ok": False, "msg": "无「补挂BOM清单」权限"}, status_code=403)
    form = await request.form()
    e = db.bom_get_entry(form.get("entryId"))
    if not e or e.get("source") != _src():
        return JSONResponse({"ok": False, "msg": "记录不存在"}, status_code=404)
    uf = None
    for _k, v in form.multi_items():
        if hasattr(v, "read"):
            uf = v
            break
    if uf is None:
        return JSONResponse({"ok": False, "msg": "请上传 BOM清单 xlsx"}, status_code=400)
    raw = await uf.read()
    fname = getattr(uf, "filename", "")
    try:
        bls = bq.parse_bom_list(raw, fname)
    except Exception as ex:
        return JSONResponse({"ok": False, "msg": "BOM清单解析失败：%s" % ex}, status_code=400)
    try:
        craft = bq.parse_craft(raw, fname)      # 同一 BOM 文件里的「工艺流程」页
    except Exception:
        craft = None
    for b in bls:
        b["craft"] = craft
    ent = bq.match_bom_entry(_rec_from_entry(e), bls)
    if ent is None:
        return JSONResponse({"ok": False, "msg": "该 BOM清单里没有与本产品（%s）对得上的物料清单页。" % e.get("product_name")}, status_code=400)
    upd = {"bom_list": ent["materials"]}
    if craft:
        upd["craft"] = craft
    db.bom_update_entry(e["id"], upd)
    db.audit(u["name"], "bom_attach_bomlist", target=str(e["id"]), detail=e.get("product_name") or "")
    finals = db.bom_finals(_src())
    return {"ok": True, "entry": _entry_view(db.bom_get_entry(e["id"]), finals)}


# ============ 复核 / 定稿 ============
@router.post("/api/bom/review")
async def bom_review(request: Request):
    """复核：改费用参数（加工费/装卸费/管理费）与渠道，逐项留痕，状态转「已复核」。"""
    u = _require_perm(request, CAP_AUDIT)
    if not u:
        return JSONResponse({"ok": False, "msg": "无「审核」权限（仅成本会计）"}, status_code=403)
    body = await request.json()
    e = db.bom_get_entry(body.get("entryId"))
    if not e or e.get("source") != _src():
        return JSONResponse({"ok": False, "msg": "记录不存在"}, status_code=404)
    old_fee = _fee_of(e)
    new_fee = body.get("fee") or {}
    fields, changed = {}, 0
    for k, label in (("mfg", "加工费"), ("load", "装卸费"), ("adm", "管理费")):
        if k in new_fee and new_fee[k] is not None:
            nv = round(float(new_fee[k]), 4)
            if abs(nv - (old_fee[k] or 0)) > 1e-9:
                fields["fee_" + k] = nv
                db.bom_add_audit(e["id"], u["name"], label, old_fee[k], nv)
                changed += 1
    new_ch = body.get("channel")
    if new_ch is not None and new_ch != (e.get("channel") or ""):
        db.bom_add_audit(e["id"], u["name"], "渠道", CH_LABELS.get(e.get("channel") or "", "—"),
                         CH_LABELS.get(new_ch, "—"))
        fields["channel"] = new_ch
        changed += 1
    # 明细税率复核：改税率 → 后端按 成本不含税=添加量×含税价÷(1+税率) 重算该料成本、小计、全成本，逐项留痕。
    rec = _rec_from_entry(e)
    mat_sub, pack_sub = e.get("mat_subtotal_excl"), e.get("pack_subtotal_excl")
    new_mats = body.get("materials")
    if isinstance(new_mats, list) and new_mats:
        inv_rules = _invoice_rules()
        old_by = {}
        for m in (e.get("materials") or []):
            old_by.setdefault((m.get("matCode") or "") + "|" + (m.get("matName") or ""), m)
        merged_mats = []
        for m in new_mats:
            om = old_by.get((m.get("matCode") or "") + "|" + (m.get("matName") or ""), {})
            m = dict(m)
            nt = m.get("taxRate")
            if nt is not None and om and abs(float(nt) - (om.get("taxRate") or 0)) > 1e-9:
                db.bom_add_audit(e["id"], u["name"], "税率·" + (m.get("matName") or ""),
                                 "%s%%" % round((om.get("taxRate") or 0) * 100, 2), "%s%%" % round(float(nt) * 100, 2))
                changed += 1
            ni = m.get("invoiceType")
            if ni is not None and om and str(ni).strip() != str(om.get("invoiceType") or "").strip():
                db.bom_add_audit(e["id"], u["name"], "发票类型·" + (m.get("matName") or ""),
                                 om.get("invoiceType") or "—", ni or "—")
                changed += 1
            q, pr, tr = m.get("qtyPerKg"), m.get("priceIncl"), m.get("taxRate")
            if q is not None and pr is not None and tr is not None:   # 权威重算：按发票类型套 N 列算法，不信前端算的成本
                m["costExcl"] = bq.invoice_cost_excl(q, pr, tr, m.get("invoiceType"), inv_rules)
            merged_mats.append(m)
        mat_sub = round(sum((x.get("costExcl") or 0) for x in merged_mats if x.get("seg") == "原料"), 4)
        pack_sub = round(sum((x.get("costExcl") or 0) for x in merged_mats if x.get("seg") == "包材"), 4)
        fields["materials"] = merged_mats
        fields["mat_subtotal_excl"] = mat_sub
        fields["pack_subtotal_excl"] = pack_sub
        rec = {**rec, "materials": merged_mats, "matSubtotal": mat_sub, "packSubtotal": pack_sub}
    if fields:
        # 全成本随费用参数/税率重算（含税五分项：料/包按新小计×1.13，费用参数用复核后值）
        merged_fee = {**old_fee, **{k[4:]: v for k, v in fields.items() if k.startswith("fee_")}}
        fields["full_cost_incl"] = bq.compose(rec, merged_fee)["full"]
    if changed:   # 改了税率/费用/渠道 → ④报价核算需重新确认；动了费用参数连②工艺流程也要重认（①BOM清单/③用量不受影响）
        rs = dict(e.get("review_steps") or {})
        rs.pop("price", None)
        if any(k.startswith("fee_") for k in fields):
            rs.pop("craft", None)
        fields["review_steps"] = rs
    fields["status"] = "已复核"
    fields["reviewed_by"] = u["name"]
    from core import _now
    fields["reviewed_at"] = _now()
    # 若在已初审/已终审基础上改了费用/税率 → 清终审戳+按id校验清定稿指针（审查 M8：否则留悬空指针、std台账显错版、卡死）
    if changed and e.get("status") in ("初审", "已审核"):
        fields["ack"] = None
        fields["finalized_by"] = ""
        fields["finalized_at"] = ""
        db.bom_clear_final_if(e.get("source"), e.get("product_key"), e["id"])
    db.bom_update_entry(e["id"], fields)
    db.audit(u["name"], "bom_review", target=str(e["id"]), detail="%d 项变更" % changed)
    finals = db.bom_finals(_src())
    return {"ok": True, "changed": changed, "entry": _entry_view(db.bom_get_entry(e["id"]), finals)}


@router.post("/api/bom/apply-goods")
async def bom_apply_goods(request: Request):
    """采纳「成本会计商品版」对本底稿(采购商务版)的价/税调整：按物料编码覆盖含税价/税率，
    权威重算成本不含税=添加量×含税价÷(1+税率)、原料/包材小计、全成本，逐项留痕。价改→价格核价步骤清空需重认。"""
    u = _require_perm(request, CAP_AUDIT)
    if not u:
        return JSONResponse({"ok": False, "msg": "无「审核」权限（仅成本会计）"}, status_code=403)
    body = await request.json()
    e = db.bom_get_entry(body.get("entryId"))
    if not e or e.get("source") != _src():
        return JSONResponse({"ok": False, "msg": "记录不存在"}, status_code=404)
    gv = e.get("goods_version") or {}
    if not gv:
        return JSONResponse({"ok": False, "msg": "本记录没有成本会计商品版可采纳。"}, status_code=400)
    if not (gv.get("diff") or {}).get("hasDiff"):
        return JSONResponse({"ok": False, "msg": "商品版与商务版价/税一致，无需采纳。"}, status_code=400)
    merged, changed = [], 0
    inv_rules = _invoice_rules()
    for m, g in bq._pair_goods(e.get("materials") or [], gv.get("materials") or []):
        m = dict(m)
        if g:
            for field, la in (("priceIncl", "含税采购价·商品版"), ("taxRate", "税率·商品版")):
                a, b = m.get(field), g.get(field)
                if a is not None and b is not None and abs(float(a) - float(b)) > 1e-9:
                    db.bom_add_audit(e["id"], u["name"], "%s·%s" % (la, m.get("matName") or ""),
                                     ("%s%%" % round(a * 100, 2)) if field == "taxRate" else a,
                                     ("%s%%" % round(b * 100, 2)) if field == "taxRate" else b)
                    m[field] = b
                    changed += 1
            q, pr, tr = m.get("qtyPerKg"), m.get("priceIncl"), m.get("taxRate")
            if q is not None and pr is not None and tr is not None:
                m["costExcl"] = bq.invoice_cost_excl(q, pr, tr, m.get("invoiceType"), inv_rules)
        merged.append(m)
    mat_sub = round(sum((x.get("costExcl") or 0) for x in merged if x.get("seg") == "原料"), 4)
    pack_sub = round(sum((x.get("costExcl") or 0) for x in merged if x.get("seg") == "包材"), 4)
    rec = {**_rec_from_entry(e), "materials": merged, "matSubtotal": mat_sub, "packSubtotal": pack_sub}
    fields = {"materials": merged, "mat_subtotal_excl": mat_sub, "pack_subtotal_excl": pack_sub,
              "full_cost_incl": bq.compose(rec, _fee_of(e))["full"]}
    gv2 = dict(gv); gv2["applied"] = True
    fields["goods_version"] = gv2
    rs = dict(e.get("review_steps") or {})
    rs.pop("price", None)                # 价改 → ④报价核算需重新确认
    fields["review_steps"] = rs
    # ⚠ 改了成本 → **失效审核态**（审查 H1）：已初审/已终审的降回「已复核」、清终审戳、清定稿指针，
    #   强制重新初审→终审，绝不让改过的成本用旧的终审戳偷偷对外开放给 BP。
    reset = _invalidate_review(e, fields)
    db.bom_update_entry(e["id"], fields)
    if reset:
        db.bom_add_audit(e["id"], u["name"], "采纳商品版·审核态失效", reset, "需重新初审/终审")
    db.audit(u["name"], "bom_apply_goods", target=str(e["id"]), detail="采纳商品版 %d 项价/税" % changed)
    finals = db.bom_finals(_src())
    return {"ok": True, "changed": changed, "reviewReset": reset, "entry": _entry_view(db.bom_get_entry(e["id"]), finals)}


@router.post("/api/bom/final-review")
async def bom_final_review(request: Request):
    """**财务BP终审**（业务方定 2026-09-04）——全流程就两个戳，别拆碎：
      成本会计**初审**（四步确认+定性 → 盖「初审」戳） → 财务BP**终审**（盖「已审核」戳）。
    ⚠ **只有终审通过的才对外开放**（BP 消费口 /final 只放已审核）；终审也能**退回**给成本会计重做。
    重新初审会清空终审戳——新版必须重新终审。"""
    u = _require_perm(request, CAP_FINAL_REVIEW)
    if not u:
        return JSONResponse({"ok": False, "msg": "无「终审」权限（该权限给财务BP/经理）"}, status_code=403)
    body = await request.json()
    e = db.bom_get_entry(body.get("entryId"))
    if not e or e.get("source") != _src():
        return JSONResponse({"ok": False, "msg": "记录不存在"}, status_code=404)
    if e.get("status") not in ("初审", "已审核"):
        return JSONResponse({"ok": False, "msg": "只有成本会计已初审的版本才轮到终审。"}, status_code=400)
    from core import _now
    note = str(body.get("note") or "").strip()
    approve = body.get("approve", True)
    if not approve and not note:
        return JSONResponse({"ok": False, "msg": "退回请写明原因（成本会计要据此改）"}, status_code=400)
    # 终审是「对外开放」的最后一道门 → **初审人不得自己终审**（代码强制，审查加固；镜像作废流的自批闸）。
    # 退回不受此限（谁都能打回）；只有「盖已审核戳·对外」这一步要求两个人。
    # 例外（业务方定 2026-09-04）：**主管理员**可自审自终——小团队常一人身兼初审+终审，
    #   与其逼着造个假的第二账号(反而污染审计)，不如允许主管理员自审、但戳与留痕明记「自审·单人」。
    self_review = approve and (e.get("finalized_by") or "") == u["name"]
    if self_review and not db.is_super(u):
        return JSONResponse({"ok": False, "msg": "本版初审人是您本人，不能自己终审——终审须由另一人（财务BP）把关，方可对外开放。（仅主管理员可自审）"}, status_code=400)
    if approve:
        ack = {"by": u["name"], "at": _now(), "note": note[:200]}
        if self_review:
            ack["selfReview"] = True          # 自审留痕：这版没经第二人，谁看台账都认得出
        db.bom_update_entry(e["id"], {"status": "已审核", "ack": ack})
        did = "已审核·对外开放" + ("·自审(主管理员单人)" if self_review else "") + ("（%s）" % note if note else "")
        db.bom_add_audit(e["id"], u["name"], "终审（财务BP）" + ("·自审" if self_review else ""), "初审", did)
        msg = "已终审通过，盖「已审核」戳，对外开放给 BP 报价" + ("（主管理员自审）" if self_review else "")
    else:
        db.bom_clear_final_if(_src(), e["product_key"], e["id"])   # 退回 → 撤下定稿指针(按id校验，审查M9)，不再供 BP 消费
        db.bom_update_entry(e["id"], {"status": "已复核", "ack": None, "finalized_by": "", "finalized_at": ""})
        db.bom_add_audit(e["id"], u["name"], "终审（财务BP）", "初审", "退回成本会计：" + note[:180])
        for rid in db.bom_clear_obsolete_by(e["id"]):                # 它替代过的旧版恢复为当前版（换码承接 V2.440）
            db.bom_add_audit(rid, u["name"], "恢复为当前版", "失效", "替代它的 #%d 被终审退回" % e["id"])
        msg = "已退回成本会计（该版撤出标准台账）"
    db.audit(u["name"], "bom_final_review", target=str(e["id"]), detail=msg)
    finals = db.bom_finals(_src())
    return {"ok": True, "msg": msg, "approved": bool(approve),
            "entry": _entry_view(db.bom_get_entry(e["id"]), finals)}


# ============ 作废：申请 + 终审批准（业务方定 2026-09-04）============
# 「作废」不是删除，是**标记**——记录留着、留痕，只是退出工作区与标准成本库。
# 因为作废等于让一份成本数据从台账消失，怕成本会计一个人把不该舍弃的舍弃掉，故做成两步：
#   ① 成本会计（bom:audit）**申请**作废，必须写理由 → 记录**照常有效**，只打「待复审作废」标；
#   ② 终审人（bom:final_review，给财务BP）批准才真作废 / 或驳回。**申请人不得自批**（代码强制）。


def _void_target(body):
    """作废对象：单条记录（entryId）或整个组（groupId+approvalNo，含「待修」批次）。"""
    eid = body.get("entryId")
    gid = str(body.get("groupId") or "").strip()
    return (int(eid) if eid else None), gid, str(body.get("approvalNo") or "").strip()


@router.post("/api/bom/void-request")
async def bom_void_request(request: Request):
    """成本会计**申请**作废（如钉钉单里有多个作废版本）。理由必填；申请期间记录照常有效。"""
    u = _require_perm(request, CAP_AUDIT)
    if not u:
        return JSONResponse({"ok": False, "msg": "无「审核」权限"}, status_code=403)
    body = await request.json()
    eid, gid, appno = _void_target(body)
    reason = str(body.get("reason") or "").strip()
    if not reason:
        return JSONResponse({"ok": False, "msg": "请写明作废理由（如「研发重发了新版，本版作废」）"}, status_code=400)
    from core import _now
    req = {"state": "pending", "reason": reason[:300], "by": u["name"], "at": _now()}
    src, n = _src(), 0
    if eid:
        e = db.bom_get_entry(eid)
        if not e or e.get("source") != src:
            return JSONResponse({"ok": False, "msg": "记录不存在"}, status_code=404)
        db.bom_update_entry(eid, {"void_req": req})
        db.bom_add_audit(eid, u["name"], "申请作废", "", reason[:200])
        n = 1
    elif gid:
        grp = db.bom_group_entries(src, gid, include_superseded=False)
        for x in grp:
            db.bom_update_entry(x["id"], {"void_req": req})
            db.bom_add_audit(x["id"], u["name"], "申请作废（整组）", "", reason[:200])
            n += 1
        # 审查 L12：整组作废可能不带 approvalNo（前端组对象无此字段）→ 从组记录/待修批次解析真实审批号，
        #   别用空串精确匹配把申请静默丢掉（存在性宽松、写入精确的口径不一致）。
        pends = db.bom_pending_list(src)
        pd = next((p for p in pends if p.get("group_id") == gid), None)
        real_ap = appno or (grp[0].get("approval_no") if grp else "") or (pd.get("approval_no") if pd else "")
        if pd:
            db.bom_pending_set_void(src, pd.get("approval_no") or real_ap, gid, req)   # 待修批次也可申请作废
            n += 1
    else:
        return JSONResponse({"ok": False, "msg": "缺作废对象"}, status_code=400)
    db.audit(u["name"], "bom_void_request", target=str(eid or gid), detail=reason[:120])
    return {"ok": True, "requested": n, "msg": "已提交作废申请，待终审人批准后才会真正作废。"}


@router.post("/api/bom/void-review")
async def bom_void_review(request: Request):
    """终审人批准/驳回作废申请。**申请人不得自批**（防一人闭环）。批准 → 真作废（active=0, voided）。"""
    u = _require_perm(request, CAP_FINAL_REVIEW)
    if not u:
        return JSONResponse({"ok": False, "msg": "无「终审」权限（该权限给财务BP/主管）"}, status_code=403)
    body = await request.json()
    eid, gid, appno = _void_target(body)
    approve = bool(body.get("approve"))
    note = str(body.get("note") or "").strip()
    if not approve and not note:
        return JSONResponse({"ok": False, "msg": "驳回请写明理由"}, status_code=400)
    from core import _now
    src = _src()
    targets = []
    if eid:
        e = db.bom_get_entry(eid)
        if e and e.get("source") == src:
            targets.append(e)
    elif gid:
        targets = db.bom_group_entries(src, gid, include_superseded=False)
    else:
        return JSONResponse({"ok": False, "msg": "缺作废对象"}, status_code=400)
    done, skipped = 0, []
    for e in targets:
        vr = e.get("void_req") or {}
        if vr.get("state") != "pending":
            skipped.append("#%d 无待终审的作废申请" % e["id"])
            continue
        self_void = (vr.get("by") or "") == u["name"]     # 硬规则：申请人不得自批（主管理员例外，同终审口径 V2.430）
        if self_void and not db.is_super(u):
            skipped.append("#%d 申请人与终审人同为 %s，不得自批" % (e["id"], u["name"]))
            continue
        vr2 = dict(vr, state=("voided" if approve else "rejected"),
                   reviewBy=u["name"], reviewAt=_now(), note=note[:300])
        if self_void:
            vr2["selfReview"] = True                       # 自批留痕：这次作废没经第二人
        if approve:
            db.bom_update_entry(e["id"], {"void_req": vr2})
            db.bom_supersede_entry(e["id"], "作废：%s（申请 %s，终审 %s%s）" % (
                vr.get("reason"), vr.get("by"), u["name"], "·自批" if self_void else ""), kind="voided")
        else:
            db.bom_update_entry(e["id"], {"void_req": vr2})
        db.bom_add_audit(e["id"], u["name"], "作废终审" + ("·自批" if self_void else ""), "申请：" + (vr.get("reason") or ""),
                         ("批准作废" if approve else "驳回") + ("（%s）" % note if note else ""))
        done += 1
    # 待修批次
    if gid:
        for p in db.bom_pending_list(src, appno or None):
            if p["group_id"] != gid:
                continue
            vr = p.get("void_req") or {}
            if vr.get("state") != "pending":
                continue
            self_void = (vr.get("by") or "") == u["name"]
            if self_void and not db.is_super(u):          # 申请人不得自批（主管理员例外）
                skipped.append("待修批次：申请人不得自批")
                continue
            db.bom_pending_set_void(src, p["approval_no"], gid,
                                    dict(vr, state=("voided" if approve else "rejected"),
                                         reviewBy=u["name"], reviewAt=_now(), note=note[:300],
                                         **({"selfReview": True} if self_void else {})))
            done += 1
    db.audit(u["name"], "bom_void_review", target=str(eid or gid),
             detail=("批准作废" if approve else "驳回") + " %d 条" % done)
    if not done:      # 一条都没处理成 → 别回「已批准」这种假话，把原因说清楚
        return JSONResponse({"ok": False, "reviewed": 0, "skipped": skipped,
                             "msg": "没有可终审的作废申请。" + ("；".join(skipped) if skipped else "")}, status_code=400)
    return {"ok": True, "reviewed": done, "skipped": skipped,
            "msg": ("已批准作废 %d 条（记录留痕、退出标准成本库）" % done if approve else "已驳回作废申请 %d 条" % done)
                   + ("；另有未处理：" + "；".join(skipped) if skipped else "")}


@router.post("/api/bom/classify")
async def bom_classify(request: Request):
    """**审核定性**（业务方 2026-09-03，点「审核」的弹窗）：成本会计指定
    ①物料类别（复配料/自产半成品/自产成品/委外半成品/委外成品）
    ②是否建议/允许对外报价——**不建议必须写明原因**（如「包材不全」「XX物料暂定」）。
    编码规律不固定，故类别以此人工定性为准；同步 semi（BP 消费口按 semi 判）并逐项留痕。定稿前必须完成。"""
    u = _require_perm(request, CAP_AUDIT)
    if not u:
        return JSONResponse({"ok": False, "msg": "无「审核」权限（仅成本会计可定性）"}, status_code=403)
    body = await request.json()
    e = db.bom_get_entry(body.get("entryId"))
    if not e or e.get("source") != _src():
        return JSONResponse({"ok": False, "msg": "记录不存在"}, status_code=404)
    cat = str(body.get("category") or "").strip()
    if cat not in bq.MAT_CATEGORIES:
        return JSONResponse({"ok": False, "msg": "请选择物料类别：%s" % " / ".join(bq.MAT_CATEGORIES)}, status_code=400)
    q = body.get("quotable")
    if q is None:
        return JSONResponse({"ok": False, "msg": "请选择是否建议/允许对外报价"}, status_code=400)
    q = bool(q)
    reason = str(body.get("reason") or "").strip()
    if not q and not reason:
        return JSONResponse({"ok": False, "msg": "不建议对外报价时，必须写明原因（如「包材不全」「XX物料暂定」）"}, status_code=400)
    if q:
        reason = ""                     # 建议报价 → 不留原因
    from core import _now
    old_cat, old_kind = (e.get("mat_category") or ""), _kind_of(e)
    new_kind = bq.cat_to_kind(cat)
    fields = {"mat_category": cat, "quotable": 1 if q else 0, "quote_reason": reason[:300],
              "semi": 0 if new_kind == "成品" else 1,
              "classified_by": u["name"], "classified_at": _now()}
    if new_kind != "成品" and (e.get("channel") or ""):
        fields["channel"] = ""          # 非成品不单独挂渠道（随上层成品走）
    db.bom_update_entry(e["id"], fields)
    if cat != old_cat:
        db.bom_add_audit(e["id"], u["name"], "物料类别（审核定性）", old_cat or ("建议:" + old_kind), cat)
    old_q = e.get("quotable")
    if old_q is None or bool(old_q) != q or (e.get("quote_reason") or "") != reason:
        db.bom_add_audit(e["id"], u["name"], "对外报价建议",
                         "未定性" if old_q is None else ("建议" if old_q else "不建议:" + (e.get("quote_reason") or "")),
                         "建议报价" if q else ("不建议报价 · " + reason))
    db.audit(u["name"], "bom_classify", target=str(e["id"]),
             detail="%s / %s%s" % (cat, "建议报价" if q else "不建议报价", "" if q else "（%s）" % reason))
    # **定性即定稿**（业务方定 2026-09-03：审核定性和定稿是一个动作，不做两个）。
    # 四步齐了就顺势定稿；没齐则只存定性、回报还缺哪步，由前端提示。
    e2 = db.bom_get_entry(e["id"])
    rs = e2.get("review_steps") or {}
    miss = ["%s%s" % (STEP_NO[k], STEP_LABELS[k]) for k in REVIEW_STEPS if not rs.get(k)]
    ub = _upstream_block(_upstream_status(e2))   # 上游未就绪 → 存定性但不定稿
    if ub:
        miss = miss + ["上游：" + "；".join(ub)]
    if not ((_net_weight(e2)[0] or 0) > 0):      # 净重闸（BP 对接 2026-09-05 §2，与勾稽同级）：单位净重空/≤0 不定稿
        miss = miss + ["单位净重(kg)未填或≤0——右栏「单位净重」填好并确认再定稿"]
    # 换码承接闸（业务方定 2026-09-05）：同 CP / 同物料编码已有审核版 → **必须先答「原版是否失效」**（confirmObsolete=true）才定稿；
    # 没答 → 只存定性，把候选回给前端弹确认。答「否」= 不定稿，先核对。
    cands = _obsolete_candidates(e2) if not miss else []
    need_confirm = bool(cands) and not bool(body.get("confirmObsolete"))
    finalized, affected = False, None
    if not miss and not need_confirm:
        prev_final = db.bom_get_final(_src(), e2["product_key"])
        db.bom_update_entry(e2["id"], {"status": "初审", "finalized_by": u["name"], "finalized_at": _now(), "ack": None,
                                       "obsolete_by": None, "obsolete_at": None, "obsolete_note": None})   # 本版重新成为当前版
        db.bom_set_final(_src(), e2["product_key"], e2["id"], u["name"])
        db.audit(u["name"], "bom_finalize", target=str(e2["id"]), detail="随审核定性定稿 · " + (e2.get("product_name") or ""))
        if cands:
            _mark_obsolete(e2, cands, u["name"])
        affected = _affected_pricing(e2)
        finalized = True
    finals = db.bom_finals(_src())
    return {"ok": True, "finalized": finalized, "missingSteps": miss,
             "needConfirm": cands if need_confirm else [], "obsoleted": cands if finalized else [],
             "affectedPricing": affected, "entry": _entry_view(db.bom_get_entry(e["id"]), finals)}


@router.post("/api/bom/confirm-step")
async def bom_confirm_step(request: Request):
    """确认/撤销复核某一步（step=qty 用量自洽 / price 价格核价）。定稿前两步都要确认。"""
    u = _require_perm(request, CAP_AUDIT)
    if not u:
        return JSONResponse({"ok": False, "msg": "无「审核」权限"}, status_code=403)
    body = await request.json()
    e = db.bom_get_entry(body.get("entryId"))
    if not e or e.get("source") != _src():
        return JSONResponse({"ok": False, "msg": "记录不存在"}, status_code=404)
    step = body.get("step")
    if step not in REVIEW_STEPS:
        return JSONResponse({"ok": False, "msg": "step 须为 %s" % "/".join(REVIEW_STEPS)}, status_code=400)
    on = body.get("on", True)
    from core import _now
    rs = dict(e.get("review_steps") or {})
    if on:
        rs[step] = {"by": u["name"], "at": _now()}
    else:
        rs.pop(step, None)
    upd = {"review_steps": rs}
    if any(rs.get(k) for k in REVIEW_STEPS) and e.get("status") == "未复核":
        upd["status"] = "已复核"
    if on and e.get("stale_note"):
        upd["stale_note"] = ""                    # 开始重新复核 → 清「上游已更新」提醒
    db.bom_update_entry(e["id"], upd)
    db.bom_add_audit(e["id"], u["name"], "确认" + STEP_LABELS.get(step, step), "",
                     "已确认" if on else "撤销确认")
    finals = db.bom_finals(_src())
    return {"ok": True, "entry": _entry_view(db.bom_get_entry(e["id"]), finals)}


@router.post("/api/bom/finalize")
async def bom_finalize(request: Request):
    """定稿：状态转「已定稿」，把该产品的定稿指针指向本版（独立表，覆盖旧指针）。前提：两步已确认。"""
    u = _require_perm(request, CAP_AUDIT)
    if not u:
        return JSONResponse({"ok": False, "msg": "无「审核」权限（仅成本会计）"}, status_code=403)
    body = await request.json()
    e = db.bom_get_entry(body.get("entryId"))
    if not e or e.get("source") != _src():
        return JSONResponse({"ok": False, "msg": "记录不存在"}, status_code=404)
    rs = e.get("review_steps") or {}
    miss = ["%s%s" % (STEP_NO[k], STEP_LABELS[k]) for k in REVIEW_STEPS if not rs.get(k)]
    if miss:
        return JSONResponse({"ok": False, "msg": "请先确认：%s（①BOM清单②工艺流程只是参考、不用确认）" % "、".join(miss)}, status_code=400)
    if not _classified(e):      # 审核定性（物料类别 + 是否允许对外报价）＝定稿动作本身
        return JSONResponse({"ok": False, "msg": "请在「审核定稿」里完成定性：物料类别 + 是否建议对外报价（不建议须写原因）。"}, status_code=400)
    ub = _upstream_block(_upstream_status(e))   # 上游链路闸：上游未定稿/价格对不上 → 下游不许先定稿
    if ub:
        return JSONResponse({"ok": False, "msg": "上游半成品/复配料未就绪，不能先定稿本品：%s" % "；".join(ub)}, status_code=400)
    if not ((_net_weight(e)[0] or 0) > 0):       # 净重闸（BP 对接 2026-09-05 §2，与勾稽同级）
        return JSONResponse({"ok": False, "msg": "单位净重(kg)未填或≤0，不能定稿——BP 定价要按袋/盒换算，请在右栏「单位净重」填好并确认。"}, status_code=400)
    cands = _obsolete_candidates(e)              # 换码承接闸（V2.440）：同 CP / 同物料编码已有审核版 → 先答「原版是否失效」
    if cands and not bool(body.get("confirmObsolete")):
        return {"ok": False, "needConfirm": cands,
                "msg": "台账里已有 %d 个同CP/同物料编码的审核版本（%s）。本版定稿后它们将失效、退出对外台账。请确认「原版本是否失效」。"
                       % (len(cands), "、".join("%s %s" % (c["cpCode"], c["auditAt"]) for c in cands))}
    from core import _now
    prev_final = db.bom_get_final(_src(), e["product_key"])
    db.bom_update_entry(e["id"], {"status": "初审", "finalized_by": u["name"], "finalized_at": _now(), "ack": None,
                                  "obsolete_by": None, "obsolete_at": None, "obsolete_note": None})
    db.bom_set_final(_src(), e["product_key"], e["id"], u["name"])
    db.audit(u["name"], "bom_finalize", target=str(e["id"]), detail=e.get("product_name") or "")
    if cands:
        _mark_obsolete(e, cands, u["name"])
    # 定稿变更通知（BP 消费提示）——留痕，前端据此弹「成本已更新，N 个定价方案受影响」。不静默变价。
    affected = _affected_pricing(e)
    finals = db.bom_finals(_src())
    return {"ok": True, "entry": _entry_view(db.bom_get_entry(e["id"]), finals),
            "replacedFinal": bool(prev_final and prev_final.get("entry_id") != e["id"]),
            "obsoleted": cands, "affectedPricing": affected}


@router.post("/api/bom/unfinalize")
async def bom_unfinalize(request: Request):
    u = _require_perm(request, CAP_AUDIT)
    if not u:
        return JSONResponse({"ok": False, "msg": "无「审核」权限"}, status_code=403)
    body = await request.json()
    e = db.bom_get_entry(body.get("entryId"))
    if not e or e.get("source") != _src():
        return JSONResponse({"ok": False, "msg": "记录不存在"}, status_code=404)
    db.bom_clear_final_if(_src(), e["product_key"], e["id"])   # 按id校验清指针(审查M9)
    db.bom_update_entry(e["id"], {"status": "已复核" if e.get("reviewed_by") else "未复核", "ack": None,
                                  "finalized_by": "", "finalized_at": ""})
    restored = db.bom_clear_obsolete_by(e["id"])               # 它替代过的旧版恢复（换码承接 V2.440）
    for rid in restored:
        db.bom_add_audit(rid, u["name"], "恢复为当前版", "失效", "替代它的 #%d 撤销定稿" % e["id"])
    db.audit(u["name"], "bom_unfinalize", target=str(e["id"]))
    finals = db.bom_finals(_src())
    return {"ok": True, "entry": _entry_view(db.bom_get_entry(e["id"]), finals)}


def _affected_pricing(e):
    """定稿会影响哪些定价方案（占位：TOB/TOC 直连=自动生效；电商/通品=显式引用）。
    正式版接 BP pricingMaterial 反查；此处按渠道回条人话提示，供前端展示，不静默变价。"""
    ch = e.get("channel") or ""
    if ch in ("tob", "toc"):
        return {"mode": "direct", "note": "%s 定价默认直连台账定稿版，本次定稿后自动生效。" % CH_LABELS.get(ch, ch)}
    if ch in ("ecom", "common"):
        return {"mode": "explicit", "note": "%s 需在定价测算里显式引用；已引用的方案将收到「成本已更新」提示，确认后刷新。" % CH_LABELS.get(ch, ch)}
    return {"mode": "none", "note": "尚未设定渠道；设定后 BP 侧才能消费本定稿版。"}


# ============ 导出 ============
def _xlsx_response(data, filename):
    from urllib.parse import quote
    return Response(content=data, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename*=UTF-8''%s" % quote(filename)})


def _xlsx_to_html(data, title=""):
    """xlsx 字节 → 只读 HTML 预览（原版/重排版通用）——服务端转，前端无需 xlsx 库、离线可用。
    尽量还原 Excel 观感（预览与下载同版）：逐格边框（无边框处不画线，对应 showGridLines=False）、字号/加粗/字色、底色、
    水平/垂直对齐与换行、行高列宽、合并 rowspan/colspan、数字格式（三段式 0 显 "-"、百分比、显式小数位、整数不加千分位护编码）、
    条件格式数据条（用渐变背景模拟）。"""
    import io
    import html as _html
    import re as _re
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    wb = load_workbook(io.BytesIO(data), data_only=True)
    esc = _html.escape

    def _dec(s):                                                    # 小数位＝小数点后连续 0 的个数
        m = _re.match(r"0*", s.split(".", 1)[1]) if "." in s else None
        return len(m.group(0)) if m else 0

    def fmt(v, nf):
        if v is None:
            return ""
        if isinstance(v, bool):
            return "是" if v else "否"
        if isinstance(v, (int, float)):
            secs = (nf or "").split(";")
            if len(secs) >= 3 and v == 0:                           # 三段式格式：0 用第三段（如 "-"）
                return secs[2].strip().strip('"')
            low = secs[0].lower()
            if "%" in low:
                return ("{:.%df}%%" % _dec(low)).format(v * 100)
            if "." in secs[0]:                                      # 显式小数位优先（1.0 按 0.0000 显 1.0000）
                return "{:.{}f}".format(v, _dec(secs[0]) or 2)
            if float(v).is_integer():
                return str(int(v))                                  # 无小数格式的整数(含 ERP 码)不加千分位
            return ("{:.4f}".format(v)).rstrip("0").rstrip(".")
        return str(v)

    def hexof(c):
        try:
            if c is None or getattr(c, "type", None) != "rgb":
                return None
            rgb = c.rgb
            return rgb[-6:].upper() if isinstance(rgb, str) and len(rgb) >= 6 else None
        except Exception:
            return None

    def style(cell, spanned, bar):
        s = []
        f = cell.font
        if f:
            if f.bold:
                s.append("font-weight:700")
            if f.sz:
                s.append("font-size:%dpx" % round(float(f.sz) * 1.33))
            fc = hexof(f.color)
            if fc and fc != "000000":
                s.append("color:#" + fc)
        bg = None
        fl = cell.fill
        if fl is not None and getattr(fl, "patternType", None) == "solid":
            bg = hexof(fl.fgColor)
            if bg == "FFFFFF":
                bg = None
        if bar is not None:                                         # 数据条：绿色占 pct，其余为原底色
            col, pct = bar
            s.append("background:linear-gradient(90deg,#%s %.0f%%,%s %.0f%%)"
                     % (col, pct * 100, ("#" + bg) if bg else "transparent", pct * 100))
        elif bg:
            s.append("background:#" + bg)
        al = cell.alignment
        if al and al.horizontal in ("left", "center", "right"):
            s.append("text-align:" + al.horizontal)
        elif isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
            s.append("text-align:right")
        if al and al.vertical in ("top", "bottom"):
            s.append("vertical-align:" + al.vertical)
        if al and al.wrap_text:
            s.append("white-space:normal")
        b = cell.border
        for side in ("left", "right", "top", "bottom"):
            sd = getattr(b, side, None)
            if sd is not None and sd.style:
                s.append("border-%s:1px solid #%s" % (side, hexof(sd.color) or "808080"))
        if spanned:
            s.append("max-width:none")
        return ";".join(s)

    out = ['<!doctype html><html lang="zh"><head><meta charset="utf-8">',
           '<meta name="viewport" content="width=device-width,initial-scale=1">',
           '<title>', esc(title or "核算表预览"), '</title><style>',
           'body{font:12px/1.35 "Microsoft YaHei",微软雅黑,-apple-system,"Segoe UI",sans-serif;margin:14px 18px;color:#1a1a1a;background:#fff}',
           'table{border-collapse:collapse;border-spacing:0;font-size:12px}',
           'td{border:0;padding:1px 6px;white-space:nowrap;max-width:320px;overflow:hidden;text-overflow:ellipsis;vertical-align:middle;font-variant-numeric:tabular-nums}',
           '.hint{color:#8a94a0;font-size:12px;margin:0 0 10px}</style></head><body>']
    if title:
        out.append('<div class="hint">%s · 只读预览（与下载 xlsx 同版；xlsx 为活公式，预览为计算值）</div>' % esc(title))
    for ws in wb.worksheets:
        skip, span = set(), {}
        for mr in ws.merged_cells.ranges:
            span[(mr.min_row, mr.min_col)] = (mr.max_row - mr.min_row + 1, mr.max_col - mr.min_col + 1)
            for r in range(mr.min_row, mr.max_row + 1):
                for c in range(mr.min_col, mr.max_col + 1):
                    if (r, c) != (mr.min_row, mr.min_col):
                        skip.add((r, c))
        bars = {}                                                   # 条件格式数据条 → {(r,c): (颜色, 0~1)}
        try:
            for cf in ws.conditional_formatting:
                for rule in cf.rules:
                    if rule.type != "dataBar" or rule.dataBar is None:
                        continue
                    col = hexof(rule.dataBar.color) or "63BE7B"
                    cfv = list(rule.dataBar.cfvo or [])
                    lo = float(cfv[0].val) if (cfv and cfv[0].type == "num" and cfv[0].val is not None) else 0.0
                    hi = float(cfv[1].val) if (len(cfv) > 1 and cfv[1].type == "num" and cfv[1].val is not None) else 1.0
                    for rng in cf.sqref.ranges:
                        for r in range(rng.min_row, rng.max_row + 1):
                            for c in range(rng.min_col, rng.max_col + 1):
                                v = ws.cell(row=r, column=c).value
                                if isinstance(v, (int, float)) and not isinstance(v, bool) and hi > lo:
                                    bars[(r, c)] = (col, max(0.0, min(1.0, (float(v) - lo) / (hi - lo))))
        except Exception:
            pass
        maxr, maxc = ws.max_row or 0, ws.max_column or 0
        out.append('<table>')
        cols = []
        for c in range(1, maxc + 1):
            dim = ws.column_dimensions.get(get_column_letter(c))
            w = int(min(340, max(18, dim.width * 7.2))) if (dim and dim.width) else 0
            cols.append('<col style="width:%dpx">' % w if w else '<col>')
        out.append('<colgroup>' + ''.join(cols) + '</colgroup>')
        for r in range(1, maxr + 1):
            rd = ws.row_dimensions.get(r)
            h = rd.height if (rd is not None and rd.height) else None
            out.append('<tr%s>' % (' style="height:%dpx"' % round(h * 1.33) if h else ''))
            for c in range(1, maxc + 1):
                if (r, c) in skip:
                    continue
                cell = ws.cell(row=r, column=c)
                sp = span.get((r, c))
                attr = ''
                if sp:
                    if sp[0] > 1:
                        attr += ' rowspan="%d"' % sp[0]
                    if sp[1] > 1:
                        attr += ' colspan="%d"' % sp[1]
                st = style(cell, bool(sp and sp[1] > 1), bars.get((r, c)))
                if st:
                    attr += ' style="%s"' % st
                out.append('<td%s>%s</td>' % (attr, esc(fmt(cell.value, cell.number_format))))
            out.append('</tr>')
        out.append('</table>')
    out.append('</body></html>')
    return "".join(out)


@router.get("/api/bom/export/pretty")
async def bom_export_pretty(request: Request, entry_id: int, preview: int = 0):
    u = _require_perm(request, CAP_EXPORT)
    if not u:
        return JSONResponse({"ok": False, "msg": "无「导出」权限"}, status_code=403)
    e = db.bom_get_entry(entry_id)
    if not e or e.get("source") != _src():
        return JSONResponse({"ok": False, "msg": "记录不存在"}, status_code=404)
    # 下载=活公式版（改黄底参数全表联动，同原版）；预览=计算值版（openpyxl 写的公式无缓存值，网页只能走值）
    data = bq.build_pretty(_rec_from_entry(e), _fee_of(e), approval=e.get("approval_no") or "", formulas=not preview,
                           rules=_invoice_rules())      # 成本不含税公式/M列下拉 按台账当前发票规则生成
    nm = (e.get("product_name") or "").strip()
    if preview:
        return HTMLResponse(_xlsx_to_html(data, "重排版核算表 · %s %s" % (e.get("cp_code") or "", nm)))
    return _xlsx_response(data, "重排版核算表_%s_%s.xlsx" % (e.get("cp_code") or "", nm))


@router.get("/api/bom/export/original")
async def bom_export_original(request: Request, entry_id: int, preview: int = 0):
    u = _require_perm(request, CAP_EXPORT)
    if not u:
        return JSONResponse({"ok": False, "msg": "无「导出」权限"}, status_code=403)
    e = db.bom_get_entry(entry_id)
    if not e or e.get("source") != _src():
        return JSONResponse({"ok": False, "msg": "记录不存在"}, status_code=404)
    pdir = os.path.join(UPLOAD_DIR, _src())
    match = None
    if os.path.isdir(pdir):
        for fn in os.listdir(pdir):
            if fn.startswith("%d__" % entry_id) and "__商品版__" not in fn:   # 优先商务版底稿，跳过商品版留档
                match = os.path.join(pdir, fn)
                break
    if not match:
        return JSONResponse({"ok": False, "msg": "源附件未留档（样例种子或旧记录可能无原文件）。可导出重排版核算表替代。"}, status_code=404)
    data = open(match, "rb").read()
    fn = os.path.basename(match).split("__", 1)[-1]
    if preview:
        return HTMLResponse(_xlsx_to_html(data, "原版核算表 · " + fn))
    return _xlsx_response(data, fn)


# ============ BP 消费口：定稿版分项（handoff GET /cost-ledger/final?channel=）============
@router.get("/api/bom/final")
async def bom_final_feed(request: Request):
    """BP 定价测算拉**终审版**：每(产品)取定稿指针那一版的含税分项。channel 可筛。共享门户登录态即可读。
    ⚠ **只放终审通过（status=已审核）的**（业务方定 2026-09-04：只有终审的才对外开放）——
    成本会计已初审但财务BP尚未终审的版本，绝不外发给 BP 报价。"""
    u = _current_user(request)
    if not u:
        return JSONResponse({"ok": False, "msg": "未登录"}, status_code=401)
    channel = (request.query_params.get("channel") or "").strip()
    src = _src()
    served, _ = _live_finals(src)               # 已审核指针 且 未被当前对外新版替代（换码承接 V2.440）
    out = []
    for e in served:
        pkey = e["product_key"]
        if channel and (e.get("channel") or "") != channel:
            continue
        comp = bq.compose(_rec_from_entry(e), _fee_of(e))
        out.append({"entryId": e["id"], "productKey": pkey, "cpCode": e.get("cp_code"),
                    "erpCode": e.get("erp_code"), "productName": (e.get("product_name") or "").strip(),
                    "customer": e.get("customer") or "", "channel": e.get("channel") or "",
                    "channelLabel": CH_LABELS.get(e.get("channel") or "", ""), "calcDate": e.get("calc_date"),
                    "approvalNo": e.get("approval_no") or "",
                    # 类型/报价建议以成本会计审核定性为准（编码规律不固定）
                    "kind": _kind_of(e), "semi": _kind_of(e) != "成品",
                    "matCategory": e.get("mat_category") or "", "outsourced": bq.is_outsourced(e.get("mat_category")),
                    "quotable": (None if e.get("quotable") is None else bool(e.get("quotable"))),
                    "quoteReason": e.get("quote_reason") or "",
                    "material": comp["mat"], "packaging": comp["pack"], "processing": comp["mfg"],
                    "loading": comp["load"], "admin": comp["adm"], "fullCostIncl": comp["full"],
                    "finalizedBy": e.get("finalized_by") or "", "finalizedAt": e.get("finalized_at") or ""})
    return {"ok": True, "source": src, "channel": channel, "items": out}


# ============ BP 消费接口（对接需求 2026-09-05 §1–§3）：GET /api/bomcost/final ============
# 与上面 /api/bom/final 同一选择口径（定稿指针 + 只放终审「已审核」），字段/参数按 BP 文档契约；
# 服务间调用用内部令牌（BP 后端同机回环调），门户登录用户亦可读。/api/bom/final 保持不动。
def _internal_token():
    """BOMCOST_INTERNAL_TOKEN：先看进程环境变量（与 DB_URL 同机制），没有则读 backend/.env 的 KEY=VALUE。凭据不进代码不进文档。"""
    t = os.environ.get("BOMCOST_INTERNAL_TOKEN", "").strip()
    if t:
        return t
    try:
        p = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), ".env")
        if os.path.exists(p):
            for line in open(p, encoding="utf-8"):
                line = line.strip()
                if line.startswith("BOMCOST_INTERNAL_TOKEN="):
                    return line.split("=", 1)[1].strip().strip('"').strip("'")
    except Exception:
        pass
    return ""


def _internal_or_user(request):
    """放行二选一：①X-Internal-Token 正确 **且** 来源回环地址（BP 后端同机调用）；②门户登录用户。→ 身份 dict 或 None。"""
    tok = (request.headers.get("X-Internal-Token") or "").strip()
    want = _internal_token()
    host = (getattr(getattr(request, "client", None), "host", "") or "")
    if tok and want and tok == want and host in ("127.0.0.1", "::1", "localhost"):
        return {"name": "bp-internal", "internal": True}
    return _current_user(request)


_CH_BY_LABEL = {v: k for k, v in CH_LABELS.items()}     # 电商/通品/TOB/TOC → ecom/common/tob/toc


@router.get("/api/bomcost/final")
async def bomcost_final(request: Request):
    """BP 定价测算消费接口（契约：对接需求_BP消费接口_20260905 §1）。只返回**终审通过(status=已审核)**的定稿版，
    同一 productKey 一版（定稿指针即最新定稿）。数值含税 元/kg，与列表页同口径。
    Query：channel（电商/通品/TOB/TOC 或 ecom/common/tob/toc）· since（与 asOf/finalizedAt 同格式字符串，只回 finalizedAt>since）·
          erpCodes（逗号分隔，按需拉少量）。
    口径：finalizedAt/By＝**终审**（成为 final 对外的时刻，since 按它增量）；firstReviewedAt/By＝初审另给。
          erpCode 空照返 ""，BP 跳过并告警，两边不猜。netWeightKg=单位净重(kg)，null=未填（老定稿行可能没有）。
          transIncl 恒 0：台账「装卸费」已含运输（源表运输+装卸合并入账），不单拆。"""
    who = _internal_or_user(request)
    if not who:
        return JSONResponse({"ok": False, "msg": "未授权：需 X-Internal-Token（回环调用）或门户登录"}, status_code=401)
    qp = request.query_params
    ch_in = (qp.get("channel") or "").strip()
    ch = _CH_BY_LABEL.get(ch_in, ch_in)
    since = (qp.get("since") or "").strip()
    codes = {c.strip() for c in (qp.get("erpCodes") or "").split(",") if c.strip()}
    src = _src()
    # 换码承接（业务方定 2026-09-05，V2.440）：只发**当前版**——被一条已终审新版替代的旧版不再出现；
    # 新版带 supersedes=[它替代的旧版]，BP 拿去翻「引用旧 CP 的定价方案」重新提示。
    served, entries = _live_finals(src)
    # 兜底去重：同一物料编码若仍有多条对外版（老数据没标替代关系 / 两边都没走确认）→ 按审核时间只发最新，其余列进 warnings，
    # 让成本会计回核算侧补「补物料编码」确认；BP 永远不会收到同一编码两条。
    by_erp, warnings, dropped = {}, [], {}
    for e in served:
        erp = (e.get("erp_code") or "").strip()
        if erp:
            by_erp.setdefault(erp, []).append(e)
    for erp, es in by_erp.items():
        if len(es) > 1:
            es.sort(key=lambda x: (_audit_at(x), x["id"]))
            dropped[es[-1]["id"]] = es[:-1]
            warnings.append("物料编码 %s 有 %d 个对外版本（%s），已按审核时间只发 %s；请成本会计在核算侧确认替代关系（补物料编码时会提示）"
                            % (erp, len(es), "/".join((x.get("cp_code") or "").strip() for x in es),
                               (es[-1].get("cp_code") or "").strip()))
    drop_ids = {x["id"] for lst in dropped.values() for x in lst}

    def _sup(x):
        return {"entryId": x["id"], "cpCode": (x.get("cp_code") or "").strip(), "productKey": x.get("product_key"),
                "erpCode": (x.get("erp_code") or "").strip(), "finalizedAt": (x.get("ack") or {}).get("at") or "",
                "obsoleteAt": x.get("obsolete_at") or ""}
    rows = []
    for e in served:
        if e["id"] in drop_ids:
            continue
        pkey = e["product_key"]
        if ch and (e.get("channel") or "") != ch:
            continue
        erp = (e.get("erp_code") or "").strip()
        if codes and erp not in codes:
            continue
        ack = e.get("ack") or {}
        fin_at = ack.get("at") or ""
        if since and not (fin_at > since):
            continue
        comp = bq.compose(_rec_from_entry(e), _fee_of(e))
        nw, nw_src = _net_weight(e)
        sups = [_sup(x) for x in entries.values() if x.get("obsolete_by") == e["id"]] + [_sup(x) for x in dropped.get(e["id"], [])]
        rows.append({
            "supersedes": sups,
            "entryId": e["id"], "productKey": pkey, "erpCode": erp,
            "cpCode": e.get("cp_code") or "", "productName": (e.get("product_name") or "").strip(),
            "customer": e.get("customer") or "",
            "channel": CH_LABELS.get(e.get("channel") or "", e.get("channel") or ""),
            "packSpec": e.get("pack_spec") or "", "netWeightKg": nw, "netWeightSrc": nw_src,
            "unit": "元/kg", "taxIncluded": True,
            "matIncl": comp["mat"], "packIncl": comp["pack"], "mfgIncl": comp["mfg"],
            "transIncl": 0.0, "loadIncl": comp["load"], "admIncl": comp["adm"], "fullIncl": comp["full"],
            "calcDate": e.get("calc_date") or "", "approvalNo": e.get("approval_no") or "",
            "status": "final",
            "finalizedBy": ack.get("by") or "", "finalizedAt": fin_at,
            "firstReviewedBy": e.get("finalized_by") or "", "firstReviewedAt": e.get("finalized_at") or "",
            "quotable": (None if e.get("quotable") is None else bool(e.get("quotable"))),
            "quoteReason": e.get("quote_reason") or "",
        })
    rows.sort(key=lambda r: (r["finalizedAt"], r["entryId"]))
    from core import _now
    return {"ok": True, "asOf": _now(), "count": len(rows), "rows": rows, "warnings": warnings}


# ============ 样例数据种子（仅本机演示，绝不入库/入 git）============
# 用途：成本会计一进页面就看到可点的台账。数据源＝交接夹「样例数据/」的真实附件——
#   **只在该目录存在时种**（开发/演示机有，服务器没有 → 服务器保持空台账，等真取数）。
#   种进的行落本地 sample_data/workbench.db（已 gitignore），源附件复制进 bom_uploads/（已 gitignore）。
#   0 数据上 git 的红线由此双保险：不 commit 任何解析结果，也不 commit 任何附件。
_SAMPLE_SRC = os.environ.get("BOM_SAMPLE_DIR",
                             r"D:\0 Claude 数据\03 财务核算工作台\_交接_成本台账_20260902\样例数据")
# 审批编号按文件名 token 映射（交接文档 §7）。核算表用「8-20」式、BOM清单用「20260820」式，两种都登记，
# 好让核算表与其配套 BOM清单 归到同一审批（同版本配对，不跨版本串台）。
_SAMPLE_APPROVALS = [("8-20", "202608201745000128676"), ("20260820", "202608201745000128676"),
                     ("8-28", "202608261641000144524"), ("20260826", "202608261641000144524"),
                     ("6-02", "202606011017000005185"), ("0529", "202606011017000005185"),
                     ("20260529", "202606011017000005185")]


def _sample_approval(src_file):
    for tok, no in _SAMPLE_APPROVALS:
        if tok in (src_file or ""):
            return no
    return ""


def seed_bom_sample():
    """首次且样例目录存在时，把 3 单真实审批的解析结果种入 sample 源。已有数据/无目录 → 跳过。"""
    try:
        if db.bom_list_entries("sample"):
            return {"seeded": 0, "note": "已有样例数据"}
        if not os.path.isdir(_SAMPLE_SRC):
            return {"seeded": 0, "note": "无样例目录（服务器正常态）"}
        import glob
        pdir = os.path.join(UPLOAD_DIR, "sample")
        os.makedirs(pdir, exist_ok=True)
        # 研发 BOM清单 **按审批分组**（不能全池混——跨版本同产品会串台）；入账时只在同审批批次内配对。
        bom_by_appr = {}
        for bp in sorted(glob.glob(os.path.join(_SAMPLE_SRC, "*BOM清单*.xlsx"))):
            ap = _sample_approval(os.path.basename(bp))
            try:
                raw = open(bp, "rb").read()
                bl = bq.parse_bom_list(raw, os.path.basename(bp))
                try:
                    cf = bq.parse_craft(raw, os.path.basename(bp))
                except Exception:
                    cf = None
                for b in bl:
                    b["craft"] = cf
                bom_by_appr.setdefault(ap, []).extend(bl)
            except Exception:
                pass
        # 一遍：按数字指纹归组，留**最全排版版**（精简版缺物料编码，价格校验按编码查金蝶，必须留完整版）
        best = {}   # fp_hash -> dict(rec/comp/approval/fname/path/rich)
        for fp in sorted(glob.glob(os.path.join(_SAMPLE_SRC, "*成本核算表*.xlsx"))):
            data = open(fp, "rb").read()
            fname = os.path.basename(fp)
            approval = _sample_approval(fname)
            try:
                recs = bq.parse_workbook(data, fname)
            except Exception:
                continue
            for rec in recs:
                if not bq.all_checks_ok(rec):
                    continue
                comp = bq.compose(rec)
                fp_hash = _num_fp(rec, comp)
                rich = bq.richness(rec)
                if fp_hash not in best or rich > best[fp_hash]["rich"]:
                    best[fp_hash] = {"rec": rec, "comp": comp, "approval": approval,
                                     "fname": fname, "path": fp, "rich": rich}
        # 组锚：与入账一致，**同一核算表文件的产品共用 group_id**（成品归组键作锚），
        # 否则同文件被拆成多组，处理页重解析时会把别组的产品误报成「未入账」。
        wb_recs = {}
        for b in best.values():
            wb_recs.setdefault((b["approval"], b["fname"]), []).append(
                {"productKey": bq.product_key(b["rec"]), "semi": bq.is_semi(b["rec"].get("cpCode"))})
        gid_of_wb = {k: _group_id("sample", k[0], _workbook_anchor(v)) for k, v in wb_recs.items()}
        # 二遍：插入每组最全的那版
        n = 0
        for fp_hash, b in best.items():
            rec, comp, approval, fname = b["rec"], b["comp"], b["approval"], b["fname"]
            fee = comp["srcFee"]
            eid = db.bom_insert_entry({
                "source": "sample", "product_key": bq.product_key(rec), "cp_code": rec.get("cpCode"),
                "bom_list": (bq.match_bom_entry(rec, bom_by_appr.get(approval, [])) or {}).get("materials"),
                "craft": (bq.match_bom_entry(rec, bom_by_appr.get(approval, [])) or {}).get("craft"),
                "erp_code": rec.get("erpCode"), "product_name": (rec.get("productName") or "").strip(),
                "customer": rec.get("customer") or "", "pack_spec": rec.get("packSpec") or "",
                "supplier": rec.get("supplier") or "", "calc_date": rec.get("calcDate") or "",
                "order_qty": rec.get("orderQty"), "channel": _default_channel(rec),
                "semi": 1 if bq.is_semi(rec.get("cpCode")) else 0, "mat_subtotal_excl": rec.get("matSubtotal"),
                "pack_subtotal_excl": rec.get("packSubtotal"), "fee_mfg": fee["mfg"], "fee_load": fee["load"],
                "fee_adm": fee["adm"], "full_cost_incl": comp["full"], "src_full": comp["srcFull"],
                "src_fee": comp["srcFee"], "summary": rec.get("summary"), "materials": rec.get("materials"),
                "checks": rec.get("checks"), "num_fp": fp_hash, "source_type": "dingtalk_form",
                "origin": "procurement", "src_label": "成本核算表（商务输出）",
                "group_id": gid_of_wb.get((approval, fname)), "active": 1,
                "approval_no": approval, "src_file": fname, "sheet": rec.get("sheet"),
                "status": "未复核", "created_by": "样例种子",
            })
            try:
                import shutil as _sh
                _sh.copy2(b["path"], os.path.join(pdir, "%d__%s" % (eid, fname)))
            except Exception:
                pass
            n += 1
        return {"seeded": n}
    except Exception as e:
        return {"seeded": 0, "note": "seed 异常：%s" % e}


try:
    seed_bom_sample()
except Exception:
    pass

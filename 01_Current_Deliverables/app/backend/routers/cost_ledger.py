# -*- coding: utf-8 -*-
# [Change Log]
# Date: 2026-08-05 | Author: Claude / c | Version: V2.172
# Description: 【成本台账】路由（V2.172 从 app.py 拆出）。
#              本文件是「成本台账」这条工具线在后端的唯一落点：改这条线的接口只动本文件，
#              不再碰 app.py —— 这样多条需求并行开发时互不冲突。
#              共享的配置/期间/权限判定见 core.py；算法在 kernels/cost_ledger.py。
#              app.py 只负责 include_router(router)，不感知本文件内部。

from fastapi import APIRouter
from fastapi import Request
from fastapi.responses import Response
from io import BytesIO
import base64
import gzip
import json
from kernels import cost_ledger as clg
import kingdee_client as kc
import urllib.parse

from core import (
    BASE, JSONResponse, _current_user, _now, _require_perm, db, os,
)

router = APIRouter()


# ============ 成本台账（存货月结核对，V2.58 前端接线）============
_CL_CFG_PATH = os.path.join(BASE, "sample_data", "cost_ledger_config.json")

# V2.122 期间化 + 封存。此前用进程内全局单槽缓存（_COST_LEDGER_CACHE），在【多人可浏览】下是错的：
#   不带期间键 → A 跑 5 期、B 跑 4 期会互相覆盖，A 点导出拿到 B 的 4 期数据（屏幕与导出对不上）；
#   不落盘 → 服务重启谁都导不出；仓库类型页的"本期有数据"跟着最后一个人跑的结果走。
# 改为复用平台现成机制：period_inputs（按期存输入·全员共享·重启不丢）+ period_snapshots（封存拍照）
#   + periods（封存状态·解封留痕）。
# **source 命名空间**：periods/period_snapshots/period_inputs 都按 source 隔离，而 periods 的唯一键是
#   (source,year,period)、不带工具维度；改表要重建（SQLite 尤其麻烦）且动的是银行对账的活数据。
#   故成本台账用【source = "cl:<账簿代码>"】自成封存线（业务方定：各封各的）——零改表、零影响银行对账。
#   不跟 CFG["source"] 走：成本台账不论平台切 sample 还是 kingdee 都真取金蝶，跟着切会把同一份
#   真数据劈到两个命名空间里。
#
# V2.126 多主体：原先把 107（孝感星期九）写死在六处。业务方指出「我这里还有不同主体」，实测金蝶 9 个
#   核算组织中【只有 107(1099万) 与 101(9,446) 有存货】，105/103 Sinkio/108 星期十 的 14xx 余额均为 0，
#   102/104/106/109 连 14xx 科目都没有。101 实测把 org 参数化即跑通（账实三科目 diff 全 0、自平通过），
#   核算体系/会计政策/币别三个常量跨主体通用。主体清单读【平台主体档案 orgs 表】（V2.125 已补齐 9 个），
#   不在本工具里写死。**每个主体一条独立封存线**：source = "cl:101" / "cl:107" …
_CL_INPUT = "cl:input"          # period_inputs.kind：本期输入（收发存/按日期/损益页/科目余额）
_CL_SNAP = "cost_ledger"        # period_snapshots.kind：封存快照（整份结果）
_CL_DEFAULT_ORG = "107"         # 默认主体＝孝感星期九（存货体量最大的那个）


def _cl_source(org):
    return "cl:" + str(org)


def _cl_orgs():
    """成本台账可选主体 → [{code(账簿代码), name(简称), full(全称)}]，取自平台主体档案（只列启用）。"""
    return [{"code": o["book_code"], "name": o["short_name"], "full": o["full_name"]}
            for o in db.list_orgs() if o.get("active") and o.get("book_code")]


def _cl_org(code):
    """账簿代码 → 主体档案那一行；查不到返回 None（路由据此拒绝，不猜）。"""
    for o in _cl_orgs():
        if str(o["code"]) == str(code):
            return o
    return None


_CL_WH_KEY = "cost_ledger_wh_attr"        # app_settings 键：仓库类型对照（业务方在「基础资料」页维护）
# app_settings 键：类别↔科目对照（V2.254 从 json 搬进数据库，同页维护）
# 值＝{"category_to_subject": {科目: [类别,…]}, "extra_subjects": [科目,…]}
_CL_CAT_KEY = "cost_ledger_cat_subject"


# ⚠**代码错误不许降级成提示**（V2.286 血的教训）。
# 取数处一律写 `except Exception` 曾把 AttributeError 一并吞掉：V2.280 改多期取数时用字符串切片
# 替换函数体，把夹在中间的 5 个函数（按日期表/成本计算单/成本科目/损益明细）**一起删掉了**，
# 而路由把 `module has no attribute 'fetch_pnl_details'` 显示成"损益归集取数失败"这一行黄字——
# 看着像金蝶那边的问题，实则是本地代码没了。V2.280–284 的包都带着这个残缺发出去过。
# 从此：金蝶/网络/数据问题 → 软失败给提示；**AttributeError/TypeError/NameError/ImportError
# ＝代码坏了 → 直接抛，让它在测试和页面上炸出来**。
_CODE_BUGS = (AttributeError, TypeError, NameError, ImportError)


class ClgBookMismatch(Exception):
    """账簿认不出来（V2.137 建，V2.253 起基本不会触发）。

    V2.137 的原始场景：靠【主体档案全称】精确匹配账簿，差一个字就全量跳过。
    V2.253 改为按【账簿代码】在金蝶端过滤后，全称不再参与匹配，这一类失败随之消失；
    异常型保留，只在"账簿代码本身在金蝶查无此账簿"时抛——那是档案里的代码填错了，
    仍需把"档案里写的"与"金蝶里有哪些"一起摆出来给人看，别丢一句让人猜。"""


class ClgConfigMissing(Exception):
    """成本台账配置文件缺失。单独立一个异常型，配 exception_handler 出可读提示——
    V2.136 前它是裸的 FileNotFoundError，且 `_cl_config()` 是在路由 try/except **之外**
    被 `_cl_payload` 调到的 → 前端只看到一个光秃秃的 500，查不出所以然。
    真实事故：历次部署包都整体排除 sample_data/（那目录混装了运行数据与本配置），
    这个文件从没上过服务器 → 服务器上成本台账一点取数就 500。"""


# ── 大块明细的压缩存取（V2.278）──────────────────────────────────
# 业务方定「导出＝结论 + 原始底表」，故原始明细必须**跟着输入一起落库**：
# 若改成导出时回金蝶重取，已封存期间导出的明细可能与当初核对的那份不一样——封存就白做了。
# 但成本计算单一期 7,794~16,379 行、裸 JSON 2.07 MB；gzip 后 0.09 MB（**压到 5%**），
# 故一律压着存。只压这几块大的，既有的 cross/gl 不动（改动既有存量的风险不值当）。
def _pack(rows):
    if not rows:
        return None
    return {"_z": base64.b64encode(gzip.compress(
        json.dumps(rows, ensure_ascii=False).encode("utf-8"), 6)).decode("ascii")}


def _unpack(v):
    """吃得下两种：压缩过的 {"_z":...} 与早期直接存的裸 list（向后兼容，老数据不用迁）。"""
    if isinstance(v, dict) and "_z" in v:
        return json.loads(gzip.decompress(base64.b64decode(v["_z"])).decode("utf-8"))
    return v


def _cl_config():
    """存货台账配置。**数据库为唯一真相源，json 只是首次使用的种子**（V2.119 仓库类型 / V2.254 类别科目）。
    业务方在「基础资料」页存过一次后，改 json 不再影响运行（避免"两个地方都能改、不知道谁生效"）。

    V2.254：`category_to_subject`/`extra_subjects` **补上数据库兜底**，与仓库类型同规格。
    在此之前这两项只有 json 一个来源——而部署包历来整体排除 `sample_data/`（那目录混装运行数据），
    该文件从没上过服务器 → 服务器上一点取数就 500（V2.132 真实事故）。
    现在：数据库里存过 → 文件缺了也照常跑；两处都没有 → 仍然报 ClgConfigMissing，
    因为没有类别↔科目对照就没法把存货类别归到总账科目，账实勾稽无从算起，**只能报错、不能猜**。"""
    cat_ov = db.get_setting(_CL_CAT_KEY)
    wh_ov = db.get_setting(_CL_WH_KEY)
    has_db_cat = isinstance(cat_ov, dict) and cat_ov.get("category_to_subject")
    if os.path.exists(_CL_CFG_PATH):
        cfg = clg.load_config(_CL_CFG_PATH)
    elif has_db_cat:
        cfg = {"category_to_subject": {}, "extra_subjects": [], "warehouse_attr": {}}
    else:
        raise ClgConfigMissing(_CL_CFG_PATH)
    if has_db_cat:
        cfg["category_to_subject"] = {k: list(v) for k, v in cat_ov["category_to_subject"].items() if v}
        cfg["extra_subjects"] = list(cat_ov.get("extra_subjects") or [])
    if isinstance(wh_ov, dict) and wh_ov.get("map") is not None:
        cfg["warehouse_attr"] = {k: v for k, v in wh_ov["map"].items() if v}
    return cfg


def _cl_period_str(y, p, org=None):
    o = _cl_org(org) if org else None
    return "%s%d年%d期" % ((o["name"] + " ") if o else "", int(y), int(p))


def _cl_closed(y, p, org):
    return db.is_closed(_cl_source(org), y, p)


def _cl_closed_block(y, p, org):
    """已封存期间的取数/上传一律拦下。返回 JSONResponse=拦截；None=放行。"""
    if not _cl_closed(y, p, org):
        return None
    i = db.get_period(_cl_source(org), y, p)
    return JSONResponse({"ok": False, "closed": True,
                         "msg": "%s 已封存（%s 于 %s 封存），本期只读。如需重取数，请先解封。"
                                % (_cl_period_str(y, p, org), i.get("封存人", "") or "?", i.get("封存时间", "") or "?")},
                        status_code=409)


def _cl_save_input(y, p, org, parts, gl, src, operator):
    """把本期【输入】整体落库（全员共享、重启不丢）。结果不存——由输入现算，避免结果与输入不同源。"""
    _o = _cl_input(y, p, org)
    old = (_o or {}).get("payload") or {}
    payload = {"cross": parts["cross"], "bydate": parts.get("bydate"),
               # ⚠三档一起存：这份 payload 是**白名单**，漏一个键就等于那一档在落库里不存在，
               #   而症状是"页面上少一块"，不会报错（V2.312 就是这么丢的「其他」档）。
               "loss": parts.get("loss"), "disp": parts.get("disp"), "oth": parts.get("oth"),
               "exc": parts.get("exc"),
               "btypes": parts.get("btypes"), "cost": parts.get("cost"),
               # V2.278 原始底表（压缩存）：导出「结论 + 原始底表」要用，且封存后必须能原样复现
               "cc_rows": _pack(parts.get("cc_rows")),
               "wip_rows": _pack(parts.get("wip_rows")),
               # **两条通道互不覆盖**——金蝶取数不带这块，若已传过就原样留着，
               # 否则"先传原表、再点一键取数"会把原表默默清掉（用户看不出来，最恶心的那种丢数据）。
               # 本期物料档案映射（V2.282）：下期比对用，当月就能发现"有人改了档案"
               "cat_map": parts.get("cat_map"),
               "pnl_detail": parts.get("pnl_detail"),
               "gl": gl, "src": src}
    meta = {"src": src, "rows_cross": len(parts["cross"]),
            "rows_bydate": len(parts.get("bydate") or []),
            "has_gl": bool(gl), "nomap": parts.get("nomap", 0), "at": _now()}
    db.set_period_input(_cl_source(org), y, p, _CL_INPUT, payload, meta, operator)
    return meta


def _cl_input(y, p, org):
    return db.get_period_input(_cl_source(org), y, p, _CL_INPUT)


def _cl_period_data_status(y, p, org):
    """成本台账某主体某期的数据状态——【说法沿用期间选择器那一套】：
    已封存 / 数据已上传 / 数据未上传。
    页面内文案叫"数据已接入/未接入"（那边讲的是取数与上传两个通道），
    但选择器是全站共用组件、颜色按这三个词分，这里就按它的词回答，不另造一套。"""
    y, p = int(y), int(p)
    if _cl_closed(y, p, org):
        return "已封存"
    return "数据已上传" if _cl_input(y, p, org) else "数据未上传"


def _cl_register_warehouses(y, p, org, cross):
    """新仓库【自动上档】（V2.124）：本期数据里出现、但仓库类型台账从没见过的仓库，
    自动登记进 app_settings 的 seen（记首次出现期间），供「仓库类型」页标「新」并优先催配。
    **只上档、不猜类型**——类型仍留空由成本会计填：猜错了会把钱默默归进错的仓库类型小计里，
    比不归更危险（同「对照缺失不硬归」的一贯口径）。"""
    whs = {r["wh"] for r in cross if r.get("wh")}
    if not whs:
        return []
    cur = db.get_setting(_CL_WH_KEY) or {}
    seen = dict(cur.get("seen") or {})
    fresh = sorted(w for w in whs if w not in seen)
    if not fresh:
        return []
    tag = "%d-%02d" % (int(y), int(p))
    for w in fresh:
        seen[w] = tag
    cur["seen"] = seen
    db.set_setting(_CL_WH_KEY, cur, operator="auto")
    return fresh


def _cl_missing_wh(res):
    """本期有数据、但没配仓库类型的仓库 → 硬校验用（卡封存、卡导出）。"""
    return list(((res or {}).get("pivot_wh_type") or {}).get("missing_attr") or [])


def _cl_result(y, p, org):
    """该主体本期结果 → (res, cross, meta, from_snapshot)；没数据 → (None,...)。
    已封存 → 读快照（拍照存死的那一版，不再碰金蝶、也不受配置改动影响）；
    进行中 → 由 period_inputs 现算（仓库类型对照改了，重进页面即刻生效）。"""
    if _cl_closed(y, p, org):
        d = db.load_snapshot(_cl_source(org), y, p, _CL_SNAP)
        if d:
            return d.get("res"), d.get("cross"), d.get("meta") or {}, True
        return None, None, {}, True        # 已封存但快照缺失（异常）→ 由调用方提示
    inp = _cl_input(y, p, org)
    if not inp:
        return None, None, {}, False
    d = inp["payload"]
    # 上期的物料档案映射：跨年时回到上一年 12 期（V2.282）
    py, pp = (y - 1, 12) if int(p) == 1 else (y, int(p) - 1)
    prev = _cl_input(py, pp, org)
    prev_map = (prev["payload"].get("cat_map") if prev else None) or None
    res = clg.build_cost_ledger(d["cross"], _cl_config(), gl_balance=d.get("gl"),
                                bydate=d.get("bydate"), loss_rows=d.get("loss"),
                                disposal_rows=d.get("disp"), other_rows=d.get("oth"),
                                excluded_rows=d.get("exc"),
                                cost_block=d.get("cost"),
                                prev_cat_map=prev_map, prev_label=f"{py}年{pp}期")
    # 事务类型汇总（V2.141）＝取数时已在服务端聚合好的 ~21 行，挂进 res：
    # 封存快照存的就是 res，自然一起拍进去，不用另改快照结构。None＝该期没取过（旧数据或🅱上传通道）。
    res["btypes"] = d.get("btypes")
    # 原始底表只挂在 res 上给导出用，**不进 _cl_payload**（前端不需要，几 MB 塞过去是灾难）
    res["_raw"] = {"bydate": d.get("bydate"), "cc_rows": _unpack(d.get("cc_rows")),
                   "wip_rows": _unpack(d.get("wip_rows")),
                   "pnl_detail": d.get("pnl_detail")}
    # 前端只要知道"原表传没传、多少行"，不要把明细塞过去
    # 行数按**整页**报（元数据行 + 表头 + 数据，含中间空行）——页面上写的数要和打开 Excel 看到的一致
    meta = dict(inp["meta"] or {})
    meta.update({"year": y, "period": p, "org": org, "source": d.get("src"),
                 "updated_by": inp["updated_by"], "updated_at": inp["updated_at"]})
    return res, d["cross"], meta, False


def _cl_payload(y, p, org, user=None):
    """给前端的一整份本期状态（页面进入/切主体或期间直接读，秒开、全员看到同一份）。
    user 传入时附当前用户的能力位——前端据此决定给不给按钮（不给点，而非点了才被拒）。"""
    o = _cl_org(org)
    closed = _cl_closed(y, p, org)
    info = db.get_period(_cl_source(org), y, p) if closed else {"status": "open", "已封存": False}
    res, cross, meta, snap = _cl_result(y, p, org)
    out = {"ok": True, "year": y, "period": p, "closed": closed, "period_info": info,
           "org": org, "org_name": (o or {}).get("name", org), "org_full": (o or {}).get("full", ""),
           "from_snapshot": snap, "has_data": res is not None,
           "can": {"fetch": bool(user and db.user_can(user, "cost_ledger_fetch")),
                   "close": bool(user and db.user_can(user, "cost_ledger_close")),
                   "wh": bool(user and db.user_can(user, "cost_ledger_wh")),
                   "reopen": bool(user and (db.is_super(user) or db.user_can(user, "manage_accounting")))}}
    if res is None:
        out["data_status"] = "已封存·快照缺失" if closed else "数据未接入"
        return out
    an = res["anomalies"]
    miss = _cl_missing_wh(res)
    out.update({"data_status": "已封存" if closed else "数据已接入",
                "missing_wh": miss, "can_seal": not miss,      # 硬校验：缺仓库类型 → 不能封存/导出
                "source": meta.get("source"), "meta": meta,
                "rows_cross": len(cross), "rows_bydate": meta.get("rows_bydate", 0),
                "nomap": meta.get("nomap", 0), "has_gl": meta.get("has_gl", False),
                "credible": res["credible"], "ties": res["ties"],
                "pivot_category": res["pivot_category"], "pivot_wh_type": res["pivot_wh_type"],
                "pivot_wh_category": res["pivot_wh_category"], "btypes": res.get("btypes"),
                "anomaly_counts": an["counts"], "anomaly_items": an["items"], "pnl": res.get("pnl"),
                # 制造费用（V2.257）：三道成本勾稽 + 车间/费用项目透视。取不到时为 None，
                # 前端据此显示"本期未取到"，不画空表。
                "cost": res.get("cost"),
                # 类别漂移（V2.282）：只报不判，不并进 credible
                "drift": res.get("drift")})
    return out


def _inv_gl_balance(year, period, org=_CL_DEFAULT_ORG):
    """某主体的存货类科目(14xx)期末余额，按(账簿,科目)去重 → {科目名称: 期末本位币}。
    去重口径同 V2.53 实证：GL_BALANCE 每(账簿×科目)返本位币汇总行+币别明细行，金额相同，留一行。

    V2.126：账簿改按【主体档案全称精确匹配】。原先是关键字 ("孝感","星期九") 模糊匹配——
    多主体下会误伤：「深圳市星期九」也含"星期九"，只是恰好不含"孝感"才没出事；
    而"深圳市星期零"与"深圳市星期零食品科技有限公司上海分公司"这种父子账簿，
    任何关键字写法都会把两个都圈进来。全称相等最稳。

    V2.137：全称匹配不上时**把话说清楚**（原先只回一句「未取到科目余额」，看不出是名字对不上
    还是这主体本来就没存货科目）。

    ⚠V2.253 起【不再用全称认账簿】——上面 V2.126/V2.137 那套已成历史，保留是为了讲清来龙去脉。
    2026-08-10 实证：GL_BALANCE 的 `FACCOUNTBOOKID.FNumber` 就是账簿代码，与主体档案 `book_code`
    同一套编码（107→孝感市星期九…），**直接在金蝶端按账簿代码过滤即可**。这条改动消掉了一个
    真实卡点：全称是主体档案里唯一要人工登记的字段，而建库种子把它插成空串（db.seed_orgs 只插
    3 条、全称留空）→ 任何新装的库开箱即坏、账实勾稽做不了，人工补一次、库一重建又掉一次。
    改后全称不再参与任何计算，只用于导出表头署名，空着也不影响勾稽。
    顺带：过滤下推到金蝶端，不再把 5 个账簿的 14xx 全拉回来在内存里挑。
    去重仍要保留——币别空的本位币汇总行与人民币明细行金额相同，不去重会翻倍（V2.53 实证）。"""
    o = _cl_org(org)
    if not o:
        raise ValueError(f"主体档案里没有账簿代码 {org}，请先到「基础数据 › 主体档案」建档")
    rows = kc.fetch_gl_balance(year, period, prefixes=("14",), book=org)
    if not rows:
        # 空有两种可能：①该主体本就没有存货类科目（正常，勾稽跳过）②档案里的账簿代码填错了。
        # 只在空的时候才多打一次账簿档案，把两种情况分开说——不让人对着"没数"猜。
        try:
            books = kc.fetch_account_books()
        except Exception:
            books = []
        if books and str(org) not in {str(b["code"]) for b in books}:
            raise ClgBookMismatch(
                "主体档案里「%s」登记的账簿代码是 %s，但金蝶没有这个账簿。\n"
                "金蝶现有账簿：%s。\n"
                "→ 到「基础数据 › 主体档案」把该主体的【账簿代码】改成金蝶里的那个。"
                % (o.get("name", org), org, "、".join(f"{b['code']} {b['name']}" for b in books)))
        return {}                        # 该主体没有 14xx 科目：勾稽跳过属正常
    seen, out = set(), {}
    for r in rows:
        book = str(r.get("账簿", "")).strip()
        code = str(r.get("科目编码", ""))[:4]
        key = (book, code)
        if key in seen:
            continue
        seen.add(key)
        nm = r.get("科目名称", "")
        out[nm] = out.get(nm, 0.0) + clg._num(r.get("期末本位币"))
    return out


def _find_header_idx(rows, *musts):
    for row in rows[:6]:
        cells = [clg._s(c) for c in row]
        if all(m in cells for m in musts):
            return {c: i for i, c in enumerate(cells)}
    return None


def _parse_loss_sheet(rows):
    """货损明细：需含 类别 + 金额 列。返回 [{cat,amount}]。"""
    idx = _find_header_idx(rows, "类别", "金额")
    if not idx:
        return None
    out = []
    for row in rows[1:]:
        if idx["类别"] >= len(row) or idx["金额"] >= len(row):
            continue
        cat = clg._s(row[idx["类别"]])
        if not cat:
            continue
        out.append({"cat": cat, "amount": clg._num(row[idx["金额"]])})
    return out


def _parse_disposal_sheet(rows):
    """资产处置(盘盈亏卡片)：需含 单据编号 + 金额，且非货损(无 类别 值)。返回 [{amount}]。"""
    idx = _find_header_idx(rows, "单据编号", "金额")
    if not idx:
        return None
    out = []
    for row in rows[1:]:
        if idx["金额"] >= len(row):
            continue
        amt = clg._num(row[idx["金额"]])
        no = clg._s(row[idx["单据编号"]]) if idx["单据编号"] < len(row) else ""
        if amt and no:
            out.append({"amount": amt})
    return out


_CC_UP_MUST = ("成本项目名称", "金额", "工单编号")






# 🅱 上传通道已于 V2.317 拆除（业务方：「删，因为那个通道只上传一个表格没啥用啊」）。
#   它只吃《存货收发存汇总表（跨维度）》一张表，而🅰一键取数本来就把这张连同另外五项一起取回来，
#   留着等于多一条要维护、要测、要在文档里解释的路。
#   ⚠**解析器 parse_cross_report 没删**——🅰 通道也走它（本文件 _cl_from_kingdee 里），删它会连金蝶取数一起废。
#   ⚠已落库的 source="upload" 历史期间**照常可读**：payload 结构与🅰完全一致，只是 meta.src 标记不同。





@router.post("/api/cost-ledger/analyze-kingdee")
def cost_ledger_analyze_kingdee(request: Request, year: int = 2026, period: int = 5, org: str = _CL_DEFAULT_ORG):
    """一键金蝶取数：存货收发存汇总表(跨维度) API 取数 + 科目余额 → 落库为本期输入 → 三道勾稽/透视/异常。
    V2.115/V2.116 起：剔小计判据订正 + 取到仓库维度(FStockId)，2026-5 期实测与成本会计底稿
    分毫不差（账实三科目 diff 全 0、44/44 仓库逐仓一致），与上传通道等价。"""
    # V2.122 补权限：此前本路由只要登录就能触发取数（上传通道却要 cost_ledger），两条通道口径不一致。
    u = _require_perm(request, "cost_ledger_fetch")
    if not u:
        return JSONResponse({"ok": False, "msg": "无取数权限（请管理员授予「存货台账·取数/上传」）——取数会覆盖全员可见的数据，故与查看分开授权"}, status_code=403)
    if not _cl_org(org):
        return {"ok": False, "msg": f"主体档案里没有账簿代码 {org}"}
    blk = _cl_closed_block(year, period, org)
    if blk:
        return blk
    try:
        inv = kc.fetch_inventory_summary(year, period, org=org)
    except kc.KingdeeError as e:
        return {"ok": False, "msg": f"金蝶取数失败：{e}"}
    except Exception as e:
        return {"ok": False, "msg": f"取数异常：{e}"}
    cross = inv["rows"]
    if not cross:
        return {"ok": False, "msg": f"{year} 年第 {period} 期收发存报表无数据"}
    gl, gl_msg = None, ""
    try:
        gl = _inv_gl_balance(year, period, org)
        if not gl:
            gl_msg = "未取到科目余额（账实勾稽跳过）"
    except ClgBookMismatch as e:
        gl_msg = str(e)                 # 已是完整诊断（两边名字都摆出来了），别再套前缀
    except Exception as e:
        gl_msg = f"科目余额取数失败（账实勾稽跳过）：{e}"
    # 勾稽①两表互勾（V2.255）：以前 🅰 通道跑不了这道，因为第二张【按日期表】只能从🅱上传的
    # 工作簿里拆。那不是技术限制、是没做——按日期表同样走 GetSysReportData（业务方 2026-08-10
    # 提供官方接口文档后实证：107/2026-5 四项金额与跨维度表**分毫不差**）。补上后 🅰 三道全跑。
    # 失败不拦取数：勾稽①缺席时内核照旧只跑后两道，页面会显示该道缺失，不伪装成通过。
    bydate, bd_msg = None, ""
    try:
        bydate = kc.fetch_inventory_bydate(year, period, org=org)
    except _CODE_BUGS:
        raise                                   # 代码坏了，别装成"金蝶取数失败"
    except Exception as e:
        bd_msg = f"按日期表取数失败（勾稽①两表互勾跳过）：{e}"
    # 损益归集（V2.256）：照成本会计的做法从【科目余额表下钻】取货损/处置明细。
    # 此前一直认为这两页是底稿里手工加的、🅰 通道永远做不了——**是错的**：它们来自
    # 管理费用/营业外支出科目的凭证分录，摘要里带单据号，机器能认，不需要人挑。
    # 实证 107：2026-3 货损 18,881.52＝底稿；2026-5 货损 -4,874.95、处置 8,912.37＝DoD 基准。
    # 失败不拦取数——损益归集缺席时第⑦步显示缺失，三道勾稽不受影响。
    loss_rows = disp_rows = oth_rows = exc_rows = None
    pnl_msg = ""
    try:
        pnl = kc.fetch_pnl_details(year, period, org=org)
        # 内核只要 {cat, amount}；cat 用单据类型（盘亏毁损单/其他入库单/卡片处置…），
        # 比底稿那列人工填的"原辅料盘盈亏/产品盘盈亏"更客观，且能追到单据号。
        # 分类用【费用项目】（产品货损/包材货损/原辅料盘盈亏…），与成本会计底稿的「类别」同口径；
        # 没挂费用项目的老凭证才退回单据类型
        loss_rows = [{"cat": x.get("item") or x["doctype"], "amount": x["amount"]} for x in pnl["loss"]]
        disp_rows = [{"amount": x["amount"]} for x in pnl["disposal"]]
        # 第三档也要给内核——它要进 res["pnl"]，前端只认那里（V2.312）
        oth_rows = [{"cat": x.get("item") or x["doctype"], "amount": x["amount"],
                     "acct": x.get("acct"), "acct_name": x.get("acct_name")} for x in (pnl.get("other") or [])]
        # 口径外（Owner 定案：6602 非货损的不算成本台账口径）——**不计入任何合计，但要看得见**
        exc_rows = [{"cat": x.get("item") or x["doctype"], "amount": x["amount"],
                     "acct": x.get("acct"), "acct_name": x.get("acct_name")} for x in (pnl.get("excluded") or [])]
        parts_pnl = pnl
    except _CODE_BUGS:
        raise
    except Exception as e:
        pnl_msg = f"损益归集取数失败（第⑦步将显示缺失）：{e}"
        parts_pnl = None
    # 成本计算单 + 成本类科目（V2.257）：制造费用三道勾稽的两端。
    # 失败不拦取数——制造费用与存货是两本账，取不到只是第⑧步不出，三道存货勾稽照跑。
    cc_recs = cost_gl = None
    cc_msg = ""
    try:
        cc_recs = kc.fetch_cost_calc(year, period, org=org)
        cost_gl = kc.fetch_cost_gl(year, period, org=org)
    except _CODE_BUGS:
        raise
    except Exception as e:
        cc_msg = f"成本计算单取数失败（制造费用勾稽将缺失）：{e}"
    # 事务类型汇总（V2.141）：流水级报表服务端聚合成 ~21 行再落库（不落 13,625 行原始流水）。
    # 失败不拦取数——事务类型是补充视角，透视屏会提示缺失，核心的三道勾稽不受影响。
    btypes, bt_msg = None, ""
    try:
        # 把损益归集查到的单据号一起递进去——流水表这一趟顺路把物料级明细也带回来，不多打接口
        bills = [x["billno"] for x in ((parts_pnl or {}).get("loss") or [])
                 + ((parts_pnl or {}).get("disposal") or [])
                 + ((parts_pnl or {}).get("other") or [])]      # V2.307 第三类也要回查物料级
        got = kc.fetch_business_type_summary(year, period, org=org, billnos=bills or None,
                                             wip_btypes=clg.WIP_CREDIT_BTYPES)
        if isinstance(got, tuple):
            btypes, pnl_rows, wip_rows = got
            if parts_pnl is not None:
                parts_pnl["rows"] = pnl_rows      # 物料级：编码/名称/规格/单位/仓库/数量/单价/金额
                # 单据备注（钉钉单号）＋单据分录行——业务方底稿「备注」列要的就是前者，
                # 后者用来补流水表不吐的零金额行（V2.314），一次回查两用、不多打接口
                try:
                    _bn = kc.fetch_bill_notes(bills)
                    parts_pnl["notes"] = _bn.get("notes") or {}
                    parts_pnl["rows"] = _merge_bill_only_rows(pnl_rows, _bn.get("lines") or {})
                except _CODE_BUGS:
                    raise
                except Exception:
                    parts_pnl["notes"] = {}       # 取不到就留空，不拦整条取数
        else:
            btypes, wip_rows = got, []
    except _CODE_BUGS:
        raise
    except Exception as e:
        bt_msg = f"事务类型取数失败（第④步透视将显示缺失）：{e}"
    parts = {"cross": cross, "bydate": bydate, "loss": loss_rows, "disp": disp_rows, "oth": oth_rows, "exc": exc_rows,
             "btypes": btypes, "nomap": inv["nomap"],
             # 成本结论块（不落 16,379 行原始明细，同事务类型的做法）
             "cost": clg.build_cost_block(cc_recs, cost_gl, btypes),
             "cat_map": clg.build_cat_map(cross),
             "cc_rows": cc_recs,
             # 勾稽②的业务侧底表（V2.310）：只存「汇报入库/生产退库/生产入库」那几种，7 月 225 行
             "wip_rows": wip_rows,
             # 明细留档：页面/导出要能列出单据号、凭证字号、摘要，光有合计追不回去
             "pnl_detail": parts_pnl}
    _cl_save_input(year, period, org, parts, gl, "kingdee", u["name"])
    fresh = _cl_register_warehouses(year, period, org, cross)
    db.audit(u["name"], "存货台账·一键金蝶取数", _cl_period_str(year, period, org),
             "跨维度 %d 行%s%s" % (len(cross), ("；按日期 %d 行" % len(bydate)) if bydate else "",
                                  ("；新仓库上档 %d 个" % len(fresh)) if fresh else ""))
    out = _cl_payload(year, period, org, u)
    out["new_wh"] = fresh
    out["gl_msg"] = gl_msg
    # 两条取数警告都要带出去，别只报一条——缺哪道勾稽得让人当场看见
    out["bd_msg"] = bd_msg
    if bt_msg:
        out["bt_msg"] = bt_msg
    if pnl_msg:
        out["pnl_msg"] = pnl_msg
    if cc_msg:
        out["cc_msg"] = cc_msg
    return out


def _cl_wh_types(attr):
    """仓库类型主档 → [{name, n(在用仓库数)}]。
    V2.121（业务方定稿）：类型就是一个**纯分类**——不带启用/禁用日期、不带操作，页面上做成横向切片器。
    仓库的启用/禁用日期与操作一并取消，合并为每个仓库一条**备注**（成本会计自由填写）。
    未建过主档时用现有对照里出现过的类型作种子；对照在用但主档缺失的类型也补进来（避免在用却不可见）。"""
    meta = db.get_setting(_CL_WH_KEY) or {}
    saved = meta.get("types")
    used_n = {}
    for t in attr.values():
        used_n[t] = used_n.get(t, 0) + 1
    names = []
    if isinstance(saved, list):
        for t in saved:                    # 兼容 V2.120 存过的 [{name,on,off}] 结构
            n = str((t.get("name") if isinstance(t, dict) else t) or "").strip()
            if n:
                names.append(n)
    else:
        names = sorted({v for v in attr.values() if v})
    out, seen = [], set()
    for n in names:
        if n in seen:
            continue
        seen.add(n)
        out.append({"name": n, "n": used_n.get(n, 0)})
    for t in sorted(used_n):
        if t not in seen:
            out.append({"name": t, "n": used_n[t]})
    return out


@router.get("/api/cost-ledger/orgs")
def cost_ledger_orgs(request: Request):
    """成本台账可选主体（取自平台主体档案，不在本工具写死）。default＝存货体量最大的孝感星期九。"""
    if not _require_perm(request, "cost_ledger"):
        return JSONResponse({"ok": False, "msg": "无存货台账权限"}, status_code=403)
    return {"ok": True, "orgs": _cl_orgs(), "default": _CL_DEFAULT_ORG}


@router.get("/api/cost-ledger/state")
def cost_ledger_state(request: Request, year: int = 2026, period: int = 5, org: str = _CL_DEFAULT_ORG):
    """本期状态 + 结果（页面进入/切主体或期间直接读）。全员看到同一份；已封存则读快照。
    只读接口，不取金蝶、不写库——没接过数就返回 has_data=false，由页面引导去取数/上传。"""
    u = _require_perm(request, "cost_ledger")
    if not u:
        return JSONResponse({"ok": False, "msg": "无存货台账权限（请管理员授予「存货台账·查看与导出」）"}, status_code=403)
    if not _cl_org(org):
        return JSONResponse({"ok": False, "msg": f"主体档案里没有账簿代码 {org}"}, status_code=400)
    return _cl_payload(int(year), int(period), org, u)


@router.get("/api/cost-ledger/detail")
def cost_ledger_detail(request: Request, year: int = 0, period: int = 0, org: str = _CL_DEFAULT_ORG,
                       wh: str = "", cat: str = "", q: str = "", neg: int = 0,
                       limit: int = 100, offset: int = 0):
    """收发存明细查询（第④步）：按 仓库 / 存货类别 / 物料关键字 / 只看负结存 筛，分页给。

    为什么走接口而不是把 cross 一次推给前端：本期明细 2,770 行（大主体更多），
    全量塞进 /state 的 JSON 里，每次切期间/切主体都要拖一遍，页面反而更慢。

    合计（sum）按【筛出来的全部行】算，不受分页影响——不然翻页时合计会变，
    看的人会以为账不对。口径与仓库透视同源（都出自同一份 cross），
    故"某仓某类别"的合计与透视格子里的数必然分毫一致。

    筛选项全部可空：都不填＝本期全部明细。"""
    u = _require_perm(request, "cost_ledger")
    if not u:
        return JSONResponse({"ok": False, "msg": "无存货台账权限（请管理员授予「存货台账·查看与导出」）"}, status_code=403)
    y, p = int(year or 0), int(period or 0)
    if not (y and p):
        return JSONResponse({"ok": False, "msg": "参数错误：明细需指定 year/period"}, status_code=400)
    if not _cl_org(org):
        return JSONResponse({"ok": False, "msg": f"主体档案里没有账簿代码 {org}"}, status_code=400)
    res, cross, meta, snap = _cl_result(y, p, org)
    if res is None:
        return JSONResponse({"ok": False, "msg": "%s 还没有数据" % _cl_period_str(y, p, org)}, status_code=400)
    kw = (q or "").strip().lower()
    rows = cross
    if wh:
        rows = [r for r in rows if r.get("wh") == wh]
    if cat:
        rows = [r for r in rows if r.get("cat") == cat]
    if neg:
        rows = [r for r in rows if (r.get("eq") or 0) < 0]
    if kw:      # 关键字同时打【物料名称】与【物料编码】——记不住编码就打名字，反之亦然
        rows = [r for r in rows
                if kw in (r.get("name") or "").lower() or kw in (r.get("code") or "").lower()]
    rows = sorted(rows, key=lambda r: -abs(r.get("ea") or 0))   # 金额绝对值降序：负结存要看，别沉底
    tot = {"ea": round(sum(r.get("ea") or 0 for r in rows), 2),
           "eq": round(sum(r.get("eq") or 0 for r in rows), 4)}
    lim = max(1, min(int(limit or 100), 1000))
    off = max(0, int(offset or 0))
    page = rows[off:off + lim]
    # V2.138：补齐业务方底稿的 17 列（加 物料分组 + 四段单价 + 期初/收入/发出数量）
    # 列序按业务方定：编码/名称/规格/类别/分组/仓库/批号/单位 → 期初·收入·发出·结存（各 数量/单价/金额）
    keys = ("code", "name", "spec", "cat", "grp", "wh", "batch", "unit",
            "oq", "op", "oa", "iq", "ip", "ia", "dq", "dp", "da", "eq", "ep", "ea")
    # 筛选下拉的候选值给前端（取自本期真实有数的仓库/类别，不是全档案——选了没数的没意义）
    return {"ok": True, "wh": wh, "cat": cat, "q": q, "neg": bool(neg),
            "total": len(rows), "sum": tot, "offset": off, "limit": lim,
            "whs": sorted({r["wh"] for r in cross if r.get("wh")}),
            "cats": sorted({r["cat"] for r in cross if r.get("cat")}),
            "rows": [{k: r.get(k) for k in keys} for r in page]}


@router.post("/api/cost-ledger/close")
def cost_ledger_close(body: dict, request: Request):
    """封存本期：三道勾稽全过 → 结果拍照落库 → 本期只读（不再取金蝶、不能重新上传）。
    未过勾稽时，主管理员/核算子管理员可填理由强制封存——真实月结总有说不清的差异。"""
    u = _require_perm(request, "cost_ledger_close")
    if not u:
        return JSONResponse({"ok": False, "msg": "无「存货台账·封存本期」权限，请联系管理员"}, status_code=403)
    y, p = int(body.get("year") or 0), int(body.get("period") or 0)
    org = str(body.get("org") or _CL_DEFAULT_ORG)
    if not (y and p):
        return {"ok": False, "msg": "参数错误：缺 year/period"}
    if not _cl_org(org):
        return {"ok": False, "msg": f"主体档案里没有账簿代码 {org}"}
    if _cl_closed(y, p, org):
        return {"ok": False, "msg": "%s 已封存，无需重复操作" % _cl_period_str(y, p, org)}
    res, cross, meta, _ = _cl_result(y, p, org)
    if res is None:
        return {"ok": False, "msg": "%s 还没有数据，先取数或上传后再封存" % _cl_period_str(y, p, org)}
    # 硬校验（V2.124 业务方定：卡封存 + 卡导出）：本期有数据的仓库必须都配了仓库类型。
    # 不给 force 绕过——缺对照会让这些仓的钱落进「（属性缺失）」，封存＝把这个状态定死，
    # 而补一条对照只要几秒钟，没有"就是配不上"的正当场景（不同于勾稽差异那种真说不清的）。
    miss = _cl_missing_wh(res)
    if miss:
        return {"ok": False, "missing_wh": miss,
                "msg": "%s 有 %d 个仓库没配仓库类型，不能封存：%s。请到「存货台账 › 基础资料」补齐后再封。"
                       % (_cl_period_str(y, p, org), len(miss), "、".join(miss[:8]) + ("…" if len(miss) > 8 else ""))}
    note = str(body.get("note", "") or "").strip()
    if not res["credible"]:
        bad = [k for k, v in res["ties"].items() if isinstance(v, dict) and v.get("pass") is False]
        if not body.get("force"):
            return {"ok": False, "msg": "三道勾稽未全过（%s），本期不可信。确要封存请填理由强制封存。"
                                        % ("、".join(bad) or "存在未过项"), "credible": False}
        if not db.can_admin_accounts(u):
            return JSONResponse({"ok": False, "msg": "勾稽未全过，只有主管理员/核算工作台子管理员能强制封存"},
                                status_code=403)
        if len(note) < 5:
            return {"ok": False, "msg": "强制封存必须填理由（至少5个字），供领导与审计核查"}
        note = "【强制封存】未过勾稽：%s。理由：%s" % ("、".join(bad) or "?", note)
    # 先存快照、再落封存标记：中途挂了本期仍是进行中，不会出现"已封存却读到空快照"
    n = db.save_snapshot(_cl_source(org), y, p, _CL_SNAP, {"res": res, "cross": cross, "meta": meta})
    info = db.close_period(_cl_source(org), y, p, u["name"], note, meta.get("updated_at", ""))
    db.audit(u["name"], "成本台账·封存", _cl_period_str(y, p, org), note or "三道勾稽全过")
    return {"ok": True, "msg": "%s 已封存，本期转为只读。" % _cl_period_str(y, p, org),
            "状态": info, "快照字节": n}


@router.post("/api/cost-ledger/reopen")
def cost_ledger_reopen(body: dict, request: Request):
    """解封：高危操作，限主管理员/核算工作台子管理员，必须填理由，全程留痕。
    快照保留不删——解封重封会覆盖，但解封前那一版留着可查。"""
    u = _current_user(request)
    if not u or not (db.is_super(u) or db.user_can(u, "manage_accounting")):
        return JSONResponse({"ok": False, "msg": "解封是高危操作，只有主管理员或核算工作台子管理员可以执行"},
                            status_code=403)
    y, p = int(body.get("year") or 0), int(body.get("period") or 0)
    org = str(body.get("org") or _CL_DEFAULT_ORG)
    if not (y and p):
        return {"ok": False, "msg": "参数错误：缺 year/period"}
    if not _cl_closed(y, p, org):
        return {"ok": False, "msg": "%s 是进行中，无需解封" % _cl_period_str(y, p, org)}
    reason = str(body.get("reason", "") or "").strip()
    if len(reason) < 5:
        return {"ok": False, "msg": "解封必须填写理由（至少5个字），供领导与审计核查"}
    info = db.reopen_period(_cl_source(org), y, p, u["name"], reason)
    db.audit(u["name"], "成本台账·解封", _cl_period_str(y, p, org), reason)
    return {"ok": True, "msg": "%s 已解封，恢复可编辑。改完记得重新封存。" % _cl_period_str(y, p, org), "状态": info}


@router.get("/api/cost-ledger/closed-periods")
def cost_ledger_closed_periods(request: Request, org: str = _CL_DEFAULT_ORG):
    """该主体的已封存期间清单。各主体一条独立封存线，互不影响。"""
    if not _require_perm(request, "cost_ledger"):
        return JSONResponse({"ok": False, "msg": "无存货台账权限"}, status_code=403)
    return {"ok": True, "org": org, "历史": db.list_closed_periods(_cl_source(org))}


@router.get("/api/cost-ledger/warehouse-types")
def cost_ledger_wh_types(request: Request):
    """「仓库类型」维护页取数：仓库类型主档 + 仓库清单（金蝶仓库档案(107) ⋃ 现有对照 ⋃ 本期出现过的仓库）。
    仓库行带：是否已配类型、是否本期有数据、金蝶档案的启用/禁用状态。不改任何数据。

    V2.142 由「仅登录」收紧为 cost_ledger_wh——三级页面的可见性按业务方新模型由动作权限管，
    菜单藏了、接口也得拦（直接打 URL 不能绕过）。只有 CostLedgerWh.jsx 调本接口，
    月结核对页不碰它，收紧不误伤查看者。"""
    u = _require_perm(request, "cost_ledger_wh")
    if not u:
        return JSONResponse({"ok": False, "msg": "无「存货台账·维护基础资料」权限（本页由该动作权限管控）"}, status_code=403)
    cfg = _cl_config()
    attr = dict(cfg.get("warehouse_attr") or {})
    # V2.124：本页不再有「本期有数据」概念——本页没有期间选择器，"本期"到底指哪一期说不清
    # （曾按"最近有数据的一期"算，实际跑到了 6 期，而月结核对页停在 5 期，两个页面对不上）。
    # 哪一期缺哪些仓库的类型，由【月结核对】页的硬校验明确拦住封存与导出，本页不重复。
    # 顺带：省掉每次进页面都要 build 一整期结果的开销。
    ks, msg = {}, ""
    try:
        # V2.126 多主体：仓库类型对照【全局共用】（业务方定），故这里取【所有主体的仓库并集】。
        #   依据：实测 101 的数据里出现 `孝感茶饮原料仓`，而它在 107 的对照里已配「工厂仓」并直接生效——
        #   说明"仓库名→类型"是【物理仓库的属性】，不是主体的属性（同一个仓谁用都是那个类型）。
        # 按名归并（收发存报表只给仓库【名称】，没有编码可用）：
        #   107 实测 143 条档案 / 140 个仓库名——3 个仓库重名双编码（旧的禁用、新的启用，如
        #   孝感半成品仓 XQJ04(启用)+XQJ17(禁用)）；跨主体也会同名（105 与 107 都有孝感成品仓）。
        #   **重名时以启用的那条为准**，否则本期在用的仓会被误标成"已禁用"。
        for o in _cl_orgs():
            for w in kc.fetch_warehouses(o["code"]):
                old = ks.get(w["name"])
                if old is None or (old.get("forbid") == "B" and w.get("forbid") != "B"):
                    ks[w["name"]] = w
    except Exception as e:
        msg = f"金蝶仓库档案取数失败，仅列出已有对照与已上档的仓库：{e}"
    _st = db.get_setting(_CL_WH_KEY) or {}
    notes = _st.get("notes") or {}
    seen = _st.get("seen") or {}          # {仓库: 首次出现期间}，取数时自动上档（V2.124）
    names = set(ks) | set(attr) | set(seen)
    names.discard("（无仓库）")
    rows = []
    for n in sorted(names):
        w = ks.get(n) or {}
        rows.append({"wh": n, "code": w.get("code", ""), "type": attr.get(n, ""),
                     "note": notes.get(n, ""), "forbid": w.get("forbid", "") == "B",
                     "in_kingdee": n in ks,
                     "since": seen.get(n, ""),                    # 首次在真实数据里出现的期间
                     "is_new": bool(seen.get(n)) and not attr.get(n)})   # 上过档但还没配类型＝新仓库待配
    meta = db.get_setting(_CL_WH_KEY) or {}
    return {"ok": True, "rows": rows, "types": _cl_wh_types(attr),
            "can_edit": bool(db.user_can(u, "cost_ledger_wh")),
            "from_db": bool(meta.get("map") is not None),
            "updated_by": meta.get("by", ""), "updated_at": meta.get("at", ""),
            "msg": msg}


@router.post("/api/cost-ledger/warehouse-types")
async def cost_ledger_wh_types_save(request: Request):
    """保存仓库类型主档 + 仓库对照 + 仓库备注（全量覆盖存 app_settings）。
    · map：{仓库: 类型}，类型留空＝不配（不硬归，仓库透视里落「（属性缺失）」）。
    · types：[类型名] —— 纯分类，无日期无状态（V2.121 定稿）。
    · notes：{仓库: 备注}，成本会计自由填写；空串不落库。"""
    u = _require_perm(request, "cost_ledger_wh")
    if not u:
        return JSONResponse({"ok": False, "msg": "无权限（请管理员授予「存货台账·维护基础资料」）"}, status_code=403)
    body = await request.json()
    m = body.get("map")
    if not isinstance(m, dict):
        return JSONResponse({"ok": False, "msg": "参数错误：map 应为 {仓库: 类型}"}, status_code=400)
    clean = {str(k).strip(): str(v).strip() for k, v in m.items() if str(k).strip() and str(v).strip()}
    notes = {str(k).strip(): str(v).strip()[:200] for k, v in (body.get("notes") or {}).items()
             if str(k).strip() and str(v).strip()}

    types, seen = [], set()
    for t in (body.get("types") or []):
        n = str((t.get("name") if isinstance(t, dict) else t) or "").strip()
        if n and n not in seen:
            seen.add(n); types.append(n)
    for t in sorted({v for v in clean.values() if v}):     # 在用类型不能凭空消失
        if t not in seen:
            seen.add(t); types.append(t)

    # seen（新仓库自动上档记录）由取数时写入，本页不提交也不能丢——原样带过去
    seen = (db.get_setting(_CL_WH_KEY) or {}).get("seen") or {}
    db.set_setting(_CL_WH_KEY, {"map": clean, "types": types, "notes": notes, "seen": seen,
                                "by": u.get("name") or u.get("username", ""),
                                "at": _now()}, operator=u.get("username", ""))
    return {"ok": True, "n": len(clean), "n_types": len(types), "n_notes": len(notes)}


def _dash_from_store(y, p, org):
    """某期【落库/封存】的口径：从当期输入里现算，不碰金蝶（V2.281）。
    返回 (subjects, flow, src) 或 None——None＝该期没数据，调用方回退实时取。"""
    inp = _cl_input(y, p, org)
    if not inp:
        return None
    d = inp["payload"]
    cross = d.get("cross") or []
    if not cross:
        return None
    gl = d.get("gl") or {}
    flow = {"ia": 0.0, "da": 0.0, "ea": 0.0, "cats": {}}
    for r in cross:
        flow["ia"] += r.get("ia") or 0.0
        flow["da"] += r.get("da") or 0.0
        flow["ea"] += r.get("ea") or 0.0
        c = flow["cats"].setdefault(r.get("cat") or "（未分类）", {"ea": 0.0, "ia": 0.0, "da": 0.0})
        c["ea"] += r.get("ea") or 0.0
        c["ia"] += r.get("ia") or 0.0
        c["da"] += r.get("da") or 0.0
    return ({k: round(v, 2) for k, v in gl.items()},
            flow, "封存快照" if _cl_closed(y, p, org) else "本期落库")


@router.get("/api/cost-ledger/dashboard")
def cost_ledger_dashboard(request: Request, year: int = 0, period: int = 0,
                          org: str = _CL_DEFAULT_ORG, months: int = 5,
                          basis: str = "current"):
    """存货看板（V2.254）：本年 1..period 各期的存货类科目余额 + 收发流量 + 周转天数。

    **两个口径同页出现，必须讲清楚，否则同一个月两个数没人对得上**：
      · 结存/构成 ＝【科目余额】口径，含在途物资/委托加工物资/材料采购（返回 `subject_total`）；
      · 收发流量/周转 ＝【收发存表】口径，不含上面那几个（返回 `flow_end`）。
      P5 实测 10,994,036.02 vs 10,958,051.26，差额就是在途与委托加工。页面上并排标注，不做加减。

    取数：科目余额**逐期取**（金蝶按期返，没有一次多期的入口）；收发存**一次取多期**
    （FIsDisplayPeriod，V2.254 实证）。故 5 期＝5 次余额 + 1 次收发存，不是 10 次。"""
    if not _require_perm(request, "cost_ledger"):
        return JSONResponse({"ok": False, "msg": "无权限"}, status_code=403)
    o = _cl_org(org)
    if not o:
        return JSONResponse({"ok": False, "msg": f"主体档案里没有账簿代码 {org}"}, status_code=400)
    y = int(year or 0)
    p = int(period or 0)
    if not (y and 1 <= p <= 12):
        return JSONResponse({"ok": False, "msg": "参数错误：year/period"}, status_code=400)
    n = max(1, min(int(months or 5), 12))
    p0 = max(1, p - n + 1)

    try:
        cfg = _cl_config()
    except ClgConfigMissing:
        return JSONResponse({"ok": False, "msg": "还没有类别↔科目对照，请先到「基础资料」建立"}, status_code=400)
    subj_order = list((cfg.get("category_to_subject") or {}).keys()) + list(cfg.get("extra_subjects") or [])

    basis = "closed" if str(basis) == "closed" else "current"
    series, gl_msg = [], ""
    stored = {}
    if basis == "closed":
        for i in range(p0, p + 1):
            got = _dash_from_store(y, i, org)
            if got:
                stored[i] = got

    # V2.298 提速：【已封存】期间的科目余额直接读落库值，不再回金蝶。
    # 业务方：「这个要现取的吗」——原来默认口径（按当前分类重算）下 7 个月＝7 次科目余额取数，
    # 进一次页面等十几秒。而金蝶科目余额没有"一次返多期"的入口，只能逐期打。
    #
    # ⚠**只对科目余额这么做，收发存那半边不动**——两者对"重分类"的敏感度完全不同：
    #   · 科目余额是**总账**的数，与存货类别无关：凭证记的是当时的科目，改物料档案不会追溯重记
    #     （同 D20/类别漂移那条）。所以封存期间的余额今天再取一遍，除非有人补记凭证，否则一模一样。
    #     而封存期间本就禁止取数/上传，补记也进不来 → **读快照与实时取等价，纯赚**。
    #   · 收发流量/类别构成**按存货类别归集**，金蝶报表按【当前】档案归类。
    #     "按当前分类重算"这个口径要的就是这个效果，读落库值会混进各月的旧口径——**不能省**。
    #     好在它一趟能取多期（FIsDisplayPeriod），本来就只有 1 次调用，不是瓶颈。
    sealed_gl = {}
    if basis == "current":
        for i in range(p0, p + 1):
            if not _cl_closed(y, i, org):
                continue
            inp = _cl_input(y, i, org)
            g = (inp or {}).get("payload", {}).get("gl") if inp else None
            if g:
                sealed_gl[i] = g
    try:
        for i in range(p0, p + 1):
            if i in stored:
                subj, _, src = stored[i]
                series.append({"period": i, "label": f"{i}月", "subjects": subj,
                               "subject_total": round(sum(subj.values()), 2), "basis": src})
                continue
            if i in sealed_gl:
                g = sealed_gl[i]
                series.append({"period": i, "label": f"{i}月",
                               "subjects": {k: round(v, 2) for k, v in g.items()},
                               "subject_total": round(sum(g.values()), 2), "basis": "封存快照"})
                continue
            gl = _inv_gl_balance(y, i, org)
            series.append({"period": i, "label": f"{i}月", "subjects": {k: round(v, 2) for k, v in gl.items()},
                           "subject_total": round(sum(gl.values()), 2),
                           "basis": "实时取数" if basis == "closed" else "当前分类"})
    except ClgBookMismatch as e:
        gl_msg = str(e)
    except Exception as e:
        gl_msg = f"取科目余额失败：{e}"

    flow, flow_msg = [], ""
    need = [i for i in range(p0, p + 1) if i not in stored]
    try:
        if need:
            flow = kc.fetch_inventory_period_totals(y, min(need), max(need), org)
    except Exception as e:
        flow_msg = f"取收发存失败：{e}"
    # 收发存的期间标签形如 "2026.05" → 对齐到期号，页面按期号 join
    by_p = {}
    for f in flow:
        try:
            by_p[int(str(f["period"]).split(".")[-1])] = f
        except (ValueError, IndexError):
            continue
    # 「库存商品」这一科目下有哪几个存货类别，取自基础资料的对照表（不写死"产成品+自制半成品+委外半成品"）
    goods_cats = set((cfg.get("category_to_subject") or {}).get("库存商品") or [])
    for s in series:
        if s["period"] in stored:
            _, sf, _ = stored[s["period"]]
            f = {"ia": round(sf["ia"], 2), "da": round(sf["da"], 2), "ea": round(sf["ea"], 2)}
            cats = sf["cats"]
        else:
            f = by_p.get(s["period"]) or {}
            cats = f.get("cats") or {}
        s["in_amt"] = f.get("ia", 0.0)
        s["out_amt"] = f.get("da", 0.0)
        s["flow_end"] = f.get("ea", 0.0)
        s["cats"] = {c: d.get("ea", 0.0) for c, d in cats.items()}
        s["goods_end"] = round(sum(d.get("ea", 0.0) for c, d in cats.items() if c in goods_cats), 2)
        s["goods_out"] = round(sum(d.get("da", 0.0) for c, d in cats.items() if c in goods_cats), 2)
        s["goods_in"] = round(sum(d.get("ia", 0.0) for c, d in cats.items() if c in goods_cats), 2)

    # 周转天数＝(期初+期末)/2 ÷ 本期发出 × 30。两条口径各算各的：
    #   turn_days      ＝全部存货；goods_turn_days ＝仅库存商品（分母是本类的发出额，不是全部发出）。
    # 发出为 0 不算——除零算出的"无穷天"摆在看板上比空着更误导（新主体、停产月都会出现）。
    # 首期没有上期期末可用，退而用本期期末当期初（会略微低估，页面上标注）。
    prev_map = {s["period"]: s for s in series}
    for s in series:
        pv = prev_map.get(s["period"] - 1)
        for beg_k, end_k, out_k, dst in (("flow_end", "flow_end", "out_amt", "turn_days"),
                                         ("goods_end", "goods_end", "goods_out", "goods_turn_days")):
            beg = pv[beg_k] if pv else s[end_k]
            s[dst] = round((beg + s[end_k]) / 2 / s[out_k] * 30, 1) if s.get(out_k) else None
        s["turn_first"] = pv is None

    cur = series[-1] if series else {}
    prev = series[-2] if len(series) > 1 else None
    vals = [s["subject_total"] for s in series if s.get("subject_total")]
    return {"ok": True, "basis": basis, "org": org, "org_name": o.get("name", org), "year": y, "period": p,
            "from_period": p0, "series": series, "subject_order": subj_order,
            "cur_total": cur.get("subject_total", 0.0),
            "mom": (round((cur["subject_total"] - prev["subject_total"]) / prev["subject_total"] * 100, 1)
                    if prev and prev.get("subject_total") else None),
            "avg_total": round(sum(vals) / len(vals), 2) if vals else 0.0,
            "turn_days": cur.get("turn_days"), "goods_turn_days": cur.get("goods_turn_days"),
            # V2.292：库存账龄/呆滞**整条砍掉**（业务方定：不做，也不走手工上传）。
            # 原先这里回一个 available:False 的占位、看板上挂一张"待账龄表"的黄卡——
            # 需求不做了，占位就该一起走：留着等于在页面上长期挂一个永远不会兑现的承诺。
            "gl_msg": gl_msg, "flow_msg": flow_msg}


@router.get("/api/cost-ledger/cat-subjects")
def cost_ledger_cat_subjects(request: Request):
    """类别↔科目对照（V2.254）。返回 {subjects:[{subject,cats:[]}], extra:[科目], src, seed_only}。
    `src`＝当前生效来源：'db'（页面存过）/'json'（还在吃种子）——页面要显式告诉人现在在用哪个，
    否则会重演"两个地方都能改、不知道谁生效"。"""
    if not _require_perm(request, "cost_ledger_wh"):
        return JSONResponse({"ok": False, "msg": "无权限（请管理员授予「存货台账·维护基础资料」）"}, status_code=403)
    ov = db.get_setting(_CL_CAT_KEY)
    from_db = bool(isinstance(ov, dict) and ov.get("category_to_subject"))
    try:
        cfg = _cl_config()
    except ClgConfigMissing:
        return {"ok": True, "subjects": [], "extra": [], "src": "none",
                "msg": "配置文件不在、数据库里也没存过——请在本页建立类别↔科目对照后保存。"}
    c2s = cfg.get("category_to_subject") or {}
    return {"ok": True, "src": "db" if from_db else "json",
            "subjects": [{"subject": k, "cats": list(v)} for k, v in c2s.items()],
            "extra": list(cfg.get("extra_subjects") or []),
            "by": (ov or {}).get("by", ""), "at": (ov or {}).get("at", "")}


@router.post("/api/cost-ledger/cat-subjects")
async def cost_ledger_cat_subjects_save(request: Request):
    """保存类别↔科目对照（全量覆盖存 app_settings）。存过之后 json 种子即失效。

    校验只做两条硬的：①至少一条科目且每条至少一个类别；②同一类别不能挂到两个科目上
    ——那会让同一笔存货被算进两个总账科目，账实勾稽必然对不上，且错得很隐蔽。
    **不校验"类别必须在金蝶存在"**：金蝶随时可能新增存货类别，卡在这里会让人改不了；
    真出现没对照上的类别，异常稽核里的「对照缺失」会点出来（那才是发现它的地方）。"""
    u = _require_perm(request, "cost_ledger_wh")
    if not u:
        return JSONResponse({"ok": False, "msg": "无权限（请管理员授予「存货台账·维护基础资料」）"}, status_code=403)
    body = await request.json()
    rows = body.get("subjects")
    if not isinstance(rows, list):
        return JSONResponse({"ok": False, "msg": "参数错误：subjects 应为 [{subject,cats}]"}, status_code=400)
    c2s, owner = {}, {}
    for r in rows:
        sub = str((r or {}).get("subject") or "").strip()
        if not sub:
            continue
        cats, seen = [], set()
        for c in ((r or {}).get("cats") or []):
            n = str(c or "").strip()
            if not n or n in seen:
                continue
            if n in owner and owner[n] != sub:
                return JSONResponse({"ok": False, "msg": "「%s」同时挂在「%s」和「%s」两个科目下。"
                                     "一个存货类别只能归一个科目，否则这部分存货会被算两遍、账实勾稽必然不平。"
                                     % (n, owner[n], sub)}, status_code=400)
            owner[n] = sub
            seen.add(n); cats.append(n)
        if cats:
            c2s[sub] = cats
    if not c2s:
        return JSONResponse({"ok": False, "msg": "至少要有一个科目、且该科目下至少一个存货类别。"
                             "没有对照就没法把存货归到总账科目，账实勾稽做不了。"}, status_code=400)
    extra, seen = [], set()
    for e in (body.get("extra") or []):
        n = str(e or "").strip()
        if n and n not in seen:
            seen.add(n); extra.append(n)
    db.set_setting(_CL_CAT_KEY, {"category_to_subject": c2s, "extra_subjects": extra,
                                 "by": u.get("name") or u.get("username", ""), "at": _now()},
                   operator=u.get("username", ""))
    db.audit(u.get("name") or u.get("username", ""), "存货台账·保存类别↔科目对照", "",
             "科目 %d 个 / 类别 %d 个 / 单列科目 %d 个" % (len(c2s), len(owner), len(extra)))
    return {"ok": True, "n_subjects": len(c2s), "n_cats": len(owner), "n_extra": len(extra)}


def _merge_bill_only_rows(flow_rows, bill_lines):
    """把**只在出库单上、流水表没吐**的分录行补进物料级明细（V2.314）。

    业务方底稿是照【出库单】做的，工具的明细页是拿单号回查【收发存流水表】——
    两者行数会差：🧪 QTCK011302 单据 123 行 / 流水 122 行，差的是
    `T00000145 口袋蛋白脆-出口版` 10.8 千克 / **金额 0.00**（整张单唯一零金额行），
    底稿 124 行 vs 工具 123 行就是它。业务方要求补齐。

    ⚠**只补、不替**：流水表返回过的 (单号,物料,仓库) 一律以流水为准——
    流水是存货账的出口，单据金额可能尚未结转成本，拿单据金额覆盖会改动已勾稽的数。
    补进来的行标 `src="bill"`，导出页在「事务类型」列写明「（出库单·流水无此行）」，
    金额一律照单据的原值（实测就是 0），**不参与任何合计的重算**——合计仍是 SUM 明细列。
    """
    if not bill_lines:
        return flow_rows
    seen = {(r.get("billno"), str(r.get("code")), r.get("wh")) for r in flow_rows}
    # 同单据的元数据（事务类型/日期/凭证字号）照该单已有的流水行填，填不出就留空、不猜
    meta = {}
    for r in flow_rows:
        meta.setdefault(r.get("billno"), r)
    out = list(flow_rows)
    for bn, lines in bill_lines.items():
        for x in lines:
            if (bn, str(x.get("code")), x.get("wh")) in seen:
                continue
            m = meta.get(bn) or {}
            out.append({"billno": bn, "btype": "（出库单·流水无此行）", "date": m.get("date") or "",
                        "code": x.get("code"), "name": x.get("name"), "spec": x.get("spec"),
                        "cat": "", "grp": "", "unit": x.get("unit") or "",
                        "wh": x.get("wh"), "batch": x.get("batch"),
                        "qty": x.get("qty"), "price": 0.0, "amount": x.get("amount") or 0.0,
                        "voucher": m.get("voucher") or "", "src": "bill"})
    # 类别/分组：这些行零金额零结存，**跨维度表里没有它们**，只能查物料档案。
    # 不补的话「类别」会按兜底规则落成"原辅料货损"——🧪 T00000145 实为产成品/植物肉，
    # 底稿写的是「产品货损」，归错档会串到分类小计里，比缺一行更糟。
    _new = [r for r in out if r.get("src") == "bill"]
    if _new:
        try:
            attrs = kc.fetch_material_attrs([r["code"] for r in _new])
            for r in _new:
                a = attrs.get(str(r["code"])) or {}
                r["cat"], r["grp"] = a.get("cat", ""), a.get("grp", "")
        except Exception:
            pass          # 查不到就留空，不猜；宁可类别空着也不填个错的
    return out


@router.get("/api/cost-ledger/export")
def cost_ledger_export(request: Request, year: int = 0, period: int = 0, org: str = _CL_DEFAULT_ORG):
    """导出《成本台账》xlsx —— 按【显式主体+期间】取数。
    V2.122 前从进程内全局单槽取，多人用时会导出别人刚跑的那一期（屏幕与导出对不上）；
    现按 org/year/period 从本期输入（或封存快照）算，导出的必然是页面上那一主体那一期。"""
    if not _require_perm(request, "cost_ledger"):
        return JSONResponse({"ok": False, "msg": "无存货台账权限（请管理员授予「存货台账·查看与导出」）"}, status_code=403)
    y, p = int(year or 0), int(period or 0)
    if not (y and p):
        return JSONResponse({"ok": False, "msg": "参数错误：导出需指定 year/period"}, status_code=400)
    o = _cl_org(org)
    if not o:
        return JSONResponse({"ok": False, "msg": f"主体档案里没有账簿代码 {org}"}, status_code=400)
    res, cross, meta, snap = _cl_result(y, p, org)
    if res is None:
        return JSONResponse({"ok": False, "msg": "%s 还没有数据，请先取数或上传后再导出" % _cl_period_str(y, p, org)},
                            status_code=400)
    # 硬校验（V2.124）：缺仓库类型不给导出——台账一旦发出去就在外面流传，
    # 里面的仓库类型小计却是残的。已封存的读快照，封存时已过校验，不再拦。
    miss = [] if snap else _cl_missing_wh(res)
    if miss:
        return JSONResponse({"ok": False, "missing_wh": miss,
                             "msg": "%s 有 %d 个仓库没配仓库类型，不能导出：%s。请到「存货台账 › 基础资料」补齐。"
                                    % (_cl_period_str(y, p, org), len(miss),
                                       "、".join(miss[:8]) + ("…" if len(miss) > 8 else ""))},
                            status_code=400)
    # 导出表头署名：全称优先，空则退简称。V2.253 起全称不再参与勾稽、常态就是空的，
    # 不给它兜底会掉进下面那个"写死孝感"的兜底里——101 的报表署名成孝感，比没名字糟得多。
    meta = dict(meta or {}); meta["org_full"] = o["full"] or o["name"]
    data = _build_cost_ledger_xlsx(res, cross, meta)
    fname = f"成本台账_{o['name']}_{y}年第{p}期.xlsx"      # 带主体：三个主体导出来不能同名
    disp = "attachment; filename=cost_ledger.xlsx; filename*=UTF-8''" + urllib.parse.quote(fname)
    return Response(content=data,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": disp})


def _build_cost_ledger_xlsx(res, cross, meta):
    """从分析结果构建《成本台账》xlsx：核对结论/账实勾稽/收发存汇总/异常清单/损益归集/收发存明细。"""
    from openpyxl import Workbook
    from copy import copy
    from openpyxl.formula.translate import Translator
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    FN = "微软雅黑"
    def F(sz=10, b=False, c="181B21"): return Font(name=FN, size=sz, bold=b, color=c)
    HEAD = PatternFill("solid", fgColor="4B53C4"); TOT = PatternFill("solid", fgColor="F0F1F5")
    GRN = PatternFill("solid", fgColor="E8F4EE"); RED = PatternFill("solid", fgColor="FBECEA")
    thin = Side(style="thin", color="DCDFE5"); BD = Border(left=thin, right=thin, top=thin, bottom=thin)
    M = '#,##0.00;[Red](#,##0.00)'
    # V2.294：**数据格一律不自动换行**（业务方：「这些不要转行，好难看」）。
    # 原先 CEN 带 wrap_text=True 且数据格也在用——「成本调整提示」在 11 字宽的列里折成两行、
    # 整行跟着变高，一页异常清单看下来高高低低。表头留换行（列窄、标题短，折一行反而齐整），
    # 数据格改用 CENN；文字说明块也去掉换行，让它平铺过去（右边本来就是空列）。
    CEN = Alignment("center", "center", wrap_text=True)      # 仅表头
    CENN = Alignment("center", "center")                      # 数据格
    RIG = Alignment("right", "center")
    wb = Workbook()
    y, p = meta["year"], meta["period"]
    # 原始底表提前取出：「损益归集」要写一行引用「原始·货损与处置明细」的勾稽公式，
    # 而那张页在后面才建——公式可以先写（Excel 打开时才解析），但**引哪一列现在就得知道**。
    raw = (res or {}).get("_raw") or {}
    _PDS = "'原始·货损与处置明细'"
    _pd0 = raw.get("pnl_detail") or {}
    _pd_has = bool(_pd0.get("loss") or _pd0.get("disposal") or _pd0.get("other"))
    # 明细页两种版式：物料级（16 列，金额在 P）／凭证级（7 列，金额在 F）。列位按实际版式取，不写死。
    _PD_AMT = "J" if _pd0.get("rows") else "F"
    # 「归属」列：物料级版式里它排在备注之后（M 列），凭证级兜底版式仍在 A 列
    _PD_TAG = "M" if _pd0.get("rows") else "A"

    def hdr(ws, r, labels):
        for i, t in enumerate(labels):
            c = ws.cell(row=r, column=i + 1, value=t); c.font = F(9.5, True, "FFFFFF"); c.fill = HEAD; c.alignment = CEN; c.border = BD

    # ── 原始底表的【金蝶元数据头】（V2.289）────────────────────────────
    # 业务方：「其他几个原始表也没有元数据噢」。金蝶自己导出的每张报表，表头之上都有两行
    # 交代**这份数是谁、哪个期间、什么币别**——手工上传的那张原样带着，接口取回的这几张却没有，
    # 同是"原始底表"两套样子。这里按金蝶原件的**逐字措辞**补齐（各报表措辞并不统一，照抄不统一）：
    #   跨维度   会计期间:2026年第3期-2026年第3期 ／ 本位币:人民币
    #   按日期   日期:【2026-03-01】-【2026-03-31】 ／ 本位币:人民币
    #   成本计算单 会计期间:2026年3期-2026年3期 ／ 币别:人民币     ← "第"字、"币别/本位币"都不一样
    # 第三行是**工具自己加的取数留痕**（金蝶原件没有）：写明来源报表、取数时间、取数人——
    # 明确标出"这行是工具写的"，免得日后被当成金蝶原文。
    _mfull = (meta or {}).get("org_full") or ""
    _mat = (meta or {}).get("updated_at") or ""
    _mby = (meta or {}).get("updated_by") or ""

    def _trace(src):
        return "〔工具留痕〕来源：%s；取数：%s%s" % (src, _mat or "—",
                                             ("　取数人：" + _mby) if _mby else "")

    def kmeta(ws, line2, src):
        """写金蝶式元数据头 → 返回表头应在的行号。
        V2.290：三行**一律黑字**（业务方定）。原先元数据用灰、留痕用更浅的灰，
        本意是"这不是数据"，实际效果是**元数据看着像被划掉的废话**——
        而它恰恰是判断"这份数是谁、哪个期间"的依据，该看清楚。"""
        r1 = ["核算体系:财务会计核算体系",
              ("核算组织:" + _mfull) if _mfull else None,
              "会计政策:中国准则会计政策"]
        for i, v in enumerate(x for x in r1 if x):
            ws.cell(row=1, column=1 + i, value=v).font = F(9)
        for i, v in enumerate(line2):
            ws.cell(row=2, column=1 + i, value=v).font = F(9)
        ws.cell(row=3, column=1, value=_trace(src)).font = F(9)
        return 4

    import calendar as _cal
    _last = _cal.monthrange(int(y), int(p))[1]
    _CUR = "人民币"

    ws = wb.active; ws.title = "核对结论"; ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 20
    for col in 'BCDE': ws.column_dimensions[col].width = 20
    ws['A1'] = "成本台账 · 存货月结核对"; ws['A1'].font = F(16, True, "4B53C4")
    # V2.126：主体全称由调用方经 meta["org_full"] 传入，不再写死孝感（多主体）
    # V2.253：兜底不能再写死孝感——多主体下把别家的报表署成孝感是错，宁可不署名。
    _full = (meta or {}).get("org_full") or ""
    ws['A2'] = " · ".join(x for x in (_full, f"{y} 年第 {p} 期", "金蝶只读 · 工具自动生成") if x)
    ws['A2'].font = F(10, False, "5A616D")
    r = 4
    ws.cell(row=r, column=1, value="三道勾稽 · 可信度报告").font = F(12, True); r += 1
    hdr(ws, r, ["勾稽项", "结果", "结论"]); r += 1
    ties = res["ties"]
    def tie_line(name, ok, detail):
        ws.cell(row=r, column=1, value=name).font = F(10, True)
        ws.cell(row=r, column=2, value=detail).font = F(9.5)
        cc = ws.cell(row=r, column=3, value=("✓ 通过" if ok else "✗ 未过")); cc.font = F(10, True, "1F7A55" if ok else "C0392B"); cc.fill = GRN if ok else RED; cc.alignment = CENN
        for c in range(1, 4): ws.cell(row=r, column=c).border = BD
    if "two_reports" in ties:
        tie_line("① 两表互勾", ties["two_reports"]["pass"], "跨维度↔按日期 四项金额合计"); r += 1
    tie_line("② 收发存自平", ties["self_balance"]["pass"], "期初＋收入−发出＝结存"); r += 1
    if "book_vs_actual" in ties:
        b = ties["book_vs_actual"]; tie_line("③ 账实勾稽", b["pass"], f"总账存货合计 {b.get('book_total',0):,.2f}"); r += 1
    r += 1
    cc = ws.cell(row=r, column=1, value=("结论：三道勾稽全部通过，本期台账可信。" if res["credible"] else "结论：存在未过勾稽，本期台账待复核。"))
    cc.font = F(11, True, "1F7A55" if res["credible"] else "A35A00")

    if "book_vs_actual" in ties:
        ws = wb.create_sheet("账实勾稽"); ws.sheet_view.showGridLines = False
        for i, w in enumerate([16, 18, 40, 18, 14, 10]): ws.column_dimensions[chr(65 + i)].width = w
        hdr(ws, 1, ["总账科目", "科目余额", "对应存货类别", "收发存结存", "差异", "状态"])
        rr = 2
        for subj, s in ties["book_vs_actual"]["subjects"].items():
            ws.cell(row=rr, column=1, value=subj).font = F(10, True)
            ws.cell(row=rr, column=2, value=s["book"]).number_format = M
            ws.cell(row=rr, column=3, value="＋".join(s["cats"])).font = F(9.5)
            ws.cell(row=rr, column=4, value=s["actual"]).number_format = M
            ws.cell(row=rr, column=5, value=s["diff"]).number_format = M
            st = ws.cell(row=rr, column=6, value="✓" if s["pass"] else "✗"); st.fill = GRN if s["pass"] else RED; st.alignment = CENN
            for c in range(1, 7): ws.cell(row=rr, column=c).border = BD
            rr += 1

    # 收发存汇总·按类别（V2.141 改为业务方指定列序：总账科目 + 类别 + 四段【数量·金额】+ 结存占比）
    # 「总账科目」＝存货类别的反查（config['category_to_subject'] 是 {科目:[类别…]}）；
    # 查不到的类别写「（未对照）」——不硬归，同 ST_NOMAP 口径；行按 科目 → 结存金额降序 排。
    PCT = '0.00%'
    QF = '#,##0.00'
    # ── 汇总块一律写成【引用原始底表的 SUMIFS 公式】（V2.291，业务方定）──────────
    # 业务方：「一般来说灰色几张表导出是怎么样就是怎么样了吧，你前面的汇总透视用 sumifs 才对啊」。
    # 对——灰色那几页是金蝶原样、不动的，那么汇总就该**从它们算出来**，而不是各算各的再摆一起：
    #   ①打开文件就能看见每个数是怎么来的（点一下就是完整口径），不用信"工具算过了"；
    #   ②任何人动了底表，汇总立刻跟着变——反而**更容易发现被改**，比存死一个数安全；
    #   ③底表本身就是留痕件，公式的分母不会跑掉。
    # ⚠**先逐格验、再写公式**：所有转公式的块都实测过与内核算出来的值一致（🧪 107/2026-3）：
    #   按类别 8 列 × 7 类别 ✓｜按仓库 8 列 × 42 仓 ✓｜车间×成本项目 19 格 + 委外 494,998.06 ✓｜费用项目 39 项 ✓
    #   验不过的**不转**（见下面「按事务类型」：它压根没有对应底表页）。
    XS = "'原始·收发存跨维度'"        # A编码 B名称 C类别 D分组 E规格 F状态 G批号 H仓库 …
    XCOL = {"oq": "L", "oa": "N", "iq": "O", "ia": "Q",     # …L期初数量 N期初金额 O收入数量 Q收入金额
            "dq": "R", "da": "T", "eq": "U", "ea": "W"}     #   R发出数量 T发出金额 U结存数量 W结存金额
    CS = "'原始·制造费用明细(接口)'"    # A车间 B产品编码 C产品名称 D工单 E单据类型 F层级 G成本项目 H子项费用项目 I金额 … L委外
    _has_x = bool(cross)
    _has_cc = bool(((res or {}).get("_raw") or {}).get("cc_rows"))

    def _q(s):
        """公式里的字符串常量：内部双引号翻倍。"""
        return '"%s"' % str(s).replace('"', '""')

    def xsumifs(key, crit_col, crit_ref):
        return "=SUMIFS(%s!%s:%s,%s!%s:%s,%s)" % (XS, XCOL[key], XCOL[key], XS, crit_col, crit_col, crit_ref)

    subj_of = clg.subject_of_category(_cl_config())
    # V2.309：按业务方的看法把【账实对比】并进本表——一张表就能看出"这几个类别加起来
    # 对不对得上总账那个一级科目"，不必在「汇总透视」与「勾稽与归集」两页之间来回跳。
    #   A–K 收发存透视（SUMIFS 活公式）｜ L 科目余额（总账取数）｜ M 差异（相减）
    # 科目余额与差异**按科目纵向合并单元格**：它是科目级的数，逐行重复会让人以为每个类别都有一个余额。
    ws = wb.create_sheet("收发存汇总·按类别"); ws.sheet_view.showGridLines = False
    for i, w in enumerate([13, 14, 13, 15, 13, 15, 13, 15, 13, 15, 11, 15, 13]):
        ws.column_dimensions[chr(65 + i)].width = w
    heads3 = ["总账科目", "存货类别", "期初数量", "期初金额", "收入数量", "收入金额",
              "发出数量", "发出金额", "结存数量", "结存金额", "结存占比",
              "科目余额（总账）", "差异"]
    hdr(ws, 1, heads3); ws.freeze_panes = "C2"; rr = 2
    _ba = (res.get("ties") or {}).get("book_vs_actual") or {}
    _subj = _ba.get("subjects") or {}
    tt = res["pivot_category"]["合计"]
    cats3 = [(c, a) for c, a in res["pivot_category"].items() if c != "合计"]
    cats3.sort(key=lambda kv: (subj_of.get(kv[0]) or "（未对照）", -abs(kv[1]["ea"])))
    NUM3 = [("oq", QF), ("oa", M), ("iq", QF), ("ia", M), ("dq", QF), ("da", M), ("eq", QF), ("ea", M)]
    first3 = rr
    _span = {}                       # 科目 → [起行, 止行]，供纵向合并
    for cat, a in cats3:
        _s = subj_of.get(cat) or "（未对照）"
        _span.setdefault(_s, [rr, rr])[1] = rr
        ws.cell(row=rr, column=1, value=_s).font = F(10)
        ws.cell(row=rr, column=2, value=cat).font = F(10)
        for j, (k, fmt) in enumerate(NUM3):
            c = ws.cell(row=rr, column=3 + j,
                        value=(xsumifs(k, "C", "B%d" % rr) if _has_x else a[k]))
            c.number_format = fmt
        sh = clg.share_of(a["ea"], tt["ea"])
        c = ws.cell(row=rr, column=11, value=sh if sh is not None else "—")
        if sh is not None: c.number_format = PCT
        c.alignment = CENN
        for c2 in range(1, 14): ws.cell(row=rr, column=c2).border = BD
        rr += 1
    # L/M 两列：每个科目只在首行写值，再纵向合并到该科目最后一行
    for _s, (r0, r1) in _span.items():
        info = _subj.get(_s) or {}
        b = ws.cell(row=r0, column=12, value=info.get("book"))
        b.number_format = M; b.alignment = CENN
        d = ws.cell(row=r0, column=13,
                    value=(round(info.get("diff"), 2) if info.get("diff") is not None else "—"))
        d.number_format = M; d.alignment = CENN
        if info.get("diff") is not None and abs(info["diff"]) > 0.005:
            d.font = F(10, True, "C0392B")          # 有差就红，别让它混在一片黑字里
        if r1 > r0:
            ws.merge_cells(start_row=r0, end_row=r1, start_column=12, end_column=12)
            ws.merge_cells(start_row=r0, end_row=r1, start_column=13, end_column=13)
    ws.cell(row=rr, column=1, value="合计").font = F(10, True)
    ws.cell(row=rr, column=2, value="%d 个类别" % len(cats3)).font = F(9.5)
    for j, (k, fmt) in enumerate(NUM3):
        col = chr(67 + j)
        c = ws.cell(row=rr, column=3 + j,
                    value=("=SUM(%s%d:%s%d)" % (col, first3, col, rr - 1) if _has_x else tt[k]))
        c.number_format = fmt; c.font = F(10, True)
    c = ws.cell(row=rr, column=11, value=1.0 if tt["ea"] else "—")
    if tt["ea"]: c.number_format = PCT
    c.font = F(10, True); c.alignment = CENN
    c = ws.cell(row=rr, column=12, value="=SUM(L%d:L%d)" % (first3, rr - 1))
    c.number_format = M; c.font = F(10, True); c.alignment = CENN
    c = ws.cell(row=rr, column=13, value="=L%d-J%d" % (rr, rr))
    c.number_format = M; c.font = F(10, True); c.alignment = CENN
    for c2 in range(1, 14):
        ws.cell(row=rr, column=c2).fill = TOT; ws.cell(row=rr, column=c2).border = BD
    rr += 1
    # 不走收发存表的单列科目（在途物资/委托加工物资/材料采购）——**放在合计之后**：
    # 它们只有总账侧、没有收发存侧，混进上面的合计会让"差异"凭空多出这几笔
    # （3 月实测：混进去差异会显示 40,347.09，而真实差异只有 ±1.92）。
    _extra = (_ba.get("extra") or {})
    if _extra:
        for _s, v in _extra.items():
            ws.cell(row=rr, column=1, value=_s).font = F(10)
            ws.cell(row=rr, column=2, value="（不走收发存表，不参与上面的合计）").font = F(9, False, "8A8F99")
            c = ws.cell(row=rr, column=12, value=v); c.number_format = M; c.alignment = CENN
            c = ws.cell(row=rr, column=13, value="—"); c.alignment = CENN
            for c2 in range(1, 14): ws.cell(row=rr, column=c2).border = BD
            rr += 1
        ws.cell(row=rr, column=1, value="存货类科目合计").font = F(10, True)
        ws.cell(row=rr, column=2, value="含在途/委托加工等单列科目").font = F(9.5)
        c = ws.cell(row=rr, column=12, value="=SUM(L%d:L%d)" % (rr - len(_extra) - 1, rr - 1))
        c.number_format = M; c.font = F(10, True); c.alignment = CENN
        for c2 in range(1, 14):
            ws.cell(row=rr, column=c2).fill = TOT; ws.cell(row=rr, column=c2).border = BD
        rr += 1
    rr += 1
    if _has_x:
        ws.cell(row=rr, column=1,
                value="来源：本块八个金额/数量列都是 **SUMIFS 公式**，实时从「原始·收发存跨维度」按【存货类别】"
                      "汇总（该页 C 列＝存货类别）——点开任一格就能看见完整口径，不必信「工具算过了」。"
                      "「总账科目」由类别↔科目对照反查（基础资料页维护），查不到写「（未对照）」、不硬归。"
                      "**L 列科目余额＝总账取数、M 列＝相减**：同一科目下各类别的结存加起来应等于科目余额，"
                      "差不为 0 会标红（多半是存货类别被改过而总账没同步结转，见「类别漂移」）。").font = F(9)
        rr += 1
    # 数量合计是跨单位相加的，必须写明——否则被当成重量看
    ws.cell(row=rr, column=1, value="注：数量列为各物料【原单位】数量之和（本期含 千克/个/Pcs/米/张 等多种单位），"
                                   "仅供分类件数参考、无物理意义；金额列才是可加的。").font = F(9, False, "8A8F99")

    # 收发存汇总·按仓库（V2.131 新增）：与「按类别」同格式 + 仓库类型 + 结存占比。
    # 按仓库类型分组、组内结存金额降序——与页面「仓库透视」的排法一致，两边对着看不会错行。
    whp = res["pivot_warehouse"]
    wtot_ea = sum(v["ea"] for v in whp.values())
    ws = wb.create_sheet("收发存汇总·按仓库"); ws.sheet_view.showGridLines = False
    for i, w in enumerate([22, 12, 16, 16, 16, 16, 12]): ws.column_dimensions[chr(65 + i)].width = w
    hdr(ws, 1, ["仓库", "仓库类型", "期初金额", "收入金额", "发出金额", "结存金额", "结存占比"])
    ws.freeze_panes = "A2"; rr = 2
    groups = {}
    for wh, v in whp.items():
        groups.setdefault(v.get("type") or "（属性缺失）", []).append((wh, v))
    for t in sorted(groups, key=lambda x: (x.startswith("（"), -sum(v["ea"] for _, v in groups[x]))):
        items = sorted(groups[t], key=lambda kv: -abs(kv[1]["ea"]))
        # 组小计行
        gsum = {k: sum(v[k] for _, v in items) for k in ("oa", "ia", "da", "ea")}
        ws.cell(row=rr, column=1, value="%s（%d 个仓）" % (t, len(items))).font = F(10, True)
        ws.cell(row=rr, column=2, value="小计").font = F(9.5)
        gfirst = rr + 1                       # 小计行在组首，加的是它下面那几行
        for j, k in enumerate(["oa", "ia", "da", "ea"]):
            col = chr(67 + j)
            c = ws.cell(row=rr, column=3 + j,
                        value=("=SUM(%s%d:%s%d)" % (col, gfirst, col, gfirst + len(items) - 1)
                               if _has_x else round(gsum[k], 2)))
            c.number_format = M; c.font = F(10, True)
        sh = clg.share_of(gsum["ea"], wtot_ea)
        c = ws.cell(row=rr, column=7, value=sh if sh is not None else "—")
        if sh is not None: c.number_format = PCT
        c.font = F(10, True); c.alignment = CENN
        for c2 in range(1, 8): ws.cell(row=rr, column=c2).fill = TOT; ws.cell(row=rr, column=c2).border = BD
        rr += 1
        for wh, v in items:
            ws.cell(row=rr, column=1, value="    " + wh).font = F(10)
            ws.cell(row=rr, column=2, value=v.get("type") or "（属性缺失）").font = F(9.5)
            for j, k in enumerate(["oa", "ia", "da", "ea"]):
                # 条件用**常量仓库名**而不是 A 列——A 列为了缩进带了 4 个空格，拿它当条件会匹配不上
                c = ws.cell(row=rr, column=3 + j,
                            value=(xsumifs(k, "H", _q(wh)) if _has_x else round(v[k], 2)))
                c.number_format = M
            sh = clg.share_of(v["ea"], wtot_ea)
            c = ws.cell(row=rr, column=7, value=sh if sh is not None else "—")
            if sh is not None: c.number_format = PCT
            c.alignment = CENN
            for c2 in range(1, 8): ws.cell(row=rr, column=c2).border = BD
            rr += 1
    ws.cell(row=rr, column=1, value="总计").font = F(10, True); ws.cell(row=rr, column=1).fill = TOT
    ws.cell(row=rr, column=2, value="%d 个仓" % len(whp)).font = F(9.5); ws.cell(row=rr, column=2).fill = TOT
    for j, k in enumerate(["oa", "ia", "da", "ea"]):
        col = chr(67 + j)
        # 总计＝各组小计之和：SUMIF「B 列＝小计」，避免把小计行和明细行一起加（会翻倍）
        c = ws.cell(row=rr, column=3 + j,
                    value=("=SUMIF(B2:B%d,\"小计\",%s2:%s%d)" % (rr - 1, col, col, rr - 1)
                           if _has_x else round(sum(v[k] for v in whp.values()), 2)))
        c.number_format = M; c.font = F(10, True); c.fill = TOT
    c = ws.cell(row=rr, column=7, value=1.0 if wtot_ea else "—")
    if wtot_ea: c.number_format = PCT
    c.font = F(10, True); c.fill = TOT; c.alignment = CENN
    for c2 in range(1, 8): ws.cell(row=rr, column=c2).border = BD
    if _has_x:
        ws.cell(row=rr + 2, column=1,
                value="来源：各仓四列均为 **SUMIFS 公式**，实时从「原始·收发存跨维度」按【仓库】汇总"
                      "（该页 H 列＝仓库）；组小计＝组内各仓求和，总计＝各组小计求和。"
                      "仓库→类型的对照在「基础资料」页维护，未配类型的仓库会卡住导出。").font = F(9)

    # 收发存·按事务类型（V2.141）：与页面第④步同款。数据随🅰取数落库；
    # 旧数据/🅱上传通道没有它 → 不出这张表（同损益归集的做法，缺就不出、不出空表）。
    bts = res.get("btypes")
    if bts:
        ws = wb.create_sheet("收发存·按事务类型"); ws.sheet_view.showGridLines = False
        for i, w in enumerate([18, 9, 14, 15, 10, 14, 15, 10]):
            ws.column_dimensions[chr(65 + i)].width = w
        hdr(ws, 1, ["事务类型", "笔数", "收入数量", "收入金额", "收入占比", "发出数量", "发出金额", "发出占比"])
        rr = 2
        t_ia = sum(a["ia"] for a in bts); t_da = sum(a["da"] for a in bts)
        for a in bts:
            ws.cell(row=rr, column=1, value=a["bt"]).font = F(10)
            ws.cell(row=rr, column=2, value=a["n"]).number_format = '#,##0'
            ws.cell(row=rr, column=3, value=a["iq"]).number_format = '#,##0.00'
            ws.cell(row=rr, column=4, value=a["ia"]).number_format = M
            sh = clg.share_of(a["ia"], t_ia)
            c = ws.cell(row=rr, column=5, value=sh if sh is not None else "—")
            if sh is not None: c.number_format = PCT
            c.alignment = CENN
            ws.cell(row=rr, column=6, value=a["dq"]).number_format = '#,##0.00'
            ws.cell(row=rr, column=7, value=a["da"]).number_format = M
            sh = clg.share_of(a["da"], t_da)
            c = ws.cell(row=rr, column=8, value=sh if sh is not None else "—")
            if sh is not None: c.number_format = PCT
            c.alignment = CENN
            if a["ia"] < 0 or a["da"] < 0:            # 红字冲销行（退货/退料）标色
                for c2 in range(1, 9): ws.cell(row=rr, column=c2).fill = RED
            for c2 in range(1, 9): ws.cell(row=rr, column=c2).border = BD
            rr += 1
        ws.cell(row=rr, column=1, value="合计").font = F(10, True)
        ws.cell(row=rr, column=2, value=sum(a["n"] for a in bts)).number_format = '#,##0'
        ws.cell(row=rr, column=3, value=round(sum(a["iq"] for a in bts), 2)).number_format = '#,##0.00'
        ws.cell(row=rr, column=4, value=round(t_ia, 2)).number_format = M
        ws.cell(row=rr, column=5, value=1.0 if t_ia else "—").number_format = PCT
        ws.cell(row=rr, column=6, value=round(sum(a["dq"] for a in bts), 2)).number_format = '#,##0.00'
        ws.cell(row=rr, column=7, value=round(t_da, 2)).number_format = M
        ws.cell(row=rr, column=8, value=1.0 if t_da else "—").number_format = PCT
        for c2 in range(1, 9):
            cell = ws.cell(row=rr, column=c2); cell.fill = TOT; cell.border = BD; cell.font = F(10, True)
        rr += 2
        # 与汇总表如实对比——不写死"同数"（5 期实测收入分毫一致、发出差 1.59 元＝金蝶流水报表自身尾差）
        tt2 = res["pivot_category"]["合计"]
        d_ia, d_da = t_ia - tt2["ia"], t_da - tt2["da"]
        cmp_txt = lambda d: "与「收发存汇总·按类别」分毫一致" if abs(d) < 0.005 \
            else "与「收发存汇总·按类别」差 %.2f 元（金蝶流水报表自身合计与成员行之和的尾差，非工具计算误差）" % d
        ws.cell(row=rr, column=1, value="注：本表按业务事务类型归集本期收入/发出【发生额】（同一本账的另一种切法）。"
                                       "收入合计%s；发出合计%s。负数行＝红字冲销（退货/退料）。数量为原单位混加，仅供参考。"
                                       % (cmp_txt(d_ia), cmp_txt(d_da))).font = F(9, False, "8A8F99")

    ws = wb.create_sheet("异常清单"); ws.sheet_view.showGridLines = False
    for i, w in enumerate([15, 12, 30, 16, 11, 13, 46]): ws.column_dimensions[chr(65 + i)].width = w
    hdr(ws, 1, ["状态", "物料编码", "物料名称", "仓库", "结存数量", "结存金额", "说明"]); rr = 2
    for it in res["anomalies"]["items"]:
        ws.cell(row=rr, column=1, value=it["status"]).alignment = CENN
        ws.cell(row=rr, column=2, value=it["code"]); ws.cell(row=rr, column=3, value=it["name"])
        ws.cell(row=rr, column=4, value=it["wh"]); ws.cell(row=rr, column=5, value=it["eq"]).number_format = '#,##0.00'
        ws.cell(row=rr, column=6, value=it["ea"]).number_format = '#,##0.0000'
        ws.cell(row=rr, column=7, value=it.get("note", "")).font = F(9.5)
        if it["status"] == clg.ST_NEG:
            for c in range(1, 8): ws.cell(row=rr, column=c).fill = RED
        for c in range(1, 8): ws.cell(row=rr, column=c).border = BD
        rr += 1

    if res.get("pnl"):
        ws = wb.create_sheet("损益归集"); ws.sheet_view.showGridLines = False
        for i, w in enumerate([24, 18]): ws.column_dimensions[chr(65 + i)].width = w
        ws['A1'] = "损益归集"; ws['A1'].font = F(13, True, "4B53C4"); rr = 3
        # V2.293：与「原始·货损与处置明细」**显式勾稽**（业务方：「这两个表应该是有勾稽的呀」）。
        # ⚠**不把本页的数换成明细页的公式**——两侧不是同一件事：
        #   本页 ＝【凭证侧】，科目余额表下钻取的管理费用/营业外支出分录，**进总账的就是它，权威**；
        #   明细页 ＝【流水侧】，拿单据号回查收发存流水得到的物料级明细，是**回查**、不是来源。
        # 回查可能对不齐（单据在流水里查不到、或流水自身四舍五入尾差），
        # 硬把本页改成明细页的 SUMIFS，等于让权威数跟着回查结果走——那是本末倒置。
        # 正确做法是并排列出两侧 + 差异，差异不为 0 就自己现形。
        ws.cell(row=rr, column=1, value="① 货损 → 管理费用（按费用项目）").font = F(11, True); rr += 1
        for cat, v in res["pnl"]["loss"]["by_cat"].items():
            ws.cell(row=rr, column=1, value=cat).font = F(10)
            ws.cell(row=rr, column=2, value=v).number_format = M
            rr += 1
        r_loss = rr
        ws.cell(row=rr, column=1, value="货损合计（凭证侧）").font = F(10, True)
        ws.cell(row=rr, column=2, value=res["pnl"]["loss"]["total"]).number_format = M
        ws.cell(row=rr, column=2).font = F(10, True); rr += 1
        if _pd_has:
            ws.cell(row=rr, column=1, value="　明细页回查合计（流水侧）").font = F(10)
            c = ws.cell(row=rr, column=2,
                        value='=SUMIFS({s}!{a}:{a},{s}!{g}:{g},"管理费用·货损")'.format(s=_PDS, a=_PD_AMT, g=_PD_TAG))
            c.number_format = M; rr += 1
            ws.cell(row=rr, column=1, value="　差异（应为 0）").font = F(10, True)
            c = ws.cell(row=rr, column=2, value="=B%d-B%d" % (r_loss, rr - 1))
            c.number_format = M; c.font = F(10, True); rr += 1
        rr += 1
        ws.cell(row=rr, column=1, value="② 资产处置 → 营业外支出").font = F(11, True); rr += 1
        r_disp = rr
        ws.cell(row=rr, column=1, value="处置合计（凭证侧）").font = F(10, True)
        ws.cell(row=rr, column=2, value=res["pnl"]["disposal"]["total"]).number_format = M
        ws.cell(row=rr, column=2).font = F(10, True); rr += 1
        if _pd_has:
            ws.cell(row=rr, column=1, value="　明细页回查合计（流水侧）").font = F(10)
            c = ws.cell(row=rr, column=2,
                        value='=SUMIFS({s}!{a}:{a},{s}!{g}:{g},"营业外支出·处置")'.format(s=_PDS, a=_PD_AMT, g=_PD_TAG))
            c.number_format = M; rr += 1
            ws.cell(row=rr, column=1, value="　差异（应为 0）").font = F(10, True)
            c = ws.cell(row=rr, column=2, value="=B%d-B%d" % (r_disp, rr - 1))
            c.number_format = M; c.font = F(10, True); rr += 2
        # V2.307 第③块：既不是货损也不是资产处置的存货出库（福利领用/捐赠…）。
        # **不计入上面两个合计，但必须列出来**——否则这几笔在台账里凭空消失，
        # 而它们确确实实是本期从存货流向损益的钱。
        _oth = ((raw.get("pnl_detail") or {}).get("other") or [])
        if _oth:
            ws.cell(row=rr, column=1, value="③ 其他存货出库（非货损、非处置）").font = F(11, True); rr += 1
            _by = {}
            for x in _oth:
                _by[x.get("doctype") or "（未注明）"] = _by.get(x.get("doctype") or "（未注明）", 0.0) + (x.get("amount") or 0.0)
            r_oth = None
            first_o = rr
            for k, v in _by.items():
                ws.cell(row=rr, column=1, value=k).font = F(10)
                ws.cell(row=rr, column=2, value=round(v, 2)).number_format = M
                rr += 1
            r_oth = rr
            ws.cell(row=rr, column=1, value="小计（不计入上面两项）").font = F(10, True)
            c = ws.cell(row=rr, column=2, value="=SUM(B%d:B%d)" % (first_o, rr - 1))
            c.number_format = M; c.font = F(10, True); rr += 1
            ws.cell(row=rr, column=1,
                    value="※ 判据＝凭证的【费用项目】（核算维度），不含「货损/盘盈亏/处置」字样的归此档，"
                          "如捐赠。既不是货损也不是资产处置，故单列——"
                          "它们仍是本期从存货流向损益的金额，只是不该混进货损合计。"
                          "明细见「原始·货损与处置明细」，归属列写「其他存货出库（非货损）」。").font = F(9)
            rr += 2
            ws.cell(row=rr, column=1,
                    value="勾稽口径：上方各行为【凭证侧】——科目余额表下钻取的管理费用/营业外支出分录，进总账的就是它。"
                          "「明细页回查合计」是 SUMIFS 公式，实时取自「原始·货损与处置明细」（A 列＝归属，%s 列＝金额），"
                          "那是拿单据号回查收发存流水得到的物料级明细，属**回查**不是来源。"
                          "差异不为 0 有两种可能：①有单据在流水里查不到（明细页会写明「仅凭证级」）；"
                          "②流水报表自身四舍五入尾差（同事务类型透视的 1.59 元）。" % _PD_AMT).font = F(9)
            rr += 1
        # 口径外（V2.314，Owner 定案「6602 福利，不算」）：**列出来但不参与任何合计**，也不进明细页。
        # 静悄悄扔掉的话，业务方拿手工表一比对不上，谁也说不清这笔去哪了——
        # V2.311 那个整页蒸发的 bug 就是这么被发现的，代价是业务方对整个工具的信任。
        _exc = (res["pnl"].get("excluded") or {})
        if _exc.get("total"):
            rr += 2
            ws.cell(row=rr, column=1, value="口径外 · 未计入（6602 管理费用，非货损非盘盈亏）").font = F(11, True, "9298A4")
            rr += 1
            for k, v in (_exc.get("by_item") or {}).items():
                ws.cell(row=rr, column=1, value="　· " + k).font = F(10, False, "9298A4")
                ws.cell(row=rr, column=2, value=v).number_format = M
                ws.cell(row=rr, column=2).font = F(10, False, "9298A4"); rr += 1
            ws.cell(row=rr, column=1, value="小计（不计入本页任何合计）").font = F(10, True, "9298A4")
            c = ws.cell(row=rr, column=2, value=round(_exc["total"], 2))
            c.number_format = M; c.font = F(10, True, "9298A4"); rr += 2
            ws.cell(row=rr, column=1,
                    value="※ 按业务方口径，6602 管理费用里既非货损也非盘盈亏的分录（如福利领用）"
                          "不属成本台账范围：不计入上面任何合计，也不进「原始·货损与处置明细」页。"
                          "此处列出仅为留痕，免得与手工表核对时找不到它的去向。").font = F(9, False, "9298A4")

    # 收发存明细：V2.138 补齐业务方底稿的 17 列（加 物料分组 + 四段单价 + 期初/收入/发出数量）。
    # 单价【单元格存金蝶原值 6 位、格式只显示 2 位】——业务方定的"显示 2 位、计算用 6 位"：
    # 点开格子看到的就是金蝶原数，拿去核对/再计算都不丢精度，表面又不刺眼。
    P2 = '#,##0.00'          # 单价显示 2 位（值仍是 6 位）
    Q = '#,##0.00'           # 数量
    ws = wb.create_sheet("收发存明细"); ws.sheet_view.showGridLines = False
    heads = ["物料编码", "物料名称", "规格型号", "存货类别", "物料分组", "仓库", "批号", "单位",
             "期初数量", "期初单价", "期初金额", "收入数量", "收入单价", "收入金额",
             "发出数量", "发出单价", "发出金额", "结存数量", "结存单价", "结存金额"]
    for i, w in enumerate([12, 26, 18, 11, 12, 14, 11, 7,
                           11, 11, 13, 11, 11, 13, 11, 11, 13, 11, 11, 13]):
        ws.column_dimensions[chr(65 + i)].width = w
    hdr(ws, 1, heads); ws.freeze_panes = "C2"; rr = 2
    TXT = ["code", "name", "spec", "cat", "grp", "wh", "batch", "unit"]
    NUMS = [("oq", Q), ("op", P2), ("oa", M), ("iq", Q), ("ip", P2), ("ia", M),
            ("dq", Q), ("dp", P2), ("da", M), ("eq", Q), ("ep", P2), ("ea", M)]
    for d in cross:
        for j, k in enumerate(TXT):
            ws.cell(row=rr, column=1 + j, value=d.get(k) or "").font = F(9)
        for j, (k, fmt) in enumerate(NUMS):
            v = d.get(k)
            # 单价为 None＝金蝶没给（数量为 0 时如此）→ 写「—」，不写 0：0 会被读成"单价真是零"
            c = ws.cell(row=rr, column=9 + j, value="—" if v is None else v)
            if v is not None:
                c.number_format = fmt
            else:
                c.alignment = CENN
            c.font = F(9)
        if (d.get("eq") or 0) < 0:
            for c in range(1, len(heads) + 1): ws.cell(row=rr, column=c).fill = RED
        rr += 1

    # ── 制造费用三张（V2.259）：与第⑧步屏一一对应。取不到成本计算单则整块不出，
    #    不出空表——空表在 Excel 里比没有更误导（看的人以为"本期没有制造费用"）。
    cost = (res or {}).get("cost") or {}
    ct = cost.get("ties") or {}
    if ct:
        ws = wb.create_sheet("成本勾稽"); ws.sheet_view.showGridLines = False
        for i, w in enumerate([16, 26, 16, 34, 16, 14, 10]): ws.column_dimensions[chr(65 + i)].width = w
        hdr(ws, 1, ["勾稽", "总账侧", "金额", "业务侧", "金额", "差异", "结论"])
        rr = 2
        for key, name, bs, zs in [
                ("mfg_collect", "① 制造费用归集", "5101 制造费用 借方", "成本计算单：制造费用＋间接材料"),
                ("complete", "② 完工结转", "5001 生产成本 贷方",
                 # ⚠事务类型清单由常量生成，别手写：WIP_CREDIT_BTYPES 是"发现一个补一个"的，
                 #   加过「生产入库」之后这行字还停在"汇报入库＋生产退库"，说明跟实算对不上了。
                 "「原始·完工入库流水」：" + "＋".join(clg.WIP_CREDIT_BTYPES)),
                ("wip_input", "③ 投入归集", "5001 生产成本 借方", "本期投入（剔委外）＋期末在产品成本调整")]:
            x = ct.get(key)
            ws.cell(row=rr, column=1, value=name).font = F(10, True)
            if not x:
                ws.cell(row=rr, column=2, value="本期未取到，该道跳过").font = F(9.5, False, "9298A4")
            else:
                ws.cell(row=rr, column=2, value=bs).font = F(9.5)
                ws.cell(row=rr, column=3, value=x["book"]).number_format = M
                ws.cell(row=rr, column=4, value=zs).font = F(9.5)
                ws.cell(row=rr, column=5, value=x["biz"]).number_format = M
                ws.cell(row=rr, column=6, value=x["diff"]).number_format = M
                c = ws.cell(row=rr, column=7, value="通过" if x["pass"] else "不平")
                c.font = F(10, True, "1E7B45" if x["pass"] else "C0392B"); c.alignment = CENN
            for c2 in range(1, 8): ws.cell(row=rr, column=c2).border = BD
            rr += 1
        cp = (ct.get("complete") or {}).get("parts") or {}
        if cp:
            rr += 1
            ws.cell(row=rr, column=1, value="②业务侧明细").font = F(9.5, True)
            ws.cell(row=rr, column=2, value="　".join(f"{k} {v:,.2f}" for k, v in cp.items())).font = F(9.5)
            rr += 1
            ws.cell(row=rr, column=2, value="「生产退库」为负、常被漏算：没有这类单据的月份等式会碰巧成立，"
                                            "有的月份才露馅（2026-3 无、4/5 月有）。").font = F(9, False, "9298A4")

        pc = cost.get("pivot_cc") or {}
        if pc.get("items"):
            ws = wb.create_sheet("车间×成本项目"); ws.sheet_view.showGridLines = False
            items = pc["items"]
            ws.column_dimensions["A"].width = 22
            for i in range(len(items) + 1): ws.column_dimensions[chr(66 + i)].width = 16
            hdr(ws, 1, ["车间（成本中心）"] + items + ["合计"])
            ws.freeze_panes = "B2"; rr = 2
            # 公式口径（🧪 19 格 + 委外 494,998.06 逐格实测一致）：
            #   本块取【成本项目层】（F 列＝层级），**不是**费用项目层——
            #   两层不能一起加（成本项目层的金额就是其下费用项目之和，会翻倍）；
            #   委外（L 列＝是）单列一行，正文行用 <>是 排除。
            def _ccf(item, outsrc, cc_name=None):
                f = ("=SUMIFS({s}!I:I,{s}!F:F,\"成本项目\",{s}!L:L,{o},{s}!G:G,{it}"
                     .format(s=CS, o=('"是"' if outsrc else '"<>是"'), it=_q(item)))
                if cc_name is not None:          # 委外那行是跨车间汇总，不加车间条件
                    f += ",{s}!A:A,{cc}".format(s=CS, cc=_q(cc_name))
                return f + ")"
            for r0 in pc.get("rows", []):
                ws.cell(row=rr, column=1, value=r0["cc"]).font = F(10)
                for j, it in enumerate(items):
                    v = r0["cells"].get(it)
                    c = ws.cell(row=rr, column=2 + j,
                                value=(_ccf(it, False, r0["cc"]) if _has_cc else (round(v, 2) if v else None)))
                    if v or _has_cc: c.number_format = M
                    else: c.alignment = CENN
                c = ws.cell(row=rr, column=2 + len(items),
                            value=("=SUM(B%d:%s%d)" % (rr, get_column_letter(1 + len(items)), rr)
                                   if _has_cc else r0["total"]))
                c.number_format = M
                for c2 in range(1, len(items) + 3): ws.cell(row=rr, column=c2).border = BD
                rr += 1
            osd = pc.get("outsourced") or {}
            if osd.get("total"):
                ws.cell(row=rr, column=1, value="委外（不走生产成本科目）").font = F(10)
                for j, it in enumerate(items):
                    v = (osd.get("cells") or {}).get(it)
                    c = ws.cell(row=rr, column=2 + j,
                                value=(_ccf(it, True) if _has_cc else (round(v, 2) if v else None)))
                    if v or _has_cc: c.number_format = M
                    else: c.alignment = CENN
                c = ws.cell(row=rr, column=2 + len(items),
                            value=("=SUM(B%d:%s%d)" % (rr, get_column_letter(1 + len(items)), rr)
                                   if _has_cc else osd["total"]))
                c.number_format = M
                for c2 in range(1, len(items) + 3): ws.cell(row=rr, column=c2).border = BD
                rr += 1
            ws.cell(row=rr, column=1, value="总计").font = F(10, True)
            for j, it in enumerate(items):
                col = get_column_letter(2 + j)
                c = ws.cell(row=rr, column=2 + j,
                            value=("=SUM(%s2:%s%d)" % (col, col, rr - 1) if _has_cc
                                   else (pc.get("item_total") or {}).get(it)))
                c.number_format = M; c.font = F(10, True)
            col = get_column_letter(2 + len(items))
            c = ws.cell(row=rr, column=2 + len(items),
                        value=("=SUM(%s2:%s%d)" % (col, col, rr - 1) if _has_cc else pc.get("total")))
            c.number_format = M; c.font = F(10, True)
            for c2 in range(1, len(items) + 3):
                ws.cell(row=rr, column=c2).fill = TOT; ws.cell(row=rr, column=c2).border = BD

        exps = cost.get("expenses") or []
        if exps:
            ws = wb.create_sheet("费用项目构成"); ws.sheet_view.showGridLines = False
            for i, w in enumerate([28, 18, 12]): ws.column_dimensions[chr(65 + i)].width = w
            hdr(ws, 1, ["费用项目", "本期投入金额", "占比"])
            ws.freeze_panes = "A2"; rr = 2
            tot_e = sum(x["amount"] for x in exps) or 0.0
            # 🧪 39 项逐项实测一致。本块取【费用项目层】（F 列＝层级），**不剔委外**——
            # 与「车间×成本项目」口径不同（那块剔了、委外另起一行），故两块合计对不上是正常的。
            for x in exps:
                ws.cell(row=rr, column=1, value=x["exp"]).font = F(10)
                c = ws.cell(row=rr, column=2,
                            value=("=SUMIFS({s}!I:I,{s}!F:F,\"费用项目\",{s}!H:H,{e})".format(s=CS, e=_q(x["exp"]))
                                   if _has_cc else x["amount"]))
                c.number_format = M
                c = ws.cell(row=rr, column=3, value=(x["amount"] / tot_e) if tot_e else "—")
                if tot_e: c.number_format = PCT
                c.alignment = CENN
                for c2 in range(1, 4): ws.cell(row=rr, column=c2).border = BD
                rr += 1
            ws.cell(row=rr, column=1, value="合计").font = F(10, True)
            ws.cell(row=rr, column=2, value=f"=SUM(B2:B{rr-1})").font = F(10, True)
            ws.cell(row=rr, column=2).number_format = M
            for c2 in range(1, 4):
                ws.cell(row=rr, column=c2).fill = TOT; ws.cell(row=rr, column=c2).border = BD

    # ── 类别漂移（V2.282）：有才出，没有就不出这张空表
    dr = (res or {}).get("drift") or {}
    if dr.get("items"):
        ws = wb.create_sheet("类别漂移"); ws.sheet_view.showGridLines = False
        for i, w in enumerate([13, 28, 12, 12, 14, 14, 14]): ws.column_dimensions[chr(65 + i)].width = w
        hdr(ws, 1, ["物料编码", "物料名称", "原存货类别", "现存货类别", "原物料分组", "现物料分组", "本期结存金额"])
        ws.freeze_panes = "A2"; rr = 2
        for x in dr["items"]:
            for i, v in enumerate([x["code"], x["name"], x["old_cat"], x["new_cat"], x["old_grp"], x["new_grp"]]):
                c = ws.cell(row=rr, column=1 + i, value=v or "")
                c.font = F(9, i in (3, 5) and (x["cat_changed"] if i == 3 else x["grp_changed"]),
                           "A35A00" if (i in (2, 3) and x["cat_changed"]) or (i in (4, 5) and x["grp_changed"]) else "181B21")
            ws.cell(row=rr, column=7, value=x["ea"]).number_format = M
            for c2 in range(1, 8): ws.cell(row=rr, column=c2).border = BD
            rr += 1
        ws.cell(row=rr + 1, column=1, value="※ 与%s的物料档案比对。金蝶报表按【当前】档案归集类别，"
                                            "改档案会追溯改变历史月份的报表；总账凭证记的是当时的科目、不会追溯变。"
                                            "此表只报不判——归正档案本身多半是对的。" % (dr.get("prev") or "上期")
                ).font = F(9, False, "9298A4")

    # ── 汇总透视合并成一张（V2.279）：业务方「这样就不会有那么多 sheet 页了」。
    #    五块汇总（按类别/按仓库/按事务类型/车间×成本项目/费用项目构成）并进一张，
    #    用 Excel 的**分级显示**（行分组）折叠，左边 +/− 一点即收起。
    #    ⚠**默认展开**，不默认折叠：折叠着打开，第一反应是"数据怎么没了"——
    #      少滚一屏的收益，抵不过让人以为数据缺失的代价。想收起点一下就行。
    #    ⚠不并明细类（收发存明细/两张原始明细）：那三张一万多行、列结构也各不相同，
    #      并进来列宽只能取一套，反而每张都难看。合并只对"块小、列数相近"的汇总有意义。
    SUM_SHEETS = ["收发存汇总·按类别", "收发存汇总·按仓库", "收发存·按事务类型",
                  "车间×成本项目", "费用项目构成"]
    _CCSRC = "「原始·制造费用明细(接口)」"
    SRC_NOTE = {
        "收发存汇总·按类别": "「原始·收发存跨维度」按【存货类别】汇总——**八个数值列全是 SUMIFS 公式**"
                             "（该页 C 列＝存货类别），点开任一格即见完整口径。",
        "收发存汇总·按仓库": "「原始·收发存跨维度」按【仓库】汇总——**四个金额列全是 SUMIFS 公式**"
                             "（该页 H 列＝仓库）；组小计＝组内各仓求和，总计＝各组小计求和。"
                             "仓库→类型的对照在「基础资料」页维护，未配类型的仓库会卡住导出。",
        "收发存·按事务类型": "金蝶存货收发存**流水**按业务类型（FBusinessType）聚合，取数时已在服务端算好。"
                             "⚠**本块没有对应的原始底表页，故只能是值、不能给公式**——"
                             "流水一期十几万行，不落库也不导出；要逐笔核请到金蝶查该期流水。",
        "车间×成本项目": "%s的【成本项目层】，各格为 SUMIFS 公式（F 列＝层级、A 列＝车间、G 列＝成本项目、I 列＝金额）。"
                         "⚠成本项目层的金额＝其下费用项目之和，**两层不可一起求和**（会翻倍），本块只取一层。"
                         "委外（L 列＝是，工单前缀 SUB）单列一行、不并进车间。" % _CCSRC,
        "费用项目构成": "%s的【费用项目层】按子项费用项目归集，各格为 SUMIFS 公式（H 列＝子项费用项目）。"
                       "⚠本块**不剔委外**，与上一块口径不同，两块合计对不上是正常的。" % _CCSRC,
    }
    def merge_sheets(title, names, subtitle, notes):
        """把若干张小表并成一张、每块可折叠。返回是否合并了（少于 2 块就不折腾）。"""
        keep = [n for n in names if n in wb.sheetnames]
        if len(keep) < 2:
            return False
        dst = wb.create_sheet(title, index=wb.sheetnames.index(keep[0]))
        dst.sheet_view.showGridLines = False
        widest = 0
        rr = 1
        dst.cell(row=rr, column=1, value="%s（%d 块）" % (title, len(keep))).font = F(13, True, "4B53C4")
        dst.cell(row=rr + 1, column=1, value=subtitle).font = F(9, False, "9298A4")
        rr = 4
        for name in keep:
            src = wb[name]
            n_col = src.max_column
            widest = max(widest, n_col)
            dst.cell(row=rr, column=1, value="■ " + name).font = F(11, True)
            for c in range(1, n_col + 1):
                dst.cell(row=rr, column=c).fill = TOT
                dst.cell(row=rr, column=c).border = BD
            rr += 1
            first = rr
            # V2.290：每块注明**这块数从哪张原始底表来的**（业务方「是不是要注明来源表格」）。
            # 写在块内、跟着一起折叠——它是这块的一部分，不该在收起后还杵在外面。
            if notes.get(name):
                dst.cell(row=rr, column=1, value="来源：" + notes[name]).font = F(9)
                rr += 1
            _src_r1 = rr          # 源页第 1 行实际落在这一行（「来源」注释行可能已经占掉了 first）
            for row in src.iter_rows(min_row=1, max_row=src.max_row, max_col=n_col):
                for c in row:
                    if c.value is None and not c.has_style:
                        continue
                    v = c.value
                    # ⚠公式必须按新位置**重算引用**（V2.290 修的老 bug，V2.279 起就错）：
                    # 「费用项目构成」的合计原是 =SUM(B2:B40)，整块搬到第 108 行之后，
                    # 公式还写着 B2:B40——加的是上面「按类别/按仓库」块的格子。
                    # 导出的汇总透视页上，那个合计**一直是错的**（源页单独看是对的，所以没被发现）。
                    if isinstance(v, str) and v.startswith("="):
                        v = Translator(v, origin=c.coordinate).translate_formula(
                            "%s%d" % (get_column_letter(c.column), rr))
                    d = dst.cell(row=rr, column=c.column, value=v)
                    d.font = copy(c.font); d.fill = copy(c.fill)
                    d.border = copy(c.border); d.alignment = copy(c.alignment)
                    d.number_format = c.number_format
                rr += 1
            # 合并单元格也得跟着搬（V2.309）：`iter_rows` 只搬值与样式，**合并范围是页级属性**，
            # 不搬的话「科目余额/差异」那两列纵向合并会在并页后全丢，变成每行一个孤零零的数。
            _off = _src_r1 - 1                   # ⚠用源页第 1 行的**实际**落点算偏移：
                                                 #   「来源」注释行会把数据整体下推一行，
                                                 #   拿 first 算会让合并范围整体高一行、压到表头上
            for _mr in list(src.merged_cells.ranges):
                dst.merge_cells(start_row=_mr.min_row + _off, end_row=_mr.max_row + _off,
                                start_column=_mr.min_col, end_column=_mr.max_col)
            # 分组：本块的数据行（含表头）可折叠；标题行留在外面，收起后仍看得见目录
            if rr - 1 >= first:
                dst.row_dimensions.group(first, rr - 1, outline_level=1, hidden=False)
            rr += 1                      # 块之间空一行
        # 列宽取各块最宽的一套（列含义不同，只能取一套，故各块表头必须留着）
        for i in range(widest):
            ws_w = max((wb[n].column_dimensions[chr(65 + i)].width or 0) for n in keep
                       if chr(65 + i) in wb[n].column_dimensions) if any(
                chr(65 + i) in wb[n].column_dimensions for n in keep) else 0
            if ws_w:
                dst.column_dimensions[chr(65 + i)].width = ws_w
        for n in keep:
            wb.remove(wb[n])
        return True

    merge_sheets("汇总透视", SUM_SHEETS,
                 "每块左侧有 −/+ 可折叠（Excel 分级显示）。各块的列含义不同，请看各自的表头行。",
                 SRC_NOTE)

    # ── 三张「对不对得上」的表并成一页（V2.290）：业务方「损益归集和账实勾稽和成本勾稽
    #    是不是可以放到一个 sheet 页」。三张都是"总账侧 vs 业务侧、差多少、结论"，
    #    看的时候本就要来回对照，分三张反而要来回切页。同 D18 的合并口径、默认展开。
    #    ⚠**顺序＝存货 → 成本 → 损益**：账实勾稽是本期能不能出台账的闸（三道勾稽之一），
    #      成本勾稽是另一套账，损益归集是归集结果。按"先判能不能信、再看具体归集"排。
    TIE_SHEETS = ["账实勾稽", "成本勾稽", "损益归集"]
    merge_sheets("勾稽与归集", TIE_SHEETS,
                 "三块都是「总账侧 ↔ 业务侧」的对账，每块左侧 −/+ 可折叠。"
                 "存货三道勾稽的结论在「核对结论」页；本页是它们的明细与另两套账。",
                 {"账实勾稽": "总账存货类科目（14xx）期末余额 ↔ 「原始·收发存跨维度」按类别汇总，"
                              "两侧按【类别↔科目对照】配对（基础资料页维护）。差异不为 0 即本期台账不可出。",
                  "成本勾稽": "总账 5001/5101 ↔ %s 与「原始·完工入库流水」。三条等式均由真数据实证，"
                              "容差 0.5 元（树形报表各层独立四舍五入）。**不并进存货那三道的可信度判定**——"
                              "两边是不同的账，一边不平不该把另一边判为不可信。"
                              "②的业务侧底表只留三种事务类型——整张流水十几万行不落库。" % _CCSRC,
                  "损益归集": "金蝶科目余额表下钻取的凭证分录（管理费用·货损 / 营业外支出·处置），"
                              "明细见「原始·货损与处置明细」。金额口径＝发出−收入，与总账借方同向（盘盈显示为负）。"})

    # ── 原始底表（V2.278）：业务方定「导出＝结论 + 原始底表」一个文件交付。
    #    前面全是核对结论与透视，这里往后是**金蝶原样的明细**，供留痕与逐行复核。
    #    数据取自落库的原始行（非导出时重取）——已封存期间导出的必须与当初核对的是同一份。
    # （raw 已在函数开头取出，见那里的说明）

    # 原样底稿：跨维度表（V2.287）。业务方底稿前两张就是**金蝶两张原表原样**——
    # 「成本X月（按时间）」＋「库存（X月）」。此前只把按日期那张原样出了，跨维度那张却被
    # 重排列序、少三列、改了列名（「基本单位」→「单位」），**同是原表却一张原样一张加工，不对称**。
    # 这里补一张真正原样的：**列序照金蝶原表 23 列**，不用第⑤步那套业务方指定的工作列序。
    # 两张表数据同源、用途不同：本页留痕，「收发存明细」供查询（列序是 D12 定的）。
    if cross:
        ws = wb.create_sheet("原始·收发存跨维度"); ws.sheet_view.showGridLines = False
        heads = ["物料编码", "物料名称", "存货类别", "物料分组", "规格型号", "库存状态", "批号", "仓库",
                 "核算范围编码", "核算范围名称", "基本单位",
                 "期初数量", "期初单价", "期初金额", "收入数量", "收入单价", "收入金额",
                 "发出数量", "发出单价", "发出金额", "结存数量", "结存单价", "结存金额"]
        for i, w in enumerate([12, 26, 11, 12, 18, 10, 12, 14, 13, 14, 8] + [11, 11, 13] * 4):
            ws.column_dimensions[chr(65 + i) if i < 26 else "A" + chr(39 + i)].width = w
        hr = kmeta(ws, ["会计期间:%d年第%d期-%d年第%d期" % (y, p, y, p), "本位币:" + _CUR],
                   "金蝶《存货收发存汇总表（跨维度）》接口只读")
        hdr(ws, hr, heads); ws.freeze_panes = "C%d" % (hr + 1); rr = hr + 1
        TXT = ["code", "name", "cat", "grp", "spec", "status", "batch", "wh",
               "rng_code", "rng_name", "unit"]
        NUMS = [("oq", Q), ("op", P2), ("oa", M), ("iq", Q), ("ip", P2), ("ia", M),
                ("dq", Q), ("dp", P2), ("da", M), ("eq", Q), ("ep", P2), ("ea", M)]
        for d0 in cross:
            for i, k in enumerate(TXT):
                ws.cell(row=rr, column=1 + i, value=d0.get(k) or "").font = F(9)
            for j, (k, fmt) in enumerate(NUMS):
                v = d0.get(k)
                c = ws.cell(row=rr, column=12 + j, value=v)
                if v is not None:
                    c.number_format = fmt
                else:
                    c.alignment = CENN
                c.font = F(9)
            rr += 1
        ws.cell(row=rr + 1, column=1, value="※ 金蝶《存货收发存汇总表（跨维度）》原样，列序同金蝶导出件。"
                                            "小计/合计行已剔除（判据＝库存状态为空，故本页保留该列供复核）。"
                                            ).font = F(9, False, "9298A4")

    bd = raw.get("bydate") or []
    if bd and isinstance(bd[0], dict) and "code" in bd[0]:
        ws = wb.create_sheet("原始·收发存按日期"); ws.sheet_view.showGridLines = False
        # V2.290 列序订正：**照金蝶原件《成本X月（按时间）》的 17 列列序**
        # ——原先写成 编码/名称/规格/类别/分组，金蝶原件是 编码/名称/**分组/规格/类别**。
        # 这页存在的理由就是"跟金蝶导出来的一样"，列序不对就不算原样（同 V2.287 跨维度那次）。
        # 批号/库存状态/单位金蝶原件没有（那三个开关业务方导的时候没打开），但接口取回来了、
        # 且对查批次有用，故**附在 17 列之后**——前 17 列与原件逐列对齐，多的排在后面不打乱它。
        heads = ["物料编码", "物料名称", "物料分组", "规格型号", "存货类别",
                 "期初数量", "期初单价", "期初金额", "收入数量", "收入单价", "收入金额",
                 "发出数量", "发出单价", "发出金额", "结存数量", "结存单价", "结存金额",
                 "批号", "库存状态", "单位"]
        for i, w in enumerate([12, 26, 12, 18, 11] + [11, 11, 13] * 4 + [12, 10, 7]):
            ws.column_dimensions[chr(65 + i) if i < 26 else "A" + chr(39 + i)].width = w
        hr = kmeta(ws, ["日期:【%d-%02d-01】-【%d-%02d-%02d】" % (y, p, y, p, _last), "本位币:" + _CUR],
                   "金蝶《存货收发存汇总表（按日期）》接口只读")
        hdr(ws, hr, heads); ws.freeze_panes = "C%d" % (hr + 1); rr = hr + 1
        TXT = ["code", "name", "grp", "spec", "cat"]
        TAIL = ["batch", "status", "unit"]
        NUMS = [("oq", Q), ("op", P2), ("oa", M), ("iq", Q), ("ip", P2), ("ia", M),
                ("dq", Q), ("dp", P2), ("da", M), ("eq", Q), ("ep", P2), ("ea", M)]
        for d0 in bd:
            for i, k in enumerate(TXT):
                ws.cell(row=rr, column=1 + i, value=d0.get(k) or "").font = F(9)
            for j, (k, fmt) in enumerate(NUMS):
                v = d0.get(k)
                c = ws.cell(row=rr, column=6 + j, value=v)
                if v is not None:
                    c.number_format = fmt
                else:
                    c.alignment = CENN
                c.font = F(9)
            for i, k in enumerate(TAIL):
                ws.cell(row=rr, column=18 + i, value=d0.get(k) or "").font = F(9)
            rr += 1
        ws.cell(row=rr + 1, column=1,
                value="※ 前 17 列＝金蝶《存货收发存汇总表（按日期）》原件列序；"
                      "「批号／库存状态／单位」金蝶原件没有（导出时那几个开关没打开），"
                      "接口能取到、对查批次有用，故附在其后。本表**无仓库维度**，"
                      "同一物料跨仓合并成一行——要按仓库看请用「原始·收发存跨维度」。").font = F(9)

    # 接口取的成本计算单（V2.291 取消 V2.290 的"退休"）。
    # V2.290 我按业务方一句"是不是可以退休了"把它撤了 —— **错的，撤之前没核数据**：
    # 业务方上传的那张只有「制造费用＋直接人工」两个成本项目（🧪 3 月合计 3,900,207.56），
    # **没有直接材料 9,787,932.32 / 间接材料 / 委外加工费**——它是筛过的制造费用部分，
    # 不是完整成本计算单。而「车间×成本项目」「费用项目构成」覆盖全部成本项目，
    # 只能由这张接口版支撑；撤了它，那两块就没有底表可引。
    # ⚠**页名用业务方的叫法，不用金蝶的报表名**（V2.295）：
    # V2.291 我把这页从「原始·制造费用明细」改成「原始·成本计算单(接口)」，本意是区分接口/上传，
    # 结果业务方拿到包一看：「你好像删错表了，我要的是制造费用，你留了成本计算单？」——
    # 页没删，是名字换了，而**业务方和他底稿里都管这张叫「制造费用」**。
    # 改名解决的是我自己的困惑（两张同源表分不清来源），代价却是把用的人认得的东西弄丢了。
    # 现改回「原始·制造费用明细(接口)」——词用他们的。
    # （后缀「(接口)」是当年为了和「(上传)」区分才加的；V2.318 撤掉上传口后只剩这一张，
    #   后缀已无区分作用，但**不改名**：改名的代价上面刚吃过一次。）
    # 报表真名写在页内的〔工具留痕〕行与页脚，出处照样一眼可查。
    cc = raw.get("cc_rows") or []
    if cc:
        ws = wb.create_sheet("原始·制造费用明细(接口)"); ws.sheet_view.showGridLines = False
        heads = ["车间（成本中心）", "产品编码", "产品名称", "工单编号", "单据类型",
                 "层级", "成本项目", "子项费用项目", "本期投入金额", "本期完工数量", "本期完工金额", "委外"]
        for i, w in enumerate([18, 13, 26, 14, 16, 8, 12, 18, 15, 13, 15, 7]):
            ws.column_dimensions[chr(65 + i)].width = w
        hr = kmeta(ws, ["会计期间:%d年%d期-%d年%d期" % (y, p, y, p), "币别:" + _CUR],
                   "金蝶《成本计算单》接口只读（聚合到 工单×成本项目×子项费用项目）")
        hdr(ws, hr, heads); ws.freeze_panes = "D%d" % (hr + 1); rr = hr + 1
        LV = {"item": "成本项目", "exp": "费用项目"}
        for d0 in cc:
            vals = [d0.get("cc") or "", d0.get("prod_no") or "", d0.get("prod_name") or "",
                    d0.get("wo") or "", d0.get("billtype") or "", LV.get(d0.get("level"), ""),
                    d0.get("item") or "", d0.get("exp") or ""]
            for i, v in enumerate(vals):
                ws.cell(row=rr, column=1 + i, value=v).font = F(9)
            for j, k in enumerate(("amt", "cqty", "camt")):
                c = ws.cell(row=rr, column=9 + j, value=d0.get(k))
                c.number_format = Q if k == "cqty" else M
                c.font = F(9)
            ws.cell(row=rr, column=12, value="是" if d0.get("outsourced") else "").font = F(9)
            rr += 1
        ws.cell(row=rr + 1, column=1,
                value="※ 本页取自金蝶《成本计算单》，**覆盖全部成本项目**（直接材料/直接人工/制造费用/"
                      "间接材料/委外加工费），不只是制造费用那一段——页名沿用工具第⑧步「制造费用」的叫法，"
                      "内容以本行为准。").font = F(9)
        ws.cell(row=rr + 2, column=1, value="※ 树形报表：「成本项目」层的金额＝其下「费用项目」之和，"
                                            "两层不可一起求和（会翻倍）。合计请只取一层。").font = F(9)

    pd = raw.get("pnl_detail") or {}
    # ⚠三档**全都要算进出页条件**（V2.311 修）：V2.307 把「福利领用/捐赠」从货损里拆成第三档
    # 「其他」，出页条件却还停在老的两档——于是**只剩「其他」的主体整页蒸发**：
    # 🧪 101 深圳星期零 6 月（26,003.58／5 条）、7 月（2,251.65／2 条）货损处置都是 0，
    #   V2.307 之前它们混在货损里、页出得来，拆完就没了，损益归集③有个小计却没有底表可查。
    # 业务方原话：「星期零的货损那个，你删除了？」——数据一条没删，是这一行的锅。
    if pd.get("loss") or pd.get("disposal") or pd.get("other"):
        ws = wb.create_sheet("原始·货损与处置明细"); ws.sheet_view.showGridLines = False
        def _loss_kind(d0, ents):
            """这一行物料该归哪个【费用项目】。单据只对一条分录就直接用；
            多条时按存货类别大类对回去（产品/包材/原辅料）——匹配不上就留原样，不硬编。"""
            items = [e.get("item") for e in ents if e.get("item")]
            if not items:
                return (ents[0].get("doctype") if ents else "") or ""
            if len(set(items)) == 1:
                return items[0]
            cat = d0.get("cat") or ""
            key = "产品" if cat in ("产成品", "自制半成品", "委外半成品") else ("包材" if cat == "包材" else "原辅料")
            for it in items:
                if it.startswith(key):
                    return it
            return items[0]

        mrows = pd.get("rows") or []
        by_bill = {}
        for x in mrows:
            by_bill.setdefault(x["billno"], []).append(x)
        # 有物料级就出物料级（业务方底稿那页要的是这个粒度）；取不到就退回凭证级，不留空
        # V2.309 列序改成业务方底稿《货损明细-管理费》那套（业务方：「抬头最好还是按照我们之前的格式」）：
        #   原单编号｜类别｜产品项目｜物料编码｜物料名称｜规格型号｜基本单位｜发出数量｜单位基本单价｜金额｜仓库｜备注
        # 两处对应关系（业务方那两列是手工填的，工具直接取数填上）：
        #   **类别**   ＝ 凭证的【费用项目】（产品货损/包材货损/原辅料货损/…）。一张单据拆成多条分录、
        #               每条费用项目不同时（7 月 QTCK011302 就是），按物料的【存货类别】大类对应回去：
        #               产成品/自制半成品/委外半成品→产品、包材→包材、其余→原辅料。
        #               🧪 7 月核对：产品 225,621.63 vs 凭证 225,621.66、包材 72,334.24 vs 72,334.27，
        #               差的几分是流水表自身尾差（同 0.04 那笔）。
        #   **产品项目** ＝ 金蝶的【物料分组】（豆蛋制品/零售其他/植物肉/零售山姆/小料…），取值与底稿一致。
        # 「归属」仍放 A 列——损益归集页的勾稽 SUMIFS 按它筛，挪走要同步改公式。
        # 前 12 列与业务方底稿**逐列一致**（业务方：「归属挪到备注后面，前面几列完全不变」），
        # 「归属」是工具自己加的分类列，排在其后，再跟其余留痕列。
        heads = (["原单编号", "类别", "产品项目", "物料编码", "物料名称", "规格型号",
                  "基本单位", "发出数量", "单位基本单价", "金额", "仓库", "备注",
                  "归属", "批号", "存货类别", "事务类型", "日期", "凭证字号"] if mrows else
                 ["归属", "日期", "凭证字号", "单据号", "单据类型", "金额", "摘要（金蝶原文）"])
        widths = ([13, 12, 12, 12, 24, 16, 8, 11, 12, 13, 14, 22, 14, 10, 10, 14, 11, 10] if mrows else
                  [14, 12, 18, 20, 16, 16, 46])
        _notes = pd.get("notes") or {}
        for i, w in enumerate(widths):
            ws.column_dimensions[chr(65 + i) if i < 26 else "A" + chr(39 + i)].width = w
        hr = kmeta(ws, ["会计期间:%d年%d期" % (y, p), "本位币:" + _CUR],
                   "金蝶科目余额表下钻取凭证分录（管理费用/营业外支出），物料级由单据号回查收发存流水")
        hdr(ws, hr, heads); ws.freeze_panes = "B%d" % (hr + 1); rr = hr + 1
        tot = 0.0
        # ⚠V2.293 修**重复写行**：原来按【凭证分录】循环，一条分录就把该单据的物料行整套写一遍。
        # 一张单据在总账里常常拆成多条分录（7 月 QTCK011302 一张出库单拆成 **13 条**，
        # 报废出库无半成品/无原料/无包材/无植物肉…各记一条），于是同一批 122 行被写了 13 遍：
        # 🧪 明细页 1,586 行、合计 4,218,939.14 ＝ 122 行 × 13 ＝ 324,533.78 × 13，
        # 而损益归集那页是 324,529.82 —— **两页差 13 倍**，业务方一眼看出"这两个表应该是有勾稽的呀"。
        # 3 月每张单只对应一条分录，所以碰巧一直是对的。
        # 改法：**物料行按单据写一次**（seen 去重），凭证字号取该单第一条分录的。
        seen_bills = set()
        for tag, key in [("管理费用·货损", "loss"), ("营业外支出·处置", "disposal"),
                         ("其他存货出库（非货损）", "other")]:
            for x in (pd.get(key) or []):
                ms = by_bill.get(x.get("billno")) if mrows else None
                if ms and x.get("billno") in seen_bills:
                    continue                     # 同一单据的第 2..n 条分录：物料行已写过，跳过
                if ms:
                    seen_bills.add(x.get("billno"))
                    # 该单据的全部凭证分录（可能多条、费用项目不同），供逐行定「类别」
                    _ents = [e for e in (pd.get(key) or []) if e.get("billno") == x.get("billno")]
                    for d0 in ms:
                        vals = [d0["billno"], _loss_kind(d0, _ents), d0["grp"],
                                d0["code"], d0["name"], d0["spec"], d0["unit"]]
                        for i, v in enumerate(vals):
                            ws.cell(row=rr, column=1 + i, value=v or "").font = F(9)
                        ws.cell(row=rr, column=8, value=d0["qty"]).number_format = Q
                        ws.cell(row=rr, column=9, value=d0["price"]).number_format = P2
                        ws.cell(row=rr, column=10, value=d0["amount"]).number_format = M
                        # 补齐行（V2.314）：流水表没吐、从出库单补来的，事务类型列已写明来源
                        for i, v in enumerate([d0["wh"], _notes.get(d0["billno"], ""), tag,
                                               d0["batch"], d0["cat"], d0["btype"], d0["date"],
                                               d0.get("voucher") or x.get("voucher")]):
                            ws.cell(row=rr, column=11 + i, value=v or "").font = F(9)
                        tot += d0["amount"]
                        for c2 in range(1, len(heads) + 1): ws.cell(row=rr, column=c2).border = BD
                        rr += 1
                elif mrows:
                    # 凭证上有、流水表里没对上（如资产处置单不走存货流水）——如实列出，不假装有明细
                    for i, v in enumerate([x.get("billno"), x.get("item") or x.get("doctype")]):
                        ws.cell(row=rr, column=1 + i, value=v or "").font = F(9)
                    ws.cell(row=rr, column=5, value="（该单据不在存货流水中，仅凭证级）").font = F(9, False, "9298A4")
                    ws.cell(row=rr, column=10, value=x.get("amount")).number_format = M
                    ws.cell(row=rr, column=12, value=_notes.get(x.get("billno"), "")).font = F(9)
                    ws.cell(row=rr, column=13, value=tag).font = F(9)
                    tot += x.get("amount") or 0.0
                    rr += 1
                else:
                    for i, v in enumerate([tag, x.get("date"), x.get("voucher"), x.get("billno"), x.get("doctype")]):
                        ws.cell(row=rr, column=1 + i, value=v or "").font = F(9)
                    ws.cell(row=rr, column=6, value=x.get("amount")).number_format = M
                    ws.cell(row=rr, column=7, value=x.get("note") or "").font = F(9)
                    tot += x.get("amount") or 0.0
                    rr += 1
        cN = len(heads)
        ws.cell(row=rr, column=1, value="合计").font = F(10, True)
        # ⚠合计落在【金额】列，不是最后一列（V2.309 列序改了之后金额是 K 列，末列成了凭证字号）；
        #   起始行跟着元数据头走（V2.289）：写死 2 会把元数据行算进合计区。
        _amtc = heads.index("金额") + 1
        c = ws.cell(row=rr, column=_amtc,
                    value="=SUM(%s%d:%s%d)" % (get_column_letter(_amtc), hr + 1,
                                               get_column_letter(_amtc), rr - 1))
        c.number_format = M; c.font = F(10, True)
        for c2 in range(1, cN + 1):
            ws.cell(row=rr, column=c2).fill = TOT; ws.cell(row=rr, column=c2).border = BD
        ws.cell(row=rr + 2, column=1, value="※ 单据来自科目余额表下钻（管理费用/营业外支出的凭证分录，判据＝摘要含单据号）；"
                                            "物料级明细由单据号回查收发存流水表得到。"
                                            "金额口径＝发出−收入，与总账借方同向：盘盈在收入侧，故显示为负。").font = F(9, False, "9298A4")

    # ── 原始·完工入库流水（V2.310）─────────────────────────────────
    # 业务方看着「成本勾稽②」那行问：「然后这几张表我们也有导出来吗？」——
    #   成本计算单侧有（就是上面那张「原始·制造费用明细(接口)」），**流水侧一直没有**，
    #   ②的业务侧 6,146,975.01 是个孤零零的数、点不进去。
    # 十几万行的收发存流水整张不落库也不导出（那条口径不变），但②只用得到三种事务类型：
    #   🧪 7 月实测 225 行 / 全表 11,984 行＝**1.9%**——这 1.9% 落库，②就有底表可查了。
    _wip = raw.get("wip_rows") or []
    if _wip:
        ws = wb.create_sheet("原始·完工入库流水"); ws.sheet_view.showGridLines = False
        heads = ["事务类型", "单据编号", "日期", "物料编码", "物料名称", "规格型号", "产品项目",
                 "基本单位", "收入数量", "单位基本单价", "金额", "仓库", "批号", "存货类别", "凭证字号"]
        for i, w in enumerate([12, 14, 11, 12, 24, 16, 12, 8, 11, 12, 13, 14, 10, 10, 10]):
            ws.column_dimensions[get_column_letter(i + 1)].width = w
        hr = kmeta(ws, ["会计期间:%d年%d期" % (y, p), "本位币:" + _CUR],
                   "收发存流水表，仅取「%s」三种事务类型" % "/".join(clg.WIP_CREDIT_BTYPES))
        hdr(ws, hr, heads); ws.freeze_panes = "C%d" % (hr + 1); rr = hr + 1
        for d0 in sorted(_wip, key=lambda x: (x["btype"], x["billno"])):
            for i, v in enumerate([d0["btype"], d0["billno"], d0["date"], d0["code"],
                                   d0["name"], d0["spec"], d0["grp"], d0["unit"]]):
                ws.cell(row=rr, column=1 + i, value=v or "").font = F(9)
            ws.cell(row=rr, column=9, value=d0["qty"]).number_format = Q
            ws.cell(row=rr, column=10, value=d0["price"]).number_format = P2
            ws.cell(row=rr, column=11, value=d0["amount"]).number_format = M
            for i, v in enumerate([d0["wh"], d0["batch"], d0["cat"], d0.get("voucher")]):
                ws.cell(row=rr, column=12 + i, value=v or "").font = F(9)
            for c2 in range(1, len(heads) + 1): ws.cell(row=rr, column=c2).border = BD
            rr += 1
        ws.cell(row=rr, column=1, value="合计").font = F(10, True)
        c = ws.cell(row=rr, column=11, value="=SUM(K%d:K%d)" % (hr + 1, rr - 1))
        c.number_format = M; c.font = F(10, True)
        for c2 in range(1, len(heads) + 1):
            ws.cell(row=rr, column=c2).fill = TOT; ws.cell(row=rr, column=c2).border = BD
        ws.cell(row=rr + 2, column=1,
                value="※ 本页是「成本勾稽②·完工结转」业务侧那个数的底表：整张流水表十几万行不落库，"
                      "只留这三种事务类型（%d 行）。金额＝收入金额，与②的业务侧口径一致。"
                      % len(_wip)).font = F(9, False, "9298A4")

    # ── 页签配色（V2.289）──────────────────────────────────────────
    # 一个文件十来张页，「结论」和「原始底表」混在一排页签里分不出来。按**来源**分色：
    #   蓝＝工具算出来的结论/透视（可以据此下判断）
    #   灰＝金蝶接口只读取回的原始底表（留痕，不该改）
    # 原本还有一档 橙＝人工上传的原表，V2.318 撤掉上传口后不存在了。
    # 只上色不改页序：页序是 D17/D18 定的「结论在前、底表在后」，配色不能顺手动它。
    # 在最后统一按页名上色，而不是各建各的时候上——「汇总透视」是把 5 张合并后新建的，
    # 建页时上色会漏掉合并出来的那张。
    TAB = {"conc": "4B53C4", "raw": "8A94A6"}
    for w in wb.worksheets:
        if w.title.startswith("原始·"):
            w.sheet_properties.tabColor = TAB["raw"]
        else:
            w.sheet_properties.tabColor = TAB["conc"]
    # 配色说明写在首页——不解释的颜色等于没有颜色
    ws0 = wb["核对结论"]
    _lg = ws0.max_row + 2
    ws0.cell(row=_lg, column=1, value="页签配色").font = F(10, True)
    for i, (t, col) in enumerate([
            ("■ 蓝　工具算出来的结论与透视", TAB["conc"]),
            ("■ 灰　金蝶接口只读取回的原始底表（留痕用，请勿改动）", TAB["raw"]),
]):
        ws0.cell(row=_lg + 1 + i, column=1, value=t).font = F(9, False, col)

    # 三张收发存长得像，用途完全不同——不写清楚就会被当成重复数据（V2.290 业务方问「有什么区别」）
    _lg += 5
    ws0.cell(row=_lg, column=1, value="三张收发存表怎么区分").font = F(10, True)
    for i, t in enumerate([
            "「原始·收发存跨维度」　金蝶原件 23 列原样，**含仓库/批号维度**，一物料一仓一批一行 —— 留痕、逐仓复核；"
            "上面所有按类别/按仓库的汇总都从它来。",
            "「原始·收发存按日期」　金蝶原件列序，**无仓库维度**，同一物料跨仓合并成一行 —— "
            "与跨维度表**互为校验**（两表四项金额合计相等＝第①道勾稽「两表互勾」）。",
            "「收发存明细」　　　　同跨维度的数据，但列序是**业务方指定的工作列序**（编码/名称/规格/类别/分组/仓库/批号/单位…）"
            "—— 供页面查询与人眼逐行看，不是原件。要对金蝶请用「原始·」那两张。"]):
        ws0.cell(row=_lg + 1 + i, column=1, value=t).font = F(9)

    # ── 去掉说明文字里的 Markdown 星号（V2.294）────────────────────
    # 代码注释里写 `**强调**` 是习惯，但这些字串会原样落进单元格——Excel 不认 Markdown，
    # 打开看到的就是「**含仓库/批号维度**」。统一在收尾处剥掉。
    # ⚠**原始· 页只剥工具自己写的那几行**（※ 说明、〔工具留痕〕、来源/注）：
    #   那些页的正文是金蝶原件/业务方上传件，逐格一致是它们的契约，
    #   哪怕真有个规格型号写成 "A**B"，也不该被这个清理动作改掉。
    _NOTE_HEAD = ("※", "〔工具留痕〕", "来源：", "注：", "勾稽口径：", "页签配色")
    for w in wb.worksheets:
        _verbatim = w.title.startswith("原始·")
        for row in w.iter_rows():
            for c in row:
                v = c.value
                if not isinstance(v, str) or "**" not in v or v.startswith("="):
                    continue
                if _verbatim and not v.startswith(_NOTE_HEAD):
                    continue
                c.value = v.replace("**", "")

    bio = BytesIO(); wb.save(bio); return bio.getvalue()

# -*- coding: utf-8 -*-
# [Change Log]
# Date: 2026-07-04 | Author: Claude / c | Version: V2.4
# Description: 科目余额表内核（账面核对）。
#   - build_rows_kingdee: GL_BALANCE 期初(按 科目+维度 去重) + 序时账本期借贷 → 还原实时科目余额表。
#     金蝶 GL_BALANCE 接口的"期末/本期发生"在凭证未过账时停在期初/返回0，不能直接用；
#     须「期初＋本期序时账」还原（与金蝶科目余额表报表同公式，V1.7 已逐户验证一致）。
#   - build_rows_sample: 样例行直读（样例数据自带期初/借/贷/期末）。
#   - parse_report_xlsx: 解析"金蝶界面导出的科目余额表 Excel"——容错认列：
#     单金额列格式(期初原币/本期借方原币/…)与 借/贷 两行表头格式都认。
#   - compare: 工具数 vs 上传数，按科目逐项核对(期初/本期借方/本期贷方/期末)，供人眼核对。
import re

MONEY_PREFIXES = ("1001", "1002", "1012", "1101")
CAT = {"1001": "库存现金", "1002": "银行存款", "1012": "其它货币资金", "1101": "交易性金融资产"}
ITEMS = ("期初", "本期借方", "本期贷方", "期末")


def to_f(v):
    if v is None or v == "":
        return 0.0
    try:
        return float(str(v).replace(",", "").replace("￥", "").replace("¥", "").strip() or 0)
    except Exception:
        return 0.0


def cat_of(code):
    for p in MONEY_PREFIXES:
        if str(code).startswith(p):
            return CAT[p]
    return ""


def build_rows_kingdee(bal_rows, vou_rows):
    """真金蝶：期初(去重) + 序时账借贷 → 每 (科目, 账户维度) 一行。返回 list[dict]。"""
    opens, seen, names, meta = {}, set(), {}, {}
    for r in bal_rows:
        code = str(r.get("科目编码") or "")
        if not code.startswith(MONEY_PREFIXES):
            continue
        dim = str(r.get("核算维度.银行账号.编码") or "").strip()
        key = (code, dim)
        if key in seen:
            continue                      # GL_BALANCE 会返回重复行，须按(科目,维度)去重（V1.7 教训）
        seen.add(key)
        opens[key] = to_f(r.get("期初原币"))
        names.setdefault(code, r.get("科目名称") or "")
        meta[key] = {"账户": (r.get("核算维度.银行账号.名称") or dim or "—"),
                     "币别": (r.get("币别") or "CNY")}
    moves = {}
    for r in vou_rows:
        code = str(r.get("科目编码") or "")
        if not code.startswith(MONEY_PREFIXES):
            continue
        dim = str(r.get("FDetailID.FF100002.FNumber") or "").strip()
        key = (code, dim)
        m = moves.setdefault(key, [0.0, 0.0])
        m[0] += to_f(r.get("FDEBIT"))
        m[1] += to_f(r.get("FCREDIT"))
        names.setdefault(code, r.get("科目名称") or "")
    rows = []
    for key in sorted(set(opens) | set(moves)):
        code, dim = key
        op = opens.get(key, 0.0)
        d, c = moves.get(key, [0.0, 0.0])
        mt = meta.get(key) or {}
        rows.append({"科目编码": code, "科目名称": names.get(code, ""), "科目大类": cat_of(code),
                     "账户": mt.get("账户") or dim or "—", "币别": mt.get("币别") or "CNY",
                     "期初": round(op, 2), "本期借方": round(d, 2), "本期贷方": round(c, 2),
                     "期末": round(op + d - c, 2)})
    return rows


def build_rows_sample(bal_rows):
    """样例模式：样例余额行直读（自带期初/借/贷/期末）。"""
    rows = []
    for r in bal_rows:
        code = str(r.get("科目编码") or "")
        if not code.startswith(MONEY_PREFIXES):
            continue
        rows.append({"科目编码": code, "科目名称": r.get("科目名称") or "", "科目大类": cat_of(code),
                     "账户": (r.get("核算维度.银行账号.名称") or "—"), "币别": r.get("币别") or "CNY",
                     "期初": round(to_f(r.get("期初原币")), 2),
                     "本期借方": round(to_f(r.get("本期借方原币")), 2),
                     "本期贷方": round(to_f(r.get("本期贷方原币")), 2),
                     "期末": round(to_f(r.get("期末原币")), 2)})
    rows.sort(key=lambda x: (x["科目编码"], x["账户"]))
    return rows


# ---------------- 上传的金蝶科目余额表 Excel：容错解析 ----------------
def _labels(grid, h):
    """表头行(可能两行：期初余额 跨列 + 下行 借方/贷方) → 每列合成标签。返回 (labels, data_start)。"""
    top = list(grid[h])
    nxt = list(grid[h + 1]) if h + 1 < len(grid) else []
    ff, last = [], ""
    for v in top:                          # 顶行向右填充（合并单元格只有左上有值）
        s = str(v).strip() if v not in (None, "") else ""
        last = s if s else last
        ff.append(last)
    two_row = any(str(v).strip() in ("借方", "贷方", "借", "贷") for v in nxt if v not in (None, ""))
    labels = []
    for j, base in enumerate(ff):
        sub = str(nxt[j]).strip() if (two_row and j < len(nxt) and nxt[j] not in (None, "")) else ""
        labels.append(base + sub)
    return labels, (h + 2 if two_row else h + 1)


def _pick(labels, must, exclude=(), prefer=()):
    """按关键词挑列：含 must 全部、不含 exclude 任何；多命中优先含 prefer 的。返回列号或 None。"""
    hits = [j for j, lb in enumerate(labels)
            if lb and all(m in lb for m in must) and not any(x in lb for x in exclude)]
    if not hits:
        return None
    for p in prefer:
        for j in hits:
            if p in labels[j]:
                return j
    return hits[0]


def parse_report_xlsx(path):
    """解析金蝶导出的科目余额表 xlsx → (per_code dict, err)。
    per_code = {科目编码: {"科目名称":…, "期初":…, "本期借方":…, "本期贷方":…, "期末":…}}（只收四类资金科目）。"""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        return None, f"打不开这个文件（请确认是 .xlsx 格式的 Excel；老 .xls 请在 Excel 里另存为 .xlsx）：{e}"
    for ws in wb.worksheets:
        grid = [list(row) for row in ws.iter_rows(values_only=True)]
        hdr = next((i for i, row in enumerate(grid[:20])
                    if any(v and "科目编码" in str(v) for v in row)), None)
        if hdr is None:
            continue
        labels, start = _labels(grid, hdr)
        c_code = _pick(labels, ["科目编码"])
        c_name = _pick(labels, ["科目名称"]) or _pick(labels, ["科目全名"])
        # 期初：单列(期初原币/期初余额) 或 借贷两列
        c_qc = _pick(labels, ["期初"], exclude=["借", "贷", "本位"], prefer=["原币"])
        c_qcj = _pick(labels, ["期初", "借"]); c_qcd = _pick(labels, ["期初", "贷"])
        # 本期发生：借/贷 各一列（"本期借方原币"或"本期发生额借方"）
        c_jf = _pick(labels, ["借"], exclude=["期初", "期末", "本年", "累计"], prefer=["本期", "原币"])
        c_df = _pick(labels, ["贷"], exclude=["期初", "期末", "本年", "累计"], prefer=["本期", "原币"])
        # 期末：单列 或 借贷两列
        c_qm = _pick(labels, ["期末"], exclude=["借", "贷", "本位"], prefer=["原币"])
        c_qmj = _pick(labels, ["期末", "借"]); c_qmd = _pick(labels, ["期末", "贷"])
        if c_code is None or (c_qc is None and c_qcj is None):
            continue
        per = {}
        for row in grid[start:]:
            code = str(row[c_code]).strip() if (c_code < len(row) and row[c_code] not in (None, "")) else ""
            if code.endswith(".0"):
                code = code[:-2]           # Excel 里科目编码若是数值格式会带 .0
            if not re.match(r"^\d{4}", code) or not code.startswith(MONEY_PREFIXES):
                continue
            g = lambda j: to_f(row[j]) if (j is not None and j < len(row)) else 0.0
            qc = g(c_qc) if c_qc is not None else g(c_qcj) - g(c_qcd)      # 资产科目：期初=借-贷
            qm = g(c_qm) if c_qm is not None else (g(c_qmj) - g(c_qmd) if c_qmj is not None else None)
            d = per.setdefault(code, {"科目名称": "", "期初": 0.0, "本期借方": 0.0, "本期贷方": 0.0, "期末": 0.0, "_qm": qm is not None})
            if c_name is not None and c_name < len(row) and row[c_name] and not d["科目名称"]:
                d["科目名称"] = str(row[c_name]).strip()
            d["期初"] += qc
            d["本期借方"] += g(c_jf)
            d["本期贷方"] += g(c_df)
            d["期末"] += (qm or 0.0)
        if per:
            for d in per.values():
                if not d.pop("_qm"):       # 表里没有期末列 → 按 期初+借-贷 补算
                    d["期末"] = d["期初"] + d["本期借方"] - d["本期贷方"]
                for k in ITEMS:
                    d[k] = round(d[k], 2)
            return per, None
    return None, "没在表里认出「科目编码」及金额列。请上传金蝶导出的《科目余额表》Excel（表头须含 科目编码、期初、借方、贷方、期末）。"


def compare(tool_rows, uploaded):
    """工具数(逐维度行) vs 上传数(按科目) → 按科目逐项核对。返回 {"rows":…, "科目数":…, "一致数":…}。"""
    tool = {}
    for r in tool_rows:
        d = tool.setdefault(r["科目编码"], {"科目名称": r.get("科目名称", ""),
                                            "期初": 0.0, "本期借方": 0.0, "本期贷方": 0.0, "期末": 0.0})
        for k in ITEMS:
            d[k] += r.get(k) or 0.0
    out = []
    for code in sorted(set(tool) | set(uploaded)):
        t, u = tool.get(code), uploaded.get(code)
        row = {"科目编码": code,
               "科目名称": (t or u or {}).get("科目名称", ""), "科目大类": cat_of(code)}
        if t and u:
            ok = True
            for k in ITEMS:
                tv, uv = round(t[k], 2), round(u[k], 2)
                row[k + "_工具"], row[k + "_报表"] = tv, uv
                row[k + "_差"] = round(tv - uv, 2)
                ok = ok and abs(tv - uv) < 0.005
            row["结果"] = "一致" if ok else "有出入"
        else:
            for k in ITEMS:
                row[k + "_工具"] = round(t[k], 2) if t else None
                row[k + "_报表"] = round(u[k], 2) if u else None
                row[k + "_差"] = None
            row["结果"] = "报表里没有此科目" if t else "工具里没有此科目"
        out.append(row)
    n_ok = sum(1 for r in out if r["结果"] == "一致")
    return {"rows": out, "科目数": len(out), "一致数": n_ok, "全部一致": n_ok == len(out) and len(out) > 0}

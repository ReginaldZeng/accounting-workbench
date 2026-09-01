# -*- coding: utf-8 -*-
# [Change Log]
# Date: 2026-09-01 | Author: Claude / c | Version: V2.416
# Description: 新增财资多份导出归并回归（TestTreasuryMultiExport）：截断版让位、同账户取最全、
#              同日同额真实重复交易不被误杀、清单说明讲人话。复刻 2026-08 真实流水包形态。
# Date: 2026-08-03 | Author: Claude / c | Version: V2.167
# Description: 银行流水导入·识别层单元测试。重点回归"改名兜底"：文件名对不上时按内容认
#   （财资平台 xlsx 表头 / 中行 HISQRY 块头），建行 csv 不被误抢、乱文件仍跳过、原名快路不读文件。
#   夹具全部合成落临时目录，不依赖真实流水。
import os
import sys
import tempfile
import unittest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from kernels import bank_import as bi

from openpyxl import Workbook


# ---- 合成夹具 ----
def make_treasury_xlsx(path):
    """财资平台导出的最小复刻：账户头两行 + 明细表头 + 两笔明细（结构对齐真实导出）。"""
    wb = Workbook()
    ws = wb.active
    ws.append(["账户明细", "账号：", "73110122000157061"])
    ws.append(["", "户名：", "测试食品科技有限公司"])
    ws.append(["", "交易时间", "本方户名", "本方账号", "收入", "支出", "账户余额",
               "对方户名", "对方账号", "币种", "交易备注", "交易类型"])
    ws.append(["", "2026-07-01 16:37:03", "测试食品科技有限公司", "73110122000157061",
               "", "2,550.00", "365,442.33", "微笑办公用品商行", "44306643601801001783",
               "人民币", "办公用品", "转账"])
    ws.append(["", "2026-07-01 19:12:01", "测试食品科技有限公司", "73110122000157061",
               "31,200.00", "", "396,642.33", "农业科技（上海）有限公司", "50131000927752036",
               "人民币", "货款", "网银转账"])
    wb.save(path)


def make_treasury_multi(path, accounts):
    """多账户财资导出夹具：一个明细表头 + 各账户若干笔。accounts = [(账号, 户名, [(日期, 收入, 支出)])]。"""
    wb = Workbook()
    ws = wb.active
    ws.append(["", "交易时间", "本方户名", "本方账号", "收入", "支出", "账户余额",
               "对方户名", "对方账号", "币种", "交易备注", "交易类型"])
    for acct, holder, rows in accounts:
        for d, inc, out in rows:
            ws.append(["", d + " 10:00:00", holder, acct, inc, out, "0.00",
                       "某某对方公司", "999", "人民币", "备注", "转账"])
    wb.save(path)


def make_hisqry_csv(path):
    """中行 HISQRY 最小复刻（V2.199 真实形态）：块头 + 明细表头(付款人/收款人名称) + 来账/往账各一笔。
    含「付款人开户行名称」陷阱列——银行名也带"公司"，修复前会被误当对方户名（需求方实查案例）。"""
    lines = [
        "查询账号,1234567890123",
        "总笔数,2",
        "借方发生总额,50.00",
        "贷方发生总额,267.10",
        "交易类型[ Transaction Type ],业务类型,付款人开户行名称,付款人账号,付款人名称[ Payer's Name ],"
        "收款人账号,收款人名称[ Payee's Name ],交易日期,交易金额,交易后余额",
        "来账,小额普通,中信银行股份有限公司东莞分行,744800018260,东莞市绿邦实业有限公司,"
        "1234567890123,深圳市星期零食品科技有限公司,20260703,+267.10,\"2,719,648.72\"",
        "往账,跨行汇款,宁波银行股份有限公司深圳分行,1234567890123,深圳市星期零食品科技有限公司,"
        "50131000927752,某某贸易有限公司,20260705,-50.00,\"2,719,598.72\"",
    ]
    with open(path, "w", encoding="gbk", newline="") as f:
        f.write("\n".join(lines))


def break_declared_dimension(path):
    """复刻银行生成器的坑：把 sheet1.xml 的 <dimension> 错报成 A1:A1。
    openpyxl read_only 模式会信这个声明、整表读成空——嗅探必须 reset_dimensions 才见内容。"""
    import re
    import zipfile
    tmp = path + ".tmp"
    with zipfile.ZipFile(path) as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.namelist():
            data = zin.read(item)
            if item == "xl/worksheets/sheet1.xml":
                data = re.sub(rb'<dimension ref="[^"]*"/>', b'<dimension ref="A1:A1"/>', data)
            zout.writestr(item, data)
    os.replace(tmp, path)


def make_ccb_csv(path):
    """建行 host-to-host 明细最小复刻（gb18030，表头含 借方发生额/贷方发生额/账号）。"""
    lines = [
        "账号,账户名称,币种,记账日期,摘要,借方发生额（支取）,贷方发生额（收入）,余额,对方户名",
        "62001234567,测试食品科技有限公司,人民币,2026-07-01,货款,0.00,100.00,100.00,某某贸易有限公司",
    ]
    with open(path, "w", encoding="gb18030", newline="") as f:
        f.write("\n".join(lines))


class TestRenameFallback(unittest.TestCase):
    """V2.167：同事随手改名后，内容还在就必须认得出（此前整包被静默标"跳过"）。"""

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.dir = self.tmp.name

    def tearDown(self):
        self.tmp.cleanup()

    # ---- 快路：原名靠文件名认，不必读文件（路径不存在也能分类）----
    def test_original_names_fast_path(self):
        self.assertEqual(bi._classify("宁波银行+招商银行20260731.xlsx"), "treasury")
        self.assertEqual(bi._classify("HISQRY_202607.csv"), "hisqry")
        self.assertEqual(bi._classify("2088123456.xls"), "alipay")
        self.assertEqual(bi._classify("1576978681_2026.csv"), "wechat")
        self.assertEqual(bi._classify("DL20260701.csv"), "douyin")

    # ---- 兜底：改名的财资平台 xlsx 按表头认回 ----
    def test_renamed_treasury_recognized(self):
        p = os.path.join(self.dir, "宁波+招商.xlsx")     # 真实案例的改名方式
        make_treasury_xlsx(p)
        self.assertTrue(bi.is_treasury_xlsx(p))
        self.assertEqual(bi._classify(p), "treasury")
        rows = bi.parse_treasury(p)
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["账号"], "73110122000157061")
        self.assertEqual(rows[0]["支出"], 2550.00)
        self.assertEqual(rows[1]["收入"], 31200.00)

    # ---- 兜底：改名 + 尺寸声明错报 A1:A1（7月真实案例的完整形态）也要认得出 ----
    def test_renamed_treasury_with_bad_dimension(self):
        p = os.path.join(self.dir, "宁波+招商.xlsx")
        make_treasury_xlsx(p)
        break_declared_dimension(p)
        self.assertTrue(bi.is_treasury_xlsx(p))
        self.assertEqual(bi._classify(p), "treasury")
        self.assertEqual(len(bi.parse_treasury(p)), 2)   # 解析器(完整模式)本就不受影响，一并锁住

    # ---- 兜底：改名的中行 HISQRY csv 按块头认回 ----
    def test_renamed_hisqry_recognized(self):
        p = os.path.join(self.dir, "中行7月流水.csv")
        make_hisqry_csv(p)
        self.assertTrue(bi.is_hisqry_csv(p))
        self.assertEqual(bi._classify(p), "hisqry")
        rows, blocks = bi.parse_hisqry(p)
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["收入"], 267.10)
        self.assertEqual(rows[1]["支出"], 50.00)
        self.assertEqual(blocks[0]["总笔数"], 2)
        # V2.199：来账对方=付款人名称（不是含"公司"的开户行名），往账对方=收款人名称（不是本方）
        self.assertEqual(rows[0]["对方户名"], "东莞市绿邦实业有限公司")
        self.assertEqual(rows[1]["对方户名"], "某某贸易有限公司")

    def test_hisqry_fallback_excludes_bank(self):
        # 无表头老格式走启发式兜底：银行/分行名让位给真正的交易对手
        self.assertEqual(bi._counterparty(
            ["来账", "汇款", "中信银行股份有限公司东莞分行", "东莞市绿邦实业有限公司"], "123"),
            "东莞市绿邦实业有限公司")

    # ---- 防误抢：建行 csv 表头是「账号」不带"查询"，仍走建行通道 ----
    def test_ccb_not_stolen_by_hisqry_sniff(self):
        p = os.path.join(self.dir, "随手改名的建行流水.csv")
        make_ccb_csv(p)
        self.assertFalse(bi.is_hisqry_csv(p))
        self.assertEqual(bi._classify(p), "skip")   # skip 后由 load_bank_dir 的建行内容检查接住
        self.assertTrue(bi.is_ccb_csv(p))

    # ---- 乱文件：内容对不上的照旧跳过，不能瞎认 ----
    def test_junk_still_skipped(self):
        px = os.path.join(self.dir, "理财持仓.xlsx")
        wb = Workbook()
        wb.active.append(["产品名称", "净值", "份额"])
        wb.save(px)
        pc = os.path.join(self.dir, "备注.csv")
        with open(pc, "w", encoding="utf-8") as f:
            f.write("a,b,c\n1,2,3")
        self.assertEqual(bi._classify(px), "skip")
        self.assertEqual(bi._classify(pc), "skip")

    # ---- 端到端：混合目录一次扫，改名件全并入、乱文件标跳过 ----
    def test_load_bank_dir_end_to_end(self):
        make_treasury_xlsx(os.path.join(self.dir, "宁波+招商.xlsx"))
        make_hisqry_csv(os.path.join(self.dir, "中行7月流水.csv"))
        make_ccb_csv(os.path.join(self.dir, "建行明细.csv"))
        with open(os.path.join(self.dir, "说明.txt"), "w", encoding="utf-8") as f:
            f.write("与流水无关的备注")
        rows, manifest = bi.load_bank_dir(self.dir)
        types = {m["文件"]: m["类型"] for m in manifest}
        self.assertEqual(types["宁波+招商.xlsx"], "财资平台·宁波+招商")
        self.assertEqual(types["中行7月流水.csv"], "中行HISQRY")
        self.assertEqual(types["建行明细.csv"], "建设银行·明细CSV")
        self.assertEqual(types["说明.txt"], "跳过")
        banks = {r["银行"] for r in rows}
        self.assertIn("中国银行", banks)
        self.assertIn("建设银行", banks)
        self.assertEqual(len(rows), 5)   # 财资2 + 中行2 + 建行1


class TestRarCmd(unittest.TestCase):
    """RAR 解压命令拼装：unar 恒带 -p（空密码探测加密包→上层能报"请填密码"而非通用天书）。
    2026-09-01 服务器实测背景：7z 无 Rar 码流、unrar-free 解不了 RAR5，unar 是唯一 RAR5 退路。"""

    def test_unar_probes_with_empty_password(self):
        self.assertEqual(bi._rar_cmd("unar", "/usr/bin/unar", "a.rar", "/d", None)[-2:], ["-p", ""])
        self.assertEqual(bi._rar_cmd("unar", "/usr/bin/unar", "a.rar", "/d", "s3cret")[-2:], ["-p", "s3cret"])

    def test_other_tools_unchanged(self):
        self.assertIn("-p", bi._rar_cmd("7z", "7z", "a.rar", "/d", None)[3])       # 7z 本就恒带 -p
        self.assertIn("-p-", bi._rar_cmd("unrar", "unrar", "a.rar", "/d", None))   # unrar 无密码用 -p-


class TestTreasuryMultiExport(unittest.TestCase):
    """财资多份导出按账户取最全（复刻 2026-08 真实流水包：出纳导了三次、一次比一次全，
    「1」是被行数上限截断的残版；旧逻辑只并第一份，静默漏账）。"""

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.dir = self.tmp.name

    def tearDown(self):
        self.tmp.cleanup()

    def test_pick_fullest_per_account(self):
        A, B, C = "73110122000157061", "86041110000117736", "755965480410106"
        make_treasury_multi(os.path.join(self.dir, "财资银行流水 1.xlsx"), [
            (A, "星期零", [("2026-08-01", "100.00", ""), ("2026-08-02", "", "50.00")]),      # 截断：少第3笔
        ])
        make_treasury_multi(os.path.join(self.dir, "财资银行流水 2.xlsx"), [
            (A, "星期零", [("2026-08-01", "100.00", ""), ("2026-08-02", "", "50.00"),
                           ("2026-08-03", "7.00", "")]),                                     # A 的全版
            (B, "星期九", [("2026-08-10", "1,865,786.33", ""), ("2026-08-10", "1,865,786.33", "")]),  # 同日同额真实两笔
        ])
        make_treasury_multi(os.path.join(self.dir, "财资银行流水 3.xlsx"), [
            (B, "星期九", [("2026-08-10", "1,865,786.33", ""), ("2026-08-10", "1,865,786.33", "")]),  # 与文件2并列
            (C, "星期十", [("2026-08-20", "", "25.00")]),                                    # 只有这份有
        ])
        rows, manifest = bi.load_bank_dir(self.dir)
        by = {}
        for x in rows:
            by.setdefault(x["账号"], []).append(x)
        self.assertEqual(len(by[A]), 3)      # 取全版，不是截断版
        self.assertEqual(len(by[B]), 2)      # 同日同额两笔都在——整户整取，没被按行"去重"误杀
        self.assertEqual(len(by[C]), 1)
        self.assertEqual(len(rows), 6)
        man = {m["文件"]: m for m in manifest}
        m1, m2, m3 = (man["财资银行流水 %d.xlsx" % i] for i in (1, 2, 3))
        self.assertFalse(m1["并入逐笔"])                              # 残版整份让位
        self.assertEqual(m1["类型"], "财资平台·宁波+招商")             # 但不再被归为"PDF/理财跳过"
        self.assertIn("逐笔核对", m1.get("说明", ""))
        self.assertTrue(m2["并入逐笔"])
        self.assertEqual((m2["笔数"], m2["账户数"]), (5, 2))          # A全版3 + B两笔（并列取排序靠前的文件）
        self.assertTrue(m3["并入逐笔"])
        self.assertEqual((m3["笔数"], m3["账户数"]), (1, 1))          # 只并 C；B 让位给文件2
        self.assertIn("更全", m3.get("说明", ""))
        self.assertTrue(bi.needs_dup_confirm(manifest))               # 出现让位 → 必须弹窗人工确认

    def test_tie_goes_to_fuller_file(self):
        # 复刻 8 月困惑点：同账户两份一模一样时，归到【整体更全的文件】名下——
        # 页面显示"全集那份并入全部、残版重复"，与人的直觉（"直接用 3"）一致
        r1, r2 = ("2026-08-01", "100.00", ""), ("2026-08-02", "", "50.00")
        make_treasury_multi(os.path.join(self.dir, "财资银行流水 1.xlsx"), [
            ("111", "星期零", [r1, r2]),                              # 残版：只有 111
        ])
        make_treasury_multi(os.path.join(self.dir, "财资银行流水 2.xlsx"), [
            ("111", "星期零", [r1, r2]),                              # 与残版一模一样
            ("222", "星期九", [("2026-08-05", "7.00", "")]),          # 多出的账户 → 整体更全
        ])
        rows, manifest = bi.load_bank_dir(self.dir)
        man = {m["文件"]: m for m in manifest}
        self.assertEqual((man["财资银行流水 2.xlsx"]["笔数"], man["财资银行流水 2.xlsx"]["账户数"]), (3, 2))
        self.assertNotIn("说明", man["财资银行流水 2.xlsx"])          # 全集干干净净一行，无让位说明
        self.assertFalse(man["财资银行流水 1.xlsx"]["并入逐笔"])       # 残版整份"重复"
        self.assertEqual(len(rows), 3)

    def test_non_subset_flags_warning(self):
        # 两次导出范围不同：让位文件里有 1 笔不在被采用的导出里 → 必须飘红提示，不许静默丢
        make_treasury_multi(os.path.join(self.dir, "财资银行流水 1.xlsx"), [
            ("111", "星期零", [("2026-08-01", "100.00", ""), ("2026-08-15", "", "88.00")]),  # 8-15 只有这份有
        ])
        make_treasury_multi(os.path.join(self.dir, "财资银行流水 2.xlsx"), [
            ("111", "星期零", [("2026-08-01", "100.00", ""), ("2026-08-02", "", "50.00"),
                               ("2026-08-03", "7.00", "")]),                                 # 更全（3笔）但缺 8-15
        ])
        rows, manifest = bi.load_bank_dir(self.dir)
        self.assertEqual(len(rows), 3)                       # 仍按整户择优并入更全那份
        man = {m["文件"]: m for m in manifest}
        m1 = man["财资银行流水 1.xlsx"]
        self.assertFalse(m1["并入逐笔"])
        self.assertIn("⚠", m1["说明"])                       # 疑漏必须飘红
        self.assertIn("1 笔", m1["说明"])
        self.assertIn("人工核对", m1["说明"])

    def test_subset_says_checked(self):
        # 真子集：让位文件的说明要讲清"已逐笔核对、无遗漏"，不能让人以为没解析
        make_treasury_multi(os.path.join(self.dir, "财资银行流水 1.xlsx"), [
            ("111", "星期零", [("2026-08-01", "100.00", "")]),
        ])
        make_treasury_multi(os.path.join(self.dir, "财资银行流水 2.xlsx"), [
            ("111", "星期零", [("2026-08-01", "100.00", ""), ("2026-08-02", "", "50.00")]),
        ])
        _rows, manifest = bi.load_bank_dir(self.dir)
        m1 = {m["文件"]: m for m in manifest}["财资银行流水 1.xlsx"]
        self.assertNotIn("⚠", m1["说明"])
        self.assertIn("逐笔核对", m1["说明"])

    def test_single_file_unchanged(self):
        # 单份导出走老口径：全并入、无让位说明
        make_treasury_multi(os.path.join(self.dir, "宁波+招商.xlsx"), [
            ("111", "星期零", [("2026-08-01", "1.00", "")]),
            ("222", "星期九", [("2026-08-02", "", "2.00")]),
        ])
        rows, manifest = bi.load_bank_dir(self.dir)
        self.assertEqual(len(rows), 2)
        m = manifest[0]
        self.assertTrue(m["并入逐笔"])
        self.assertEqual((m["笔数"], m["账户数"]), (2, 2))
        self.assertNotIn("说明", m)
        self.assertFalse(bi.needs_dup_confirm(manifest))              # 单份干净导出 → 不弹确认窗


if __name__ == "__main__":
    unittest.main(verbosity=2)

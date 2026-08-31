# -*- coding: utf-8 -*-
# [Change Log]
# Date: 2026-08-06 | Author: Claude / c | Version: V2.195
# Description: 物流计提·账单直采内核（重构方案 v2.1 §9，2026-08-05 POC 实证工程化）。
#   物流部上传「核对后账单」（对账时已逐票标注 类型/主体/金蝶单号）→ 按文件名认商 → 各家适配器解析成票级记录
#   → 类型标注翻译（logistics_type_map 精确键优先，内核规则兜底，再不行进待人工）
#   → 聚合成计提明细行（主体×物流商×费用归属×业务线×描述，一行=一张凭证，D-2）
#   → 预填做账维度（logistics_fee_map 四级取数 + logistics_bizline 产品维度）→ 活表行（前端可改，服务端以行值为准）。
#   验收基线（2026-07 真账单包 POC）：7 文件 208 票 212,120.96 → 28 行，待人工 2 行 1,786.08。
#   票级明细（带金蝶单号）随行返回存档——对账线三期"回填单号直查比对"的数据源，两线源头合流。
import re
from collections import defaultdict

from kernels import logistics_accrual as la

# 文件名关键词 → (简称, 全名, 渠道)。全名=金蝶 BD_Supplier 口径（极鲜达/顺丰冷运已账证核对；
# 其余转录自 v1.8 供应商列表/账单抬头，录入时 check_voucher_for_post 对金蝶档案兜底拦截）。
# TODO(P2 后续)：并入供应商维表页面维护，此表退役为种子。
CARRIERS = [
    ("丰源", "丰源", "湖北丰源物流供应链管理有限公司", "线下"),
    ("易风达", "易风达", "武汉易风达冷链物流有限公司", "线下"),
    ("极鲜达", "极鲜达", "湖北极鲜达供应链有限责任公司", "线下"),
    ("跨越", "跨越物流", "深圳市跨越速运有限公司", "线下"),
    ("链盟", "链盟", "东莞市链盟供应链有限公司", "线下"),
    ("顺丰冷运", "顺丰冷运", "上海顺丰冷运供应链有限公司", "线下"),
    ("顺丰速运", "顺丰速运", "深圳顺丰速运有限公司", "线下"),
    ("比翼", "比翼电商仓", "厦门比翼信息科技有限公司", "线上"),
    ("顺鸽", "顺鸽电商仓", "武汉顺鸽科技有限公司", "线上"),
]
SUBJECT_ALIAS = {"深圳星期零": "深圳星期零", "孝感星期九": "孝感星期九", "深圳星期九": "深圳星期九",
                 "星期零": "深圳星期零"}   # "星期九"歧义（D-10）不猜——落待确认


def _s(v):
    if v is None:
        return ""
    try:
        import math
        if isinstance(v, float) and math.isnan(v):
            return ""
    except Exception:
        pass
    return str(v).strip()


def _num(v):
    try:
        f = float(str(v).replace(",", "").strip())
        return f if f == f else None
    except (TypeError, ValueError):
        return None


_SUFFIXES = ("电商仓", "物流", "速运", "冷运", "快运")


def identify_carrier(filename, suppliers=None):
    """按文件名认商。suppliers=维表行列表(V2.198 起路由传 db.list_logi_suppliers())；None=用内置 CARRIERS 兜底。
    二段匹配：①全简称包含 ②去后缀核心词包含(≥2字，简称长者优先——防"顺丰"抢在"顺丰冷运"前)。"""
    if suppliers is None:
        for kw, short, full, channel in CARRIERS:
            if kw in filename:
                return {"简称": short, "全名": full, "渠道": channel}
        return None
    rows = sorted(suppliers, key=lambda r: -len(r.get("short") or ""))
    for r in rows:
        if r.get("short") and r["short"] in filename:
            return {"简称": r["short"], "全名": r.get("full") or r["short"], "渠道": r.get("channel") or "线下"}
    for r in rows:
        core = r.get("short") or ""
        for suf in _SUFFIXES:
            core = core.replace(suf, "")
        if len(core) >= 2 and core in filename:
            return {"简称": r["short"], "全名": r.get("full") or r["short"], "渠道": r.get("channel") or "线下"}
    return None


# ---------------- 各家适配器：DataFrame(dict of sheets) -> 票级记录 ----------------
# 票级记录: {单号, 类型标注, 主体标注, 金额, 特殊, 日期, 说明}
def _stop(cell0):
    return _s(cell0).replace(" ", "") in ("合计", "总计")


def parse_fengyuan(sheets, month=0):
    tickets = []
    for name, df in sheets.items():
        if df.shape[1] < 16:
            continue
        for r in range(3, len(df)):
            if _s(df.iat[r, 2]) in ("合计", "总计"):
                break
            amt = _num(df.iat[r, 15])
            if amt:
                tickets.append({"单号": _s(df.iat[r, 1])[:10], "类型标注": "", "主体标注": "深圳星期零",
                                "金额": amt, "特殊": "无标注", "说明": _s(df.iat[r, 8])})
    return tickets


def parse_yifengda(sheets, month=0):
    tickets = []
    mm = str(int(month)) if month else ""
    for name, df in sheets.items():
        nm = name.strip()
        if "运费" in nm:
            if mm and f"{mm}月" not in nm:     # 账单常带补账月份 sheet,只取本账期(V2.191 解析坑)
                continue
            for r in range(3, len(df)):
                if _stop(df.iat[r, 0]):
                    break
                amt = _num(df.iat[r, 14]) if df.shape[1] > 14 else None
                if amt:
                    tickets.append({"单号": _s(df.iat[r, 16]) if df.shape[1] > 16 else "",
                                    "类型标注": _s(df.iat[r, 17]) if df.shape[1] > 17 else "",
                                    "主体标注": _s(df.iat[r, 18]) if df.shape[1] > 18 else "",
                                    "金额": amt, "特殊": "", "说明": ""})
        elif nm == f"{mm}月" or nm == f"{mm}月仓储费" or (not mm and ("仓储" in nm or re.fullmatch(r"\d+月", nm))):
            # 仓储日表：板位费(c7)+装卸费(c8) 按列求和（勿抓尾行——V2.191 解析坑）；「N月」=玉湖仓、「N月仓储费」=山绿仓
            wname = "山绿仓" if "仓储" in nm else "玉湖仓"
            fee = 0.0
            for r in range(2, len(df)):
                for c in (7, 8):
                    v = _num(df.iat[r, c]) if df.shape[1] > c else None
                    if v and _s(df.iat[r, 0]) not in ("", "上月结余"):
                        fee += v
            if fee > 0.005:
                tickets.append({"单号": "", "类型标注": f"仓储费·{wname}",
                                "主体标注": "孝感星期九",     # 账证:易风达仓储做账=孝感簿(账单表头"星期零"具迷惑性)
                                "金额": round(fee, 2), "特殊": "", "说明": nm})
    return tickets


def parse_jixianda(sheets, month=0):
    tickets = []
    for name, df in sheets.items():
        if "零担" in name or "快运" in name:
            for r in range(2, len(df)):
                if _stop(df.iat[r, 0]) or "总计" in _s(df.iat[r, 16]):
                    break                      # 合计行以下=开票信息等附注,不再扫(V2.195 修:continue 会捡尾部杂行)
                amt = _num(df.iat[r, 16])
                if amt:
                    sp = "到付" if "到付" in _s(df.iat[r, 17]) else ""
                    tickets.append({"单号": _s(df.iat[r, 18]), "类型标注": _s(df.iat[r, 19]),
                                    "主体标注": _s(df.iat[r, 20]), "金额": amt, "特殊": sp,
                                    "说明": _s(df.iat[r, 3])})
        elif "山姆" in name:             # sheet 级规则：整表=山姆送仓（无类型列）
            for r in range(2, len(df)):
                if _stop(df.iat[r, 0]):
                    break
                amt = _num(df.iat[r, 14])
                if amt:
                    tickets.append({"单号": _s(df.iat[r, 1]), "类型标注": "山姆送仓",
                                    "主体标注": "深圳星期零", "金额": amt, "特殊": "",
                                    "说明": _s(df.iat[r, 3])})
    return tickets


def parse_kuayue(sheets, month=0):
    tickets = []
    for name, df in sheets.items():
        if df.shape[1] < 15:
            continue
        for r in range(4, len(df)):
            if "总计" in _s(df.iat[r, 12]) or _stop(df.iat[r, 0]):
                break
            amt = _num(df.iat[r, 12])
            if amt is not None and _s(df.iat[r, 3]):
                subj = "深圳星期零" if "星期零" in _s(df.iat[r, 1]) else _s(df.iat[r, 1])
                tickets.append({"单号": _s(df.iat[r, 13]), "类型标注": _s(df.iat[r, 14]),
                                "主体标注": subj, "金额": amt, "特殊": "", "说明": ""})
    return tickets


def parse_lianmeng(sheets, month=0):
    tickets = []
    for name, df in sheets.items():
        if df.shape[1] < 14:
            continue
        for r in range(3, len(df)):
            amt = _num(df.iat[r, 8])
            if amt and _s(df.iat[r, 0]):
                desc = "路凯卡板" if "路凯" in _s(df.iat[r, 3]) else ""
                tickets.append({"单号": _s(df.iat[r, 11]), "类型标注": _s(df.iat[r, 12]),
                                "主体标注": _s(df.iat[r, 13]), "金额": amt, "特殊": "",
                                "说明": desc or _s(df.iat[r, 3])})
    return tickets


def parse_sf_lengyun(sheets, month=0):
    tickets = []
    for name, df in sheets.items():
        if name.strip() not in SUBJECT_ALIAS:          # 主体名 sheet 才是明细（月结卡=上海分公司,费用主体按 sheet 分）
            continue
        for r in range(2, len(df)):
            if _stop(df.iat[r, 0]):
                break
            amt = _num(df.iat[r, 11])
            if amt:
                tickets.append({"单号": _s(df.iat[r, 14]), "类型标注": _s(df.iat[r, 15]),
                                "主体标注": name.strip(), "金额": amt, "特殊": "",
                                "说明": _s(df.iat[r, 13])})
    return tickets


def parse_sf_suyun(sheets, month=0):
    tickets = []
    for name, df in sheets.items():
        if "明细" not in name:
            continue
        for r in range(2, len(df)):
            if _stop(df.iat[r, 0]):
                break
            amt = _num(df.iat[r, 11])
            if amt and _s(df.iat[r, 2]):
                # 账单卡=深圳星期零；经手人列标"星期九"（D-10 主体歧义，不猜，备注带出待确认）
                tickets.append({"单号": _s(df.iat[r, 14]), "类型标注": _s(df.iat[r, 15]),
                                "主体标注": "深圳星期零", "金额": amt, "特殊": "",
                                "说明": "经手人标注:" + _s(df.iat[r, 12])})
    return tickets


ADAPTERS = {"丰源": parse_fengyuan, "易风达": parse_yifengda, "极鲜达": parse_jixianda,
            "跨越物流": parse_kuayue, "链盟": parse_lianmeng,
            "顺丰冷运": parse_sf_lengyun, "顺丰速运": parse_sf_suyun}


# ---------------- 类型标注翻译：维表精确键 → 内核规则兜底 → 待人工 ----------------
_BIZWORDS = [("山姆零售", "山姆零售"), ("零售其他", "零售"), ("kikiherb", "kikiherb"),
             ("植物肉", "植物肉"), ("豆蛋制品", "豆蛋制品"), ("小料", "小料"), ("鲜食", "鲜食"),
             ("电商", "电商"), ("海外", "海外"), ("零售", "零售")]


def _bizline_of(t):
    for k, v in _BIZWORDS:
        if k in t:
            return v
    return "—"


def translate_rule(t):
    """规则兜底（维表查不到时）。返回 (fee, bizline, descr) 或 None。
    ⚠判定顺序=特称先于泛称（V2.237 修）：研发/设备 要在"调拨"之前（"设备调拨-需求部门"曾被泛"调拨"
    抢判成成品调拨）；原料调拨/原料仓储/入库装卸 同理要在各自泛称之前。标注规范的自由后缀
    （设备调拨-需求部门X、研发费用-项目名）都靠这里兜——后缀随意写，前缀定类。"""
    if not t:
        return None
    if "设备调拨" in t:   # 要在"研发"之前——需求部门名可能带"研发"（如 设备调拨-永续研发中心）
        return ("设备调拨费用", "—", t)
    if "研发" in t:
        return ("研发设备采购", "—", t)
    if "设备" in t:
        return ("设备调拨费用", "—", t)
    if "原料入库装卸" in t:
        return ("原料入库装卸费用", _bizline_of(t), t)
    if "成品入库装卸" in t:
        return ("成品入库装卸费用", _bizline_of(t), t)
    if "装卸" in t:
        return ("出库装卸费用", _bizline_of(t), t)
    if "原料仓储" in t:
        return ("原料仓储费用", _bizline_of(t.replace("原料仓储", "")), "")
    if "仓储" in t:
        return ("成品仓储费用", _bizline_of(t), t)
    if "原料入库" in t:
        return ("原料入库费用", _bizline_of(t.replace("原料入库", "")), "")
    if "成品入库" in t:
        return ("成品入库费用", _bizline_of(t.replace("成品入库", "")), "")
    if "原料调拨" in t:
        return ("原料调拨费用", _bizline_of(t.replace("原料调拨", "")), "")
    if "调拨" in t:
        return ("成品调拨费用", _bizline_of(t), "")
    if "样品" in t:
        return ("销售出库费用", _bizline_of(t.replace("样品", "")), "样品")
    if "其它出库" in t or "其他出库" in t:
        return ("销售出库费用", _bizline_of(t), "其它出库")
    if "销售" in t or "出库" in t:
        return ("销售出库费用", _bizline_of(t), "")
    return None


def translate(typ, special, tmap):
    """(费用归属, 业务线, 描述, 翻译来源)。费用归属空=待人工。"""
    t = _s(typ)
    if special == "到付":
        return ("", _bizline_of(t), "到付·是否计提待定", "到付")
    if not t or special == "无标注":
        return ("", "—", "账单无类型标注", "无标注")
    hit = (tmap or {}).get(t)
    if hit:
        fee, biz, desc = hit
        return (fee, biz or "—", desc, "翻译表")
    r = translate_rule(t)
    if r:
        return (r[0], r[1], r[2], "规则")
    return ("", "—", f"标注无法翻译:{t}", "翻译失败")


# ---------------- 主流程：文件集 -> 活表行 + 票级明细 ----------------
def parse_bill_files(files, month, rates=None, fee_lk=None, biz_lk=None, tmap=None, suppliers=None):
    """files: [(filename, bytes)]。返回 {rows, tickets, stats, unknown_files}。
    rows=活表行（vouchers 同构+维度预填），前端可改维度；服务端录入时以行值为准重建凭证。"""
    import io as _io
    import pandas as pd
    all_tickets = []
    unknown = []
    per_file = []
    for fn, data in files:
        car = identify_carrier(fn, suppliers)
        if not car:
            unknown.append(fn)
            continue
        if car["简称"] not in ADAPTERS:
            per_file.append({"文件": fn, "物流商": car["简称"], "状态": "暂无解析器(新商需开发适配)", "票数": 0, "金额": 0})
            continue
        try:
            sheets = pd.read_excel(_io.BytesIO(data), sheet_name=None, header=None)
        except Exception as e:
            msg = str(e)
            if "xlrd" in msg:   # 服务器环境缺 .xls 读取组件（2026-08-06 部署首日实测：链盟 .xls 全军覆没）
                msg = "服务器缺 .xls 读取组件——请管理员在宝塔 Python 项目执行 pip install xlrd 并重启，再重传本文件（或把账单另存为 .xlsx 上传）"
            per_file.append({"文件": fn, "物流商": car["简称"], "状态": f"读取失败:{msg}", "票数": 0, "金额": 0})
            continue
        tk = ADAPTERS[car["简称"]](sheets, month)
        for t in tk:
            t["物流商"] = car["简称"]
            t["公司全名"] = car["全名"]
            t["渠道"] = car["渠道"]
        all_tickets += tk
        per_file.append({"文件": fn, "物流商": car["简称"], "状态": "已解析",
                         "票数": len(tk), "金额": round(sum(t["金额"] for t in tk), 2)})

    # 聚合：主体×商×费用归属×业务线×描述（一行=一张凭证）
    agg = defaultdict(lambda: {"金额": 0.0, "票数": 0, "标注": set(), "特殊": set(), "说明": set(), "单号数": 0})
    for t in all_tickets:
        fee, biz, desc, src = translate(t["类型标注"], t["特殊"], tmap)
        subj = SUBJECT_ALIAS.get(t["主体标注"], t["主体标注"])
        key = (subj, t["物流商"], t["公司全名"], t["渠道"], fee, biz, desc)
        a = agg[key]
        a["金额"] += t["金额"]
        a["票数"] += 1
        if t["类型标注"]:
            a["标注"].add(t["类型标注"])
        if t["特殊"]:
            a["特殊"].add(t["特殊"])
        note = t.get("说明") or ""
        if note.startswith("经手人标注"):
            a["说明"].add(note)
        if t.get("单号"):
            a["单号数"] += 1

    rows = []
    for (subj, short, full, channel, fee, biz, desc), a in agg.items():
        gross = round(a["金额"], 2)
        remark = f"账单直采·{a['票数']}票·源标注[{'/'.join(sorted(a['标注'])) or '无'}]"
        if a["特殊"]:
            remark += "·" + "/".join(sorted(a["特殊"]))
        if a["说明"]:
            remark += "·" + ";".join(sorted(a["说明"])) + "(主体待确认)"
        rows.append(build_row(subj, short, full, channel, fee, biz, desc, gross, month,
                              rates=rates, fee_lk=fee_lk, biz_lk=biz_lk,
                              tickets=a["票数"], remark=remark))
    rows.sort(key=lambda x: (x["物流商"], x["费用归属"], x["业务线"]))
    n_pend = sum(1 for r in rows if not r["费用归属"])
    stats = {"文件数": len(per_file), "票数": len(all_tickets),
             "含税合计": round(sum(t["金额"] for t in all_tickets), 2),
             "明细行数": len(rows), "待人工行": n_pend,
             "待人工金额": round(sum(r["含税"] for r in rows if not r["费用归属"]), 2)}
    return {"rows": rows, "per_file": per_file, "tickets": all_tickets,
            "stats": stats, "unknown_files": unknown}


# ---------------- B 期·费用率（V2.197 第一刀）：分母=BP 工作台各事业单元不含税收入 ----------------
# BP /api/performance/actuals 的 byUnit 键 → 物流业务线（2026-08-06 实测 BP 返回的单元名）。
# 口径：BP 收入=应收单不含税(FNoTaxAmountFor_D)；费率=未税物流费÷不含税收入——分子分母天然同口径。
# TODO(后续)：并入维表页面维护；BP 新增单元名会落「未对照」显性放出，不静默丢。
BP_UNIT_MAP = {
    "山姆 鲜食": "鲜食", "山姆 零售": "山姆零售", "其它线下零售渠道": "零售",
    "小料": "小料", "植物肉": "植物肉", "豆蛋制品": "豆蛋制品",
    "星期零电商自营店": "电商", "星期零电商一件代发": "电商",
    "Kiki Herb电商自营店": "kikiherb",
}


# 豆蛋制品系列关键词：BP 管报把豆蛋并进「植物肉」事业单元（业务方 2026-08-06 指出），
# 费率端按 byUnitBrandSeries 的系列名拆回独立业务线。关键词口径待业务方核（TODO 维表化）。
_EGG_SERIES_KW = ("鸡蛋豆腐", "双蛋白")


def fetch_bp_revenue(year, bp_base=None, timeout=6):
    """调 BP 工作台拿全年逐月各业务线不含税收入。返回 {period: {byBiz, total, unmapped, guming, egg_split}}；
    BP 不可达 → None（降级，不拖垮页面）。两个口径修正（业务方 2026-08-06）：
    ①豆蛋从 BP 植物肉单元按系列关键词拆出独立线；②古茗（自提、收入过大稀释小料费率）单独取出供"非古茗"口径。"""
    import json as _json
    import urllib.request
    import db as _db
    base = bp_base or _db.BP_API_BASE
    try:
        with urllib.request.urlopen(f"{base}/api/performance/actuals?year={int(year)}", timeout=timeout) as r:
            d = _json.loads(r.read().decode("utf-8"))
    except Exception:
        return None
    out = {}
    for period, m in (d.get("months") or {}).items():
        by = {}
        unmapped = 0.0
        for unit, amt in (m.get("byUnit") or {}).items():
            biz = BP_UNIT_MAP.get(unit)
            if biz:
                by[biz] = round(by.get(biz, 0) + float(amt or 0), 2)
            else:
                unmapped += float(amt or 0)
        # ① 豆蛋从植物肉单元拆出（按品牌×系列，系列名命中关键词）
        egg = 0.0
        for _brand, smap in ((m.get("byUnitBrandSeries") or {}).get("植物肉") or {}).items():
            for series, amt in (smap or {}).items():
                if any(kw in series for kw in _EGG_SERIES_KW):
                    egg += float(amt or 0)
        if egg:
            by["豆蛋制品"] = round(by.get("豆蛋制品", 0) + egg, 2)
            by["植物肉"] = round(by.get("植物肉", 0) - egg, 2)
        # ② 古茗收入（小料单元下的古茗品牌）
        gm = float((((m.get("byUnitBrand") or {}).get("小料") or {}).get("古茗")) or 0)
        pend = m.get("pending") or {}
        una = m.get("unassigned") or {}
        out[period] = {"byBiz": by, "total": float(m.get("total") or 0), "unmapped": round(unmapped, 2),
                       "guming": round(gm, 2), "egg_split": round(egg, 2),
                       # 数据新鲜度与未归线口径（V2.200：BP 批次停在月中导致"收入对不上300万"的实战教训）
                       "fetched_at": m.get("fetchedAt") or "", "batch": m.get("batchId"),
                       "pending": round(float(pend.get("amount") or 0), 2), "pending_rows": pend.get("rows") or 0,
                       "unassigned": round(float(una.get("amount") or 0), 2)}
    return out


# 金蝶序时账查上月已录物流计提费用（环比基数）。物流费用项目编码前缀（V2.192 教训：按费用项目锁定，
# 不按摘要——"计提%"混着返利/代运营）。产品分类 CPFL011 双义：带 CPXM017=山姆零售、不带=零售。
_LOGI_ITEM_PREFIX = ("FYXM008.002", "FYXM002.002", "FYXM005.002", "FYXM005.003")
_CPFL_TO_BIZ = {"CPFL007": "植物肉", "CPFL010": "鲜食", "CPFL008": "豆蛋制品",
                "CPFL009": "小料", "CPFL002": "电商", "CPFL013": "kikiherb"}


def fetch_ledger_expense(year, period):
    """序时账查某期已录物流计提费用(未税=借方)，按业务线聚合。返回 {业务线: 金额, "—": 无业务线合计}；金蝶不可达→None。"""
    import kingdee_client as kc
    FIELDS = [("FACCOUNTID.FNumber", "科目"), ("FDEBIT", "借"), ("FEXPLANATION", "摘要"),
              ("FDetailID.FFLEX9.FNumber", "费项"), ("FDetailID.FF100010.FNumber", "产分"),
              ("FDetailID.FF100006.FNumber", "产项")]
    try:
        s, conf = kc.login()
        rows = kc._query(s, conf, "GL_VOUCHER", FIELDS,
                         f"FYear={int(year)} AND FPeriod={int(period)} AND FEXPLANATION LIKE '计提%'")
    except Exception:
        return None
    out = {}
    for r in rows:
        debit = float(r["借"] or 0)
        item = str(r["费项"] or "")
        acc = str(r["科目"] or "")[:4]
        if debit <= 0 or acc not in ("6601", "6401", "5101", "6604", "6402"):
            continue
        if not any(item.startswith(p) for p in _LOGI_ITEM_PREFIX):
            continue
        cpfl = str(r["产分"] or "")
        cpxm = str(r["产项"] or "")
        if cpfl == "CPFL011":
            biz = "山姆零售" if cpxm == "CPXM017" else "零售"
        else:
            biz = _CPFL_TO_BIZ.get(cpfl, "—")
        out[biz] = round(out.get(biz, 0) + debit, 2)
    return out


def compute_expense_ratio(rows, year, month, bp_rev=None, prev_exp=None):
    """费率表（列=业务方 2026-08-06 定稿：本期收入/本期费用/本期费率/上期收入/上期费用/上期费率/提示）。
    提示=环比 pp（百分点差）+相对百分比，超±30%加⚠；另两类异常：有费用没收入/有收入没费用。
    小料出两行：合计 + 非古茗（古茗自提无物流费、收入过大稀释费率——分母剔古茗，分子不变）。"""
    cur_p = f"{int(year)}-{int(month):02d}"
    pm_y, pm_m = (int(year) - 1, 12) if int(month) == 1 else (int(year), int(month) - 1)
    prev_p = f"{pm_y}-{pm_m:02d}"
    exp = {}
    for v in rows or []:
        if not v.get("费用归属") or v.get("未税") is None:
            continue
        biz = v.get("业务线") or "—"
        exp[biz] = round(exp.get(biz, 0) + float(v["未税"]), 2)
    cur = (bp_rev or {}).get(cur_p, {}) or {}
    prv_m = (bp_rev or {}).get(prev_p, {}) or {}
    rev_cur, rev_prev = cur.get("byBiz", {}), prv_m.get("byBiz", {})
    gm_cur, gm_prev = cur.get("guming", 0), prv_m.get("guming", 0)
    prev_exp = prev_exp or {}

    def mk(biz, e, rv, pe, prv, note=""):
        ratio = (e / rv) if rv else None
        pratio = (pe / prv) if (pe is not None and prv) else None
        tips = []
        if biz == "—":
            tips.append("无业务线（研发/设备/入库等），不算费率")
        elif rv is None or rv == 0:
            tips.append("⚠有费用没收入——查业务线是否选错，或 BP 该线本月收入未归集")
        elif e == 0:
            tips.append("⚠有收入没费用——该线本月没计提物流费？")
        elif ratio is not None and pratio is not None:
            pp = (ratio - pratio) * 100
            pct = (ratio - pratio) / pratio if pratio else 0
            arrow = "＋" if pp >= 0 else "－"
            head = "⚠" if abs(pct) > 0.30 else ""
            tips.append(f"{head}环比{arrow}{abs(pp):.2f}pp（{'↑' if pct >= 0 else '↓'}{abs(pct):.0%}）")
        if note:
            tips.append(note)
        return {"业务线": biz, "本期收入": rv, "本期费用": e, "本期费率": ratio,
                "上期收入": prv, "上期费用": pe, "上期费率": pratio, "提示": "；".join(t for t in tips if t)}

    lines = []
    for biz in sorted(set(exp) | set(rev_cur), key=lambda b: -(exp.get(b, 0))):
        e, rv = exp.get(biz, 0), rev_cur.get(biz)
        pe, prv = prev_exp.get(biz), rev_prev.get(biz)
        note = ""
        if biz == "豆蛋制品" and (cur.get("egg_split") or prv_m.get("egg_split")):
            note = "收入按系列(鸡蛋豆腐/双蛋白)从BP植物肉单元拆出"
        if biz == "小料" and gm_cur:
            note = f"含古茗自提收入 {gm_cur:,.0f}（费率被稀释，看下面非古茗行）"
        lines.append(mk(biz, e, rv, pe, prv, note))
        if biz == "小料" and (gm_cur or gm_prev):
            rv2 = None if rv is None else round(rv - gm_cur, 2)
            prv2 = None if prv is None else round(prv - gm_prev, 2)
            lines.append(mk("小料（非古茗）", e, rv2, pe, prv2,
                            "古茗自提不产生物流费——分母剔古茗、费用不变（物流部 Sheet1 同口径）"))
    tot_e = round(sum(exp.values()), 2)
    biz_e = round(sum(v for k, v in exp.items() if k != "—"), 2)
    tot_rv = round(sum(rev_cur.values()), 2)
    tot_rv_ng = round(tot_rv - gm_cur, 2)
    return {"period": cur_p, "prev_period": prev_p, "lines": lines,
            "合计": {"费用": tot_e, "业务线费用": biz_e, "收入": tot_rv,
                    "总费率": (biz_e / tot_rv) if tot_rv else None,
                    "收入非古茗": tot_rv_ng,
                    "总费率非古茗": (biz_e / tot_rv_ng) if tot_rv_ng else None}}


# ---------------- 长表上传（V2.224 核算组定稿流程：物流部上传 24 列长表模板） ----------------
_LF_FEE13 = {"销售出库费用", "成品入库费用", "原料入库费用", "成品仓储费用", "原料仓储费用",
             "成品调拨费用", "原料调拨费用", "出库装卸费用", "成品入库装卸费用", "原料入库装卸费用",
             "研发设备采购", "设备调拨费用", "其它"}
_LF_BIZ = {"植物肉", "鲜食", "零售", "小料", "豆蛋制品", "电商", "山姆零售", "kikiherb", "海外", "—", ""}
_LF_SUBJ = {"深圳星期零", "孝感星期九", "深圳星期九"}


def parse_long_form(data, month, rates=None, fee_lk=None, biz_lk=None, suppliers=None):
    """解析物流部手填的 24 列长表（计提表模板 v2.1「计提明细表」sheet）。
    返回 {rows(活表行,一行一凭证 D-2), new_suppliers(维表没有的商), dirty(不干净行清单), stats}。
    维度列(T-X)有值=人工已定优先采用；空则按映射维表预填。质检口径（核算组定稿"检测计提表是否干净"）：
    主体/简称/费用归属/业务线枚举合法、含税为数、税率可取、维度齐全。"""
    import io as _io
    from openpyxl import load_workbook
    wb = load_workbook(_io.BytesIO(data), data_only=True)
    ws = None
    for nm in wb.sheetnames:
        if "计提明细" in nm:
            ws = wb[nm]
            break
    if ws is None:
        ws = wb[wb.sheetnames[0]]
    # 表头校验（防传错文件——账单/别的表进来给人话报错）
    h = [str(ws.cell(row=3, column=c).value or "").strip() for c in (1, 2, 3, 4)]
    if h[1] != "付款主体" or h[2] != "物流商简称":
        raise ValueError("这不像计提表长表模板（第 3 行表头对不上：应为 账期/付款主体/物流商简称/费用归属）。请用「基础数据」签发的模板填报。")

    sup_by_short = {r["short"]: r for r in (suppliers or [])}
    rows_out, dirty = [], []
    new_sup = {}
    r = 4
    while r <= ws.max_row:
        a = str(ws.cell(row=r, column=1).value or "").strip()
        subj = str(ws.cell(row=r, column=2).value or "").strip()
        short = str(ws.cell(row=r, column=3).value or "").strip()
        gross_v = ws.cell(row=r, column=9).value
        if a == "合计":
            break
        if not (subj or short or gross_v):
            r += 1
            continue
        fee = str(ws.cell(row=r, column=4).value or "").strip()
        biz = str(ws.cell(row=r, column=5).value or "").strip()
        desc = str(ws.cell(row=r, column=6).value or "").strip()
        settle = str(ws.cell(row=r, column=7).value or "月结").strip() or "月结"
        handler = str(ws.cell(row=r, column=8).value or "").strip()
        probs = []
        if subj not in _LF_SUBJ:
            probs.append(f"付款主体「{subj or '(空)'}」不认识（应为 深圳星期零/孝感星期九/深圳星期九）")
        if not short:
            probs.append("物流商简称没填")
        if fee and fee not in _LF_FEE13:
            probs.append(f"费用归属「{fee}」不在 13 类里（请用模板下拉）")
        if not fee:
            probs.append("费用归属没选")
        if biz not in _LF_BIZ:
            probs.append(f"业务线「{biz}」不认识（请用模板下拉）")
        try:
            gross = round(float(str(gross_v).replace(",", "")), 2)
            if gross == 0:
                probs.append("含税金额为 0")
        except (TypeError, ValueError):
            gross = 0.0
            probs.append(f"含税金额「{gross_v}」不是数字")
        sp = sup_by_short.get(short)
        if short and not sp:
            new_sup.setdefault(short, {"简称": short, "行数": 0, "金额": 0.0})
            new_sup[short]["行数"] += 1
            new_sup[short]["金额"] = round(new_sup[short]["金额"] + gross, 2)
        full = (sp or {}).get("full") or short
        channel = (sp or {}).get("channel") or "线下"
        row = build_row(subj if subj in _LF_SUBJ else "", short, full, channel,
                        fee if fee in _LF_FEE13 else "", biz, desc, gross, month,
                        rates=rates, fee_lk=fee_lk, biz_lk=biz_lk, tickets=1,
                        remark=f"长表第{r}行" + (f"·经办{handler}" if handler else ""))
        row["结算类型"] = settle
        # 维度列 T-X：物流部/财务在 Excel 里已定的值优先（活表同规矩：人工优先于映射）
        manual_dims = {"科目": 20, "部门": 21, "费用项目": 22, "产品分类编码": 23, "产品项目": 24}
        overridden = False
        for k, c in manual_dims.items():
            v = str(ws.cell(row=r, column=c).value or "").strip()
            if v and not v.startswith("🖐") and v != row.get(k, ""):
                row[k] = v
                overridden = True
        if overridden:
            row = finalize_row(row, month, rates)
        if row.get("税率来源") == "缺税率" and short and sp:
            probs.append(f"缺税率（{short}×{fee}）——去「基础数据」补")
        if probs:
            dirty.append({"行": r, "物流商": short or "(空)", "金额": gross, "问题": probs})
        rows_out.append(row)
        r += 1

    n_bad = len(dirty)
    stats = {"文件数": 1, "票数": len(rows_out), "含税合计": round(sum(x["含税"] for x in rows_out), 2),
             "明细行数": len(rows_out), "待人工行": sum(1 for x in rows_out if not x["可录入"]),
             "待人工金额": round(sum(x["含税"] for x in rows_out if not x["可录入"]), 2),
             "不干净行": n_bad, "新供应商": len(new_sup)}
    return {"rows": rows_out, "new_suppliers": sorted(new_sup.values(), key=lambda x: -x["金额"]),
            "dirty": dirty, "stats": stats, "clean": n_bad == 0 and not new_sup}


def make_summary(full, month, channel, bizline, descr, sword):
    """摘要=计提+全称+月+渠道+业务线或描述+摘要用语（与模板 N 列公式/真凭证逐字一致——防重幂等键）。"""
    seg = descr or ("" if bizline in ("", "—") else bizline)
    return f"计提{full}{int(month)}月{channel or '线下'}{seg}{sword}"


def build_row(subj, short, full, channel, fee, biz, desc, gross, month,
              rates=None, fee_lk=None, biz_lk=None, tickets=0, remark=""):
    """一行计提明细（=A期 voucher 同构 + 维度预填 + 活表元信息）。费用归属空 → 待人工行。"""
    import db as _db
    row = {"主体": subj, "物流商": short, "公司全名": full, "渠道": channel,
           "费用归属": fee, "业务线": biz or "—", "业务描述": desc, "结算类型": "月结",
           "含税": round(gross, 2), "票数": tickets, "备注": remark,
           "科目": "", "部门": "", "费用项目": "", "摘要用语": "", "产品分类编码": "", "产品项目": "",
           "manual": False, "映射层级": "", "摘要": "", "税率": None, "税率来源": "",
           "未税": None, "税额": None, "可录入": False, "分录": []}
    if not fee:
        return row
    # 维度预填：fee_map 四级取数
    m, tier = _db.resolve_fee_map(fee_lk or {}, fee, subj, row["业务线"])
    if m:
        row.update(科目=m["account"] or "", 部门=m["dept"] or "", 费用项目=m["item"] or "",
                   摘要用语=m["sword"] or "", manual=bool(m["manual"]), 映射层级=tier)
    # 产品维度：科目 6601/6401 才挂（账证维度配置）
    acc4 = (row["科目"] or "")[:4]
    if biz_lk and row["业务线"] in biz_lk and acc4 in ("6601", "6401"):
        cpfl, cpxm = biz_lk[row["业务线"]]
        row["产品分类编码"], row["产品项目"] = cpfl, cpxm
    return finalize_row(row, month, rates)


def finalize_row(row, month, rates=None):
    """按行当前维度算 税率/未税/税额/摘要/分录/可录入。活表改维度后服务端重算也走这里（唯一事实源）。"""
    fee = row.get("费用归属") or ""
    if not fee:
        row.update(税率=None, 税率来源="", 未税=None, 税额=None, 摘要="", 分录=[], 可录入=False)
        return row
    rate, rate_src = la.resolve_rate(rates, row["公司全名"], fee, row.get("税率"))
    ok = rate is not None
    rate = rate or 0.0
    gross = round(float(row["含税"]), 2)
    net = round(gross / (1 + rate), 2)
    tax = round(gross - net, 2)
    row.update(税率=rate, 税率来源=rate_src, 未税=net, 税额=tax)
    row["摘要"] = make_summary(row["公司全名"], month, row.get("渠道"), row.get("业务线"),
                               row.get("业务描述"), row.get("摘要用语"))
    manual_gap = row.get("manual") and (not row.get("科目") or not row.get("部门"))
    row["可录入"] = bool(ok and row.get("科目") and row.get("部门") and row.get("费用项目") and not manual_gap)
    biz_dim = ("/" + row["业务线"]) if row["业务线"] not in ("", "—") else ""
    toc_dim = "/TO C" if row.get("产品项目") else ""
    row["分录"] = [
        {"方向": "借", "科目": row["科目"], "维度": f'{row["部门"]}/{row["费用项目"]}{biz_dim}{toc_dim}',
         "借方": net, "贷方": 0},
        {"方向": "借", "科目": la.ACC_TAX, "维度": row["公司全名"], "借方": tax, "贷方": 0},
        {"方向": "贷", "科目": la.ACC_PAYABLE, "维度": row["公司全名"], "借方": 0, "贷方": gross},
    ]
    return row

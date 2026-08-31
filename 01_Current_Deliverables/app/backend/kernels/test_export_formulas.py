# -*- coding: utf-8 -*-
# [Change Log]
# Date: 2026-08-13 | Author: Claude / c | Version: V2.291
# Description: 导出文件里【公式算出来的数】与【内核算出来的数】必须一致——逐格验。
#   起因：V2.291 起业务方定"灰表原样、汇总用 SUMIFS"，汇总透视各块改成引用原始底表的公式。
#   公式一旦写错（列字母、条件文本、页名、合并后错位），Excel 里显示的就是另一个数，
#   而 openpyxl 只写不算、本机也没有 LibreOffice——**不主动验就等于不知道对不对**。
#   V2.290 刚出过一次同类事故：合并透视时公式没重定位，B149=SUM(B2:B40) 加的是别的块，
#   从 V2.279 起错了十来个版本没人发现（源页单独看是对的，而源页合并后就被删了）。
#   本测试：合成数据 → 建 res → 出 xlsx → 把公式真算一遍 → 与 res 比。
import os
import sys
import unittest
from io import BytesIO

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault("KD_CONF_PATH", os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "conf.ini"))

import openpyxl

from kernels import cost_ledger as cl
from kernels.xlsx_formula_eval import evaluate, evaluate_all
from routers.cost_ledger import _build_cost_ledger_xlsx

CFG = cl.load_config(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                                  "sample_data", "cost_ledger_config.json"))
HEADER = ["物料编码", "物料名称", "存货类别", "物料分组", "规格型号", "库存状态", "批号", "仓库",
          "核算范围编码", "核算范围名称", "基本单位",
          "期初数量", "期初单价", "期初金额", "收入数量", "收入单价", "收入金额",
          "发出数量", "发出单价", "发出金额", "结存数量", "结存单价", "结存金额"]


def _row(code, name, cat, wh, oq, oa, iq, ia, dq, da, eq, ea):
    return [code, name, cat, "", "", "可用", "2026", wh, "HSFW", "孝感星期九", "千克",
            oq, 0, oa, iq, 0, ia, dq, 0, da, eq, 0, ea]


def _book():
    """两个仓、三个类别，且**同一类别跨两个仓**——这样按类别与按仓库不会退化成同一组数，
    列字母/条件写反了才会被抓出来。"""
    rows = [["核算体系:财务会计核算体系"], ["会计期间:2026年第5期"], HEADER,
            _row("A", "甲", "产成品", "成品仓", 100, 1000, 50, 500, 30, 300, 120, 1200),
            _row("B", "乙", "原材料", "原料仓", 10, 90, 20, 210, 5, 50, 25, 250),
            _row("C", "丙", "原材料", "成品仓", 4, 40, 1, 10, 2, 20, 3, 30),
            _row("D", "丁", "包材", "原料仓", 7, 70, 3, 30, 1, 10, 9, 90)]
    cross = cl.parse_cross_report(rows)
    res = cl.build_cost_ledger(cross, CFG)
    res.setdefault("_raw", {})["bydate"] = None
    meta = {"year": 2026, "period": 5, "org": "107", "org_full": "测试主体",
            "updated_at": "2026-08-13 00:00", "updated_by": "ut"}
    wb = openpyxl.load_workbook(BytesIO(_build_cost_ledger_xlsx(res, cross, meta)))
    return wb, res


class TestExportFormulas(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.wb, cls.res = _book()
        cls.ws = cls.wb["汇总透视"] if "汇总透视" in cls.wb.sheetnames else None

    def _blocks(self):
        out = {}
        for row in self.ws.iter_rows(min_col=1, max_col=1):
            v = row[0].value
            if isinstance(v, str) and v.startswith("■"):
                out[v[2:]] = row[0].row
        return out

    def test_every_formula_is_understood_and_computes(self):
        """全簿公式必须**全部**能求值——出现没见过的形态就是新写了没验过的公式。"""
        vals = evaluate_all(self.wb)
        self.assertGreater(len(vals), 0)
        self.assertTrue(all(v is not None for v in vals.values()))

    def test_category_block_matches_kernel(self):
        """按类别块八列 SUMIFS ＝ 内核 pivot_category。"""
        ws, blk = self.ws, self._blocks()
        r = blk["收发存汇总·按类别"] + 3            # 标题 + 来源行 + 表头
        KS = ["oq", "oa", "iq", "ia", "dq", "da", "eq", "ea"]
        seen = 0
        while ws.cell(row=r, column=2).value and ws.cell(row=r, column=1).value != "合计":
            cat = ws.cell(row=r, column=2).value
            a = self.res["pivot_category"][cat]
            for j, k in enumerate(KS):
                got = evaluate(self.wb, ws, ws.cell(row=r, column=3 + j).value)
                self.assertAlmostEqual(got, a[k], places=2, msg="%s %s" % (cat, k))
            seen += 1
            r += 1
        self.assertGreaterEqual(seen, 3)          # 三个类别都验到了，没有提前退出

    def test_warehouse_block_matches_kernel(self):
        """按仓库块四列 SUMIFS ＝ 内核 pivot_warehouse；组小计与总计不重复计。"""
        ws, blk = self.ws, self._blocks()
        whp = self.res["pivot_warehouse"]
        r, seen = blk["收发存汇总·按仓库"] + 3, 0
        while ws.cell(row=r, column=1).value and ws.cell(row=r, column=1).value != "总计":
            nm = str(ws.cell(row=r, column=1).value).strip()
            if nm in whp:
                for j, k in enumerate(["oa", "ia", "da", "ea"]):
                    got = evaluate(self.wb, ws, ws.cell(row=r, column=3 + j).value)
                    self.assertAlmostEqual(got, whp[nm][k], places=2, msg="%s %s" % (nm, k))
                seen += 1
            r += 1
        self.assertEqual(seen, len(whp))
        # 总计行：必须等于各仓之和（错写成"连小计一起加"会翻倍）
        for j, k in enumerate(["oa", "ia", "da", "ea"]):
            got = evaluate(self.wb, ws, ws.cell(row=r, column=3 + j).value)
            self.assertAlmostEqual(got, sum(v[k] for v in whp.values()), places=2)

    def test_merged_formula_is_relocated(self):
        """合并进汇总透视后，`=SUM(...)` 的引用范围必须**落在本块之内**——V2.290 事故的护栏
        （那次是 B149=SUM(B2:B40)，跨过块边界加到别的块去了）。
        ⚠不能断言"只能往上加"：组小计行在**组首**，加的是它下面那几行，往下是对的。"""
        import re
        starts = sorted(self._blocks().values())
        for row in self.ws.iter_rows():
            for c in row:
                if not (isinstance(c.value, str) and c.value.startswith("=SUM(")):
                    continue
                m = re.match(r"=SUM\([A-Z]+(\d+):[A-Z]+(\d+)\)$", c.value)
                if not m:
                    continue
                lo, hi = int(m.group(1)), int(m.group(2))
                own = max((s for s in starts if s <= c.row), default=0)
                nxt = min([s for s in starts if s > c.row] or [self.ws.max_row + 1])
                self.assertGreater(lo, own, "%s 的 %s 引到了本块之前" % (c.coordinate, c.value))
                self.assertLess(hi, nxt, "%s 的 %s 引到了下一块" % (c.coordinate, c.value))
                self.assertNotIn(c.row, range(lo, hi + 1), "%s 的 %s 把自己也加进去了" % (c.coordinate, c.value))

class TestPnlDetailTie(unittest.TestCase):
    """损益归集 ↔ 原始·货损与处置明细 必须勾稽得上（V2.293 事故的护栏）。

    事故：明细页按【凭证分录】循环，一条分录就把该单据的物料行整套写一遍。
    一张单据在总账里常拆成多条分录（🧪 2026-7 的 QTCK011302 拆成 13 条），
    于是同一批 122 行被写了 13 遍 —— 明细页 1,586 行 / 4,218,939.14，
    而损益归集 324,529.82，**两页差 13 倍**。3 月每单只对应一条分录，碰巧一直是对的，
    所以本测试**必须造"一单多分录"**，否则测不出来。"""

    BILL = "QTCK000001"

    def _res(self):
        rows = [{"billno": self.BILL, "btype": "标准其他出库单", "date": "2026-07-01",
                 "voucher": "记570", "code": "M%03d" % i, "name": "料%d" % i, "spec": "",
                 "cat": "原材料", "grp": "", "unit": "kg", "wh": "W", "batch": "",
                 "qty": 1.0, "price": 100.0, "amount": 100.0} for i in range(5)]
        # 一张单据 → 3 条凭证分录（合计 500 ＝ 5 行物料 × 100）
        loss = [{"billno": self.BILL, "doctype": "报废出库无原料的其他出库单",
                 "amount": a, "date": "2026-07-01", "voucher": "记570"} for a in (200.0, 200.0, 100.0)]
        return {"ties": {"self_balance": {"pass": True}}, "credible": True,
                "anomalies": {"counts": {}, "items": []},
                "pivot_category": {"合计": {k: 0 for k in
                                            ("oq", "oa", "iq", "ia", "dq", "da", "eq", "ea")}},
                "pivot_warehouse": {}, "pivot_wh_type": {"by_type": {}, "missing_attr": []},
                "pnl": {"loss": {"by_cat": {"报废出库无原料的其他出库单": 500.0}, "total": 500.0},
                        "disposal": {"total": 0.0}},
                "_raw": {"pnl_detail": {"loss": loss, "disposal": [], "rows": rows}}}

    def test_sheet_survives_when_only_other_bucket(self):
        """只剩「其他」那一档时，明细页**仍然要出**。

        事故（V2.311）：V2.307 把福利领用/捐赠从货损里拆成第三档「其他」，
        出页条件却还是 `if loss or disposal` —— 🧪 101 深圳星期零 6 月货损/处置皆 0、
        只有 5 条其他（26,003.58），于是**整页蒸发**，损益归集③有小计却没有底表。
        业务方发现的方式是「星期零的货损那个，你删除了？」——没有任何测试挡住它。"""
        res = self._res()
        oth = res["_raw"]["pnl_detail"].pop("loss")
        res["_raw"]["pnl_detail"]["loss"] = []
        res["_raw"]["pnl_detail"]["other"] = oth
        res["pnl"] = {"loss": {"by_cat": {}, "total": 0.0}, "disposal": {"total": 0.0},
                      "other": {"total": 500.0}}
        meta = {"year": 2026, "period": 6, "org": "101", "org_full": "测试主体",
                "updated_at": "2026-08-18 00:00", "updated_by": "ut"}
        wb = openpyxl.load_workbook(BytesIO(_build_cost_ledger_xlsx(res, [], meta)))
        self.assertIn("原始·货损与处置明细", wb.sheetnames,
                      "只有「其他」这一档时明细页整页丢了——那些钱仍是从存货流向损益的")
        ws = wb["原始·货损与处置明细"]
        col = next(c.column for row in ws.iter_rows(max_row=12) for c in row if c.value == "归属")
        n = sum(1 for r in ws.iter_rows(min_col=col, max_col=col)
                if str(r[0].value or "") == "其他存货出库（非货损）")
        self.assertEqual(n, 5, "「其他」的物料行没写进明细页")

    def setUp(self):
        meta = {"year": 2026, "period": 7, "org": "107", "org_full": "测试主体",
                "updated_at": "2026-08-13 00:00", "updated_by": "ut"}
        self.wb = openpyxl.load_workbook(BytesIO(_build_cost_ledger_xlsx(self._res(), [], meta)))

    def test_material_rows_written_once_per_bill(self):
        """一单三分录 → 物料行只能写 5 行，不是 15 行。"""
        ws = self.wb["原始·货损与处置明细"]
        # ⚠按**表头名**定位「归属」列，别写死 A 列：V2.309 列序改成业务方底稿那套之后
        # 归属挪到了 M 列，写死列号的断言会变成"永远数到 0 行"——**假绿更糟**。
        col = next(c.column for row in ws.iter_rows(max_row=12) for c in row if c.value == "归属")
        n = sum(1 for r in ws.iter_rows(min_col=col, max_col=col)
                if str(r[0].value or "") == "管理费用·货损")
        self.assertEqual(n, 5, "同一单据的物料行被按分录重复写了")

    def test_pnl_page_ties_to_detail(self):
        """损益归集的「差异」行必须算出 0。"""
        ws = self.wb["勾稽与归集"] if "勾稽与归集" in self.wb.sheetnames else self.wb["损益归集"]
        hit = 0
        for row in ws.iter_rows(min_col=1, max_col=2):
            if "差异（应为" in str(row[0].value or ""):
                self.assertAlmostEqual(evaluate(self.wb, ws, row[1].value), 0.0, places=2)
                hit += 1
        self.assertGreaterEqual(hit, 1, "没找到勾稽差异行——两页又变成各写各的了")



if __name__ == "__main__":
    unittest.main(verbosity=2)

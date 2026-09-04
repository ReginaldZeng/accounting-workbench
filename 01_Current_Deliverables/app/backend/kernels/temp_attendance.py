# -*- coding: utf-8 -*-
# [Change Log]
# Date: 2026-08-18 | Author: Claude / c | Version: V2.318
# Description: 【临时工考勤】内核——把「人力上报工时」与「打卡记录」按公司口径逐日重算比对。
#              口径来源三方互证（见 03_Source_Materials/20260810_临时工考勤/现状分析与口径探查 v1.0）：
#                ① 人力 2026-08-10 答复：下班打卡−上班打卡，半小时取整，白班扣1小时/夜班扣0.5小时
#                ② 成本会计 6 月复核底表公式：FLOOR(...,0.5) 与 IF(...,1,0.5)
#                ③ ⚠ 曾以「保洁岗 53 个出勤日 53/53 吻合」佐证，已于 2026-08-18 撤回：
#                   那 53 天的打卡跨度全在 10.10–10.43，向下取整后同落 10.0 档，
#                   「按口径算」与「固定 9 小时排班」结果相同，这份数据根本不具判别力。
#              **定位：偏离监控，不是重算工资**（2026-08-18 业务定案）。
#              实证表明上报工时不是从打卡算出来的——同一个打卡跨度档内，上报工时能出现 3–4 种取值
#              （如跨度 12.5 档 48 天里，上报有 10.0/10.5/11.0/11.5 四种），它更像排班班次时长。
#              故工具不产出「应付多少」，只指认偏离：
#                ⚠ 超出弹性的【多记】＝ 唯一必须查的异常（公司多付钱）
#                △ 【少记】＝ 财务可接受，只作提示，不要求补付
#              另附结算风险核对：同名跨派遣方、同名重复行、结算归属与打卡部门不符。
#              本文件只做算法，不碰 HTTP/DB；路由在 routers/temp_attendance.py。

import datetime
import io
import math
import re

try:
    from openpyxl import load_workbook
except Exception:                                    # 环境缺 openpyxl 时不炸整站，调用处报错
    load_workbook = None


# ==================== 口径参数 ====================
# 全部可由前端覆盖。夜班切班窗口尤其重要：窗口定义错一点，异常清单就会假报一大片
# （实测：用「次日06:00–11:00 首卡」收班，6 月异常多记假报 117 人日；改成「次日11:30 前末卡」后收敛到 3 人日）。
DEFAULT_PARAMS = {
    "round_step": 0.5,        # 取整粒度（小时）
    "round_mode": "floor",    # floor=向下取整（实证口径）｜round=四舍五入
    "day_break": 1.0,         # 白班扣减（午饭）
    "night_break": 0.5,       # 夜班扣减（夜宵）
    "tolerance": 0.5,         # 多记弹性（小时/天）：≤ 此值视为正常波动，> 此值判异常
    "night_start_from": 16 * 60,      # 夜班上班窗口起点（当日 16:00 之后的首卡＝上班）
    "night_end_by": 11 * 60 + 30,     # 夜班下班窗口终点（次日 11:30 之前的末卡＝下班）
    "adj_ratio_cap": 0.20,    # 奖/罚/补贴单笔占该人当月工资的提示线（实测 1 月桂丽 75/135＝55.6%）
    # ⚠ 上报工时按什么口径记——这一条决定整张判定表问的是什么问题：
    #   shift＝**排班班次时长**（2026-06 全量 448 人实证：打卡跨度 12.5h 和 13.0h 的日子，
    #          上报**同为 11.0h**；夜班 452/542 天报 11.0、白班 590/900 天报 11.0。
    #          11h ＝ 标准班 08:00–20:00 或 20:00–08:00 扣 1 小时休息）
    #          → 判定问「**打卡撑不撑得起上报的班次**」，撑得住就不是问题；
    #   punch＝上报应当等于打卡重算（人力 2026-08-10 口头答复的口径，实证不成立，保留可切回）
    #          → 判定问「**差了多少**」，少记/多记分档。
    # 选 punch 会让 6 月 2,841 人日里 1,410 条报「少记」——那不是人力少记，是两边算法不同。
    "report_basis": "shift",
}

# ==================== 单价：合同价是唯一基准 ====================
# 为什么不能是一个全局单价：员工工资、管理费、白班夜班**三个维度都按派遣方分档**，
# 成本会计的复核底表就是栽在这里——全员写死 17 元员工 + 4 元派遣费，既不分派遣方也不认保洁的 15 元，
# 7 月因此把应付算成 62,727 元，比按合同单价的 54,840 元多了 7,887 元（工时反而是对的）。
# 结构：派遣方 → 岗位 → {"day": (员工工资, 管理费), "night": (员工工资, 管理费)}；night=None 表示该档无夜班。
# **单价是「该派遣方的该岗位」，不是全局岗位**——规则原文写的就是「锦绣保洁」，不是「保洁」。
# 早先把保洁做成跨派遣方的全局覆盖是错的：万一华顺也有保洁，会被套上锦绣的 15 元。
#
# 本工具里单价只有三个来源，各有各的用处，**不混用**（V2.346 定案）：
#   合同价   ＝ 成本会计在「合同价登记表」登记的行（按行带生效期）。**唯一的核对基准。**
#              没登记的格子就是「缺档」——应付算不出来，结论是「待核」，不是正常也不是异常。
#   表上单价 ＝ 结算表里人力给每个人实际套用的单价。它是**被核对的一方**；
#              同时也是偏离计价的依据——公司就是按这个价付的，多记 1 小时就多付这么多。
#   表头解析 ＝ 汇总表表头那段计价规则文字解析出来的表。人力自己写的，**只作参考展示**。
POST_DEFAULT = "普工"          # 岗位列为空＝普工（绝大多数人）

# 单价比对的容差：**逐格（contract_vs_actual）与逐人（rate_gaps）必须用同一个**，
# 否则会出现「②说这一格不符，④一个人也标不出」——两处各写一个默认值就是这么来的（V2.349 审出）。
RATE_TOL = 0.005

# ⚠ 下面这张表**不参与任何判定、任何计价**。它是 2026 年从人力结算数据里观察到的计价档位，
# 只留两个用处：单元测试的夹具、给没见过这套数据的人一个量级参考。
# V2.342–V2.345 曾拿它给合同价兜底，后果是成本会计一行没登记也满屏「一致」——
# 它本来就是从人力数据反推的，拿它当基准＝人力跟自己比。这个坑别再踩第三次。
RATE_TABLE_OBSERVED = {
    "华顺": {POST_DEFAULT: {"day": (16.5, 2.5), "night": (19.0, 3.0)}},
    "恒祺": {POST_DEFAULT: {"day": (16.5, 2.5), "night": (19.0, 3.0)}},
    "锦绣": {POST_DEFAULT: {"day": (16.5, 2.5), "night": (19.0, 3.0)},
             "保洁": {"day": (15.0, 0.0), "night": None}},      # 15 元/小时，无管理费、无夜班
    "成达": {POST_DEFAULT: {"day": (17.0, 2.0), "night": (19.0, 3.0)}},
    "广才": {POST_DEFAULT: {"day": (17.0, 2.0), "night": (19.0, 3.0)}},
    "天幕": {POST_DEFAULT: {"day": (17.0, 2.0), "night": (19.0, 3.0)}},
    "天募": {POST_DEFAULT: {"day": (17.0, 2.0), "night": (19.0, 3.0)}},   # 汇总表里「天幕/天募」两种写法都出现过
    "鑫路达": {POST_DEFAULT: {"day": (17.0, 2.0), "night": (19.0, 3.0)}},
}

_RULE_DAYNIGHT = re.compile(
    r"(?P<names>[一-龥、，,\s]+?)白班员工工资\s*(?P<dw>[\d.]+)\s*元\s*/\s*[Hh小时]+"
    r"[，,、；;\s]*管理费\s*(?P<dm>[\d.]+)\s*元\s*/\s*[Hh小时]+"
    r"[；;，,\s]*夜班员工工资\s*(?P<nw>[\d.]+)\s*元\s*/\s*[Hh小时]+"
    r"[；;，,\s]*管理费\s*(?P<nm>[\d.]+)\s*元\s*/\s*[Hh小时]+")
# 表头首行是「对外总价」：白班19元/小时·人、夜班22元/小时·人。
# 它不是某一家的单价，而是「员工工资＋管理费」的合计口径——正好拿来交叉校验分档表有没有被改错。
_RULE_HEADLINE = re.compile(
    r"白班\s*(?P<day>[\d.]+)\s*元?\s*/\s*小时.*?夜班\D*(?P<night>[\d.]+)\s*元?\s*/\s*小时")
_RULE_KIND = re.compile(
    r"(?P<agency>[一-龥]{2,4}?)(?P<kind>保洁|保安|司机)\s*[：:]\s*"
    r"员工工资\s*(?P<w>[\d.]+)\s*元\s*/\s*[Hh小时]+(?P<rest>.*)")


def parse_rate_rules(text):
    """把汇总表表头第 2 行的规则原文解析成单价表（派遣方 → 岗位 → 档）。
    原文形如：
      「华顺、恒祺、锦绣白班员工工资16.5元/H，管理费2.5元/H；夜班员工工资19元/H；管理费3元/H.」→ 三家的普工档
      「锦绣保洁：员工工资15元/小时，无管理费无夜班」→ **锦绣名下的保洁岗**，不是全局保洁
    返回 (table, notes, headline)。notes 记下没看懂的行——宁可标出来交人工，也不静默按默认值算钱。"""
    table, notes, headline = {}, [], {}
    for raw in str(text or "").splitlines():
        line = raw.strip().rstrip("。.；;")
        if not line:
            continue
        m = _RULE_DAYNIGHT.search(line)
        if not m:
            mh = _RULE_HEADLINE.search(line)
            if mh and "员工工资" not in line:
                headline = {"day": float(mh.group("day")), "night": float(mh.group("night"))}
                continue
        if m:
            for n in [x for x in re.split(r"[、，,\s]+", m.group("names").strip()) if x]:
                table.setdefault(n, {})[POST_DEFAULT] = {
                    "day": (float(m.group("dw")), float(m.group("dm"))),
                    "night": (float(m.group("nw")), float(m.group("nm")))}
            continue
        m = _RULE_KIND.search(line)
        if m:
            rest, agency, post = m.group("rest") or "", m.group("agency"), m.group("kind")
            mgmt = 0.0 if "无管理费" in rest else None
            if mgmt is None:
                notes.append(f"「{line}」没写管理费单价，已按 0 处理，请确认")
                mgmt = 0.0
            night = None
            if "无夜班" not in rest:
                notes.append(f"「{line}」没说夜班怎么算，该岗位夜班单价留空")
            table.setdefault(agency, {})[post] = {"day": (float(m.group("w")), mgmt), "night": night}
            continue
        if re.search(r"\d+\s*元|\d+\s*/\s*小时", line):     # 像是价钱但没解析出来
            notes.append(f"这一行没看懂，未纳入单价表：「{line}」")
    return table, notes, headline


def _band(v):
    """把前端/存档传来的一档规范成 {"day":(w,m)|None, "night":(w,m)|None}。
    白班缺失和夜班缺失**一样是 None**——早先把白班补成 (0,0)，结果只登记夜班价的行会被当成
    「白班登记了 0 元」：对比页报「不符 合同 0+0」、应付算成 0，而不是老实说「没有白班价」。"""
    d = v.get("day")
    n = v.get("night")
    return {"day": None if d in (None, "none", "") else tuple(d),
            "night": None if n in (None, "none", "") else tuple(n)}


def _as_nested(x):
    """把外来的单价表规范成 {派遣方: {岗位: 档}}。要认三种形态，因为它们都真实出现过：

      ① 新结构      {派遣方: {岗位: 档}}
      ② 页面整包     {"table": {...}, "默认岗位": "普工", "解析自表头": true, ...}
                     —— 前端把整个 state 发了过来。非字典的字段（如 "默认岗位": "普工"）
                     若当成派遣方去遍历，会炸 'str' object has no attribute 'items'
                     （2026-08-18 服务器实测复现：一上传就「解析失败」）。
      ③ 旧存档      {"agencies": {派遣方:档}, "kinds": {岗位:档}}（岗位是跨派遣方的全局覆盖）
                     旧的 kinds 无法知道属于哪一家，一律并到「锦绣」——历史上只有锦绣有保洁，
                     且规则原文写的就是「锦绣保洁」。并完在页面原样显示，人能看见能改。

    另：任何一层遇到非字典值一律跳过，不再抛异常。单价表是人手填的，宁可少收一格也不能整页打不开。"""
    if not isinstance(x, dict) or not x:
        return {}
    if isinstance(x.get("table"), dict):          # ② 页面整包 → 只取 table
        x = x["table"]
    if "agencies" in x or "kinds" in x:           # ③ 旧存档
        out = {}
        for a, v in (x.get("agencies") or {}).items():
            if isinstance(v, dict):
                out[a] = {POST_DEFAULT: _band(v)}
        for post, v in (x.get("kinds") or {}).items():
            if isinstance(v, dict):
                out.setdefault("锦绣", {})[post] = _band(v)
        return out
    out = {}                                       # ① 新结构
    for a, posts in x.items():
        if not isinstance(posts, dict):            # 「默认岗位」这类标量字段，跳过
            continue
        cell = {}
        for post, b in posts.items():
            if isinstance(b, dict):
                cell[post] = _band(b)
        if cell:
            out[a] = cell
    return out


def contract_only(saved):
    """合同价表 ＝ **成本会计登记了什么就是什么**，登记表之外一律不补。

    ⚠ 不拿任何表兜底。V2.345 之前这里以 RATE_TABLE_OBSERVED 打底，
    后果是成本会计一行没登记也满屏「一致」——那张表本来就是从人力数据反推的，
    拿它当基准＝人力跟自己比（跟"不能拿表头当合同价"是同一个坑，只是换了扇门）。

    没登记就该老老实实报「合同缺档」。
    """
    return _as_nested(saved)


# 判定档位（前缀即分档，前端按前缀上色）
J_OK = "✓ 与口径一致"
J_UNDER = "△ 少记"
J_OVER_IN = "○ 多记"
J_OVER_OUT = "⚠ 多记"
J_NO_PUNCH = "⚠ 记了工时但无打卡"
J_NO_HOUR = "⚠ 有打卡但未记工时"
J_THIN = "△ 仅1次打卡且未记工时,疑似无效"
# 白夜混合：**shift 口径下已按切班窗口逐日切开、正常判档**（V2.369，规则见 compute_shifts）。
# 这一档只在 punch 口径下还用得上——那个口径要精确到小时的少记/多记，混合日的重算值撑不起那种精度。
J_MIXED = "◇ 白夜混合，本口径下不逐日判，待人工"
# 有打卡、当天却没算临时工工时。**这多半不是漏记**：打卡表是全厂的门禁数据
# （6 月 710 人 vs 结算表 448 人），同一个人有很多天在别的名目下上班。
# 早先归进「待查」并标红，全量数据一上就淹没真问题（6 月 1,183 条待查里它占 1,011 条）。
J_UNBILLED = "◇ 有打卡·当天未计临时工工时"
J_BACKED = "✓ 打卡撑得住上报"

_PERIOD_CN = re.compile(r"(20\d{2})\s*年\s*(\d{1,2})\s*月")
_PERIOD_ISO = re.compile(r"(20\d{2})[-/.](\d{1,2})[-/.](\d{1,2})")


def _period_from(text, iso=False):
    """从标题里取年月。汇总表标题写「2026年7月…」，打卡表写「统计日期：2026-07-01 至 …」。
    两边都取出来是为了对一下——成本会计那份 7 月复核版，汇总表标题还写着「2026年5月」（从旧文件带过来没改）。"""
    t = str(text or "")
    m = (_PERIOD_ISO if iso else _PERIOD_CN).search(t)
    if not m:
        m = (_PERIOD_CN if iso else _PERIOD_ISO).search(t)
    if not m:
        return ""
    return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}"


_NAME_TAIL = re.compile(r"[（(].*?[)）]|[A-Za-z0-9]+")
_TIME_RE = re.compile(r"(次日)?(\d{1,2}):(\d{2})")


_PAREN_ONLY = re.compile(r"[（(][^）)]*[)）]")


def same_person_key(s):
    """判「是不是同一个人」的键：**只去掉括号备注，保留字母数字后缀**。

    两头都不能错：
      · 夏菊红 / 夏菊红（离职）  → 同一个人，人力只是在其中一行标了「离职」。
        用原始姓名当键的话并不上，同一批打卡会摆两遍（使用者实测 12 行里 6 行是重复的）。
      · 张博G / 张博J、黄亚军0415 / 黄亚军7327 → **不同的人**，公司就是用这种后缀区分同名者。
        用 norm_name（连后缀一起抹掉）当键的话会把两个人并成一个，那是更严重的错。
    """
    return _PAREN_ONLY.sub("", str(s or "")).strip()


def norm_name(s):
    """姓名归一：去掉「（离职）」「（驻场）」这类括号后缀和「黄亚军0415」这类数字/字母尾巴。
    两张表的同一个人写法常不一致，不归一会大面积匹配不上（7 月 40 人里有 4 个）。"""
    return _NAME_TAIL.sub("", str(s or "")).strip()


def parse_punch_cell(cell):
    """一个打卡格 → 当日打卡时刻列表（分钟）。「次日07:52」加 1440。去重并升序。"""
    if cell in (None, ""):
        return []
    out = set()
    for m in _TIME_RE.finditer(str(cell)):
        out.add(int(m.group(2)) * 60 + int(m.group(3)) + (1440 if m.group(1) else 0))
    return sorted(out)


def round_step(x, params):
    """按口径取整。floor＝向下（实证口径）；round＝四舍五入（备选，参数可切）。"""
    step = float(params.get("round_step") or 0.5)
    if step <= 0:
        return x
    if params.get("round_mode") == "round":
        return round(x / step) * step
    return math.floor(x / step + 1e-9) * step


def fmt_hm(minutes):
    """分钟 → HH:MM（跨零点显示成 次日HH:MM，跟打卡表原样一致，人一眼能对上）。"""
    if minutes is None:
        return ""
    nxt, m = divmod(int(minutes), 1440)
    return ("次日" if nxt else "") + f"{m // 60:02d}:{m % 60:02d}"


# ==================== 解析：打卡表 ====================
# 跨月边界两天在打卡表里是「上月末 / 次月初」两列。读进 rec["bnd"]（不进 days），
# 逐日判定一律不碰，只有导出的⑧原始表用。哨兵键放在 1..31 之外。
_BND_PREV, _BND_NEXT = 0, 32


def parse_punch(data):
    """打卡时刻表（考勤系统导出）→ {归一姓名: {"raw": 原名, "组": 考勤组, "部门": 部门, "days": {日: [分钟...]}}}
    识别方式：找到含「姓名」的表头行，其下一行是日期行（1..31 或「六」「日」）。
    同名多行单独记 dup，交调用方提示人工指认——静默合并会串行（6 月底表里丁菊华/洪杰各有 2 行）。"""
    if load_workbook is None:
        raise RuntimeError("缺少 openpyxl，无法解析打卡表")
    wb = load_workbook(io.BytesIO(data), data_only=True)
    ws = wb.worksheets[0]
    hdr = _find_header_row(ws, must=("姓名",), scan=12)
    if not hdr:
        raise ValueError("打卡表里找不到「姓名」表头行，请确认上传的是考勤系统导出的打卡时刻表")
    name_col = _col_of(ws, hdr, "姓名")
    grp_col = _col_of(ws, hdr, "考勤组")
    dept_col = _col_of(ws, hdr, "部门")
    # 可选的身份列：钉钉取数生成的打卡表会带「手机尾号」。有它，同名的人就能硬判是不是一个人
    # ——不必再靠「打卡行对不对得齐」猜（张博G / 张博J 就是靠这个才分得开）。人力手工导的表没有这列，不影响。
    id_col = _col_of(ws, hdr, "手机尾号") or _col_of(ws, hdr, "工号")
    day_cols = _day_columns(ws, hdr + 1,
                            start_after=max(c for c in (name_col, grp_col, dept_col, id_col) if c))
    if not day_cols:
        raise ValueError("打卡表里找不到 1–31 的日期列")
    # 边界列（钉钉整月取数会多两列）：预先定位，逐行读进 bnd
    bnd_cols = {}
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(hdr + 1, c).value or "").strip()
        if h == "上月末":
            bnd_cols[c] = _BND_PREV
        elif h == "次月初":
            bnd_cols[c] = _BND_NEXT
    by_raw, by_key = {}, {}
    for r in range(hdr + 2, ws.max_row + 1):
        raw = ws.cell(r, name_col).value
        if raw in (None, ""):
            continue
        days = {}
        for d, c in day_cols.items():
            ts = parse_punch_cell(ws.cell(r, c).value)
            if ts:
                days[d] = ts
        bnd = {}
        for c, k in bnd_cols.items():
            bt = parse_punch_cell(ws.cell(r, c).value)
            if bt:
                bnd[k] = bt
        rec = {"raw": str(raw).strip(), "row": r, "组": _txt(ws, r, grp_col),
               "部门": (_txt(ws, r, dept_col).splitlines() or [""])[0], "days": days,
               "bnd": bnd, "标识": _txt(ws, r, id_col) if id_col else ""}
        by_raw.setdefault(rec["raw"], []).append(rec)
        by_key.setdefault(norm_name(raw), []).append(rec)
    return {"by_raw": by_raw, "by_key": by_key,
            "period": _period_from(ws.cell(1, 1).value, iso=True),
            "title": str(ws.cell(1, 1).value or "").strip(),
            "dup_raw": sorted(k for k, v in by_raw.items() if len(v) > 1),
            "dup_key": sorted(k for k, v in by_key.items() if len(v) > 1)}


def match_punch(punch, name):
    """结算表姓名 → 打卡行。三段式，**歧义时不猜**：
       ① 原名完全一致且唯一 → 用它（「黄亚军0415」在打卡表里原样存在时直接命中）
       ② 去尾归一后唯一 → 用它（覆盖「鲁保军」↔「鲁保军（离职）」这类）
       ③ 归一后撞上多个人 → 返回 None + 歧义候选，交人工指认。
    归一是把不同写法拉到一起，但也会把「张博」和「张博G（离职）」拉成一个——那是两个人，猜错就是算错工资。"""
    raw = str(name or "").strip()
    hit = punch["by_raw"].get(raw)
    if hit and len(hit) == 1:
        return hit[0], None
    cand = punch["by_key"].get(norm_name(raw)) or []
    if len(cand) == 1:
        return cand[0], None
    if not cand:
        return None, None
    return None, [c["raw"] for c in cand]


# ==================== 解析：人力上报汇总表 ====================
_PAY_LABELS = {
    "白班工资": ("工时工资", "白班工资"), "夜班工资": ("工时工资", "夜班工资"),
    "白班工资单价": ("工时工资", "白班单价"), "夜班工资单价": ("工时工资", "夜班单价"),
    "白班管理费": ("管理费", "白班管理费"), "夜班管理费": ("管理费", "夜班管理费"),
    "白班管理费单价": ("管理费", "白班单价"), "夜班管理费单价": ("管理费", "夜班单价"),
}


def _pay_cols(ws, hdr):
    """定位结算表右侧的金额区。**这些列是人力自己填的应付金额，不是我们算的**——
    有了它们，复核结论才是「表上写的 vs 按合同价应付」，而不是我们另立一套账。

    列头是两行：hdr 行是大类（工时工资 / 蒸练补贴 / 奖 / 罚 / 员工工资 / 管理费 / 合计，跨列合并），
    hdr+1 行是小类（白班单价 / 白班工资 / 夜班单价 / 夜班工资 …）。
    ⚠「白班单价」在两个大类下各出现一次（工资一次、管理费一次），只认小类必然张冠李戴，
    所以要拿「本列往左最近的一个大类标题」把它们区分开。

    列缺失一律返回 None，下游降级成「只重算、不比对」——不是每个月的表都长一样。"""
    def _t(r, c):
        return str(ws.cell(r, c).value or "").replace("\n", "").strip()

    grp, cur = {}, ""
    for c in range(1, ws.max_column + 1):
        t = _t(hdr, c)
        if t:
            cur = t
        grp[c] = cur
    out = {}
    for c in range(1, ws.max_column + 1):
        sub, top = _t(hdr + 1, c), _t(hdr, c)
        for key, (want_top, want_sub) in _PAY_LABELS.items():
            if sub == want_sub and grp.get(c) == want_top and key not in out:
                out[key] = c
        for key, label in (("补贴", "蒸练补贴"), ("奖", "奖"), ("罚", "罚"),
                           ("员工工资", "员工工资"), ("合计", "合计")):
            if top == label and key not in out:
                out[key] = c
    return out



def parse_summary(data, sheet=None):
    """临时工劳务明细汇总表 →
       {"people":[{no,name,key,dept,agency,kind,days:{日:工时},白班,夜班,总工时}...], "rules": 表头规则原文}
    两种版式都认：全量表（部门/归属/性质）与按派遣方拆分页（部门/归属/备注）。"""
    if load_workbook is None:
        raise RuntimeError("缺少 openpyxl，无法解析汇总表")
    wb = load_workbook(io.BytesIO(data), data_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else _pick_summary_sheet(wb)
    hdr = _find_header_row(ws, must=("姓名", "归属"), scan=12)
    if not hdr:
        raise ValueError("汇总表里找不到含「姓名」「归属」的表头行")
    c_no = _col_of(ws, hdr, "序号")
    c_name = _col_of(ws, hdr, "姓名")
    c_dept = _col_of(ws, hdr, "部门")
    c_agency = _col_of(ws, hdr, "归属")
    c_kind = _col_of(ws, hdr, "性质") or _col_of(ws, hdr, "备注")
    c_day_h = _col_of(ws, hdr, "白班", contains=True)
    c_night_h = _col_of(ws, hdr, "夜班", contains=True)
    day_cols = _day_columns(ws, hdr + 1, start_after=max(x for x in (c_no, c_name, c_dept, c_agency, c_kind) if x))
    pay_cols = _pay_cols(ws, hdr)
    if not day_cols:
        raise ValueError("汇总表里找不到 1–31 的日期列")
    rules = ""
    for r in range(1, hdr):
        v = ws.cell(r, 1).value
        if v and "小时" in str(v):
            rules = str(v).strip()
            break
    out = []
    for r in range(hdr + 2, ws.max_row + 1):
        raw = ws.cell(r, c_name).value if c_name else None
        if raw in (None, ""):
            continue
        days = {}
        for d, c in day_cols.items():
            v = ws.cell(r, c).value
            if isinstance(v, (int, float)) and v:
                days[d] = float(v)
        out.append({
            "no": ws.cell(r, c_no).value if c_no else len(out) + 1,
            "name": str(raw).strip(), "key": norm_name(raw),
            "dept": _txt(ws, r, c_dept), "agency": _txt(ws, r, c_agency), "kind": _txt(ws, r, c_kind),
            "days": days,
            "白班": _num(ws, r, c_day_h), "夜班": _num(ws, r, c_night_h),
            "总工时": round(sum(days.values()), 2),
            # 表上金额原样带出，一个字都不改。「罚」在表里就是负数，直接加，别再减一次
            "表上": {k: _num(ws, r, c) for k, c in pay_cols.items()} if pay_cols else {},
        })
    # 表头规则栏是人力自己写的计价说明：解析出来只作参考展示（rates.表头解析），**不参与任何计算**——合同价只认成本会计登记表
    rt, rn, rh = parse_rate_rules(rules)
    period = _period_from(ws.cell(1, 1).value)
    return {"people": out, "rules": rules, "sheet": ws.title, "sheets": wb.sheetnames,
            "rate_table": rt, "rate_notes": rn, "rate_headline": rh,
            "period": period, "title": str(ws.cell(1, 1).value or "").strip()}


def _pick_summary_sheet(wb):
    """优先取「明细」页（拆分版主表）；没有就取第一个含「姓名」的页。"""
    for ws in wb.worksheets:
        if "明细" in ws.title:
            return ws
    for ws in wb.worksheets:
        if _find_header_row(ws, must=("姓名",), scan=12):
            return ws
    return wb.worksheets[0]


def _find_header_row(ws, must=("姓名",), scan=12):
    for r in range(1, min(ws.max_row, scan) + 1):
        vals = {str(ws.cell(r, c).value or "").replace("\n", "") for c in range(1, min(ws.max_column, 60) + 1)}
        if all(any(m == v or m in v for v in vals) for m in must):
            return r
    return None


def _col_of(ws, row, label, contains=False):
    for c in range(1, ws.max_column + 1):
        v = str(ws.cell(row, c).value or "").replace("\n", "").strip()
        if v == label or (contains and label in v and "工时" in v):
            return c
    return None


def _day_columns(ws, row, start_after=0):
    """日期行 → {日: 列}。
    ⚠ 不能「只认数字列」：打卡表把周末的表头写成「六」「日」而不是日期数字（1,2,3,六,日,6,7…）。
    只挑数字列会让周末整列丢失，后面所有日期错位——上线前实测：40 人里 25 人被误报「记了工时但无打卡」。
    正确做法是按位置定锚：用数字标签算出 列号−日号 的固定偏移，再把整段日期区按偏移铺满。"""
    marks = []
    for c in range(start_after + 1, ws.max_column + 1):
        try:
            d = int(str(ws.cell(row, c).value).strip())
        except (TypeError, ValueError):
            continue
        if 1 <= d <= 31:
            marks.append((c, d))
    if not marks:
        return {}
    offsets = {}
    for c, d in marks:
        offsets[c - d] = offsets.get(c - d, 0) + 1
    offset = max(offsets, key=offsets.get)            # 取众数，容忍个别串位
    end = max(c for c, d in marks if c - d == offset)  # 日期区末列＝最后一个吻合偏移的数字列
    return {c - offset: c for c in range(offset + 1, end + 1) if 1 <= c - offset <= 31}


def _txt(ws, r, c):
    return "" if not c else str(ws.cell(r, c).value or "").strip()


def _num(ws, r, c):
    v = ws.cell(r, c).value if c else None
    return float(v) if isinstance(v, (int, float)) else 0.0


# ==================== 切班 ====================
def shift_type(person):
    """按人力上报的白/夜工时判班型。排班信息只有人力有，打卡本身推不出来，故以人力标记为准。
    day=纯白班｜night=纯夜班｜mixed=同月既有白班又有夜班（逐日按切班窗口各自切，见 compute_shifts）。"""
    d, n = float(person.get("白班") or 0), float(person.get("夜班") or 0)
    if n > 0 and d > 0:
        return "mixed"
    return "night" if n > 0 else "day"


class _Shifts(dict):
    """{归班日: 班次} 的 dict，另挂两样：

    · consumed —— 整天都被前一晚那个夜班用掉的**日子**
    · used_pts —— 被别的日子那一班用掉的**具体打卡**，形如 {(日, 分钟)}

    为什么要按「张」记而不是按「天」记：夜班跨零点，次日凌晨的宵夜卡、早上的下班卡
    都属于起点日那一班，可它们记在次日的格子里。只按天记的话，
    次日若还有当晚的上班卡（黄寒寒 6/9：00:15、00:42、08:32 属于 8 日那班，20:03 是当晚新上班），
    这一天就既不能整天划走、又会把 4 张卡全算到自己头上——
    于是 8 日那班显示「次数 1」，9 日显示「次数 4、跨度对不上」。按张记才分得开。"""
    consumed = frozenset()
    used_pts = frozenset()


def compute_shifts(days, kind, params, only=None):
    """打卡 {日:[分钟]} → {归班日: {"start","end","span","hours"}}。
    白班：当日首卡→当日末卡，扣 day_break。
    夜班：当日 night_start_from 之后的首卡 → 次日 night_end_by 之前的末卡，扣 night_break，**归班日＝上班日**。
          （已被人力数据反证：黄寒寒 6 月 10 个夜班起点，人力就记在起点日）
    only：**只为这几天切班**（一般传「这一行报了工时的日子」）。仍然能读到次日的打卡去找夜班的下班卡，
          只是不为次日单独造一个班。为的是拆行的人：黄相元同月有「植物肉白班」和「小料夜班」两行，
          两行若都拿整月打卡去切，白班那行会把 6 日夜班的宵夜卡（7 日 00:22/08:36）当成自己的一天，
          切出 19.4 小时的假班；夜班那行又会把 3 日 17:30 当成上班卡、抢走 4 日早上的 07:25。
          **白/夜怎么判完全不变**——kind 来自汇总表那一行自己的白班/夜班工时列，
          白夜混合仍按首卡落在哪个窗口逐日判。改的只是「这一行管哪几天」。

    白夜混合：**逐日按首卡落在哪个窗口决定这天是白是夜**，再各自按上面那套切。
          切班规则是 2026-06 全量实证出来的：白班 07:20–08:10 上班、20:20–20:45 下班；
          夜班 19:30–20:10 上班、23:00–00:50 是宵夜卡、**次日 08:20–08:40 才下班**。
          原来混合的人整档走白班分支，夜班起点日只算到宵夜卡（丁玲 5/27 上班 20:01、
          末卡次日 00:54 → 在厂 3.5h vs 上报 11h），于是被判「撑不起」——那是切错了，不是真异常。"""
    out, used, pts = {}, set(), set()
    if kind == "mixed":
        consumed = set()                     # 已被前一夜班用掉的卡：(日, 分钟)
        for d in sorted(days):
            if only is not None and d not in only:
                continue
            ts = [t for t in days[d] if (d, t) not in consumed]
            if not ts:
                continue
            if ts[0] >= params["night_start_from"]:          # 首卡在夜班窗口 → 这天是夜班
                s0 = ts[0]
                nxt = days.get(d + 1) or []
                tail = [t for t in nxt if t <= params["night_end_by"]]
                if tail:
                    e0 = tail[-1] + 1440
                    for t in tail:
                        consumed.add((d + 1, t))             # 次日早上那几张是这一班的下班卡
                else:
                    later = [t for t in ts if t > s0]
                    if not later:
                        continue
                    e0 = later[-1]
                brk = params["night_break"]
            else:                                            # 否则按白班切
                if len(ts) < 2:
                    continue
                s0, e0, brk = ts[0], ts[-1], params["day_break"]
            span = (e0 - s0) / 60.0
            if span > 0:
                _all = sorted([t for t in ts if s0 <= t <= e0]
                              + [t + 1440 for t in (days.get(d + 1) or [])
                                 if (d + 1, t) in consumed])
                out[d] = {"start": s0, "end": e0, "span": span, "n": len(_all),
                          "mid": _all[1:-1],
                          "hours": max(round_step(span - brk, params), 0.0)}
        res = _Shifts(out)
        res.used_pts = frozenset(consumed)
        # 混合的人是按 (日, 分钟) 记的，换算成「整天都被吃掉」的日子
        res.consumed = frozenset(
            d2 for d2 in days
            if days[d2] and all((d2, t) in consumed for t in days[d2]))
        return res
    if kind == "night":
        for d in sorted(days):
            if d in used or (only is not None and d not in only):
                continue
            st = [t for t in days[d] if t >= params["night_start_from"]]
            if not st:
                continue
            s = st[0]
            nxt = days.get(d + 1, [])
            tail = [t for t in nxt if t <= params["night_end_by"]]
            if tail:
                e = tail[-1] + 1440
                if len(tail) == len(nxt):        # 次日整天都属于这一班，别再单独成一天
                    used.add(d + 1)
            else:
                later = [t for t in days[d] if t > s]
                if not later:
                    continue
                e = later[-1]
            span = (e - s) / 60.0
            if span > 0:
                own = [t for t in days[d] if t >= s]          # 起点日里属于这一班的（上班卡起）
                allp = sorted(own + [t + 1440 for t in tail])
                out[d] = {"start": s, "end": e, "span": span, "n": len(allp),
                          "mid": allp[1:-1],                    # 上下班之外的（宵夜卡等）
                          "hours": max(round_step(span - params["night_break"], params), 0.0)}
                pts |= {(d + 1, t) for t in tail}             # 次日那几张归这一班，不再算次日的
        res = _Shifts(out)
        res.consumed = frozenset(used)
        res.used_pts = frozenset(pts)
        return res
    for d, ts in days.items():
        if only is not None and d not in only:
            continue
        if len(ts) < 2:
            continue
        span = (ts[-1] - ts[0]) / 60.0
        if span <= 0:
            continue
        out[d] = {"start": ts[0], "end": ts[-1], "span": span, "n": len(ts), "mid": ts[1:-1],
                  "hours": max(round_step(span - params["day_break"], params), 0.0)}
    return _Shifts(out)


def post_of(person):
    """人的岗位。汇总表里这一列（备注/性质）空着＝普工，绝大多数人如此。"""
    return (person.get("kind") or "").strip() or POST_DEFAULT


def rate_of(person, shift, table):
    """按「派遣方 × 岗位 × 班次」在给定的表里查单价（纯查表，不带任何默认值）。
    返回 (员工工资单价, 管理费单价, 含管理费总价, 来源说明)。查不到返回 0 并写明原因——
    宁可标出来交人工，也不拿别的档顶上去：拿白班顶夜班、或拿普工价顶保洁，都是直接算错钱。"""
    table = table or {}
    shift = "night" if shift == "night" else "day"
    cn = "夜班" if shift == "night" else "白班"
    agency = (person.get("agency") or "").strip()
    post = post_of(person)
    posts = table.get(agency)
    if not posts:
        return 0.0, 0.0, 0.0, f"派遣方「{agency or '空'}」不在单价表里"
    band = posts.get(post)
    if band is None:
        # 岗位不在该派遣方名下：**不回落到普工**。保洁按普工价算就是每小时多算 4 元。
        return 0.0, 0.0, 0.0, f"{agency}名下没有「{post}」岗的单价，请在单价表里补一行"
    v = band.get(shift)
    if v is None:
        return 0.0, 0.0, 0.0, f"{agency}·{post}无{cn}单价"
    return v[0], v[1], v[0] + v[1], f"{agency}·{post}·{cn}"


def contract_rate(person, shift, contract):
    """合同价登记表里这个人这一班的单价。查不到返回 0 并把原因带出去——原因会原样出现在页面和报告里。"""
    shift = "night" if shift == "night" else "day"
    cn = "夜班" if shift == "night" else "白班"
    agency = (person.get("agency") or "").strip()
    post = post_of(person)
    posts = (contract or {}).get(agency)
    if not posts:
        return 0.0, 0.0, 0.0, f"合同价登记表里没有「{agency or '（空）'}」覆盖本期的行"
    band = posts.get(post)
    if band is None:
        return 0.0, 0.0, 0.0, f"合同价登记表里{agency}名下没有「{post}」岗"
    v = band.get(shift)
    if v is None:
        return 0.0, 0.0, 0.0, f"合同价登记表里{agency}·{post}没有{cn}价"
    return float(v[0]), float(v[1]), float(v[0]) + float(v[1]), f"合同价 {agency}·{post}·{cn}"


def contract_band(person, shift, contract):
    """合同价登记表里这个人这一班的 (员工工资, 管理费)；**没登记返回 None**。
    「登记了 0」也是登记（锦绣保洁管理费就是 0）——判有没有登记只能看这里是不是 None，
    不能看「价>0」：否则保洁档人力多收了管理费，逐人金额核对会因为"合同是 0"而跳过。"""
    shift = "night" if shift == "night" else "day"
    posts = (contract or {}).get((person.get("agency") or "").strip())
    band = posts.get(post_of(person)) if posts else None
    v = band.get(shift) if band else None
    return (float(v[0]), float(v[1])) if v else None


def actual_rate(person, shift):
    """结算表上这个人实际套用的单价（人力自己填的那几列）。表里没有金额列时返回 None。"""
    t = person.get("表上") or {}
    if not t:
        return None
    cn = "夜班" if shift == "night" else "白班"
    w, m = float(t.get(f"{cn}工资单价") or 0.0), float(t.get(f"{cn}管理费单价") or 0.0)
    if w <= 0 and m <= 0:
        return None
    return w, m, w + m, f"结算表上该人的{cn}单价"


def dev_rate(person, shift, contract):
    """给【偏离】计价用的单价：多记 1 小时值多少钱。
    取**结算表上该人实际套用的单价**——公司就是按这个价付的，多记 1 小时公司就多付这么多；
    结算表没有金额列才退到合同价；两者都没有记 0，并把原因带出去。
    ⚠ 不拿任何反推出来的表兜底。"""
    a = actual_rate(person, shift)
    if a:
        return a
    return contract_rate(person, shift, contract)


# ==================== 主流程 ====================
def compute(summary, punch, params=None, contract=None):
    """逐日比对 + 四档判定 + 逐人汇总 + 全表统计。summary/punch 为上面两个 parse_* 的返回值。

    contract：本期适用的合同价表 {派遣方: {岗位: 档}}，来自成本会计的合同价登记表（路由层按生效期挑好再传进来）。
    ⚠ 单价只有两个来路，各司其职：
      · 应付 / 结算表自查 / 合同价核对 → **只认 contract**，缺档就是缺档，不拿别的表顶上
      · 偏离计价（多记 1 小时值多少钱） → 结算表上该人实际套用的单价，没有才退到 contract
    汇总表表头解析出来的规则文字**不参与计算**，只原样带回页面作参考（它是人力自己写的）。"""
    p = dict(DEFAULT_PARAMS)
    p.update({k: v for k, v in (params or {}).items() if v is not None})
    tol = float(p["tolerance"])
    basis = "punch" if str(p.get("report_basis")) == "punch" else "shift"
    contract = contract_only(contract)
    contract_src = ("成本会计在「合同价（成本会计维护）」页登记的行" if contract
                    else "本期一行合同价都没登记")
    rows, people, unmatched, ambiguous, no_contract, no_dev_rate = [], [], [], [], [], []
    # 同一个人当月在两个车间干过，工资要分摊，汇总表就把他拆成两行（「按业务线拆行」，6 月有 117 组）。
    # 拆出来的两行各自只在自己那几天有工时，另一行那几天是 0。
    # 逐日比对若照单全收，就会在「另一行的日子」上报一条「◇有打卡·当天未计临时工工时」——
    # **那天他明明记了工时，只是记在另一行**，这是拆行造出来的假象。
    # 实测 2026-06：1,242 条「未计工时」里 946 条是这么来的，真实只有 306 条。
    # 所以先记下「这个人这一天在**任何一行**上有没有工时」，有就别在别的行上重复摆一遍。
    # 每一行用掉了哪几张卡，先扫一遍记下来（按行记，不并成一份）。
    # ⚠ 这一步必须在「每行只切自己报工的日子」之后才安全：早先每行都拿整月打卡去切，
    #   夜班那行会把 3 日 17:30 当成上班卡、吃掉 4 日早上的 07:25，
    #   合并归属后白班那行就少了上班卡（重算 9.0 → 5.0，撑不起 4 → 164）。踩过一次，别再颠倒顺序。
    _used_by_row, _rows_of = {}, {}
    for _p in summary["people"]:
        _rows_of.setdefault(same_person_key(_p["name"]), []).append(id(_p))
        _rec, _ = match_punch(punch, _p["name"])
        if not _rec:
            continue
        _own = {int(d) for d, h in (_p.get("days") or {}).items() if h and float(h) > 0}
        _sh = compute_shifts(_rec["days"], shift_type(_p), p, only=_own)
        _used_by_row[id(_p)] = set(getattr(_sh, "used_pts", ()))

    _shown_empty = set()          # 已经摆过一条「这天没上报工时」的 (姓名, 日)，别重复
    _hours_that_day = {}
    for _p in summary["people"]:
        for _d, _h in (_p.get("days") or {}).items():
            if _h and float(_h) > 0:
                _hours_that_day[(same_person_key(_p["name"]), int(_d))] = True
    for person in summary["people"]:
        rec, cand = match_punch(punch, person["name"])
        if cand:
            ambiguous.append({"姓名": person["name"], "候选": cand})
        elif not rec:
            unmatched.append(person["name"])
        kind = shift_type(person)
        pdays = rec["days"] if rec else {}
        # 这一行只为**自己报了工时的日子**切班。没人报工时的日子照样会摆一条中性行
        # （有打卡·未计工时 / 仅1次卡），只是不再由某一行硬切出一个班次来。
        _own_days = {int(d) for d, h in (person.get("days") or {}).items() if h and float(h) > 0}
        shifts = compute_shifts(pdays, kind, p, only=_own_days)
        # 偏离计价单价按这个人本月的班型取；白夜混合按白班取并在页面单列提示。
        # （逐日的白/夜已经切得出来了，但这个人整月的「代表单价」仍取白班；
        #   应付另按结算表上的白班/夜班工时分开乘各自单价算，不受这里影响）
        w, m, rate, src = dev_rate(person, "night" if kind == "night" else "day", contract)
        if rate <= 0:
            no_dev_rate.append({"姓名": person["name"], "归属": person["agency"],
                                "岗位": post_of(person), "原因": src})
        pay = payable(person, contract, kind)
        gaps = rate_gaps(person, contract)
        if pay["合同缺档"]:
            no_contract.append({"姓名": person["name"], "归属": person["agency"],
                                "岗位": post_of(person),
                                "班型": {"day": "白班", "night": "夜班", "mixed": "白夜混合"}[kind],
                                "原因": pay["合同缺档"]})
        agg = {"少记日": 0, "少记时": 0.0, "弹性内日": 0, "异常日": 0, "异常时": 0.0,
               "一致日": 0, "硬伤日": 0, "混合日": 0, "未计日": 0, "薄卡日": 0,
               "多出日": 0, "多出时": 0.0, "上报": 0.0, "重算": 0.0,
               # 净多记＝「判过的班日」上 Σ(上报−重算)，>0 即公司整期多付；已消化日＝逐日超弹性但整期没超、被降级的天
               "净多记": 0.0, "已消化日": 0}
        _row_start = len(rows)          # 本人在 rows 里的起点，供整期判定后回头给逐日降级用
        for d in sorted(set(person["days"]) | set(shifts) | set(pdays)):
            rep = float(person["days"].get(d, 0.0))
            sh = shifts.get(d)
            cal = float(sh["hours"]) if sh else 0.0
            # 这一天「自己的卡」＝当天格子里的卡，减去被前一晚那一班用掉的那几张。
            # 不减的话，夜班次日会背着上一班的宵夜卡和下班卡，次数和跨度都对不上（黄寒寒 6/9 实测）。
            # 这一天「自己的卡」＝当天格子里的，减去**任何一行**（含自己）那一班跨日用掉的。
            # 拆行的人尤其要这样：黄相元 7 日的 00:22/08:36 是小料夜班行 6 日那一班的卡，
            # 植物肉白班行不该再拿它们单独摆一天。
            _used = set(getattr(shifts, "used_pts", ()))
            for _rid in _rows_of.get(same_person_key(person["name"]), ()):
                _used |= _used_by_row.get(_rid, set())
            ts = [t for t in pdays.get(d, []) if (d, t) not in _used]
            if rep <= 0 and not ts:
                continue
            # 本行这天没工时，但同一个人另一行有 → 那天归另一行管，这里跳过，别重复摆一条
            if rep <= 0 and _hours_that_day.get((same_person_key(person["name"]), int(d))):
                continue
            # 这一天的打卡整天都被前一晚那个夜班用掉了（次日早上那张下班卡），
            # 班已经算在起点日了，这里不该再摆一行——否则夜班的人每天都多出一条
            # 「仅1次卡·疑似无效」或「有打卡·未计工时」（2026-06 实测 142+126 条）。
            # 但**如果这天上报了工时就必须摆**：那是真冲突，得让人看见。
            if rep <= 0 and d in getattr(shifts, "consumed", ()):
                continue
            # 这一天没上报工时、当天又切不出班次，而所有打卡都落在**夜班下班窗口之内**
            # （凌晨到 11:30 之前）→ 这是前一晚那个夜班的下班卡。
            # 上面那条 consumed 只盖得住「数据里有前一晚」的情形；跨月的盖不住——
            # 何青建、李仲 1 日凌晨那几张卡，属于上个月 31 日的夜班，本月数据里根本没有那一晚。
            # ⚠ 只对**夜班/白夜混合**的人生效。白班的人早上那张孤零零的上班卡也落在这个窗口里，
            #   但它是上班卡不是下班卡，正是该报的「仅1次卡·疑似无效」——
            #   不加这个限定会把苏兵 7/7 那条真发现吃掉（单元测试当场抓住）。
            if (rep <= 0 and not sh and ts and kind in ("night", "mixed")
                    and max(ts) <= p["night_end_by"]):
                continue
            # 同一个人同一天、几行都没上报工时（按业务线拆行的人在双方都没排班的日子）→ 只摆一条
            if rep <= 0:
                _k = (same_person_key(person["name"]), int(d))
                if _k in _shown_empty:
                    continue
                _shown_empty.add(_k)
            diff = round(cal - rep, 2)
            # basis=shift：上报是排班班次时长，本就＜在厂时长，「差多少」没有意义；
            #              只问一件事——**打卡撑不撑得起这个班**。撑不起才是公司可能多付了钱。
            # basis=punch：上报应当等于打卡重算，少记/多记分档（原口径）。
            if kind == "mixed" and basis == "punch":
                # punch 口径下混合日的重算值不可信（少记/多记要精确到小时），仍旧只摆不判。
                # shift 口径不需要——它只问「在厂撑不撑得起上报」，而 compute_shifts 已按
                # 切班窗口逐日切过白/夜了，所以下面照常判。
                judge, cls = J_MIXED, "mixed"
            elif not ts and rep > 0:
                judge, cls = J_NO_PUNCH, "hard"
            elif rep <= 0 and len(ts) <= 1:
                judge, cls = J_THIN, "thin"
            elif rep <= 0:
                judge, cls = J_UNBILLED, "unbilled"
            elif basis == "shift":
                if diff < -tol - 1e-9:      # 在厂时长不够覆盖上报的班次
                    judge, cls = f"{J_OVER_OUT} {-diff:.1f} 小时(撑不起上报，整期超弹性)", "over_out"
                else:
                    judge, cls = (f"{J_BACKED}（在厂 {cal:.1f}h ≥ 上报 {rep:.1f}h）"
                                  if diff > 1e-9 else J_BACKED), "ok"
            elif diff > 1e-9:
                judge, cls = f"{J_UNDER} {diff:.1f} 小时", "under"
            elif diff < -tol - 1e-9:
                judge, cls = f"{J_OVER_OUT} {-diff:.1f} 小时(超弹性,异常)", "over_out"
            elif diff < -1e-9:
                judge, cls = f"{J_OVER_IN} {-diff:.1f} 小时(弹性内)", "over_in"
            else:
                judge, cls = J_OK, "ok"
            # 切不出班次却有好几张卡：说清为什么，别让人对着「跨度 —」猜
            # （黄寒寒 6/9：00:15/00:42/08:32 是 8 日夜班的下班卡，20:03 是又刷了一次没下班）
            if not sh and len(ts) >= 2 and cls in ("unbilled", "thin"):
                judge += f"（当天 {len(ts)} 次卡分属前后两个班，切不出完整班次）"
            agg["上报"] += rep
            agg["重算"] += cal
            if cls in ("ok", "over_out"):     # 只累「判过的班日」（上报>0、比过在厂）；中性天不参与整期净额相抵
                agg["净多记"] += (rep - cal)
            agg[{"under": "少记日", "over_in": "弹性内日", "over_out": "异常日",
                 "ok": "一致日", "mixed": "混合日", "unbilled": "未计日",
                 "thin": "薄卡日"}.get(cls, "硬伤日")] += 1
            if cls == "ok" and diff > 1e-9:      # 参考量：人在厂里、但没算进上报的时间
                agg["多出日"] += 1
                agg["多出时"] += diff
            if cls == "under":
                agg["少记时"] += diff
            if cls == "over_out":
                agg["异常时"] += -diff
            rows.append({
                "姓名": person["name"], "部门": person["dept"], "归属": person["agency"],
                "岗位": person["kind"] or "普工", "班型": {"day": "白班", "night": "夜班", "mixed": "白夜混合"}[kind],
                "日": d, "上班打卡": fmt_hm(sh["start"]) if sh else (fmt_hm(ts[0]) if ts else ""),
                "下班打卡": fmt_hm(sh["end"]) if sh else (fmt_hm(ts[-1]) if len(ts) > 1 else ""),
                # ⚠ 切不出班次时**别硬凑一个跨度**：这天的卡可能分属前后两个班
                # （黄寒寒 6/9 的 00:15/00:42/08:32 是 8 日夜班的下班卡，20:03 是又刷了一次没下班），
                # 拿首末两张算出来的「跨度」既不是在厂时长、也不对应任何一个班，摆出来只会误导。
                # 次数按「这一班用了几张卡」算，不按日历天数——夜班的宵夜卡和次日下班卡都算进起点日
                "打卡次数": (sh.get("n") or len(ts)) if sh else len(ts),
                # 上下班之外的那些卡（午休/宵夜等）。它们不参与在厂时长，但摆出来行才读得通：
                # 不然「上班 19:51 → 下班 次日08:32、次数 4」中间那两张去哪了，看的人只能猜。
                "无效卡": [fmt_hm(t) for t in (sh.get("mid") if sh else ts[1:-1])],
                "跨度": round(sh["span"], 2) if sh else None,
                "上报工时": rep, "重算工时": cal, "差异": diff,
                "单价": rate, "单价来源": src,
                # 「金额」＝差异 × 单价，只在**上报了工时**的日子才有意义（多记 1 小时就多付这么多）。
                # 上报为 0 的中性档（有打卡未计工时 / 仅1次卡），差异＝整段在厂时长，
                # 乘出来的数不是任何意义上的风险敞口——黄相元 6/7 因此显示「¥342」，
                # 而那天公司一分钱没付。这种数摆在金额列里只会误导，一律记 0。
                "金额影响": round(diff * rate, 2) if rep > 0 else 0.0,
                "判定": judge, "档": cls,
            })
        # ── 人级「整期总量」判定（与成本会计底表一致：按人合起来比，弹性 0.5，白/夜各自成型）──
        # 逐日照算差异；异常与否看**这个人整期净多记**是否超弹性。整期没超 → 把逐日那几天的
        # 「超弹性异常」降级为中性（仍显示多记几小时，但不红、不必查、不计异常金额）。
        # 白夜混合的人在厂重算本就只摆不判，整期净额不可靠 → 维持「单独交人工」，不进此判定。
        _net = round(agg["净多记"], 2)
        _over = kind != "mixed" and _net > tol + 1e-9
        if not _over:
            for _r in rows[_row_start:]:
                if _r["档"] == "over_out":
                    _r["档"] = "over_absorbed"
                    _r["判定"] = f"○ 多记 {-_r['差异']:.1f} 小时(撑不起上报，整期已消化)"
                    agg["已消化日"] += 1
            agg["异常日"] = 0
            agg["异常时"] = 0.0
        people.append({
            "序号": person["no"], "姓名": person["name"], "部门": person["dept"], "归属": person["agency"],
            "岗位": person["kind"] or "普工", "班型": {"day": "白班", "night": "夜班", "mixed": "白夜混合"}[kind],
            "匹配打卡": bool(rec), "考勤组": rec["组"] if rec else "",
            "上报总工时": round(agg["上报"], 2), "重算总工时": round(agg["重算"], 2),
            "差异": round(agg["重算"] - agg["上报"], 2),
            "差额金额": round((agg["重算"] - agg["上报"]) * rate, 2),
            "少记日次": agg["少记日"], "少记小时": round(agg["少记时"], 2),
            "弹性内多记日次": agg["弹性内日"],
            # 整期口径：超弹性＝这个人整期净多记 > 弹性（白夜混合除外，维持交人工）。
            "超弹性": _over, "整期净多记小时": round(max(0.0, _net), 2),
            "异常多记日次": agg["异常日"],          # 仍在异常档的逐日天数（整期没超时已全部降级为 0）
            "整期已消化多记日次": agg["已消化日"],   # 逐日冒尖、但整期没超被降级的天
            "异常多记小时": round(_net, 2) if _over else 0.0,      # 超弹性时＝整期净多记（公司整期多付的量）
            "异常多记金额": round(_net * rate, 2) if _over else 0.0,
            "一致日次": agg["一致日"], "待查日次": agg["硬伤日"], "白夜混合日次": agg["混合日"],
            "未计工时日次": agg["未计日"] + agg["薄卡日"],
            "打卡多于上报日次": agg["多出日"], "打卡多于上报小时": round(agg["多出时"], 2),
            "员工单价": w, "管理费单价": m, "含管理费单价": rate, "单价来源": src,
            # 单价不符逐人挂一份：②总览只说「N 格不符、涉及 M 人」，点进逐人核对得能认出是哪几个人
            "单价不符": gaps,
            # 应付按**上报工时 × 合同价**算，与上面的偏离统计各走各的——一个是要付的钱，一个是要查的差
            **pay,
        })
    stats = _stats(rows, people, summary, punch, p, unmatched, ambiguous, no_contract, no_dev_rate)
    stats["单价不符人数"] = sum(1 for x in people if x.get("单价不符"))
    # ⚠ 「按人去重」的口径只有这一个。contract_vs_actual 里那个「人数」是**按格×班次累加的人次**，
    #    白夜混合的人两班都不符就会被数两次——页面报「涉及 N 人」必须用下面这个，别拿格里的加总
    stats["单价不符金额"] = round(sum(rate_gap_amount(x) for x in people), 2)
    stats["单价不符多付"] = round(sum(a for a in (rate_gap_amount(x) for x in people) if a > 0), 2)
    stats["单价不符少付"] = round(sum(a for a in (rate_gap_amount(x) for x in people) if a < 0), 2)
    stats["单价不符工时"] = round(sum(float(x.get("上报总工时") or 0)
                                      for x in people if x.get("单价不符")), 1)
    stats["金额核对"] = pay_check(summary["people"], contract)
    stats["金额核对条数"] = len(stats["金额核对"])
    stats["有表上金额"] = sum(1 for x in people if x.get("表上合计") is not None)
    # 奖罚补贴单列一段：它们是全表唯一没有对照源的钱，只能列出来 + 验符号/占比，验不了金额本身
    stats["合同外调整"] = adjust_check(summary["people"], p.get("adj_ratio_cap", ADJ_RATIO_CAP))
    stats["合同外调整合计"] = adjust_total(stats["合同外调整"])
    # 合同价 vs 人力实际计价：第③步「对比」页签要的就是这个
    stats["单价核对"] = contract_vs_actual(summary["people"], contract)
    stats["单价核对"]["合同来源"] = contract_src
    def _ser(tbl):
        return {a: {post: {"day": list(b["day"]) if b.get("day") else None,
                           "night": list(b["night"]) if b.get("night") else None}
                    for post, b in posts.items()}
                for a, posts in (tbl or {}).items()}
    return {"params": p, "rows": rows, "people": people, "stats": stats,
            "settle": settle(people),
            "rates": {"合同表": _ser(contract), "合同来源": contract_src, "默认岗位": POST_DEFAULT,
                      # 下面三项是人力表头写的，只作参考展示，不参与任何计算
                      "表头解析": _ser(_as_nested(summary.get("rate_table"))),
                      "对外总价": summary.get("rate_headline") or {},
                      "表头未解析行": summary.get("rate_notes") or []}}


def _stats(rows, people, summary, punch, p, unmatched, ambiguous, no_contract=None, no_dev_rate=None):
    _dups = cross_agency(summary, punch)
    _AGM = agency_mismatch(summary, punch)
    def s(cls, field):
        return round(sum(abs(r[field]) for r in rows if r["档"] == cls), 2)
    mixed = [x["姓名"] for x in people if x["班型"] == "白夜混合"]
    return {
        "人数": len(people), "比对人日": len(rows),
        "上报总工时": round(sum(x["上报总工时"] for x in people), 2),
        "重算总工时": round(sum(x["重算总工时"] for x in people), 2),
        "差异小时": round(sum(x["差异"] for x in people), 2),
        "差额金额": round(sum(x["差额金额"] for x in people), 2),
        "一致日次": sum(1 for r in rows if r["档"] == "ok"),
        "少记日次": sum(1 for r in rows if r["档"] == "under"),
        "少记小时": s("under", "差异"), "少记金额": s("under", "金额影响"),
        "弹性内多记日次": sum(1 for r in rows if r["档"] == "over_in"),
        "弹性内多记小时": s("over_in", "差异"), "弹性内多记金额": s("over_in", "金额影响"),
        # 整期口径（与成本会计一致）：异常＝按人整期净多记超弹性。逐日日次仍给（供下钻定位），
        # 但小时/金额按**人级整期净多记**算，不再逐日累加。
        "超弹性人数": sum(1 for x in people if x.get("超弹性")),
        "异常多记日次": sum(1 for r in rows if r["档"] == "over_out"),
        "异常多记小时": round(sum(x["异常多记小时"] for x in people), 2),
        "异常多记金额": round(sum(x["异常多记金额"] for x in people), 2),
        "整期已消化多记日次": sum(1 for r in rows if r["档"] == "over_absorbed"),
        # 待查只留**真该查的**：报了工时、却一次卡都没有。「没记工时」的那些归下面「未计工时」
        "待查日次": sum(1 for r in rows if r["档"] == "hard"),
        "白夜混合日次": sum(1 for r in rows if r["档"] == "mixed"),
        # 有打卡没算工时：单独一档，**不计入待查**——打卡表是全厂的，这些天多半是这人在别的名目下上班
        "未计工时日次": sum(1 for r in rows if r["档"] in ("unbilled", "thin")),
        "未计工时人数": len({r["姓名"] for r in rows if r["档"] in ("unbilled", "thin")}),
        "未计工时·仅1次卡": sum(1 for r in rows if r["档"] == "thin"),
        # 参考量：打卡撑得住、但在厂比上报多出来的时间（班次制下这是常态，不是少记）
        "打卡多于上报日次": sum(1 for r in rows if r["档"] == "ok" and r["差异"] > 0),
        "打卡多于上报小时": round(sum(r["差异"] for r in rows if r["档"] == "ok" and r["差异"] > 0), 2),
        "上报口径": p.get("report_basis"),
        "白夜混合人数": len(mixed), "白夜混合名单": mixed[:50],
        "未匹配打卡": unmatched[:50], "未匹配人数": len(unmatched),
        "待人工指认": ambiguous[:50], "待指认人数": len(ambiguous),
        "缺合同价": (no_contract or [])[:80], "缺合同价人数": len(no_contract or []),
        "偏离未计价": (no_dev_rate or [])[:50], "偏离未计价人数": len(no_dev_rate or []),
        "打卡表重名": punch.get("dup_raw") or [],
        "汇总表月份": summary.get("period", ""), "打卡表月份": punch.get("period", ""),
        "汇总表标题": summary.get("title", ""), "打卡表标题": punch.get("title", ""),
        "月份不一致": bool(summary.get("period") and punch.get("period")
                          and summary.get("period") != punch.get("period")),
        "计价规则": summary.get("rules", ""),
        # 「按业务线拆行」是正常成本归集，不进需要人工看的清单（6 月 116/117 组是这一类）
        "同名多行": [x for x in _dups if not x.get("按业务线拆行")],
        "同名·按业务线拆行": sum(1 for x in _dups if x.get("按业务线拆行")),
        "同名跨派遣方数": sum(1 for x in _dups if x["跨派遣方"] and not x.get("按业务线拆行")),
        "同名高风险数": sum(1 for x in _dups if x["高风险"]),
        "同名同日重叠数": sum(1 for x in _dups if x.get("重叠日")),
        "归属与打卡不符": _AGM,
        # 打卡表没给部门的人，这一项**没查**——用钉钉取数时离职员工就是这样（2026-06 有 169 人）。
        # 必须报出来：没查和没事是两回事。
        "归属无法核对人数": len(getattr(_AGM, "blind", ())),
        "归属无法核对名单": list(getattr(_AGM, "blind", ()))[:50],
    }


# ==================== 结算风险核对 ====================
def cross_agency(summary, punch=None):
    """同一个月里，同名的人出现在两家及以上派遣方，或在同一派遣方出现多行。
    这是**结算风险**：一个人被两家同时计费，或一行被录了两遍——两种都直接多付钱。

    ⚠ 注意「同名」是**归一之后**的同名：公司本来就用后缀区分同名者（张博G / 张博J、黄亚军0415 / 黄亚军7327），
    是归一把后缀抹掉才撞到一起的。所以光看名字判不了，要另找硬判据：

    **首选：打卡表里的「手机尾号」列**（钉钉取数生成的表带这一列，人力手工导的没有）。
      手机号是钉钉侧的身份，比什么都硬：
      · 每行尾号都不同 → **确认是不同的人，直接不报**（张博G 3038 / 张博J 9170，实测）
      · 尾号相同 → 确认是同一个人，再看日期有没有重叠
      没有这一列时才退回下面这套弱判据。

    次选（没有手机尾号时）——**看打卡行**：
      · 两个名字在打卡表里各有各的行 → 两个人，正常，降为提示
      · 两个名字指向**同一行打卡** → 同一个人被登记了两次，高风险
      · 打卡表里查不到 → 存疑，交人工

    ⚠ 同一派遣方内的多行，还要再问一句「**钱会不会被算两次**」：
      工资按业务线分摊到车间，同一个人当月在小料和植物肉都干过，**就必须拆成两行**——
      这是正常的成本归集，不是重复录入。判据是**日期有没有重叠**：
      · 同一天两行都记了工时 → 钱可能算两次 → 高风险
      · 日期完全不重叠、部门又不同 → 按业务线拆行，正当，不报风险
    （2026-06 全量实证：116 组同名同派遣方多行，全部是「小料+植物肉」且日期不重叠，
      0 组重叠。7 月那张拆分表只有植物肉一个车间，所以这条从没触发过。）
    """
    by = {}
    for p_ in summary["people"]:
        by.setdefault(norm_name(p_["name"]), []).append(p_)
    out = []
    for key, rows in by.items():
        if len(rows) < 2:
            continue
        # ── 第一步：有手机尾号就先按身份把这一组拆开 ──
        # 一个归一名下可能同时装着「另一个人」和「同一个人的两行」：
        # 张博G（尾号 3038）是别人，两行张博J（尾号 9170）才是同一个人按车间拆行。
        # 整组一起判会把它们混成一坨，反而升成高风险（实测踩过）。
        # 拆完之后：只出现一次的身份不是风险，丢掉；留下「同一身份出现多行」那一簇再往下判。
        idents = []
        if punch:
            for r in rows:
                rec, _ = match_punch(punch, r["name"])
                idents.append(str((rec or {}).get("标识") or "").strip())
            if idents and all(idents) and len(set(idents)) > 1:
                cl = {}
                for r, i in zip(rows, idents):
                    cl.setdefault(i, []).append(r)
                multi = [v for v in cl.values() if len(v) > 1]
                if not multi:
                    continue                      # 每个身份各一行 → 全是不同的人，不是风险
                rows = max(multi, key=len)

        # ⚠ 这三个必须在拆组**之后**算——rows 可能已经换成其中一簇了
        agencies = sorted({(r["agency"] or "").strip() for r in rows})
        raws = [r["name"] for r in rows]
        same_raw = len(set(raws)) == 1
        pk_rows, pk_note, idents = [], "", []
        if punch:
            for r in rows:
                rec, _ = match_punch(punch, r["name"])
                pk_rows.append(rec.get("row") if rec else None)
                idents.append(str((rec or {}).get("标识") or "").strip())
            # 尾号齐全：两两不同＝不同的人（不是风险）；相同＝钉钉确认是同一个人
            if idents and all(idents):
                if len(set(idents)) == len(idents):
                    continue
                pk_note = "手机尾号相同 → 钉钉确认是同一个人"
            got = [x for x in pk_rows if x]
            if pk_note:
                pass
            elif len(got) == len(rows) and len(set(got)) == 1:
                pk_note = "两行指向同一条打卡记录 → 很可能是同一个人被登记了两次"
            elif len(got) == len(rows) and len(set(got)) == len(got):
                pk_note = "两行在打卡表里各有各的记录 → 大概率是两个不同的人"
            else:
                pk_note = "打卡表里对不齐，无法判断是不是同一个人"
        # 逐日看：同一天有没有两行都记了工时——这是「钱被算两次」的唯一硬证据。
        # ⚠ 只在**同一个原名**内部比。公司本来就用后缀区分同名者（黄亚军0415 / 黄亚军7327），
        #   跨原名比等于拿甲的上班日去撞乙的上班日，必然误报（2026-06 实测踩过）。
        #   两个原名是不是同一个人，交给下面的「打卡判据」去说。
        # 分组用的键：有手机尾号就用它（同一个人可能写成两个名字），否则退回原名
        _key = (lambda i, r: idents[i]) if (idents and all(idents)) else (lambda i, r: r["name"])
        _byraw = {}
        for i, r in enumerate(rows):
            _byraw.setdefault(_key(i, r), []).append(r)
        overlap = set()
        for _same in _byraw.values():
            _ds = [{d for d, h in (r.get("days") or {}).items() if h and float(h) > 0} for r in _same]
            for i in range(len(_ds)):
                for j in range(i + 1, len(_ds)):
                    overlap |= _ds[i] & _ds[j]
        depts = sorted({r["dept"] for r in rows if r["dept"]})
        # 同一派遣方 + 同一原名内日期不重叠 + 部门不同 ＝ 按业务线拆行，正当，不是风险
        split_by_line = (len(agencies) == 1 and not overlap and len(depts) > 1
                         and all(len({x["dept"] for x in v}) == len(v) for v in _byraw.values()))
        # 高风险＝原名完全相同、或打卡指向同一个人；但**按业务线拆行的一律不算**
        high = (same_raw or pk_note.startswith("两行指向同一条")
                or pk_note.startswith("手机尾号相同")) and not split_by_line
        if split_by_line:
            risk = ("按业务线拆行（%s）——同一个人当月在多个车间干过，工资要分摊到各车间，"
                    "日期不重叠、没有哪一天被算两次，属正常成本归集" % "＋".join(depts))
        elif overlap:
            risk = ("⚠ **同一个人**（同一原名）多行且同一天都记了工时（%s 日）——这几天的钱可能被算了两次"
                    % "、".join(str(d) for d in sorted(overlap)[:8]))
        elif len(agencies) > 1:
            risk = "同名挂在两家及以上派遣方——同一个人被重复计费？"
        else:
            risk = "同名在同一派遣方出现多行，部门也相同——为什么拆两行？"
        out.append({
            "归一姓名": key, "原名": raws, "原名完全相同": same_raw,
            "派遣方": agencies, "跨派遣方": len(agencies) > 1,
            "部门": depts, "重叠日": sorted(overlap), "按业务线拆行": bool(split_by_line),
            "岗位": sorted({post_of(r) for r in rows}),
            "各行工时": [r["总工时"] for r in rows],
            "合计工时": round(sum(r["总工时"] for r in rows), 2),
            "打卡判据": pk_note, "高风险": bool(high), "风险": risk,
        })
    out.sort(key=lambda x: (not x["高风险"], not x["跨派遣方"], -x["合计工时"]))
    return out


_PK_AGENCY = re.compile(r"临时普工-(.+?)人力")


def agency_of_dept(dept):
    """从打卡表的「部门」里认出派遣方。**两种写法都要认**：
      · 人力导出的是全路径：「孝感…-生产制造部-临时普工-锦绣人力」
      · 钉钉取数给的是部门树叶子名：「锦绣人力」——没有「临时普工-」前缀，原来一个都认不出来
    认不出来就返回空串（比如离职员工在钉钉里查不到部门），由调用方算进「无法核对」。"""
    d = str(dept or "").strip()
    if not d:
        return ""
    m = _PK_AGENCY.search(d)
    if m:
        return m.group(1)
    seg = d.split("-")[-1].strip()
    if seg.endswith("人力") and len(seg) > 2:
        return seg[:-2].strip()
    # ⚠ 只要末段非空就返回它，**别再要求带「-」**：钉钉给的是叶子名，本来就没有「-」，
    #   加了这个条件会把「浮动组」「中段车间」这类正常部门当成「查不到部门」，
    #   于是该报的「归属不符」没报，还被算进「没查成」的人数里（实测 15 人里有 12 个是这么来的）。
    return seg


class _MismatchList(list):
    """就是个 list，多挂一个 blind——「打卡表没给部门、这一项根本没查」的人。
    不能把这些人悄悄算成「没问题」：没查和没事是两回事。"""
    blind = ()


def agency_mismatch(summary, punch):
    """结算表的「归属」与打卡表部门里写的派遣方对不上。
    对不上不等于错——打卡部门可能写「浮动组」这类调配名目（7 月两位保洁就是），
    但结算按哪家出钱、考勤挂在哪家，两边长期不一致就该问一句。

    ⚠ 打卡表没给部门的人**核不了**，名单塞在返回值的 blind 上由调用方报出来。
      这些人的「归属不符」这一项是**没查**，不是「没问题」——两者差别很大，页面必须写明。
      2026-06 实测：人力导出那张表有 8 行没写部门（陈双珍/冯必志/李翠娥/李苏/马福），
      钉钉取数那份 0 行——**别以为没部门是钉钉的毛病，恰恰相反**。"""
    out, blind = [], []
    for p in summary["people"]:
        rec, _ = match_punch(punch, p["name"])
        if not rec:
            continue
        dept = rec.get("部门") or ""
        pk_ag = agency_of_dept(dept)
        ag = (p["agency"] or "").strip()
        if not pk_ag:
            blind.append(p["name"])        # 打卡表没给部门（钉钉取数时离职员工就是这样），核不了
        elif ag and pk_ag != ag:
            out.append({"姓名": p["name"], "结算归属": ag, "打卡部门派遣方": pk_ag,
                        "打卡部门原文": dept, "总工时": p["总工时"],
                        # 手机尾号带出来给页面拼进名字：同名的两个人（陈晶/代进莉）光看名字分不清，
                        # 加「（手机尾号XXXX）」才知道是不是同一个人（钉钉取数才有，人力导出没有）
                        "手机尾号": rec.get("标识") or ""})
    res = _MismatchList(out)               # list 的子类：调用方当列表用，额外带一份「没查成的人」
    res.blind = blind
    return res


# ==================== 应付薪资与复核结论 ====================
def payable(person, contract=None, kind=None):
    """这个人本期**按合同价**应付多少。

    ⚠ 基数是**人力上报的工时**，不是打卡重算的工时。付款本来就按上报的数走，本工具是查偏离的，
    不是重新发一遍工资（2026-08-18 业务定案）。拿重算工时去算应付，等于自说自话另立一套账。

    ⚠ 单价**只认合同价登记表**。这个人用到的班次里有任何一档没登记，应付就是 None——
    「算不出来」必须老老实实是空，不能拿别的价顶上去算出一个看似正常的数
    （V2.345 之前正是这么干的，结果没登记的派遣方全部显示「正常」）。

    工资与管理费必须**分开算**：19 元/时里 16.5 是工人拿的工资、2.5 是付给派遣方的管理费，
    合同里本来就是两条，付款、入账、发票口径都不同，合成一个数就没法用了。
    白班夜班分开乘各自单价——白夜混合的人只有这样才算得对。"""
    dh = float(person.get("白班") or 0.0)
    nh = float(person.get("夜班") or 0.0)
    if dh <= 0 and nh <= 0:
        # 表上没分白夜（只有总工时）：按班型把总工时整体归一边。
        # 班型得用调用方从打卡推出来的 kind——shift_type 只看白/夜两列，这种版式下它永远答「白班」
        kind = kind or shift_type(person)
        if kind == "night":
            nh = float(person.get("总工时") or 0.0)
        else:
            dh = float(person.get("总工时") or 0.0)
    dw, dm, _, dsrc = contract_rate(person, "day", contract)
    nw, nm, _, nsrc = contract_rate(person, "night", contract)
    missing = []
    if dh > 0 and contract_band(person, "day", contract) is None:
        missing.append(dsrc)
    if nh > 0 and contract_band(person, "night", contract) is None:
        missing.append(nsrc)
    if missing:
        wage = mgmt = None
    else:
        wage = dh * dw + nh * nw
        mgmt = dh * dm + nh * nm
    t = person.get("表上") or {}
    got = bool(t)
    extra = (t.get("补贴", 0.0) + t.get("奖", 0.0) + t.get("罚", 0.0)) if got else 0.0
    t_wage = t.get("员工工资", 0.0) if got else None
    t_mgmt = (t.get("白班管理费", 0.0) + t.get("夜班管理费", 0.0)) if got else None
    t_all = t.get("合计", 0.0) if got else None
    ad, an = actual_rate(person, "day"), actual_rate(person, "night")
    r2 = lambda v: None if v is None else round(v, 2)
    return {
        "上报白班工时": round(dh, 2), "上报夜班工时": round(nh, 2),
        # 合同价（缺档为 0）；表上单价是人力实际套用的含管理费总价（给看板拆白/夜金额用）
        "白班工资单价": dw, "白班管理费单价": dm, "夜班工资单价": nw, "夜班管理费单价": nm,
        "表上白班单价": ad[2] if ad else None, "表上夜班单价": an[2] if an else None,
        "应付工资": r2(wage), "应付管理费": r2(mgmt),
        "应付合计": r2(wage + mgmt) if wage is not None else None,
        "合同缺档": "；".join(missing) if missing else None,
        "补贴奖罚": round(extra, 2) if got else None,
        "表上工资": round(t_wage, 2) if got else None,
        "表上管理费": round(t_mgmt, 2) if got else None,
        "表上合计": round(t_all, 2) if got else None,
        # 差额剔掉补贴奖罚——那是合同外的另计项，不是单价算错
        "应付偏差": round(t_all - extra - wage - mgmt, 2) if (got and wage is not None) else None,
        "应付单价来源": ("；".join(missing) if missing
                         else (dsrc if nh <= 0 else (nsrc if dh <= 0 else f"{dsrc}／{nsrc}"))),
    }


def rate_gaps(person, contract=None, tol=RATE_TOL):
    """这个人身上**表上单价 vs 合同价**对不上的那几项，逐项列出来。

    ⚠ 只比这个人**真上过的班**（没上夜班就不谈他的夜班价），
    合同没登记的档**跳过**——那是「缺档」，另有判定，不能混成「不符」。
    「登记为 0」也是登记（锦绣保洁管理费就是 0），照比。

    工资与管理费**分开比**：16.5+2.5 与 17+2 合计都是 19，但在合同、入账、发票上是两条不同的线。
    这个函数是「单价对不对」的唯一实现——pay_check 的第①道和逐人核对页的标记都走它，免得两处判得不一样。
    """
    t = person.get("表上") or {}
    if not t:
        return []
    out = []
    for cn, shift in (("白班", "day"), ("夜班", "night")):
        if float(person.get(cn) or 0) <= 0:
            continue
        band = contract_band(person, shift, contract)
        if band is None:
            continue
        # ⚠ 「表上没有这一班的单价列」≠「表上单价是 0」。_pay_cols 允许部分命中（不是每个月的表都长一样），
        #    那时 t 里根本没有这两个键，取到 0 去跟合同价比会把全表误报成「单价不符」。
        #    判据与 actual_rate 一致：两项都 ≤0 就当这一班没有单价列，跳过。
        act = actual_rate(person, shift)
        if act is None:
            continue
        for i, label in enumerate(("工资单价", "管理费单价")):
            tv, cv = float(t.get(f"{cn}{label}") or 0.0), float(band[i])
            if abs(tv - cv) > tol:
                out.append({"项目": f"{cn}{label}", "表上": round(tv, 2),
                            "合同": round(cv, 2), "差": round(tv - cv, 2),
                            # 这一项差价 × 这个人这一班的工时 ＝ 单价算错造成的钱，**只归因于单价**，
                            # 不要拿「应付偏差」当它——那一项还含补贴奖罚之外的其它原因，且多付少付会相抵
                            "金额": round((tv - cv) * float(person.get(cn) or 0), 2)})
    return out


def rate_gap_amount(person):
    """这个人因为单价算错多付（正）／少付（负）了多少。逐项差价 × 该班工时，不掺别的原因。"""
    return round(sum(g.get("金额") or 0.0 for g in (person.get("单价不符") or [])), 2)


def pay_check(people, contract=None, tol=0.01):
    """对结算表**自己填的金额**做三道核对。工具只指出对不上，不改数——改数是人力和成本会计的事。

      ① 单价核对：表上单价 vs **合同价登记表**。成本会计 6 月底表就是这里错的（17+4=21，合同是 19）。
         合同价没登记的档跳过——没基准就没法核，缺档由「单价核对」那张表按格报出来。
      ② 金额核对：表上金额 vs 表上工时 × 表上单价。用**表上自己的单价**乘，专抓公式拖漏、手改数。
      ③ 勾稽核对：员工工资 ＝ 白工资+夜工资+补贴+奖+罚（罚在表里已是负数，直接加，别再减一次）；
                  合计 ＝ 员工工资 + 管理费。

    表里没有金额列（比如按派遣方拆出来的简表）就整段跳过，不报假异常。"""
    out = []
    for p in people:
        t = p.get("表上") or {}
        if not t:
            continue
        nm, ag = p["name"], p.get("agency", "")
        dh, nh = float(p.get("白班") or 0.0), float(p.get("夜班") or 0.0)
        def add(kind, item, tv, ev, note=""):
            out.append({"姓名": nm, "归属": ag, "岗位": post_of(p), "类型": kind,
                        "项目": item, "表上": round(tv, 2), "应为": round(ev, 2),
                        "差": round(tv - ev, 2), "说明": note})
        for g in rate_gaps(p, contract):        # 单价那一道用 RATE_TOL，与逐格比对同一把尺
            add("单价", g["项目"], g["表上"], g["合同"], "与合同价登记表不一致")
        for h, tag, up, key in ((dh, "白班", t.get("白班工资单价", 0.0), "白班工资"),
                                (nh, "夜班", t.get("夜班工资单价", 0.0), "夜班工资"),
                                (dh, "白班", t.get("白班管理费单价", 0.0), "白班管理费"),
                                (nh, "夜班", t.get("夜班管理费单价", 0.0), "夜班管理费")):
            if abs(t.get(key, 0.0) - h * up) > tol:
                add("金额", key, t.get(key, 0.0), h * up, f"应＝{tag}工时 {h:g} × 表上单价 {up:g}")
        w = (t.get("白班工资", 0.0) + t.get("夜班工资", 0.0)
             + t.get("补贴", 0.0) + t.get("奖", 0.0) + t.get("罚", 0.0))
        if abs(t.get("员工工资", 0.0) - w) > tol:
            add("勾稽", "员工工资", t.get("员工工资", 0.0), w, "应＝白工资+夜工资+补贴+奖+罚")
        a = t.get("员工工资", 0.0) + t.get("白班管理费", 0.0) + t.get("夜班管理费", 0.0)
        if abs(t.get("合计", 0.0) - a) > tol:
            add("勾稽", "合计", t.get("合计", 0.0), a, "应＝员工工资+管理费")
    return out


ADJ_ITEMS = (("补贴", "蒸练补贴", +1), ("奖", "奖", +1), ("罚", "罚", -1))
ADJ_RATIO_CAP = 0.20          # 单笔占该人当月工资的比例上限，超过就提请人工看一眼


def adjust_check(people, ratio_cap=ADJ_RATIO_CAP):
    """奖 / 罚 / 蒸练补贴 —— 合同外调整项的核对。

    ⚠ **先说工具做不到什么**：这三项在两张表里都没有第二个出处。工时有打卡可比、单价有合同可比，
    奖罚补贴只有结算表这一处孤证，金额对不对工具**验不了**，只能靠审批单/处罚通知。
    所以本函数不判金额对错，只做三件事：逐笔列出来、验符号、验占比。过了检查 ≠ 奖罚是对的。

    ① 符号：罚 ≤ 0，奖 ≥ 0，补贴 ≥ 0。
       「罚记作负数、直接加进员工工资」是**表自己的约定**，不是我拍的——全年 22 笔罚里 21 笔为负，
       唯一那笔正数（8 月 李秀英 +5）正是错的：本该扣 5，结果多发 5，一来一回差 10 元。
       金额虽小，但说明这个符号没人管，换成 −500 写成 +500 也一样漏得过去。
    ② 有调整无工资：当月工资为 0 却有奖金（实测 1 月 胡桂华 奖 60、工资 0）。
    ③ 占比：单笔金额占该人当月工资超过 ratio_cap（实测 1 月 桂丽 奖 75、工资 135，占 55.6%）。

    金额量级供参考：全年 12 个月合计 奖 1,360 元、罚 −230 元、蒸练补贴 0 元，
    对员工工资 629 万只占 0.025%。**蒸练补贴这一列建了一整年一次没用过**，该问一句还留不留。"""
    out = []
    for p in people:
        t = p.get("表上") or {}
        if not t:
            continue
        wage = float(t.get("员工工资") or 0.0)
        for key, label, sign in ADJ_ITEMS:
            v = float(t.get(key) or 0.0)
            if not v:
                continue
            row = {"姓名": p["name"], "归属": p.get("agency", ""), "部门": p.get("dept", ""),
                   "岗位": post_of(p), "项目": label, "金额": round(v, 2),
                   "当月工资": round(wage, 2),
                   "占工资": round(abs(v) / wage * 100, 1) if wage else None,
                   "级别": "提示", "说明": "合同外调整，工具验不了金额，需附审批依据"}
            if v * sign < 0:
                row["级别"] = "异常"
                row["说明"] = (f"「{label}」符号反了：本应{'≤' if sign < 0 else '≥'} 0，实际 {v:g}。"
                               + ("罚款记成正数＝不扣反发，一来一回差 " + f"{abs(v) * 2:g} 元"
                                  if sign < 0 else "奖金记成负数＝反而扣钱"))
            elif wage <= 0:
                row["级别"] = "异常"
                row["说明"] = f"当月工资为 0 却有「{label}」{v:g} 元——是漏了工时，还是这笔不该给？"
            elif ratio_cap and abs(v) / wage > ratio_cap:
                row["级别"] = "存疑"
                row["说明"] = (f"单笔占当月工资 {abs(v) / wage * 100:.1f}%，超过 {ratio_cap * 100:.0f}% 提示线，"
                               f"请核对审批依据")
            out.append(row)
    rank = {"异常": 0, "存疑": 1, "提示": 2}
    out.sort(key=lambda x: (rank[x["级别"]], -abs(x["金额"])))
    return out


def adjust_total(rows):
    """奖罚补贴的分项合计 + 异常笔数，给页面顶部一行摘要用。"""
    s = {"补贴": 0.0, "奖": 0.0, "罚": 0.0}
    for r in rows:
        for key, label, _ in ADJ_ITEMS:
            if r["项目"] == label:
                s[key] += r["金额"]
    return {"蒸练补贴": round(s["补贴"], 2), "奖": round(s["奖"], 2), "罚": round(s["罚"], 2),
            "净额": round(s["补贴"] + s["奖"] + s["罚"], 2), "笔数": len(rows),
            "异常": sum(1 for r in rows if r["级别"] == "异常"),
            "存疑": sum(1 for r in rows if r["级别"] == "存疑")}


def settle(people):
    """复核结论：每家派遣方、每条业务线该付多少工资、多少管理费。
    分组＝派遣方 × 部门（业务线）× 岗位——岗位不能并，锦绣的保洁 15 元和普工 19 元并一起就看不出单价了。
    同时给出派遣方小计和全表合计，好直接对着请款单逐家核。"""
    keyf = lambda x: (x["归属"] or "（空）", x["部门"] or "（空）", x["岗位"] or POST_DEFAULT)
    NUM = ["上报白班工时", "上报夜班工时", "上报总工时", "应付工资", "应付管理费", "应付合计",
           "补贴奖罚", "表上工资", "表上管理费", "表上合计", "应付偏差"]
    agg = {}
    for x in people:
        k = keyf(x)
        cell = agg.setdefault(k, dict.fromkeys(NUM, 0.0))
        cell["人数"] = cell.get("人数", 0) + 1
        cell["有表上金额"] = cell.get("有表上金额", 0) + (1 if x.get("表上合计") is not None else 0)
        cell["缺合同价人数"] = cell.get("缺合同价人数", 0) + (1 if x.get("合同缺档") else 0)
        for f in NUM:
            v = x.get(f)
            if isinstance(v, (int, float)):
                cell[f] += v
        for f in ("白班工资单价", "白班管理费单价", "夜班工资单价", "夜班管理费单价"):
            cell.setdefault(f, x.get(f))
            if cell[f] != x.get(f):
                cell[f] = None                    # 同一格里单价不唯一 → 留空，别造一个假单价出来
    rows = []
    for (ag, dept, post), v in sorted(agg.items()):
        r = {"归属": ag, "部门": dept, "岗位": post}
        r.update({k: (round(v[k], 2) if isinstance(v.get(k), (int, float)) else v.get(k)) for k in v})
        for f in ("表上工资", "表上管理费", "表上合计", "应付偏差"):
            r[f] = r[f] if v["有表上金额"] else None
        _blank_if_missing(r)
        rows.append(r)
    def roll(sel, label):
        keys = sorted({r[label] for r in rows})
        out = []
        for k in keys:
            grp = [r for r in rows if r[label] == k]
            c = {label: k, "人数": sum(r["人数"] for r in grp)}
            for f in NUM:
                c[f] = round(sum(r[f] or 0 for r in grp), 2)
            c["有表上金额"] = sum(r["有表上金额"] for r in grp)
            c["缺合同价人数"] = sum(r.get("缺合同价人数") or 0 for r in grp)
            _blank_if_missing(c)
            out.append(c)
        return out
    total = {"人数": sum(r["人数"] for r in rows), "有表上金额": sum(r["有表上金额"] for r in rows),
             "缺合同价人数": sum(r.get("缺合同价人数") or 0 for r in rows)}
    for f in NUM:
        total[f] = round(sum(r[f] or 0 for r in rows), 2)
    _blank_if_missing(total)
    return {"明细": rows, "派遣方小计": roll(None, "归属"), "业务线小计": roll(None, "部门"), "合计": total}


def _blank_if_missing(row):
    """这一格/这一组里只要有人的合同价没登记，按合同应付就不是一个完整的数——
    留空，别拿少算了一家的部分和冒充"合计"。
    同理，整组都没有结算表金额（简表）时，「表上」四项也留空——路由和看板用 is None 判「有没有请款金额」，
    小计/合计若给 0.0 会被当成"请款 0 元"。"""
    if row.get("缺合同价人数"):
        for f in ("应付工资", "应付管理费", "应付合计", "应付偏差"):
            row[f] = None
    if not row.get("有表上金额"):
        for f in ("表上工资", "表上管理费", "表上合计", "应付偏差"):
            row[f] = None


def outsiders(summary, punch, keyword="临时普工"):
    """打卡表里标了「临时普工」、本月有打卡、却不在结算名单上的人。只提示不下结论：
    打卡表对临时工的部门只写到派遣方，看不出车间，工具无从判断是别的车间还是漏人。"""
    listed = {x["key"] for x in summary["people"]}
    out = []
    for key, recs in punch["by_key"].items():
        if key in listed:
            continue
        # 逐行看，不能只看首行：同名多行时首行可能是正式工，只看它会漏掉真正的临时普工那一行
        for rec in recs:
            if keyword not in (rec.get("部门") or "") or not rec["days"]:
                continue
            thin = sum(1 for ts in rec["days"].values() if len(ts) <= 1)
            out.append({"姓名": rec["raw"], "考勤组": rec["组"],
                        "派遣方": (rec["部门"] or "").split("-")[-1],
                        "出勤天数": len(rec["days"]), "仅1-2次打卡天数": thin})
    out.sort(key=lambda x: -x["出勤天数"])
    return out


# ==================== 全年工资结构表（看板取数）====================
# 输入是《临工结构》里的「工资结构综合」页：一行 = 派遣公司 × 部门 × 性质(新/旧) × 类别(白班/夜班/餐补/其他补贴)，
# 横向每个月三列（工时 / 工时单价 / 工资金额）。补贴行有金额无工时，算有效单价时要排除。
_STRUCT_HEAD = ("劳务派遣公司", "部门", "性质", "类别")


def contract_vs_actual(people, contract, tol=RATE_TOL):
    """**合同价 vs 人力实际计价**，逐格比（派遣方 × 岗位 × 班次）。

    这是第③步「对比」页签最该回答的问题：**人力到底有没有按合同的价算。**
    - 合同价 ＝ **成本会计在「合同价登记表」登记的行**（按行带生效期，由路由层按期挑好传进来）
    - 人力实际 ＝ 结算表右侧金额区里，每个人身上真正套用的那几列单价

    ⚠ 合同价**绝不能取汇总表表头解析出来的那张**——表头那段计价规则是**人力自己写的**，
    拿它当合同价就成了「人力跟人力自己比」，永远一致，这个核对就白做了。
    （表头解析只原样带回页面和报告作参考，不参与任何计算；它是待核对的一方，不是基准。）

    两边都按「员工工资 + 管理费」两条分开比——**合计相等不代表没事**：
    16.5+2.5 与 17+2 合计都是 19，但工资与管理费在合同、入账、发票上是两条不同的线。

    状态四种：
      一致        两边都有且分毫不差
      ⚠不符      两边都有但对不上（这一档是真要查的）
      合同缺档    人力用了这一档，合同规则里没有 —— 依据不明
      本期无人    合同有这一档，本期没人套用 —— 不是问题，只是没用到
    """
    act = {}
    for p in people:
        t = p.get("表上") or {}
        if not t:
            continue
        agency, post = (p.get("agency") or "（空）"), post_of(p)
        for hkey, cn, wk, mk in (("白班", "白班", "白班工资单价", "白班管理费单价"),
                                 ("夜班", "夜班", "夜班工资单价", "夜班管理费单价")):
            if float(p.get(hkey) or 0) <= 0:
                continue                        # 这个人这一班没工时，谈不上「用了什么价」
            w, m = float(t.get(wk) or 0.0), float(t.get(mk) or 0.0)
            cell = act.setdefault((agency, post, cn), {"值": set(), "人数": 0, "工时": 0.0})
            cell["值"].add((round(w, 2), round(m, 2)))
            cell["人数"] += 1
            cell["工时"] += float(p.get(hkey) or 0)

    con = {}
    for a, posts in (contract or {}).items():
        if not isinstance(posts, dict):
            continue
        for post, band in posts.items():
            if not isinstance(band, dict):
                continue
            for shift, cn in (("day", "白班"), ("night", "夜班")):
                v = band.get(shift)
                if v:
                    con[(a, post, cn)] = (float(v[0]), float(v[1]))

    def money(w, m):
        return {"员工工资": w, "管理费": m, "合计": round(w + m, 2)}

    rows = []
    for k in sorted(set(con) | set(act)):
        agency, post, shift = k
        c, a = con.get(k), act.get(k)
        row = {"派遣方": agency, "岗位": post, "班次": shift,
               "合同": money(*c) if c else None,
               "人数": a["人数"] if a else 0,
               "工时": round(a["工时"], 1) if a else 0.0}
        if not a:
            row.update({"人力": None, "状态": "本期无人", "差额": None})
        elif len(a["值"]) > 1:
            # 同一格里人力用了不止一种价——这本身就是问题，但**不是「与合同价不符」**：
            # 它可能每一种都对得上合同（比如期中调过价），也可能都对不上。单独成一档，
            # 免得②按格报的「不符人数」把整格的人算进去、而④逐人一个也标不出（V2.349 审出）。
            vs = sorted(a["值"])
            row.update({"人力": money(*vs[0]), "状态": "⚠同格多价", "差额": None,
                        "说明": "同一格里人力用了 %d 种单价：%s——先查清为什么，再谈跟合同价对不对" %
                                (len(vs), "、".join(f"{w}+{m}" for w, m in vs))})
        else:
            w, m = next(iter(a["值"]))
            row["人力"] = money(w, m)
            if not c:
                row.update({"状态": "合同缺档", "差额": None,
                            "说明": "人力按 %s+%s 算了，但合同价登记表里没有覆盖本期的这一档，依据不明" % (w, m)})
            elif abs(w - c[0]) > tol or abs(m - c[1]) > tol:
                row.update({"状态": "⚠不符",
                            "差额": {"员工工资": round(w - c[0], 2), "管理费": round(m - c[1], 2),
                                     "合计": round((w + m) - (c[0] + c[1]), 2)}})
            else:
                row.update({"状态": "一致", "差额": None})
        rows.append(row)

    bad = [r for r in rows if r["状态"] not in ("一致", "本期无人") and r["人数"] > 0]
    return {"明细": rows,
            "有人力数据": bool(act),
            "一致": sum(1 for r in rows if r["状态"] == "一致"),
            "不符": sum(1 for r in rows if r["状态"] == "⚠不符"),
            "同格多价": sum(1 for r in rows if r["状态"] == "⚠同格多价"),
            "合同缺档": sum(1 for r in rows if r["状态"] == "合同缺档"),
            "本期无人": sum(1 for r in rows if r["状态"] == "本期无人"),
            "全对": not bad}


def settle_verdict(res, tol=0.01):
    """给复核结论表的每一行下一句结论：**正常 / 异常**，异常的写清楚为什么。

    需求方 2026-08-22：「这里应该有结论，正常还是异常」——原来只给一列「偏差」，
    要人自己看到 ¥0 再推断「那就是没问题吧」。**结论该由工具给，不该让人二次推断。**

    ⚠ 结论不能只看偏差。一格偏差为 0，可能同时存在：单价与合同不符、奖罚符号反了、
    某人某天超弹性多记、同名挂在两家派遣方——这些都是「这家这条线本期不干净」。
    所以把本期所有检查按 **归属 × 部门 × 岗位** 归位，逐格汇总。

    ⚠ 已认定（V2.335）的不算异常——人已经看过并说明过了，再报一次就是噪音。
    但认定过的会在原因里留一句「另有 N 项已认定」，不让它彻底消失。"""
    se = res.get("settle") or {}
    st = res.get("stats") or {}
    rows = se.get("明细") or []
    if not rows:
        return res

    # (姓名, 归属) → 这个人落在哪一格。**必须带归属**：同名跨派遣方（张博·广才 / 张博·锦绣）正是
    # 结算风险专门盯的场景，只按姓名定位会把华顺那条多记记到锦绣头上——有问题那家判正常、没问题那家判异常。
    where, where_n = {}, {}
    for p in res.get("people") or []:
        cell = (p.get("归属") or "（空）", p.get("部门") or "（空）", p.get("岗位") or POST_DEFAULT)
        where.setdefault((p.get("姓名"), cell[0]), cell)
        where_n.setdefault(p.get("姓名"), cell)
    issues, acked = {}, {}

    def locate(name, agency):
        if agency and (name, agency) in where:
            return where[(name, agency)]
        c = where_n.get(name)
        if c and (not agency or c[0] == agency):
            return c
        if agency:
            return (agency, None, None)           # 定位不到人时退到派遣方级
        return c

    def put(name, label, agency_only=None):
        cell = locate(name, agency_only)
        if not cell:
            return
        issues.setdefault(cell, {}).setdefault(label, 0)
        issues[cell][label] += 1

    def put_acked(name, agency_only=None):
        cell = locate(name, agency_only)
        if cell:
            acked[cell] = acked.get(cell, 0) + 1

    for x in st.get("金额核对") or []:
        if x.get("已认定"):
            put_acked(x.get("姓名"), x.get("归属"))
        else:
            put(x.get("姓名"), "单价/金额对不上", x.get("归属"))
    for x in st.get("合同外调整") or []:
        if x.get("级别") != "异常":
            continue
        if x.get("已认定"):
            put_acked(x.get("姓名"), x.get("归属"))
        else:
            put(x.get("姓名"), "奖罚异常", x.get("归属"))
    for r in res.get("rows") or []:
        if r.get("档") != "over_out":
            continue
        if r.get("已认定"):
            put_acked(r.get("姓名"), r.get("归属"))
        else:
            put(r.get("姓名"), "超弹性多记", r.get("归属"))
    for x in st.get("归属与打卡不符") or []:
        if x.get("已认定"):
            put_acked(x.get("姓名"), x.get("结算归属"))
        else:
            put(x.get("姓名"), "归属与打卡不符", x.get("结算归属"))
    for x in st.get("同名多行") or []:
        # 同名挂在两家派遣方：两家的格子都要记一笔（每格一次，别按 原名×派遣方 叠加）
        cells = {locate(nm, ag) for nm in (x.get("原名") or []) for ag in (x.get("派遣方") or [None])}
        for cell in cells:
            if not cell:
                continue
            if x.get("已认定"):
                acked[cell] = acked.get(cell, 0) + 1
            else:
                issues.setdefault(cell, {}).setdefault("同名重复计费存疑", 0)
                issues[cell]["同名重复计费存疑"] += 1

    def collect(keys):
        """把落在这些格里的问题并起来（派遣方级的条目用 (agency, None, None) 兜底）。"""
        bag, ak = {}, 0
        for k in keys:
            for lab, n in (issues.get(k) or {}).items():
                bag[lab] = bag.get(lab, 0) + n
            ak += acked.get(k, 0)
        return bag, ak

    def verdict(row, keys, dev):
        """三态：异常（有任何一项对不上）／ 待核（没发现问题，但合同价缺档，应付没法核）／ 正常。
        「待核」不能并进「正常」——没登记合同价就说正常，正是 V2.345 之前那个满屏绿色的错。"""
        bag, ak = collect(keys)
        why = []
        if dev is not None and abs(dev) > tol:
            why.append(f"结算与按合同价重算差 {dev:+,.2f} 元")
        why += [f"{lab} {n} 处" for lab, n in sorted(bag.items())]
        miss = int(row.get("缺合同价人数") or 0)
        row["结论"] = "异常" if why else ("待核" if miss else "正常")
        if miss:
            why.append(f"合同价缺档（{miss} 人），应付无法按合同价重算")
        row["异常原因"] = why
        if ak:
            row["已认定"] = ak
        return row["结论"]

    bad_agency, wait_agency = {}, {}
    for r in rows:
        a, d, po = r.get("归属"), r.get("部门"), r.get("岗位")
        keys = [(a, d, po)]
        v = verdict(r, keys, r.get("应付偏差"))
        if v == "异常":
            bad_agency[a] = True
        elif v == "待核":
            wait_agency[a] = True
        dev = r.get("应付偏差")
        if dev is not None and abs(dev) > tol:
            # 明细格的差异要记成一条问题，小计/合计才收得到——两格一正一负抵成 0，小计不能装作没事
            issues.setdefault((a, d, po), {}).setdefault("明细格结算与合同价有差", 0)
            issues[(a, d, po)]["明细格结算与合同价有差"] += 1

    for r in se.get("派遣方小计") or []:
        a = r.get("归属")
        keys = [k for k in issues if k[0] == a] + [(a, None, None)]
        verdict(r, list(dict.fromkeys(keys)), r.get("应付偏差"))
    for r in se.get("业务线小计") or []:
        d = r.get("部门")
        verdict(r, [k for k in issues if k[1] == d], r.get("应付偏差"))
    tot = se.get("合计")
    if tot is not None:
        verdict(tot, list(issues), tot.get("应付偏差"))
        tot["异常派遣方"] = sorted(bad_agency)
        tot["待核派遣方"] = sorted(a for a in wait_agency if a not in bad_agency)
    return res


# ==================== 看板数据源：从历次复核留档直接汇总 ====================
def board_from_periods(periods):
    """把「已核期次」的留档结果汇总成看板数据。

    ⚠ 看板的数据源是**每期复核的结果**，不是另外上传一张《临工结构》表（2026-08-22 定案）。
    理由：结构表得有人按月手工维护，维护的人一停看板就悄悄过期；而复核每月都要做，
    结果本来就在库里——**让看板吃已经产生的数，别再让人喂第二遍**。

    输出结构与 parse_structure 完全一致（months/depts/monthly/company/kpi/标准单价/残缺月），
    所以前端那五张图一行都不用改；只是多带一个「期次」段，标出每期的异常数与跑批信息。

    periods: [(月份, 该期 compute() 的完整结果)]，按月份升序。

    金额口径：**取结算表自己的金额（＝请款额）**，与复核结论页主列一致；
    该期没有金额列时才回落到按合同价重算的值——两者混在一张图里会看不出差别，故在「期次」里标明来源。
    白/夜班金额按各人各自的单价拆：白班工时 ×（白班工资单价＋白班管理费单价），夜班同理。"""
    periods = sorted(periods, key=lambda x: x[0])
    months = [m for m, _ in periods]
    monthly, dept_amt, comp_rows, price_w, status = [], {}, {}, {"白班": {}, "夜班": {}}, []

    for m, res in periods:
        people = res.get("people") or []
        st = res.get("stats") or {}
        tot = (res.get("settle") or {}).get("合计") or {}
        用表上 = tot.get("表上合计") is not None
        da = na = sa = dh = nh = 0.0
        unpriced_h = 0.0                       # 结算表没金额、合同价也缺的工时——钱拆不出来，不能静默按 0
        depts_m, comps_m = {}, {}
        for p in people:
            _dh = float(p.get("上报白班工时") or 0)
            _nh = float(p.get("上报夜班工时") or 0)
            # 拆白/夜金额用的单价：结算表上该人实际的价 → 合同价。两者都没有这一段就拆不出（记 0）
            _dp = float(p.get("表上白班单价") or 0) or (float(p.get("白班工资单价") or 0) + float(p.get("白班管理费单价") or 0))
            _np = float(p.get("表上夜班单价") or 0) or (float(p.get("夜班工资单价") or 0) + float(p.get("夜班管理费单价") or 0))
            _da, _na = _dh * _dp, _nh * _np
            if p.get("表上合计") is None:
                unpriced_h += (_dh if _dp <= 0 else 0.0) + (_nh if _np <= 0 else 0.0)
            _extra = float(p.get("补贴奖罚") or 0)
            da += _da; na += _na; sa += _extra; dh += _dh; nh += _nh
            # 这个人本期的钱：优先结算表自己的合计，没有就用工时×单价
            amt = p.get("表上合计")
            amt = float(amt) if amt is not None else _da + _na
            depts_m[p.get("部门") or "（空）"] = depts_m.get(p.get("部门") or "（空）", 0.0) + amt
            c = p.get("归属") or "（空）"
            cell = comps_m.setdefault(c, {"金额": 0.0, "白工时": 0.0, "夜工时": 0.0})
            cell["金额"] += amt; cell["白工时"] += _dh; cell["夜工时"] += _nh
            for cat, price, weight in (("白班", _dp, _da), ("夜班", _np, _na)):
                if price > 0 and weight > 0:
                    price_w[cat][round(price, 2)] = price_w[cat].get(round(price, 2), 0.0) + weight
        th = dh + nh
        total_amt = float(tot.get("表上合计") if 用表上 else (tot.get("应付合计") or 0)) or (da + na)
        for d, v in depts_m.items():
            dept_amt[d] = dept_amt.get(d, 0.0) + v
        for c, v in comps_m.items():
            comp_rows.setdefault(c, {})[m] = v
        monthly.append({
            "m": m, "白班金额": round(da), "夜班金额": round(na), "补贴金额": round(sa),
            "合计金额": round(total_amt), "白班工时": round(dh, 1), "夜班工时": round(nh, 1),
            "总工时": round(th, 1),
            "夜班工时占比": round(nh / th, 4) if th else 0,
            "有效单价": round(total_amt / th, 2) if th else 0,
            "部门": {d: round(v) for d, v in depts_m.items()},
        })
        # 异常不藏——月度看板把出过问题的月份盖掉，比看不到还糟
        status.append({
            "m": m, "人数": st.get("人数"), "比对人日": st.get("比对人日"),
            "异常多记日次": st.get("异常多记日次") or 0,
            "金额核对条数": st.get("金额核对条数") or 0,
            "奖罚异常": (st.get("合同外调整合计") or {}).get("异常") or 0,
            "同名跨派遣方": st.get("同名跨派遣方数") or 0,
            "金额来源": "结算表" if 用表上 else "按合同价重算",
            # 复核结论三态与合同价缺档也要上看板——整期「待核」的月份不能长得跟正常月一样
            "结论": tot.get("结论") or "—",
            "缺合同价人数": st.get("缺合同价人数") or 0,
            "未计价工时": round(unpriced_h, 1),
        })

    depts = sorted(dept_amt, key=lambda d: -dept_amt[d])
    for x in monthly:                                   # 每月补齐所有部门，缺的补 0，堆叠图才不会错位
        for d in depts:
            x["部门"].setdefault(d, 0)

    company = []
    for c, per_m in comp_rows.items():
        per = [round(per_m.get(m, {}).get("金额", 0)) for m in months]
        if not sum(per):
            continue
        h = sum(v["白工时"] + v["夜工时"] for v in per_m.values())
        nh_ = sum(v["夜工时"] for v in per_m.values())
        company.append({"c": c, "月": per, "合计": sum(per), "工时": round(h, 1),
                        "有效单价": round(sum(per) / h, 2) if h else 0,
                        "夜班占比": round(nh_ / h, 4) if h else 0,
                        "活跃月数": sum(1 for v in per if v > 0)})
    company.sort(key=lambda x: -x["合计"])

    std = {cat: (max(w, key=w.get) if w else 0.0) for cat, w in price_w.items()}
    ta = sum(x["合计金额"] for x in monthly)
    th = sum(x["总工时"] for x in monthly)
    peak = max(monthly, key=lambda x: x["合计金额"]) if monthly else {"m": "", "合计金额": 0}
    avg = ta / max(1, sum(1 for x in monthly if x["合计金额"] > 0))
    thin = [x["m"] for x in monthly if 0 < x["合计金额"] < avg * 0.25]
    return {
        "months": months, "depts": depts, "monthly": monthly, "company": company,
        "残缺月": thin, "标准单价": std, "期次": status, "来源": "留档",
        "kpi": {"全年金额": round(ta), "全年工时": round(th, 1),
                "有效单价": round(ta / th, 2) if th else 0,
                "夜班金额占比": round(sum(x["夜班金额"] for x in monthly) / ta, 4) if ta else 0,
                "夜班工时占比": round(sum(x["夜班工时"] for x in monthly) / th, 4) if th else 0,
                "头部": company[0]["c"] if company else "",
                "头部占比": round(company[0]["合计"] / ta, 4) if company and ta else 0,
                "派遣方家数": len(company), "峰值月": peak["m"], "峰值金额": peak["合计金额"],
                "期数": len(months)},
    }


def parse_structure(data, sheet=None):
    """→ {"months":[...], "monthly":[...], "company":[...], "kpi":{...}, "rows": 明细}
    月份分组取自表头上一行（每 3 列一个月）。找不到就按「1月…12月」顺排。"""
    if load_workbook is None:
        raise RuntimeError("缺少 openpyxl，无法解析结构表")
    wb = load_workbook(io.BytesIO(data), data_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else _pick_struct_sheet(wb)
    hdr = _find_header_row(ws, must=("部门", "类别"), scan=12)
    if not hdr:
        raise ValueError("结构表里找不到含「部门」「类别」的表头行，请确认上传的是《临工结构》的「工资结构综合」页")
    cols = {}
    for c in range(1, ws.max_column + 1):
        v = str(ws.cell(hdr, c).value or "").replace("\n", "").strip()
        if v in _STRUCT_HEAD and v not in cols:
            cols[v] = c
    first = max(cols.values()) + 1 if cols else 5
    months = []
    for c in range(first, ws.max_column + 1, 3):
        lab = str(ws.cell(hdr - 1, c).value or "").strip() or f"{len(months) + 1}月"
        months.append((lab, c))
        if len(months) >= 12:
            break
    if not months:
        raise ValueError("结构表里找不到按月分组的「工时/单价/金额」列")

    rows = []
    for r in range(hdr + 1, ws.max_row + 1):
        comp = ws.cell(r, cols.get("劳务派遣公司", 1)).value
        if not comp:
            continue
        rec = {"公司": str(comp).strip(),
               "部门": _txt(ws, r, cols.get("部门")), "性质": _txt(ws, r, cols.get("性质")),
               "类别": _txt(ws, r, cols.get("类别")), "月": {}}
        for lab, c in months:
            rec["月"][lab] = {
                "工时": float(ws.cell(r, c).value or 0),
                "单价": float(ws.cell(r, c + 1).value or 0),
                "金额": float(ws.cell(r, c + 2).value or 0)}
        rows.append(rec)
    return _struct_series(rows, [m[0] for m in months], ws.title, wb.sheetnames)


def _pick_struct_sheet(wb):
    for ws in wb.worksheets:
        if "结构" in ws.title and "综合" in ws.title:
            return ws
    for ws in wb.worksheets:
        if _find_header_row(ws, must=("部门", "类别"), scan=12):
            return ws
    return wb.worksheets[0]


def _struct_series(rows, months, sheet, sheets):
    def S(key, m, f=lambda r: True):
        return sum(r["月"][m][key] for r in rows if f(r))

    WORK = ("白班", "夜班")
    depts = sorted({r["部门"] for r in rows if r["部门"]},
                   key=lambda d: -sum(r["月"][m]["金额"] for r in rows if r["部门"] == d for m in months))
    monthly = []
    for m in months:
        da, na = S("金额", m, lambda r: r["类别"] == "白班"), S("金额", m, lambda r: r["类别"] == "夜班")
        sa = S("金额", m, lambda r: r["类别"] not in WORK)
        dh, nh = S("工时", m, lambda r: r["类别"] == "白班"), S("工时", m, lambda r: r["类别"] == "夜班")
        th = dh + nh
        row = {"m": m, "白班金额": round(da), "夜班金额": round(na), "补贴金额": round(sa),
               "合计金额": round(da + na + sa), "白班工时": round(dh, 1), "夜班工时": round(nh, 1),
               "总工时": round(th, 1),
               "夜班工时占比": round(nh / th, 4) if th else 0,
               "有效单价": round((da + na + sa) / th, 2) if th else 0, "部门": {}}
        for d in depts:
            row["部门"][d] = round(S("金额", m, lambda r, d=d: r["部门"] == d))
        monthly.append(row)

    comps = sorted({r["公司"] for r in rows},
                   key=lambda c: -sum(r["月"][m]["金额"] for r in rows if r["公司"] == c for m in months))
    company = []
    for c in comps:
        per = [round(S("金额", m, lambda r, c=c: r["公司"] == c)) for m in months]
        if not sum(per):
            continue                       # 全年为 0 的派遣方不入图，省得占一整行空白
        h = sum(r["月"][m]["工时"] for r in rows if r["公司"] == c and r["类别"] in WORK for m in months)
        a = sum(r["月"][m]["金额"] for r in rows if r["公司"] == c and r["类别"] in WORK for m in months)
        dh = sum(r["月"][m]["工时"] for r in rows if r["公司"] == c and r["类别"] == "白班" for m in months)
        nh = sum(r["月"][m]["工时"] for r in rows if r["公司"] == c and r["类别"] == "夜班" for m in months)
        company.append({"c": c, "月": per, "合计": sum(per), "工时": round(h, 1),
                        "有效单价": round(a / h, 2) if h else 0,
                        "夜班占比": round(nh / (dh + nh), 4) if dh + nh else 0,
                        "活跃月数": sum(1 for v in per if v > 0)})

    # 白班/夜班的标准单价直接从表里的「工时单价」列取众数——别在前端写死 19/22，
    # 那样调价之后理论线就悄悄错了。按金额加权取众数，避免个别加权行（如混了保洁的锦绣）当选。
    def _mode_price(cat):
        w = {}
        for r in rows:
            if r["类别"] != cat:
                continue
            for m in months:
                p_, a_ = r["月"][m]["单价"], r["月"][m]["金额"]
                if p_ > 0 and a_ > 0:
                    w[round(p_, 2)] = w.get(round(p_, 2), 0) + a_
        return max(w, key=w.get) if w else 0.0

    std = {"白班": _mode_price("白班"), "夜班": _mode_price("夜班")}
    ta = sum(x["合计金额"] for x in monthly)
    th = sum(x["总工时"] for x in monthly)
    peak = max(monthly, key=lambda x: x["合计金额"]) if monthly else {"m": "", "合计金额": 0}
    # 残缺月：金额不足全年月均的 25%——7 月只并了植物肉 40 人、小料整月为 0，图上必须标出来，
    # 否则谁看都以为那个月业务停了。
    avg = ta / max(1, sum(1 for x in monthly if x["合计金额"] > 0))
    thin = [x["m"] for x in monthly if 0 < x["合计金额"] < avg * 0.25]
    return {
        "months": months, "depts": depts, "monthly": monthly, "company": company,
        "sheet": sheet, "sheets": sheets, "残缺月": thin, "标准单价": std,
        "kpi": {"全年金额": round(ta), "全年工时": round(th, 1),
                "有效单价": round(ta / th, 2) if th else 0,
                "夜班金额占比": round(sum(x["夜班金额"] for x in monthly) / ta, 4) if ta else 0,
                "夜班工时占比": round(sum(x["夜班工时"] for x in monthly) / th, 4) if th else 0,
                "头部": company[0]["c"] if company else "",
                "头部占比": round(company[0]["合计"] / ta, 4) if company and ta else 0,
                "派遣方家数": len(company), "峰值月": peak["m"], "峰值金额": peak["合计金额"]},
    }

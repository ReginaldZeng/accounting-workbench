# -*- coding: utf-8 -*-
# [Change Log]
# Date: 2026-07-08 | Author: Claude / c | Version: V2.44
# Description: 物流对账内核（A 期第一增量，按需求确认书 v1.2 §6）。
#   账单解析（极鲜达 4-sheet 模板）→ 五信号匹配（S1 内部单剔除 / S2 收货地址线路 / S3 日期±2 /
#   S4 按客户计费重折算 / S5 客户别名表）→ 价格重算（零担/快运/山姆/首衡仓储，零尾差）→
#   9 态归因 → 三道勾稽（笔数/金额/覆盖率）→ 《差异清单》《单号回填版账单》导出。
#   铁律：宁标"待人工确认"不猜；规则未配置不硬算；别名未确认只给候选不自动锁定。
#   金蝶侧只读，取数走 kingdee_client.fetch_outbound_docs（四种出库/调拨单）。
import datetime
import json
import os
import re

EPOCH = datetime.date(1899, 12, 30)          # Excel 序列日期基准

# 9 态（需求确认书 v1.2 §6）
ST_PRICE = "多收·单价"
ST_SURCH = "多收·附加费"
ST_WEIGHT = "重量不符"
ST_QTY = "数量不符"
ST_BILL_ONLY = "账单单边"
ST_ERP_ONLY = "ERP单边"
ST_NO_RULE = "规则未配置"
ST_MANUAL = "待人工配对"
ST_OK = "核对一致"
ALL_STATES = [ST_PRICE, ST_SURCH, ST_WEIGHT, ST_QTY, ST_BILL_ONLY, ST_ERP_ONLY, ST_NO_RULE, ST_MANUAL, ST_OK]


def load_config(path):
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def to_date(v):
    """账单日期兼容：Excel 序列数 / datetime / 'YYYY-MM-DD' 字符串；解不出返回 None。"""
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    try:
        f = float(v)
        if f != f:                      # NaN
            return None
        if f > 40000:                   # Excel 序列
            return EPOCH + datetime.timedelta(days=int(f))
    except (TypeError, ValueError):
        pass
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", str(v))
    if m:
        return datetime.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    return None


def _num(v, default=0.0):
    try:
        f = float(v)
        return default if f != f else f
    except (TypeError, ValueError):
        return default


def _s(v):
    return "" if v is None else str(v).strip()


# ---------------- 账单解析（极鲜达模板：汇总 / 零担、快运 / 山姆 / 首衡外仓） ----------------

def parse_jxd_bill(path):
    """解析极鲜达月度对账单 → {'summary': 汇总数, 'rows': 标准化行, 'warnings': [...]}."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    out = {"summary": {}, "rows": [], "warnings": []}

    if "汇总" in wb.sheetnames:
        for row in wb["汇总"].iter_rows(min_row=2, max_row=8, values_only=True):
            if row and _s(row[0]) and "-" in _s(row[0]) and _num(row[5], None) is not None:
                out["summary"] = {"周期": _s(row[0]), "零担快运": _num(row[1]), "山姆": _num(row[2]),
                                  "首衡": _num(row[3]), "合计": _num(row[5])}
                break

    if "零担、快运" in wb.sheetnames:
        ws = wb["零担、快运"]
        for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            if not row or _s(row[0]) in ("", "合计"):
                continue
            r = {"sheet": "零担快运", "line": i, "origin": _s(row[0]), "biz": _s(row[1]),
                 "waybill": _s(row[2]), "cust": _s(row[3]), "dest": _s(row[4]),
                 "ship_date": to_date(row[5]), "svc": _s(row[6]), "pieces": _num(row[7]),
                 "weight": _num(row[8], None), "unit_price": _num(row[9], None),
                 "freight": _num(row[10]), "delivery": _num(row[11]), "pickup": _num(row[12]),
                 "unload": _num(row[13]), "upstairs": _num(row[14]), "other": _num(row[15]),
                 "total": _num(row[16]), "note": _s(row[17]) if len(row) > 17 else ""}
            r["inbound"] = ("孝感" in r["dest"]) and (r["origin"] not in ("孝感", ""))
            out["rows"].append(r)

    if "山姆" in wb.sheetnames:
        ws = wb["山姆"]
        for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            if not row or _s(row[0]) in ("", "合计"):
                continue
            out["rows"].append({
                "sheet": "山姆", "line": i, "origin": _s(row[0]), "waybill": _s(row[1]),
                "product": _s(row[2]), "dc": _s(row[3]), "ship_date": to_date(row[4]),
                "arrive_date": to_date(row[5]), "pieces": _num(row[6]),
                "pallets": _num(row[7], None), "weight": _num(row[8], None),
                "unit_price": _num(row[9], None), "trunk": _num(row[10]),
                "delivery": _num(row[11]), "unload": _num(row[12]), "storage": _num(row[13]),
                "total": _num(row[14]), "note": _s(row[15]) if len(row) > 15 else "",
                "backfill_no": _s(row[16]) if len(row) > 16 else "", "inbound": False})

    if "首衡外仓" in wb.sheetnames:
        ws = wb["首衡外仓"]
        for i, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
            if not row or (to_date(row[0]) is None and not _s(row[1])):
                continue
            out["rows"].append({
                "sheet": "首衡外仓", "line": i, "date": to_date(row[0]), "item": _s(row[1]),
                "in_tons": _num(row[7], None), "in_price": _num(row[8], None),
                "out_tons": _num(row[12], None), "out_price": _num(row[13], None),
                "handling": _num(row[14]), "storage_fee": _num(row[15]),
                "shuttle_type": _s(row[16]), "shuttle_n": _num(row[18], None),
                "shuttle_price": _num(row[19], None), "shuttle_fee": _num(row[20]),
                "total": _num(row[15]) + _num(row[20]),   # 仓储费用列+短驳运费列（=汇总页首衡口径）
                "inbound": False})
    return out


# ---------------- 诚煜账单解析（二期：单 sheet 扁平表，自带回填单号 C16） ----------------

_CY_HEADERS = {           # 表头文字 → 标准键（按文字定位列，容列序变动）
    "发车日期": "ship_date", "交货日期": "deliver_date", "客户": "cust",
    "货物名称": "product", "起运地": "origin", "目的地": "dest", "数量": "pieces",
    "重量/t": "weight_t", "重量": "weight_t", "体积m³": "volume", "体积": "volume",
    "干线单价": "unit_price", "干线运费": "freight", "提货费": "pickup",
    "送货费": "delivery", "操作费": "ops", "收入总计": "total", "备注": "note", "单号": "backfill_no",
}
_CY_NUM = {"pieces", "weight_t", "volume", "unit_price", "freight", "pickup", "delivery", "ops", "total"}
_CY_DATE = {"ship_date", "deliver_date"}


def parse_chengyu_bill(path):
    """解析诚煜物流月度账单（.xls 单表）→ {'summary','rows','warnings'}。

    按表头文字定位列（容列序变动）；合计行捕获作勾稽基准；单号列(收入总计后)=物流部门
    已回填的金蝶单据号（XSCKD 销售出库 / FBDC 分步式调出）。无金额或单号="6月结算"→ 待人工。
    """
    import xlrd
    wb = xlrd.open_workbook(file_contents=path) if isinstance(path, (bytes, bytearray)) else xlrd.open_workbook(path)
    sh = wb.sheet_by_index(0)
    out = {"summary": {}, "rows": [], "warnings": []}

    # 定位表头行（含"单号"和"收入总计"）
    hdr_row, colmap = None, {}
    for r in range(min(10, sh.nrows)):
        texts = {_s(sh.cell_value(r, c)): c for c in range(sh.ncols)}
        if "单号" in texts and ("收入总计" in texts or "客户" in texts):
            hdr_row = r
            for txt, c in texts.items():
                if txt in _CY_HEADERS:
                    colmap[_CY_HEADERS[txt]] = c
            break
    if hdr_row is None:
        out["warnings"].append("未找到表头行（缺'单号'/'收入总计'），非诚煜账单格式？")
        return out

    def cell(r, key):
        c = colmap.get(key)
        return sh.cell_value(r, c) if c is not None else None

    total_sum = 0.0
    for r in range(hdr_row + 1, sh.nrows):
        no = _s(cell(r, "backfill_no"))
        cust = _s(cell(r, "cust"))
        billed = _num(cell(r, "total"), None)
        # 合计行：客户空、单号空，但金额列有值
        if not cust and not no and billed is not None:
            out["summary"]["合计"] = billed
            continue
        if not cust and no in ("", "6月结算") and billed is None:
            continue                                  # 纯空行
        row = {"sheet": "诚煜", "line": r + 1, "backfill_no": no}
        for key in _CY_HEADERS.values():
            v = cell(r, key)
            row[key] = to_date(v) if key in _CY_DATE else (_num(v, None) if key in _CY_NUM else _s(v))
        # 待人工：无金蝶单号（"6月结算"占位或空）或无金额
        row["pending"] = (no in ("", "6月结算")) or (billed is None)
        if billed:
            total_sum += billed
        out["rows"].append(row)
    out["summary"]["明细求和"] = round(total_sum, 2)
    return out


# ---------------- 跨越物流账单解析（多主体分 sheet，自带回填单号 C17） ----------------

def parse_kuayue_bill(path):
    """解析跨越物流月结清单（.xlsx，一主体一 sheet）→ {'summary','rows','warnings'}。

    每个 sheet=一个付款主体（sheet 名=主体+应付总额）；表头行含"序号"；只认数字序号为明细行；
    C4=跨越运单号(KY)、C6=计费重量、C15=应付金额、C17=物流部门回填的金蝶单号(可一格多单/换行)、
    C18=类型(业务描述)。合计＝各 sheet 应付总额之和，供勾稽。
    """
    import openpyxl
    from io import BytesIO
    wb = openpyxl.load_workbook(BytesIO(path) if isinstance(path, (bytes, bytearray)) else path, data_only=True)
    out = {"summary": {}, "rows": [], "warnings": []}
    declared, row_sum = 0.0, 0.0
    for sn in wb.sheetnames:
        if sn.strip().lower() == "sheet1":
            continue
        ws = wb[sn]
        hdr = None
        for r in range(1, min(9, ws.max_row + 1)):
            if any("序号" in _s(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)):
                hdr = r
                break
        if hdr is None:
            out["warnings"].append(f"sheet「{sn}」找不到表头（缺『序号』），非跨越模板？")
            continue
        # 主体＝sheet 名去掉末尾金额（付款主体；R2「客户简称」是总客户深圳星期零、非主体，勿用）
        mname = re.match(r"^(.*?)([\d,]+\.\d{2})$", sn)
        subj = mname.group(1).strip() if mname else sn
        # 应付总额：R3「应付总额：X元」优先，否则 sheet 名末尾数字
        stotal = None
        for r in range(1, hdr):
            mt = re.search(r"应付总额[：:]\s*([\d,]+\.?\d*)", _s(ws.cell(r, 1).value))
            if mt:
                stotal = float(mt.group(1).replace(",", ""))
                break
        if stotal is None and mname:
            stotal = float(mname.group(2).replace(",", ""))
        if stotal:
            declared += stotal
        for r in range(hdr + 1, ws.max_row + 1):
            c1 = ws.cell(r, 1).value
            if not isinstance(c1, (int, float)):          # 只认数字序号=真明细，跳过合计/页脚文字
                continue
            no = _s(ws.cell(r, 17).value)
            billed = _num(ws.cell(r, 15).value, None)
            if billed:
                row_sum += billed
            # 主体＝C19「主体」列(深零/孝九/深九)；C2寄件公司在各 sheet 恒为深圳星期零、不能区分主体
            zhuti = _s(ws.cell(r, 19).value) or subj
            out["rows"].append({
                "sheet": "跨越", "主体": zhuti, "line": r, "seq": int(c1),
                "waybill": _s(ws.cell(r, 4).value), "寄件人": _s(ws.cell(r, 16).value),
                "pieces": _num(ws.cell(r, 5).value),
                "weight": _num(ws.cell(r, 6).value, None), "total": billed,
                "desc": _s(ws.cell(r, 18).value), "backfill_no": no, "cust": zhuti, "dest": "",
                "pending": (not no) or (billed is None)})
    out["summary"] = {"合计": round(declared, 2) if declared else None, "明细求和": round(row_sum, 2)}
    return out


# ---------------- 天鹰物流账单解析（多文件·多sheet；按吨数×18.5元/吨计费，单号向下填充） ----------------
# 表头文字 → 标准键（按子串匹配，容两种布局与括号差异）。天鹰每月拆多个文件（销售装货/分步调拨/
# 原辅料卸货退料…），文件数不定；两种列布局：装货明细含「收货仓+数量(箱)」、单号在后列；卸货明细为
# 「数量(KG)」。故一律按表头文字定位列，不写死列号。匹配顺序：先专一键，避免"数量"误吞"吨数"。
_TY_HEADER_RULES = [
    ("金蝶", "backfill_no"),     # 金蝶单据编号 / 金蝶入库单号
    ("金额", "total"),           # 含税金额（18.5元/吨）
    ("物料编码", "mat_code"),
    ("产品名称", "product"),
    ("规格", "spec"),
    ("收货仓", "dest"),
    ("物流", "truck"),           # 物流车信息 / 物流信息
    ("吨数", "tons"),
    ("数量", "qty"),             # 数量(箱)/数量(KG)/数量(千克)
    ("日期", "date"),            # 装货日期 / 卸货日期
    ("序号", "seq"),
    ("备注", "note"),
]
# 单号前缀 → 单据类型（供人看；均在 fetch_docs_by_nos 的 7 类取数内）
_TY_FORM = {"XSCKD": "销售出库单", "QTCK": "其他出库单", "FBDC": "分步式调出单",
            "CGRK": "采购入库单", "CGTL": "采购退料单"}


def _ty_colmap(ws, hdr_row):
    cm = {}
    for c in range(1, ws.max_column + 1):
        txt = _s(ws.cell(hdr_row, c).value)
        if not txt:
            continue
        for kw, key in _TY_HEADER_RULES:
            if kw in txt and key not in cm:
                cm[key] = c
                break
    return cm


def parse_tianying_bill(files):
    """解析天鹰物流月度费用明细（.xlsx，多文件多sheet）→ {'summary','rows','warnings'}。

    入参 files：单个(bytes/路径) 或 列表——天鹰每月账单拆成若干文件，文件数不定。逐文件逐 sheet：
    前 6 行内含"序号"+"金蝶"的行＝表头，按文字定位列；数字序号＝明细行，含"合计"的行＝勾稽基准。
    计费口径＝吨数×18.5元/吨＝含税金额（账单自洽，核价不依赖金蝶）。
    **单号向下填充**：无金蝶单号的续行并入上方最近有单号的行（同一单据拆行＝对账组，业务方口径）；
    填充后仍无单号的行才标 pending（真·待人工，无从直查金蝶）。有单号但无金额的行标 no_amount＝True——
    仍参与核量（组内重量要算全），只是不参与核价（V2.151：修此前"无金额即踢出核量"的错）。
    weight 存 **KG**（吨数×1000），与金蝶基本单位数量(kg)同量纲，供 reconcile 出毛净比/每公斤费用。
    **整表判重**：按 sheet 明细行指纹去重——账单里存在内容完全相同的重复 sheet（物流部门留"一份带图、
    一份不带图"），逐 sheet 全读会把同一批货算两遍；多文件上传后同一文件被选两次同理。重复的整表
    跳过并记入 `notices`（提示，不阻断解析；`warnings` 才是阻断级）。
    """
    import openpyxl
    from io import BytesIO
    if not isinstance(files, (list, tuple)):
        files = [files]
    out = {"summary": {}, "rows": [], "warnings": [], "notices": []}
    declared_amt, declared_tons, row_amt, row_tons, sheets_seen = 0.0, 0.0, 0.0, 0.0, 0
    seen_fp = {}          # 整表指纹 → 首次出现位置。实盘教训：账单里存在内容完全相同的重复 sheet
                          # （物流部门留「一份带图、一份不带图」），逐 sheet 全读会把同一批货算两遍——
                          # 金额与重量双双翻倍且合计行同步翻倍、勾稽照样"通过"，属静默错账。多文件上传后
                          # 还多一层风险：同一个文件被选中两次。故按明细行指纹去重，跳过并在 notices 明示。
    for fi, f in enumerate(files, start=1):
        wb = openpyxl.load_workbook(BytesIO(f) if isinstance(f, (bytes, bytearray)) else f, data_only=True)
        for sn in wb.sheetnames:
            ws = wb[sn]
            if ws.max_row < 2 or ws.max_column < 2:
                continue                                  # 空 sheet（如 Sheet1）
            hdr = None
            for r in range(1, min(7, ws.max_row + 1)):
                texts = [_s(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
                if any("序号" in t for t in texts) and any("金蝶" in t for t in texts):
                    hdr = r
                    break
            if hdr is None:
                continue                                  # 非天鹰明细 sheet，跳过不报错
            cm = _ty_colmap(ws, hdr)
            missing = [k for k in ("seq", "backfill_no", "total", "tons") if k not in cm]
            if missing:
                out["warnings"].append(f"sheet「{sn}」缺列{missing}，疑似天鹰模板被改")
                continue
            def cell(r, key):
                c = cm.get(key)
                return ws.cell(r, c).value if c is not None else None

            # 先把本 sheet 读进临时缓冲（含它自己的合计），指纹判重后再决定是否并入总账
            srows, sdecl_amt, sdecl_tons, samt, stons = [], 0.0, 0.0, 0.0, 0.0
            last_no = ""                                  # 单号向下填充：续行并入上一行单号
            for r in range(hdr + 1, ws.max_row + 1):
                rowtxt = [_s(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
                seqv = cell(r, "seq")
                raw_no = _s(cell(r, "backfill_no"))
                billed_p = _num(cell(r, "total"), None)
                tons_p = _num(cell(r, "tons"), None)
                # 行性质按【实质内容】判，不靠序号——实盘遇到有单号有金额却没填序号的真明细行
                # （被旧的"序号必须是数字"规则静默丢弃＝漏计费用），也遇到不写"合计"二字的合计行。
                has_id = bool(_s(cell(r, "product")) or _s(cell(r, "mat_code")) or raw_no)  # 有货物标识
                has_val = (tons_p is not None) or (billed_p is not None)
                if any("合计" in t for t in rowtxt) or (has_val and not has_id):
                    sdecl_amt += _num(cell(r, "total"))    # 合计行＝勾稽基准（显式"合计"或"有数无货"）
                    sdecl_tons += _num(cell(r, "tons"))
                    continue
                if not has_id or not (has_val or isinstance(seqv, (int, float))):
                    continue                              # 空行/页脚/说明行
                if raw_no:
                    last_no = raw_no
                no = raw_no or last_no                    # 续行继承最近单号
                billed, tons = billed_p, tons_p
                if billed:
                    samt += billed
                if tons:
                    stons += tons
                pref = re.match(r"^[A-Za-z]+", no)
                srows.append({
                    # 序号可能没填（不影响解析，仅作展示）；行号 line 才是定位账单行的依据
                    "sheet": "天鹰-" + sn.strip(), "line": r,
                    "seq": int(seqv) if isinstance(seqv, (int, float)) else None,
                    "mat_code": _s(cell(r, "mat_code")), "product": _s(cell(r, "product")),
                    "spec": _s(cell(r, "spec")), "pieces": _num(cell(r, "qty"), None),
                    "weight": round(tons * 1000, 3) if tons else None,   # 吨→KG，对齐金蝶基本单位数量
                    "tons": tons, "total": billed,
                    "dest": _s(cell(r, "dest")), "truck": _s(cell(r, "truck")),
                    "date": to_date(cell(r, "date")), "note": _s(cell(r, "note")),
                    "backfill_no": no, "backfill_raw": raw_no, "inherited": bool(not raw_no and last_no),
                    "cust": "", "主体": "", "waybill": "", "desc": _s(cell(r, "product")),
                    "form_hint": _TY_FORM.get(pref.group(0) if pref else "", ""),
                    "no_amount": billed is None,              # 有单号但无金额：参与核量、不参与核价
                    # pending＝无单号可直查金蝶（真·待人工）。有单号的行即便无金额也要进核量分组——
                    # 否则同一单据的拆行(如某行有单号无额)被踢出，组内重量少算，与金蝶对不上（V2.151 修）。
                    "pending": not no})

            if not srows:
                continue
            # 指纹＝本 sheet 全部明细行的(序号,单号,物料编码,吨数,金额)。同 sheet 内序号唯一，
            # 故只有"整表重复"才会撞指纹，不会误伤同表内合法的同料同重两笔。
            fp = tuple((x["seq"], x["backfill_no"], x["mat_code"], x["tons"], x["total"]) for x in srows)
            here = f"文件{fi}·{sn.strip()}"
            if fp in seen_fp:
                out["notices"].append(
                    f"「{here}」与「{seen_fp[fp]}」内容完全相同（{len(srows)} 行，"
                    f"¥{round(samt, 2)}／{round(stons, 3)}吨），已跳过不重复计算")
                continue
            seen_fp[fp] = here
            sheets_seen += 1
            out["rows"].extend(srows)
            declared_amt += sdecl_amt
            declared_tons += sdecl_tons
            row_amt += samt
            row_tons += stons
    if sheets_seen == 0:
        out["warnings"].append("未找到任何天鹰明细 sheet（每个 sheet 需含'序号'+'金蝶单号'表头）——非天鹰账单？")
    out["summary"] = {
        "合计": round(declared_amt, 2) if declared_amt else None,     # 供 tieout 勾稽（键名对齐诚煜/跨越）
        "明细求和": round(row_amt, 2),
        "合计吨数": round(declared_tons, 6) if declared_tons else None,
        "明细吨数求和": round(row_tons, 6)}
    # 吨数勾稽（金额勾稽由 reconcile 的 tieout 出）：账单自己的合计吨数与明细求和不符时提示。
    # 实盘遇到：某 sheet 合计行的吨数漏加了一行、金额却算了——账单内部矛盾，付钱按金额不受损，
    # 但要让人知道承运商的表有错，别默默吞掉。
    if declared_tons and abs(row_tons - declared_tons) > 0.001:
        out["notices"].append(
            f"账单吨数不自洽：明细求和 {round(row_tons, 3)} 吨 vs 账单合计行 {round(declared_tons, 3)} 吨"
            f"（差 {round(row_tons - declared_tons, 3)} 吨）——承运商合计行漏加或多加，请其核对")
    return out


# ---------------- 解析方案注册表（承运商 → 解析器；显式选承运商，不自动识别） ----------------
# 加新承运商＝在此注册一行；前端按此渲染上传按钮网格，报错也引用方案名，杜绝"没做方案 vs 模板被改"歧义。
PARSERS = {
    "诚煜物流": {"fn": parse_chengyu_bill, "name": "诚煜解析方案", "version": "v1", "backfill": True,
             "match": "诚煜",                # 与金蝶/税率维表全称匹配的关键词（全称含此词=此承运商）
             "format": "单 sheet 扁平表；关键列：单号(回填金蝶单据号)、收入总计、客户、起运地/目的地、数量",
             "docs": "XSCKD 销售出库单 / FBDC 分步式调出单", "qty_check": True},   # 账单数量=产品数量，可核量
    "跨越物流": {"fn": parse_kuayue_bill, "name": "跨越解析方案", "version": "v1", "backfill": True,
             "match": "跨越",
             "format": "多主体分 sheet；关键列：序号、单号(运单KY)、计费重量、应付金额、C17金蝶单号、类型",
             "docs": "QTCK 其他出库 / XSCKD 销售出库 / CGRK 采购入库 / CGTL 采购退料",
             "qty_check": False},   # 账单是件数(包裹数)≠产品数量，只核单号命中，不核量
    "天鹰物流": {"fn": parse_tianying_bill, "name": "天鹰解析方案", "version": "v1", "backfill": True,
             "match": "天鹰", "multi": True,     # multi＝一次可传多个账单文件（天鹰每月拆多文件，文件数不定）
             "format": "多文件·多sheet；按吨数×18.5元/吨计费(账单自洽)；单号向下填充续行；"
                       "两种布局(装货明细含收货仓/数量箱、卸货明细为数量KG)，按表头文字定位列",
             "docs": "XSCKD 销售出库 / QTCK 其他出库 / FBDC 分步式调出 / CGRK 采购入库 / CGTL 采购退料",
             "qty_check": False,    # 按吨计费，箱数≠金蝶数量；不核件数
             "mat_check": True},    # 走物料级核量（单号×物料编码 比 kg，金蝶为准）——见 reconcile_by_material
}


def list_parsers():
    """已做解析方案的承运商清单（供前端渲染上传按钮）。"""
    return [{"carrier": c, "name": p["name"], "version": p["version"],
             "format": p["format"], "docs": p.get("docs", ""),
             "multi": bool(p.get("multi"))} for c, p in PARSERS.items()]


_ACCR_KW = ("运费", "仓储费", "装卸", "搬运", "物流")
_ACCR_RE = re.compile(r"计提([一-龥A-Za-z（）()]+?(?:公司|中心|部|厂|物流|供应链))")


def accrued_by_carrier(voucher_rows):
    """从金蝶 2241 供应商往来凭证行汇总物流计提数：摘要含"计提"+物流费用词，按承运商全称累加贷方。"""
    acc = {}
    for r in voucher_rows:
        z = str(r.get("FEXPLANATION") or "")
        cr = r.get("FCREDIT") or 0
        if not cr or "计提" not in z or not any(k in z for k in _ACCR_KW):
            continue
        m = _ACCR_RE.search(z)
        if m:
            acc[m.group(1)] = acc.get(m.group(1), 0.0) + float(cr)
    return acc


def parse_bill(carrier, data):
    """按承运商显式选解析方案解析（不自动识别）。未注册→明确报"未做方案"；
    解析器报警(缺列等)→包成"按【X方案】解析失败…疑似模板被改"，指向模板而非方案缺失。

    data＝单个文件(bytes/路径) 或 文件列表。multi 方案(天鹰)吃整个列表；单文件方案(诚煜/跨越)
    只吃一个——传多个即明确报错，杜绝"多传了却只解析第一个"的静默漏账。"""
    p = PARSERS.get(carrier)
    if not p:
        raise ValueError(f"承运商「{carrier}」未做解析方案，无法解析（请选已支持的承运商）")
    files = list(data) if isinstance(data, (list, tuple)) else [data]
    if not files:
        raise ValueError("空文件（请选择物流账单 .xls/.xlsx）")
    if p.get("multi"):
        bill = p["fn"](files)
    else:
        if len(files) > 1:
            raise ValueError(f"【{p['name']}】只接受单个账单文件，本次却传了 {len(files)} 个——"
                             f"请每次只传一个（该承运商账单为单文件）")
        bill = p["fn"](files[0])
    if bill.get("warnings"):
        raise ValueError(f"按【{p['name']} {p['version']}】解析失败：{'；'.join(bill['warnings'])}"
                         f"——疑似账单模板被改。预期格式：{p['format']}")
    return bill


# ---------------- 二期：按回填单号直查金蝶比对（不启用匹配引擎） ----------------

def _agg_docs_by_no(docs):
    """金蝶分录级取数 → 按单号聚合到单据级：{单号: {form,往来,数量,kg,物料集,行数}}。"""
    agg = {}
    for d in docs:
        no = _s(d.get("单号"))
        if not no:
            continue
        a = agg.setdefault(no, {"form": d.get("form_name") or d.get("form"), "party": None,
                                "qty": 0.0, "kg": 0.0, "mats": set(), "lines": 0})
        a["party"] = a["party"] or _s(d.get("客户")) or _s(d.get("供应商")) or None
        a["qty"] += _num(d.get("数量"))
        a["kg"] += _num(d.get("kg"))
        if _s(d.get("物料")):
            a["mats"].add(_s(d.get("物料")))
        a["lines"] += 1
    return agg


def _split_nos(no):
    """拆一行多单号：'A+B+C' / 'A/B' / 'A，B' → ['A','B','C']。"""
    parts = re.split(r"[+/,，、;；\s]+", no or "")
    return [p for p in parts if p and p != "6月结算"]


def reconcile_by_backfill(bill, docs, qty_tol=1.0, qty_check=True):
    """二期核心：账单按回填单号直查金蝶 → 对账组核量 + 校验 → 汇总。不调匹配引擎。

    对账组＝按"单号集合"分组（一行多单号拆+号；一单号多行合并求量），组级比 账单数量 vs
    金蝶单据数量。三校验：单号查无 / 部分查无 / (金蝶无数量→待核)。返回 {'rows','stats','tieout'}。
    """
    agg = _agg_docs_by_no(docs)
    from collections import defaultdict
    groups, singles = defaultdict(list), []
    for r in bill["rows"]:
        nos = [] if r.get("pending") else _split_nos(r.get("backfill_no", ""))
        (singles if not nos else groups[tuple(sorted(set(nos)))]).append(r)

    def enrich(rs, found):
        parties = [agg[n].get("party") for n in found if agg[n].get("party")]
        mats = set()
        for n in found:
            mats |= agg[n].get("mats", set())
        mats = [m for m in mats if m]
        return {"主体": _s(rs[0].get("主体")) or _s(rs[0].get("cust")),
                "类型": _s(rs[0].get("desc")) or _s(rs[0].get("product")),
                "运单号": _s(rs[0].get("waybill")), "寄件人": _s(rs[0].get("寄件人")),
                "客户": parties[0] if parties else "",
                "货物": "、".join(mats[:2]) + ("…" if len(mats) > 2 else "")}

    rows, stats = [], {"组数": 0, "一致": 0, "单号命中": 0, "数量不符": 0, "单号查无": 0,
                       "部分查无": 0, "待核·金蝶无数量": 0, "待人工": 0}
    for r in singles:
        bw = _num(r.get("weight"))
        rows.append({"lines": str(r["line"]), "nos": r.get("backfill_no", ""), "state": ST_MANUAL,
                     "billed": r.get("total"), "bill_qty": r.get("pieces"),
                     "bill_wt": round(bw, 3) if bw else None, "gn_ratio": None,
                     "per_kg": round(_num(r.get("total")) / bw, 3) if bw else None,
                     "主体": _s(r.get("主体")) or _s(r.get("cust")),
                     "类型": _s(r.get("desc")) or _s(r.get("product")),
                     "运单号": _s(r.get("waybill")), "寄件人": _s(r.get("寄件人")), "客户": "", "货物": ""})
        stats["待人工"] += 1
    for noset, rs in sorted(groups.items()):
        stats["组数"] += 1
        bill_qty = sum(_num(x.get("pieces")) for x in rs)
        bill_wt = round(sum(_num(x.get("weight")) for x in rs), 3)
        bill_amt = round(sum(_num(x.get("total")) for x in rs), 2)
        found = [n for n in noset if n in agg]
        missing = [n for n in noset if n not in agg]
        kd_qty = sum(agg[n]["qty"] for n in found)
        kd_kg = sum(agg[n]["kg"] for n in found)
        forms = "/".join(sorted({agg[n]["form"] for n in found})) or "—"
        # 拆行明细：一行一个金蝶单号；账单费用按各单号金蝶重量(缺则数量)占比分摊
        wbase = sum(agg[n]["kg"] for n in found) or sum(agg[n]["qty"] for n in found)
        doc_details = []
        for n in found:
            b = agg[n]["kg"] or agg[n]["qty"]
            share = (b / wbase) if wbase else (1.0 / len(found) if found else 0)
            mats = [m for m in agg[n].get("mats", set()) if m]
            doc_details.append({"单号": n, "单据类型": agg[n]["form"], "客户": agg[n].get("party") or "",
                                "货物": "、".join(mats[:2]) + ("…" if len(mats) > 2 else ""),
                                "金蝶量": agg[n]["qty"], "金蝶重量": round(agg[n]["kg"], 3),
                                "费用分摊": round(bill_amt * share, 2)})
        for n in missing:
            doc_details.append({"单号": n, "单据类型": "查无", "客户": "", "货物": "",
                                "金蝶量": None, "金蝶重量": None, "费用分摊": None})
        res = {"lines": ",".join(str(x["line"]) for x in rs), "nos": "+".join(noset),
               "dest": rs[0].get("dest"), "bill_qty": bill_qty, "billed": bill_amt,
               "bill_wt": bill_wt or None, "kd_qty": kd_qty, "kd_kg": round(kd_kg, 3),
               "gn_ratio": round(bill_wt / kd_kg, 3) if (bill_wt and kd_kg) else None,   # 毛重净重比=计费重量/金蝶基本单位数量(KG)
               "per_kg": round(bill_amt / bill_wt, 3) if bill_wt else None,                # 每公斤费用=费用/计费重量
               "kd_forms": forms, "missing": "+".join(missing), "docs": doc_details}
        res.update(enrich(rs, found))
        if not found:
            res["state"] = "单号查无"; stats["单号查无"] += 1
        elif missing:
            res["state"] = "部分查无"; stats["部分查无"] += 1
        elif not qty_check:                            # 该承运商账单件数≠产品数量，只核单号命中、不核量
            res["state"] = "单号命中"; stats["单号命中"] += 1
        elif not kd_qty:
            res["state"] = "待核·金蝶无数量"; stats["待核·金蝶无数量"] += 1
        elif abs(bill_qty - kd_qty) > qty_tol:
            res["state"] = ST_QTY; stats["数量不符"] += 1
        else:
            res["state"] = ST_OK; stats["一致"] += 1
        rows.append(res)

    tie = {"账单合计": bill.get("summary", {}).get("合计"),
           "明细求和": bill.get("summary", {}).get("明细求和")}
    tie["勾稽"] = (tie["账单合计"] is not None and tie["明细求和"] is not None
                 and abs(tie["账单合计"] - tie["明细求和"]) < 0.01)
    return {"rows": rows, "stats": stats, "tieout": tie}


# ---------------- 物料级核量（V2.152：金蝶为准，不一致提醒人工复核） ----------------
# 为何要到物料级：按单号汇总整张单据核重量会大面积误报——①一个金蝶单号常含多个物料，而承运商
# 只拉其中子集（如 QTCK011172 金蝶12行1017kg、天鹰只拉4个物料370kg）；②部分物料金蝶基本单位
# 不是千克（印刷内袋＝Pcs、椒麻汁＝包），拿账单kg比件数无意义。故按 单号×物料编码 逐行比，
# 并先看基本单位能不能用重量核。

ST_MAT_OK = "核对一致"
ST_MAT_REVIEW = "需人工复核"          # 仅账单多报＝可能多付运费，要人去追
ST_MAT_UNDER = "账单少报·我方有利"     # 账单少报＝承运商少收我方钱，只标记留痕、不强提醒（业务方定）
ST_MAT_UNIT = "ERP非kg计量·重量无法核"
ST_MAT_NOMAT = "金蝶无此物料"
ST_MAT_NONO = "单号查无"
ST_MAT_MANUAL = "待人工配对"

_KG_UNITS = ("千克", "公斤", "KG", "Kg", "kg")     # 基本单位属这些才可用重量核


def _norm_code(v):
    """物料编码归一：Excel 数值单元格可能读成 305000001.0 / 3.05e8，一律回正为纯数字串。"""
    s = _s(v)
    if not s:
        return ""
    if s.endswith(".0"):
        s = s[:-2]
    return s


def _is_kg_unit(u):
    u = _s(u)
    return bool(u) and any(k.lower() == u.lower() for k in ("千克", "公斤", "kg"))


def reconcile_by_material(bill, docs, tol_kg=0.0, tol_pct=0.0):
    """物料级核量：按 单号×物料编码 比 账单重量(kg) vs 金蝶重量(kg)。金蝶为准。

    **容差默认 0**（业务方 2026-07-21 定：不一致就要提醒人工复核，含承运商吨数舍位造成的小差）。
    比对前把差异按**克级(3 位小数)四舍五入**——这不是容差，是避开二进制浮点噪声（0.0482×1000
    可能得 48.20000000000001，真·零容差会把纯计算尾巴也报成差异）。仍留 tol_kg/tol_pct 形参，
    日后要放宽改调用处即可，不动内核。
    超容差即落「需人工复核」并标方向（账单多报＝多收运费的风险 / 账单少报）。基本单位非千克的
    物料不判差异，落「ERP非kg计量」如实说明无法核。返回 {'rows','stats','tieout'}，一行＝一个 单号×物料。
    """
    from collections import defaultdict
    # 金蝶侧：(单号, 物料编码) → kg/数量/单位/名称
    kd = {}
    kd_by_no = defaultdict(list)
    for d in docs:
        no = _s(d.get("单号"))
        if not no:
            continue
        code = _norm_code(d.get("物料编码"))
        kd_by_no[no].append(d)
        k = (no, code)
        e = kd.setdefault(k, {"kg": 0.0, "qty": 0.0, "unit": _s(d.get("基本单位")),
                              "name": _s(d.get("物料")), "form": d.get("form_name") or d.get("form")})
        e["kg"] += _num(d.get("kg"))
        e["qty"] += _num(d.get("数量"))
        if not e["unit"]:
            e["unit"] = _s(d.get("基本单位"))

    # 账单侧：(单号, 物料编码) → 重量/金额/行号
    bl = {}
    manual = []
    for r in bill["rows"]:
        if r.get("pending"):
            manual.append(r)
            continue
        nos = _split_nos(r.get("backfill_no", ""))
        if not nos:
            manual.append(r)
            continue
        code = _norm_code(r.get("mat_code"))
        for no in nos:                                   # 一行多单号：重量按单号数均摊（天鹰实测未出现）
            k = (no, code)
            e = bl.setdefault(k, {"kg": 0.0, "amt": 0.0, "lines": [], "name": _s(r.get("product"))})
            e["kg"] += _num(r.get("weight")) / len(nos)
            e["amt"] += _num(r.get("total")) / len(nos)
            e["lines"].append(r.get("line"))

    rows, stats = [], {}

    def bump(st):
        stats[st] = stats.get(st, 0) + 1

    for (no, code), b in sorted(bl.items()):
        kdoc = kd.get((no, code))
        base = {"单号": no, "物料编码": code, "物料": b["name"], "账单kg": round(b["kg"], 3),
                "账单金额": round(b["amt"], 2), "行号": ",".join(str(x) for x in b["lines"]),
                "金蝶kg": None, "金蝶数量": None, "基本单位": "", "差异kg": None,
                "方向": "", "单据类型": ""}
        if no not in kd_by_no:                           # 整个单号金蝶查不到
            base["state"] = ST_MAT_NONO
        elif kdoc is None:                               # 单号在，但这张单里没有这个物料
            base["state"] = ST_MAT_NOMAT
            base["单据类型"] = (kd_by_no[no][0].get("form_name") or "")
        else:
            base.update({"金蝶kg": round(kdoc["kg"], 3), "金蝶数量": round(kdoc["qty"], 3),
                         "基本单位": kdoc["unit"], "单据类型": kdoc["form"]})
            if not _is_kg_unit(kdoc["unit"]):
                base["state"] = ST_MAT_UNIT              # 基本单位非千克→重量不可比，如实说明
            else:
                diff = round(b["kg"] - kdoc["kg"], 3)     # 克级取整＝避浮点噪声，非容差
                base["差异kg"] = diff
                tol = max(tol_kg, tol_pct * abs(kdoc["kg"]))
                if abs(diff) <= tol:
                    base["state"] = ST_MAT_OK
                elif diff > 0:                            # 账单 > 金蝶：可能多付运费，要人去追
                    base["state"] = ST_MAT_REVIEW
                    base["方向"] = "账单多报"
                else:                                     # 账单 < 金蝶：承运商少收我方钱，只标记
                    base["state"] = ST_MAT_UNDER
                    base["方向"] = "账单少报"
        bump(base["state"])
        rows.append(base)

    for r in manual:                                     # 无单号可直查
        rows.append({"单号": "", "物料编码": _norm_code(r.get("mat_code")), "物料": _s(r.get("product")),
                     "账单kg": _num(r.get("weight"), None), "账单金额": r.get("total"),
                     "行号": str(r.get("line")), "金蝶kg": None, "金蝶数量": None, "基本单位": "",
                     "差异kg": None, "方向": "", "单据类型": "", "state": ST_MAT_MANUAL})
        bump(ST_MAT_MANUAL)

    # 只在"可核"口径内汇总重量（非kg计量/查无 不计入，否则总比值又是假象）
    _CMP = (ST_MAT_OK, ST_MAT_REVIEW, ST_MAT_UNDER)
    cmp_bill = sum(r["账单kg"] for r in rows if r["state"] in _CMP)
    cmp_kd = sum(r["金蝶kg"] for r in rows if r["state"] in _CMP)
    tie = {"账单合计": bill.get("summary", {}).get("合计"),
           "明细求和": bill.get("summary", {}).get("明细求和"),
           "可核账单kg": round(cmp_bill, 3), "可核金蝶kg": round(cmp_kd, 3),
           "可核差异kg": round(cmp_bill - cmp_kd, 3)}
    tie["勾稽"] = (tie["账单合计"] is not None and tie["明细求和"] is not None
                 and abs(tie["账单合计"] - tie["明细求和"]) < 0.01)
    return {"rows": rows, "stats": stats, "tieout": tie}


# ---------------- 单位费用分析报表（从对账结果派生：单据×物流商→费用/元每KG） ----------------

def unit_cost_report(bill, docs, carrier):
    """金蝶单据锚定的物流成本明细：一行一张金蝶单据 → 承运商 / 费用 / 单位费用(元/KG)。

    账单费用按金蝶重量摊到它回填引用的每个单号（一行多单号=按重量分摊；无重量则均摊；
    一单号多账单行=费用累加）。单位费用 = 摊得费用 ÷ 金蝶重量。附单据类型小计。纯只读派生。
    """
    from collections import defaultdict
    agg = _agg_docs_by_no(docs)                         # 单号 → {form,party,qty,kg,...}
    fee_by_no = defaultdict(float)
    for r in bill.get("rows", []):
        if r.get("pending"):
            continue
        found = [n for n in _split_nos(r.get("backfill_no", "")) if n in agg]
        if not found:
            continue
        fee = _num(r.get("total"))
        wsum = sum(agg[n]["kg"] for n in found)
        for n in found:
            share = (agg[n]["kg"] / wsum) if wsum else (1.0 / len(found))
            fee_by_no[n] += fee * share

    docs_out = []
    for no, a in agg.items():
        fee = round(fee_by_no.get(no, 0.0), 2)
        kg = round(a["kg"], 3)
        docs_out.append({"单号": no, "单据类型": a["form"], "承运商": carrier,
                         "往来": a["party"], "数量": a["qty"], "重量KG": kg, "费用": fee,
                         "元每KG": round(fee / kg, 4) if kg else None})
    docs_out.sort(key=lambda d: (d["单据类型"], d["单号"]))

    by_form = defaultdict(lambda: {"单据数": 0, "费用": 0.0, "KG": 0.0, "缺重量": 0})
    for d in docs_out:
        b = by_form[d["单据类型"]]
        b["单据数"] += 1
        b["费用"] += d["费用"]
        b["KG"] += d["重量KG"]
        if not d["重量KG"]:
            b["缺重量"] += 1
    summary = [{"单据类型": k, "单据数": v["单据数"], "费用": round(v["费用"], 2),
                "KG": round(v["KG"], 3), "元每KG": round(v["费用"] / v["KG"], 4) if v["KG"] else None,
                "缺重量": v["缺重量"]} for k, v in sorted(by_form.items())]
    tot_fee = round(sum(s["费用"] for s in summary), 2)
    tot_kg = round(sum(s["KG"] for s in summary), 3)
    return {"carrier": carrier, "docs": docs_out, "summary": summary,
            "total": {"单据数": len(docs_out), "费用": tot_fee, "KG": tot_kg,
                      "元每KG": round(tot_fee / tot_kg, 4) if tot_kg else None}}


# ---------------- 匹配引擎（五信号，§6） ----------------

def prep_docs(docs, cfg):
    """S1 内部单剔除 + 单据级聚合。docs 为 kingdee_client.fetch_outbound_docs 的分录级输出。"""
    internal = tuple(cfg["internal_names"])
    agg = {}
    for r in docs:
        cust = r.get("客户") or ""
        d = agg.setdefault(r["单号"], {
            "no": r["单号"], "form": r.get("form", "SAL_OUTSTOCK"), "date": to_date(r["日期"]),
            "cust": cust, "addr": r.get("收货地址") or "", "kg": 0.0, "bags": 0.0, "mats": set(),
            "internal": any(t in cust for t in internal)})
        d["kg"] += _num(r.get("kg"))
        d["bags"] += _num(r.get("数量"))
        if r.get("物料"):
            d["mats"].add(r["物料"])
    return list(agg.values())


def _doc_cases(doc, cfg):
    """ERP 袋数 → 箱数（按物料箱规，仅当单据物料箱规全部已知）。"""
    sizes = cfg.get("case_size_bags", {})
    if doc["mats"] and all(m in sizes for m in doc["mats"]):
        size = sizes[next(iter(doc["mats"]))]
        return doc["bags"] / size if size else None
    return None


def _conv_factor(cust, cfg):
    c = cfg.get("conversions", {})
    rule = c.get(cust, c.get("_default", {"factor": 1.0}))
    return rule.get("factor", 1.0)


def match_row(row, docs, cfg):
    """给一行账单找出库单。返回 {'match': doc|None, 'cands': [...], 'method': str}."""
    win = cfg.get("date_window_days", 2)
    tol_abs, tol_pct = cfg.get("match_weight_tol_abs", 0.5), cfg.get("match_weight_tol_pct", 0.005)

    if row.get("inbound") or row["sheet"] == "首衡外仓":
        return {"match": None, "cands": [], "method": "入向/仓储·不进出向匹配"}
    d0 = row.get("ship_date")
    if d0 is None:
        return {"match": None, "cands": [], "method": "无发货日期"}

    if row["sheet"] == "山姆":
        # 山姆走全局分桶配对（match_sam_global），单行调用只兜底给候选
        city = cfg["sam"]["dc_city"].get(row["dc"])
        pool = [d for d in docs if not d["internal"] and d["cust"] == cfg["sam"]["customer"]
                and d["date"] and abs((d["date"] - d0).days) <= 5
                and (not city or city in d["addr"] or not d["addr"])]
        return {"match": None, "cands": pool[:5], "method": "山姆·多候选" if pool else "山姆·无候选"}

    # 零担快运：S2 线路 + S3 日期 + S4 折算重量 + S5 别名
    dest, w = row.get("dest") or "", row.get("weight")
    scored = []
    for d in docs:
        if d["internal"] or d["date"] is None or abs((d["date"] - d0).days) > win:
            continue
        sc, why = 0, []
        if dest and dest in d["addr"]:
            sc += 2
            why.append("线路")
        alias_hit, alias_conf = _alias_hit(row.get("cust", ""), d["cust"], cfg)
        if alias_hit:
            sc += 2 if alias_conf else 1
            why.append("别名" if alias_conf else "别名?")
        if w:
            f = _conv_factor(d["cust"], cfg)
            if abs(d["kg"] * f - w) <= max(tol_abs, w * tol_pct):
                sc += 2
                why.append("重量")
        if sc >= 2:
            scored.append((sc, d, "+".join(why)))
    scored.sort(key=lambda t: -t[0])
    if scored and (len(scored) == 1 or scored[0][0] > scored[1][0]) and scored[0][0] >= 3:
        return {"match": scored[0][1], "cands": [t[1] for t in scored[:5]], "method": scored[0][2]}

    # 对账组（1:N）：同客户多张出库单合计 ≈ 账单一行（样品/汇总发货场景）
    if w and scored:
        by_cust = {}
        for sc, d, _ in scored:
            by_cust.setdefault(d["cust"], []).append(d)
        for cust, ds in by_cust.items():
            if len(ds) < 2:
                continue
            f = _conv_factor(cust, cfg)
            if abs(sum(d["kg"] for d in ds) * f - w) <= max(tol_abs, w * tol_pct):
                return {"match": None, "group": ds, "cands": ds,
                        "method": f"对账组·{len(ds)}单合计重量"}
    return {"match": None, "cands": [t[1] for t in scored[:5]], "method": "多候选" if scored else "无候选"}


def match_sam_global(rows, docs, cfg):
    """山姆全局分桶配对：桶=(城市, 件数, 品规山姆/京东)，桶内账单行与出库单各按日期排序对位。

    依据：账单"发货时间"与金蝶单据日期有 ±2~4 天漂移，逐行就近会抢错单；
    同桶按时间顺序对位与物流部门人工回填逻辑一致（6月账单实证）。
    返回 {账单line: mres}。"""
    sam = cfg["sam"]
    cities = set(sam["dc_city"].values())

    def mat_type(mats):
        return "京东" if any("京东" in m for m in mats) else "山姆"

    buckets = {}
    for d in docs:
        if d["internal"] or d["cust"] != sam["customer"] or d["date"] is None:
            continue
        cases = _doc_cases(d, cfg)
        if cases is None:
            continue
        city = next((c for c in cities if c in d["addr"]), None)
        buckets.setdefault((city, round(cases, 2), mat_type(d["mats"])), []).append(d)
    for v in buckets.values():
        v.sort(key=lambda d: d["date"])

    rows_by = {}
    for r in rows:
        key = (sam["dc_city"].get(r["dc"]), round(r["pieces"], 2),
               "京东" if "京东" in r["dc"] else "山姆")
        rows_by.setdefault(key, []).append(r)

    out = {}
    for key, rs in rows_by.items():
        rs.sort(key=lambda r: r["ship_date"] or datetime.date.min)
        ds = buckets.get(key, [])
        if len(ds) == len(rs):                       # 桶内数量齐 → 按序对位
            for r, d in zip(rs, ds):
                out[r["line"]] = {"match": d, "cands": [d], "method": "山姆·城市+件数+顺序"}
        else:                                        # 不齐（窗口缓冲混入多余单据）→ 保序最小日期差对齐
            for r, d in zip(rs, _align_by_date(rs, ds)):
                if d is not None:
                    out[r["line"]] = {"match": d, "cands": [d], "method": "山姆·顺序对齐"}
                else:
                    out[r["line"]] = {"match": None, "cands": ds[:5],
                                      "method": "山姆·多候选" if ds else "山姆·无候选"}
    return out


def _align_by_date(rows, ds, max_gap=6):
    """保序对齐：rows（按日期升序）在 ds（按日期升序）里选同序子序列，总日期差最小。

    多批次同规格连续发货时，账单顺序=发货顺序，贪心就近会抢错，序列对齐不会。
    返回与 rows 等长的 doc 列表（无解位置为 None）。"""
    k, n = len(rows), len(ds)
    if k > n:
        return [None] * k
    INF = 10 ** 9

    def cost(r, d):
        if r["ship_date"] is None or d["date"] is None:
            return INF
        gap = abs((d["date"] - r["ship_date"]).days)
        return gap if gap <= max_gap else INF

    dp = [[INF] * (n + 1) for _ in range(k + 1)]
    dp[0] = [0] * (n + 1)
    for i in range(1, k + 1):
        for j in range(i, n + 1):
            skip = dp[i][j - 1]
            take = dp[i - 1][j - 1] + cost(rows[i - 1], ds[j - 1])
            dp[i][j] = min(skip, take)
    if dp[k][n] >= INF:
        return [None] * k
    picks, j = [None] * k, n
    for i in range(k, 0, -1):
        while j > i and dp[i][j] == dp[i][j - 1]:
            j -= 1
        picks[i - 1] = ds[j - 1]
        j -= 1
    return picks


def _alias_hit(bill_cust, erp_cust, cfg):
    """账单简称 ↔ 金蝶工商名。返回 (命中?, 已人工确认?)。"""
    for brand, a in cfg.get("aliases", {}).items():
        if brand in bill_cust and erp_cust in a.get("names", []):
            return True, bool(a.get("confirmed"))
    core = re.sub(r"^(上海|深圳|广州|北京|杭州|成都|东莞|南京|无锡|南通|武汉|天津|郑州|长沙|昆山|江门|佛山|西安|海口|新疆|山东|昌邑|乌鲁木齐)", "", bill_cust)
    return (len(core) >= 2 and core in erp_cust), False


# ---------------- 价格重算（零尾差） ----------------

def price_ltl(row, cfg):
    """零担/加急零担：干线=计重×城市价(+加急0.1) + 送货费按服务方式。价格卡仅覆盖孝感始发。"""
    ltl = cfg["ltl"]
    if row.get("origin") and row["origin"] != ltl.get("origin", "孝感"):
        return None                      # 非孝感始发（如山东工厂直发）= 一单一价，规则未配置
    rate = ltl["rates"].get(row["dest"])
    if rate is None:
        return None
    if row["biz"].startswith("加急"):
        rate = round(rate + ltl["urgent_uplift"], 4)
    fee = ltl["delivery_fee"].get(row["svc"])
    return {"rate": rate, "trunk": row["weight"] * rate if row["weight"] else None, "delivery": fee}


def price_express(row, cfg):
    """快运：max(最低收费, 计重×分档价)；送货费按服务方式。"""
    ex = cfg["express"]
    c = ex["cities"].get(row["dest"])
    if c is None or not row["weight"]:
        return None
    rate = None
    for ceil_kg, r in c["tiers"]:
        if ceil_kg is None or row["weight"] <= ceil_kg:
            rate = r
            break
    linear = row["weight"] * rate
    if row["biz"].startswith("加急"):
        linear = row["weight"] * (rate)          # 加急快运干线无书面加价规律，先按标准价
    fee = ex["delivery_fee"].get(row["svc"])
    if linear < c["min"]:
        return {"rate": None, "trunk": c["min"], "delivery": fee, "min_applied": True}
    return {"rate": rate, "trunk": linear, "delivery": fee, "min_applied": False}


def price_sam(row, cfg):
    """山姆：结算重量=板数×650（京东=件数×13）；干线=结算重量×(零担价+0.1)；武汉特例。"""
    sam, ltl = cfg["sam"], cfg["ltl"]
    city = sam["dc_city"].get(row["dc"])
    if city is None:
        return None
    is_jd = "京东" in row["dc"]
    if is_jd:
        exp_weight = row["pieces"] * sam["jd_case_kg"]
    elif row["pallets"]:
        exp_weight = row["pallets"] * sam["pallet_kg"]
    else:
        exp_weight = None
    if city == "武汉":
        return {"rate": None, "trunk": sam["wuhan"]["trunk"], "delivery": sam["wuhan"]["delivery_fee"],
                "exp_weight": exp_weight}
    base = ltl["rates"].get(city)
    if base is None:
        return None
    rate = round(base + sam["uplift_over_ltl"], 4)
    fee = sam["delivery_fee_urgent"] if "加急" in row.get("note", "") else sam["delivery_fee"]
    w = row["weight"] if row["weight"] is not None else exp_weight
    return {"rate": rate, "trunk": (w * rate) if w is not None else None, "delivery": fee, "exp_weight": exp_weight}


# ---------------- 逐行归因（9 态） ----------------

def classify_row(row, mres, cfg):
    """核价+核量 → 恰归一态。返回 {'state','reason','expect','diff', ...}。金额零尾差、重量零容差。"""
    tol = cfg.get("compare_tol_amount", 0.005)
    group = mres.get("group") or []
    out = {"state": ST_OK, "reason": "", "expect": None, "diff": 0.0,
           "match_no": mres["match"]["no"] if mres["match"] else "+".join(d["no"] for d in group),
           "method": mres["method"], "cands": [d["no"] for d in mres["cands"]]}

    # —— 天然人工项 ——
    if row.get("inbound"):
        out.update(state=ST_MANUAL, reason="入向·回程运费（对应采购入库，A期人工核）")
        return out
    if row["sheet"] == "首衡外仓":
        return _classify_warehouse(row, cfg, out)
    if row.get("ship_date") is None or _s(row.get("svc")) == "补送" or "补" in _s(row.get("biz")):
        out.update(state=ST_MANUAL, reason="补送/缺发货日期，账单信息不全")
        return out

    # —— 核价 ——
    if row["sheet"] == "山姆":
        p = price_sam(row, cfg)
    elif row.get("biz", "").endswith("零担"):
        p = price_ltl(row, cfg)
    elif row.get("biz", "").endswith("快运"):
        p = price_express(row, cfg)
    else:
        p = None
    if p is None:
        out.update(state=ST_NO_RULE, reason=f"价格卡无规则：{row.get('dest') or row.get('dc')} / {row.get('biz') or '山姆'}")
        return out

    if row["sheet"] == "山姆":
        billed_trunk, billed_fee = row["trunk"], row["delivery"]
    else:
        billed_trunk, billed_fee = row["freight"], row["delivery"]
    exp_trunk, exp_fee = p.get("trunk"), p.get("delivery")
    out["expect"] = exp_trunk

    if exp_trunk is not None and abs(billed_trunk - exp_trunk) > tol:
        out.update(state=ST_PRICE, diff=round(billed_trunk - exp_trunk, 4),
                   reason=f"干线运费：合同应收 {exp_trunk:.4f}，账单 {billed_trunk:.4f}"
                          + ("（少收）" if billed_trunk < exp_trunk else ""))
        return out
    if exp_fee is not None and billed_fee and abs(billed_fee - exp_fee) > tol:
        out.update(state=ST_SURCH, diff=round(billed_fee - exp_fee, 4),
                   reason=f"送货/入仓费：约定 {exp_fee}，账单 {billed_fee}")
        return out

    # —— 核量（需已匹配上 ERP；对账组按多单合计）——
    doc = mres["match"]
    if group and not doc:
        doc = {"no": out["match_no"], "cust": group[0]["cust"],
               "kg": sum(d["kg"] for d in group), "mats": set().union(*(d["mats"] for d in group))}
    if doc:
        if row["sheet"] == "山姆":
            cases = _doc_cases(doc, cfg)
            if cases is not None and abs(cases - row["pieces"]) > 0.01:
                out.update(state=ST_QTY, diff=row["pieces"] - cases,
                           reason=f"件数：账单 {row['pieces']:.0f} 箱 vs ERP {cases:.1f} 箱")
                return out
            ew = p.get("exp_weight")
            if ew is not None and row["weight"] is not None and abs(row["weight"] - ew) > tol:
                out.update(state=ST_WEIGHT, diff=round(row["weight"] - ew, 4),
                           reason=f"结算重量 {row['weight']} ≠ 板/箱折重 {ew}")
                return out
        elif row.get("weight"):
            w = row["weight"]
            if re.search(r"\d+\*\d+", row.get("note", "")) or round(w, 2) != round(w, 6):
                # 体积重计费行（备注带箱规，或计重小数位>2=体积折算特征）：
                # ERP 物料主数据无体积，无法独立重算 → 降级仅提示（R4）
                out["reason"] = "体积重计费（箱规折算），ERP 无体积数据，重量核对降级仅提示"
            else:
                f = _conv_factor(doc["cust"], cfg)
                conv = doc["kg"] * f
                if abs(conv - row["weight"]) > max(0.5, row["weight"] * cfg.get("compare_tol_weight", 0.005)):
                    out.update(state=ST_WEIGHT, diff=round(row["weight"] - conv, 4),
                               reason=f"账单计重 {row['weight']:.2f} vs ERP {doc['kg']:.2f}×{f}={conv:.2f}")
                    return out
    else:
        out.update(state=ST_MANUAL, reason=f"价格核对一致，但未锁定出库单（{mres['method']}）")
        return out
    out["reason"] = "价格与数量重量核对一致"
    return out


def _classify_warehouse(row, cfg, out):
    """首衡外仓：出/入库操作费按吨、短驳按车型一口价核对。"""
    wh, tol = cfg["warehouse"], cfg.get("compare_tol_amount", 0.005)
    diffs = []
    if row.get("out_price") is not None and abs(row["out_price"] - wh["out_op_per_ton"]) > tol:
        diffs.append((round((row["out_price"] - wh["out_op_per_ton"]) * (row.get("out_tons") or 0), 2),
                      f"出库操作费 {row['out_price']}/吨 vs 约定 {wh['out_op_per_ton']}/吨"))
    if row.get("in_price") is not None and abs(row["in_price"] - wh["in_op_per_ton"]) > tol:
        diffs.append((round((row["in_price"] - wh["in_op_per_ton"]) * (row.get("in_tons") or 0), 2),
                      f"入库操作费 {row['in_price']}/吨 vs 约定 {wh['in_op_per_ton']}/吨"))
    if row.get("shuttle_price") is not None and row.get("shuttle_type"):
        std = next((v for k, v in wh["shuttle_per_trip"].items() if k in row["shuttle_type"]), None)
        if std is not None and abs(row["shuttle_price"] - std) > tol:
            diffs.append((round((row["shuttle_price"] - std) * (row.get("shuttle_n") or 1), 2),
                          f"短驳 {row['shuttle_type']} {row['shuttle_price']}/车 vs 约定 {std}/车"))
    if diffs:
        out.update(state=ST_SURCH, diff=round(sum(d for d, _ in diffs), 2),
                   reason="；".join(r for _, r in diffs))
    else:
        out.update(state=ST_OK, reason="仓储费用与报价一致")
    return out


# ---------------- 主流程 + 三道勾稽 ----------------

def reconcile(bill, docs, cfg):
    """bill=parse_jxd_bill 输出；docs=fetch_outbound_docs 分录级。返回 {'rows','guards','stats'}。"""
    pool = prep_docs(docs, cfg)
    sam_matches = match_sam_global([r for r in bill["rows"] if r["sheet"] == "山姆"], pool, cfg)
    results = []
    for row in bill["rows"]:
        mres = sam_matches.get(row["line"]) if row["sheet"] == "山姆" else None
        if mres is None:
            mres = match_row(row, pool, cfg)
        cls = classify_row(row, mres, cfg)
        results.append({**row, **cls})

    # 勾稽一：笔数——每行恰归一态
    counts = {st: sum(1 for r in results if r["state"] == st) for st in ALL_STATES}
    g1 = sum(counts.values()) == len(results)

    # 勾稽二：金额——账单明细合计 ↔ 汇总页
    amt = {"零担快运": sum(r["total"] for r in results if r["sheet"] == "零担快运"),
           "山姆": sum(r["total"] for r in results if r["sheet"] == "山姆"),
           "首衡": sum(r.get("total") or 0 for r in results if r["sheet"] == "首衡外仓")}
    g2, g2_detail = True, {}
    for k, v in (bill.get("summary") or {}).items():
        if k in amt:
            ok = abs(round(amt[k], 2) - v) <= 0.01
            g2 = g2 and ok
            g2_detail[k] = {"明细": round(amt[k], 2), "汇总页": v, "ok": ok}

    # 勾稽三：覆盖率（按金额）
    total = sum(r.get("total") or 0 for r in results)
    auto = sum(r.get("total") or 0 for r in results if r["state"] not in (ST_MANUAL, ST_NO_RULE))
    manual = sum(r.get("total") or 0 for r in results if r["state"] in (ST_MANUAL, ST_NO_RULE))
    guards = {"笔数勾稽": g1, "金额勾稽": g2, "金额勾稽明细": g2_detail,
              "覆盖率": {"自动核金额": round(auto, 2), "人工核金额": round(manual, 2),
                        "自动核占比": round(auto / total, 4) if total else None}}
    matched = sum(1 for r in results if r["match_no"])
    stats = {"行数": len(results), "各态": counts, "已匹配出库单": matched,
             "差异合计": round(sum(r["diff"] for r in results
                                if r["state"] in (ST_PRICE, ST_SURCH)), 2)}
    return {"rows": results, "guards": guards, "stats": stats}


# ---------------- 成品导出 ----------------

def export_diff_excel(res, path, carrier="极鲜达", period=""):
    """《差异清单》：非一致行，按差异金额降序。"""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "差异清单"
    ws.append([f"{carrier} 差异清单 {period}（工具自动归因，供人工判断）"])
    ws.append(["sheet", "行", "状态", "客户/仓", "目的地", "发货日", "账单金额", "合同应收(干线)",
               "差额", "原因", "匹配出库单", "匹配方式", "备注"])
    rows = [r for r in res["rows"] if r["state"] != ST_OK]
    rows.sort(key=lambda r: -abs(r.get("diff") or 0))
    for r in rows:
        ws.append([r["sheet"], r["line"], r["state"], r.get("cust") or r.get("dc") or r.get("item"),
                   r.get("dest") or "", str(r.get("ship_date") or r.get("date") or ""),
                   r.get("total"), r.get("expect"), r.get("diff"), r["reason"],
                   r["match_no"] or "/".join(r["cands"][:3]), r["method"], r.get("note", "")])
    ws.append([])
    ws.append(["三道勾稽", json.dumps(res["guards"], ensure_ascii=False, default=str)])
    ws.append(["统计", json.dumps(res["stats"], ensure_ascii=False, default=str)])
    wb.save(path)
    return path


def export_backfill_excel(res, path, carrier="极鲜达", period=""):
    """《单号回填版账单》：明细行 + 回填出库单号/匹配方式/状态。"""
    import openpyxl
    wb = openpyxl.Workbook()
    for sheet in ("零担快运", "山姆"):
        ws = wb.create_sheet(sheet)
        ws.append([f"{carrier} {period} {sheet}（工具自动回填出库单号）"])
        ws.append(["行", "客户/仓", "目的地", "发货日", "件数", "计重/结算重量", "总运费",
                   "回填出库单号", "匹配方式", "候选", "状态"])
        for r in res["rows"]:
            if r["sheet"] != sheet:
                continue
            ws.append([r["line"], r.get("cust") or r.get("dc"), r.get("dest") or "",
                       str(r.get("ship_date") or ""), r.get("pieces"), r.get("weight"),
                       r.get("total"), r["match_no"], r["method"],
                       "/".join(r["cands"][:3]) if not r["match_no"] else "", r["state"]])
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    wb.save(path)
    return path

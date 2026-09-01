# -*- coding: utf-8 -*-
# [Change Log]
# Date: 2026-09-01 | Author: Claude / c | Version: V2.418
# Description: 服务器 RAR5 实测（8月真实包）：7z 无 Rar 码流、unrar-free 解不了 RAR5，unar 是唯一退路。
#              unar 改恒带 -p（空密码探测）：加密包未填密码时能被识别成"请填密码"而非通用失败天书。
# Date: 2026-09-01 | Author: Claude / c | Version: V2.416
# Description: 财资平台多份导出不再只认第一份——按【账户】整户择优并入（笔数最多、并列看收支合计）。
#              实证（2026-08 流水包）：出纳把财资流水导了三次（财资银行流水 1/2/3.xlsx），一次比一次全；
#              「1」只有 1 个账户且被 200 行导出上限截断，「3」才是 11 个账户的全集。旧的 seen_treasury
#              单例逻辑恰好只并入了最残的「1」，其余两份掉进"PDF/理财等跳过"，静默漏 10 个账户。
#              整户整取、绝不跨文件按行拼——按行去重会误杀同日同额的真实重复交易（如同额来回划转）。
# Date: 2026-08-06 | Author: Claude / c | Version: V2.199
# Description: 修中行HISQRY对方户名读错——开户行名称(含"公司")抢先被当对手。改按块内表头认列：
#              来账=付款人名称、往账=收款人名称（个人名也能取对）；无表头退回启发式并排除银行名。
# Date: 2026-08-03 | Author: Claude / c | Version: V2.167
# Description: 识别加内容兜底——文件名对不上的 xlsx/csv 再翻内容认（财资平台表头/中行HISQRY块头），
#              出纳随手改名后整包流水不再被静默标"跳过"；建行/花旗原本就按内容认，不动。
# Date: 2026-07-04 | Author: Claude / c | Version: V1.2
# Description: 银行流水导入/归一化（task①）。把核算组各家银行的原始导出解析成统一 schema，
#              喂逐笔稽核引擎(reconcile.bank_to_recs)。当前吃三类干净结构化源：
#                ① 财资平台"一键下载明细"  宁波+招商  (多账户分块 xlsx)
#                ② 中国银行 HISQRY          (多账户块 csv，来账=收/往账=支，金额带±号，自带借贷发生总额可自校验)
#                ③ 支付宝/微信/抖音          第三方支付(1012)——不进逐笔，只登记，走"渠道总额勾稽"(口径已定)
#              花旗/建行等 PDF 源本轮不解析（待核算组电子明细 / 单独适配）。
#              统一输出行 schema（对齐 sample_data.sample_bank_rows / reconcile.bank_to_recs）：
#                账号 / 户名 / 交易日期 / 摘要 / 对方户名 / 收入 / 支出 / 来源文件 / 银行
"""银行流水导入器 bank_import

用法：
    rows, manifest = load_bank_dir(r"C:\\...\\202606\\...\\6月流水")
    # rows: 逐笔稽核用的规范银行行(仅宁波/招商/中行等真实资金账户)
    # manifest: 每个文件 → 类型/账户数/笔数 或 跳过原因（透明可审计）
"""
from __future__ import annotations
import os
import re
import csv
import glob
import zipfile
import subprocess
import shutil as _shutil
import datetime

try:                                  # 花旗 PDF 对账单解析（V2.26）；未装 pymupdf 时降级为不识别
    from kernels import citi_pdf
except ImportError:                   # 扁平运行上下文（单测/脚本）兜底
    try:
        import citi_pdf
    except ImportError:
        citi_pdf = None

try:
    from openpyxl import load_workbook
except Exception:  # 引擎单测环境可能无 openpyxl；xlsx 解析用到时才需要
    load_workbook = None


# ----------------------------- 工具 -----------------------------
def _s(x) -> str:
    return "" if x is None else str(x).strip()


def _num(x) -> float:
    """'1,234.56' / '-464.90' / '+3.39' / '' → float。"""
    s = _s(x).replace(",", "").replace("¥", "").replace("￥", "")
    if s in ("", "-", "—"):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _is_date(v) -> bool:
    if isinstance(v, (datetime.datetime, datetime.date)):
        return True
    return bool(re.match(r"^20\d\d[-/.]?\d\d[-/.]?\d\d", _s(v)))


def _date_iso(v) -> str:
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d") if not isinstance(v, datetime.date) or isinstance(v, datetime.datetime) else v.isoformat()
    s = _s(v)[:10]
    m = re.match(r"^(20\d\d)[-/.]?(\d\d)[-/.]?(\d\d)", s)
    return f"{m.group(1)}-{m.group(2)}-{m.group(3)}" if m else s


def _row(acct, holder, d, memo, cp, inflow, outflow, src, bank, bal=None):
    return {"账号": _s(acct), "户名": _s(holder), "交易日期": _date_iso(d),
            "摘要": _s(memo), "对方户名": _s(cp),
            "收入": round(inflow, 2), "支出": round(outflow, 2),
            "余额": (_num(bal) if bal not in (None, "") else None),   # 账户余额(取末笔=期末，供余额调节)
            "来源文件": os.path.basename(str(src)), "银行": bank}


# ----------------------------- ① 财资平台 一键下载明细 -----------------------------
def parse_treasury(path):
    """财资平台"宁波银行+招商银行.xlsx"：多账户分块，每块 账号头 + 明细表。
    明细表头列固定：交易时间/本方户名/本方账号/收入/支出/账户余额/对方户名/对方账号/…/本方开户行。"""
    if load_workbook is None:
        raise RuntimeError("需要 openpyxl 才能解析财资平台 xlsx")
    ws = load_workbook(path, read_only=False).active
    rows = [[c.value for c in r] for r in ws.iter_rows()]
    # 定位明细表头列（各块一致，取首个）
    col = None
    for r in rows:
        names = [_s(c) for c in r]
        if "交易时间" in names and "本方账号" in names:
            col = {n: i for i, n in enumerate(names) if n}
            break
    if not col:
        return []
    def g(r, name):
        i = col.get(name)
        return r[i] if (i is not None and i < len(r)) else None
    out = []
    for r in rows:
        if not _is_date(g(r, "交易时间")):
            continue
        bank = _s(g(r, "本方开户行")) or ("宁波银行" if "宁波" in _s(g(r, "本方开户行")) else "")
        out.append(_row(
            acct=g(r, "本方账号"), holder=g(r, "本方户名"), d=g(r, "交易时间"),
            memo=_s(g(r, "交易备注")) or _s(g(r, "交易类型")),
            cp=g(r, "对方户名"),
            inflow=_num(g(r, "收入")), outflow=_num(g(r, "支出")),
            src=path, bank=bank or "宁波/招商", bal=g(r, "账户余额")))
    return out


# ----------------------------- ② 中国银行 HISQRY -----------------------------
_AMT_SIGNED = re.compile(r"^[+-][\d,]+\.\d{2}$")
_AMT_UNSIGNED = re.compile(r"^[\d,]+\.\d{2}$")
_DATE8 = re.compile(r"^20\d{6}$")


def parse_hisqry(path):
    """中行 HISQRY 多账户块 csv。每块：查询账号 / 总笔数 / 借贷发生总额 / 明细表头 / 明细行。
    明细行首列 来账=收 / 往账=支；金额取带±号的单元；日期取 8 位 20yymmdd；摘要用业务类型(第2列)。
    返回 (rows, blocks) —— blocks 供自校验：{账号,总笔数,借方发生总额,贷方发生总额,解析收,解析支,解析笔数}。"""
    text = None
    for enc in ("utf-8-sig", "gbk", "utf-16", "utf-8"):
        try:
            text = open(path, encoding=enc).read()
            break
        except Exception:
            continue
    if text is None:
        return [], []
    csv_rows = list(csv.reader(io_str(text)))
    out, blocks = [], []
    cur = None  # 当前账户块
    pay_i = payee_i = None   # V2.199 表头认列：付款人名称/收款人名称 所在列（每块重认）
    for r in csv_rows:
        if not r:
            continue
        head = _s(r[0])
        if any("付款人名称" in _s(c) for c in r):     # 明细表头行：记住对手户名两列的位置
            for i, c in enumerate(r):
                if "付款人名称" in _s(c):
                    pay_i = i
                if "收款人名称" in _s(c):
                    payee_i = i
            continue
        if "查询账号" in head:
            pay_i = payee_i = None
            cur = {"账号": re.sub(r"\s", "", _s(r[1])), "总笔数": None,
                   "借方发生总额": 0.0, "贷方发生总额": 0.0,
                   "解析收": 0.0, "解析支": 0.0, "解析笔数": 0}
            blocks.append(cur)
            continue
        if cur is None:
            continue
        if head.startswith("总笔数"):
            cur["总笔数"] = int(re.sub(r"\D", "", _s(r[1])) or 0)
            continue
        if head.startswith("借方发生总额"):
            cur["借方发生总额"] = _num(r[1]); continue
        if head.startswith("贷方发生总额"):
            cur["贷方发生总额"] = _num(r[1]); continue
        if head in ("来账", "往账"):
            cells = [_s(c) for c in r]
            ai = next((i for i, c in enumerate(cells) if _AMT_SIGNED.match(c)), -1)
            amt = _num(cells[ai]) if ai >= 0 else 0.0
            bal = next((c for c in cells[ai + 1:] if _AMT_UNSIGNED.match(c)), None) if ai >= 0 else None
            d8 = next((c for c in cells if _DATE8.match(c)), "")
            memo = cells[1] if len(cells) > 1 else ""     # 业务类型：结息/收费/大额支付…
            # V2.199 对方户名按表头列取：来账=付款人名称、往账=收款人名称（个人名也能取对）；
            # 无表头(老格式)退回启发式。此前"第一个含公司的格子"会被【付款人开户行名称】抢走——
            # 开户行名"中信银行股份有限公司东莞分行"也含"公司"（需求方实查：应为 东莞市绿邦实业有限公司）。
            ci_cp = pay_i if head == "来账" else payee_i
            cp = cells[ci_cp] if (ci_cp is not None and ci_cp < len(cells)) else ""
            if not cp:
                cp = _counterparty(cells, cur["账号"])
            inflow = abs(amt) if head == "来账" else 0.0   # 来账=收
            outflow = abs(amt) if head == "往账" else 0.0  # 往账=支
            out.append(_row(cur["账号"], "", d8, memo, cp, inflow, outflow, path, "中国银行", bal=bal))
            cur["解析收"] += inflow; cur["解析支"] += outflow; cur["解析笔数"] += 1
    return out, blocks


def _counterparty(cells, own_acct):
    """启发式兜底取对方户名（无表头老格式用）：含公司/有限的单元里挑最像交易对手的。
    V2.199：排除 银行/分行/支行/信用社——开户行名称"××银行股份有限公司××分行"也含"公司"，会抢先。"""
    cands = [c for c in cells if ("公司" in c or "有限" in c) and own_acct not in c]
    non_bank = [c for c in cands if not any(k in c for k in ("银行", "分行", "支行", "信用社"))]
    return (non_bank or cands or [""])[0]


def io_str(text):
    import io
    return io.StringIO(text)


# ----------------------------- ②.5 建设银行 host-to-host 明细 CSV -----------------------------
def is_ccb_csv(path):
    """建行 host-to-host 导出明细（gb18030，表头含 借方发生额（支取）/贷方发生额（收入）；文件名无规律，靠内容认）。"""
    if not str(path).lower().endswith(".csv"):
        return False
    try:
        with open(path, "rb") as f:
            head = f.read(600).decode("gb18030", "ignore")
    except Exception:
        return False
    return ("借方发生额" in head) and ("贷方发生额" in head) and ("账号" in head)


def parse_ccb_csv(path):
    """建行明细 CSV → (rows, report)。按表头名取列（列多且含个性化信息，靠名不靠位）。
    余额连续性自校验：同户按记账日期序，相邻余额差应=本笔净额（收−支）。"""
    with open(path, encoding="gb18030", errors="replace", newline="") as f:
        data = [r for r in csv.reader(f)]
    if len(data) < 2:
        return [], []
    hdr = [c.strip().strip("\t") for c in data[0]]
    def col(*names):
        for i, h in enumerate(hdr):
            if any(h.startswith(n) for n in names):
                return i
        return -1
    ci = {"acct": col("账号"), "name": col("账户名称"), "deb": col("借方发生额"),
          "cred": col("贷方发生额"), "bal": col("余额"), "cp": col("对方户名"),
          "memo": col("摘要"), "date": col("记账日期", "交易日期", "交易时间"), "ccy": col("币种")}
    def g(row, key):
        i = ci[key]
        return row[i].strip().strip("\t") if 0 <= i < len(row) else ""
    rows, byacct = [], {}
    for r in data[1:]:
        if not any(c.strip() for c in r):
            continue
        acct = g(r, "acct")
        if not acct:
            continue
        deb, cred = _num(g(r, "deb")), _num(g(r, "cred"))
        bal = _num(g(r, "bal")) if g(r, "bal") else None
        row = _row(acct, g(r, "name"), g(r, "date"), g(r, "memo"), g(r, "cp"), cred, deb, path, "建设银行", bal)
        row["币种"] = g(r, "ccy")
        rows.append(row)
        byacct.setdefault(acct, []).append((g(r, "date"), round(cred - deb, 2), bal))
    report = []
    for a, lst in byacct.items():
        lst.sort(key=lambda x: x[0])
        ok = True
        for i in range(1, len(lst)):
            if lst[i][2] is not None and lst[i - 1][2] is not None and abs((lst[i][2] - lst[i - 1][2]) - lst[i][1]) > 0.01:
                ok = False
                break
        report.append({"账号": a, "笔数": len(lst), "自校验": "OK" if ok else "余额不连续"})
    return rows, report


# ----------------------------- ③ 支付宝/电商渠道（第三方，1012，只做渠道级余额勾稽）-----------------------------
def _parse_alipay_stream(fileobj):
    """流式解析支付宝"账务组合查询"SpreadsheetML（文件 25MB+，不整体入内存、不留逐笔）。
    返回 {email, acct, 收, 支, 期末余额, 期末日期, 笔数}。行倒序，期末余额=最晚一笔的账户余额。"""
    import xml.etree.ElementTree as ET
    email = ""; acct = ""; idx = {}
    sin = sout = 0.0; cnt = 0; last_dt = ""; last_bal = None
    for _ev, elem in ET.iterparse(fileobj, events=("end",)):
        if elem.tag.split("}")[-1] != "Row":
            continue
        cells = []
        for c in elem:
            if c.tag.split("}")[-1] != "Cell":
                continue
            t = ""
            for d in c:
                if d.tag.split("}")[-1] == "Data":
                    t = d.text or ""; break
            cells.append(t)
        elem.clear()
        if not cells:
            continue
        c0 = _s(cells[0])
        if c0.startswith("#账号"):
            m = re.search(r"#账号[:：]\s*([^\[\]]+?)\s*(?:\[(\d+)\])?$", c0)
            if m:
                email = _s(m.group(1)); acct = m.group(2) or ""
            continue
        if c0 == "序号":
            idx = {name: i for i, name in enumerate(cells)}
            continue
        if not idx or not re.match(r"^\d+$", c0):
            continue
        def g(name):
            i = idx.get(name)
            return cells[i] if (i is not None and i < len(cells)) else ""
        sin += _num(g("收入（+元）")); sout += _num(g("支出（-元）")); cnt += 1
        dt = _s(g("入账时间"))
        if dt >= last_dt:
            last_dt = dt; b = g("账户余额（元）")
            last_bal = _num(b) if _s(b) else last_bal
    return {"email": email, "acct": acct, "收": round(sin, 2), "支": round(sout, 2),
            "期末余额": last_bal, "期末日期": last_dt[:10], "笔数": cnt}


def _read_broken_zip(path):
    """手工解压单文件 zip：部分支付宝 .xls.zip 的中文内名被错标 UTF-8 标志，zipfile 打开即崩。
    直接从本地文件头之后取 deflate 流解压（绕过文件名解码），返回内文件 bytes。"""
    import struct
    import zlib
    with open(path, "rb") as f:
        raw = f.read()
    if raw[:4] != b"PK\x03\x04":
        return None
    method = struct.unpack("<H", raw[8:10])[0]
    fn_len = struct.unpack("<H", raw[26:28])[0]
    ex_len = struct.unpack("<H", raw[28:30])[0]
    comp = raw[30 + fn_len + ex_len:]
    if method == 8:                       # deflate：decompressobj 会在流结束处自然停止
        return zlib.decompressobj(-15).decompress(comp)
    if method == 0:                       # stored
        return comp
    return None


def parse_alipay(path):
    """解析单个支付宝文件（.xls=SpreadsheetML / .xls.zip=内含同格式）。"""
    if path.lower().endswith(".zip"):
        try:
            with zipfile.ZipFile(path) as z:
                inner = [n for n in z.namelist() if n.lower().endswith((".xls", ".xml"))] or z.namelist()
                if not inner:
                    return None
                with z.open(inner[0]) as f:
                    return _parse_alipay_stream(f)
        except Exception:
            data = _read_broken_zip(path)     # 坏 zip 兜底
            if not data:
                return None
            import io as _io
            return _parse_alipay_stream(_io.BytesIO(data))
    with open(path, "rb") as f:
        return _parse_alipay_stream(f)


def parse_channels(folder):
    """扫描目录里的支付宝(2088*)文件，按查询段去重(同段的 .xls 与 .xls.zip 只取一次)，
    按账户合并(多日期段)→ 每渠道账户 收/支合计 + 期末余额。返回 list[dict]。"""
    files = []
    for root, _d, fnames in os.walk(folder):
        for fn in fnames:
            if fn.startswith("2088") and (fn.lower().endswith(".xls") or fn.lower().endswith(".xls.zip")):
                files.append(os.path.join(root, fn))
    # 去重：同一查询段(账户-日期-查询ID 前3段)优先 .xls，其次 .xls.zip
    by_seg = {}
    for p in files:
        base = os.path.basename(p)
        seg = "-".join(base.split("-")[:3])
        if seg not in by_seg or (by_seg[seg].lower().endswith(".zip") and not p.lower().endswith(".zip")):
            by_seg[seg] = p
    agg = {}
    for p in by_seg.values():
        try:
            r = parse_alipay(p)
        except Exception:
            r = None
        if not r:
            continue
        key = r["acct"] or r["email"]
        g = agg.setdefault(key, {"渠道": "支付宝", "email": r["email"], "acct": r["acct"],
                                 "收": 0.0, "支": 0.0, "期末余额": None, "期末日期": "", "笔数": 0, "文件段数": 0})
        g["收"] = round(g["收"] + r["收"], 2); g["支"] = round(g["支"] + r["支"], 2)
        g["笔数"] += r["笔数"]; g["文件段数"] += 1
        g["email"] = g["email"] or r["email"]
        if r["期末余额"] is not None and r["期末日期"] >= g["期末日期"]:
            g["期末日期"] = r["期末日期"]; g["期末余额"] = r["期末余额"]
    return list(agg.values())


# ----------------------------- 内容嗅探（改名兜底，V2.167）-----------------------------
def is_treasury_xlsx(path):
    """财资平台 xlsx 内容嗅探：前 50 行内出现「交易时间+本方账号」同行的明细表头即认
    （与 parse_treasury 定位表头的条件保持一致）。出纳转发常随手改名，文件名只当快路不当门槛。"""
    if load_workbook is None or not str(path).lower().endswith(".xlsx"):
        return False
    try:
        wb = load_workbook(path, read_only=True)
        try:
            ws = wb.active
            ws.reset_dimensions()   # 银行生成器常把尺寸错报成 A1:A1，read_only 会信以为真读到空——重算才见内容
            for r in ws.iter_rows(max_row=50, values_only=True):
                names = [_s(c) for c in r]
                if "交易时间" in names and "本方账号" in names:
                    return True
        finally:
            wb.close()
    except Exception:
        return False
    return False


def is_hisqry_csv(path):
    """中行 HISQRY 内容嗅探：头部含「查询账号」块头，且伴 来账/往账/总笔数 之一即认。
    「查询账号」是 HISQRY 独有块头；建行明细表头是「账号」不带"查询"，不会被误抢。"""
    if not str(path).lower().endswith(".csv"):
        return False
    try:
        with open(path, "rb") as f:
            head = f.read(4096)
    except Exception:
        return False
    for enc in ("utf-8-sig", "gbk", "utf-16"):
        t = head.decode(enc, "ignore")
        if "查询账号" in t and ("来账" in t or "往账" in t or "总笔数" in t):
            return True
    return False


# ----------------------------- 目录分派 -----------------------------
def _classify(fname):
    """先按文件名认（快路），认不出的 xlsx/csv 再翻内容认——
    改名只影响按名认的快路，内容在就不该整包漏账（V2.167）。"""
    n = os.path.basename(fname)
    low = n.lower()
    if n.startswith("宁波银行+招商银行") and low.endswith(".xlsx"):
        return "treasury"
    if n.startswith("HISQRY") and low.endswith(".csv"):
        return "hisqry"
    if n.startswith("2088") and (low.endswith(".xls") or low.endswith(".xls.zip")):
        return "alipay"          # 支付宝 → 1012，不进逐笔
    if n.startswith("1576978681"):
        return "wechat"          # 微信 → 1012
    if n.startswith("DL") and low.endswith(".csv"):
        return "douyin"          # 抖音 → 1012
    if low.endswith(".xlsx") and is_treasury_xlsx(fname):
        return "treasury"        # 改名的财资平台导出，按表头认回
    if low.endswith(".csv") and is_hisqry_csv(fname):
        return "hisqry"          # 改名的中行 HISQRY，按块头认回
    return "skip"


def _open_zip(zip_path, password):
    """返回 (打开的 zip 对象, 密码bytes或None)。
    有密码时优先用 pyzipper(同时支持 AES 与老式 ZipCrypto)；无 pyzipper 时退回标准库
    (仅能解老式 ZipCrypto，遇 AES 会在读取时报错，提示装 pyzipper)。无密码走标准库不变。"""
    if not password:
        return zipfile.ZipFile(zip_path), None
    pwd = password.encode("utf-8")
    try:
        import pyzipper
        return pyzipper.AESZipFile(zip_path), pwd
    except ImportError:
        return zipfile.ZipFile(zip_path), pwd


def extract_zip(zip_path, dest, password=None):
    """解压出纳流水包到 dest。支持带密码的加密包（password 非空时解密）。
    修正 Windows zip 中文文件名乱码（无 UTF-8 标志位时按 GBK 还原），
    否则文件名乱码会导致分类器认不出 宁波银行+招商银行.xlsx / HISQRY 等。返回 dest。"""
    os.makedirs(dest, exist_ok=True)
    z, pwd = _open_zip(zip_path, password)
    with z:
        if pwd is not None:
            z.setpassword(pwd)
        for info in z.infolist():
            name = info.filename
            if not (info.flag_bits & 0x800):        # 无 UTF-8 标志 → 多为 cp437 承载的 GBK
                for enc in ("gbk", "utf-8"):
                    try:
                        name = info.filename.encode("cp437").decode(enc)
                        break
                    except Exception:
                        continue
            name = name.replace("\\", "/")
            target = os.path.normpath(os.path.join(dest, name))
            if not target.startswith(os.path.normpath(dest)):
                continue                             # 防 zip 路径穿越
            if info.is_dir() or name.endswith("/"):
                os.makedirs(target, exist_ok=True)
                continue
            os.makedirs(os.path.dirname(target), exist_ok=True)
            with z.open(info) as src, open(target, "wb") as out:
                out.write(src.read())
    return dest


def _find_7z():
    """找 7-Zip 可执行文件（能解 rar/zip/7z/加密，跨平台首选）。找不到返回 None。"""
    for name in ("7z", "7za", "7zr", "7zz"):
        p = _shutil.which(name)
        if p:
            return p
    for p in (r"C:\Program Files\7-Zip\7z.exe", r"C:\Program Files (x86)\7-Zip\7z.exe",
              "/usr/bin/7z", "/usr/bin/7za", "/usr/lib/p7zip/7z", "/usr/local/bin/7z", "/usr/bin/7zz"):
        if os.path.exists(p):
            return p
    return None


def _find_unrar():
    """退路：unrar / unar（部分 Linux 只装了这个）。"""
    for name in ("unrar", "unar", "bsdtar"):
        p = _shutil.which(name)
        if p:
            return name, p
    return None


def _rar_cmd(name, exe, path, dest, password):
    if name == "7z":
        return [exe, "x", "-y", "-p" + (password or ""), "-o" + dest, path]
    if name == "unrar":
        return [exe, "x", "-y", "-p" + (password or "-"), path, dest + os.sep]
    if name == "unar":
        # 恒带 -p（无密码时传空串）：加密包+空密码 → unar 逐文件报 "Missing or wrong password"，
        # 上层据此识别"包已加密"并提示用户填密码；不带 -p 时 unar 的失败输出没有密码字样，
        # 会掉进天书般的通用失败（2026-09-01 服务器实测：7z 无 Rar 码流、unrar-free 解不了 RAR5，
        # unar 是唯一能解 RAR5 的退路，这条探测路径必须可靠）。空密码对未加密包无副作用。
        return [exe, "-f", "-o", dest, path, "-p", password or ""]
    return [exe, "-xf", path, "-C", dest] + (["--passphrase", password] if password else [])  # bsdtar


def _extract_rar(path, dest, password):
    """解 RAR 包到 dest。依次尝试 7z → unrar → unar → bsdtar，谁成算谁——
    因为服务器 7z 常不带 RAR 编解码（`7z i` 无 Rar），此时需退到 unrar 才能解。
    password 为空表示无密码；密码错立即抛；全部工具都失败才报错。stdin 关掉防交互卡住。"""
    os.makedirs(dest, exist_ok=True)
    tools = []
    seven = _find_7z()
    if seven:
        tools.append(("7z", seven))
    for name in ("unrar", "unar", "bsdtar"):   # 全部收进来逐个试（unrar-free 解不了RAR5时可退到 unar）
        exe = _shutil.which(name)
        if exe:
            tools.append((name, exe))
    if not tools:
        raise RuntimeError("NO_RAR_TOOL")
    errs = []
    for name, exe in tools:
        r = subprocess.run(_rar_cmd(name, exe, path, dest, password),
                           stdin=subprocess.DEVNULL, capture_output=True, text=True, errors="ignore")
        if r.returncode == 0:
            return dest
        out = ((r.stdout or "") + (r.stderr or "")).lower()
        if "wrong password" in out or "encrypted" in out or (password and "password" in out):
            raise RuntimeError("Wrong password for rar")
        errs.append(f"{name}:{((r.stderr or r.stdout or '').strip()[:120])}")
    raise RuntimeError("rar 解压失败（已试 " + "、".join(t[0] for t in tools) + "）：" + " | ".join(errs)[:240])


def _sniff_kind(path):
    """按文件头判断真实格式（上传统一存 upload.zip，不能只看扩展名）。"""
    try:
        with open(path, "rb") as f:
            head = f.read(8)
    except Exception:
        return "unknown"
    if head[:4] == b"PK\x03\x04" or head[:4] == b"PK\x05\x06":
        return "zip"
    if head[:6] == b"Rar!\x1a\x07":          # RAR4 / RAR5 同前缀
        return "rar"
    if head[:2] == b"7z" or head[:6] == b"7z\xbc\xaf\x27\x1c":
        return "7z"
    return "unknown"


def extract_archive(path, dest, password=None):
    """统一入口：按文件头识别 ZIP / RAR / 7z，自动选解压方式（都支持密码）。
    ZIP 走内置 pyzipper（零系统依赖）；RAR/7z 走 7-Zip / unrar（需系统工具）。返回 dest。"""
    kind = _sniff_kind(path)
    if kind == "zip":
        return extract_zip(path, dest, password=password)
    if kind in ("rar", "7z"):
        return _extract_rar(path, dest, password)
    # 兜底：先按 zip 试，不行再按 rar 试
    try:
        return extract_zip(path, dest, password=password)
    except Exception:
        return _extract_rar(path, dest, password)


def _merge_treasury(paths):
    """归并同一包里的多份财资平台导出。出纳会把财资流水导出多次（每次勾选范围不同，且单次导出
    有行数上限会截断），同一账户在不同文件里可能一份全、一份残——【每份都解析】，按【账户】整户择优：
    取笔数最多的那份（并列比收支合计；再并列取【整体更全的文件】——让"3 是全集"时页面直接显示
    "3 并入全部、1/2 重复"，与人的直觉一致；最后才看文件名排序），整户整取、
    绝不跨文件按行拼行——按行去重会误杀同日同额的真实重复交易（如同额来回划转）。
    让位的每一份还要过【逐笔查重】：其每笔都必须能在被采用的那份里找到（多重集包含），
    找不到的计入"疑漏笔数"——防"两次导出范围不同"时整户择优悄悄丢行，疑漏必须飘红给人看。
    返回 (rows, stats)。stats: {path: {"账户数","笔数","让位账户数","疑漏笔数"} 或 {"error": 原因}}。"""
    parsed, stats = [], {}
    for p in paths:
        try:
            by_acct = {}
            for r in parse_treasury(p):
                by_acct.setdefault(r["账号"], []).append(r)
            parsed.append((p, by_acct))
        except Exception as e:
            stats[p] = {"error": str(e)[:120]}
    totals = [sum(len(rs) for rs in by_acct.values()) for _p, by_acct in parsed]   # 各文件整体笔数
    best = {}   # 账号 -> (完整度, 文件序)；完整度 = (账户笔数, 账户收支合计, 文件整体笔数)
    for i, (_p, by_acct) in enumerate(parsed):
        for acct, rs in by_acct.items():
            score = (len(rs), sum(x["收入"] + x["支出"] for x in rs), totals[i])
            if acct not in best or score > best[acct][0]:
                best[acct] = (score, i)
    def _key(r):        # 逐笔查重键：除来源文件外的全部字段（连余额都相同才算同一笔）
        return (r["交易日期"], r["收入"], r["支出"], r["对方户名"], r["摘要"], r.get("余额"))

    from collections import Counter
    chosen_keys = {}    # 账号 -> Counter(被采用那份的逐笔键)
    for acct, (_s, i) in best.items():
        chosen_keys[acct] = Counter(_key(r) for r in parsed[i][1][acct])
    rows = []
    for i, (p, by_acct) in enumerate(parsed):
        take = [a for a in by_acct if best[a][1] == i]
        for a in take:
            rows += by_acct[a]
        missing = 0     # 让位账户逐笔查重：此文件独有、未被并入的笔数
        for a in by_acct:
            if best[a][1] != i:
                extra = Counter(_key(r) for r in by_acct[a]) - chosen_keys[a]
                missing += sum(extra.values())
        stats[p] = {"账户数": len(take), "笔数": sum(len(by_acct[a]) for a in take),
                    "让位账户数": len(by_acct) - len(take), "疑漏笔数": missing}
    return rows, stats


def needs_dup_confirm(manifest) -> bool:
    """财资归并出现"让位/重复/疑漏"时为 True——需求方定的责任口径（2026-09-01）：
    查重判定只是系统初核，必须弹窗经【人工确认】并留痕（确认人/时间），系统不独自背锅。"""
    for m in manifest or []:
        if str(m.get("类型", "")).startswith("财资平台"):
            if (not m.get("并入逐笔")) or m.get("说明"):
                return True
    return False


def load_bank_dir(folder, include_hisqry=True):
    """扫描目录(递归，兼容上传zip解压后的嵌套子目录)，解析可逐笔的结构化源(财资/中行)，
    返回 (rows, manifest)。财资平台允许多份导出，按账户取最全的并入（_merge_treasury）；
    支付宝/微信/抖音只登记不并入(走渠道总额)；其余(PDF/理财/余额表)跳过。"""
    rows, manifest = [], []
    all_files = []
    for root, _dirs, fnames in os.walk(folder):
        for fn in fnames:
            all_files.append(os.path.join(root, fn))
    all_files = sorted(all_files)
    kinds = {p: _classify(p) for p in all_files}
    t_rows, t_stats = _merge_treasury([p for p in all_files if kinds[p] == "treasury"])
    rows += t_rows
    for path in all_files:
        if os.path.isdir(path):
            continue
        kind = kinds[path]
        name = os.path.basename(path)
        try:
            if kind == "treasury":
                st = t_stats.get(path) or {}
                if "error" in st:
                    manifest.append({"文件": name, "类型": "解析失败", "并入逐笔": False,
                                     "说明": st["error"]})
                elif st.get("笔数"):
                    m = {"文件": name, "类型": "财资平台·宁波+招商", "并入逐笔": True,
                         "笔数": st["笔数"], "账户数": st["账户数"]}
                    if st.get("让位账户数"):
                        m["说明"] = ("还有 %d 个账户这份里也有，但别的文件里更全——那部分按更全的并入了"
                                     "（逐笔核对过，一笔没丢）" % st["让位账户数"])
                        if st.get("疑漏笔数"):
                            m["说明"] = ("还有 %d 个账户按别的文件并入；⚠ 但有 %d 笔只有这份里有、没被并入"
                                         "——两次导出范围可能不同，请人工核对是否漏账"
                                         % (st["让位账户数"], st["疑漏笔数"]))
                    manifest.append(m)
                else:
                    if st.get("疑漏笔数"):
                        note = ("⚠ 解析核对后发现：有 %d 笔只有这份里有、没被并入——两次导出范围可能不同，"
                                "请人工核对是否漏账" % st["疑漏笔数"])
                    else:
                        note = "这份里的每一笔，别的文件里都有（已逐笔核对）——为防同一笔记两遍，不再并入"
                    manifest.append({"文件": name, "类型": "财资平台·宁波+招商", "并入逐笔": False,
                                     "笔数": 0, "说明": note})
            elif kind == "hisqry" and include_hisqry:
                r, blocks = parse_hisqry(path)
                rows += r
                bad = [b for b in blocks if b["总笔数"] is not None and b["解析笔数"] != b["总笔数"]]
                manifest.append({"文件": name, "类型": "中行HISQRY", "并入逐笔": True,
                                 "笔数": len(r), "账户数": len(blocks),
                                 "自校验": ("笔数不符:" + str(bad)) if bad else "OK"})
            elif kind in ("alipay", "wechat", "douyin"):
                manifest.append({"文件": name, "类型": {"alipay": "支付宝", "wechat": "微信",
                                 "douyin": "抖音"}[kind], "并入逐笔": False, "笔数": "—",
                                 "说明": "第三方支付(1012)·走渠道总额，不逐笔"})
            elif is_ccb_csv(path):
                r, rep = parse_ccb_csv(path)         # 建行明细·余额连续性自校验
                rows += r
                chk = " / ".join(f"{a['账号']}·{a['自校验']}" for a in rep)
                manifest.append({"文件": name, "类型": "建设银行·明细CSV", "并入逐笔": True,
                                 "笔数": len(r), "账户数": len(rep), "自校验": chk})
            elif citi_pdf is not None and citi_pdf.is_citi_pdf(path):
                r, rep = citi_pdf.parse_citi_pdf(path)   # 逐户余额自校验，仅并入通过的账户
                rows += r
                chk = " / ".join(f"{a['账号']}·{a['币种']}·{a['自校验']}" for a in rep)
                bad = [a for a in rep if a["自校验"] != "OK"]
                manifest.append({"文件": name, "类型": "花旗银行·PDF对账单", "并入逐笔": True,
                                 "笔数": len(r), "账户数": len(rep),
                                 "自校验": (chk if not bad else "⚠ 有账户未过自校验·未并入：" + chk)})
            else:
                manifest.append({"文件": name, "类型": "跳过", "并入逐笔": False,
                                 "说明": "PDF/理财/余额/资产证明等，本轮不解析"})
        except Exception as e:
            manifest.append({"文件": name, "类型": "解析失败", "并入逐笔": False, "说明": str(e)[:120]})
    return rows, manifest

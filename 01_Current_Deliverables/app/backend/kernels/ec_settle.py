# -*- coding: utf-8 -*-
# [Change Log]
# Date: 2026-08-11 | Author: Claude / c | Version: V2.250
# Description: 【电商对账】收款核销内核（条目⑤一期，天猫渠道）。
#              规则全部来自 2026-05/07 两月真实底稿端到端复现（确认书⑤ §三~五）：
#              5 月 1,219 单核销+44 笔差异 100% 复现；7 月 39,212 单以 5 月规则零修改跑通。
#              流水解析按【列名】识别不按列号（支付渠道列 5 月有 7 月无——版式漂移实证）；
#              金额一律按分位取整后比对（确认书⑤ D11）。对金蝶只读。
"""
三层流水线（确认书⑤ §三）：
  ① 支付宝流水 → 按费目码归集 → 两资金账户（支付渠道=聚合结算渠道→聚合区，其余→支付宝区）
  ② 归集 → 凭证数字（费目→科目按 ec_fee_map；新费目码=待定科目，报警不静默）
  ③ 流水按订单归集平台收入(0010001−0020001) → 对金蝶应收(蓝Text6/红Text4双字段) → 差异分桶

五条会算错钱的规则（确认书⑤ §五）在此固化：
  红字字段左移双字段反查 / 应收期间闸(由取数层保证) / 三类剔除留痕 / 退款不退货=调节项 / 串单±配对
"""
import io
import json
from collections import defaultdict

import openpyxl


def _num(v):
    try:
        f = float(v or 0)
        return 0.0 if f != f else f
    except (TypeError, ValueError):
        return 0.0


def _r2(x):
    return round(x + 0.0, 2)


def _key(v):
    return str(v).strip().lstrip("'").strip() if v is not None else ""


def _code_of(desc):
    """业务描述 '0030003|软件服务费-…' → ('0030003','软件服务费-…')；无竖线返回 ('', 原文)。"""
    s = str(desc or "").strip()
    if "|" in s:
        c, _, label = s.partition("|")
        return c.strip(), label.strip()
    return "", s


# ---------------- 解析：支付宝流水（按列名识别，防版式漂移） ----------------
# 表头行特征：同一行内出现「入账时间」与「账务类型」。7 月单店 95,558 行——
# read_only 流式逐行，不整表载入。
_FLOW_COLS = {"入账时间": "ts", "支付宝交易号": "txn", "支付宝流水号": "serial",
              "商户订单号": "mch_no", "账务类型": "btype", "收入（+元）": "income",
              "支出（-元）": "outgo", "支付渠道": "chan", "商品名称": "goods",
              "对方名称": "peer", "业务基础订单号": "order_no", "业务描述": "desc"}


def _pick_sheet(wb, must_heads):
    """在整本工作簿里找含指定表头的工作表；多张命中取行数最多的那张。
    实测坑：wb.active 在春艳的多页工作簿里是「凭证」页——不能假设流水在首页。
    多店工作簿（星期零+Kiki 两张流水页）按行数取大的一张——**分店上传，一次跑一个店**。"""
    best, best_rows = None, -1
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=8, values_only=True):
            heads = {str(c).strip() for c in row if c is not None}
            if all(h in heads for h in must_heads):
                if ws.max_row > best_rows:
                    best, best_rows = ws.title, ws.max_row
                break
    return best


def parse_flow(data):
    """bytes → 流水行 list[dict]。找不到表头即报错（不猜列）。"""
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    title = _pick_sheet(wb, ("入账时间", "账务类型", "业务基础订单号"))
    if not title:
        wb.close()
        raise ValueError("没找到流水表头（应含「入账时间」「账务类型」「业务基础订单号」列）"
                         "——这不是支付宝流水导出？")
    ws = wb[title]
    idx, rows = None, []
    for row in ws.iter_rows(values_only=True):
        if idx is None:
            heads = {str(c).strip(): i for i, c in enumerate(row) if c is not None}
            if "入账时间" in heads and "账务类型" in heads and "业务基础订单号" in heads:
                idx = {k: heads[cn] for cn, k in _FLOW_COLS.items() if cn in heads}
                miss = [cn for cn in ("入账时间", "账务类型", "收入（+元）", "支出（-元）",
                                      "业务基础订单号", "业务描述") if cn not in heads]
                if miss:
                    raise ValueError("流水表缺少必需列：%s（版式变了，先人工确认）" % "、".join(miss))
            continue
        if row[idx["ts"]] is None:
            continue
        r = {k: (row[i] if i < len(row) else None) for k, i in idx.items()}
        rows.append(r)
    wb.close()
    return rows


# ---------------- 解析：支付宝 SpreadsheetML（银行对账流水包里的 2088* 原始文件，V2.255） ----------------
# 出纳每月给银行对账上传的流水包里就有支付宝「账务组合查询」全量文件（25MB+ 逐笔）——
# 本函数把它逐行读全（bank_import._parse_alipay_stream 只留渠道汇总，这里要喂核销引擎）。
def parse_flow_sml(fileobj):
    """SpreadsheetML 文件对象 → (流水行 list[dict]（与 parse_flow 同构）, 账号 acct, 邮箱 email)。"""
    import xml.etree.ElementTree as ET
    email = ""
    acct = ""
    idx = None
    rows = []
    for _ev, elem in ET.iterparse(fileobj, events=("end",)):
        if elem.tag.split("}")[-1] != "Row":
            continue
        cells = []
        for c in elem:
            if c.tag.split("}")[-1] != "Cell":
                continue
            t = ""
            for d in c:
                if d.tag.split("}")[-1] == "Data":
                    t = d.text or ""
                    break
            cells.append(t)
        elem.clear()
        if not cells:
            continue
        c0 = str(cells[0]).strip()
        if c0.startswith("#账号"):
            import re
            m = re.search(r"#账号[:：]\s*([^\[\]]+?)\s*(?:\[(\d+)\])?$", c0)
            if m:
                email = m.group(1).strip()
                acct = m.group(2) or ""
            continue
        if c0 == "序号":
            heads = {str(h).strip(): i for i, h in enumerate(cells)}
            idx = {k: heads[cn] for cn, k in _FLOW_COLS.items() if cn in heads}
            continue
        if idx is None or not c0.isdigit():
            continue
        rows.append({k: (cells[i] if i < len(cells) else None) for k, i in idx.items()})
    return rows, acct, email


def parse_flow_any(data):
    """bytes → 流水行。自动识别 xlsx(PK) / zip内SpreadsheetML / 裸SpreadsheetML(.xls)。"""
    import io as _io
    if data[:2] == b"PK":
        try:
            return parse_flow(data)                          # 先按 xlsx 试
        except Exception:
            import zipfile
            try:                                             # 再按 zip 内含 .xls(SpreadsheetML) 试
                with zipfile.ZipFile(_io.BytesIO(data)) as z:
                    inner = z.read(z.namelist()[0])
            except Exception:                                # 支付宝坏 zip（中文内名错标）→ 复用银行对账的手工解压
                from kernels import bank_import as _bi
                import tempfile, os as _os
                with tempfile.NamedTemporaryFile(delete=False, suffix=".zip") as tf:
                    tf.write(data)
                try:
                    inner = _bi._read_broken_zip(tf.name)
                finally:
                    _os.unlink(tf.name)
            rows, _a, _e = parse_flow_sml(_io.BytesIO(inner))
            return rows
    if b"<?xml" in data[:200] or b"Workbook" in data[:400]:
        rows, _a, _e = parse_flow_sml(_io.BytesIO(data))
        return rows
    return parse_flow(data)


# ---------------- 文件自动识别（V2.258：多文件拖入，按表头认类型——不猜不静默） ----------------
# 会计把旺店通/卖家中心导出的文件全选拖入，工具按表头特征归类；认不出的单独列出问人。
_SNIFF_RULES = [
    ("退款不退货", {"原始单号", "分摊退款金额"}),
    ("销售退货", {"入库单号", "入库量", "入库总额"}),
    ("销售出库", {"出库单编号", "货品成交总价"}),
    # ④平台数据是多文件类别（需求方定四通道分类学：①外部资金账户 ②金蝶 ③旺店通 ④平台导出）：
    # 订单导出(必) + 退款明细(推荐,货物状态列=未发货退款解释器) + 价保赔付(可选)
    # + 保证金流水明细(推荐,保证金退款暗道的逐单解释器——2026-08-11 拆 5646 实证)
    ("平台订单", {"子订单编号", "买家应付货款"}),
    ("平台退款", {"退款编号", "货物状态"}),
    ("平台保证金", {"完成时间", "操作类型", "收支金额(元)", "来源账户"}),
    ("平台价保", {"主订单ID"}),
]


# ---------------- 事件索引（V2.269：抽屉时间链的业务全链——下单/付款/发货） ----------------
def parse_platform_events(data):
    """④平台订单导出 → {主订单号: {created, paid}}。51k 行扫一遍，上传后后台建索引用。"""
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    ws = wb.active
    hdr = None
    out = {}
    for row in ws.iter_rows(values_only=True):
        if hdr is None:
            hdr = {str(h).strip(): i for i, h in enumerate(row) if h is not None}
            if "主订单编号" not in hdr:
                hdr = None
            continue
        g = lambda n: row[hdr[n]] if n in hdr and hdr[n] < len(row) else None
        o = _key(g("主订单编号"))
        if o and o not in out:
            out[o] = {"created": str(g("订单创建时间") or "")[:16], "paid": str(g("订单付款时间") or "")[:16]}
    wb.close()
    return out


def parse_ship_events(data):
    """③旺店通销售出库 → {平台单号: {shipped, ck, jy}}。
    原始单号可能是分号连双单号（合并发货实证）——逐个拆开都建键；子单原始单号也建键。"""
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    title = _pick_sheet(wb, ("出库单编号", "货品成交总价"))
    if not title:
        wb.close()
        return {}
    ws = wb[title]
    hdr = None
    out = {}
    for row in ws.iter_rows(values_only=True):
        if hdr is None:
            heads = {str(h).strip(): i for i, h in enumerate(row) if h is not None}
            if "出库单编号" in heads:
                hdr = heads
            continue
        g = lambda n: _key(row[hdr[n]]) if n in hdr and hdr[n] < len(row) else ""
        ev = {"shipped": str(g("发货时间"))[:16], "ck": g("出库单编号"), "jy": g("订单编号")}
        keys = set()
        for col in ("原始单号", "子单原始单号", "原始子订单号"):
            for part in g(col).replace("；", ";").split(";"):
                if part.strip():
                    keys.add(part.strip())
        for k in keys:
            out.setdefault(k, ev)
    wb.close()
    return out


def parse_deposit(data):
    """平台保证金流水明细 → {订单号: 保证金退款合计}。
    只取「出账·给消费者」的退款行（0020002/交易售后，负数取绝对值）；补缴充值行不计（那是池子回血）。"""
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    title = _pick_sheet(wb, ("完成时间", "操作类型", "收支金额(元)"))
    if not title:
        wb.close()
        raise ValueError("没找到保证金流水表头（应含「完成时间」「操作类型」「收支金额(元)」列）")
    ws = wb[title]
    hdr = None
    out = defaultdict(float)
    n = 0
    for row in ws.iter_rows(values_only=True):
        if hdr is None:
            heads = {str(c).strip(): i for i, c in enumerate(row) if c is not None}
            if "完成时间" in heads and "操作类型" in heads:
                hdr = heads
            continue
        g = lambda name: row[hdr[name]] if name in hdr and hdr[name] < len(row) else None
        if str(g("操作类型") or "").strip() != "出账":
            continue
        # 订单编号列：7 月导出该列【表头为空】（5 月有列名——版式漂移实证 2026-08-12）。
        # 有列名用列名；没有就从右往左找 15 位以上纯数字串当订单号。
        o = _key(g("订单编号"))
        if not o:
            import re as _re
            for cell in reversed(row):
                v = _key(cell)
                if _re.match(r"^\d{15,25}$", v):
                    o = v
                    break
        amt = _num(g("收支金额(元)"))
        desc = str(g("业务描述") or "")
        if o and amt < 0 and ("0020002" in desc or "退款" in desc or "售后" in str(g("原因") or "")):
            out[o] += -amt
            n += 1
    wb.close()
    return dict(out), n


def sniff_type(data):
    """bytes(xlsx) → 类型名 或 None。扫每张表前几行找表头，按特征列组合判型。"""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception:
        return None
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows(min_row=1, max_row=8, values_only=True):
                heads = {str(c).strip() for c in row if c is not None}
                for name, need in _SNIFF_RULES:
                    if need <= heads:
                        return name
    finally:
        wb.close()
    return None


# ---------------- 凭证报文（V2.257，写金蝶=草稿 only；配置驱动，配不齐不出报文） ----------------
# 结构参照物流线 2026-07-06 控制测试验证的配方（kernels/logistics_accrual.build_kd_model，只参照不改动）。
# ⚠⚠ 两个实证坑（缺一即录错月份）：① 字段名必须 FDate（非全大写）；② 金蝶 Save 对 JSON 字段顺序
# 敏感——FDate 必须排在 FACCOUNTBOOKID/FEntity 之后，否则被忽略、凭证钉到当前会计期间。
def build_settle_vouchers(fees, stats, cfg, period):
    """费目归集 → 两张凭证 Model（扣款项 / 收款核销）。cfg 缺项时 raise ValueError（不出半张报文）。

    cfg（基础资料·凭证配置）：book_code 账簿 / voucher_group 凭证字 / currency 币别 /
        cash_acct 其他货币资金-支付宝科目编码 / ar_acct 应收账款科目编码；
    fee_map 行须有 kd_code（各费目的借方科目编码）。"""
    need = [k for k in ("book_code", "voucher_group", "currency", "cash_acct", "ar_acct") if not str(cfg.get(k) or "").strip()]
    if need:
        raise ValueError("凭证配置缺项：%s（基础资料 › 凭证配置）" % "、".join(need))
    nocode = sorted({f["code"] for f in fees
                     if f["code"] not in ("0010001", "0020001") and (f["income"] or f["outgo"])
                     and not str(f.get("kd_code") or "").strip()})
    if nocode:
        raise ValueError("以下费目缺科目编码：%s（基础资料 › 费目科目映射）" % "、".join(nocode[:8]))
    y, m = int(period[:4]), int(period[5:7])
    import calendar
    fdate = "%04d-%02d-%02d" % (y, m, calendar.monthrange(y, m)[1])
    base = {"FCURRENCYID": {"FNumber": cfg["currency"]},
            "FEXCHANGERATETYPE": {"FNumber": cfg.get("rate_type") or "HLTX01_SYS"}, "FEXCHANGERATE": 1.0}

    def model(entries):
        return {"FACCOUNTBOOKID": {"FNumber": cfg["book_code"]},
                "FVOUCHERGROUPID": {"FNumber": cfg["voucher_group"]},
                "FEntity": entries,
                "FDate": fdate, "FYear": y, "FPeriod": m}

    # 凭证一 · 平台扣款项：借 各费目（净支出=支−收），贷 其他货币资金
    exp_entries = []
    total = 0.0
    for f in fees:
        if f["code"] in ("0010001", "0020001"):
            continue
        amt = _r2(f["outgo"] - f["income"])
        if amt == 0:
            continue
        total = _r2(total + amt)
        exp_entries.append(dict(base, FEXPLANATION="%s年%s月结算单扣款项 %s%s元" % (y, m, f["label"], amt),
                                FACCOUNTID={"FNumber": f["kd_code"]}, FDEBIT=amt, FCREDIT=0))
    exp_entries.append(dict(base, FEXPLANATION="%s年%s月结算单扣款项" % (y, m),
                            FACCOUNTID={"FNumber": cfg["cash_acct"]}, FDEBIT=0, FCREDIT=total))
    # 凭证二 · 收款核销：借 其他货币资金，贷 应收账款（收款−余额退款）
    recv = _r2(sum(f["income"] for f in fees if f["code"] == "0010001")
               - sum(f["outgo"] for f in fees if f["code"] == "0020001"))
    settle_entries = [
        dict(base, FEXPLANATION="%s年%s月平台结算收款" % (y, m),
             FACCOUNTID={"FNumber": cfg["cash_acct"]}, FDEBIT=recv, FCREDIT=0),
        dict(base, FEXPLANATION="%s年%s月平台结算收款·核销应收" % (y, m),
             FACCOUNTID={"FNumber": cfg["ar_acct"]}, FDEBIT=0, FCREDIT=recv),
    ]
    return [{"kind": "扣款项", "amount": total, "model": model(exp_entries)},
            {"kind": "收款核销", "amount": recv, "model": model(settle_entries)}]


# ---------------- 解析：退款不退货表 ----------------
def parse_refunds(data):
    """bytes → {原始单号: 分摊退款合计}。表头按列名找（原始单号/分摊退款金额）。"""
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    title = _pick_sheet(wb, ("原始单号", "分摊退款金额"))
    if not title:
        wb.close()
        raise ValueError("没找到退款不退货表头（应含「原始单号」「分摊退款金额」列）")
    ws = wb[title]
    io_, ia = None, None
    rk = defaultdict(float)
    n = 0
    for row in ws.iter_rows(values_only=True):
        if io_ is None:
            heads = {str(c).strip(): i for i, c in enumerate(row) if c is not None}
            if "原始单号" in heads and "分摊退款金额" in heads:
                io_, ia = heads["原始单号"], heads["分摊退款金额"]
            continue
        k = _key(row[io_])
        if k:
            rk[k] += _num(row[ia])
            n += 1
    wb.close()
    if io_ is None:
        raise ValueError("没找到退款不退货表头（应含「原始单号」「分摊退款金额」列）")
    return dict(rk), n


# ---------------- AR 索引：红字字段左移，双字段 ----------------
def index_receivables(ar_rows):
    """金蝶应收行 → {平台原始单号: {"amt", "nos", "has_red"}}。蓝字取 Text6；红字 Text6 空→取 Text4。
    has_red＝名下有红字行（重复登记检测用：退款不退货登记 + 已红冲退货并存＝疑重复）。"""
    ar = defaultdict(lambda: {"amt": 0.0, "nos": set(), "has_red": False})
    for r in ar_rows:
        k = _key(r.get("Text6")) or _key(r.get("Text4"))
        if not k:
            continue
        amt = _num(r.get("价税合计"))
        ar[k]["amt"] += amt
        if amt < 0:
            ar[k]["has_red"] = True
        ar[k]["nos"].add(str(r.get("单据编号", "")).split("-")[0])
    return ar


# ---------------- 主引擎 ----------------
DEFAULT_RULES = {
    "ufirst_max": 1.0,          # U先识别兜底：单收入 < 此值 且 无应收
    "inner_min": 1000.0,        # 空描述大额转账 ≥ 此值 → 内部划转剔除（Kiki 开户垫资 5000 实证）
    "qr_goods": "收钱码收款",    # 线下扫码直付（5 月线下活动 5,323 实证）→ 剔除留痕
}
# U先识别主信号：订单流水里带 U先专属费目（2026-08-11 首跑实证 0170155T 9,066 笔）。
# 金额档兜底会漏多件U先单（0.19×6件>1元），费目信号不受件数影响。
UFIRST_FEE_CODES = ("0170155T", "0170144T")


def run(flow_rows, refund_map, ar_index, fee_map, rules=None, shop="", deposit_map=None):
    """核销主引擎。输入全部为已解析结构，输出 dict（orders/fees/excluded/stats/vouchers）。

    fee_map: {code: {"label","account"}}——科目映射；不在表内的费目码 account="待定"（前端红标报警，D9）。
    deposit_map: {订单号: 保证金退款额}——④平台保证金明细解析所得。交易结算后客服再退款走保证金池
    （0020002），不在支付宝交易退款(0020001)里——不并入则此类单永远假差异（5646 实证 2026-08-11）。
    """
    deposit_map = deposit_map or {}
    ru = dict(DEFAULT_RULES)
    ru.update(rules or {})

    zones = defaultdict(lambda: [0.0, 0.0])         # (区,费目码) -> [收,支]
    zone_label = {}
    plat = defaultdict(float)                       # 订单 -> 平台收入
    meta = {}                                       # 订单 -> {channel, arrive, serial}
    ufirst_set = set()                              # 带 U先专属费目的订单（识别主信号）
    excluded = defaultdict(lambda: {"cnt": 0, "amount": 0.0, "detail": []})

    for r in flow_rows:
        chan_raw = str(r.get("chan") or "").strip()
        zone = "聚合" if chan_raw == "聚合结算渠道" else "支付宝"
        code, label = _code_of(r.get("desc"))
        inc, out = _num(r.get("income")), _num(r.get("outgo"))
        k = _key(r.get("order_no"))

        # ③ 层：按订单归集平台收入（0010001 − 0020001）
        if code == "0010001" and k and k != "0":
            plat[k] += inc
            m = meta.setdefault(k, {})
            m["channel"] = "聚合" if zone == "聚合" else "支付宝"
            m["arrive"] = str(r.get("ts") or "")[:19]
            # 聚合渠道无支付宝流水号，以聚合交易号定位（2026-08-11 实证，回查走天猫结算中心）
            m["serial"] = _key(r.get("serial")) or _key(r.get("txn"))
        elif code == "0020001" and k and k != "0":
            plat[k] -= out
        elif code in UFIRST_FEE_CODES and k and k != "0":
            ufirst_set.add(k)

        # ① 层：费目归集（含三类剔除，确认书⑤ 5.3——剔了必须留痕）
        if not code:
            btype = str(r.get("btype") or "")
            if btype == "在线支付" and str(r.get("goods") or "") == ru["qr_goods"]:
                _excl(excluded, "收钱码收款（线下扫码直付）", r, inc)
                continue
            if btype == "转账" and inc >= ru["inner_min"]:
                _excl(excluded, "内部划转", r, inc)
                continue
            if label:                                # 描述有字但无费目码：万相台充值等划转类
                _excl(excluded, "充值划转（%s）" % label[:24], r, inc + out)
                continue
            code, label = "空格扣款", '"业务描述=空格"扣款'
        zones[(zone, code)][0] += inc
        zones[(zone, code)][1] += out
        zone_label.setdefault(code, label)

    # ② 层：凭证数字（费目→科目；无映射=待定，报警不静默）
    fees = []
    unmapped = 0
    for (zone, code), (inc, out) in sorted(zones.items()):
        fm = fee_map.get(code) or {}
        account = fm.get("account") or "待定"
        if account == "待定":
            unmapped += 1
        fees.append({"zone": zone, "code": code,
                     "label": fm.get("label") or zone_label.get(code, ""),
                     "income": _r2(inc), "outgo": _r2(out), "account": account})

    # ③ 层：逐单核销 + 分桶
    orders = []
    for k in sorted(set(plat) | set(refund_map)):
        in_flow = k in plat                          # 本期支付宝有没有这单的收付
        p = _r2(plat.get(k, 0.0))
        a_ent = ar_index.get(k)
        a = _r2(a_ent["amt"]) if a_ent else 0.0
        nos = ",".join(sorted(a_ent["nos"])) if a_ent else ""
        rk = _r2(refund_map.get(k, 0.0))
        dep = _r2(deposit_map.get(k, 0.0))           # 保证金通道退款（④平台保证金明细）
        m = meta.get(k, {})
        o = {"order_no": k, "shop": shop, "channel": m.get("channel", ""),
             "arrive_time": m.get("arrive", ""), "serial_no": m.get("serial", ""),
             "plat_amt": p, "ar_no": nos, "ar_amt": a, "rk_amt": rk,
             "diff": _r2(p - dep - a + rk), "bucket": "", "note": ""}
        if dep:
            o["note"] = "含保证金通道退款 −%.2f（结算后客服退款走保证金池，见④平台保证金明细）" % dep
        # 重复登记检测（5646 实证）：退款不退货登记 + 名下已有红字退货并存 → 疑同一退款两套记录
        if rk > 0 and a_ent and a_ent.get("has_red"):
            o["note"] = (o["note"] + "；" if o["note"] else "") + \
                "⚠退款不退货登记与已红冲退货并存——疑同一笔退款重复记录，核实货是否实际回仓后剔除一边"
        # 跨期单（2026-08-11 九笔排查落定）：本期流水里根本没这单、只有退款不退货登记——
        # 其应收已在往期结算核销，7月再计全额应收是引擎粗糙。只列调节、不计差异。
        if not in_flow and rk > 0:
            o["bucket"] = "carry"
            o["diff"] = 0.0
            o["note"] = ("跨期：本期仅退款不退货登记 %.2f，无本期到账；应收 %.2f（%s）已在往期结算核销——不计差异，调节随售后通道处理"
                         % (rk, a, nos or "查无"))
        orders.append(o)

    _bucketize(orders, ru, ufirst_set)

    # 汇总
    stats = {"orders": len(orders), "plat_amt": _r2(sum(o["plat_amt"] for o in orders)),
             "ar_amt": _r2(sum(o["ar_amt"] for o in orders)),
             "rk_amt": _r2(sum(o["rk_amt"] for o in orders)),
             "fee_out": _r2(sum(f["outgo"] for f in fees if f["code"] not in ("0010001", "0020001"))),
             "unmapped_codes": unmapped,
             "buckets": {}}
    for o in orders:
        b = stats["buckets"].setdefault(o["bucket"], {"cnt": 0, "amount": 0.0})
        b["cnt"] += 1
        b["amount"] = _r2(b["amount"] + o["diff"])
    excl_rows = [{"kind": kd, "cnt": v["cnt"], "amount": _r2(v["amount"]),
                  "detail": json.dumps(v["detail"], ensure_ascii=False)}   # V2.273 全量（_excl 内 500 上限）
                 for kd, v in excluded.items()]
    return {"orders": orders, "fees": fees, "excluded": excl_rows, "stats": stats}


def _excl(excluded, kind, r, amount):
    e = excluded[kind]
    e["cnt"] += 1
    e["amount"] = _r2(e["amount"] + amount)
    # V2.273 全量留痕（需求方定：剔除项要逐笔定性——对得上活动的正常，对不上=违规资金流出）。
    # 上限 500 只防极端撑爆存储，截断时前端按 cnt>len(detail) 明示，不静默。
    if len(e["detail"]) < 500:
        e["detail"].append({"时间": str(r.get("ts") or "")[:19], "金额": _num(r.get("income")) or -_num(r.get("outgo")),
                            "流水号": _key(r.get("serial")) or _key(r.get("txn")),
                            "对方": str(r.get("peer") or "")[:20], "描述": str(r.get("desc") or "")[:80]})


def _bucketize(orders, ru, ufirst_set=()):
    """分桶（确认书⑤ D11）：ok / ufirst / crossed（串单±配对）/ carry（跨期调节）/ real（真差异，要人看）。
    U先＝费目信号（主）+ 金额档（兜底）；串单＝±金额精确对冲自动配对（两月实证）。"""
    by_diff = defaultdict(list)
    for o in orders:
        if o["bucket"]:                              # 跨期等已在上游定桶
            continue
        if abs(o["diff"]) <= 0.005:
            o["bucket"] = "ok"
        elif o["ar_amt"] == 0 and (o["order_no"] in ufirst_set
                                   or 0 < o["plat_amt"] < ru["ufirst_max"]):
            o["bucket"] = "ufirst"                   # U先：月末总账一笔汇总，逐单跳过、每月对总数
        else:
            o["bucket"] = "real"
            by_diff[abs(o["diff"])].append(o)
    for amt, group in by_diff.items():
        pos = [o for o in group if o["diff"] > 0]
        neg = [o for o in group if o["diff"] < 0]
        for a, b in zip(pos, neg):                   # ±精确对冲 → 配成串单对
            a["bucket"] = b["bucket"] = "crossed"
            a["note"] = "串单嫌疑：与 %s 对冲 ±%.2f（改单号前禁止下推）" % (b["order_no"], amt)
            b["note"] = "串单嫌疑：与 %s 对冲 ±%.2f（改单号前禁止下推）" % (a["order_no"], amt)
    # 真差异·机器初判（2026-08-11 九笔五路排查归纳的四类形态，只提示不定论）
    for o in orders:
        if o["bucket"] != "real" or o["note"]:
            continue
        p, a, rk = o["plat_amt"], o["ar_amt"], o["rk_amt"]
        if a < 0:
            o["note"] = "仅红字应收——蓝字疑串至合并发货的另一单（同捡货合并单查旺店通出库；改单号后自平）"
        elif a == 0 and rk > 0 and p > 0:
            o["note"] = "金蝶蓝红已对冲清零、退款不退货又登记调节——疑重复登记，核实该单是否实际退了货"
        elif a > p and rk == 0:
            o["note"] = "应收大于到账——疑换货/补发挂原单多生成应收，或退货红冲未出（催退货单审批后自平）"
        elif a == 0 and p > 0:
            o["note"] = "查无应收——疑串单蓝字挂他单，或旺店通未推送生成应收"
        elif a > 0 and p > 0 and a < p:
            o["note"] = "到账大于应收——查平台退款明细是否有金蝶未红冲的部分退款/口径差（退款含运费等）"


# ---------------- 费目映射种子（2026-05/07 两月凭证区实证提取，D9 受控起点） ----------------
FEE_MAP_SEED = [
    ("0010001", "交易收款-交易收款", "应收账款"),
    ("0020001", "交易退款-余额退款", "应收账款"),
    ("0020002", "交易退款-保证金退款", "费用"),
    ("0030001", "软件服务费-天猫返点积分", "费用"),
    ("0030002", "软件服务费-天猫年费", "费用"),
    ("0030003", "软件服务费-类目软件服务费（原天猫佣金）", "费用"),
    ("0030007", "软件服务费-聚划算", "费用"),
    ("0030018", "软件服务费-天猫返点积分（退款）", "费用"),
    ("0030038", "软件服务费-支付宝服务费", "费用"),
    ("0030039", "软件服务费-每日必买", "费用"),
    ("0030129", "软件服务费-热浪引擎第三方服务商服务费（平台开票）", "费用"),
    ("0030130", "软件服务费-基础软件服务费", "费用"),
    ("0030153T", "软件服务费-淘宝营销活动软件服务费", "费用"),
    ("0030162T", "软件服务费-淘金币软件服务费", "费用"),
    ("0050005", "保险支出-卖家版运费险", "费用"),
    ("0060011", "营销支出-淘宝客佣金", "费用"),
    ("0060081T", "营销支出-食品生鲜美食多买多省商家红包出资", "费用"),
    ("0060092T", "营销支出-消费券代付资金扣回", "费用"),
    ("0060096T", "营销支出-淘特营销托管推广服务费", "费用"),
    ("0060105T", "营销支出-食品生鲜美食多买多省商家加补券出资", "费用"),
    ("0070023", "其他支出-百亿补贴预收", "费用"),
    ("0110001", "公益性捐赠支出-公益宝贝", "费用"),
    ("0170125T", "服务费-消费者体验提升计划服务费", "费用"),
    ("0240004T", "其他收入-百亿补贴激励前返", "费用"),
    ("0530294T", "技术&服务费-限时红包代商家垫付扣回", "费用"),
    ("008000200003", "保证金-天猫-扣除转移", "费用"),
    ("008002800014", "保证金-天猫-出账缴存", "费用"),
    ("008002800015", "保证金-淘宝-额度补齐缴存", "费用"),
    # 7 月流水新见三码（2026-08-11 正式内核首跑补录；0170155T=U先供应链服务费 9,066 笔，
    # 亦是 U先单的费目级识别信号）
    ("0170144T", "服务费-天猫U先物流理赔费用", "费用"),
    ("0170155T", "服务费-天猫U先供应链管理服务费", "费用"),
    ("0530288T", "技术&服务费-大服饰跨境服务增值费", "费用"),
    ("空格扣款", '"业务描述=空格"扣款', "费用"),
]

# 店铺对照种子（确认书① D8：3 处已知不一致 + 页签差字实证）
SHOP_MAP_SEED = [
    ("星期零旗舰店（抖音）（头条放心购）", "星期零STARFIELD 抖音官旗店", "抖音"),
    ("星期零食品旗舰店（天猫）", "星期零_starfield", "淘宝"),
    ("旺店通线下手工店铺", "星期零线下手工店铺", "线下"),
    ("星期零STARFIELD 天猫官旗店", "星期零STARFIELD 天猫官旗店", "天猫"),
    ("Kiki Herb天猫官方旗舰店", "Kiki Herb天猫官方旗舰店", "天猫"),
    ("Kiki Herb小红书官方旗舰店", "Kiki Herb小红书官方旗舰店", "小红书"),
]

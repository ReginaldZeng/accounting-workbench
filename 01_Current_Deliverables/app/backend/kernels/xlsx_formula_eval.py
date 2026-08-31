# -*- coding: utf-8 -*-
# [Change Log]
# Date: 2026-08-13 | Author: Claude / c | Version: V2.291
# Description: 极小的 Excel 公式求值器，**只为验自己写出去的公式**（不是通用引擎）。
#   起因：V2.291 起「汇总透视」各块改成引用原始底表的 SUMIFS 公式——
#   openpyxl 写得出、读不回值，本机又没有 LibreOffice，
#   于是"公式对不对"这件事无法验证；而 V2.290 刚出过一次**合并后公式指错行**的错数事故
#   （汇总透视!B149=SUM(B2:B40)，加的是别的块），教训是"公式一旦不验，就是不知道对不对的数"。
#   本模块把导出文件当输入真算一遍，与内核算出来的值逐格比。
#   ⚠**只支持本工具实际生成的三种形态**，多一种都不认（认不出就抛，不许静默返回 None——
#     静默返回会让"没验到"看起来像"验过了"）：
#       =SUMIFS('页'!X:X, '页'!Y:Y, 条件[, '页'!Z:Z, 条件]…)   条件＝"常量" / "<>常量" / 同页单元格引用
#       =SUM(A1:A9)
#       =SUMIF(B2:B9,"小计",C2:C9)
#       =B12-B13                                             两格相减（损益归集的「差异」行）
import re

from openpyxl.utils import column_index_from_string as _ci

_RANGE = re.compile(r"'([^']+)'!([A-Z]+):[A-Z]+$")
_SUM = re.compile(r"SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)$")
_SUMIF = re.compile(r'SUMIF\(([A-Z]+)(\d+):([A-Z]+)(\d+),"([^"]*)",([A-Z]+)(\d+):([A-Z]+)(\d+)\)$')
_DIFF = re.compile(r"([A-Z]+\d+)-([A-Z]+\d+)$")


class UnsupportedFormula(Exception):
    pass


def _match(v, crit):
    """Excel SUMIFS 的条件语义（本工具只用到 相等 与 <>）。空单元格按空串比。"""
    s = "" if v is None else str(v)
    if isinstance(crit, str) and crit.startswith("<>"):
        return s != crit[2:]
    return s == ("" if crit is None else str(crit))


def _num(wb, ws, cell, seen=None):
    """取一格的数值。**这格本身是公式就递归求**——「按仓库」的总计是
    `SUMIF(…,"小计",…)`，而小计格又是 `SUM(…)`；不解嵌套会把它当文本跳过、静静地算出 0。"""
    v = cell.value
    if isinstance(v, str) and v.startswith("="):
        seen = seen or set()
        key = (ws.title, cell.coordinate)
        if key in seen:
            raise UnsupportedFormula("循环引用 %s!%s" % key)
        seen.add(key)
        return evaluate(wb, ws, v, seen)
    return v if isinstance(v, (int, float)) else None


def evaluate(wb, ws, formula, _seen=None):
    """求一个公式的值。wb=Workbook，ws=公式所在页，formula 以 '=' 开头。"""
    f = formula[1:].strip()

    if f.startswith("SUMIFS("):
        parts = [p.strip() for p in f[len("SUMIFS("):-1].split(",")]
        m = _RANGE.match(parts[0])
        if not m or len(parts) < 3 or (len(parts) - 1) % 2:
            raise UnsupportedFormula(formula)
        sheet, sum_col = m.group(1), m.group(2)
        crits = []
        for i in range(1, len(parts), 2):
            m2 = _RANGE.match(parts[i])
            if not m2:
                raise UnsupportedFormula(formula)
            c = parts[i + 1]
            if c.startswith('"'):
                c = c[1:-1]
            else:                       # 同页单元格引用（如 B7）
                c = ws[c].value
            crits.append((m2.group(2), c))
        src = wb[sheet]
        cols = {sum_col: _ci(sum_col)}
        for cc, _ in crits:
            cols[cc] = _ci(cc)
        tot = 0.0
        for r in range(1, src.max_row + 1):
            if all(_match(src.cell(row=r, column=cols[cc]).value, cv) for cc, cv in crits):
                v = _num(wb, src, src.cell(row=r, column=cols[sum_col]), _seen)
                if v is not None:
                    tot += v
        return tot

    m = _SUM.match(f)
    if m:
        j = _ci(m.group(1))
        return sum(_num(wb, ws, ws.cell(row=r, column=j), _seen) or 0
                   for r in range(int(m.group(2)), int(m.group(4)) + 1))

    m = _SUMIF.match(f)
    if m:
        cj, sj = _ci(m.group(1)), _ci(m.group(6))
        tot = 0.0
        for r in range(int(m.group(2)), int(m.group(4)) + 1):
            if str(ws.cell(row=r, column=cj).value or "") == m.group(5):
                v = _num(wb, ws, ws.cell(row=r, column=sj), _seen)
                if v is not None:
                    tot += v
        return tot

    m = _DIFF.match(f)
    if m:
        a = _num(wb, ws, ws[m.group(1)], _seen) or 0
        b = _num(wb, ws, ws[m.group(2)], _seen) or 0
        return a - b

    raise UnsupportedFormula(formula)


def evaluate_all(wb):
    """全簿求值 → {(页名, 坐标): 值}。任何一个认不出就抛。"""
    out = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    out[(ws.title, c.coordinate)] = evaluate(wb, ws, c.value)
    return out

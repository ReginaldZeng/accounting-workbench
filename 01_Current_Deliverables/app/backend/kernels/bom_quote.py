# -*- coding: utf-8 -*-
# [Change Log]
# Date: 2026-09-02 | Author: Claude / c | Version: V-draft(BOM报价审核)
# Description: 【BOM报价审核】内核——星期九「成本核算样表」解析器 + 台账口径（含税五分项）+ 归组键 + 重排版导出。
#              纯算法层：吃 xlsx 字节 / 路径，吐结构化记录；不碰 FastAPI、不碰 DB、不联网。
#              移植自交接夹 tools/parse_cost_workbook.py（标签定位、绝不写死列号）+ build_pretty_sheet.py。
#
#              一处比样机更稳的口径决定（记下来免得日后当 BUG）：
#              加工费(mfg) = (制造费用小计 + 工厂费用小计) × 1.13。样机只算了制造费用；
#              但源表勾稽 #4 的成本合计里明确含「工厂费用小计」，样例里它恒为 0 故看不出差别，
#              一旦某单工厂费用≠0，只算制造费用会让「源表对账·差异」凭空非零。把工厂费用并进加工费，
#              既保留样机的五列口径，又让 full == 全成本含税（勾稽全平时差异恒 0）。人工成本不单独加——
#              它本就含在制造费用小计里（否则勾稽 #4 过不了，样例 6 项全平已反证）。

import io
import os
import re
from openpyxl import load_workbook

# 含税毛化系数：源表自身惯例 含税 = 不含税 × 1.13（quirk #3，人工含税原值与此精确一致）
TAX_GROSS = 1.13
# 成本涨跌显示阈值 元/kg（低于此不标三角，与样机一致）
DELTA_EPS = 0.005

HEAD_LABELS = ["供应商名称", "产品名称", "物料编码", "产品编号", "包装规格", "核算日期"]
COL_ALIASES = {  # 明细表头 -> 规范字段
    "物料编码": "matCode", "包材编码": "matCode", "物料名称": "matName",
    "型号": "model", "规格": "spec", "单位": "unit", "品牌": "brand",
    "添加量（kg)": "qtyPerKg", "添加量（kg）": "qtyPerKg",
    "税率": "taxRate", "发票类别": "invoiceType", "占比": "share",
    "报价说明": "priceNote", "订单对应用量": "orderQty", "起订量": "moq",
}

_DATE_RE = re.compile(r"(20\d{2})[-.年/](\d{1,2})[-.月/](\d{1,2})")


def norm(v):
    return str(v).strip() if v is not None else ""


def num(v):
    return float(v) if isinstance(v, (int, float)) else None


def _date_from_name(name):
    """文件名回推核算日期（半成品页表头常把「核算日期」抓歪，见 quirk 容错层）。→ 'YYYY-MM-DD' 或 ''。"""
    m = _DATE_RE.search(name or "")
    if not m:
        return ""
    return "%04d-%02d-%02d" % (int(m.group(1)), int(m.group(2)), int(m.group(3)))


def _clean_date(raw, src_file):
    """核算日期容错：是日期取前 10 位；抓到中文/客户名等非日期文本 → 从文件名回推。"""
    s = str(raw)[:10] if raw is not None else ""
    if re.match(r"^20\d{2}-\d{1,2}-\d{1,2}", s) or re.match(r"^20\d{2}/\d{1,2}/\d{1,2}", s):
        return s.replace("/", "-")
    return _date_from_name(src_file)


def _clean_erp(raw):
    """ERP 码容错：抓到中文（半成品页把「物料名称」当成了 erpCode）→ 置空。"""
    s = norm(raw)
    if re.search(r"[一-鿿]", s):     # 含中文 = 抓歪了，宁可空
        return ""
    return s


def parse_sheet(ws, src_file):
    """解析一个 worksheet；不是成本核算页返回 None。字段口径见文件头与交接文档 §4。"""
    grid = [[c.value for c in row] for row in ws.iter_rows()]
    R = len(grid)
    C = max((len(r) for r in grid), default=0)

    def cell(r, c):
        return grid[r][c] if 0 <= r < R and 0 <= c < len(grid[r]) else None

    # 是成本核算页吗（前 4 行找「成本核算样表」标记；BOM 清单等非核算表自动跳过）
    if not any("成本核算样表" in norm(v) for row in grid[:4] for v in row):
        return None

    head = {}
    for r in range(min(6, R)):
        for c in range(C):
            lab = norm(cell(r, c)).replace(" ", "")
            if lab in HEAD_LABELS and lab not in head:
                for cc in range(c + 1, min(c + 6, C)):
                    v = cell(r, cc)
                    if norm(v):
                        head[lab] = v
                        break
        for c in range(C):     # 订单量 / 客户 藏在右侧
            t = norm(cell(r, c))
            if t.startswith("订单量"):
                v = cell(r + 1, c)
                if num(v) is not None:
                    head["订单量kg"] = v
            if t.startswith("客户："):
                head.setdefault("客户", t.replace("客户：", "").strip())

    def find_row(pred, start=0):
        for r in range(start, R):
            if pred(r):
                return r
        return None

    def row_has(r, text):
        return any(norm(cell(r, c)).replace("\n", "").replace(" ", "").startswith(text) for c in range(C))

    r_var = find_row(lambda r: row_has(r, "1、变动成本"))

    def header_map(r):
        m = {}
        if r is None:
            return m
        for c in range(C):
            t = norm(cell(r, c)).replace("\n", "")
            for k, f in COL_ALIASES.items():
                if t == k or t.startswith(k):
                    m.setdefault(f, c)
            if "采购价" in t:
                m["priceIncl"] = c
            if t.startswith("成本不含税"):
                m["costExcl"] = c
        return m

    def read_block(hdr_row, seg):
        if hdr_row is None:
            return [], None, hdr_row
        m = header_map(hdr_row)
        rows, r, subtotal = [], hdr_row + 1, None
        while r < R:
            if any(norm(cell(r, c)) == "小计" for c in range(C)) and num(cell(r, m.get("costExcl", -1))) is not None:
                subtotal = num(cell(r, m["costExcl"]))
                break
            name = cell(r, m["matName"]) if "matName" in m else None
            if norm(name):
                rows.append({
                    "seg": seg,
                    "matCode": norm(cell(r, m.get("matCode", -1))),
                    "matName": norm(name),
                    "model": norm(cell(r, m["model"])) if "model" in m else "",
                    "spec": norm(cell(r, m["spec"])) if "spec" in m else "",
                    "unit": norm(cell(r, m.get("unit", -1))),
                    "brand": norm(cell(r, m["brand"])) if "brand" in m else "",
                    "qtyPerKg": num(cell(r, m.get("qtyPerKg", -1))),
                    "priceIncl": num(cell(r, m.get("priceIncl", -1))),
                    "taxRate": num(cell(r, m.get("taxRate", -1))),
                    "invoiceType": norm(cell(r, m.get("invoiceType", -1))),
                    "costExcl": num(cell(r, m.get("costExcl", -1))),
                    "share": num(cell(r, m.get("share", -1))),
                    "priceNote": norm(cell(r, m.get("priceNote", -1))),
                    "moq": norm(cell(r, m.get("moq", -1))) if "moq" in m else "",
                })
            r += 1
        return rows, subtotal, r

    mats, sub_mat, r_end = read_block(r_var, "原料")
    r_pack_hdr = find_row(lambda r: any("包材" in norm(cell(r, c)) and "编码" in norm(cell(r, c))
                                        for c in range(C)), r_end or 0)
    packs, sub_pack, _ = read_block(r_pack_hdr, "包材") if r_pack_hdr else ([], None, r_end)

    def label_value(text, want_col=None, start=0):
        for r in range(start, R):
            for c in range(C):
                t = norm(cell(r, c)).replace("\n", "").replace(" ", "")
                if t.startswith(text):
                    if want_col is not None:
                        return num(cell(r, want_col))
                    for cc in range(c + 1, C):
                        v = num(cell(r, cc))
                        if v is not None:
                            return v
                    return None
        return None

    ce = header_map(r_var).get("costExcl")
    summary = {
        "变动小计不含税": label_value("变动含税合计", ce),   # quirk #1 模板标签笔误：值实为不含税
        "人工成本不含税": label_value("人工成本", ce),
        "制造费用小计不含税": None,
        "工厂费用小计不含税": None,
        "运输费用不含税": label_value("4、运输费用", ce) or 0,
        "装卸费不含税": label_value("5、装卸费", ce) or 0,
        "成本合计不含税": label_value("6、成本合计"),
        "增值税合计": label_value("7、增值税合计"),
        "成本合计含税": label_value("8、成本合计"),
        "管理费加成含税": label_value("9、管理费"),
        "全成本含税": label_value("10、全成本"),
    }
    r_mfg = find_row(lambda r: row_has(r, "2、制造费用"))
    r_fac = find_row(lambda r: row_has(r, "3、工厂费用"))
    r_trans = find_row(lambda r: row_has(r, "4、运输费用"))

    def next_subtotal(r0, r1):
        if ce is None:               # 成本不含税列没定位到 → 别 cell(r, None) 崩（审查 M10）
            return None
        for r in range(r0 or 0, r1 or R):
            if any(norm(cell(r, c)) == "小计" for c in range(C)):
                v = num(cell(r, ce))
                if v is not None:
                    return v
        return None

    summary["制造费用小计不含税"] = next_subtotal(r_mfg, r_fac)
    summary["工厂费用小计不含税"] = next_subtotal(r_fac, r_trans)

    checks = []

    def chk(name, a, b, tol=0.01):
        checks.append({"check": name, "a": a, "b": b,
                       "ok": a is not None and b is not None and abs(a - b) < tol})

    chk("原料小计=Σ明细", sub_mat, sum(x["costExcl"] or 0 for x in mats))
    chk("包材小计=Σ明细", sub_pack, sum(x["costExcl"] or 0 for x in packs))
    chk("变动=原料+包材", summary["变动小计不含税"], (sub_mat or 0) + (sub_pack or 0))
    chk("成本合计=变动+制造+工厂+运输+装卸", summary["成本合计不含税"],
        (summary["变动小计不含税"] or 0) + (summary["制造费用小计不含税"] or 0)
        + (summary["工厂费用小计不含税"] or 0) + (summary["运输费用不含税"] or 0) + (summary["装卸费不含税"] or 0))
    chk("含税=不含税+增值税", summary["成本合计含税"],
        (summary["成本合计不含税"] or 0) + (summary["增值税合计"] or 0))
    chk("全成本=含税+管理费", summary["全成本含税"],
        (summary["成本合计含税"] or 0) + (summary["管理费加成含税"] or 0))

    return {
        "srcFile": os.path.basename(src_file), "sheet": ws.title.strip(),
        "supplier": norm(head.get("供应商名称")), "productName": norm(head.get("产品名称")),
        "erpCode": _clean_erp(head.get("物料编码")), "cpCode": norm(head.get("产品编号")),
        "packSpec": norm(head.get("包装规格")), "calcDate": _clean_date(head.get("核算日期"), src_file),
        "orderQty": head.get("订单量kg"), "customer": norm(head.get("客户", "")),
        "matSubtotal": sub_mat, "packSubtotal": sub_pack, "summary": summary,
        "materials": mats + packs, "checks": checks,
    }


def parse_workbook(data, src_filename):
    """吃 xlsx 字节（或路径）→ 该文件里所有成本核算页的记录列表（一单 N 产品 = N 条）。
    data_only=True：读 Excel 缓存的公式值；文件必须是 Excel 存过的（否则公式值为 None）。"""
    src = data if isinstance(data, (bytes, bytearray)) else open(data, "rb").read()
    name = src_filename or (data if isinstance(data, str) else "")
    wb = load_workbook(io.BytesIO(src), data_only=True, read_only=True)
    out, sheet_errors = [], []
    for ws in wb.worksheets:
        try:                        # 逐 sheet 隔离（审查 M10）：一页版式异常不该毁掉整份 N 产品文件
            rec = parse_sheet(ws, name)
        except Exception as ex:
            sheet_errors.append("%s：%s" % (ws.title, ex))
            continue
        if rec:
            out.append(rec)
    wb.close()
    if sheet_errors and not out:    # 全崩才抛（让调用方回报）；部分崩则跳过坏页、留下能解析的
        raise RuntimeError("；".join(sheet_errors))
    return out


# ---------------- 台账口径：归组键 · 分类判定 · 含税五分项 · 涨跌 ----------------
# ⚠ 分类口径（业务方 2026-09-03 定）：**编码规律不固定了**，编码判定只作**建议值**，
#   最终以**成本会计人工指定**为准（entry.kind_override，见 router 的 /set-kind）。
#   实证反例（202609011316000251965）：CP04118003 名字带「半成品」却是 CP0→按编码判成品；
#   CP21120601 是真成品（120g规格）却是 CP2→按编码判半成品。故编码不可作唯一依据。
# 建议值仍按原规律给（好让多数单不用手点）：SZ→复配料；CP2→半成品；其它 CP→成品；名字兜底。
def classify(cp_code, product_name=""):
    """产品分类**建议值** → '成品' / '半成品' / '复配料'。⚠ 非权威：人工指定(kind_override)优先。
    **名字里写死的分类比编码可靠**（业务方 2026-09-04：251965 里 CP0 的「蘑力辣丝丝半成品」编码判成品、
    CP2 的真成品「蘑力辣丝丝」编码判半成品——都判反了）。故名字有明确信号时以名字为准，编码只在名字无信号时兜底。
    ⚠ 但**成品/半成品的顺序/分类以组内 BOM 结构最准**（见 group_kind），本函数是无组上下文时的单品建议值。"""
    nm = str(product_name or "")
    if "复配料" in nm or "复合调味" in nm:
        return "复配料"
    if "半成品" in nm:
        return "半成品"
    c = str(cp_code or "").strip().upper()
    if c.startswith("SZ"):
        return "复配料"
    if c.startswith("CP"):
        return "半成品" if c[2:3] == "2" else "成品"
    return "成品"


# ---------------- 审核定性：物料类别（5 类）+ 是否允许对外报价 ----------------
# 业务方 2026-09-03 定：点「审核」弹窗里由成本会计指定——物料类别五选一 + 是否建议/允许对外报价
# （不建议必须写原因，如「包材不全」「XX物料暂定」）。编码判定只作建议值，不再权威。
MAT_CATEGORIES = ("复配料", "自产半成品", "自产成品", "委外半成品", "委外成品")
# 五类 → 台账基础三分类（决定 semi：非成品可作原料进上层、不单独挂渠道）
_CAT_TO_KIND = {"复配料": "复配料", "自产半成品": "半成品", "委外半成品": "半成品",
                "自产成品": "成品", "委外成品": "成品"}
KINDS = ("成品", "半成品", "复配料")


def cat_to_kind(category):
    """物料类别（5 类）→ 基础三分类（成品/半成品/复配料）。认不出 → None。"""
    return _CAT_TO_KIND.get(str(category or "").strip())


def is_outsourced(category):
    """委外（代工厂生产）？——五类里「委外」打头的两类。"""
    return str(category or "").strip().startswith("委外")


def classify_by_name(product_name):
    """只按**产品名**判基础分类（名字里明写了半成品/复配料时）。给不出信号 → None。
    用途：与编码判定对照——两者打架时前端标「疑似分类不符，请成本会计确认」。"""
    nm = str(product_name or "")
    if "复配料" in nm or "复合调味" in nm:
        return "复配料"
    if "半成品" in nm:
        return "半成品"
    return None


def effective_kind(cp_code, product_name="", category=None):
    """最终基础分类：**人工定性的物料类别优先**，否则用编码/名字的建议值（编码不固定，2026-09-03）。"""
    k = cat_to_kind(category)
    return k or classify(cp_code, product_name)


def suggest_category(cp_code, product_name="", supplier=""):
    """物料类别建议值（仅供弹窗预选，成本会计须确认）：基础分类由编码/名字建议；
    自产/委外这一维**编码里没有**——核算表带「生产工厂/供应商」时倾向委外（星期九等代工厂），否则自产。
    复配料不分自产委外。"""
    base = classify(cp_code, product_name)
    if base == "复配料":
        return "复配料"
    pref = "委外" if str(supplier or "").strip() else "自产"
    return pref + ("半成品" if base == "半成品" else "成品")


def kind_doubt(cp_code, product_name="", category=None):
    """建议值可疑吗：名字明写的分类与编码判出的不一致，且尚未人工定性 → True（前端标⚠请确认）。"""
    if cat_to_kind(category):
        return False
    by_name = classify_by_name(product_name)
    return bool(by_name) and by_name != classify(cp_code, product_name)


# ---------------- 组内结构判分类（比编码可靠，业务方 2026-09-04 纠偏）----------------
# 实证反例（251965）：编码把「蘑力辣丝丝半成品」(CP0) 判成品、把真成品「蘑力辣丝丝」(CP2) 判半成品——判反了。
# BOM 依赖结构不会骗人：**没人引用它的是最终成品；既引用下游又被上层当原料的是半成品；被上层引用而自己
# 不引用下游的是最底复配料（原料）**。再叠加名字里写死的「半成品/复配料」。uses/used_by＝组内产品名列表。
def group_kind(product_name, uses, used_by):
    """组内结构优先的基础分类建议 → '成品'/'半成品'/'复配料'。名字写死的最优先，其次按 BOM 依赖结构。"""
    nm = str(product_name or "")
    if "半成品" in nm:
        return "半成品"
    if "复配料" in nm or "复合调味" in nm:
        return "复配料"
    u = bool(uses)           # 引用了下游（把半成品/复配料当原料）
    b = bool(used_by)        # 被上层当原料引用
    if b and u:
        return "半成品"      # 中间层
    if b and not u:
        return "复配料"      # 最底原料/复配（喂给别人、自己不引用下游）
    return "成品"            # 顶层（没人用它）或独立单品


def group_category(product_name, uses, used_by, supplier=""):
    """组内结构优先的物料类别建议（5 类，供定性弹窗预选）：基础分类走 group_kind，
    自产/委外维沿用 suggest_category 口径（带供应商/生产工厂→委外，否则自产；复配料不分）。"""
    base = group_kind(product_name, uses, used_by)
    if base == "复配料":
        return "复配料"
    pref = "委外" if str(supplier or "").strip() else "自产"
    return pref + ("半成品" if base == "半成品" else "成品")


def is_semi(cp_code, product_name=""):
    """非成品（半成品 + 复配料）即视作 semi：不默认渠道、可作原料进上层。CP 码可含全角括号（SN2）。
    ⚠ 只用编码建议值——已入账记录请用 effective_kind(含人工指定) 判。"""
    return classify(cp_code, product_name) != "成品"


# ---------------- 发票类型 → 成本不含税 算法（基础数据可维护）----------------
# 对应成本核算表 N 列公式（业务方 2026-09-04 给）：K=含税价 L=税率 J=添加量 →
#   专票:          K/(1+L)*J                              价税分离·进项按票面税率抵扣
#   自产自销农产品:  (K − K×9%)*J                            买价×(1−扣除率)，扣除率默认 9%
#   普票:          K*J                                    不抵扣，全额计入成本
#   农产品专票:     L>1% ? (K − K/(1+L)×9%)*J : K/(1+L)*J    有票面税率则先价税分离再计算抵扣，否则纯价税分离
# 每种发票 = {type, mode, rate}，可在基础数据增改；mode 决定算法，rate 是扣除率（农产品类用）。
INVOICE_MODES = ("价税分离", "全额", "买价扣除", "农产品专票")
INVOICE_RULE_DEFAULTS = [
    {"type": "专票", "mode": "价税分离", "rate": 0.0},
    {"type": "普票", "mode": "全额", "rate": 0.0},
    {"type": "自产自销农产品", "mode": "买价扣除", "rate": 0.09},
    {"type": "农产品专票", "mode": "农产品专票", "rate": 0.09},
]


def invoice_unit_excl(price, tax, mode, rate=0.0):
    """单位成本不含税（不乘添加量）。price=含税价, tax=税率, rate=扣除率。未知 mode → 价税分离兜底。"""
    p = float(price or 0)
    t = float(tax or 0)
    r = float(rate or 0)
    if mode == "全额":
        return p
    if mode == "买价扣除":
        return p * (1 - r)
    if mode == "农产品专票":
        return (p - p / (1 + t) * r) if t > 0.01 else (p / (1 + t) if (1 + t) else p)
    return p / (1 + t) if (1 + t) else p            # 价税分离（专票，默认）


def invoice_cost_excl(qty, price, tax, invoice_type, rules=None):
    """成本不含税 = 单位成本不含税 × 添加量，按发票类型选算法。未匹配到类型 → 价税分离兜底。"""
    rules = rules or INVOICE_RULE_DEFAULTS
    rule = next((x for x in rules if str(x.get("type") or "") == str(invoice_type or "")), None)
    mode = (rule or {}).get("mode", "价税分离")
    rate = (rule or {}).get("rate", 0.0)
    return round(invoice_unit_excl(price, tax, mode, rate) * float(qty or 0), 4)


def product_key(rec):
    """产品身份键（业务方 2026-09-04 定：**CP 码就是产品身份**，改口径）：
    有 CP 码 → 「产品名|CP码」——不同 CP（变体如印刷袋 …-2 / 空白袋 …、或改配方跳号）**各自成产品、各自定稿入库**，
    互不顶替；**同 CP** 才是同产品的版本链（如同一核算表重算/替换）。CP 码为空 → 退回「产品名|客户」（同名双产品靠客户分）。
    ⚠ 旧记录存的是老键(名+客户)；改键后**新入账/重连才走新键**，两者不互成版本——现存撞车单需重新立项/重连一次。"""
    pn = norm(rec.get("productName"))
    cp = norm(rec.get("cpCode"))
    if cp:
        return pn + "|" + cp
    return pn + "|" + norm(rec.get("customer"))


def fee_from_summary(summary):
    """从源表汇总推台账费用参数（全部 含税 元/kg，与样机 feeOf 一致，另并入工厂费用见文件头）。"""
    s = summary or {}
    mfg = round(((s.get("制造费用小计不含税") or 0) + (s.get("工厂费用小计不含税") or 0)) * TAX_GROSS, 4)
    load = round(((s.get("装卸费不含税") or 0) + (s.get("运输费用不含税") or 0)) * TAX_GROSS, 4)
    adm = round(s.get("管理费加成含税") or 0, 4)
    return {"mfg": mfg, "load": load, "adm": adm}


def compose(rec, fee=None):
    """含税五分项 + 全成本。fee=None 时用源表推的默认参数（复核可覆盖 mfg/load/adm）。
    返回 {mat, pack, mfg, load, adm, full, srcFull, diff}；full==srcFull（勾稽全平时 diff≈0）。"""
    src_fee = fee_from_summary(rec.get("summary"))
    f = {**src_fee, **(fee or {})}
    mat = round((rec.get("matSubtotal") or 0) * TAX_GROSS, 4)
    pack = round((rec.get("packSubtotal") or 0) * TAX_GROSS, 4)
    full = round(mat + pack + (f["mfg"] or 0) + (f["load"] or 0) + (f["adm"] or 0), 4)
    src_full = (rec.get("summary") or {}).get("全成本含税")
    return {"mat": mat, "pack": pack, "mfg": f["mfg"], "load": f["load"], "adm": f["adm"],
            "full": full, "srcFull": src_full,
            "diff": round(full - src_full, 4) if src_full is not None else None,
            "srcFee": src_fee}


def all_checks_ok(rec):
    """6 项勾稽全平才算可入账（红线：勾稽不平不准静默入账）。"""
    ck = rec.get("checks") or []
    return len(ck) >= 6 and all(c.get("ok") for c in ck)


def upstream_refs(rec, recs):
    """本产品的物料行里，哪些是**引用了同一核算表工作簿里另一张表**（半成品/复配料作原料进上层）。
    按「物料名 == 上游产品名」精确配（同一工作簿里研发就是这么写的：成品料行「蘑力辣丝丝半成品」= 半成品页产品名）。
    口径 quirk#5：**下层「全成本含税」＝上层料行里的「含税价」**——两者对不上也要报。
    返回 [{matName, qtyPerKg, priceUsed, upCpCode, upProductName, upFull, priceOk, upChecksOk}]。"""
    me = product_key(rec)
    by_name = {}
    for r in recs or []:
        pn = norm(r.get("productName"))
        if pn and product_key(r) != me:
            by_name.setdefault(pn, r)
    out = []
    for m in (rec.get("materials") or []):
        up = by_name.get(norm(m.get("matName")))
        if not up:
            continue
        up_full = (up.get("summary") or {}).get("全成本含税")
        price = m.get("priceIncl")
        ok = (up_full is not None and price is not None and abs(float(up_full) - float(price)) < 0.01)
        out.append({"matName": norm(m.get("matName")), "qtyPerKg": m.get("qtyPerKg"), "priceUsed": price,
                    "upCpCode": up.get("cpCode"), "upProductName": norm(up.get("productName")),
                    "upFull": up_full, "priceOk": ok, "upChecksOk": all_checks_ok(up)})
    return out


def upstream_bad_chain(rec, recs):
    """**传导式**收本品所有（含隔级）勾稽不平的上游产品名。用于连带拦截：
    复配料C不平 → 半成品S(自身平但用了C) → 成品P(自身平但用了S)，P 的 bad_chain = [C]（S 自身平但站在错数上，其上游C要报）。
    只在同一核算表工作簿(recs)内按页名精确配上下游（无歧义）；跨文件上游不在此列（属另案）。"""
    by_name = {}
    for r in recs or []:
        pn = norm(r.get("productName"))
        if pn:
            by_name.setdefault(pn, r)
    bad, seen = [], set()

    def walk(r):
        for u in upstream_refs(r, recs):
            nm = u["upProductName"]
            if nm in seen:
                continue
            seen.add(nm)
            up = by_name.get(nm)
            if not u["upChecksOk"]:          # 直接上游自身不平 → 记
                if nm not in bad:
                    bad.append(nm)
            elif up:                          # 直接上游自身平 → 继续查它的上游（隔级传导）
                walk(up)
    walk(rec)
    return bad


def failed_checks(rec):
    """不平的勾稽项 + 人话诊断。对「原料/包材小计=Σ明细」不平，进一步侦测**疑似漏加的料**——
    源表常见错：底部新增了料但「小计」公式求和范围没往下拉（申报小计恰=前 N 味之和，后面的被漏）。
    返回 [{check, a(申报), b(Σ明细), diff, missing:[{matName,costExcl}]}]。全平→[]。"""
    out = []
    for c in (rec.get("checks") or []):
        if c.get("ok"):
            continue
        a, b = c.get("a"), c.get("b")
        item = {"check": c.get("check"), "a": a, "b": b,
                "diff": round((b or 0) - (a or 0), 4) if (a is not None and b is not None) else None}
        chk = c.get("check") or ""
        if "小计=Σ明细" in chk and a is not None:
            seg = "包材" if chk.startswith("包材") else "原料"
            mats = [m for m in (rec.get("materials") or []) if m.get("seg") == seg]
            run = 0.0
            for i, m in enumerate(mats):
                run += (m.get("costExcl") or 0)
                if abs(run - a) < 0.01:      # 申报小计恰=前 i+1 味之和 → 后面的疑似被漏加
                    item["missing"] = [{"matName": x.get("matName"), "costExcl": x.get("costExcl")}
                                       for x in mats[i + 1:]]
                    break
        out.append(item)
    return out


def richness(rec):
    """排版丰富度：有编码/型号的物料数。同数字的排版差异版去重时留最全的那版——
    精简版常缺物料编码，而价格校验按编码查金蝶，留精简版会让价格校验形同虚设。"""
    mats = rec.get("materials") or []
    return sum(1 for m in mats if str(m.get("matCode") or "").strip()) \
        + sum(1 for m in mats if str(m.get("model") or "").strip())


# ---------------- 重排版核算表导出（移植 tools/build_pretty_sheet.py，参数可覆盖为复核后值）----------------
def build_pretty(rec, fee=None, approval="", formulas=True, rules=None):
    """一条记录 → 重排版核算表 xlsx 字节，对齐星期九「成本核算表（财务版本）」原版。
    formulas=True（下载）：派生数字全写 **Excel 活公式**，参数一改全表联动；formulas=False（网页预览）：同布局写计算值。
    配色（业务方定 2026-09-05）：A6A6A6＝标题+公式格；D9D9D9＝填写格；**按汇总链路着色**——汇入「变动成本合计=N14+N19」的
    原料/包材小计 N 格 FDE9D9；汇入「6、成本合计=N20+N30+N33+N34+N35」的 变动成本合计/制造小计/工厂小计/运输/装卸 N 格 FABF8F；
    变动成本合计行 A6A6A6（仅 N 格 FABF8F）；6~10 FFFF00；预估订单需求红字。
    成本不含税：按台账发票规则（rules=[{type,mode,rate}]，同 invoice_cost_excl）生成 Excel 嵌套 IF 公式引 M 列发票类型，
    M 列数据验证下拉限选配置类型——下载者改发票类型，公式自动切算法（成本会计口径不丢）。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import DataBarRule
    from openpyxl.worksheet.datavalidation import DataValidation

    F = "微软雅黑"
    INK, MUTED, BLUE = "1A1A1A", "6F7A86", "2D70C9"
    HEAD, INPUT, ORANGE, PEACH, TOTAL, RED = "A6A6A6", "D9D9D9", "FABF8F", "FDE9D9", "FFFF00", "FF0000"
    thin = Side(style="thin", color="595959")
    BOX = Border(thin, thin, thin, thin)
    FREIGHT_TAX, LOAD_TAX = 0.09, 0.06     # 运输/装卸 税率：星期九模板惯例（台账未存，按原版填）
    NUM, NUM0, PCT, QTY = "0.00", '0.00;-0.00;"-"', "0.00%", "0.0000"

    def ff(sz=10, b=False, c=INK):
        return Font(name=F, size=sz, bold=b, color=c)

    def fill(hexc):
        return PatternFill("solid", fgColor=hexc)
    LFT, CEN, RGT = Alignment("left", "center"), Alignment("center", "center"), Alignment("right", "center")
    CENW = Alignment("center", "center", wrap_text=True)
    LFTW = Alignment("left", "center", wrap_text=True)    # B 列分区大标题：左对齐、垂直居中（业务方定）

    def num(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return 0.0

    def txt(v):
        s = norm(v)
        return "" if s in ("", "0", "0.0", "—", "-", "None") else s

    def put(r, c, formula, value, fmt=None, font=None, align=None, bg=None):
        """公式/值双写：formulas 模式写 Excel 公式，否则写计算值。"""
        cell = ws.cell(r, c, formula if (formulas and formula) else value)
        if fmt:
            cell.number_format = fmt
        cell.font = font or ff(9)
        cell.alignment = align or RGT
        if bg:
            cell.fill = fill(bg)
        cell.border = BOX
        return cell

    def paint(r, c1, c2, bg):
        for c in range(c1, c2 + 1):
            ws.cell(r, c).fill = fill(bg)

    fe = {**fee_from_summary(rec.get("summary")), **(fee or {})}
    mfg, load, adm = num(fe.get("mfg")), num(fe.get("load")), num(fe.get("adm"))
    p = rec
    semi = is_semi(p.get("cpCode"))
    name = norm(p.get("productName")) or "核算表"
    order_qty = num(p.get("orderQty"))
    mats = [m for m in p.get("materials", []) if m.get("seg") == "原料"]
    packs = [m for m in p.get("materials", []) if m.get("seg") == "包材"]

    rules = rules or INVOICE_RULE_DEFAULTS                # 发票类型→成本不含税 算法（台账基础设置，成本会计口径）

    def cost_excl(m):
        return invoice_cost_excl(num(m.get("qtyPerKg")), num(m.get("priceIncl")), num(m.get("taxRate")),
                                 m.get("invoiceType") or "专票", rules)

    def excl_formula(r):
        """成本不含税 Excel 公式：按 M{r} 发票类型分支，与 invoice_cost_excl 同算法；未匹配→价税分离；ROUND 4 位同内核。"""
        I_, K_, L_, M_ = "I%d" % r, "K%d" % r, "L%d" % r, "M%d" % r
        base = "%s/(1+%s)*%s" % (K_, L_, I_)                                  # 价税分离（专票/默认）
        expr = base
        for rule in reversed([x for x in rules if x.get("mode") != "价税分离"]):
            rate, mode = float(rule.get("rate") or 0), rule.get("mode")
            if mode == "全额":
                br = "%s*%s" % (K_, I_)
            elif mode == "买价扣除":
                br = "%s*(1-%s)*%s" % (K_, rate, I_)
            elif mode == "农产品专票":
                br = "IF(%s>0.01,(%s-%s/(1+%s)*%s)*%s,%s)" % (L_, K_, K_, L_, rate, I_, base)
            else:
                continue
            expr = 'IF(%s="%s",%s,%s)' % (M_, str(rule.get("type") or ""), br, expr)
        return "=ROUND(%s,4)" % expr
    mat_excl = sum(cost_excl(m) for m in mats)
    pack_excl = sum(cost_excl(m) for m in packs)
    mfg_x = mfg / TAX_GROSS
    load_x = invoice_cost_excl(1.0, load, LOAD_TAX, "专票", rules)          # 装卸：与公式同走发票规则（专票→价税分离）
    var_x = mat_excl + pack_excl
    total_x = var_x + mfg_x + load_x
    total_incl = total_x * TAX_GROSS
    full = total_incl + adm

    def pct(v):
        return (v / full) if full else 0.0

    # ── 行号显式布局 ──
    nm, npk = len(mats), len(packs)
    R_HDR = 8
    R_MAT0 = 9;              R_MATSUB = R_MAT0 + nm
    R_PACK0 = R_MATSUB + 1;  R_PACKSUB = R_PACK0 + npk
    R_VAR = R_PACKSUB + 1
    R_FEEHDR = R_VAR + 2
    R_LAB0 = R_FEEHDR + 1;   R_LAB2 = R_LAB0 + 2
    R_MAN0 = R_LAB2 + 1;     R_MAN3 = R_MAN0 + 3
    R_MFGSUB = R_MAN3 + 1
    R_FAC0 = R_MFGSUB + 1;   R_FAC1 = R_FAC0 + 1
    R_FACSUB = R_FAC1 + 1
    R_FR = R_FACSUB + 1
    R_LD = R_FR + 1
    R6 = R_LD + 1; R7, R8, R9, R10 = R6 + 1, R6 + 2, R6 + 3, R6 + 4
    R_SRC = R10 + 1
    FULL = "$C$%d" % R10
    OQ = "$K$6"

    WIDTHS = [2.5, 21, 11.5, 13, 26, 14, 24, 10, 11, 8, 15, 7, 10, 12, 10, 28, 13]
    NC = len(WIDTHS)
    DATA_H, HEAD_H = 18, 30

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(name[:28])
    for i, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.sheet_view.showGridLines = False

    def mc(r, c1, c2):
        ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)

    def box_row(r, c1=2, c2=None):
        for c in range(c1, (c2 or NC) + 1):
            ws.cell(r, c).border = BOX

    def vlabel(c, r1, r2, text, sz=10, align=None):
        ws.merge_cells(start_row=r1, start_column=c, end_row=r2, end_column=c)
        cc = ws.cell(r1, c, text); cc.font = ff(sz, True); cc.alignment = align or CENW; cc.fill = fill(HEAD)
        for rr in range(r1, r2 + 1):
            ws.cell(rr, c).border = BOX

    # ── 标题条（标题＝A6A6A6）──
    mc(1, 2, NC)
    t = ws.cell(1, 2, "成本核算表（财务版本）｜%s" % name); t.font = ff(14, True); t.alignment = CEN
    paint(1, 2, NC, HEAD); ws.row_dimensions[1].height = 28
    mc(2, 2, NC)
    s = ws.cell(2, 2, "%s　·　来源：钉钉审批 %s　·　%s［%s］　·　核算工作台程序生成"
                % ("半成品·作原料进入成品" if semi else "成品", approval or "—", norm(p.get("srcFile")), norm(p.get("sheet"))))
    s.font = ff(9, False, "3A3A3A"); s.alignment = LFT
    paint(2, 2, NC, HEAD); ws.row_dimensions[2].height = 16

    # ── 信息区（行3-6）：标签 A6A6A6 / 值 D9D9D9 ──
    def info(r, lb, lv, rb, rv, rv_fmt=None, unit=None, rv_font=None):
        a = ws.cell(r, 2, lb); a.font = ff(10, True); a.fill = fill(HEAD); a.alignment = LFT
        mc(r, 3, 8)
        cv = ws.cell(r, 3, lv); cv.font = ff(10, True); cv.alignment = CEN
        paint(r, 3, 8, INPUT)
        mc(r, 9, 10)                                     # 右标签 I:J 合并（同原版 I3:J3…I6:J6）
        b = ws.cell(r, 9, rb); b.font = ff(10, True); b.fill = fill(HEAD); b.alignment = CEN
        if unit is None:
            mc(r, 11, NC); paint(r, 11, NC, INPUT)
        else:
            mc(r, 11, NC - 1); paint(r, 11, NC, INPUT)
            u = ws.cell(r, NC, unit); u.font = rv_font or ff(10, True); u.alignment = LFT
        rc = ws.cell(r, 11, rv); rc.font = rv_font or ff(10, True); rc.alignment = CEN
        if rv_fmt:
            rc.number_format = rv_fmt
        box_row(r)
        ws.row_dimensions[r].height = DATA_H + 2
    info(3, "供应商名称", txt(p.get("supplier")), "审核日期", txt(p.get("calcDate")))
    info(4, "产品名称", name, "包装规格", txt(p.get("packSpec")))
    info(5, "物料编码", txt(p.get("erpCode")), "研发编码", txt(p.get("cpCode")))
    info(6, "客户", txt(p.get("customer")), "预估订单需求", (order_qty if order_qty else None),
         rv_fmt="#,##0.00", unit="kg", rv_font=ff(10, True, RED))            # 预估订单需求：红字

    # ── ①变动成本 表头（A6A6A6）──
    for c, h in [(3, "类别"), (4, "物料编码"), (5, "物料名称"), (6, "型号"), (7, "规格"), (8, "品牌"),
                 (9, "添加量(kg)"), (10, "单位"), (11, "含税含运采购价\n(元/kg)"), (12, "税率"), (13, "发票类别"),
                 (14, "成本不含税\n(元/kg)"), (15, "占比"), (16, "报价说明"), (17, "订单对应用量")]:
        cell = ws.cell(R_HDR, c, h); cell.font = ff(9.5, True); cell.alignment = CENW
    paint(R_HDR, 2, NC, HEAD)
    box_row(R_HDR)
    ws.row_dimensions[R_HDR].height = HEAD_H

    def seg_block(items, r0, r_sub, cat_label):
        for i, m in enumerate(items):
            r = r0 + i
            q, pr, t = num(m.get("qtyPerKg")), num(m.get("priceIncl")), num(m.get("taxRate"))
            ce = cost_excl(m)
            oq_val = num(m.get("orderQty")) or (q * order_qty if order_qty else None)
            note = "　".join(x for x in [("起订：%s" % txt(m.get("moq"))) if txt(m.get("moq")) else "",
                                        txt(m.get("priceNote")).replace("|", "；")] if x)
            # 填写格 D9D9D9
            for c, v in [(4, txt(m.get("matCode"))), (5, txt(m.get("matName"))), (6, txt(m.get("model"))), (7, txt(m.get("spec"))),
                         (8, txt(m.get("brand"))), (10, txt(m.get("unit")) or "kg"), (13, txt(m.get("invoiceType")) or "专票"), (16, note)]:
                cell = ws.cell(r, c, v); cell.border = BOX; cell.font = ff(9, c == 5); cell.fill = fill(INPUT)
                cell.alignment = LFT if c in (5, 6, 7, 16) else CEN
            put(r, 9, None, q, QTY, bg=INPUT)
            put(r, 11, None, pr, NUM, bg=INPUT)
            put(r, 12, None, t, "0%", align=CEN, bg=INPUT)
            # 公式格 A6A6A6
            put(r, 14, excl_formula(r), ce, NUM, bg=HEAD)                      # 成本不含税：按 M 列发票类型分支
            put(r, 15, "=N%d/%s" % (r, FULL), pct(ce), PCT, bg=HEAD)
            put(r, 17, '=IF(%s="","",I%d*%s)' % (OQ, r, OQ), (oq_val if oq_val else ""), NUM, bg=HEAD)
            ws.cell(r, 3).border = BOX
            ws.row_dimensions[r].height = DATA_H
        if items:
            vlabel(3, r0, r0 + len(items) - 1, cat_label)
        # 小计行：整行 FF0000
        r = r_sub
        sub = sum(cost_excl(m) for m in items)
        rng = (r0, r0 + len(items) - 1) if items else (r0, r0)
        ws.cell(r, 3, "小计").font = ff(9.5, True); ws.cell(r, 3).alignment = CEN
        put(r, 9, "=SUM(I%d:I%d)" % rng, sum(num(m.get("qtyPerKg")) for m in items), QTY, font=ff(9.5, True))
        put(r, 14, "=SUM(N%d:N%d)" % rng, sub, NUM, font=ff(9.5, True))
        put(r, 15, "=N%d/%s" % (r, FULL), pct(sub), PCT, font=ff(9.5, True))
        paint(r, 3, NC, HEAD); ws.cell(r, 14).fill = fill(PEACH)      # 原料/包材小计 N 格 FDE9D9（汇入 变动成本合计=N14+N19），行其余 A6A6A6
        box_row(r, 3)
        ws.row_dimensions[r].height = DATA_H

    seg_block(mats, R_MAT0, R_MATSUB, "原料成本")
    seg_block(packs, R_PACK0, R_PACKSUB, "包装成本")
    # 合计行：整行 FFFF00
    ws.cell(R_VAR, 3, "变动成本合计").font = ff(10, True); ws.cell(R_VAR, 3).alignment = CEN
    put(R_VAR, 14, "=N%d+N%d" % (R_MATSUB, R_PACKSUB), var_x, NUM, font=ff(10, True))
    put(R_VAR, 15, "=N%d/%s" % (R_VAR, FULL), pct(var_x), PCT, font=ff(10, True))
    paint(R_VAR, 3, NC, HEAD); ws.cell(R_VAR, 14).fill = fill(ORANGE)   # 变动成本合计：行 A6A6A6，仅 N 格 FABF8F（汇入 6、成本合计）
    box_row(R_VAR, 3)
    ws.row_dimensions[R_VAR].height = DATA_H
    vlabel(2, R_HDR, R_VAR, "1、变动成本", align=LFTW)

    # ── 费用段表头（A6A6A6）──
    # 消耗量放 I 列与①「添加量(kg)」同列对齐（业务方定：去掉"每公斤产品"、左移一列）；P/Q 表头与①一致
    for c, h in [(3, "类别"), (5, "项目"), (9, "消耗量"), (11, "单价含税\n元/单位"), (14, "成本不含税\n(元/kg)"), (15, "占比"),
                 (16, "报价说明"), (17, "订单对应用量")]:
        cell = ws.cell(R_FEEHDR, c, h); cell.font = ff(9.5, True); cell.alignment = CENW
    paint(R_FEEHDR, 3, NC, HEAD)
    box_row(R_FEEHDR, 3)
    ws.row_dimensions[R_FEEHDR].height = HEAD_H

    def na_row(r, item):
        a = ws.cell(r, 4, "N/A"); a.font = ff(9, False, MUTED); a.alignment = CEN
        b = ws.cell(r, 5, item); b.font = ff(9); b.alignment = LFT
        paint(r, 4, 13, INPUT)                           # 填写区 D..M
        ws.cell(r, 3).fill = fill(HEAD)                  # 类别列 C 无子类别文字也铺灰（同原版整条灰带，避免白洞）
        ws.cell(r, 16).fill = fill(INPUT)                # P 报价说明＝填写格 D9D9D9（与①一致）
        ws.cell(r, 17).fill = fill(HEAD)                 # Q 订单对应用量＝公式格 A6A6A6（与①一致）
        box_row(r, 3)
        ws.row_dimensions[r].height = DATA_H

    def zero_row(r):
        put(r, 14, None, 0.0, NUM0, bg=HEAD)
        put(r, 15, "=N%d/%s" % (r, FULL), 0.0, PCT, bg=HEAD)

    def subtotal(r, formula, value):
        ws.cell(r, 3, "小计").font = ff(9.5, True); ws.cell(r, 3).alignment = CEN
        put(r, 14, formula, value, NUM0, font=ff(9.5, True))
        put(r, 15, "=N%d/%s" % (r, FULL), pct(value), PCT, font=ff(9.5, True))
        paint(r, 3, NC, HEAD); ws.cell(r, 14).fill = fill(ORANGE)     # 制造/工厂费用小计 N 格 FABF8F（汇入 6、成本合计），行其余 A6A6A6
        box_row(r, 3)
        ws.row_dimensions[r].height = DATA_H

    def fee_tax_cols(r, tax):
        put(r, 9, None, 1.0, QTY, bg=INPUT)              # 消耗量在 I 列（与①添加量同列）
        put(r, 12, None, tax, "0%", align=CEN, bg=INPUT)
        cell = ws.cell(r, 13, "专票"); cell.alignment = CEN; cell.font = ff(9); cell.border = BOX; cell.fill = fill(INPUT)

    # 2、制造费用
    for i, item in enumerate(["生产人员数量", "月均工资", "产量（kg/天）"]):
        na_row(R_LAB0 + i, item)
    vlabel(3, R_LAB0, R_LAB2, "人工成本", sz=9.5)
    for c, formula, value, fmt, bg in [(11, None, mfg, NUM, INPUT),                          # K 单价含税＝加工费（填写）
                                       (14, "=K%d/1.13" % R_LAB0, mfg_x, NUM, HEAD),          # N 公式格 A6A6A6
                                       (15, "=N%d/%s" % (R_LAB0, FULL), pct(mfg_x), PCT, HEAD)]:
        ws.merge_cells(start_row=R_LAB0, start_column=c, end_row=R_LAB2, end_column=c)
        put(R_LAB0, c, formula, value, fmt, bg=bg)
        for rr in range(R_LAB0, R_LAB2 + 1):
            ws.cell(rr, c).border = BOX
    for i, item in enumerate(["水-m³", "电-度", "燃气", "折旧"]):
        na_row(R_MAN0 + i, item); zero_row(R_MAN0 + i)
    vlabel(3, R_MAN0, R_MAN3, "制造成本", sz=9.5)
    subtotal(R_MFGSUB, "=N%d+SUM(N%d:N%d)" % (R_LAB0, R_MAN0, R_MAN3), mfg_x)
    vlabel(2, R_FEEHDR, R_MFGSUB, "2、制造费用", align=LFTW)

    # 3、工厂费用
    for i, item in enumerate(["管理费用", "厂商利润"]):
        na_row(R_FAC0 + i, item); zero_row(R_FAC0 + i)
    subtotal(R_FACSUB, "=SUM(N%d:N%d)" % (R_FAC0, R_FAC1), 0.0)
    vlabel(2, R_FAC0, R_FACSUB, "3、工厂费用", align=LFTW)

    # 4、运输费用 / 5、装卸费：填写行——填写格 D9D9D9、公式格 A6A6A6（B 标签 A6A6A6），不铺黄
    def side_label(r, text):
        a = ws.cell(r, 2, text); a.font = ff(10, True); a.fill = fill(HEAD); a.alignment = LFT; a.border = BOX

    na_row(R_FR, "运输费用"); fee_tax_cols(R_FR, FREIGHT_TAX)
    put(R_FR, 14, excl_formula(R_FR), 0.0, NUM0, bg=ORANGE)
    put(R_FR, 15, "=N%d/%s" % (R_FR, FULL), 0.0, PCT, bg=HEAD)
    side_label(R_FR, "4、运输费用")

    na_row(R_LD, "装卸费用"); fee_tax_cols(R_LD, LOAD_TAX)
    put(R_LD, 11, None, load, NUM, bg=INPUT)
    put(R_LD, 14, excl_formula(R_LD), load_x, NUM, font=ff(9, True), bg=ORANGE)
    put(R_LD, 15, "=N%d/%s" % (R_LD, FULL), pct(load_x), PCT, bg=HEAD)
    side_label(R_LD, "5、装卸费")

    # 发票类型下拉（M 列）：选项＝台账配置的发票类型；改类型后成本不含税公式自动切算法
    dv = DataValidation(type="list", allow_blank=True,
                        formula1='"%s"' % ",".join(str(x.get("type") or "") for x in rules if x.get("type")))
    dv.errorTitle, dv.error = "发票类型", "请选择台账「基础设置」里配置的发票类型"
    ws.add_data_validation(dv)
    if nm:
        dv.add("M%d:M%d" % (R_MAT0, R_MAT0 + nm - 1))
    if npk:
        dv.add("M%d:M%d" % (R_PACK0, R_PACK0 + npk - 1))
    dv.add("M%d:M%d" % (R_FR, R_LD))

    # ── 6~10：B 标签 A6A6A6 / C:N 合并值 + O + P 整行 FFFF00 ──
    summ = [(R6, "6、成本合计（不含税）", "=N%d+N%d+N%d+N%d+N%d" % (R_VAR, R_MFGSUB, R_FACSUB, R_FR, R_LD), total_x, None),
            (R7, "7、增值税合计", "=C%d*0.13" % R6, total_x * 0.13, "＝不含税 × 13%"),
            (R8, "8、成本合计（含税）", "=C%d*1.13" % R6, total_incl, "＝不含税 × 1.13"),
            (R9, "9、管理费（含税）", None, adm, "台账固定加成（参数）"),
            (R10, "10、全成本（含税）", "=C%d+C%d" % (R8, R9), full, "＝ 8 + 9")]
    for r, lb, formula, value, note in summ:
        last = (r == R10)
        side_label(r, lb)
        ws.cell(r, 2).font = ff(10, True, BLUE if last else INK)
        mc(r, 3, 14)
        put(r, 3, formula, value, NUM, font=ff(12 if last else 10.5, True, BLUE if last else INK), align=CEN)
        put(r, 15, "=C%d/%s" % (r, FULL), (1.0 if last else pct(value)), PCT, font=ff(9.5, True))
        mc(r, 16, NC)
        nc_ = ws.cell(r, 16, note or ""); nc_.font = ff(8.5, False, "5A5A5A"); nc_.alignment = LFT
        paint(r, 3, NC, TOTAL)
        box_row(r)
        ws.row_dimensions[r].height = DATA_H + 2

    src_full = num((p.get("summary") or {}).get("全成本含税"))
    if src_full:
        a = ws.cell(R_SRC, 2, "源表全成本（对账）"); a.font = ff(9, False, MUTED); a.alignment = LFT
        put(R_SRC, 3, None, src_full, NUM, font=ff(9, False, MUTED), align=CEN)
        mc(R_SRC, 16, NC)
        put(R_SRC, 16, '="差异 "&TEXT(C%d-C%d,"0.00")&"（费用参数与源表不同时才有差）"' % (R10, R_SRC),
            "差异 %+.2f（费用参数与源表不同时才有差）" % (full - src_full), font=ff(8.5, False, MUTED), align=LFT)
        ws.cell(R_SRC, 16).border = Border()
        ws.row_dimensions[R_SRC].height = DATA_H

    ws.conditional_formatting.add("O%d:O%d" % (R_MAT0, R10),
                                  DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color="63BE7B", showValue=True))

    # ── 脚注 ──
    r = R_SRC + 2
    for tnote in [
        "口径：①成本不含税按发票类型取算法（专票价税分离÷(1+税率)／普票全额／自产自销农产品买价扣除／农产品专票先分离再扣）×添加量，发票类型可下拉改、公式自动切换；②加工费不含税＝含税÷1.13，装卸不含税＝含税÷1.06；成本合计(含税)＝不含税合计×1.13。",
        "③本表为「含税不含运」标准成本＝原料+包材+加工费+装卸+管理费（含税）；制造/工厂/运输为原版分区，台账未拆存者以 N/A 占位、加工费归入制造费用。",
        "④灰底(D9D9D9)＝填写格，深灰(A6A6A6)＝标题与公式格；改加工费/装卸单价/管理费/预估订单需求，全表公式联动。"
        + "　生成：核算工作台·BOM报价审核　审批 %s　核算日期 %s" % (approval or "—", txt(p.get("calcDate")) or "—"),
    ]:
        mc(r, 2, NC)
        ws.cell(r, 2, tnote).font = ff(8.5, False, "8A94A0"); ws.cell(r, 2).alignment = LFT
        r += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------- 研发 BOM清单 解析 + 自洽校验（核算表 vs BOM清单：用量/结构）----------------
# BOM清单 与 成本核算表 同为研发出品：前者是「官方物料清单」（结构+用量、无价），后者是「成本核算版」（用量+研发填的价）。
# 成本会计需确认这两份研发文件逐料对得上——用量不符 / 核算表缺料 / 核算表多料 都要标出来（缺料多料=BOM结构不一致）。
# 采购侧数据暂不开放，故当前只做研发内部自洽；比对引擎与将来接采购同一套（换一个源即可升级成 研发vs采购）。
_BOM_QTY_TOL_ABS = 1e-4        # 用量绝对容差（防两份文件显示位数不同造成的伪差异）
_BOM_QTY_TOL_REL = 0.005       # 用量相对容差 0.5%


def _parse_bom_sheet(ws):
    """解析 BOM清单 的一个 sheet（成品页「BOM清单」或半成品页「复合调味料」）→ {sheet, productName, materials} 或 None。"""
    grid = [[c.value for c in row] for row in ws.iter_rows()]
    R = len(grid)
    C = max((len(r) for r in grid), default=0)

    def cell(r, c):
        return grid[r][c] if 0 <= r < R and 0 <= c < len(grid[r]) else None

    hdr = None
    for r in range(min(R, 20)):
        texts = [norm(cell(r, c)).replace("\n", "") for c in range(C)]
        if any(t == "用量" for t in texts) and any(("原料名称" in t or "物料名称" in t) for t in texts):
            hdr = r
            break
    if hdr is None:
        return None
    m = {}
    for c in range(C):
        t = norm(cell(hdr, c)).replace("\n", "")
        # ⚠ 型号/规格/名称先判：否则「原料型号」会被下面「原料…号」的编码分支吃掉（"型号"里含"号"），
        #    实测漏读整列型号（G100/GT-1200/规格尺寸…全空，业务方 2026-09-04 发现）。
        if "物料类型" in t:
            m["type"] = c
        elif t.startswith("原料型号") or t == "型号":
            m["model"] = c
        elif "规格" in t:
            m["spec"] = c
        elif "原料名称" in t or "物料名称" in t:
            m["name"] = c
        elif t.startswith("原料") and ("号" in t):     # 原料/复配料号、原料号（型号已在上面拦掉）
            m.setdefault("code", c)
        elif "物料编码" in t or "包材编码" in t:
            m.setdefault("code", c)
        elif t == "单位":
            m["unit"] = c
        elif "品牌" in t or "供应商" in t:
            m["brand"] = c
        elif t == "用量":
            m["qty"] = c
    if "name" not in m or "qty" not in m:
        return None
    hdr_info = {}
    _WANT = {"产品中文名称": "productName", "产品名称": "productName", "客户": "customer",
             "客  户": "customer", "客     户": "customer", "产品编号": "cpCode"}
    for r in range(min(R, 10)):
        for c in range(C):
            t = norm(cell(r, c)).replace("\n", "").replace(" ", "")
            key = _WANT.get(t)
            if key and key not in hdr_info:
                for cc in range(c + 1, min(c + 4, C)):
                    v = norm(cell(r, cc))
                    if v:
                        hdr_info[key] = v
                        break
    pname = hdr_info.get("productName") or ws.title.strip()
    mats = []
    for r in range(hdr + 1, R):
        name = norm(cell(r, m["name"]))
        qty = num(cell(r, m.get("qty", -1)))
        if not name or qty is None or name in ("得率", "合计", "小计"):
            continue
        mtype = norm(cell(r, m.get("type", -1)))
        mats.append({
            "seg": "包材" if "包材" in mtype else "原料",
            "matType": mtype,
            "matCode": norm(cell(r, m.get("code", -1))),
            "matName": name,
            "model": norm(cell(r, m.get("model", -1))),
            "spec": norm(cell(r, m.get("spec", -1))),
            "unit": norm(cell(r, m.get("unit", -1))),
            "brand": norm(cell(r, m.get("brand", -1))),
            "qty": qty,
        })
    if not mats:
        return None
    return {"sheet": ws.title.strip(), "productName": pname,
            "customer": hdr_info.get("customer", ""), "cpCode": hdr_info.get("cpCode", ""),
            "materials": mats}


def _parse_craft_sheet(ws):
    """解析 BOM 文件里的「工艺流程」sheet（实证：研发 BOM 工作簿含该页，工序+工艺细节两列）。
    → {steps:[{step, detail}], head:{产品编号/客户/审核人…}, imageCount} 或 None（不是工艺流程页）。
    ⚠ 工艺细节里可能嵌图片（工艺照片）——文本照常解析，图片只报数量，看图请下载 BOM 原件。"""
    grid = [[c.value for c in row] for row in ws.iter_rows()]
    R = len(grid)
    C = max((len(r) for r in grid), default=0)

    def cell(r, c):
        return grid[r][c] if 0 <= r < R and 0 <= c < len(grid[r]) else None

    hdr = None
    for r in range(min(R, 25)):
        texts = [norm(cell(r, c)).replace("\n", "").replace(" ", "") for c in range(C)]
        if any(t == "工艺流程" for t in texts) and any("工艺细节" in t or "细节" == t for t in texts):
            hdr = r
            break
    if hdr is None:
        return None
    c_step = c_detail = None
    for c in range(C):
        t = norm(cell(hdr, c)).replace("\n", "").replace(" ", "")
        if t == "工艺流程" and c_step is None:
            c_step = c
        elif ("工艺细节" in t or t == "细节") and c_detail is None:
            c_detail = c
    if c_step is None or c_detail is None:
        return None
    # 表头信息（产品编号/客户/审核人等，在工艺表上方）
    head, _WANT = {}, {"产品中文名称": "productName", "产品编号": "cpCode", "客户": "customer",
                       "审核人": "reviewer", "批准人": "approver", "开发者（产品负责人）": "developer",
                       "本次编写时间": "writtenAt", "上次修订时间": "revisedAt", "执行标准": "std"}
    for r in range(min(hdr, 12)):
        for c in range(C):
            t = norm(cell(r, c)).replace("\n", "").replace(" ", "")
            key = _WANT.get(t)
            if key and key not in head:
                for cc in range(c + 1, min(c + 4, C)):
                    v = norm(cell(r, cc))
                    if v:
                        head[key] = v
                        break
    # ⚠ 合并单元格错位：表头「工艺流程」在 C2、值却落在 C3（实证 251965）。故按表头列**向右开窗**找值，别死用表头列。
    def win(r, c0, c1):
        for c in range(c0, min(c1, C)):
            v = norm(cell(r, c))
            if v:
                return v
        return ""

    steps = []
    for r in range(hdr + 1, R):
        st = win(r, c_step, c_detail)                 # 工序名：表头列 → 细节列之间
        dt = win(r, c_detail, c_detail + 6)           # 工艺细节：细节列起向右几列
        if not st and not dt:
            continue
        if st in ("工艺流程", "工艺细节"):
            continue
        if steps and not st and dt:                   # 细节续行 → 并进上一道工序
            steps[-1]["detail"] = (steps[-1]["detail"] + "\n" + dt).strip()
            continue
        if st:
            steps.append({"step": st, "detail": dt})
    if not steps:
        return None
    try:
        n_img = len(getattr(ws, "_images", []) or [])
    except Exception:
        n_img = 0
    return {"steps": steps, "head": head, "imageCount": n_img, "sheet": ws.title.strip()}


def parse_craft(data, src_filename=None):
    """吃 BOM xlsx 字节/路径 → 工艺流程 {steps, head, imageCount} 或 None。
    ⚠ 不用 read_only（要数图片）；文件大时略慢，只在解析 BOM 时调一次。"""
    src = data if isinstance(data, (bytes, bytearray)) else open(data, "rb").read()
    try:
        wb = load_workbook(io.BytesIO(src), data_only=True)
    except Exception:
        return None
    out = None
    for ws in wb.worksheets:
        if "工艺" in (ws.title or ""):
            out = _parse_craft_sheet(ws)
            if out:
                break
    if out is None:                     # 页名没写「工艺」时兜底全扫
        for ws in wb.worksheets:
            out = _parse_craft_sheet(ws)
            if out:
                break
    try:
        wb.close()
    except Exception:
        pass
    return out


def parse_bom_list(data, src_filename=None):
    """吃 BOM清单 xlsx 字节/路径 → [{sheet, productName, materials}]（成品页 + 复合调味料页各一条）。
    非 BOM清单文件（找不到「用量」表头）返回 []。"""
    src = data if isinstance(data, (bytes, bytearray)) else open(data, "rb").read()
    wb = load_workbook(io.BytesIO(src), data_only=True, read_only=True)
    out = []
    for ws in wb.worksheets:
        e = _parse_bom_sheet(ws)
        if e:
            out.append(e)
    wb.close()
    return out


def match_bom_entry(rec, bom_lists):
    """同 match_bom_list，但返回**整条 BOM 条目**（含 materials 与该文件的 craft 工艺流程）。找不到 → None。
    ⚠ **CP 码精确匹配放最前**（业务方 2026-09-04 实证 059207）：同名同客户的两个**包装版本**——
    「空白袋+工厂打印标签」(CP…) 与「印刷袋」(CP…-2)，归组键(名+客户)会撞、把 -2 配到空白袋那份 BOM → 假报缺/多料。
    CP 码(…vs …-2)是唯一能区分两版的键，必须先按它配。"""
    pn = norm(rec.get("productName"))
    if not pn:
        return None
    bl = bom_lists or []
    cp = norm(rec.get("cpCode"))
    if cp:                                          # ① CP码精确对齐（区分同名不同包装版）
        for b in bl:
            if norm(b.get("cpCode")) == cp:
                return b
    pk = product_key(rec)
    for b in bl:                                    # ② 产品身份键兜底（CP 空时按名+客户；同口径 product_key）
        if product_key(b) == pk:
            return b
    for b in bl:                                    # ③ 产品名/页名对齐
        if norm(b.get("productName")) == pn or norm(b.get("sheet")) == pn:
            return b
    return None


def match_bom_list(rec, bom_lists):
    """给一条核算表记录，从**同一审批批次**的 BOM清单 里找对得上的那张物料清单。找不到 → None。
    ⚠ 只在同批（一单一版）里配对：跨版本同产品各有各的 BOM清单，靠调用方按批传入隔离，本函数不判版本。
    对齐用归组键(产品名,客户)——同日同名双产品(木薯 鱼你/通品)靠客户区分；半成品按名兜底。
    ⚠ 不做「唯一清单兜底」：一单多产品但只附了一份 BOM清单时（实测 202609011151000012138 有此情况），
    兜底会把清单错配给没有清单的产品 → 假「有差异」。宁可返回 None 显「未附BOM清单」，也不乱配。"""
    b = match_bom_entry(rec, bom_lists)
    return b["materials"] if b else None


# ---------------- 来源方判定 + 成本会计商品版 价/税 diff ----------------
# 一份 BOM报价审批里的数据来自不同角色，成本会计要一眼知道每条数据谁出的（可信度不同）：
#   研发BOM(research)：研发出的 BOM清单/原料信息（用量/结构，无价）；
#   采购商务版(procurement)：采购在「商务输出」槽上传的成本核算表（全量，含供应商/型号/规格/实价/发票税率）——复核底稿；
#   成本会计商品版(costacct)：成本会计在「商品版本」槽输出的脱敏公开版（删了型号/规格/供应商三列，可能调过价/税）；
#   手工上传(manual)/评论区上传(comment)：渠道兜底。
# 判定优先级：控件标注含「商品版」→costacct、含「商务」→procurement；否则按解析类型（BOM清单→research、核算表→procurement）。
ORIGIN_LABELS = {"research": "研发BOM", "procurement": "采购商务版", "costacct": "成本会计商品版",
                 "manual": "手工上传", "comment": "评论区上传"}


def origin_from_label(label, source_type="", is_bom_list=False):
    """按钉钉控件标注 + 渠道 + 解析类型，判来源方代码（见 ORIGIN_LABELS）。"""
    lb = str(label or "")
    if "商品版" in lb:
        return "costacct"
    if "商务" in lb:
        return "procurement"
    if source_type == "dingtalk_comment":
        return "comment"
    if source_type == "manual_upload":
        return "manual"
    if is_bom_list:
        return "research"
    # 表单里的核算表但无「商务/商品」标注：核算表本就是采购/工厂侧产物，归采购商务版
    return "procurement"


def _pair_goods(base_mats, goods_mats):
    """把商务版底稿逐料对到商品版对应料。商品版=商务版删列不删行、行序一致 → **主用行位置对齐**，
    物料名一致时采信；名字对不上（行序意外错动）则退回按名字全表查找。
    ⚠ 不能按物料编码对齐：源表编码列常是「103系列」这类系列占位，非唯一。返回 [(base_mat, goods_mat|None)]。"""
    gm = goods_mats or []
    used = set()
    pairs = []
    for i, m in enumerate(base_mats or []):
        nm = norm(m.get("matName"))
        g = None
        if i < len(gm) and norm(gm[i].get("matName")) == nm and i not in used:
            g, gi = gm[i], i
        else:
            gi = next((j for j, x in enumerate(gm)
                       if j not in used and nm and norm(x.get("matName")) == nm), None)
            g = gm[gi] if gi is not None else None
        if g is not None:
            used.add(gi)
        pairs.append((m, g))
    return pairs


def diff_goods(base_mats, goods_mats):
    """成本会计商品版 vs 采购商务版 的**价/税**差异（商品版可能调过单价或税率）。逐料按 _pair_goods 对齐。
    返回 {rows:[{matCode,matName,field,from,to}], hasDiff, count}。无差异→hasDiff=False（静默过，不弹确认闸）。"""
    rows = []
    for m, g in _pair_goods(base_mats, goods_mats):
        if not g:
            continue
        for field, la in (("priceIncl", "含税采购价"), ("taxRate", "税率")):
            a, b = m.get(field), g.get(field)
            if a is None or b is None:
                continue
            if abs(float(a) - float(b)) > 1e-9:
                rows.append({"matCode": norm(m.get("matCode")), "matName": norm(m.get("matName")),
                             "field": field, "fieldLabel": la, "from": a, "to": b})
    return {"rows": rows, "hasDiff": bool(rows), "count": len(rows)}


def _real_code(m):
    """物料真实编码：非空、且**不是占位码**（如「402系列」「108系列」——研发临时占位、非唯一）→ 返回；否则 None。
    实证 251965/403530 源表：占位「XX系列」会撞多料、且同料在核算表是真码、在研发BOM是占位（梅子粒 108010128 vs 108系列），
    所以占位码不能当身份。"""
    c = norm(m.get("matCode"))
    return None if (not c or "系列" in c) else c


def compare_bom(calc_mats, bom_mats):
    """核算表(采购·添加量) vs BOM清单(研发·用量) 逐料自洽校验。业务方 2026-09-04 定权责：
    **编码取研发、分类(原料/包材)取采购**。
    · **编码**：优先研发 BOM 的真实编码；研发是占位(XX系列)/空 → **回退采购核算表的真实编码**并标 `codeNote=研发编码待补`
      （下游查金蝶实采价/BOM反查要真码）。两边都无真码 → 显占位/空、同样标待补。
    · **分类**：段一律取**采购(核算表)**为准；研发归了不同段（如纱布包袋研发记原料、采购记包材）只在 `segNote`
      灰字提示、不算差异、不拦定稿。
    对齐键：真实编码两侧都出现 → 按编码（同名不同码如「大豆分离蛋白」101020006/101200006 正确分开）；否则按名字
    （占位/空码/单侧缺对码，如梅子粒 采购 108010128 vs 研发 108系列）。同键多行按用量求和（审查 L13 食盐 3+2 vs 5）。"""
    calc, bom = list(calc_mats or []), list(bom_mats or [])    # calc=采购(核算表), bom=研发(BOM清单)
    ccodes = {c for c in (_real_code(m) for m in calc) if c}
    bcodes = {c for c in (_real_code(m) for m in bom) if c}
    shared = ccodes & bcodes            # 只有两侧都出现的真实编码，才够格当「跨表身份」

    def key_of(m):
        rc = _real_code(m)
        if rc and rc in shared:
            return ("code", rc)                                 # 编码优先对齐
        base = norm(m.get("matName")) or rc or norm(m.get("matCode"))
        return ("name", base) if base else None                 # 退回名字（占位/空码/单侧缺对码）

    def agg(mats, qf):
        d = {}
        for m in mats:
            k = key_of(m)
            if k is None:
                continue
            cur = d.setdefault(k, {"name": m.get("matName") or m.get("matCode") or "",
                                   "realCode": None, "rawCode": "", "segs": set(), "qty": 0.0, "n": 0})
            rc = _real_code(m)
            if rc and not cur["realCode"]:
                cur["realCode"] = rc
            if not cur["rawCode"]:
                cur["rawCode"] = norm(m.get("matCode")) or ""
            if m.get("seg"):
                cur["segs"].add(m.get("seg"))
            v = qf(m)
            if v is not None:
                cur["qty"] += v
            cur["n"] += 1
        return d

    calc_ix = agg(calc, lambda m: m.get("qtyPerKg"))            # 采购
    bom_ix = agg(bom, lambda m: m.get("qty"))                   # 研发
    rows = []
    order = list(calc_ix.keys()) + [k for k in bom_ix if k not in calc_ix]  # 核算表在先、BOM独有在后
    for k in order:
        c, d = calc_ix.get(k), bom_ix.get(k)                    # c=采购, d=研发
        a = c["qty"] if c else None
        bq = d["qty"] if d else None
        csegs = c["segs"] if c else set()
        bsegs = d["segs"] if d else set()
        seg = (sorted(csegs) or sorted(bsegs) or [""])[0]      # 分类取采购(核算表)为准
        seg_note = ""
        if csegs and bsegs and csegs != bsegs:                  # 研发归了别的段 → 灰字提示、不当差异
            seg_note = "研发BOM记「%s」· 分类以采购为准" % "/".join(sorted(bsegs))
        dev_rc = d["realCode"] if d else None                   # 研发真码
        pur_rc = c["realCode"] if c else None                   # 采购真码
        if dev_rc:
            code, code_note = dev_rc, ""                        # 取研发真码
        elif pur_rc:
            code, code_note = pur_rc, "研发编码待补"             # 研发占位/空 → 回退采购真码
        else:
            code = (d or c or {}).get("rawCode") or ""          # 两边都无真码 → 显占位/空
            code_note = "研发编码待补" if code else ""
        if a is not None and bq is not None:
            diff = round(a - bq, 6)
            mism = abs(a - bq) > max(_BOM_QTY_TOL_ABS, _BOM_QTY_TOL_REL * max(abs(a), abs(bq)))
            st = "用量不符" if mism else "一致"
        elif bq is not None:
            diff, st = None, "核算表缺料"
        else:
            diff, st = None, "核算表多料"
        rows.append({"matName": (c or d)["name"], "matCode": code, "codeNote": code_note,
                     "matchBy": "编码" if k[0] == "code" else "名称", "seg": seg, "segNote": seg_note,
                     "calcQty": a, "bomQty": bq, "diff": diff, "status": st})
    summ = {"total": len(rows),
            "consistent": sum(1 for r in rows if r["status"] == "一致"),
            "qtyMismatch": sum(1 for r in rows if r["status"] == "用量不符"),
            "missing": sum(1 for r in rows if r["status"] == "核算表缺料"),
            "extra": sum(1 for r in rows if r["status"] == "核算表多料"),
            "segDiff": sum(1 for r in rows if r["segNote"]),
            "codeTBD": sum(1 for r in rows if r["codeNote"])}
    summ["ok"] = summ["qtyMismatch"] == 0 and summ["missing"] == 0 and summ["extra"] == 0
    return {"rows": rows, "summary": summ}

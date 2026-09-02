# -*- coding: utf-8 -*-
# [Change Log]
# Date: 2026-08-29 | Author: Claude / c | Version: V2.354
# Description: 【临时工考勤】从钉钉取打卡数据，产出一张**与人力手工导出格式完全一致**的打卡表。
#
#   为什么绕这一道、不直接把数据塞进内核：
#     取数结果要能人眼核对。生成一张跟人力导出的那张长得一样的 Excel，
#     使用者可以下载来逐格比对，也可以拿它走原来的上传流程——
#     下游 parse_punch / 留档 / 导出 / 重跑 一行都不用改。
#     「系统取数留 API + 手工上传双模式」是这条线一开始就定的规矩。
#
#   实证记录（2026-08-29，全部在生产钉钉上跑过）：
#     · 打卡归属与 Excel 完全一致：钉钉 work_date=25 返回的正是 Excel 25 日格子里那串
#       （含次日 08:09）。跨零点的夜班不会错位。
#     · ⚠⚠ 取 **check_record_list**（原始打卡流水），不是 attendance_result_list（考勤结果）。
#       同一个返回里两者都有，差别是致命的：考勤结果每班只给上/下班两条，中间的宵夜卡、
#       多次刷卡全没有。而中间那几次是**承重的**——朱普奇 6/6 原始「00:10 00:21 08:34 19:17 23:46」
#       正是靠中间的 08:34/19:17 才切得出白班；只留首末，内核看到一段 23.6 小时的荒谬跨度，算不出班次。
#       实测：拿人力原始表只保留每天首末两次，红灯从 17 条暴涨到 411 条。
#       check_record_list 与人力导出**逐格一致**（6/6 抽样全中，含跨零点的夜班）。
#     · ⚠ 不能用 listschedule 批量取。它返回的是**排班计划时间**不是实际打卡：
#       陈志元 6/25 计划「25日12:00→26日00:00」，实际「25日12:38→26日08:09」，下班差 8 小时。
#       抽样时若只抽到「不定时打卡」组的人会看到计划＝实际，从而误判可用——踩过，别再试。
#     · 逐人逐日 getupdatedata 是唯一可靠的口。4 线程实测 30 次/秒且不触发限流，
#       6 线程开始出 errcode 88。一个月约 1 万次调用、5 分钟。
#     · ⚠⚠ 返回里混着**系统占位记录**，必须滤掉，否则会凭空造出在厂时长：
#       没打卡的日子钉钉会补一条 source_type=SYSTEM / time_result=NotSigned 的记录，
#       时间取排班计划值（09:00、18:00 这种整点）。丁玲 2026-06 全月 57 条里有 19 条是这种，
#       害她凭空多出 9 个「在厂 9 小时」的日子；11 日更糟——真打卡 19:57 被占位的 09:00 顶掉，
#       在厂时长直接从 4.6h 变成 15.6h。判据用 time_result != NotSigned：
#       既滤掉占位，又保留审批补卡（补卡是人力认可的出勤，Excel 导出里也在）。
#     · 钉钉只保留最近 180 天：2026-08-29 实测 2026-03-05 可查、2026-03-01 报 850002。
#       所以这条路只能**当月取、及时取**，补不了太老的账。
#     · 重名的人靠**手机号**区分（智能人事 sys00-mobile，在职离职都给，实测覆盖 99%、
#       全公司抽 600 人零重号）。2026-06 的 46 组重名里：
#         37 组手机号各异 → 真的是不同的人，要定人；
#          8 组手机号相同 → **同一个人在钉钉有两个账号**（离职再入职会新建 userid），
#            必须把两个账号的打卡**合并**，否则丢掉其中一个账号那几天的记录；
#          1 组缺手机号 → 定不了，交人工。
#       user/get 不返回 mobile，只能走智能人事那个口；临时工没有工号（在职样本 0/25），
#       所以工号这条路走不通。

import re
import threading
import time
from calendar import monthrange
from datetime import date, timedelta, datetime
from concurrent.futures import ThreadPoolExecutor
from io import BytesIO

import requests

try:
    from openpyxl import Workbook
except ImportError:                                   # pragma: no cover
    Workbook = None

import db
import notifier

_OAPI = "https://oapi.dingtalk.com/"
# 并发档位实测（2026-08-29）：1 线程 7.6 次/秒｜3 线程 23.7｜4 线程 30.8（零限流）
# ｜6 线程 37.5（6 次退避重试、零失败）｜8 线程起大面积 errcode 88。
# 取 6：限流由 _api 退避重试兜住，比 4 快约 1/4。
WORKERS = 6

# ⚠ 试过用 listschedule 预筛「这天谁有打卡」来少打电话，**没用，反而更慢**（401 秒 vs 357 秒）：
#   它列的是当天**有排班**的人，固定班组的人不打卡也有排班条目，
#   8,550 条只筛掉 650 条，却先花了约 1,800 次去问。别再试这条路。
_ERR_BUSY = (88, 90018)        # 限流，退避重试
_local = threading.local()


class DingError(RuntimeError):
    pass


def configured():
    return bool(notifier.load_dingtalk_conf())


def _sess():
    s = getattr(_local, "s", None)
    if s is None:
        s = _local.s = requests.Session()
    return s


def _api(path, body, tries=5, result_key="result"):
    """POST 一个 oapi 接口。限流自动退避重试；其它错误原样抛出，不吞。
    result_key：取结果的顶层字段名——多数接口是 result，listRecord 的记录在 recordresult。"""
    conf = notifier.load_dingtalk_conf()
    if not conf:
        raise DingError("钉钉未配置：conf.ini 缺 [dingtalk] 段的 appkey/appsecret/agentid")
    tok = notifier._dt_token(conf)
    if not tok:
        raise DingError("取钉钉 access_token 失败，请检查 appkey/appsecret")
    last = ""
    for i in range(tries):
        try:
            r = _sess().post(_OAPI + path, params={"access_token": tok},
                             json=body, timeout=40).json()
        except Exception as e:                          # 网络抖动
            last = str(e)
            time.sleep(0.4 * (i + 1))
            continue
        code = r.get("errcode")
        if code == 0:
            return r.get(result_key)
        last = f"errcode={code} {r.get('sub_msg') or r.get('errmsg')}"
        if code in _ERR_BUSY and "权限" not in str(r.get("sub_msg") or ""):
            time.sleep(0.3 * (i + 1))                   # 限流：退避重试
            continue
        raise DingError(f"{path} 调用失败：{last}")
    raise DingError(f"{path} 重试 {tries} 次仍失败：{last}")


# ==================== 权限体检 ====================
_CHECKS = (
    ("通讯录·部门树", "topapi/v2/department/listsub", {"dept_id": 1},
     "qyapi_get_department_list", "读不到部门，就建不出花名册"),
    ("通讯录·部门成员", "topapi/user/listsimple", {"dept_id": 1, "cursor": 0, "size": 1},
     "qyapi_get_department_member", "读不到在职员工姓名"),
    ("智能人事·离职名单", "topapi/smartwork/hrm/employee/querydimission", {"offset": 0, "size": 1},
     "qyapi_hrm_read_user", "临时工流动大，离职的占花名册 87%；不开这项，历史月份一半人对不上名字"),
    ("考勤·打卡明细", "topapi/attendance/getupdatedata",
     {"userid": "__probe__", "work_date": None}, "", "取不到打卡就没法用"),
)

MAX_BACK_DAYS = 175        # 钉钉只留最近 180 天，留几天余量


def _probe_date():
    """体检要用**最近**的日期：写死一个旧日期必然撞上「禁止查询半年以前」，
    把一条本来通着的接口报成不通（自己踩过）。"""
    return (date.today() - timedelta(days=1)).strftime("%Y-%m-%d 00:00:00")


def month_reachable(month):
    """这一期还能不能从钉钉取。返回 (能否, 说明)。"""
    try:
        y, mo = (int(x) for x in str(month).split("-"))
        last = date(y, mo, monthrange(y, mo)[1])
    except Exception:
        return False, f"期次「{month}」不是 YYYY-MM"
    gap = (date.today() - date(y, mo, 1)).days
    if gap > MAX_BACK_DAYS:
        return False, (f"钉钉只保留最近约半年的考勤，{month} 已超出（距今 {gap} 天）。"
                       f"这一期只能用人力导出的打卡表。往后请当月取、别拖。")
    if last > date.today():
        return True, f"{month} 尚未结束，取到今天为止"
    return True, ""


def probe():
    """逐项体检，返回每项通不通。给页面显示「差哪一项、找谁开」用。"""
    if not configured():
        return {"ok": False, "配置": False, "项": [],
                "说明": "conf.ini 里没有 [dingtalk] 段，本机没配钉钉"}
    conf = notifier.load_dingtalk_conf()
    out, ok = [], True
    for name, path, body, scope, why in _CHECKS:
        body = dict(body)
        if body.get("work_date") is None and "work_date" in body:
            body["work_date"] = _probe_date()
        try:
            _api(path, body, tries=2)
            out.append({"项": name, "通": True})
        except DingError as e:
            msg = str(e)
            # 「找不到该用户」说明接口本身是通的，只是我拿假 userid 去探
            passed = "找不到该用户" in msg or "60121" in msg
            out.append({"项": name, "通": passed, "报错": "" if passed else msg,
                        "权限点": scope, "为什么要": why,
                        "申请链接": (f"https://open-dev.dingtalk.com/appscope/apply"
                                     f"?content={conf['appkey']}%23{scope}") if scope and not passed else ""})
            ok &= passed
    return {"ok": ok, "配置": True, "项": out}


# ==================== 花名册（在职 + 离职） ====================
# 为什么要缓存：汇总表上只有姓名，要换成 userid；而钉钉只能**反着查**
# （给 userid 才给姓名），没有按姓名搜人的接口。所以为了找到当月那一百多个
# 已离职的人，第一次不得不把全部离职账号的名字扫一遍（2026-08 是 9,050 个 / 453 次调用）。
#
# 但**离职员工的姓名和手机号不会再变**——人都走了。存下来，以后每次只补新增的那几个，
# 这一步就从 453 次降到接近 0。在职名单每次仍然重建（会变，且只要约 200 次）。
_CACHE_KEY = "tempatt_ding_roster"      # 离职段：{uid: {"姓名","手机"}}，人走了就不再变
_ROSTER_KEY = "tempatt_ding_roster_all"  # 整份花名册（在职+离职）：{uid: {...}}，带建表日期
# ⚠ 缓存里存的东西**结构一变就必须换这个号**，否则旧缓存会被当成好的接着用。
# V2.374 给离职员工补了「部门」字段，忘了换号，结果 215 人的盲区看着像没修好——踩过。
_ROSTER_VER = 3        # V3：部门存全路径（公司-…-派遣方），不再只存叶子名


def _cache_load():
    v = db.get_setting(_CACHE_KEY) or {}
    d = v.get("离职")
    return (d if isinstance(d, dict) else {}), v.get("更新") or ""


def _cache_save(left, who="系统"):
    try:
        db.set_setting(_CACHE_KEY, {"离职": left,
                                    "更新": time.strftime("%Y-%m-%d %H:%M:%S")}, who)
    except Exception:
        pass                                        # 缓存写不进去不该让取数失败


def _roster_load():
    v = db.get_setting(_ROSTER_KEY) or {}
    if int(v.get("版本") or 0) != _ROSTER_VER:      # 结构变了，旧的一律不认
        return {}, "", ""
    d = v.get("人")
    return (d if isinstance(d, dict) else {}), (v.get("建于") or ""), (v.get("耗时") or "")


def _roster_save(roster, cost="", who="系统"):
    try:
        db.set_setting(_ROSTER_KEY, {"人": roster, "耗时": cost, "版本": _ROSTER_VER,
                                     "建于": time.strftime("%Y-%m-%d %H:%M:%S")}, who)
    except Exception:
        pass


def cache_info():
    left, at = _cache_load()
    full, built, cost = _roster_load()
    return {"离职缓存": len(left), "更新于": at,
            "花名册缓存": len(full), "建于": built, "上次耗时": cost}


def cache_clear(who="系统"):
    """疑心缓存脏了就清掉，下次重建。"""
    db.set_setting(_CACHE_KEY, {"离职": {}, "更新": ""}, who)
    db.set_setting(_ROSTER_KEY, {"人": {}, "建于": ""}, who)

def build_roster(progress=None, force=False):
    """{userid: {"姓名":…, "部门":[…]}}。

    必须带上离职员工：2026-06 实测，汇总表 332 个名字里 159 个在通讯录查不到，
    而这 159 个**全部**出现在打卡表里——他们打了卡、只是人已经走了。
    只用在职花名册命中率 52%，加上离职是 100%。
    """
    say = progress or (lambda *a, **k: None)
    say("正在建「姓名→钉钉ID」对照表（钉钉只能按 ID 反查，没有按姓名搜人的接口）…", 3)
    # 部门要连**全路径**存（「公司-生产制造部-临时普工-华顺人力」），跟考勤系统导出的写法一致，
    # 而不是只存叶子名「华顺人力」——使用者 2026-08-29 要「保留这个组织架构」。
    # listsub(d) 返回的是 d 的直接子部门，所以子部门的父亲就是 d；顺着父链往上拼即得全称。
    dept_name, dept_parent, stack, seen = {}, {}, [1], {1}
    try:                                               # 根节点（公司）的名字，全路径从它起头
        _root = _api("topapi/v2/department/get", {"dept_id": 1}) or {}
        dept_name[1] = (_root.get("name") if isinstance(_root, dict) else "") or ""
    except DingError:
        dept_name[1] = ""                              # 取不到就路径不带公司名，不致命
    while stack:
        d = stack.pop()
        for x in (_api("topapi/v2/department/listsub", {"dept_id": d}) or []):
            i = x.get("dept_id")
            if i and i not in seen:
                seen.add(i); dept_name[i] = x.get("name") or ""
                dept_parent[i] = d; stack.append(i)

    def _full_path(i):
        parts, guard = [], 0
        while i and guard < 40:                        # guard 防环，正常撑死几层
            if dept_name.get(i):
                parts.append(dept_name[i])
            i = dept_parent.get(i)
            guard += 1
        return "-".join(reversed(parts))

    depts = {i: _full_path(i) for i in dept_name if i != 1}   # 部门id → 全称
    # 叶子名 → 全称（只在叶子名唯一时才建）：离职员工只拿得到叶子名，用它补全路径
    _leaf2full = {}
    for full in depts.values():
        _leaf2full.setdefault(full.rsplit("-", 1)[-1], set()).add(full)
    _leaf2full = {k: next(iter(v)) for k, v in _leaf2full.items() if len(v) == 1}
    say(f"部门 {len(depts)} 个（已按全路径归一）", 5)

    roster = {}
    for did in [1] + list(depts):
        cur = 0
        while True:
            res = _api("topapi/user/listsimple",
                       {"dept_id": did, "cursor": cur, "size": 100}) or {}
            for x in (res.get("list") or []):
                r = roster.setdefault(x["userid"], {"姓名": x.get("name"), "部门": []})
                nm = depts.get(did, "")
                if nm and nm not in r["部门"]:
                    r["部门"].append(nm)
            if not res.get("has_more"):
                break
            cur = res.get("next_cursor") or 0
    say(f"对照表：在职 {len(roster)} 人（这一步不取打卡，只取姓名）", 15)

    dim, off = [], 0
    while off < 60000:
        res = _api("topapi/smartwork/hrm/employee/querydimission",
                   {"offset": off, "size": 50}) or {}
        b = res.get("data_list") or []
        if not b:
            break
        dim += b
        off += len(b)
        if len(b) < 50:
            break

    cached, at = _cache_load()
    # 「没见过的」＋「见过但没存部门的」都要补。部门是后加的字段（V2.374），
    # 早先缓存的 9 千条只有姓名和手机号，缺了它「归属与打卡不符」对离职的人就查不了。
    todo = [u for u in dim if u not in cached or "部门" not in cached[u]]
    if todo:
        say(f"离职 {len(dim)} 人，其中 {len(todo)} 个是新的，正在取姓名…", 20)
        for i in range(0, len(todo), 20):              # 这个接口一次最多 20 个
            for e in (_api("topapi/smartwork/hrm/employee/list",
                           {"userid_list": ",".join(todo[i:i + 20]),
                            "field_filter_list": "sys00-name,sys00-mobile,sys00-dept"}) or []):
                f = {x.get("field_code"): x.get("value") for x in (e.get("field_list") or [])}
                if f.get("sys00-name"):
                    # sys00-dept 对离职员工照样有值（实测 12/12 是「锦绣人力」「天幕人力」这种），
                    # 正是「归属与打卡不符」要比的那个派遣方——上次抽样抽到几个测试账号，
                    # 误以为拿不到，白白留了 215 人的盲区。
                    cached[e["userid"]] = {"姓名": f["sys00-name"],
                                           "手机": _mobile(f.get("sys00-mobile")),
                                           "部门": str(f.get("sys00-dept") or "").strip()}
            if i % 400 == 0:
                say(f"补新离职人员 {i}/{len(todo)}", 20 + int(15 * i / max(1, len(todo))))
        _cache_save(cached)
    else:
        say(f"离职 {len(dim)} 人，全部已缓存（{at}），跳过", 33)

    for u in dim:
        c = cached.get(u)
        if c and u not in roster:
            # 离职缓存里的部门是叶子名（sys00-dept 只给这个），映射成全称与在职的写法对齐；映不到就留原样
            _dep = c.get("部门")
            _dep = _leaf2full.get(_dep, _dep) if _dep else _dep
            roster[u] = {"姓名": c["姓名"], "已离职": True, "手机": c.get("手机") or "",
                         "部门": [x for x in [_dep] if x]}
    say(f"花名册 {len(roster)} 人（离职缓存 {len(cached)}）", 35)
    return roster


def _mobile(v):
    return str(v or "").replace("+86-", "").replace("+86", "").strip()


def fetch_mobiles(uids):
    """{userid: 手机号}。只有智能人事那个口给手机号，user/get 不返回。一次最多 20 个。
    离职的人手机号已经在缓存里，直接用，不再问一次。"""
    cached, _ = _cache_load()
    out = {u: cached[u]["手机"] for u in uids
           if cached.get(u) and cached[u].get("手机")}
    uids = [u for u in uids if u not in out]
    for i in range(0, len(uids), 20):
        try:
            res = _api("topapi/smartwork/hrm/employee/list",
                       {"userid_list": ",".join(uids[i:i + 20]),
                        "field_filter_list": "sys00-mobile"}) or []
        except DingError:
            continue                                    # 拿不到手机号不该中断取数，退化成按打卡定人
        for e in res:
            v = next((f.get("value") for f in (e.get("field_list") or [])
                      if f.get("field_code") == "sys00-mobile"), None)
            if v:
                out[e["userid"]] = _mobile(v)
    return out


def tail(v, k=4):
    """手机号尾号，给人核对用——整串号码没必要摊在页面上。"""
    v = str(v or "")
    return v[-k:] if len(v) >= k else v


def roster_for(names, progress=None, force=False):
    """给这批姓名找一份够用的花名册。**缓存优先，缺谁才回钉钉重扫。**

    为什么值得这么做：建整份花名册要走 101 个部门 + 离职名单，约 380 次调用、90 秒；
    而 2026-07 那种单车间月份只有 40 个人、打卡本身才 1,240 次 40 秒——
    建表比取数还慢，使用者一眼就看出不对（2026-08-29 反馈「怎么还是取那么多人」）。

    但**不能无条件用缓存**：新入职的人不在里面，静默漏掉比慢更糟。
    所以判据是「这批名字缓存里齐不齐」——
      · 全都找得到 → 直接用，零调用；
      · 有找不到的 → 说明有新人，重建一次再匹配（顺带把缓存刷新）。
    这样稳态几乎不花时间，来了新人也不会漏。
    """
    say = progress or (lambda *a, **k: None)
    if not force:
        cached, built, _ = _roster_load()
        if cached:
            mt = match_names(names, cached)
            if not mt["查无"]:
                say(f"花名册用缓存（{len(cached)} 人，建于 {built}），"
                    f"{len(names)} 个名字全部对上，跳过重建", 35)
                return cached
            say(f"缓存里有 {len(mt['查无'])} 个名字找不到（多半是新入职），重建花名册…", 3)
    t0 = time.time()
    roster = build_roster(say, force=force)
    _roster_save(roster, f"{time.time() - t0:.0f} 秒")
    return roster


# ==================== 姓名 → userid ====================
_PAREN = re.compile(r"[（(][^）)]*[)）]")


def _strip(s):
    """只去掉括号备注（（离职）（驻场）），保留「张丽T」这种人力自己编的尾缀。"""
    return _PAREN.sub("", str(s or "")).strip()


def match_names(names, roster):
    """把汇总表上的姓名对到 userid。

    三种结局，撞名的**不猜**：整份花名册 650 个名字被 1466 个人共用
    （张浩 6 个、李婷 6 个），猜错就是把甲的打卡记到乙头上。
    """
    exact, loose = {}, {}
    for uid, r in roster.items():
        exact.setdefault(r["姓名"], []).append(uid)
        loose.setdefault(_strip(r["姓名"]), []).append(uid)
        # 「张丽T」这类尾缀，内核的 norm_name 会去掉，这里也备一份
        loose.setdefault(_norm(r["姓名"]), []).append(uid)

    hit, dup, miss = {}, {}, []
    for n in names:
        for tbl, key in ((exact, n), (loose, _strip(n)), (loose, _norm(n))):
            got = tbl.get(key)
            if got:
                uids = sorted(set(got))
                if len(uids) == 1:
                    hit[n] = uids[0]
                else:
                    dup[n] = uids
                break
        else:
            miss.append(n)
    return {"唯一": hit, "撞名": dup, "查无": sorted(miss)}


_TAIL = re.compile(r"[（(].*?[)）]|[A-Za-z0-9]+")


def _norm(s):
    """与内核 temp_attendance.norm_name 同规则，避免两处口径不一致。"""
    return _TAIL.sub("", str(s or "")).strip()


# ==================== 取打卡 ====================
def _minutes(work_date, stamp):
    """打卡时刻 → 当日 0 点起的分钟数。跨到次日加 1440（与 Excel 的「次日07:52」同义）。"""
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})[ T](\d{1,2}):(\d{2})", str(stamp or ""))
    if not m:
        return None
    y, mo, d, hh, mi = (int(x) for x in m.groups())
    delta = (date(y, mo, d) - work_date).days
    if delta < 0 or delta > 2:                         # 离谱的丢掉，别污染判定
        return None
    return delta * 1440 + hh * 60 + mi


def _score(worked, punched):
    """候选人的打卡日 vs 汇总表上说他上工的日子，算个 F1。

    两头都要看：
      · 覆盖率（上工的日子他有没有打卡）——只看这个，一个天天在岗的正式工也能拿满分
      · 准确率（他打卡的日子是不是都在上工日里）——临时工只在上工那几天刷卡，
        正式工整月都刷，这一项能把两者分开（李志鹏：临时工 1/1，同名正式工 1/21）
    """
    if not worked or not punched:
        return 0.0
    hit = len(worked & punched)
    if not hit:
        return 0.0
    cov, pre = hit / len(worked), hit / len(punched)
    return 2 * cov * pre / (cov + pre)


def resolve_dups(dup, month, worked_days=None, progress=None):
    """重名的人怎么办：先用**手机号**分清是几个人，再用**上工日**定是哪一个。

    为什么必须处理而不是跳过：跳过等于这个人在打卡表里凭空消失，
    下游会把他所有上工日判成「报了工时却没打卡」——2026-06 实测，
    46 个重名的人这样一跳，红灯从 17 条涨到 538 条，全是假的。

    两步：
      ① 手机号相同的候选＝**同一个人的多个钉钉账号**（离职再入职会新建 userid）。
         合并，不是二选一——只选一个会丢掉另一个账号那几天的打卡。2026-06 有 8 组是这种。
      ② 剩下手机号不同的，才是真的不同的人。按「打卡日 vs 汇总表上工日」算 F1，
         要求 ≥0.5 且甩开第二名一倍才认；拿不准宁可空着交人工——
         猜错就是把甲的工时记到乙头上。

    **无论自动定成什么，都要原样报给成本会计复核**（返回值里的「记录」），
    页面和生成的打卡表里都会列出来：选中谁、手机尾号、几天打卡、判据是什么。
    """
    say = progress or (lambda *a, **k: None)
    worked_days = worked_days or {}
    y, mo = (int(x) for x in month.split("-"))
    last = monthrange(y, mo)[1]
    cands = sorted({u for v in dup.values() for u in v})
    if not cands:
        return {}, {}, {}, []
    say(f"重名 {len(dup)} 组，取手机号分辨…", 34)
    mob = fetch_mobiles(cands)
    say(f"重名 {len(dup)} 组，取当月打卡定人（{len(cands)} 个候选）…", 36)
    got = fetch_punches([(u, date(y, mo, d)) for u in cands for d in range(1, last + 1)])

    hit, still, rec = {}, {}, []
    for nm, uids in sorted(dup.items()):
        # ① 手机号相同的并成一个人（没手机号的各算一个，别乱并）
        groups, solo = {}, []
        for u in uids:
            m = mob.get(u)
            if m:
                groups.setdefault(m, []).append(u)
            else:
                solo.append(u)                          # 没手机号的各算一个，别乱并
        people = list(groups.values()) + [[u] for u in solo]

        w = set(worked_days.get(nm) or ())
        scored = []
        for grp in people:
            days = {}
            for u in grp:
                days.update(got.get(u) or {})           # 合并同一个人多个账号的打卡
            scored.append((_score(w, set(days)), grp, days))
        scored.sort(key=lambda x: -x[0])
        best = scored[0]
        second = scored[1][0] if len(scored) > 1 else 0.0
        ok = len(scored) == 1 or (best[0] >= 0.5 and best[0] >= 2 * second)

        item = {"姓名": nm, "候选人数": len(people), "钉钉账号数": len(uids),
                "上工日数": len(w), "已定": bool(ok),
                "合并账号": sum(1 for g in people if len(g) > 1),
                "候选": [{"账号": g, "手机尾号": tail(mob.get(g[0])),
                          "打卡日数": len(d), "命中上工日": len(w & set(d)),
                          "得分": round(sc, 2), "选中": bool(ok and g is best[1])}
                         for sc, g, d in scored]}
        rec.append(item)
        if ok:
            hit[nm] = {"账号": best[1], "days": best[2]}
        else:
            still[nm] = item
    say(f"重名自动定人 {len(hit)}／仍需人工 {len(still)}", 40)
    return hit, still, got, rec


# 跨月边界两天的键：上月最后一天 / 次月第一天。**故意放在 1..31 之外**，
# 这样逐日判定（compute 只认月内 1..last）永远碰不到它们，只有导出的⑧原始表会取。
BND_PREV, BND_NEXT = 0, 32


def _beijing(ms):
    """钉钉返回的 epoch 毫秒 → 北京时刻（绝对时间 +8，不看服务器时区，服务器设 UTC 也不会错）。"""
    return datetime.utcfromtimestamp(ms / 1000) + timedelta(hours=8)


def fetch_punches(uid_days, progress=None):
    """[(userid, date[, 键])] → {userid: {键: [分钟…]}}。
    优先走**批量** listRecord（一批人一段日期一次取，比逐人逐日快约两个数量级）；
    批量口不可用（无权限/报错）时自动**回退逐日** getupdatedata，功能不受影响。"""
    say = progress or (lambda *a, **k: None)
    try:
        return _fetch_punches_batch(uid_days, say)
    except Exception as e:                              # 批量口有任何闪失都不影响出数
        say(f"批量取数不可用（{e}），改用逐日口…", 40)
        return _fetch_punches_perday(uid_days, say)


def _fetch_punches_batch(uid_days, say):
    """批量口 listRecord：每条记录自带 workDate（考勤日/班次归属），按它归格——
    夜班次日凌晨的卡会归到上班那天，与逐日 getupdatedata 逐格一致（2026-08 六人全月实测 114/114）。
    调用数：从「人×天」（整月全量 9000+ 次）降到「人批(≤50) × 日窗(≤7天)」（整月约几十次）。
    输出契约与逐日口完全相同：{userid: {键: [分钟…]}}。"""
    want = {}                                           # (uid, date) → 键（日号 或 BND_*）
    for job in uid_days:
        want[(str(job[0]), job[1])] = job[2] if len(job) > 2 else job[1].day
    if not want:
        return {}
    uids = sorted({u for u, _ in want})
    ds = sorted({d for _, d in want})
    lo, hi = ds[0], ds[-1] + timedelta(days=1)          # +1 天：接住末日夜班落在次日的下班卡
    wins, s = [], lo                                    # ≤7 天日窗（listRecord 跨度上限）
    while s <= hi:
        e = min(s + timedelta(days=6), hi)
        wins.append((s, e)); s = e + timedelta(days=1)
    ubs = [uids[i:i + 50] for i in range(0, len(uids), 50)]   # 人按 50 分批
    calls = [(ub, a, b) for ub in ubs for (a, b) in wins]
    out, lock, done = {}, threading.Lock(), [0]
    total = max(1, len(calls))

    def one(job):
        ub, a, b = job
        recs = _api("attendance/listRecord",
                    {"userIds": ub,
                     "checkDateFrom": a.strftime("%Y-%m-%d 00:00:00"),
                     "checkDateTo": b.strftime("%Y-%m-%d 23:59:59")},
                    result_key="recordresult") or []
        with lock:
            for r in recs:
                u = str(r.get("userId"))
                wd = _beijing(r["workDate"]).date()
                key = want.get((u, wd))                 # 只留请求过的(人,考勤日)，多余记录自然丢弃
                if key is None:
                    continue
                ct = _beijing(r["userCheckTime"])
                mins = (ct.date() - wd).days * 1440 + ct.hour * 60 + ct.minute
                out.setdefault(u, {}).setdefault(key, set()).add(mins)
            done[0] += 1
            say(f"取打卡 {done[0]}/{total} 批", 40 + int(50 * done[0] / total))

    with ThreadPoolExecutor(max_workers=WORKERS) as ex:
        list(ex.map(one, calls))
    return {u: {k: sorted(v) for k, v in days.items()} for u, days in out.items()}


def _fetch_punches_perday(uid_days, say):
    """逐日口 getupdatedata（批量口不可用时的兜底）。一人一天一调，4–6 线程并发、限流退避。"""
    out, lock = {}, threading.Lock()
    done = [0]
    total = max(1, len(uid_days))

    def one(job):
        uid, day = job[0], job[1]
        key = job[2] if len(job) > 2 else day.day   # 三元组可显式给键，用于跨月边界
        try:
            res = _api("topapi/attendance/getupdatedata",
                       {"userid": uid, "work_date": day.strftime("%Y-%m-%d 00:00:00")}) or {}
        except DingError:
            res = {}                                    # 单点失败不炸整批，缺的天体现为「无打卡」
        ts = []
        for a in (res.get("check_record_list") or []):
            # 这里是原始流水，没有 SYSTEM 占位那种脏数据，不需要再按 time_result 过滤；
            # valid_matched=False 的也要（人力导出里同样有，比如丁玲 6/11 的 19:57）
            v = _minutes(day, a.get("user_check_time"))
            if v is not None:
                ts.append(v)
        with lock:
            if ts:
                out.setdefault(uid, {})[key] = sorted(set(ts))
            done[0] += 1
            if done[0] % 200 == 0 or done[0] == total:
                say(f"取打卡 {done[0]}/{total}", 40 + int(50 * done[0] / total))

    with ThreadPoolExecutor(max_workers=WORKERS) as ex:
        list(ex.map(one, uid_days))
    return out


# ==================== 产出打卡表 Excel ====================
def _cell(ts):
    """分钟列表 → 打卡格文本，与人力导出的写法一致（换行分隔，次日加前缀）。"""
    out = []
    for t in sorted(set(ts)):
        d, hm = divmod(t, 1440)
        out.append(("次日" if d else "") + f"{hm // 60:02d}:{hm % 60:02d}")
    return "\n".join(out)


def build_punch_xlsx(month, rows, dup_rec=None):
    """rows: [{"姓名","部门","考勤组","days":{日:[分钟]}}] → 打卡表 xlsx 字节。

    版式必须和人力从考勤系统导出的那张一致，否则 parse_punch 认不出：
      第 1 行 标题（含统计区间）／第 2 行 表头「姓名 部门 考勤组 …」／
      第 3 行 日期号 1..N（**表头的下一行**）／第 4 行起 数据。
    """
    if Workbook is None:
        raise DingError("缺少 openpyxl，无法生成打卡表")
    y, mo = (int(x) for x in month.split("-"))
    last = monthrange(y, mo)[1]
    nxt = f"{y + 1}-01-01" if mo == 12 else f"{y}-{mo + 1:02d}-01"

    wb = Workbook()
    ws = wb.active
    ws.title = "打卡时间"
    # ⚠ 标题**必须**以本月 1 日起头：期次识别（_period_from / 前端文件名提示）读的就是标题里的
    #   第一个日期。V2.394 一度把起点写成上月末（2026-05-31），结果六月的表被认成 2026-05，
    #   跳出「两张表月份对不上」的假警告。跨月边界那两天靠「上月末/次月初」两列自证，不靠标题。
    ws.cell(1, 1, f"打卡时间 统计日期：{y}-{mo:02d}-01 至 {nxt}")
    # 多一列「手机尾号」：复核那边靠它硬判同名的人是不是同一个（parse_punch 会读，没有也能跑）
    for c, v in enumerate(("姓名", "部门", "考勤组", "手机尾号"), start=1):
        ws.cell(2, c, v)
    for d in range(1, last + 1):
        ws.cell(3, 4 + d, d)                            # 日期行＝表头下一行，从第 5 列起
    # 跨月边界两列：只有整月取数（带 BND_PREV/BND_NEXT）时才摆，摆在日号区之后。
    # 表头写「上月末」「次月初」（非数字），parse_punch 的 _day_columns 只认 1..31，
    # 不会把它们当日期，两边互不打架；parse 另有一段专读这两列。
    _has_prev = any(BND_PREV in (r.get("days") or {}) for r in rows)
    _has_next = any(BND_NEXT in (r.get("days") or {}) for r in rows)
    _cprev = 4 + last + 1 if _has_prev else 0
    _cnext = 4 + last + (2 if _has_prev else 1) if _has_next else 0
    if _cprev:
        ws.cell(3, _cprev, "上月末")
    if _cnext:
        ws.cell(3, _cnext, "次月初")
    for i, r in enumerate(rows):
        rr = 4 + i
        ws.cell(rr, 1, r.get("姓名") or "")
        ws.cell(rr, 2, r.get("部门") or "")
        ws.cell(rr, 3, r.get("考勤组") or "")
        ws.cell(rr, 4, r.get("手机尾号") or "")
        _days = r.get("days") or {}
        for d, ts in _days.items():
            if 1 <= int(d) <= last and ts:
                ws.cell(rr, 4 + int(d), _cell(ts))
        if _cprev and _days.get(BND_PREV):
            ws.cell(rr, _cprev, _cell(_days[BND_PREV]))
        if _cnext and _days.get(BND_NEXT):
            ws.cell(rr, _cnext, _cell(_days[BND_NEXT]))
    ws.freeze_panes = "E4"
    for col, w in (("A", 14), ("B", 30), ("C", 18), ("D", 10)):
        ws.column_dimensions[col].width = w
    from openpyxl.utils import get_column_letter as _gcl
    for _c in (c for c in (_cprev, _cnext) if c):
        ws.column_dimensions[_gcl(_c)].width = 16

    # 第二页：重名定人底稿，交成本会计复核。
    # parse_punch 只读第一个工作表，所以加这一页不影响解析；
    # 而它跟着文件走，下载下来就能核、能存档——不能让「工具替人做了选择」这件事没有痕迹。
    if dup_rec:
        w2 = wb.create_sheet("重名定人·待成本会计复核")
        w2.cell(1, 1, f"{month} 重名定人底稿——同名的人，工具选了谁、凭什么选的。请逐行核对。")
        w2.cell(2, 1, "手机号相同＝同一个人在钉钉有两个账号（离职再入职），已合并打卡；"
                      "手机号不同＝真的是不同的人，按「打卡日和汇总表上工日对不对得上」定的。")
        head = ("姓名", "结论", "候选", "手机尾号", "钉钉账号", "当月打卡天数",
                "命中上工日", "上工日数", "吻合度", "是否选中")
        for c, v in enumerate(head, start=1):
            w2.cell(4, c, v)
        r = 5
        for it in dup_rec:
            for i, cd in enumerate(it["候选"]):
                w2.cell(r, 1, it["姓名"] if i == 0 else "")
                w2.cell(r, 2, ("已定人" if it["已定"] else "⚠ 定不了，需人工") if i == 0 else "")
                w2.cell(r, 3, f"候选{i + 1}" + ("（同一人的%d个账号已合并）" % len(cd["账号"])
                                                if len(cd["账号"]) > 1 else ""))
                w2.cell(r, 4, cd["手机尾号"] or "（无）")
                w2.cell(r, 5, "、".join(cd["账号"]))
                w2.cell(r, 6, cd["打卡日数"])
                w2.cell(r, 7, cd["命中上工日"])
                w2.cell(r, 8, it["上工日数"] if i == 0 else "")
                w2.cell(r, 9, cd["得分"])
                w2.cell(r, 10, "✔ 选中" if cd["选中"] else "")
                r += 1
        w2.freeze_panes = "A5"
        for col, wd in (("A", 12), ("B", 16), ("C", 24), ("D", 10), ("E", 44),
                        ("F", 12), ("G", 12), ("H", 10), ("I", 9), ("J", 9)):
            w2.column_dimensions[col].width = wd
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ==================== 一条龙 ====================
def pull_month(month, names, progress=None, roster=None, worked_days=None,
               force_roster=False, scope="worked"):
    """给定期次与汇总表上的姓名清单 → 打卡表 xlsx + 对名结果。

    只取汇总表上有的人：全厂 1300+ 人全取要 4 万次调用，而复核只关心工资表上这些人。
    代价是「有打卡但不在工资表上的人」查不出来——那一档本来就是中性提示。

    scope 控制取哪些天，**这是快慢的主要开关**：
      · "worked"（默认，快）：只取汇总表上**报了工时的那些天**。2026-06 是 2,841 人日，
        而整月是 8,550——少 2/3。**所有红灯一条不少**：「打卡撑不起上报」「报了工时却没打卡」
        判的都是有上报的日子。代价只是「◇ 有打卡·未计工时」这一中性档统计不出来
        （那一档本来就写着「不是问题」），页面会标注本次未取。
      · "full"（慢）：整月每一天都取，中性档也齐。

    worked_days: {姓名: [日,…]}，汇总表上这个人上了工的日子。撞名定人要靠它，
    不传也能跑，但撞名会大量退回人工。
    """
    say = progress or (lambda *a, **k: None)
    okm, why = month_reachable(month)
    if not okm:
        raise DingError(why)
    y, mo = (int(x) for x in month.split("-"))
    last = monthrange(y, mo)[1]

    roster = roster or roster_for(names, say, force=force_roster)
    mt = match_names(names, roster)
    say(f"对上 {len(mt['唯一'])} 人｜撞名 {len(mt['撞名'])}｜查无 {len(mt['查无'])}", 38)

    fixed, still, cached, rec = resolve_dups(mt["撞名"], month, worked_days, say)

    uids = sorted(set(mt["唯一"].values()))
    if scope == "worked":
        # 只取报了工时的那些天：红灯判的都是有上报的日子，一条不少
        # ⚠ 必须连**前后各一天**一起取：夜班的下班卡落在次日、次日凌晨的卡又可能属于前一晚，
        #   compute_shifts 要拿相邻日才切得出班。只取上报当天的话，夜班在厂时长算不出来——
        #   实测 2026-06 会凭空多出 188 条假「撑不起上报」（4 → 192）。
        n2u = {n: u for n, u in mt["唯一"].items()}
        jobs = sorted({(n2u[n], date(y, mo, dd2))
                       for n, ds in (worked_days or {}).items() if n in n2u
                       for d in ds for dd2 in (int(d) - 1, int(d), int(d) + 1)
                       if 1 <= dd2 <= last})
        say(f"只取报了工时的日子：{len(uids)} 人 / {len(jobs)} 人日"
            f"（整月要 {len(uids) * last} 次，省下 {max(0, len(uids) * last - len(jobs))} 次）", 42)
    else:
        # 整月每一天 + **跨月边界两天**：夜班跨零点，上月末那班的下班卡落在 1 日凌晨、
        # 本月末那班的下班卡落在次月 1 日——不取这两天，边界上的夜班在原始表里就是残的
        # （使用者 2026-08-29：「还是要导出全月的」「甚至要将 T-1 和 T+1 也导出来」）。
        d_prev = date(y, mo, 1) - timedelta(days=1)
        d_next = date(y, mo, last) + timedelta(days=1)
        jobs = [(u, date(y, mo, d), d) for u in uids for d in range(1, last + 1)]
        jobs += [(u, d_prev, BND_PREV) for u in uids]
        jobs += [(u, d_next, BND_NEXT) for u in uids]
        say(f"整月全取 + 跨月边界：{len(uids)} 人 ×（{last}+2）天 = {len(jobs)} 次…", 42)
    got = fetch_punches(jobs, say)

    say("正在生成打卡表…", 92)
    _tag = "钉钉取数" if scope != "worked" else "钉钉取数·仅上报日±1天"
    # 全员的手机尾号：332 人只要 17 次调用，换来同名风险能硬判，很划算
    _all_uids = sorted(set(mt["唯一"].values()) | {u for v in fixed.values() for u in v["账号"]})
    _mob = fetch_mobiles(_all_uids)
    _tail = lambda u: tail(_mob.get(u))
    rows = []
    for nm in sorted(mt["唯一"]):
        uid = mt["唯一"][nm]
        r = roster.get(uid) or {}
        rows.append({"姓名": nm,                         # 用汇总表上的写法，方便逐格核对
                     "部门": "、".join(r.get("部门") or []),
                     "考勤组": _tag, "手机尾号": _tail(uid),
                     "days": got.get(uid) or {}})
    for nm, v in sorted(fixed.items()):
        # 合并了多个账号的人：取**第一个有部门的**账号，别死认第一个——
        # 离职那个账号常常没部门，取到它就等于把部门丢了（陈双珍/李苏/马福 实测）
        r = next((roster.get(u) or {} for u in v["账号"] if (roster.get(u) or {}).get("部门")),
                 roster.get(v["账号"][0]) or {})
        rows.append({"姓名": nm, "部门": "、".join(r.get("部门") or []),
                     "考勤组": _tag + "·重名已定人", "手机尾号": _tail(v["账号"][0]),
                     "days": v["days"]})
    rows.sort(key=lambda x: x["姓名"])
    xlsx = build_punch_xlsx(month, rows, dup_rec=rec)
    say("完成", 100)
    # 「未取到」必须原样报出去：这些人不在打卡表里，下游会把他们判成「没打卡」。
    # 不讲清楚，使用者会把工具的盲区当成员工的问题。
    miss = sorted(set(still) | set(mt["查无"]))
    return {"xlsx": xlsx, "人数": len(rows), "取数范围": scope,
            "对上": len(mt["唯一"]) + len(fixed), "撞名自动定人": len(fixed),
            "重名记录": rec, "合并账号组数": sum(1 for x in rec if x["合并账号"]),
            "未取到": miss, "未取到明细": still, "查无此人": mt["查无"],
            "有打卡人数": sum(1 for r in rows if r["days"]),
            "打卡日次": sum(len(r["days"]) for r in rows),
            "调用次数": len(jobs) + len(cached) * last}
